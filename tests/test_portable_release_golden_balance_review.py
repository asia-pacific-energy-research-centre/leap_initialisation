"""Golden regression test for the USA 2022 balance-review workbook.

This pins the balance-review command against a known-good historical case:

    outputs/leap_exports/supply_reconciliation/supporting_files/
      baseline_seed_balance_diagnostics/results_update_preview_20260803_usa_tgt/
      comparison_workbooks/balance_review_20_USA_tgt_2022.xlsx

The test rebuilds that workbook **into a temporary directory** from the same
inputs and compares its structural contract and selected core values against the
golden expectation recorded below. It never writes to the historical workbook,
the diagnostics directory, or the LEAP balance export — and it asserts that,
by re-hashing both source inputs afterwards.

The inputs are large local artifacts that are not tracked in Git, so the test
skips cleanly on a machine that does not have them.
"""

from __future__ import annotations

import hashlib
from pathlib import Path

import pytest

from codebase.portable_release.commands import run_balance_review
from codebase.portable_release.runtime import RuntimeContext


REPO_ROOT = Path(__file__).resolve().parents[1]

GOLDEN_DIAGNOSTICS_DIR = (
    REPO_ROOT
    / "outputs"
    / "leap_exports"
    / "supply_reconciliation"
    / "supporting_files"
    / "baseline_seed_balance_diagnostics"
    / "results_update_preview_20260803_usa_tgt"
)
GOLDEN_REFERENCE_WORKBOOK = (
    GOLDEN_DIAGNOSTICS_DIR / "comparison_workbooks" / "balance_review_20_USA_tgt_2022.xlsx"
)
GOLDEN_BALANCE_EXPORT = (
    REPO_ROOT / "data" / "leap balances exports" / "20_USA" / "TGT 0308.xlsx"
)

GOLDEN_ECONOMY = "20_USA"
GOLDEN_SCENARIO = "Target"
GOLDEN_YEAR = 2022

#: Sheet order and names are the workbook's contract with its readers.
GOLDEN_SHEET_NAMES = [
    "LEAP Values",
    "LEAP - Source Error",
    "Full Expected Source",
]
GOLDEN_SOURCE_SHAPE = {"rows": 79, "columns": 49}

GOLDEN_STATUS_COUNTS = {"value_mismatch": 180, "reference_unavailable": 14, "match": 162}
GOLDEN_ISSUE_COUNTS = {"missing_esto_pair": 11}
GOLDEN_COMPARISON_STATE_COUNTS = {
    "mapped": 316,
    "no_direct_projection_comparator": 12,
    "reference_unavailable": 2,
    "missing_in_leap": 0,
    "missing_visible_structure": 26,
    "ambiguous_structure_resolution": 0,
}
GOLDEN_MISSING_AUDIT_ROWS = 51

GOLDEN_METADATA = {
    "title": 'Energy Balance for Area "USA clean slate 29_07"',
    "scenario": "Target",
    "year": 2022,
    "sourceUnits": "Petajoule",
    "outputUnits": "Petajoule",
    "sourceUnitMultiplier": 1.0,
}

#: A few core values, chosen to cover a supply row, an import row, a sign
#: change, and the summary block. ``address`` is on the diagnostic error sheet;
#: ``leap_value`` is the corresponding cell on the LEAP Values sheet.
GOLDEN_CELLS = [
    {"address": "AK4", "row": "Production", "fuel": "Natural gas liquids",
     "error": -8.102158, "leap_value": 7415.522195999019},
    {"address": "B5", "row": "Imports", "fuel": "Electricity",
     "error": 6.266241, "leap_value": 211.359331},
    {"address": "F5", "row": "Imports", "fuel": "Crude oil",
     "error": -531.299682, "leap_value": 13055.64209776145},
    {"address": "O5", "row": "Imports", "fuel": "Bitumen",
     "error": 99.2477, "leap_value": 241.8549805826103},
]
TOLERANCE = 1e-6


pytestmark = pytest.mark.skipif(
    not (GOLDEN_DIAGNOSTICS_DIR.is_dir() and GOLDEN_BALANCE_EXPORT.is_file()),
    reason=(
        "The USA 2022 golden inputs are large local artifacts that are not tracked "
        "in Git. Restore them from the shared archive to run this test."
    ),
)


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _context(tmp_path: Path) -> RuntimeContext:
    """A minimal runtime context; balance-review needs no configuration assets."""
    return RuntimeContext(
        mode="developer",
        release_name="leap-review-tools",
        release_version="0.0.0+test",
        package_root=tmp_path,
        config_root=tmp_path / "config",
        output_root=tmp_path / "output",
        log_root=tmp_path / "logs",
        input_root=tmp_path / "input",
        repository_roots={"leap_initialisation": REPO_ROOT},
    )


@pytest.fixture(scope="module")
def golden_result(tmp_path_factory: pytest.TempPathFactory):
    """Rebuild the golden workbook once, into a temporary directory."""
    tmp_path = tmp_path_factory.mktemp("golden_balance_review")
    inputs_before = {
        path: _sha256(path)
        for path in [
            GOLDEN_BALANCE_EXPORT,
            GOLDEN_REFERENCE_WORKBOOK,
            *sorted(GOLDEN_DIAGNOSTICS_DIR.glob("leap_balance_*.csv")),
        ]
        if path.is_file()
    }
    result = run_balance_review(
        _context(tmp_path),
        economy=GOLDEN_ECONOMY,
        scenario=GOLDEN_SCENARIO,
        year=GOLDEN_YEAR,
        balance_export_workbook=GOLDEN_BALANCE_EXPORT,
        diagnostics_directory=GOLDEN_DIAGNOSTICS_DIR,
        run_label="golden",
    )
    assert result.ok, result.error
    return result, inputs_before


def test_source_inputs_were_not_modified(golden_result) -> None:
    _, inputs_before = golden_result
    for path, digest in inputs_before.items():
        assert _sha256(path) == digest, f"{path} was modified by the build"


def test_output_is_written_only_under_the_temporary_run_directory(golden_result) -> None:
    result, _ = golden_result
    workbook = Path(result.outputs["workbook"])
    assert workbook.is_file()
    # Deliverables land in the per-economy folder; the manifest and validation
    # report live in a run_records sub-folder beneath it.
    assert workbook.parent == result.output_directory
    assert result.run_directory.parent.parent == result.output_directory
    assert workbook.parent.name == "balance_review"
    assert workbook.parent.parent.name == "20_USA"
    # Whatever the layout, nothing may be written over the historical reference.
    assert GOLDEN_REFERENCE_WORKBOOK.resolve() != workbook.resolve()


def test_build_summary_matches_the_golden_expectation(golden_result) -> None:
    result, _ = golden_result
    build = result.outputs["build_result"]
    assert build["sourceSheet"] == "2022"
    assert build["sourceShape"] == GOLDEN_SOURCE_SHAPE
    assert build["metadata"] == GOLDEN_METADATA
    assert build["statusCounts"] == GOLDEN_STATUS_COUNTS
    assert build["issueCounts"] == GOLDEN_ISSUE_COUNTS
    assert build["comparisonStateCounts"] == GOLDEN_COMPARISON_STATE_COUNTS
    assert build["missingAuditRows"] == GOLDEN_MISSING_AUDIT_ROWS
    assert build["formulaErrorCells"] == []


def test_workbook_structural_contract(golden_result) -> None:
    from openpyxl import load_workbook

    result, _ = golden_result
    workbook = load_workbook(Path(result.outputs["workbook"]))
    try:
        assert workbook.sheetnames == GOLDEN_SHEET_NAMES
        leap_values = workbook["LEAP Values"]
        assert leap_values.max_row == GOLDEN_SOURCE_SHAPE["rows"]
        assert leap_values.max_column == GOLDEN_SOURCE_SHAPE["columns"]
        assert leap_values.freeze_panes == "B4"
        assert str(leap_values["A2"].value or "").endswith("Units: Petajoule")

    finally:
        workbook.close()


def test_selected_core_values_match_the_golden_expectation(golden_result) -> None:
    from openpyxl import load_workbook

    result, _ = golden_result
    workbook = load_workbook(Path(result.outputs["workbook"]))
    try:
        leap_values = workbook["LEAP Values"]
        error_sheet = workbook["LEAP - Source Error"]
        full_expected_sheet = workbook["Full Expected Source"]
        for cell in GOLDEN_CELLS:
            address = cell["address"]
            column = error_sheet[address].column
            row = error_sheet[address].row
            assert leap_values.cell(row, 1).value.strip() == cell["row"], address
            assert leap_values.cell(3, column).value == cell["fuel"], address
            assert leap_values[address].value == pytest.approx(
                cell["leap_value"], abs=TOLERANCE
            ), address
            assert error_sheet[address].value == pytest.approx(
                cell["error"], abs=TOLERANCE
            ), address
            assert full_expected_sheet[address].value == pytest.approx(
                cell["leap_value"] - cell["error"], abs=TOLERANCE
            ), address
    finally:
        workbook.close()


def test_rebuild_matches_the_historical_reference_workbook(golden_result) -> None:
    """The current code still reproduces the named historical artifact.

    The reference workbook is opened read-only and never written. Only the
    diagnostic values are compared: styling and Excel's own bookkeeping are not
    part of the contract this test defends.
    """
    from openpyxl import load_workbook

    if not GOLDEN_REFERENCE_WORKBOOK.is_file():
        pytest.skip("The historical reference workbook is not present here.")

    result, _ = golden_result
    rebuilt = load_workbook(Path(result.outputs["workbook"]))
    reference = load_workbook(GOLDEN_REFERENCE_WORKBOOK)
    try:
        assert rebuilt.sheetnames == GOLDEN_SHEET_NAMES
        assert all(name in reference.sheetnames for name in GOLDEN_SHEET_NAMES)
        for sheet_name in ["LEAP Values", "LEAP - Source Error"]:
            new_sheet = rebuilt[sheet_name]
            old_sheet = reference[sheet_name]
            assert new_sheet.max_row == old_sheet.max_row, sheet_name
            assert new_sheet.max_column == old_sheet.max_column, sheet_name
            differences: list[str] = []
            for row in range(4, old_sheet.max_row + 1):
                for column in range(2, old_sheet.max_column + 1):
                    new_value = new_sheet.cell(row, column).value
                    old_value = old_sheet.cell(row, column).value
                    if isinstance(old_value, (int, float)) and isinstance(
                        new_value, (int, float)
                    ):
                        if abs(float(new_value) - float(old_value)) > TOLERANCE:
                            differences.append(
                                f"{sheet_name}!{new_sheet.cell(row, column).coordinate}: "
                                f"{old_value} -> {new_value}"
                            )
                    elif new_value != old_value:
                        differences.append(
                            f"{sheet_name}!{new_sheet.cell(row, column).coordinate}: "
                            f"{old_value!r} -> {new_value!r}"
                        )
            assert not differences, "\n".join(differences[:20])
    finally:
        rebuilt.close()
        reference.close()


def test_run_manifest_records_inputs_and_outputs(golden_result) -> None:
    result, _ = golden_result
    manifest = result.run_manifest
    assert manifest.status == "succeeded"
    assert manifest.command == "balance-review"
    input_roles = {record.role for record in manifest.inputs}
    assert "input:balance_export_workbook" in input_roles
    assert any(role.startswith("input:diagnostic_artifact") for role in input_roles)
    assert all(record.sha256 for record in manifest.inputs if record.exists)
    assert [record.role for record in manifest.outputs] == [
        "output:balance_review_workbook"
    ]
    assert manifest.validation["ok"] is True
    assert (result.run_directory / "run_manifest.json").is_file()
    assert (result.run_directory / "run_manifest.txt").is_file()
    assert (result.run_directory / "validation_report.txt").is_file()


def test_invalid_input_fails_cleanly_with_a_manifest(tmp_path: Path) -> None:
    result = run_balance_review(
        _context(tmp_path),
        economy=GOLDEN_ECONOMY,
        scenario=GOLDEN_SCENARIO,
        year=1234,
        balance_export_workbook=GOLDEN_BALANCE_EXPORT,
        diagnostics_directory=GOLDEN_DIAGNOSTICS_DIR,
        run_label="invalid",
    )
    assert not result.ok
    assert "outside the supported range" in (result.error or "")
    assert result.run_manifest.status == "failed"
    assert (result.run_directory / "run_manifest.json").is_file()
    assert (result.run_directory / "validation_report.txt").is_file()
    assert not list(result.run_directory.glob("*.xlsx"))
