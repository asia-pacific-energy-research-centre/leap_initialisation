"""Tests for the workbook-backed missing-branch warning ledger."""
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import Workbook, load_workbook

from codebase.functions.baseline_seed_validation_exceptions import (
    REQUIRED_COLUMNS,
    NINTH_VALUE_COLUMN,
    refresh_exception_materiality,
    apply_zero_filter,
    audit_exception_relevance,
    load_enabled_exception_notes,
    register_material_missing_branch_findings,
    register_missing_branch_paths,
)
from codebase.mapping_tools.missing_branch_registry_materiality_workflow import (
    _esto_base_materiality,
    _registry_source_keys,
    _source_sector_candidates,
)
import codebase.mapping_tools.missing_branch_registry_materiality_workflow as materiality


def _workbook(path: Path, branch_path: str = "") -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "branch_exceptions"
    sheet.append(REQUIRED_COLUMNS)
    if branch_path:
        sheet.append([True, branch_path, "", "", "", *([""] * (len(REQUIRED_COLUMNS) - 5))])
    workbook.save(path)


def test_registers_new_paths_enabled_with_observed_economy(tmp_path: Path) -> None:
    path = tmp_path / "exceptions.xlsx"
    _workbook(path)

    added = register_missing_branch_paths(
        ["Demand\\Other loss and own use\\Coal mines\\BKB and PB"],
        economy="20_USA", workbook_path=path,
    )

    assert added == ["Demand\\Other loss and own use\\Coal mines\\BKB and PB"]
    values = list(load_workbook(path)["branch_exceptions"].values)
    assert values[1][0] is True
    assert values[1][3] == "20_USA"


def test_loading_exception_notes_can_skip_materiality_refresh(tmp_path: Path, monkeypatch) -> None:
    path = tmp_path / "exceptions.xlsx"
    _workbook(path, "Demand\\Example")
    calls: list[Path] = []

    def fail_if_refreshed(workbook_path, **kwargs):
        calls.append(Path(workbook_path))
        raise AssertionError("reader unexpectedly refreshed the shared workbook")

    monkeypatch.setattr(
        "codebase.functions.baseline_seed_validation_exceptions.refresh_exception_materiality",
        fail_if_refreshed,
    )

    assert load_enabled_exception_notes(path, refresh_materiality=False) == {"Demand\\Example": ""}
    assert calls == []


def test_refresh_uses_last_esto_year_and_reference_ninth_average(tmp_path: Path) -> None:
    branch = "Demand\\Other loss and own use\\Coal mines\\BKB and PB"
    workbook = tmp_path / "exceptions.xlsx"
    _workbook(workbook, branch)
    esto = tmp_path / "esto.csv"
    pd.DataFrame([{
        "economy": "01AUS", "flows": "10.01.06 Coal mines", "products": "02.08 BKB/PB",
        "is_subtotal": False, "2022": -2.0,
    }]).to_csv(esto, index=False)
    ninth = tmp_path / "ninth.csv"
    pd.DataFrame([
        {"scenarios": scenario, "sectors": "x", "sub1sectors": "x", "sub2sectors": "10_01_06_coal_mines",
         "fuels": "02_coal_products", "subfuels": "02_08_bkb_pb", "subtotal_layout": False,
         "subtotal_results": False, "2023": value, "2024": value}
        for scenario, value in (("reference", -3.0), ("target", -9.0))
    ]).to_csv(ninth, index=False)
    rows = pd.read_excel(workbook).fillna("")
    rows.loc[0, "relevance_audit"] = "ESTO 2026: triggered for 20_USA."
    rows.loc[0, "zero filter"] = "MAPPING INCOMPLETE — seed triggered"
    rows.to_excel(workbook, sheet_name="branch_exceptions", index=False)

    refreshed = refresh_exception_materiality(
        workbook, esto_vintages={"2024": (esto, 2022)}, ninth_path=ninth,
        projection_start_year=2023, projection_final_year=2024,
        retry_mapping_incomplete=True,
    )

    assert refreshed.at[0, "esto_2024_last_year_signed_pj_all_economies"] == -2.0
    assert refreshed.at[0, NINTH_VALUE_COLUMN] == -3.0


def test_final_stage_registers_material_economies_and_prunes_only_after_all_vintages(tmp_path: Path) -> None:
    path = tmp_path / "exceptions.xlsx"
    branch = r"Demand\Other loss and own use\Coal mines\BKB and PB"
    _workbook(path, branch)
    register_material_missing_branch_findings(
        pd.DataFrame([{"economy": "20_USA", "branch_path": branch}]), workbook_path=path,
    )
    audit_exception_relevance(
        {"ESTO 2026": pd.DataFrame([{"economy": "20_USA", "branch_path": branch}])},
        workbook_path=path, prune_after_all_vintages=True,
    )
    row = pd.read_excel(path).iloc[0]
    assert row["enabled"]
    assert row["economies_that_need_it"] == "20_USA"
    assert "ESTO 2026: triggered for 20_USA" in row["notes"]

    audit_exception_relevance(
        {"ESTO 2026": pd.DataFrame(columns=["economy", "branch_path"])},
        workbook_path=path, prune_after_all_vintages=True,
    )
    row = pd.read_excel(path).iloc[0]
    assert not row["enabled"]
    assert pd.isna(row["economies_that_need_it"])  # blank Excel cell reads as NaN
    assert "Disabled after the completed all-vintage relevance audit" in row["notes"]


def test_zero_filter_blanks_false_mapping_zeros_when_seed_audit_triggered() -> None:
    rows = pd.DataFrame([{
        "enabled": True,
        "branch_path": r"Demand\Other\Fuel",
        "relevance_audit": "ESTO 2026: triggered for 20_USA.",
        "esto_2024_last_year_signed_pj_all_economies": 0.0,
        "esto_2025_last_year_signed_pj_all_economies": 0.0,
        "esto_2026_last_year_signed_pj_all_economies": 0.0,
        "ninth_reference_average_pj_per_year_all_economies": 0.0,
    }])

    filtered = apply_zero_filter(rows)

    assert filtered.loc[0, "zero filter"] == "MAPPING INCOMPLETE — seed triggered"
    assert filtered.loc[0, "esto_2024_last_year_signed_pj_all_economies"] == ""


def test_materiality_refresh_does_not_refill_a_seed_triggered_mapping_gap(tmp_path: Path) -> None:
    workbook = tmp_path / "exceptions.xlsx"
    _workbook(workbook, r"Demand\Other loss and own use\Coal mines\BKB and PB")
    rows = pd.read_excel(workbook).fillna("")
    rows.loc[0, "relevance_audit"] = "ESTO 2026: triggered for 20_USA."
    rows.loc[0, "zero filter"] = "MAPPING INCOMPLETE — seed triggered"
    rows.to_excel(workbook, sheet_name="branch_exceptions", index=False)

    refreshed = refresh_exception_materiality(
        workbook,
        esto_vintages={},
        ninth_path=tmp_path / "not_needed.csv",
    )

    assert refreshed.loc[0, "zero filter"] == "MAPPING INCOMPLETE — seed triggered"
    assert refreshed.loc[0, NINTH_VALUE_COLUMN] == ""


def test_canonical_single_axis_mapping_resolves_nonenergy_hydrogen() -> None:
    keys = _registry_source_keys([{
        "branch_path": r"Demand\All demand aggregated\Non Energy Use\Hydrogen",
    }])

    assert keys.loc[0, "ninth_sector"] == "17_nonenergy_use"
    assert keys.loc[0, "ninth_fuel"] == "16_x_hydrogen"
    assert keys.loc[0, "esto_flow"] == "17 Non-energy use"
    assert keys.loc[0, "esto_product"] == "16.12 Hydrogen"


def test_composes_unique_sector_and_fuel_axes_for_missing_interim_leaf(monkeypatch) -> None:
    path = r"Demand\Other loss and own use\Coal mines\Petroleum coke"
    empty_direct = pd.DataFrame(columns=[
        "leap_sector_name_full_path", "raw_leap_fuel_name", "esto_flow",
        "esto_product", "ninth_sector", "ninth_fuel",
    ])
    esto_axis = pd.DataFrame([
        {"leap_sector_name_full_path": "Other loss and own use/Coal mines", "raw_leap_fuel_name": "Anthracite", "esto_flow": "10.01.06 Coal mines", "esto_product": "01.04 Anthracite"},
        {"leap_sector_name_full_path": "Industry", "raw_leap_fuel_name": "Petroleum coke", "esto_flow": "14 Industry sector", "esto_product": "07.16 Petroleum coke"},
    ])
    ninth_axis = pd.DataFrame([
        {"leap_sector_name_full_path": "Other loss and own use/Coal mines", "raw_leap_fuel_name": "Anthracite", "ninth_sector": "10_01_06_coal_mines", "ninth_fuel": "01_x_thermal_coal"},
        {"leap_sector_name_full_path": "Industry", "raw_leap_fuel_name": "Petroleum coke", "ninth_sector": "14_industry_sector", "ninth_fuel": "07_x_other_petroleum_products"},
    ])
    monkeypatch.setattr(materiality, "_canonical_leaf_relationships", lambda: empty_direct)
    monkeypatch.setattr(materiality, "_canonical_axis_relationships", lambda: (esto_axis, ninth_axis))

    keys = _registry_source_keys([{"branch_path": path}])

    assert keys.loc[0, "esto_flow"] == "10.01.06 Coal mines"
    assert keys.loc[0, "esto_product"] == "07.16 Petroleum coke"
    assert keys.loc[0, "ninth_sector"] == "10_01_06_coal_mines"
    assert keys.loc[0, "ninth_fuel"] == "07_x_other_petroleum_products"


def test_rejects_flat_aggregated_demand_path_before_mapping() -> None:
    with pytest.raises(ValueError, match="needs a sector child"):
        _registry_source_keys([{
            "branch_path": r"Demand\All demand aggregated\Black liquor",
        }])


def test_normalises_transformation_grouping_nodes_to_canonical_process_path() -> None:
    parts = r"Transformation\CHP interim\Processes\CHP interim\Feedstock Fuels\Petroleum coke".split("\\")

    candidates = _source_sector_candidates(parts)

    assert candidates[0] == "CHP interim/CHP interim"


def test_uses_transfers_workflow_source_boundary(monkeypatch) -> None:
    path = r"Transformation\Transfers unallocated\Output Fuels\Natural gas liquids"
    esto_axis = pd.DataFrame([
        {"leap_sector_name_full_path": "Industry", "raw_leap_fuel_name": "Natural gas liquids", "esto_flow": "14 Industry sector", "esto_product": "06.02 Natural gas liquids"},
    ])
    ninth_axis = pd.DataFrame([
        {"leap_sector_name_full_path": "Industry", "raw_leap_fuel_name": "Natural gas liquids", "ninth_sector": "14_industry_sector", "ninth_fuel": "06_02_natural_gas_liquids"},
    ])
    monkeypatch.setattr(materiality, "_canonical_axis_relationships", lambda: (esto_axis, ninth_axis))

    keys = _registry_source_keys([{"branch_path": path}])

    assert keys.loc[0, "esto_flow"] == "08 Transfers"
    assert keys.loc[0, "ninth_sector"] == "08_transfers"
    assert keys.loc[0, "esto_product"] == "06.02 Natural gas liquids"
    assert keys.loc[0, "ninth_fuel"] == "06_02_natural_gas_liquids"


@pytest.mark.parametrize(
    ("path", "expected_esto_flow", "expected_ninth_sector", "fuel", "esto_product", "ninth_fuel"),
    [
        (
            r"Transformation\Coke ovens\Processes\Coke ovens\Feedstock Fuels\Coke oven coke",
            "09.08.01 Coke ovens",
            "09_08_01_coke_ovens",
            "Coke oven coke",
            "02.01 Coke oven coke",
            "02_01_coke_oven_coke",
        ),
        (
            r"Transformation\Gas to liquids plants\Output Fuels\Refinery gas not liquefied",
            "09.06.04 Gas-to-liquids plants",
            "09_06_04_gastoliquids_plants",
            "Refinery gas not liquefied",
            "07.10 Refinery gas not liquefied",
            "07_10_refinery_gas_not_liquefied",
        ),
    ],
)
def test_uses_detailed_transformation_process_boundary(
    monkeypatch,
    path: str,
    expected_esto_flow: str,
    expected_ninth_sector: str,
    fuel: str,
    esto_product: str,
    ninth_fuel: str,
) -> None:
    esto_axis = pd.DataFrame([{
        "leap_sector_name_full_path": "Industry",
        "raw_leap_fuel_name": fuel,
        "esto_flow": "14 Industry sector",
        "esto_product": esto_product,
    }])
    ninth_axis = pd.DataFrame([{
        "leap_sector_name_full_path": "Industry",
        "raw_leap_fuel_name": fuel,
        "ninth_sector": "14_industry_sector",
        "ninth_fuel": ninth_fuel,
    }])
    monkeypatch.setattr(materiality, "_canonical_axis_relationships", lambda: (esto_axis, ninth_axis))

    keys = _registry_source_keys([{"branch_path": path}])

    assert keys.loc[0, "esto_flow"] == expected_esto_flow
    assert keys.loc[0, "ninth_sector"] == expected_ninth_sector
    assert keys.loc[0, "esto_product"] == esto_product
    assert keys.loc[0, "ninth_fuel"] == ninth_fuel


def test_esto_materiality_uses_subtotal_only_when_no_detailed_match(tmp_path: Path) -> None:
    source = tmp_path / "esto.csv"
    pd.DataFrame([
        {"flows": "17 Non-energy use", "products": "08.03 Gas works gas", "is_subtotal": True, "2023": 43.0},
        {"flows": "17 Non-energy use", "products": "08.03 Gas works gas", "is_subtotal": False, "2023": 2.0},
    ]).to_csv(source, index=False)
    keys = pd.DataFrame([{
        "branch_path": r"Demand\All demand aggregated\Non Energy Use\Gas works gas",
        "esto_flow": "17 Non-energy use",
        "esto_product": "08.03 Gas works gas",
    }])

    values = _esto_base_materiality(keys, esto_path=source, base_year=2023)
    assert values[keys.loc[0, "branch_path"]][0] == 2.0

    pd.DataFrame([
        {"flows": "17 Non-energy use", "products": "08.03 Gas works gas", "is_subtotal": True, "2023": 43.0},
    ]).to_csv(source, index=False)
    values = _esto_base_materiality(keys, esto_path=source, base_year=2023)
    assert values[keys.loc[0, "branch_path"]][0] == 43.0
