"""Focused tests for the compressed results-update preflight.

The compressed results-update preflight (``supply_preflight.run_preflight_compressed_results_update``)
exercises the majority of the results_update path against the real 20_USA LEAP
balance-export structure compressed to two effective years. These tests cover the
building blocks (reduced REF/TGT balance workbooks, scenario-separated 9th
compression, the isolated config broadcast, and the deterministic issue report)
plus an opt-in end-to-end smoke test.

Fast unit tests run by default. Two groups are opt-in because they read the real
multi-hundred-MB source workbooks / 9th CSV:

* ``RUN_PREFLIGHT_RU_WORKBOOK=1`` -- build reduced workbooks from the real 20_USA
  REF/TGT source workbooks.
* ``RUN_PREFLIGHT_RU_SMOKE=1`` -- run the whole preflight orchestration with the
  heavy builders and the workflow runner stubbed out, to assert config routing
  and state restoration.
"""
from __future__ import annotations

import os
from pathlib import Path
from types import SimpleNamespace

import pandas as pd
import pytest

import codebase.functions.supply_preflight as sp


# --- Grid summation: signs, abs diagnostics, structural mismatch --------------


def _grid(header_fuels: list[str], rows: list[tuple[str, list[float]]]) -> list[list[object]]:
    """Build a minimal LEAP-balance-shaped grid (title, subtitle, header, data)."""
    grid: list[list[object]] = [
        ["Energy Balance", None] + [None] * len(header_fuels),
        ["Scenario: Reference", None] + [None] * len(header_fuels),
        [None] + list(header_fuels),
    ]
    for label, values in rows:
        grid.append([label] + list(values))
    return grid


def test_signed_values_preserved_when_summed() -> None:
    # Fix 3: signed sum, not absolute sum, for the synthetic future values.
    g1 = _grid(["Electricity", "Gas"], [("Production", [10.0, -5.0]), ("Imports", [2.0, 3.0])])
    g2 = _grid(["Electricity", "Gas"], [("Production", [-4.0, 5.0]), ("Imports", [1.0, -3.0])])
    signed, abs_grid, method = sp._sum_future_balance_grids([(2023, g1), (2024, g2)])
    # Production: Electricity 10 + -4 = 6 ; Gas -5 + 5 = 0 (perfect cancellation)
    assert signed[3][1] == 6.0
    assert signed[3][2] == 0.0
    # Structural cells preserved.
    assert signed[2][1] == "Electricity"
    assert signed[3][0] == "Production"
    assert method == "positional_identical"


def test_absolute_sum_diagnostic_exposes_cancellation() -> None:
    # Fix 4: the abs-sum diagnostic exposes a category zeroed by signed cancellation.
    g1 = _grid(["Gas"], [("Production", [10.0])])
    g2 = _grid(["Gas"], [("Production", [-10.0])])
    signed, abs_grid, _ = sp._sum_future_balance_grids([(2023, g1), (2024, g2)])
    assert signed[3][1] == 0.0  # signed cancels
    assert abs_grid[3][1] == 20.0  # abs-sum keeps it visible


def test_structural_mismatch_raises_clear_failure() -> None:
    # Fix 5: differing balance-row identities must fail, not corrupt positionally.
    g1 = _grid(["Gas"], [("Production", [1.0]), ("Imports", [1.0])])
    g_bad = _grid(["Gas"], [("Production", [1.0]), ("DIFFERENT ROW", [1.0])])
    with pytest.raises(ValueError, match="balance-row identities differ"):
        sp._sum_future_balance_grids([(2023, g1), (2024, g_bad)])


def test_varying_fuel_columns_union_aligned_not_dropped() -> None:
    # Point 8: a fuel present in only one year must not be dropped or corrupted.
    g1 = _grid(["Electricity", "Gas"], [("Production", [1.0, 1.0])])
    g2 = _grid(["Electricity", "Gas", "Coal"], [("Production", [2.0, 2.0, 7.0])])
    signed, _, method = sp._sum_future_balance_grids([(2023, g1), (2024, g2)])
    header = [sp._normalize_balance_grid_label(v) for v in signed[2]]
    assert "coal" in header
    coal_col = header.index("coal")
    assert signed[3][coal_col] == 7.0  # only present in g2, contributes 7
    assert method == "label_union_aligned"


# --- Reduced workbook build (real source workbooks; opt-in) -------------------

_WORKBOOK_ENABLED = os.environ.get("RUN_PREFLIGHT_RU_WORKBOOK") == "1"
_USA_DIR = Path(__file__).resolve().parents[1] / "data" / "leap balances exports" / "20_USA"


@pytest.mark.skipif(not _WORKBOOK_ENABLED, reason="reads real 20_USA workbooks; set RUN_PREFLIGHT_RU_WORKBOOK=1")
@pytest.mark.parametrize(
    "code, has_literal_2023",
    [("REF", True), ("TGT", False)],
)
def test_real_reduction_produces_base_and_synthetic_sheets(tmp_path, code, has_literal_2023) -> None:
    # Fix 1 (REF base+synthetic) and Fix 2 (TGT works without a literal EBal|2023).
    src = next(_USA_DIR.glob(f"*{code}.xlsx"))
    import openpyxl

    wb = openpyxl.load_workbook(src, read_only=True, data_only=True)
    literal_2023 = "EBal|2023" in wb.sheetnames
    wb.close()
    assert literal_2023 == has_literal_2023

    result = sp._build_reduced_preflight_balance_workbook(
        source_path=src,
        output_path=tmp_path / f"reduced_{code}.xlsx",
        base_year=2022,
        synthetic_year=2023,
        scenario_code=code,
        abs_diagnostic_path=tmp_path / f"reduced_{code}_abs.csv",
    )
    out = openpyxl.load_workbook(result["workbook_path"], read_only=True)
    assert out.sheetnames == ["EBal|2022", "EBal|2023"]
    out.close()
    assert result["abs_diagnostic_path"].exists()
    # The source workbook must be untouched.
    assert src.parent == _USA_DIR


# --- Scenario-separated 9th compression ---------------------------------------


def _fake_energy_source(tmp_path: Path) -> SimpleNamespace:
    esto = tmp_path / "esto.csv"
    esto.write_text("economy,flows,products,2022\n20USA,Production,Electricity,5\n", encoding="utf-8")
    ninth = tmp_path / "ninth.csv"
    rows = [
        # scenarios, economy, sectors, fuels, 2022, 2023, 2024
        ("reference", "20_USA", "14_industry", "electricity", 100, 3, 4),
        ("target", "20_USA", "14_industry", "electricity", 100, 30, 40),
        ("reference", "01_AUS", "14_industry", "electricity", 100, 999, 999),
    ]
    header = "scenarios,economy,sectors,fuels,2022,2023,2024\n"
    body = "\n".join(",".join(str(v) for v in r) for r in rows) + "\n"
    ninth.write_text(header + body, encoding="utf-8")
    return SimpleNamespace(
        esto_base_year=2022,
        esto_base_table_path=str(esto),
        ninth_projection_table_path=str(ninth),
    )


def test_scenario_separated_compression_keeps_reference_and_target_separate(tmp_path, monkeypatch) -> None:
    # Fix 6: Reference source rows compress into Reference, Target into Target;
    # never summed together and replicated.
    cfg = _fake_energy_source(tmp_path)
    monkeypatch.setattr(sp.workflow_cfg, "get_energy_source_config", lambda: cfg)

    result = sp._create_preflight_compressed_source_files(
        output_dir=tmp_path / "sources",
        scenario_names=["Reference", "Target"],
        preserve_source_scenarios=True,
        economy_filter=["20_USA"],
        file_prefix="test_ru",
    )
    ninth = pd.read_csv(result["ninth_path"])
    # Economy filter drops 01_AUS.
    assert set(ninth["economy"].unique()) == {"20_USA"}
    ref_val = ninth[ninth["scenarios"] == "reference"]["2023"].iloc[0]
    tgt_val = ninth[ninth["scenarios"] == "target"]["2023"].iloc[0]
    # Signed sum across future years within scenario: ref=3+4=7, tgt=30+40=70.
    assert ref_val == 7
    assert tgt_val == 70
    # Crucially not summed together (would be 77 in both).
    assert ref_val != tgt_val


def test_default_projection_compression_still_replicates_across_scenarios(tmp_path, monkeypatch) -> None:
    # Fix 14: existing projection-preflight behaviour intact -- the default mode
    # sums across scenarios and replicates the combined series into each scenario.
    cfg = _fake_energy_source(tmp_path)
    monkeypatch.setattr(sp.workflow_cfg, "get_energy_source_config", lambda: cfg)

    result = sp._create_preflight_compressed_source_files(
        output_dir=tmp_path / "sources",
        scenario_names=["Reference", "Target"],
    )
    ninth = pd.read_csv(result["ninth_path"])
    # All economies retained (no economy filter), and per group the value is the
    # same across the replicated scenarios.
    industry = ninth[(ninth["economy"] == "20_USA") & (ninth["sectors"] == "14_industry")]
    vals = set(industry["2023"].tolist())
    # reference(3+4=7) + target(30+40=70) summed across scenarios = 77, replicated.
    assert vals == {77}


# --- Balance-workbook resolution routing --------------------------------------


def test_temporary_workbook_resolution_used_by_balance_loader(monkeypatch, tmp_path) -> None:
    # Fix 7: load_balance_demand_inputs(..., allow_projection_only=False) resolves
    # the temporary reduced REF/TGT workbooks via the explicit path overrides.
    import codebase.functions.supply_demand_mapping as sdm

    ref = tmp_path / "reduced_REF.xlsx"
    tgt = tmp_path / "reduced_TGT.xlsx"
    monkeypatch.setattr(sdm, "DIRECT_DEMAND_PROJECTION_ECONOMY", "20_USA")
    monkeypatch.setattr(sdm, "BALANCE_DEMAND_REF_WORKBOOK_PATH", ref)
    monkeypatch.setattr(sdm, "BALANCE_DEMAND_TGT_WORKBOOK_PATH", tgt)

    resolved_ref, resolved_tgt = sdm._resolve_balance_demand_workbooks_for_economy("20_USA")
    assert resolved_ref == sdm._resolve(ref)
    assert resolved_tgt == sdm._resolve(tgt)


# --- Config broadcast: isolation + restoration --------------------------------


def test_config_broadcast_round_trips_across_modules() -> None:
    # Fixes 10/12 (mechanism): overrides reach every consuming module and are
    # restored exactly afterwards, so production state is untouched.
    import codebase.functions.supply_results_saver as srs
    import codebase.functions.supply_demand_mapping as sdm

    before_srs = srs.CAPACITY_UNMET_PASS_MODE
    before_sdm_econ = sdm.DIRECT_DEMAND_PROJECTION_ECONOMY

    snapshot = sp._broadcast_config_overrides(
        {"CAPACITY_UNMET_PASS_MODE": "results_update", "DIRECT_DEMAND_PROJECTION_ECONOMY": "20_USA"}
    )
    try:
        assert srs.CAPACITY_UNMET_PASS_MODE == "results_update"
        assert sdm.DIRECT_DEMAND_PROJECTION_ECONOMY == "20_USA"
    finally:
        sp._restore_config_overrides(snapshot)

    assert srs.CAPACITY_UNMET_PASS_MODE == before_srs
    assert sdm.DIRECT_DEMAND_PROJECTION_ECONOMY == before_sdm_econ


def test_broadcast_does_not_create_missing_attributes() -> None:
    snapshot = sp._broadcast_config_overrides({"__DEFINITELY_NOT_A_CONFIG_NAME__": 123})
    try:
        assert snapshot == []
    finally:
        sp._restore_config_overrides(snapshot)


# --- Deterministic issue report -----------------------------------------------


def test_zero_issues_writes_header_only_report_with_current_schema(tmp_path) -> None:
    # Fix 13: no issues -> a fresh header-only CSV using the current schema,
    # never a stale report; obsolete issue_fuel_is_do_not_use is not used.
    summary = sp._finalize_balance_demand_issue_report(checks_dir=tmp_path)
    report = tmp_path / sp.RESULTS_BALANCE_DEMAND_ISSUES_FILENAME
    assert report.exists()
    cols = list(pd.read_csv(report).columns)
    assert "issue_fuel_is_non_actionable" in cols
    assert "demand_relevant" in cols
    assert "demand_relevance_basis" in cols
    assert "issue_fuel_is_do_not_use" not in cols
    assert summary["total_issue_rows"] == 0
    assert summary["actionable_demand_issue_rows"] == 0


def test_issue_report_summary_counts_actionable_and_ignored(tmp_path) -> None:
    report = tmp_path / sp.RESULTS_BALANCE_DEMAND_ISSUES_FILENAME
    pd.DataFrame(
        [
            {
                "reason": "missing_esto_pair",
                "issue_sector_key": "Industry",
                "issue_fuel_key": "Electricity",
                "issue_fuel_is_non_actionable": False,
                "demand_relevant": True,
            },
            {
                "reason": "missing_esto_pair",
                "issue_sector_key": "Blast furnaces",
                "issue_fuel_key": "Total",
                "issue_fuel_is_non_actionable": False,
                "demand_relevant": False,
            },
        ]
    ).to_csv(report, index=False)

    summary = sp._finalize_balance_demand_issue_report(checks_dir=tmp_path)
    assert summary["total_issue_rows"] == 2
    assert summary["actionable_demand_issue_rows"] == 1
    assert summary["ignored_non_demand_rows"] == 1
    assert summary["total_fuel_rows"] == 1  # the 'Total' fuel row
    assert summary["unique_sector_fuel_keys"] == 2


# --- End-to-end orchestration (heavy builders + runner stubbed; opt-in) -------

_SMOKE_ENABLED = os.environ.get("RUN_PREFLIGHT_RU_SMOKE") == "1"


@pytest.mark.skipif(not _SMOKE_ENABLED, reason="runs preflight orchestration; set RUN_PREFLIGHT_RU_SMOKE=1")
@pytest.mark.parametrize("runner_fails", [False, True])
def test_results_update_preflight_routing_and_restoration(tmp_path, monkeypatch, runner_fails) -> None:
    # Fixes 8/9/10/11/12: forces 20_USA + results_update, disables cache and LEAP
    # imports, routes to the temporary workbooks, isolates outputs, and restores
    # production state after both success and failure.
    import codebase.supply_reconciliation_workflow as workflow
    import codebase.functions.supply_results_saver as srs
    import codebase.functions.supply_demand_mapping as sdm

    # Stub the heavy builders so the test does not read the real source workbooks.
    reduced_ref = tmp_path / "reduced_REF.xlsx"
    reduced_tgt = tmp_path / "reduced_TGT.xlsx"
    reduced_ref.write_bytes(b"")
    reduced_tgt.write_bytes(b"")
    monkeypatch.setattr(
        sp,
        "_build_reduced_preflight_balance_workbook",
        lambda **k: {
            "workbook_path": reduced_ref if k["scenario_code"] == "REF" else reduced_tgt,
            "sheet_names": ["EBal|2022", "EBal|2023"],
            "future_years": [2023, 2024],
            "sum_method": "label_union_aligned",
            "scenario_code": k["scenario_code"],
        },
    )
    monkeypatch.setattr(
        sp,
        "_create_preflight_compressed_source_files",
        lambda **k: {
            "esto_path": tmp_path / "esto.csv",
            "ninth_path": tmp_path / "ninth.csv",
            "ninth_abs_diagnostics_path": tmp_path / "ninth_abs.csv",
            "base_year": 2022,
            "compressed_year": 2023,
        },
    )

    observed: dict[str, object] = {}

    def _fake_runner(**kwargs):
        # Capture the config the pipeline actually sees at call time.
        observed["economies_arg"] = kwargs.get("economies")
        observed["pass_mode"] = srs.CAPACITY_UNMET_PASS_MODE
        observed["cache"] = srs.TRANSFORMATION_SUPPLY_CACHE_ENABLED
        observed["leap_supply"] = srs.LEAP_IMPORT_SUPPLY_TO_LEAP
        observed["econ"] = srs.ECONOMIES
        observed["ref_wb"] = sdm.BALANCE_DEMAND_REF_WORKBOOK_PATH
        observed["checks"] = srs.RESULTS_CHECKS_DIR
        observed["validation_years"] = workflow_cfg.get_baseline_seed_validation_years(
            ["Reference", "Target"]
        )
        if runner_fails:
            raise RuntimeError("simulated pipeline failure")
        return {}

    monkeypatch.setattr(workflow, "run_results_linked_transformation_supply_workflow", _fake_runner)

    prod_pass_mode = srs.CAPACITY_UNMET_PASS_MODE
    prod_econ = srs.ECONOMIES
    prod_ref_wb = sdm.BALANCE_DEMAND_REF_WORKBOOK_PATH

    if runner_fails:
        with pytest.raises(RuntimeError, match="simulated pipeline failure"):
            sp.run_preflight_compressed_results_update(scenario_names=["Reference", "Target"])
    else:
        sp.run_preflight_compressed_results_update(scenario_names=["Reference", "Target"])

    # Routing observed during the run.
    assert observed["economies_arg"] == ["20_USA"]
    assert observed["econ"] == ["20_USA"]
    assert observed["pass_mode"] == "results_update"
    assert observed["cache"] is False
    assert observed["leap_supply"] is False
    assert observed["ref_wb"] == reduced_ref
    assert "preflight_compressed_results_update" in str(observed["checks"])
    assert observed["validation_years"] == {
        "Reference": [2023],
        "Target": [2023],
    }

    # Production state restored after success AND failure.
    assert srs.CAPACITY_UNMET_PASS_MODE == prod_pass_mode
    assert srs.ECONOMIES == prod_econ
    assert sdm.BALANCE_DEMAND_REF_WORKBOOK_PATH == prod_ref_wb
