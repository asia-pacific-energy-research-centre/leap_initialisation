"""Acceptance coverage for the central post-write baseline-seed artifact gate."""

from __future__ import annotations

import json
from pathlib import Path

import openpyxl
import pandas as pd
import pytest

from codebase.functions.baseline_seed_artifact_validation import (
    CHECK_IDS,
    FINDING_COLUMNS,
    run_baseline_seed_artifact_validation,
)
from codebase.functions.baseline_seed_validation import validate_seed_rows
from codebase.functions.leap_excel_io import add_leap_preamble


ECONOMY = "20_USA"
REGION = "United States of America"
BASE_PATH = r"Resources\Primary\Coal"


def _row(
    *,
    branch_path: str = BASE_PATH,
    variable: str = "Production",
    scenario: str = "Reference",
    expression: str = "Data(2023,10)",
    branch_id: int = 10,
    variable_id: int = 20,
    scenario_id: int = 30,
    region_id: int = 1,
) -> dict[str, object]:
    return {
        "BranchID": branch_id,
        "VariableID": variable_id,
        "ScenarioID": scenario_id,
        "RegionID": region_id,
        "Branch Path": branch_path,
        "Variable": variable,
        "Scenario": scenario,
        "Region": REGION,
        "Scale": "",
        "Units": "PJ",
        "Per...": "",
        "Expression": expression,
        "Level 1": branch_path.split("\\")[0],
    }


def _write_template(path: Path, rows: pd.DataFrame) -> Path:
    export = add_leap_preamble(rows)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        export.to_excel(writer, sheet_name="Export", index=False, header=False)
    return path


def _viewing_rows(rows: pd.DataFrame, years: tuple[int, ...] = (2023,)) -> pd.DataFrame:
    viewing = rows.drop(columns=["Expression"], errors="ignore").copy()
    viewing["Method"] = "Data"
    for year in years:
        values = []
        for expression in rows["Expression"]:
            body = str(expression).replace("Data(", "").replace(")", "")
            tokens = [token.strip() for token in body.split(",")]
            by_year = {
                int(float(tokens[index])): float(tokens[index + 1])
                for index in range(0, len(tokens), 2)
            }
            values.append(by_year.get(year, pd.NA))
        viewing[str(year)] = values
    return viewing


def _write_candidate(
    path: Path,
    rows: pd.DataFrame,
    *,
    include_leap: bool = True,
    include_viewing: bool = True,
    damaged_preamble: bool = False,
) -> Path:
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        if include_leap:
            leap = add_leap_preamble(rows)
            if damaged_preamble:
                leap.iloc[0] = pd.NA
            leap.to_excel(writer, sheet_name="LEAP", index=False, header=False)
        if include_viewing:
            viewing = add_leap_preamble(_viewing_rows(rows))
            viewing.to_excel(writer, sheet_name="FOR_VIEWING", index=False, header=False)
    return path


def _write_producer_workbook(path: Path, rows: pd.DataFrame) -> Path:
    """Write a native standalone producer artifact (LEAP sheet only)."""
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        add_leap_preamble(rows).to_excel(
            writer, sheet_name="LEAP", index=False, header=False
        )
    return path


def _empty_zero_manifest() -> pd.DataFrame:
    return pd.DataFrame(columns=[*[
        "Branch Path", "Variable", "Scenario", "Region"
    ], "authorized", "source_workflow", "exception_id"])


def _run(
    tmp_path: Path,
    *,
    rows: pd.DataFrame | None = None,
    template_rows: pd.DataFrame | None = None,
    candidate: Path | None = None,
    expected_rows: pd.DataFrame | None | object = ...,  # Ellipsis means use rows.
    zero_manifest: pd.DataFrame | None | object = ...,
    required_diagnostics: list[Path] | None = None,
    expected_scenarios: list[str] | None = None,
    expected_years: dict[str, list[int]] | None = None,
    enforcement: dict[str, str] | None = None,
    validation_exceptions: list[dict[str, object]] | None = None,
    producer_workbooks: dict[str, dict[str, list[Path]]] | None = None,
    check_functions=None,
    output_name: str = "audit",
):
    rows = rows.copy() if rows is not None else pd.DataFrame([_row()])
    template_rows = template_rows.copy() if template_rows is not None else rows.copy()
    template = _write_template(tmp_path / f"template_{output_name}.xlsx", template_rows)
    candidate = candidate or _write_candidate(tmp_path / f"seed_{output_name}.xlsx", rows)
    diagnostic = tmp_path / "required_diagnostic.csv"
    if required_diagnostics is None:
        diagnostic.write_text("status\npass\n", encoding="utf-8")
        required_diagnostics = [diagnostic]
    resolved_expected = rows if expected_rows is ... else expected_rows
    resolved_zero = _empty_zero_manifest() if zero_manifest is ... else zero_manifest
    return run_baseline_seed_artifact_validation(
        run_id="RUN-001",
        candidate_workbooks={ECONOMY: candidate},
        expected_economies=[ECONOMY],
        template_paths_by_economy={ECONOMY: template},
        expected_scenarios=expected_scenarios or ["Reference"],
        expected_years_by_scenario=expected_years or {"Reference": [2023]},
        expected_producers=["supply_workflow"],
        producer_artifacts_by_producer={"supply_workflow": [candidate]},
        producer_workbooks_by_economy=producer_workbooks,
        source_rows_by_economy={} if resolved_expected is None else {ECONOMY: resolved_expected},
        zero_scope_manifests_by_economy={} if resolved_zero is None else {ECONOMY: resolved_zero},
        required_diagnostics=required_diagnostics,
        output_dir=tmp_path / output_name,
        enforcement_by_check=enforcement,
        validation_exceptions=validation_exceptions,
        check_functions=check_functions,
    )


def _failures(result, check_id: str) -> pd.DataFrame:
    return result.findings[
        result.findings["check_id"].eq(check_id)
        & result.findings["status"].isin({"FAIL", "INCOMPLETE", "CHECK_ERROR"})
    ]


def test_valid_complete_artifact_set_writes_shadow_pass_package(tmp_path: Path) -> None:
    result = _run(tmp_path)

    assert result.shadow_status == "SHADOW_PASS"
    assert result.accepted is True
    assert result.findings_path.exists()
    assert result.summary_path.exists()
    assert result.manifest_path.exists()
    assert set(result.findings["check_id"]) == set(CHECK_IDS)
    assert not result.findings["would_block"].any()


def test_missing_workbook_for_expected_economy_is_reported(tmp_path: Path) -> None:
    missing = tmp_path / "missing.xlsx"
    result = _run(tmp_path, candidate=missing)

    failure = _failures(result, "BSA-001").iloc[0]
    assert failure["contract_severity"] == "hard"
    assert failure["enforcement_mode"] == "audit"
    assert bool(failure["would_block"]) is True
    assert bool(failure["run_was_blocked"]) is False


def test_unreadable_workbook_is_not_a_pass(tmp_path: Path) -> None:
    bad = tmp_path / "bad.xlsx"
    bad.write_text("not an xlsx", encoding="utf-8")
    result = _run(tmp_path, candidate=bad)

    assert not _failures(result, "BSA-002").empty
    assert result.shadow_status == "SHADOW_INCOMPLETE"


@pytest.mark.parametrize("missing_sheet", ["LEAP", "FOR_VIEWING"])
def test_missing_required_sheet_is_reported(tmp_path: Path, missing_sheet: str) -> None:
    rows = pd.DataFrame([_row()])
    candidate = _write_candidate(
        tmp_path / f"missing_{missing_sheet}.xlsx",
        rows,
        include_leap=missing_sheet != "LEAP",
        include_viewing=missing_sheet != "FOR_VIEWING",
    )
    result = _run(tmp_path, rows=rows, candidate=candidate)

    assert _failures(result, "BSA-002")["actual"].str.contains(missing_sheet).any()


@pytest.mark.parametrize("damage", ["preamble", "column"])
def test_damaged_preamble_or_missing_required_column_is_reported(
    tmp_path: Path,
    damage: str,
) -> None:
    rows = pd.DataFrame([_row()])
    if damage == "column":
        rows = rows.drop(columns=["Units"])
    candidate = _write_candidate(
        tmp_path / f"damaged_{damage}.xlsx",
        rows,
        damaged_preamble=damage == "preamble",
    )
    result = _run(tmp_path, rows=rows, template_rows=pd.DataFrame([_row()]), candidate=candidate)

    assert not _failures(result, "BSA-002").empty


def test_duplicate_final_logical_keys_are_reported(tmp_path: Path) -> None:
    rows = pd.DataFrame([_row(), _row(expression="Data(2023,11)")])
    result = _run(tmp_path, rows=rows, template_rows=pd.DataFrame([_row()]))

    assert _failures(result, "BSA-003")["evidence"].str.contains("SEED-001|SEED-002").any()


def test_id_label_mismatch_against_target_template_is_reported(tmp_path: Path) -> None:
    rows = pd.DataFrame([_row(branch_id=999)])
    result = _run(tmp_path, rows=rows, template_rows=pd.DataFrame([_row()]))

    failure = _failures(result, "BSA-004")
    assert failure["expected"].str.contains('"BranchID": 10').any()
    assert failure["actual"].str.contains('"BranchID": 999').any()


def test_nonzero_unresolved_new_structure_is_prominent_but_nonblocking(tmp_path: Path) -> None:
    unknown = _row(branch_path=r"Resources\Primary\Unknown", branch_id=-1)
    rows = pd.DataFrame([unknown])
    result = _run(tmp_path, rows=rows, template_rows=pd.DataFrame([_row()]))

    warnings = result.findings[
        result.findings["check_id"].eq("BSA-005")
        & result.findings["status"].eq("WARN")
    ]
    assert warnings["evidence"].str.contains("SEED-004").any()
    assert set(warnings["structure_migration_classification"]) == {"new_migration_candidate"}
    assert warnings["would_block_without_migration_policy"].any()
    assert not warnings["would_block"].any()


def test_aggregate_demand_placeholder_warning_never_blocks(tmp_path: Path) -> None:
    """A retained missing LEAP branch stays visible without becoming a BSA failure."""
    placeholder = _row(
        branch_path=r"Demand\All demand aggregated\Industry\Geothermal",
        variable="Activity Level",
        expression="Data(2023,4.2703)",
        branch_id=-1,
    )
    rows = pd.DataFrame([placeholder])
    result = _run(
        tmp_path,
        rows=rows,
        template_rows=pd.DataFrame([_row()]),
        enforcement={"BSA-005": "block"},
    )

    warning = result.findings[
        result.findings["check_id"].eq("BSA-005")
        & result.findings["status"].eq("WARN")
    ]
    assert warning["evidence"].str.contains("SEED-004").any()
    assert not warning["would_block"].any()
    assert not warning["run_was_blocked"].any()
    assert result.accepted is True
    assert result.shadow_status == "SHADOW_WARN"


def test_standalone_chp_missing_petroleum_coke_is_new_migration_with_provenance(
    tmp_path: Path,
) -> None:
    branch_path = r"Transformation\CHP interim\Processes\CHP interim\Feedstock Fuels\Petroleum coke"
    producer = _write_producer_workbook(
        tmp_path / "electricity_heat_interim_20_USA_Target_Reference_Current_Accounts.xlsx",
        pd.DataFrame([_row(branch_path=branch_path, expression="Data(2023,4)", branch_id=-1, variable_id=-1, scenario_id=-1, region_id=-1)]),
    )
    result = _run(
        tmp_path,
        producer_workbooks={ECONOMY: {"electricity_heat_interim_workflow": [producer]}},
        enforcement={"BSA-004": "block", "BSA-005": "block"},
    )

    findings = result.findings[
        result.findings["workbook"].eq(str(producer.resolve()))
        & result.findings["branch_path"].eq(branch_path)
    ]
    assert set(findings["source_workflow"]) == {"electricity_heat_interim_workflow"}
    assert findings["evidence"].str.contains("SEED-003|SEED-004|SEED-011").any()
    assert set(findings["structure_migration_classification"]) == {"new_migration_candidate"}
    assert findings["would_block_without_migration_policy"].any()
    assert not findings["would_block"].any()
    assert result.accepted


def test_valid_standalone_producer_has_no_shared_seed_failures(tmp_path: Path) -> None:
    row = pd.DataFrame([_row()])
    producer = _write_producer_workbook(tmp_path / "supply_20_USA.xlsx", row)
    result = _run(
        tmp_path,
        producer_workbooks={ECONOMY: {"supply_workflow": [producer]}},
    )

    producer_failures = result.findings[
        result.findings["workbook"].eq(str(producer.resolve()))
        & result.findings["status"].isin({"FAIL", "CHECK_ERROR", "INCOMPLETE"})
    ]
    assert producer_failures.empty


def test_standalone_aggregate_placeholder_remains_warning_only(tmp_path: Path) -> None:
    placeholder = pd.DataFrame([_row(
        branch_path=r"Demand\All demand aggregated\Industry\Geothermal",
        variable="Activity Level",
        expression="Data(2023,4)",
        branch_id=-1,
    )])
    producer = _write_producer_workbook(tmp_path / "aggregated_demand_20_USA.xlsx", placeholder)
    result = _run(
        tmp_path,
        producer_workbooks={ECONOMY: {"aggregated_demand_workflow": [producer]}},
        enforcement={"BSA-004": "block", "BSA-005": "block"},
    )

    warnings = result.findings[result.findings["workbook"].eq(str(producer.resolve()))]
    assert "WARN" in set(warnings["status"])
    assert not warnings["would_block"].any()


def test_zero_only_unmatched_transformation_row_remains_nonblocking(tmp_path: Path) -> None:
    zero_row = pd.DataFrame([_row(
        branch_path=r"Resources\Primary\Unused fuel",
        variable="Imports",
        expression="Data(2023,0)",
        branch_id=-1,
        variable_id=-1,
        scenario_id=-1,
        region_id=-1,
    )])
    producer = _write_producer_workbook(tmp_path / "transformation_20_USA.xlsx", zero_row)
    result = _run(
        tmp_path,
        producer_workbooks={ECONOMY: {"transformation_workflow": [producer]}},
    )

    producer_findings = result.findings[result.findings["workbook"].eq(str(producer.resolve()))]
    assert not producer_findings["would_block"].any()


def test_multiple_standalone_producers_keep_workflow_and_workbook_provenance(tmp_path: Path) -> None:
    supply = _write_producer_workbook(tmp_path / "supply_20_USA.xlsx", pd.DataFrame([_row()]))
    transfers = _write_producer_workbook(tmp_path / "transfers_20_USA.xlsx", pd.DataFrame([_row()]))
    result = _run(
        tmp_path,
        producer_workbooks={ECONOMY: {
            "supply_workflow": [supply],
            "transfers_workflow": [transfers],
        }},
    )

    producer_passes = result.findings[
        result.findings["workbook"].isin([str(supply.resolve()), str(transfers.resolve())])
        & result.findings["status"].eq("PASS")
    ]
    assert set(producer_passes["source_workflow"]) == {
        "supply_workflow", "transfers_workflow"
    }


def _share_rows(values: tuple[float, ...], *, include_second: bool = True) -> pd.DataFrame:
    paths = [
        r"Transformation\Plant\Output Fuels\Coal",
        r"Transformation\Plant\Output Fuels\Gas",
    ]
    rows = []
    for index, value in enumerate(values):
        if index == 1 and not include_second:
            continue
        rows.append(_row(
            branch_path=paths[index],
            variable="Output Share",
            expression=f"Data(2023,{value})",
            branch_id=100 + index,
            variable_id=200 + index,
        ))
    return pd.DataFrame(rows)


def test_incomplete_share_siblings_are_reported(tmp_path: Path) -> None:
    template = _share_rows((50, 50))
    rows = _share_rows((100, 0), include_second=False)
    result = _run(tmp_path, rows=rows, template_rows=template)

    assert not _failures(result, "BSA-006").empty


def test_final_share_sum_other_than_100_is_reported(tmp_path: Path) -> None:
    template = _share_rows((50, 50))
    rows = _share_rows((60, 30))
    result = _run(tmp_path, rows=rows, template_rows=template)

    assert _failures(result, "BSA-006")["evidence"].str.contains("SEED-006").any()


@pytest.mark.parametrize("missing_kind", ["scenario", "year"])
def test_missing_scenario_or_required_year_is_reported(tmp_path: Path, missing_kind: str) -> None:
    if missing_kind == "scenario":
        expected_scenarios = ["Reference", "Target"]
        expected_years = {"Reference": [2023], "Target": [2023]}
    else:
        expected_scenarios = ["Reference"]
        expected_years = {"Reference": [2023, 2024]}
    result = _run(
        tmp_path,
        expected_scenarios=expected_scenarios,
        expected_years=expected_years,
    )

    assert not _failures(result, "BSA-007").empty


def test_unauthorized_zero_row_is_reported(tmp_path: Path) -> None:
    rows = pd.DataFrame([_row(expression="Data(2023,0)")])
    zero_manifest = rows[["Branch Path", "Variable", "Scenario", "Region"]].copy()
    zero_manifest["authorized"] = False
    zero_manifest["source_workflow"] = "demand_zeroing_workflow"
    result = _run(tmp_path, rows=rows, zero_manifest=zero_manifest)

    assert _failures(result, "BSA-008")["actual"].eq("authorized=false").any()


def test_source_to_serialized_value_conservation_loss_is_reported(tmp_path: Path) -> None:
    rows = pd.DataFrame([_row(expression="Data(2023,10)")])
    expected = pd.DataFrame([_row(expression="Data(2023,12)")])
    result = _run(tmp_path, rows=rows, expected_rows=expected)

    failure = _failures(result, "BSA-009")
    assert failure["actual"].str.contains("LEAP=10.0").any()


def test_serialized_value_conservation_accepts_numeric_viewing_year_header(
    tmp_path: Path,
) -> None:
    rows = pd.DataFrame([_row(expression="Data(2023,10)")])
    candidate = _write_candidate(tmp_path / "numeric_viewing_year.xlsx", rows)
    workbook = openpyxl.load_workbook(candidate)
    viewing = workbook["FOR_VIEWING"]
    for cell in viewing[3]:
        if str(cell.value) == "2023":
            cell.value = 2023.0
            break
    workbook.save(candidate)

    result = _run(tmp_path, rows=rows, candidate=candidate)

    assert _failures(result, "BSA-009").empty


def test_missing_required_diagnostic_is_incomplete(tmp_path: Path) -> None:
    result = _run(tmp_path, required_diagnostics=[tmp_path / "missing.csv"])

    assert not _failures(result, "BSA-010").empty
    assert result.manifest["missing_diagnostics"]
    assert result.shadow_status == "SHADOW_INCOMPLETE"


def test_unexpected_check_exception_produces_shadow_incomplete(tmp_path: Path) -> None:
    def explode(**_kwargs):
        raise RuntimeError("injected validator defect")

    result = _run(tmp_path, check_functions={"BSA-004": explode})

    failure = _failures(result, "BSA-004").iloc[0]
    assert failure["status"] == "CHECK_ERROR"
    assert "injected validator defect" in failure["evidence"]
    assert result.shadow_status == "SHADOW_INCOMPLETE"


def test_audit_mode_hard_findings_allow_run_to_finish(tmp_path: Path) -> None:
    result = _run(tmp_path, candidate=tmp_path / "missing.xlsx")

    assert result.accepted is True
    assert result.manifest["would_block_count"] > 0
    assert result.manifest["run_was_blocked_count"] == 0


def test_block_mode_unit_behavior_prevents_acceptance(tmp_path: Path) -> None:
    result = _run(
        tmp_path,
        candidate=tmp_path / "missing.xlsx",
        enforcement={"BSA-001": "block"},
    )

    assert result.accepted is False
    assert _failures(result, "BSA-001")["run_was_blocked"].all()


def test_post_write_corruption_is_detected_after_prewrite_rows_pass(tmp_path: Path) -> None:
    rows = pd.DataFrame([_row(expression="Data(2023,10)")])
    template = _write_template(tmp_path / "template_corrupt.xlsx", rows)
    assert validate_seed_rows(
        rows,
        template_path=template,
        required_years_by_scenario={"Reference": [2023]},
        required_scenarios=["Reference"],
    ).blocking_findings.empty
    candidate = _write_candidate(tmp_path / "corrupt.xlsx", rows)
    workbook = openpyxl.load_workbook(candidate)
    sheet = workbook["LEAP"]
    headers = {cell.value: cell.column for cell in sheet[3]}
    sheet.cell(row=4, column=headers["Expression"], value="Data(2023,99)")
    workbook.save(candidate)

    result = run_baseline_seed_artifact_validation(
        run_id="RUN-CORRUPT",
        candidate_workbooks={ECONOMY: candidate},
        expected_economies=[ECONOMY],
        template_paths_by_economy={ECONOMY: template},
        expected_scenarios=["Reference"],
        expected_years_by_scenario={"Reference": [2023]},
        source_rows_by_economy={ECONOMY: rows},
        zero_scope_manifests_by_economy={ECONOMY: _empty_zero_manifest()},
        output_dir=tmp_path / "corrupt_audit",
    )

    assert not _failures(result, "BSA-009").empty


def test_shared_validator_produces_identical_local_and_final_share_rule(tmp_path: Path) -> None:
    template_rows = _share_rows((50, 50))
    rows = _share_rows((60, 30))
    template = _write_template(tmp_path / "template_parity.xlsx", template_rows)
    local = validate_seed_rows(
        rows,
        template_path=template,
        required_years_by_scenario={"Reference": [2023]},
        required_scenarios=["Reference"],
    )
    local_rules = set(local.blocking_findings["rule_id"])
    result = _run(tmp_path, rows=rows, template_rows=template_rows)
    final_rules = {
        evidence.split(":", 1)[0]
        for evidence in _failures(result, "BSA-006")["evidence"]
    }

    assert "SEED-006" in local_rules
    assert final_rules == (local_rules & {"SEED-006", "SEED-007", "SEED-008"})


def test_manifest_is_complete_and_deterministic(tmp_path: Path) -> None:
    rows = pd.DataFrame([_row()])
    candidate = _write_candidate(tmp_path / "seed_deterministic.xlsx", rows)
    template = _write_template(tmp_path / "template_deterministic.xlsx", rows)
    diagnostic = tmp_path / "deterministic_diagnostic.csv"
    diagnostic.write_text("status\npass\n", encoding="utf-8")

    def run(output_name: str):
        return run_baseline_seed_artifact_validation(
            run_id="RUN-DETERMINISTIC",
            candidate_workbooks={ECONOMY: candidate},
            expected_economies=[ECONOMY],
            template_paths_by_economy={ECONOMY: template},
            expected_scenarios=["Reference"],
            expected_years_by_scenario={"Reference": [2023]},
            expected_producers=["supply_workflow"],
            producer_artifacts_by_producer={"supply_workflow": [candidate]},
            source_rows_by_economy={ECONOMY: rows},
            zero_scope_manifests_by_economy={ECONOMY: _empty_zero_manifest()},
            required_diagnostics=[diagnostic],
            output_dir=tmp_path / output_name,
        )

    first = run("audit_one")
    second = run("audit_two")

    first_manifest = json.loads(first.manifest_path.read_text(encoding="utf-8"))
    second_manifest = json.loads(second.manifest_path.read_text(encoding="utf-8"))
    assert first_manifest == second_manifest
    assert first.findings_path.read_text(encoding="utf-8") == second.findings_path.read_text(encoding="utf-8")
    assert first.summary_path.read_text(encoding="utf-8") == second.summary_path.read_text(encoding="utf-8")
    assert set(first.findings.columns) == set(FINDING_COLUMNS)
    assert {item["check_id"] for item in first.manifest["configured_checks"]} == set(CHECK_IDS)


def test_manifest_records_applied_seed_exceptions(tmp_path: Path) -> None:
    unknown_path = r"Resources\Primary\Unknown"
    rows = pd.DataFrame([_row(branch_path=unknown_path, branch_id=-1)])
    result = _run(
        tmp_path,
        rows=rows,
        template_rows=pd.DataFrame([_row()]),
        validation_exceptions=[{
            "rule_id": "SEED-004",
            "Branch Path": unknown_path,
            "exception_id": "EX-REVIEWED-001",
            "reason": "Focused manifest-recording fixture.",
        }],
    )

    assert "EX-REVIEWED-001" in result.manifest["applied_exceptions"]
    excepted = result.findings[
        result.findings["check_id"].eq("BSA-005")
        & result.findings["status"].eq("EXCEPTED")
    ]
    assert not excepted.empty
