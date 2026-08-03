from __future__ import annotations

import csv
import hashlib
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from codebase.functions.balance_review_workbook_builder import (
    BLUE_FILL,
    ERROR_SHEET_NAME,
    FULL_EXPECTED_SHEET_NAME,
    LEAP_SHEET_NAME,
    MISSING_SHEET_NAME,
    NO_COMPARATOR_FILL,
    RED_FILL,
    build_balance_structure_review_workbook,
)
from codebase.functions.balance_review_workbooks import (
    build_balance_review_workbooks,
)
from codebase.utilities.leap_balance_export_resolver import BalanceExportSheet


def _sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def _write_csv(path: Path, rows: list[dict[str, object]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = list(rows[0])
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def _write_source_workbook(path: Path, *, units: str = "Petajoule") -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "2022"
    sheet["A1"] = 'Energy Balance for Area "Test area"'
    sheet["A2"] = f"Scenario: Target, Year: 2022, Units: {units}"
    sheet.append(["", "", ""])
    sheet.append(["Production", 10, 2])
    sheet.append(["Imports", 1, 5])
    sheet.append(["Exports", -2, -1])
    sheet.append(["  Process", -3, -2])
    sheet["A3"] = "Balance row"
    sheet["B3"] = "Electricity"
    sheet["C3"] = "Natural gas"
    sheet["A3"].font = Font(bold=True, color="FFFFFF")
    sheet["A3"].fill = PatternFill("solid", fgColor="1F4E78")
    sheet.column_dimensions["A"].width = 24
    sheet.column_dimensions["B"].width = 14
    sheet.freeze_panes = "B4"
    workbook.create_sheet("Unused 2023")
    workbook.save(path)


def _write_diagnostics(directory: Path) -> None:
    base_rows = [
        {
            "economy": "20_USA",
            "scenario": "Target",
            "year": 2022,
            "esto_flow": "01 Production",
            "esto_product": "17 Electricity",
            "leap_sector_names": "Production",
            "leap_fuel_names": "Electricity",
            "leap_value_pj": 10,
            "source_value_pj": 8,
            "difference_pj": 2,
            "status": "value_mismatch",
            "leap_balance_row": "Production",
            "leap_balance_fuel": "Electricity",
            "no_direct_projection_comparator": False,
            "primary_classification": "approved_results_update",
            "evidence_note": "",
            "next_action": "",
            "preliminary_owner": "supply",
            "fuel_label": "Electricity",
        },
        {
            "economy": "20_USA",
            "scenario": "Target",
            "year": 2022,
            "esto_flow": "02 Imports",
            "esto_product": "08.01 Natural gas",
            "leap_sector_names": "Imports",
            "leap_fuel_names": "Natural gas",
            "leap_value_pj": 5,
            "source_value_pj": 5,
            "difference_pj": 0,
            "status": "match",
            "leap_balance_row": "Imports",
            "leap_balance_fuel": "Natural gas",
            "no_direct_projection_comparator": False,
            "primary_classification": "match",
            "evidence_note": "",
            "next_action": "",
            "preliminary_owner": "supply",
            "fuel_label": "Natural gas",
        },
        {
            "economy": "20_USA",
            "scenario": "Target",
            "year": 2022,
            "esto_flow": "09 Transformation",
            "esto_product": "17 Electricity",
            "leap_sector_names": "Transformation/Process",
            "leap_fuel_names": "Electricity",
            "leap_value_pj": -3,
            "source_value_pj": "",
            "difference_pj": "",
            "status": "reference_unavailable",
            "leap_balance_row": "Transformation/Process",
            "leap_balance_fuel": "Electricity",
            "no_direct_projection_comparator": True,
            "primary_classification": "seed_carry_forward",
            "evidence_note": "No direct comparator.",
            "next_action": "Keep the seed value.",
            "preliminary_owner": "transformation",
            "fuel_label": "Electricity",
        },
    ]
    _write_csv(directory / "leap_balance_source_review.csv", base_rows)
    difference_fields = [
        "economy",
        "scenario",
        "year",
        "esto_flow",
        "esto_product",
        "leap_sector_names",
        "leap_fuel_names",
        "leap_value_pj",
        "source_value_pj",
        "difference_pj",
        "status",
    ]
    _write_csv(
        directory / "leap_balance_source_differences.csv",
        [{field: row[field] for field in difference_fields} for row in base_rows],
    )
    _write_csv(
        directory / "leap_balance_mapping_issues.csv",
        [
            {
                "reason": "missing_esto_pair",
                "details": "Mapping is absent.",
                "scenario": "Target",
                "year": 2022,
                "leap_sector_name_full_path": "Imports",
                "leap_flow": "Imports",
                "leap_flow_name": "Imports",
                "leap_product": "Electricity",
                "leap_product_name": "Electricity",
                "value_petajoule": 1,
                "economy": "20_USA",
            }
        ],
    )


def test_python_builder_preserves_layout_and_writes_diagnostics(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    diagnostics = tmp_path / "diagnostics"
    output = tmp_path / "review.xlsx"
    _write_source_workbook(source)
    _write_diagnostics(diagnostics)
    source_hash = _sha256(source)

    result = build_balance_structure_review_workbook(
        economy="20_USA",
        source_workbook=source,
        source_sheet_name="2022",
        diagnostics_directory=diagnostics,
        output_workbook=output,
    )

    assert _sha256(source) == source_hash
    assert result["formulaErrorCells"] == []
    assert result["comparisonStateCounts"]["mapped"] == 2
    workbook = load_workbook(output, data_only=False)
    assert workbook.sheetnames == [
        LEAP_SHEET_NAME,
        ERROR_SHEET_NAME,
        "Correct Source Values",
        FULL_EXPECTED_SHEET_NAME,
        MISSING_SHEET_NAME,
    ]
    assert workbook[LEAP_SHEET_NAME].freeze_panes == "B4"
    assert workbook[ERROR_SHEET_NAME]["B4"].value == 2
    assert workbook[ERROR_SHEET_NAME]["B4"].fill.fgColor.rgb.endswith(RED_FILL)
    assert workbook["Correct Source Values"]["B4"].value == (
        "='LEAP Values'!B4-'LEAP - Source Error'!B4"
    )
    assert workbook["Correct Source Values"]["C5"].value == 5
    assert workbook["Correct Source Values"]["C5"].fill.fgColor.rgb.endswith(
        BLUE_FILL
    )
    assert workbook[LEAP_SHEET_NAME]["B7"].fill.fgColor.rgb.endswith(
        NO_COMPARATOR_FILL
    )
    assert "MissingCombinationsTable" in workbook[MISSING_SHEET_NAME].tables
    workbook.close()


def test_python_builder_normalizes_thousand_petajoule(tmp_path: Path) -> None:
    source = tmp_path / "source_thousand_pj.xlsx"
    diagnostics = tmp_path / "diagnostics"
    output = tmp_path / "review.xlsx"
    _write_source_workbook(source, units="Thousand Petajoule")
    _write_diagnostics(diagnostics)

    result = build_balance_structure_review_workbook(
        economy="20_USA",
        source_workbook=source,
        source_sheet_name="2022",
        diagnostics_directory=diagnostics,
        output_workbook=output,
    )

    workbook = load_workbook(output, data_only=False)
    assert workbook[LEAP_SHEET_NAME]["B4"].value == 10000
    assert "Units: Petajoule" in workbook[LEAP_SHEET_NAME]["A2"].value
    assert result["metadata"]["sourceUnitMultiplier"] == 1000
    workbook.close()


def test_public_builder_no_longer_requires_node(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    diagnostics = tmp_path / "diagnostics"
    _write_source_workbook(source)
    _write_diagnostics(diagnostics)
    selected = BalanceExportSheet(
        path=source,
        sheet_name="2022",
        scenario="Target",
        scenario_code="TGT",
        year=2022,
        units="Petajoule",
    )

    result = build_balance_review_workbooks(
        diagnostic_results={
            "20_USA": {"selected_balance_sheets": [selected]}
        },
        diagnostics_directory=diagnostics,
        output_directory=tmp_path / "comparison_workbooks",
    )

    assert Path(result[0]["outputWorkbook"]).exists()
