from zipfile import ZIP_DEFLATED, ZipFile
from types import SimpleNamespace
import re

import pytest
from openpyxl import Workbook, load_workbook

import codebase.mapping_tools.add_validation_exception_template_rows as placeholder_module
from codebase.mapping_tools.add_validation_exception_template_rows import (
    PLACEHOLDER_BRANCH_ID,
    SHEET_XML_PATH,
    insert_missing_exception_rows,
    sync_exception_resolution_status,
    validate_exception_placeholder_rows,
    apply_material_exception_placeholders,
    preview_material_exception_placeholder_rows,
)


def _write_template(
    path,
    sibling_profiles: dict[str, list[tuple[str, str, str]]],
    worksheet_attributes: str = "",
) -> None:
    headers = (
        '<row r="3"><c r="A3" t="inlineStr"><is><t>BranchID</t></is></c>'
        '<c r="E3" t="inlineStr"><is><t>Branch Path</t></is></c>'
        '<c r="F3" t="inlineStr"><is><t>Variable</t></is></c>'
        '<c r="G3" t="inlineStr"><is><t>Scenario</t></is></c>'
        '<c r="H3" t="inlineStr"><is><t>Region</t></is></c>'
        '<c r="I3" t="inlineStr"><is><t>Level 1</t></is></c>'
        '<c r="J3" t="inlineStr"><is><t>Level 2</t></is></c>'
        '<c r="K3" t="inlineStr"><is><t>Level 3</t></is></c></row>'
    )
    rows = []
    row_number = 4
    for branch_path, profile in sibling_profiles.items():
        for variable, scenario, region in profile:
            levels = branch_path.split("\\")
            rows.append(
                '<row r="{row}"><c r="A{row}" t="n"><v>1200</v></c>'
                '<c r="E{row}" t="inlineStr"><is><t>{path}</t></is></c>'
                '<c r="F{row}" t="inlineStr"><is><t>{variable}</t></is></c>'
                '<c r="G{row}" t="inlineStr"><is><t>{scenario}</t></is></c>'
                '<c r="H{row}" t="inlineStr"><is><t>{region}</t></is></c>'
                '<c r="I{row}" t="inlineStr"><is><t>{level1}</t></is></c>'
                '<c r="J{row}" t="inlineStr"><is><t>{level2}</t></is></c>'
                '<c r="K{row}" t="inlineStr"><is><t>{level3}</t></is></c></row>'.format(
                    row=row_number, path=branch_path, variable=variable, scenario=scenario,
                    region=region, level1=levels[0], level2=levels[1], level3=levels[2],
                )
            )
            row_number += 1
    xml = '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"{}><sheetData>{}</sheetData></worksheet>'.format(
        worksheet_attributes,
        headers + "".join(rows),
    )
    with ZipFile(path, "w", ZIP_DEFLATED) as archive:
        archive.writestr(SHEET_XML_PATH, xml)


def test_inserts_complete_largest_sibling_profile_with_branch_id_99(tmp_path):
    template_path = tmp_path / "template.xlsx"
    target_path = r"Demand\Other\Missing fuel"
    profile = [("Activity Level", "Reference", "Test"), ("Activity Level", "Target", "Test")]
    _write_template(template_path, {
        r"Demand\Other\Existing fuel": profile,
        r"Demand\Other\Sparse fuel": profile[:1],
    })

    result = insert_missing_exception_rows(template_path, {target_path}, apply_changes=True)

    assert result == {"template": "template.xlsx", "rows_added": 2}
    validate_exception_placeholder_rows(template_path, {target_path})
    with ZipFile(template_path) as archive:
        xml = archive.read(SHEET_XML_PATH).decode("utf-8")
    assert xml.count(target_path) == 2
    assert f">{PLACEHOLDER_BRANCH_ID}<" in xml


def test_refuses_ambiguous_largest_sibling_profiles(tmp_path):
    template_path = tmp_path / "template.xlsx"
    _write_template(template_path, {
        r"Demand\Other\Existing fuel": [("Activity Level", "Reference", "Test")],
        r"Demand\Other\Different fuel": [("Energy Intensity", "Reference", "Test")],
    })

    with pytest.raises(ValueError, match="Ambiguous sibling profiles"):
        insert_missing_exception_rows(template_path, {r"Demand\Other\Missing fuel"})


def test_insertion_preserves_excel_namespace_declarations(tmp_path):
    template_path = tmp_path / "template.xlsx"
    _write_template(
        template_path,
        {r"Demand\Other\Existing fuel": [("Activity Level", "Reference", "Test")]},
        ' xmlns:mc="http://schemas.openxmlformats.org/markup-compatibility/2006" '
        'xmlns:x14ac="http://schemas.microsoft.com/office/spreadsheetml/2009/9/ac" '
        'mc:Ignorable="x14ac"',
    )

    insert_missing_exception_rows(
        template_path,
        {r"Demand\Other\Missing fuel"},
        apply_changes=True,
    )

    with ZipFile(template_path) as archive:
        xml = archive.read(SHEET_XML_PATH).decode("utf-8")
    assert 'xmlns:mc="http://schemas.openxmlformats.org/markup-compatibility/2006"' in xml
    assert 'xmlns:x14ac="http://schemas.microsoft.com/office/spreadsheetml/2009/9/ac"' in xml
    assert 'mc:Ignorable="x14ac"' in xml


def test_insertion_preserves_prefixed_export_sheet_xml(tmp_path):
    template_path = tmp_path / "template.xlsx"
    _write_template(
        template_path,
        {r"Demand\Other\Existing fuel": [("Activity Level", "Reference", "Test")]},
    )
    with ZipFile(template_path) as archive:
        xml = archive.read(SHEET_XML_PATH).decode("utf-8")
        contents = {
            item.filename: archive.read(item.filename)
            for item in archive.infolist()
        }
    xml = re.sub(
        r"<(/?)(worksheet|sheetData|row|c|v|is|t)(?=[\s>/])",
        r"<\1s:\2",
        xml,
    ).replace(
        'xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"',
        'xmlns:s="http://schemas.openxmlformats.org/spreadsheetml/2006/main"',
    )
    with ZipFile(template_path, "w", ZIP_DEFLATED) as archive:
        for filename, content in contents.items():
            archive.writestr(filename, xml.encode("utf-8") if filename == SHEET_XML_PATH else content)

    insert_missing_exception_rows(
        template_path,
        {r"Demand\Other\Missing fuel"},
        apply_changes=True,
    )

    with ZipFile(template_path) as archive:
        updated_xml = archive.read(SHEET_XML_PATH).decode("utf-8")
    assert '<s:c r="A' in updated_xml
    assert "<s:v>99</s:v>" in updated_xml


def test_sync_records_per_economy_coverage_without_disabling_rows(tmp_path, monkeypatch):
    template_paths = [tmp_path / "one.xlsx", tmp_path / "two.xlsx"]
    real_path = r"Demand\Other\Existing fuel"
    placeholder_path = r"Demand\Other\Missing fuel"
    profile = [("Activity Level", "Reference", "Test")]
    for template_path in template_paths:
        _write_template(template_path, {real_path: profile})
        insert_missing_exception_rows(template_path, {placeholder_path}, apply_changes=True)

    exception_path = tmp_path / "exceptions.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "branch_exceptions"
    sheet.append(["enabled", "branch_path", "notes"])
    sheet.append([True, real_path, "Already real everywhere."])
    sheet.append([True, placeholder_path, "Still represented by ID 99."])
    workbook.save(exception_path)
    monkeypatch.setattr(
        placeholder_module.leap_export_template_resolver,
        "iter_leap_export_templates",
        lambda _root: [SimpleNamespace(path=path, economy=f"0{index}_TST") for index, path in enumerate(template_paths, start=1)],
    )

    status = sync_exception_resolution_status(
        exception_workbook_path=exception_path,
        templates_root=tmp_path,
    )

    assert status == {real_path: True, placeholder_path: False}
    values = list(load_workbook(exception_path)["branch_exceptions"].values)
    assert values[0][:5] == (
        "enabled", "branch_path", "notes", "economies_that_need_it",
        "economies_resolved_in_templates",
    )
    assert values[1][0] is True and values[1][4] == "all"
    # Template coverage does not invent materiality: this list is populated by
    # an observed material missing-path finding, not every template gap.
    assert values[2][0] is True and values[2][3] is None


def test_material_placeholder_operation_targets_every_economy(tmp_path, monkeypatch):
    template_paths = [tmp_path / "one.xlsx", tmp_path / "two.xlsx"]
    target_path = r"Demand\Other\Missing fuel"
    profile = [("Activity Level", "Reference", "Test")]
    for template_path in template_paths:
        _write_template(template_path, {r"Demand\Other\Existing fuel": profile})

    exception_path = tmp_path / "exceptions.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "branch_exceptions"
    sheet.append(["enabled", "branch_path", "notes", "economies_that_need_it"])
    sheet.append([True, target_path, "Material test path.", "01_TST"])
    workbook.save(exception_path)
    monkeypatch.setattr(
        placeholder_module.leap_export_template_resolver,
        "iter_leap_export_templates",
        lambda _root: [SimpleNamespace(path=path, economy=f"0{index}_TST") for index, path in enumerate(template_paths, start=1)],
    )

    plan = apply_material_exception_placeholders(
        exception_workbook_path=exception_path, templates_root=tmp_path,
    )
    assert plan.set_index("economy")["rows_added"].to_dict() == {"01_TST": 1, "02_TST": 1}

    apply_material_exception_placeholders(
        apply_changes=True, exception_workbook_path=exception_path, templates_root=tmp_path,
    )
    validate_exception_placeholder_rows(template_paths[0], {target_path})
    validate_exception_placeholder_rows(template_paths[1], {target_path})


def test_material_placeholder_dry_run_reports_blockers_without_writing(tmp_path, monkeypatch):
    template_path = tmp_path / "template.xlsx"
    _write_template(template_path, {r"Demand\Other\Existing fuel": [("Activity Level", "Reference", "Test")]})
    exception_path = tmp_path / "exceptions.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "branch_exceptions"
    sheet.append(["enabled", "branch_path", "notes", "economies_that_need_it"])
    sheet.append([True, r"Demand\Missing parent\Fuel", "Blocked test path.", "01_TST"])
    workbook.save(exception_path)
    monkeypatch.setattr(
        placeholder_module.leap_export_template_resolver,
        "iter_leap_export_templates",
        lambda _root: [SimpleNamespace(path=template_path, economy="01_TST")],
    )

    plan = apply_material_exception_placeholders(
        exception_workbook_path=exception_path, templates_root=tmp_path,
    )
    assert plan.at[0, "status"] == "blocked"
    assert "No direct sibling leaf exists" in plan.at[0, "blocker"]
    assert not placeholder_module._exception_rows(template_path, {r"Demand\Missing parent\Fuel"})
    with pytest.raises(ValueError, match="no templates changed"):
        apply_material_exception_placeholders(
            apply_changes=True, exception_workbook_path=exception_path, templates_root=tmp_path,
        )


def test_placeholder_preview_exposes_the_exact_cloned_row(tmp_path, monkeypatch):
    template_path = tmp_path / "template.xlsx"
    source_path = r"Demand\Other\Existing fuel"
    target_path = r"Demand\Other\Missing fuel"
    _write_template(template_path, {source_path: [("Activity Level", "Reference", "Test")]})
    exception_path = tmp_path / "exceptions.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "branch_exceptions"
    sheet.append(["enabled", "branch_path", "notes"])
    sheet.append([True, target_path, "Preview test path."])
    workbook.save(exception_path)
    monkeypatch.setattr(
        placeholder_module.leap_export_template_resolver,
        "iter_leap_export_templates",
        lambda _root: [SimpleNamespace(path=template_path, economy="01_TST")],
    )

    preview = preview_material_exception_placeholder_rows(
        exception_workbook_path=exception_path, templates_root=tmp_path,
    )
    assert preview[["source_branch_path", "placeholder_branch_path", "placeholder_branch_id", "Variable"]].to_dict("records") == [{
        "source_branch_path": source_path,
        "placeholder_branch_path": target_path,
        "placeholder_branch_id": 99,
        "Variable": "Activity Level",
    }]
