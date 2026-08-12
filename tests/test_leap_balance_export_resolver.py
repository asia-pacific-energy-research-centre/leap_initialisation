from __future__ import annotations

from datetime import datetime
import os
from pathlib import Path

from openpyxl import Workbook
import pytest

from codebase.utilities.leap_balance_export_resolver import (
    balance_export_unit_to_petajoule_multiplier,
    is_leap_balance_own_use_or_loss_row,
    inspect_balance_export_detail,
    load_leap_balance_activity_table,
    require_level2_balance_export_detail,
    resolve_balance_export_workbook,
    select_balance_export_sheets,
)


@pytest.mark.parametrize(
    ("units", "expected"),
    [
        ("Joule", 1e-15),
        ("Thousand Kilojoule", 1e-9),
        ("Million Megajoule", 1e-3),
        ("Billion Gigajoule", 1e3),
        ("Trillion Terajoule", 1e9),
        ("Petajoule", 1.0),
        ("Exajoule", 1e3),
        ("Million Gigajoules", 1.0),
    ],
)
def test_balance_export_unit_to_petajoule_multiplier(
    units: str,
    expected: float,
) -> None:
    assert balance_export_unit_to_petajoule_multiplier(units) == pytest.approx(expected)


@pytest.mark.parametrize(
    "units",
    [
        "British Thermal Unit",
        "Gigawatt-Hour",
        "Barrel of Oil Equivalent",
        "Tonnes of Coal Equivalent",
    ],
)
def test_balance_export_unit_rejects_non_joule_units(units: str) -> None:
    with pytest.raises(ValueError, match="None.*Petajoule"):
        balance_export_unit_to_petajoule_multiplier(units)


def test_non_specified_own_use_is_normalized_as_negative_consumption() -> None:
    assert is_leap_balance_own_use_or_loss_row("Non specified own uses")
    assert is_leap_balance_own_use_or_loss_row("Non-specified own uses")


def _touch(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("", encoding="utf-8")


def test_resolve_balance_export_workbook_uses_latest_date_id(tmp_path: Path) -> None:
    export_dir = tmp_path / "20_USA"
    _touch(export_dir / "full model output all years 492026 TGT.xlsx")
    expected = export_dir / "full model output all years 4212026 TGT.xlsx"
    _touch(expected)

    resolved = resolve_balance_export_workbook(
        economy="20_USA",
        scenario="Target",
        exports_root=tmp_path,
    )

    assert resolved == expected


def test_resolve_balance_export_workbook_honors_explicit_date_id(tmp_path: Path) -> None:
    export_dir = tmp_path / "20_USA"
    expected = export_dir / "full model output all years 492026 REF.xlsx"
    _touch(expected)
    _touch(export_dir / "full model output all years 4212026 REF.xlsx")

    resolved = resolve_balance_export_workbook(
        economy="20_USA",
        scenario="ref",
        date_id="492026",
        exports_root=tmp_path,
    )

    assert resolved == expected


def test_resolve_balance_export_workbook_accepts_current_scenario_first_name(
    tmp_path: Path,
) -> None:
    export_dir = tmp_path / "01_AUS"
    _touch(export_dir / "REF 27072026 AUS.xlsx")
    expected = export_dir / "REF 28072026 AUS.xlsx"
    _touch(expected)

    resolved = resolve_balance_export_workbook(
        economy="01_AUS",
        scenario="Reference",
        exports_root=tmp_path,
    )

    assert resolved == expected


def test_resolve_balance_export_workbook_accepts_four_digit_day_month(
    tmp_path: Path,
) -> None:
    export_dir = tmp_path / "20_USA"
    _touch(export_dir / "REF 28072026 USA.xlsx")
    expected = export_dir / "REF 2907 USA.xlsx"
    _touch(expected)
    modified = datetime(2026, 7, 29).timestamp()
    os.utime(expected, (modified, modified))

    resolved = resolve_balance_export_workbook(
        economy="20_USA",
        scenario="REF",
        exports_root=tmp_path,
    )

    assert resolved == expected


def test_resolve_balance_export_workbook_validates_nonstandard_filename_from_headers(
    tmp_path: Path,
) -> None:
    export_dir = tmp_path / "01_AUS"
    expected = export_dir / "3007 REF.xlsx"
    export_dir.mkdir()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "2060"
    sheet.append(['Energy Balance for Area "aus clean slate 29_07"'])
    sheet.append(["Scenario: Reference, Year: 2060, Units: Petajoule"])
    sheet.append([None, "Electricity"])
    workbook.save(expected)

    resolved = resolve_balance_export_workbook(
        economy="01_AUS",
        scenario="Reference",
        exports_root=tmp_path,
    )

    assert resolved == expected


def test_resolve_balance_export_workbook_reports_missing_match(tmp_path: Path) -> None:
    try:
        resolve_balance_export_workbook(
            economy="20_USA",
            scenario="REF",
            exports_root=tmp_path,
        )
    except FileNotFoundError as exc:
        assert "20_USA" in str(exc)
        assert "REF" in str(exc)
    else:
        raise AssertionError("missing balance-export workbook did not raise")


def _write_balance_workbook(path: Path, *, units: str, electricity_value: float) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "EBal|2060"
    sheet.append(['Energy Balance for Area "Test"', None, None])
    sheet.append([f"Scenario: Target, Year: 2060, Units: {units}", None, None])
    sheet.append([None, "Electricity", "Natural gas"])
    sheet.append(["Imports", electricity_value, 2.0])
    sheet.append(["Production", 3.0, 4.0])
    workbook.save(path)


def test_select_balance_export_sheets_uses_exact_requested_years(
    tmp_path: Path,
) -> None:
    path = tmp_path / "all_years.xlsx"
    workbook = Workbook()
    first = workbook.active
    first.title = "2022"
    first.append(['Energy Balance for Area "Test"'])
    first.append(["Scenario: Reference, Year: 2022, Units: Petajoule"])
    second = workbook.create_sheet("2023")
    second.append(['Energy Balance for Area "Test"'])
    second.append(["Scenario: Reference, Year: 2023, Units: Petajoule"])
    workbook.save(path)

    selected = select_balance_export_sheets(
        path,
        years=[2022],
        scenarios=["Reference"],
    )

    assert [(sheet.sheet_name, sheet.year) for sheet in selected] == [("2022", 2022)]

    with pytest.raises(ValueError, match="missing"):
        select_balance_export_sheets(
            path,
            years=[2024],
            scenarios=["Reference"],
        )


def test_inspect_balance_export_detail_distinguishes_level1_and_level2(
    tmp_path: Path,
) -> None:
    level1_path = tmp_path / "level1.xlsx"
    level2_path = tmp_path / "level2.xlsx"
    _write_balance_workbook(
        level1_path,
        units="Petajoule",
        electricity_value=1.0,
    )
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Energy Balance"
    sheet.append(['Energy Balance for Area "Test"', None])
    sheet.append(["Scenario: Reference, Year: 2022, Units: Petajoule", None])
    sheet.append([None, "Electricity"])
    sheet.append(["Oil Refining", 1.0])
    sheet.append(["   Refinery process", 1.0])
    workbook.save(level2_path)

    level1 = inspect_balance_export_detail(level1_path)
    level2 = inspect_balance_export_detail(level2_path)

    assert level1.detected_level_label == "Level 1"
    assert level1.has_level2_detail is False
    assert level2.detected_level_label == "Level 2+"
    assert level2.has_level2_detail is True
    assert level2.sample_indented_label == "Refinery process"


def test_require_level2_balance_export_detail_rejects_level1(tmp_path: Path) -> None:
    path = tmp_path / "level1.xlsx"
    _write_balance_workbook(path, units="Petajoule", electricity_value=1.0)

    try:
        require_level2_balance_export_detail([path])
    except ValueError as exc:
        assert "Level 1" in str(exc)
        assert "at least Level 2" in str(exc)
        assert str(path) in str(exc)
    else:
        raise AssertionError("Level 1 balance export did not raise")


def test_load_leap_balance_activity_table_normalizes_thousand_petajoule_to_pj(tmp_path: Path) -> None:
    pj_path = tmp_path / "pj.xlsx"
    thousand_pj_path = tmp_path / "thousand_pj.xlsx"
    _write_balance_workbook(pj_path, units="Petajoule", electricity_value=1200.0)
    _write_balance_workbook(thousand_pj_path, units="Thousand Petajoule", electricity_value=1.2)

    pj = load_leap_balance_activity_table(
        pj_path,
        balance_rows=["Imports"],
        fuels=["Electricity"],
    )
    thousand_pj = load_leap_balance_activity_table(
        thousand_pj_path,
        balance_rows=["Imports"],
        fuels=["Electricity"],
    )

    assert pj.loc[0, "value"] == 1200.0
    assert thousand_pj.loc[0, "value"] == 1200.0


def test_load_leap_balance_activity_table_normalizes_billion_gigajoule_to_pj(
    tmp_path: Path,
) -> None:
    path = tmp_path / "billion_gigajoule.xlsx"
    _write_balance_workbook(
        path,
        units="Billion Gigajoule",
        electricity_value=1.2,
    )

    result = load_leap_balance_activity_table(
        path,
        balance_rows=["Imports"],
        fuels=["Electricity"],
    )

    assert result.loc[0, "value"] == pytest.approx(1200.0)
