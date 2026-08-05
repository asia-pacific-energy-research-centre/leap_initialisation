#%%
"""Small Gradio web application for the Python balance-review builder.

The app deliberately imports the repository's existing builder rather than
reimplementing workbook logic.  When this repository is updated and the Space
is rebuilt from the updated checkout, the web app uses the same source file as
the local workflows and portable release.
"""

from __future__ import annotations

import json
import os
import re
import shutil
import subprocess
import sys
import tempfile
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


# Resolve the repository root from this file so the app works regardless of
# the directory from which it is launched.
REPO_ROOT = Path(__file__).resolve().parents[1]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from codebase.functions.balance_review_workbook_builder import (  # noqa: E402
    build_balance_structure_review_workbook,
)


REQUIRED_DIAGNOSTIC_FILES = {
    "leap_balance_source_differences.csv",
    "leap_balance_source_review.csv",
}
OPTIONAL_DIAGNOSTIC_FILES = {
    "leap_balance_mapping_issues.csv",
    "ninth_projection_allocation_diagnostics.csv",
}
ALLOWED_DIAGNOSTIC_FILES = REQUIRED_DIAGNOSTIC_FILES | OPTIONAL_DIAGNOSTIC_FILES


def _path_from_gradio_file(value: object) -> Path:
    """Return a safe local path from a Gradio File component value."""
    if value is None or str(value).strip() == "":
        raise ValueError("Please upload the LEAP balance export workbook.")
    raw_path = getattr(value, "name", value)
    path = Path(str(raw_path))
    if not path.is_file():
        raise FileNotFoundError(f"Uploaded file was not found: {path.name}")
    return path


def _uploaded_paths(values: object) -> list[Path]:
    """Normalize Gradio's single/multiple file values to existing Paths."""
    if values is None:
        return []
    if isinstance(values, (str, Path)) or hasattr(values, "name"):
        values = [values]
    paths = []
    for value in values:  # type: ignore[union-attr]
        path = _path_from_gradio_file(value)
        paths.append(path)
    return paths


def _select_source_sheet(workbook_path: Path, requested_year: int) -> str:
    """Select the exact year sheet used by the existing builder."""
    workbook = load_workbook(workbook_path, read_only=True, data_only=False)
    try:
        sheet_names = list(workbook.sheetnames)
        exact_name = str(requested_year)
        if exact_name in sheet_names:
            return exact_name

        for sheet_name in sheet_names:
            sheet = workbook[sheet_name]
            metadata = str(sheet["A2"].value or "")
            if f"Year: {requested_year}" in metadata:
                return sheet_name
    finally:
        workbook.close()

    available = ", ".join(sheet_names) or "none"
    raise ValueError(
        f"Could not find a workbook sheet for year {requested_year}. "
        f"Available sheets: {available}"
    )


def _validate_source_scenario(workbook_path: Path, sheet_name: str, scenario: str) -> None:
    """Reject a UI scenario that disagrees with the workbook metadata."""
    workbook = load_workbook(workbook_path, read_only=True, data_only=False)
    try:
        metadata = str(workbook[sheet_name]["A2"].value or "")
    finally:
        workbook.close()
    match = re.search(r"Scenario:\s*([^,]+)", metadata, flags=re.IGNORECASE)
    if match and match.group(1).strip().lower() != scenario.strip().lower():
        raise ValueError(
            f"The workbook sheet declares scenario {match.group(1).strip()!r}, "
            f"but the form contains {scenario!r}."
        )


def _copy_diagnostics(paths: Iterable[Path], destination: Path) -> list[str]:
    """Copy only recognised diagnostic filenames into the run workspace."""
    copied: set[str] = set()
    for source_path in paths:
        filename = source_path.name
        if filename not in ALLOWED_DIAGNOSTIC_FILES:
            continue
        if filename in copied:
            raise ValueError(f"Diagnostic file uploaded more than once: {filename}")
        shutil.copy2(source_path, destination / filename)
        copied.add(filename)

    missing = sorted(REQUIRED_DIAGNOSTIC_FILES - copied)
    if missing:
        raise ValueError(
            "Upload both required diagnostic files: " + ", ".join(missing)
        )
    return sorted(copied)


def _source_commit() -> str:
    """Return a lightweight source stamp without requiring Git at runtime."""
    try:
        result = subprocess.run(
            ["git", "rev-parse", "HEAD"],
            cwd=REPO_ROOT,
            capture_output=True,
            check=True,
            text=True,
        )
        commit = result.stdout.strip()
        if commit:
            return commit
    except (OSError, subprocess.CalledProcessError):
        pass

    manifest_path = REPO_ROOT / "release_build" / "distribution" / "leap-review-tools-0.1.0" / "release_manifest.json"
    if manifest_path.is_file():
        try:
            manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
            commit = manifest.get("repositories", {}).get("leap_initialisation", {}).get("commit")
            if commit:
                return str(commit)
        except (OSError, json.JSONDecodeError):
            pass
    return "live repository checkout"


def _safe_filename_token(value: object) -> str:
    token = re.sub(r"[^A-Za-z0-9_-]+", "_", str(value).strip())
    return token.strip("_") or "unknown"


def build_review(
    economy: str,
    scenario: str,
    year: float,
    source_workbook: object,
    diagnostic_files: object,
) -> tuple[str, str, str | None]:
    """Build one review workbook and return a user-facing summary and file."""
    try:
        economy_value = str(economy or "").strip()
        scenario_value = str(scenario or "").strip()
        if not economy_value:
            raise ValueError("Enter an economy code, for example 20_USA.")
        if not scenario_value:
            raise ValueError("Enter the scenario, for example Target.")
        if year is None or int(year) != year:
            raise ValueError("Enter a four-digit review year, for example 2022.")
        year_value = int(year)
        workbook_path = _path_from_gradio_file(source_workbook)
        diagnostic_paths = _uploaded_paths(diagnostic_files)

        with tempfile.TemporaryDirectory(prefix="leap_balance_review_") as run_dir_text:
            run_dir = Path(run_dir_text)
            local_workbook = run_dir / workbook_path.name
            shutil.copy2(workbook_path, local_workbook)
            diagnostics_dir = run_dir / "diagnostics"
            diagnostics_dir.mkdir()
            copied_files = _copy_diagnostics(diagnostic_paths, diagnostics_dir)
            source_sheet_name = _select_source_sheet(local_workbook, year_value)
            _validate_source_scenario(local_workbook, source_sheet_name, scenario_value)
            output_path = run_dir / (
                f"balance_review_{_safe_filename_token(economy_value)}_"
                f"{_safe_filename_token(scenario_value)}_{year_value}.xlsx"
            )

            result = build_balance_structure_review_workbook(
                economy=economy_value,
                source_workbook=local_workbook,
                source_sheet_name=source_sheet_name,
                diagnostics_directory=diagnostics_dir,
                output_workbook=output_path,
            )
            summary = {
                "status": "succeeded",
                "source_commit": _source_commit(),
                "economy": economy_value,
                "scenario": scenario_value,
                "year": year_value,
                "source_sheet": source_sheet_name,
                "diagnostic_files": copied_files,
                "comparison_state_counts": result["comparisonStateCounts"],
                "missing_audit_rows": result["missingAuditRows"],
                "formula_error_cells": result["formulaErrorCells"],
            }
            # Copy the result outside the temporary directory so Gradio can
            # serve it after this function returns; the caller owns cleanup.
            persistent_output = Path(tempfile.mkdtemp(prefix="leap_balance_review_output_")) / output_path.name
            shutil.copy2(output_path, persistent_output)

        return json.dumps(summary, indent=2), "Review workbook created.", str(persistent_output)
    except Exception as error:  # Gradio should show a plain-language failure.
        return "", f"Build failed: {error}", None


def create_app():
    """Create the web interface for local or Hugging Face execution."""
    import gradio as gr

    with gr.Blocks(title="LEAP Balance Review") as app:
        gr.Markdown(
            """# LEAP Balance Review

Upload one LEAP balance export workbook and the two required diagnostic CSVs.
The app calls the repository's Python workbook builder and returns the same
five-sheet `.xlsx` review workbook produced by the desktop release.

Uploads are copied into a temporary run directory and are not intentionally
retained by the application.
"""
        )
        with gr.Row():
            economy = gr.Textbox(label="Economy", value="20_USA")
            scenario = gr.Textbox(label="Scenario", value="Target")
            year = gr.Number(label="Review year", value=2022, precision=0)
        source_workbook = gr.File(
            label="LEAP balance export workbook (.xlsx)",
            file_types=[".xlsx", ".xlsm"],
            type="filepath",
        )
        diagnostic_files = gr.File(
            label="Diagnostic CSV files (select both required files; optional files are accepted)",
            file_count="multiple",
            file_types=[".csv"],
            type="filepath",
        )
        run_button = gr.Button("Build review workbook", variant="primary")
        status = gr.Textbox(label="Status", interactive=False)
        summary = gr.Code(label="Run summary", language="json", interactive=False)
        output = gr.File(label="Download workbook")
        run_button.click(
            fn=build_review,
            inputs=[economy, scenario, year, source_workbook, diagnostic_files],
            outputs=[summary, status, output],
        )
    return app


#%%
if __name__ == "__main__":
    APP = create_app()
    APP.launch(
        server_name=os.getenv("GRADIO_SERVER_NAME", "127.0.0.1"),
        server_port=int(os.getenv("GRADIO_SERVER_PORT", "7860")),
    )

#%%
