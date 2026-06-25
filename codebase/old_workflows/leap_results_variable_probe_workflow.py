from __future__ import annotations

import sys
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

REPO_ROOT = Path(__file__).resolve().parents[1]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from codebase import leap_results_workflow as lrw
from codebase.functions.analysis_input_write_dispatcher import (
    ensure_analysis_view_api_read_allowed,
    ensure_api_write_allowed,
)
from codebase.functions.leap_api_guard import ensure_leap_api_allowed


def _resolve(path: Path | str) -> Path:
    raw = str(path).replace("\\", "/")
    candidate = Path(raw)
    return candidate if candidate.is_absolute() else (REPO_ROOT / candidate)


#%%
# Probe config

PROBE_TEMPLATE_PATHS = [
    REPO_ROOT / "data" / "leap results tables" / "transformation_results_20_USA_Reference.xlsx",
    REPO_ROOT / "data" / "leap results tables" / "transformation_results_20_USA_Target.xlsx",
]
PROBE_OUTPUT_DIR = REPO_ROOT / "outputs" / "leap_results" / "variable_probe"
PROBE_OUTPUT_CSV = PROBE_OUTPUT_DIR / "supporting_files" / "template_variable_probe.csv"


def connect_to_leap_probe():
    ensure_analysis_view_api_read_allowed(
        "leap_results_variable_probe_workflow.connect_to_leap_probe"
    )
    ensure_leap_api_allowed("leap_results_variable_probe_workflow.connect_to_leap_probe")
    from win32com.client import Dispatch, GetActiveObject

    try:
        app = GetActiveObject("LEAP.LEAPApplication")
        print("[INFO] Connected to existing LEAP instance")
    except Exception:
        app = Dispatch("LEAP.LEAPApplication")
        print("[INFO] Created new LEAP instance")
    print(f"[INFO] Active area: {app.ActiveArea}")
    return app


def _branch_exists(app, branch_path: str) -> bool:
    try:
        return bool(app.Branches.Exists(branch_path))
    except Exception:
        return False


def _available_variable_names(app, branch_path: str) -> list[str]:
    names: list[str] = []
    try:
        branch_obj = app.Branches.Item(branch_path)
        variables = branch_obj.Variables
        count = int(getattr(variables, "Count", 0))
        for idx in range(1, count + 1):
            name = str(getattr(variables.Item(idx), "Name", "") or "").strip()
            if name:
                names.append(name)
    except Exception:
        return []
    return names


def probe_template_variable_coverage(
    template_paths: list[Path | str] | None = None,
    *,
    output_csv: Path | str = PROBE_OUTPUT_CSV,
) -> pd.DataFrame:
    """
    Check each template sheet's requested variable against live LEAP branch variables.

    Outputs one row per sheet with:
    - strict status (requested variable exists or not)
    - fallback variable that the old substitution logic would have used
    """
    selected_paths = [_resolve(path) for path in (template_paths or PROBE_TEMPLATE_PATHS)]
    app = connect_to_leap_probe()

    rows: list[dict[str, object]] = []
    for template_path in selected_paths:
        wb = load_workbook(template_path, data_only=False)
        print("\n" + "#" * 79)
        print(f"# TEMPLATE: {template_path}")
        print("#" * 79)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            meta = lrw.parse_template_worksheet(ws)
            requested_variable = str(meta.get("variable") or "").strip()
            branch_path = str(meta.get("branch") or "").strip()
            scenario = str(meta.get("scenario") or "").strip()
            region = str(meta.get("region") or "").strip()

            branch_exists = _branch_exists(app, branch_path) if branch_path else False
            available_variables = _available_variable_names(app, branch_path) if branch_exists else []
            available_text = " | ".join(available_variables)

            strict_status = "ok"
            strict_error = ""
            fallback_variable = ""
            fallback_changed_measure = False

            if not branch_path:
                strict_status = "missing_branch_path"
                strict_error = "Template sheet branch path is blank"
            elif not branch_exists:
                strict_status = "missing_branch"
                strict_error = f"Branch does not exist in LEAP: {branch_path}"
            else:
                try:
                    strict_obj = lrw._resolve_branch_variable(
                        app,
                        branch_path,
                        requested_variable,
                        allow_substitution=False,
                    )
                    strict_name = str(getattr(strict_obj, "Name", "") or "").strip()
                    fallback_variable = strict_name
                except Exception as exc:
                    strict_status = "missing_requested_variable"
                    strict_error = str(exc)

                try:
                    fallback_obj = lrw._resolve_branch_variable(
                        app,
                        branch_path,
                        requested_variable,
                        allow_substitution=True,
                    )
                    fallback_name = str(getattr(fallback_obj, "Name", "") or "").strip()
                    if fallback_name:
                        fallback_variable = fallback_name
                    fallback_changed_measure = bool(
                        fallback_name
                        and requested_variable
                        and fallback_name.strip().lower() != requested_variable.strip().lower()
                    )
                except Exception:
                    pass

            rows.append(
                {
                    "template": str(template_path),
                    "sheet": sheet_name,
                    "scenario": scenario,
                    "region": region,
                    "branch": branch_path,
                    "requested_variable": requested_variable,
                    "branch_exists": branch_exists,
                    "strict_status": strict_status,
                    "strict_error": strict_error,
                    "fallback_variable": fallback_variable,
                    "fallback_changed_measure": fallback_changed_measure,
                    "available_variables": available_text,
                }
            )

    out = pd.DataFrame(rows)
    output_path = _resolve(output_csv)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    out.to_csv(output_path, index=False)

    print("\n" + "#" * 79)
    print("# TEMPLATE VARIABLE PROBE SUMMARY")
    print("#" * 79)
    print(f"[INFO] Rows analysed: {len(out)}")
    if not out.empty:
        issue_df = out[out["strict_status"] != "ok"].copy()
        print(f"[INFO] Strict failures: {len(issue_df)}")
        if not issue_df.empty:
            display_cols = [
                "template",
                "sheet",
                "branch",
                "requested_variable",
                "strict_status",
                "fallback_variable",
                "fallback_changed_measure",
            ]
            print(issue_df[display_cols].to_string(index=False))
    print(f"[INFO] Saved probe CSV: {output_path}")
    return out


#%%
# Example run (uncomment in notebook/script execution)

# probe_df = probe_template_variable_coverage(PROBE_TEMPLATE_PATHS)
# probe_df.head()


#%%
# Transformation import workbook probe config

LEAP_IMPORT_WORKBOOK_PATH = (
    REPO_ROOT
    / "outputs"
    / "supply_reconciliation"
    / "leap_exports"
    / "transformation_leap_imports_20_USA_Reference.xlsx"
)
PROBE_SCENARIO = "Reference"
PROBE_REGION = "United States"
PROBE_BRANCH_FILTER = None
PROBE_VARIABLE_FILTER = None
PROBE_LIMIT = 20


def _show_analysis_view(app) -> None:
    for method_name in ("ShowAnalysisView", "ShowAnalysisViewTree", "ShowAnalysisViewBranches"):
        method = getattr(app, method_name, None)
        if callable(method):
            try:
                method()
                print(f"[INFO] View reset via {method_name}()")
                break
            except Exception as exc:
                print(f"[WARN] View reset failed via {method_name}(): {exc}")


def _head(text: object, n: int = 120) -> str:
    value = str(text or "")
    return value[:n] + ("..." if len(value) > n else "")


def _load_leap_sheet(export_path: Path | str) -> pd.DataFrame:
    export_path = _resolve(export_path)
    df = pd.read_excel(export_path, sheet_name="LEAP", header=2)
    required = {"Branch Path", "Variable", "Scenario", "Region"}
    missing = required.difference(df.columns)
    if missing:
        raise ValueError(f"Missing required columns in {export_path}: {sorted(missing)}")
    return df


def _load_output_fuel_names_from_import_workbook(
    export_path: Path | str,
    *,
    sector_name: str,
    scenario: str,
    region: str,
) -> list[str]:
    """Return output-fuel names from Transformation import workbook rows."""
    try:
        df = _load_leap_sheet(export_path)
    except Exception:
        return []
    df = df[df["Scenario"].astype(str).str.strip() == str(scenario).strip()].copy()
    df = df[df["Region"].astype(str).str.strip() == str(region).strip()].copy()
    prefix = f"Transformation\\{str(sector_name).strip()}\\Output Fuels\\"
    df = df[df["Branch Path"].astype(str).str.startswith(prefix)].copy()
    fuels = (
        df["Branch Path"]
        .astype(str)
        .str.replace(prefix, "", regex=False)
        .str.strip()
        .replace("", pd.NA)
        .dropna()
        .drop_duplicates()
        .sort_values()
        .tolist()
    )
    return fuels


def list_transformation_sectors(
    export_path: Path | str,
    *,
    scenario: str = PROBE_SCENARIO,
    region: str = PROBE_REGION,
) -> list[str]:
    export_path = _resolve(export_path)
    df = _load_leap_sheet(export_path)
    df = df[df["Scenario"].astype(str).str.strip() == str(scenario).strip()].copy()
    df = df[df["Region"].astype(str).str.strip() == str(region).strip()].copy()
    df = df[df["Branch Path"].astype(str).str.startswith("Transformation\\")].copy()
    sectors = (
        df["Branch Path"]
        .astype(str)
        .str.extract(r"^Transformation\\([^\\]+)", expand=False)
        .dropna()
        .astype(str)
        .drop_duplicates()
        .sort_values()
        .tolist()
    )
    print("\n" + "#" * 79)
    print("# TRANSFORMATION SECTORS")
    print("#" * 79)
    for sector in sectors:
        print(sector)
    return sectors


def _prepare_probe_rows(
    export_path: Path | str,
    *,
    scenario: str,
    region: str,
    branch_filter: str | None = None,
    variable_filter: str | None = None,
    limit: int | None = None,
) -> pd.DataFrame:
    export_path = _resolve(export_path)
    df = _load_leap_sheet(export_path)
    df = df[df["Scenario"].astype(str).str.strip() == str(scenario).strip()].copy()
    df = df[df["Region"].astype(str).str.strip() == str(region).strip()].copy()
    df = df[df["Branch Path"].astype(str).str.startswith("Transformation\\")].copy()
    if branch_filter:
        token = str(branch_filter).strip().lower()
        df = df[df["Branch Path"].astype(str).str.lower().str.contains(token, na=False, regex=False)].copy()
    if variable_filter:
        token = str(variable_filter).strip().lower()
        df = df[df["Variable"].astype(str).str.lower() == token].copy()
    df = df.reset_index(drop=True)
    if limit is not None:
        df = df.head(int(limit)).copy()
    return df


def probe_transformation_workbook(
    export_path: Path | str,
    *,
    scenario: str = PROBE_SCENARIO,
    region: str = PROBE_REGION,
    branch_filter: str | None = None,
    variable_filter: str | None = None,
    limit: int | None = 20,
    write_expression: bool = False,
) -> list[dict[str, object]]:
    export_path = _resolve(export_path)
    rows_df = _prepare_probe_rows(
        export_path,
        scenario=scenario,
        region=region,
        branch_filter=branch_filter,
        variable_filter=variable_filter,
        limit=limit,
    )

    app = connect_to_leap_probe()
    try:
        app.ActiveScenario = scenario
    except Exception as exc:
        print(f"[WARN] Failed to set ActiveScenario='{scenario}': {exc}")
    try:
        app.ActiveRegion = region
    except Exception as exc:
        print(f"[WARN] Failed to set ActiveRegion='{region}': {exc}")
    _show_analysis_view(app)
    try:
        app.ActiveBranch = "Transformation"
    except Exception as exc:
        print(f"[WARN] Failed to set ActiveBranch to Transformation: {exc}")

    results: list[dict[str, object]] = []
    print(f"[INFO] Probing {len(rows_df)} transformation row(s) from {export_path.name}")
    for idx, row in rows_df.iterrows():
        branch_path = str(row["Branch Path"]).strip()
        requested_variable = str(row["Variable"]).strip()
        expression = row["Expression"] if "Expression" in row.index else None
        print(f"\n[{idx + 1}/{len(rows_df)}] {branch_path} / {requested_variable}")
        outcome: dict[str, object] = {
            "branch_path": branch_path,
            "requested_variable": requested_variable,
            "resolved_variable": "",
            "exists": False,
            "branch_ok": False,
            "variable_ok": False,
            "write_ok": None,
            "error": "",
        }
        try:
            exists = bool(app.Branches.Exists(branch_path))
            outcome["exists"] = exists
            print(f"  exists: {exists}")
            if not exists:
                results.append(outcome)
                continue

            branch_obj = app.Branches.Item(branch_path)
            outcome["branch_ok"] = True
            print(f"  branch full: {_head(getattr(branch_obj, 'FullName', branch_path))}")

            try:
                variable_obj = lrw._resolve_branch_variable(
                    app,
                    branch_path,
                    requested_variable,
                    allow_substitution=False,
                )
                resolved_name = str(getattr(variable_obj, "Name", "") or "").strip()
                outcome["resolved_variable"] = resolved_name
                outcome["variable_ok"] = True
                print(f"  variable: {resolved_name}")
            except Exception as exc:
                outcome["error"] = str(exc)
                print(f"  variable resolve failed: {exc}")
                results.append(outcome)
                continue

            try:
                print(f"  current expr: {_head(getattr(variable_obj, 'Expression', ''))}")
            except Exception as exc:
                print(f"  current expr read failed: {exc}")

            if write_expression:
                try:
                    if pd.isna(expression):
                        expression_to_write = ""
                    else:
                        expression_to_write = str(expression)
                    ensure_api_write_allowed(
                        "leap_results_variable_probe_workflow.probe_transformation_workbook"
                    )
                    variable_obj.Expression = expression_to_write
                    outcome["write_ok"] = True
                    print("  write: ok")
                except Exception as exc:
                    outcome["write_ok"] = False
                    outcome["error"] = str(exc)
                    print(f"  write failed: {exc}")
            results.append(outcome)
        except Exception as exc:
            outcome["error"] = str(exc)
            print(f"  probe failed: {exc}")
            results.append(outcome)
    return results


def probe_transformation_sector(
    export_path: Path | str,
    *,
    sector_name: str,
    scenario: str = PROBE_SCENARIO,
    region: str = PROBE_REGION,
    variable_filter: str | None = None,
    limit: int | None = None,
    write_expression: bool = False,
) -> list[dict[str, object]]:
    sector_token = str(sector_name).strip()
    if not sector_token:
        raise ValueError("sector_name is required")
    return probe_transformation_workbook(
        export_path,
        scenario=scenario,
        region=region,
        branch_filter=f"Transformation\\{sector_token}\\",
        variable_filter=variable_filter,
        limit=limit,
        write_expression=write_expression,
    )


def probe_all_transformation_sectors(
    export_path: Path | str,
    *,
    scenario: str = PROBE_SCENARIO,
    region: str = PROBE_REGION,
    variable_filter: str | None = None,
    limit_per_sector: int | None = None,
    write_expression: bool = False,
) -> pd.DataFrame:
    sectors = list_transformation_sectors(export_path, scenario=scenario, region=region)
    summary_rows: list[dict[str, object]] = []
    for sector in sectors:
        print("\n" + "=" * 79)
        print(f"SECTOR PROBE: {sector}")
        print("=" * 79)
        sector_results = probe_transformation_sector(
            export_path,
            sector_name=sector,
            scenario=scenario,
            region=region,
            variable_filter=variable_filter,
            limit=limit_per_sector,
            write_expression=write_expression,
        )
        failed = [
            row
            for row in sector_results
            if (row.get("variable_ok") is False) or (row.get("write_ok") is False)
        ]
        summary_rows.append(
            {
                "sector": sector,
                "rows_seen": len(sector_results),
                "rows_failed": len(failed),
                "first_failed_branch": failed[0]["branch_path"] if failed else "",
                "first_failed_variable": failed[0]["requested_variable"] if failed else "",
                "status": "failed" if failed else "ok",
            }
        )
    summary = pd.DataFrame(summary_rows).sort_values(["status", "sector"]).reset_index(drop=True)
    print("\n" + "#" * 79)
    print("# ALL-SECTOR PROBE SUMMARY")
    print("#" * 79)
    if not summary.empty:
        print(summary.to_string(index=False))
    return summary


#%%
# PART 3A: CONNECT TO LEAP ONLY

_probe_app = connect_to_leap_probe()


#%%
# PART 3B: PROBE TRANSFORMATION ROWS WITHOUT WRITING
#
# probe_transformation_workbook(
#     LEAP_IMPORT_WORKBOOK_PATH,
#     scenario=PROBE_SCENARIO,
#     region=PROBE_REGION,
#     branch_filter=PROBE_BRANCH_FILTER,
#     variable_filter=PROBE_VARIABLE_FILTER,
#     limit=PROBE_LIMIT,
#     write_expression=False,
# )


#%%
# PART 3C: PROBE TRANSFORMATION ROWS WITH WRITES ENABLED
#
# probe_transformation_workbook(
#     LEAP_IMPORT_WORKBOOK_PATH,
#     scenario=PROBE_SCENARIO,
#     region=PROBE_REGION,
#     branch_filter=PROBE_BRANCH_FILTER,
#     variable_filter=PROBE_VARIABLE_FILTER,
#     limit=PROBE_LIMIT,
#     write_expression=True,
# )


#%%
# PART 3D: BLAST FURNACES
#
# probe_transformation_sector(
#     LEAP_IMPORT_WORKBOOK_PATH,
#     sector_name="Blast furnaces",
#     scenario=PROBE_SCENARIO,
#     region=PROBE_REGION,
#     variable_filter=PROBE_VARIABLE_FILTER,
#     limit=PROBE_LIMIT,
#     write_expression=True,
# )


#%%
# PART 3E: COKE OVENS
#
# probe_transformation_sector(
#     LEAP_IMPORT_WORKBOOK_PATH,
#     sector_name="Coke ovens",
#     scenario=PROBE_SCENARIO,
#     region=PROBE_REGION,
#     variable_filter=PROBE_VARIABLE_FILTER,
#     limit=PROBE_LIMIT,
#     write_expression=True,
# )


#%%
# PART 3F: GAS WORKS PLANTS
#
# probe_transformation_sector(
#     LEAP_IMPORT_WORKBOOK_PATH,
#     sector_name="Gas works plants",
#     scenario=PROBE_SCENARIO,
#     region=PROBE_REGION,
#     variable_filter=PROBE_VARIABLE_FILTER,
#     limit=PROBE_LIMIT,
#     write_expression=True,
# )


#%%
# PART 3G: HYDROGEN TRANSFORMATION
#
# probe_transformation_sector(
#     LEAP_IMPORT_WORKBOOK_PATH,
#     sector_name="Hydrogen transformation",
#     scenario=PROBE_SCENARIO,
#     region=PROBE_REGION,
#     variable_filter=PROBE_VARIABLE_FILTER,
#     limit=PROBE_LIMIT,
#     write_expression=True,
# )


#%%
# PART 3H: NG LIQUEFACTION
#
# probe_transformation_sector(
#     LEAP_IMPORT_WORKBOOK_PATH,
#     sector_name="NG Liquefaction",
#     scenario=PROBE_SCENARIO,
#     region=PROBE_REGION,
#     variable_filter=PROBE_VARIABLE_FILTER,
#     limit=PROBE_LIMIT,
#     write_expression=True,
# )


#%%
# PART 3I: NATURAL GAS BLENDING PLANTS
#
# probe_transformation_sector(
#     LEAP_IMPORT_WORKBOOK_PATH,
#     sector_name="Natural gas blending plants",
#     scenario=PROBE_SCENARIO,
#     region=PROBE_REGION,
#     variable_filter=PROBE_VARIABLE_FILTER,
#     limit=PROBE_LIMIT,
#     write_expression=True,
# )


#%%
# PART 3J: NON SPECIFIED TRANSFORMATION
#
# probe_transformation_sector(
#     LEAP_IMPORT_WORKBOOK_PATH,
#     sector_name="Non specified transformation",
#     scenario=PROBE_SCENARIO,
#     region=PROBE_REGION,
#     variable_filter=PROBE_VARIABLE_FILTER,
#     limit=PROBE_LIMIT,
#     write_expression=True,
# )


#%%
# PART 3K: RUN ALL-SECTOR PROBE
#
# probe_all_transformation_sectors(
#     LEAP_IMPORT_WORKBOOK_PATH,
#     scenario=PROBE_SCENARIO,
#     region=PROBE_REGION,
#     variable_filter=PROBE_VARIABLE_FILTER,
#     limit_per_sector=None,
#     write_expression=True,
# )


#%%
# PART 4: PROCESS-LEVEL OUTPUT FUEL AXIS PROBE

AXIS_PROBE_SECTOR = "Blast furnaces"
AXIS_PROBE_PROCESS = "Blast furnaces"
AXIS_PROBE_VARIABLE = "Outputs by Output Fuel"
AXIS_PROBE_YEAR = 2024
AXIS_PROBE_MAX_FUELS = 20


def probe_process_output_fuel_axis(
    *,
    sector_name: str = AXIS_PROBE_SECTOR,
    process_name: str = AXIS_PROBE_PROCESS,
    variable_name: str = AXIS_PROBE_VARIABLE,
    scenario: str = PROBE_SCENARIO,
    region: str = PROBE_REGION,
    year: int = AXIS_PROBE_YEAR,
    max_fuels: int = AXIS_PROBE_MAX_FUELS,
) -> pd.DataFrame:
    """
    Probe output fuels from the process branch axis instead of fuel-child branches.

    This is useful when LEAP reports fuel-child branches as not visible in a given
    region/scenario but the process-level Results variable still exposes fuel values
    through the legend axis.
    """
    app = connect_to_leap_probe()
    try:
        app.ActiveScenario = scenario
    except Exception as exc:
        print(f"[WARN] Failed to set ActiveScenario='{scenario}': {exc}")
    try:
        app.ActiveRegion = region
    except Exception as exc:
        print(f"[WARN] Failed to set ActiveRegion='{region}': {exc}")
    _show_analysis_view(app)

    sector_root_branch_path = f"Transformation\\{str(sector_name).strip()}"
    process_collection_branch_path = f"{sector_root_branch_path}\\Processes"
    process_leaf_branch_path = f"{process_collection_branch_path}\\{str(process_name).strip()}"
    output_fuels_parent = f"Transformation\\{str(sector_name).strip()}\\Output Fuels"
    candidate_branch_paths = [
        process_leaf_branch_path,
        process_collection_branch_path,
        sector_root_branch_path,
    ]

    resolved_branch = ""
    variable_obj = None
    dim_name = None
    members: list[str] = []
    last_error = ""

    for branch_candidate in candidate_branch_paths:
        try:
            resolved_candidate = lrw._resolve_existing_branch_path(app, branch_candidate)
            variable_candidate = lrw._resolve_branch_variable(
                app,
                resolved_candidate,
                variable_name,
                allow_substitution=False,
            )
            try:
                app.ShowResultsViewTable()
            except Exception:
                pass
            lrw.set_axes(app, x_axis="Years", legend="Fuel")
            lrw.set_context(
                app,
                scenario=scenario,
                region=region,
                branch_path=resolved_candidate,
            )
            dim_candidate, members_candidate = lrw.discover_legend_members_from_api(
                app,
                "Fuel",
                preferred_dimension_names=["Output Fuel", "Fuel"],
            )
            if members_candidate:
                resolved_branch = resolved_candidate
                variable_obj = variable_candidate
                dim_name = dim_candidate
                members = members_candidate
                break
            # Keep this candidate if we have not resolved anything yet.
            if not resolved_branch:
                resolved_branch = resolved_candidate
                variable_obj = variable_candidate
                dim_name = dim_candidate
                members = members_candidate
        except Exception as exc:
            last_error = str(exc)
            continue

    if variable_obj is None or not resolved_branch:
        raise RuntimeError(
            "Unable to resolve probe branch/variable for process output fuel axis. "
            f"Last error: {last_error}"
        )
    print("\n" + "#" * 79)
    print("# PROCESS OUTPUT FUEL AXIS PROBE")
    print("#" * 79)
    print(f"[INFO] Process branch: {resolved_branch}")
    print(f"[INFO] Process label hint: {process_name}")
    print(f"[INFO] Variable: {getattr(variable_obj, 'Name', variable_name)}")
    print(f"[INFO] Scenario/Region/Year: {scenario} / {region} / {year}")

    lrw.set_axes(app, x_axis="Years", legend="Fuel")
    lrw.set_context(
        app,
        scenario=scenario,
        region=region,
        branch_path=resolved_branch,
    )

    if not members:
        dim_name, members = lrw.discover_legend_members_from_api(
            app,
            "Fuel",
            preferred_dimension_names=["Output Fuel", "Fuel"],
        )
    fallback_mode = False
    fallback_rows: list[dict[str, object]] = []
    if not members:
        fallback_mode = True
        known_fuels = _load_output_fuel_names_from_import_workbook(
            LEAP_IMPORT_WORKBOOK_PATH,
            sector_name=sector_name,
            scenario=scenario,
            region=region,
        )
        print(f"[INFO] Fallback known fuels from import workbook ({len(known_fuels)}): {known_fuels}")
        for fuel in known_fuels[: int(max_fuels)]:
            best_value = None
            best_dim = ""
            best_error = "all filters failed"
            for dim_label in ("Output Fuel", "Fuel", "Feedstock Fuel"):
                filter_str = f"{dim_label}={fuel}"
                try:
                    value = variable_obj.ValueRS(region, scenario, int(year), "", filter_str)
                    best_value = value
                    best_dim = dim_label
                    best_error = ""
                    break
                except Exception as exc:
                    best_error = str(exc)
            child_branch_path = f"{output_fuels_parent}\\{fuel}"
            try:
                child_visible = bool(app.Branches.Exists(child_branch_path))
            except Exception:
                child_visible = False
            fallback_rows.append(
                {
                    "fuel": fuel,
                    "year": int(year),
                    "value_rs": best_value,
                    "filter_dimension_used": best_dim,
                    "child_branch_exists": child_visible,
                    "child_branch_path": child_branch_path,
                    "error": best_error,
                }
            )
    members = members[: int(max_fuels)]
    print(f"[INFO] Legend dimension resolved: {dim_name or '<none>'}")
    print(f"[INFO] Candidate fuels ({len(members)}): {members}")

    rows: list[dict[str, object]] = []
    for fuel in members:
        filter_str = f"{dim_name}={fuel}" if dim_name else f"Fuel={fuel}"
        try:
            value = variable_obj.ValueRS(region, scenario, int(year), "", filter_str)
        except Exception as exc:
            value = None
            error = str(exc)
        else:
            error = ""
        child_branch_path = f"{output_fuels_parent}\\{fuel}"
        try:
            child_visible = bool(app.Branches.Exists(child_branch_path))
        except Exception:
            child_visible = False
        rows.append(
            {
                "fuel": fuel,
                "year": int(year),
                "value_rs": value,
                "filter_dimension_used": dim_name or "Fuel",
                "child_branch_exists": child_visible,
                "child_branch_path": child_branch_path,
                "error": error,
            }
        )

    out = pd.DataFrame(rows if rows else fallback_rows)
    if not out.empty:
        out["value_rs"] = pd.to_numeric(out["value_rs"], errors="coerce")
        out = out.sort_values(["child_branch_exists", "fuel"], ascending=[False, True]).reset_index(drop=True)
    if fallback_mode:
        print("[INFO] Using fallback fuel probing mode because legend members were empty.")
    print(out.to_string(index=False))
    return out


#%%
# PART 4A: BLAST FURNACES PROCESS OUTPUT-FUEL AXIS PROBE
#
# probe_process_output_fuel_axis(
#     sector_name="Blast furnaces",
#     process_name="Blast furnaces",
#     variable_name="Outputs by Output Fuel",
#     scenario=PROBE_SCENARIO,
#     region=PROBE_REGION,
#     year=2024,
#     max_fuels=20,
# )
#%%


#%%
# PART 5: REPRODUCE INVALID ACTIVEVARIABLE POPUP

INVALID_VAR_PROBE_VARIABLE = "Outputs by Output Fuel"
INVALID_VAR_PROBE_SECTOR = "Hydrogen transformation"
INVALID_VAR_PROBE_LIMIT = 40


def reproduce_invalid_activevariable_popup(
    export_path: Path | str,
    *,
    variable_name: str = INVALID_VAR_PROBE_VARIABLE,
    scenario: str = PROBE_SCENARIO,
    region: str = PROBE_REGION,
    sector_name: str | None = INVALID_VAR_PROBE_SECTOR,
    limit: int | None = INVALID_VAR_PROBE_LIMIT,
) -> pd.DataFrame:
    """
    Reproduce where LEAP rejects ActiveVariable for transformation branches.

    This function attempts to set ActiveVariable on each candidate branch and logs
    the exact branch that fails with "Invalid variable".
    """
    app = connect_to_leap_probe()
    try:
        app.ActiveScenario = scenario
    except Exception as exc:
        print(f"[WARN] Failed to set ActiveScenario='{scenario}': {exc}")
    try:
        app.ActiveRegion = region
    except Exception as exc:
        print(f"[WARN] Failed to set ActiveRegion='{region}': {exc}")
    _show_analysis_view(app)
    try:
        app.ShowResultsViewTable()
    except Exception as exc:
        print(f"[WARN] Failed to switch to results table: {exc}")

    branch_filter = None
    if sector_name:
        branch_filter = f"Transformation\\{str(sector_name).strip()}\\"
    rows_df = _prepare_probe_rows(
        export_path,
        scenario=scenario,
        region=region,
        branch_filter=branch_filter,
        variable_filter=None,
        limit=limit,
    )

    print("\n" + "#" * 79)
    print("# INVALID ACTIVEVARIABLE REPROBE")
    print("#" * 79)
    print(f"[INFO] Variable under test: {variable_name}")
    print(f"[INFO] Scenario/Region: {scenario} / {region}")
    print(f"[INFO] Source rows: {len(rows_df)}")

    records: list[dict[str, object]] = []
    seen_candidates: set[str] = set()

    for _, row in rows_df.iterrows():
        raw_branch = str(row["Branch Path"]).strip()
        pieces = [part for part in raw_branch.split("\\") if part]
        candidates: list[str] = []
        if len(pieces) >= 4 and pieces[0] == "Transformation" and pieces[2] == "Processes":
            candidates.append("\\".join(pieces[:4]))
        if len(pieces) >= 3 and pieces[0] == "Transformation" and pieces[2] == "Processes":
            candidates.append("\\".join(pieces[:3]))
        if len(pieces) >= 2 and pieces[0] == "Transformation":
            candidates.append("\\".join(pieces[:2]))
        candidates.append(raw_branch)

        for candidate in candidates:
            if candidate in seen_candidates:
                continue
            seen_candidates.add(candidate)
            branch_exists = _branch_exists(app, candidate)
            available_variables = _available_variable_names(app, candidate) if branch_exists else []
            has_variable = any(
                str(name).strip().lower() == str(variable_name).strip().lower()
                for name in available_variables
            )
            set_ok = False
            set_error = ""
            if branch_exists:
                try:
                    lrw.set_context(
                        app,
                        scenario=scenario,
                        region=region,
                        branch_path=candidate,
                    )
                    # Avoid ActiveVariable assignment to prevent LEAP modal errors.
                    # Treat strict availability + context set as success for probing.
                    set_ok = bool(has_variable)
                except Exception as exc:
                    set_error = str(exc)
            else:
                set_error = "branch_not_found"

            records.append(
                {
                    "branch_path": candidate,
                    "branch_exists": branch_exists,
                    "has_requested_variable": has_variable,
                    "activevariable_set_ok": set_ok,
                    "error": set_error,
                    "available_variables": " | ".join(available_variables),
                }
            )

    out = pd.DataFrame(records)
    if not out.empty:
        out = out.sort_values(
            ["activevariable_set_ok", "has_requested_variable", "branch_path"],
            ascending=[True, True, True],
        ).reset_index(drop=True)
    print(out.to_string(index=False))
    return out


#%%
# PART 5A: RUN INVALID ACTIVEVARIABLE REPROBE
#
# reproduce_invalid_activevariable_popup(
#     LEAP_IMPORT_WORKBOOK_PATH,
#     variable_name="Outputs by Output Fuel",
#     scenario=PROBE_SCENARIO,
#     region=PROBE_REGION,
#     sector_name="Hydrogen transformation",
#     limit=80,
# )
#%%


#%%
# PART 6: MULTI-PROCESS EXTRACTION PROBE

PROCESS_PROBE_SECTOR = "Hydrogen transformation"
PROCESS_PROBE_LIST = ["Electrolysers", "SMR with CCS"]
PROCESS_PROBE_YEAR = 2024
PROCESS_PROBE_VARIABLES = [
    "Outputs by Output Fuel",
    "Inputs",
    "Outputs by Feedstock Fuel",
]


def probe_transformation_processes(
    *,
    sector_name: str = PROCESS_PROBE_SECTOR,
    process_names: list[str] | None = None,
    scenario: str = PROBE_SCENARIO,
    region: str = PROBE_REGION,
    year: int = PROCESS_PROBE_YEAR,
    variables: list[str] | None = None,
) -> pd.DataFrame:
    """
    Probe multiple process branches under one transformation sector.

    Checks strict variable availability and attempts ValueRS reads from each
    process branch for each requested variable.
    """
    selected_processes = [str(name).strip() for name in (process_names or PROCESS_PROBE_LIST) if str(name).strip()]
    selected_variables = [str(name).strip() for name in (variables or PROCESS_PROBE_VARIABLES) if str(name).strip()]
    if not selected_processes:
        raise ValueError("process_names is required")
    if not selected_variables:
        raise ValueError("variables is required")

    app = connect_to_leap_probe()
    try:
        app.ActiveScenario = scenario
    except Exception as exc:
        print(f"[WARN] Failed to set ActiveScenario='{scenario}': {exc}")
    try:
        app.ActiveRegion = region
    except Exception as exc:
        print(f"[WARN] Failed to set ActiveRegion='{region}': {exc}")
    _show_analysis_view(app)
    try:
        app.ShowResultsViewTable()
    except Exception:
        pass

    sector_root = f"Transformation\\{str(sector_name).strip()}"
    print("\n" + "#" * 79)
    print("# MULTI-PROCESS EXTRACTION PROBE")
    print("#" * 79)
    print(f"[INFO] Sector: {sector_root}")
    print(f"[INFO] Processes: {selected_processes}")
    print(f"[INFO] Scenario/Region/Year: {scenario} / {region} / {year}")
    print(f"[INFO] Variables: {selected_variables}")

    rows: list[dict[str, object]] = []
    for process_name in selected_processes:
        process_branch = f"{sector_root}\\Processes\\{process_name}"
        branch_exists = _branch_exists(app, process_branch)
        available_variables = _available_variable_names(app, process_branch) if branch_exists else []
        for variable_name in selected_variables:
            has_requested = any(
                str(v).strip().lower() == variable_name.strip().lower()
                for v in available_variables
            )
            value_rs = None
            error = ""
            if not branch_exists:
                error = "branch_not_found"
            elif not has_requested:
                error = "requested_variable_not_on_branch"
            else:
                try:
                    variable_obj = lrw._resolve_branch_variable(
                        app,
                        process_branch,
                        variable_name,
                        allow_substitution=False,
                    )
                    lrw.set_context(
                        app,
                        scenario=scenario,
                        region=region,
                        branch_path=process_branch,
                    )
                    # Do not force app.ActiveVariable here: LEAP UI validation can
                    # raise modal popups for process variables. Query directly.
                    try:
                        value_rs = variable_obj.ValueRS(region, scenario, int(year), "", "")
                    except Exception:
                        value_rs = variable_obj.Value(int(year))
                except Exception as exc:
                    error = str(exc)
            rows.append(
                {
                    "sector": sector_name,
                    "process": process_name,
                    "branch_path": process_branch,
                    "branch_exists": branch_exists,
                    "variable": variable_name,
                    "has_requested_variable": has_requested,
                    "year": int(year),
                    "value_rs": value_rs,
                    "error": error,
                    "available_variables": " | ".join(available_variables),
                }
            )

    out = pd.DataFrame(rows)
    if not out.empty:
        out["value_rs"] = pd.to_numeric(out["value_rs"], errors="coerce")
        out = out.sort_values(["process", "variable"]).reset_index(drop=True)
    print(out.to_string(index=False))
    return out


#%%
# PART 6A: HYDROGEN ELECTROLYSERS + SMR WITH CCS
#
# probe_transformation_processes(
#     sector_name="Hydrogen transformation",
#     process_names=["Electrolysers", "SMR with CCS"],
#     scenario=PROBE_SCENARIO,
#     region=PROBE_REGION,
#     year=2024,
# )
#%%


#%%
# PART 7: NON-SPECIFIED BRANCH-MATCH REPROBE

NONSPEC_TEMPLATE_PATH = (
    REPO_ROOT / "data" / "leap results tables" / "transformation_results_20_USA_Reference.xlsx"
)


def probe_nonspecified_branch_matching(
    *,
    template_path: Path | str = NONSPEC_TEMPLATE_PATH,
    raw_sector_name: str = "Non-specified transformation",
    raw_process_name: str = "09.12 Non-specified transformation",
    scenario: str = PROBE_SCENARIO,
    region: str = PROBE_REGION,
) -> pd.DataFrame:
    """
    Reproduce raw-sector vs template-sector branch candidate behavior.
    """
    template = _resolve(template_path)
    wb = load_workbook(template, data_only=False)

    template_sector_name = ""
    template_variables: list[str] = []
    for sheet in wb.sheetnames:
        meta = lrw.parse_template_worksheet(wb[sheet])
        branch = str(meta.get("branch") or "").strip()
        variable = str(meta.get("variable") or "").strip()
        if branch.startswith("Transformation\\Non") and variable:
            bits = [part for part in branch.split("\\") if part]
            if len(bits) >= 2 and bits[0] == "Transformation":
                template_sector_name = bits[1]
                template_variables.append(variable)

    template_variables = sorted(set(template_variables))
    if not template_sector_name:
        raise RuntimeError(f"Could not locate Non-specified template sector in {template}")

    app = connect_to_leap_probe()
    try:
        app.ActiveScenario = scenario
    except Exception:
        pass
    try:
        app.ActiveRegion = region
    except Exception:
        pass
    _show_analysis_view(app)

    raw_sector_branch = f"Transformation\\{raw_sector_name}"
    raw_candidates = [
        f"{raw_sector_branch}\\Processes\\{raw_process_name}",
        f"{raw_sector_branch}\\Processes",
        raw_sector_branch,
    ]
    template_sector_branch = f"Transformation\\{template_sector_name}"
    template_candidates = [
        f"{template_sector_branch}\\Processes\\{raw_process_name}",
        f"{template_sector_branch}\\Processes",
        template_sector_branch,
    ]

    rows: list[dict[str, object]] = []
    for source, candidates in (
        ("raw_names", raw_candidates),
        ("template_sector_name", template_candidates),
    ):
        for branch_path in candidates:
            exists = _branch_exists(app, branch_path)
            available_variables = _available_variable_names(app, branch_path) if exists else []
            has_outputs_by_output_fuel = any(
                str(v).strip().lower() == "outputs by output fuel"
                for v in available_variables
            )
            rows.append(
                {
                    "source": source,
                    "branch_path": branch_path,
                    "branch_exists": exists,
                    "has_outputs_by_output_fuel": has_outputs_by_output_fuel,
                    "available_variables": " | ".join(available_variables),
                    "template_sector_name": template_sector_name,
                    "template_variables": " | ".join(template_variables),
                }
            )

    out = pd.DataFrame(rows)
    print("\n" + "#" * 79)
    print("# NON-SPECIFIED BRANCH-MATCH REPROBE")
    print("#" * 79)
    print(f"[INFO] Template path: {template}")
    print(f"[INFO] Raw sector/process: {raw_sector_name} / {raw_process_name}")
    print(f"[INFO] Template sector name: {template_sector_name}")
    print(out.to_string(index=False))
    return out


#%%
# PART 7A: RUN NON-SPECIFIED BRANCH-MATCH REPROBE
#
# probe_nonspecified_branch_matching(
#     template_path=NONSPEC_TEMPLATE_PATH,
#     raw_sector_name="Non-specified transformation",
#     raw_process_name="09.12 Non-specified transformation",
#     scenario=PROBE_SCENARIO,
#     region=PROBE_REGION,
# )
#%%


try:
    from codebase.utilities.workflow_common import emit_completion_beep as _emit_completion_beep
except Exception:  # pragma: no cover
    def _emit_completion_beep(*, success: bool = True) -> None:  # noqa: ARG001
        return


if __name__ == "__main__":  # pragma: no cover
    _emit_completion_beep(success=True, style="chime")
