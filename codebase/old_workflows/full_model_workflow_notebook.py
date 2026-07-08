#%%
"""
Full model workflow notebook runner.

This notebook-style script coordinates the main sector workflows to build a
full LEAP model and then points users to the manual post-run checklist.
It centralizes the economy, scenario, import, and workbook-combine settings
used when running the whole model rather than one sector at a time.

Open this file as a notebook (VS Code: Python Interactive) and run all cells
to build a full LEAP model in one go.

After running, use `codebase/full_model_workflow_notebook_post_run_guide.md`
to complete manual LEAP checks/actions (units, skipped variables, branch gaps).
"""
from __future__ import annotations

import os
import sys
import time
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

#%%
# --- Repo setup ---
REPO_ROOT = Path(__file__).resolve().parents[1]
if Path.cwd() != REPO_ROOT:
    os.chdir(REPO_ROOT)
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

#%%
# --- Global toggles ---
# Most user-editable settings live in `codebase/workflow_config.py`.
from codebase import leap_api
from codebase.functions.analysis_input_write_dispatcher import get_analysis_input_write_mode
from codebase.utilities import workflow_common
from codebase.configuration import workflow_config as workflow_cfg
from codebase.utilities.output_paths import COMBINED_LEAP_EXPORTS_ROOT, STANDALONE_LEAP_EXPORTS_ROOT

LEAP_API_AVAILABLE = leap_api.is_available()
ANALYSIS_INPUT_WRITE_MODE = get_analysis_input_write_mode()

RUN_TRANSFORMATION_WORKFLOW = workflow_cfg.FULL_MODEL_RUN_TRANSFORMATION_WORKFLOW
RUN_HYDROGEN_TRANSFORMATION_WORKFLOW = (
    workflow_cfg.FULL_MODEL_RUN_HYDROGEN_TRANSFORMATION_WORKFLOW
)
RUN_SUPPLY_WORKFLOW = workflow_cfg.FULL_MODEL_RUN_SUPPLY_WORKFLOW
RUN_TRANSFERS_WORKFLOW = workflow_cfg.FULL_MODEL_RUN_TRANSFERS_WORKFLOW
RUN_MINOR_DEMAND_WORKFLOW = workflow_cfg.FULL_MODEL_RUN_MINOR_DEMAND_WORKFLOW
RUN_INDUSTRY_MAPPING_WORKFLOW = workflow_cfg.FULL_MODEL_RUN_INDUSTRY_MAPPING_WORKFLOW

# Replace standalone supply (and optional standalone transfers) with the
# integrated results-supply-link runner.
USE_supply_reconciliation_FOR_SUPPLY = True

#%%
# --- Feedstock method (applies to transformation + hydrogen + transfers) ---
# Options:
# - "single_feedstock_aux_others"
# - "split_processes_per_feedstock"
# - "multi_feedstock_single_process"
# Use None to keep the module default.
FEEDSTOCK_METHOD = workflow_cfg.FULL_MODEL_FEEDSTOCK_METHOD


def _default_import_scenarios(scenarios: list[str]) -> list[str]:
    """Return non-current-account scenarios as lowercase labels."""
    current_accounts_labels = {"current accounts", "current account"}
    return [
        str(scenario).strip().lower()
        for scenario in scenarios
        if str(scenario).strip()
        and str(scenario).strip().lower() not in current_accounts_labels
    ]

#%%
# --- Transformation workflow config ---
TRANSFORMATION_ECONOMIES = workflow_cfg.FULL_MODEL_TRANSFORMATION_ECONOMIES
TRANSFORMATION_SCENARIOS = list(workflow_cfg.FULL_MODEL_TRANSFORMATION_SCENARIOS)
TRANSFORMATION_INCLUDE_LEAP_IMPORT = (
    (LEAP_API_AVAILABLE if ANALYSIS_INPUT_WRITE_MODE == "api" else True)
    if workflow_cfg.FULL_MODEL_TRANSFORMATION_INCLUDE_LEAP_IMPORT is None
    else workflow_cfg.FULL_MODEL_TRANSFORMATION_INCLUDE_LEAP_IMPORT
)
TRANSFORMATION_IMPORT_SCENARIOS = _default_import_scenarios(TRANSFORMATION_SCENARIOS)
TRANSFORMATION_AGGREGATE_ECONOMY_LABEL = (
    workflow_cfg.FULL_MODEL_TRANSFORMATION_AGGREGATE_ECONOMY_LABEL
)
TRANSFORMATION_EXPORT_DIR = workflow_cfg.FULL_MODEL_TRANSFORMATION_EXPORT_DIR
TRANSFORMATION_FILENAME_TEMPLATE = workflow_cfg.FULL_MODEL_TRANSFORMATION_FILENAME_TEMPLATE

#%%
# --- Hydrogen transformation workflow config ---
HYDROGEN_TRANSFORMATION_ECONOMIES = (
    workflow_cfg.FULL_MODEL_HYDROGEN_TRANSFORMATION_ECONOMIES
)
HYDROGEN_TRANSFORMATION_SCENARIOS = list(
    workflow_cfg.FULL_MODEL_HYDROGEN_TRANSFORMATION_SCENARIOS
)
HYDROGEN_TRANSFORMATION_INCLUDE_LEAP_IMPORT = (
    (LEAP_API_AVAILABLE if ANALYSIS_INPUT_WRITE_MODE == "api" else True)
    if workflow_cfg.FULL_MODEL_HYDROGEN_TRANSFORMATION_INCLUDE_LEAP_IMPORT is None
    else workflow_cfg.FULL_MODEL_HYDROGEN_TRANSFORMATION_INCLUDE_LEAP_IMPORT
)
HYDROGEN_TRANSFORMATION_IMPORT_SCENARIOS = _default_import_scenarios(HYDROGEN_TRANSFORMATION_SCENARIOS)
HYDROGEN_TRANSFORMATION_HANDLE_CURRENT_ACCOUNTS = (
    workflow_cfg.FULL_MODEL_HYDROGEN_TRANSFORMATION_HANDLE_CURRENT_ACCOUNTS
)
HYDROGEN_TRANSFORMATION_AGGREGATE_ECONOMY_LABEL = (
    workflow_cfg.FULL_MODEL_HYDROGEN_TRANSFORMATION_AGGREGATE_ECONOMY_LABEL
)
HYDROGEN_TRANSFORMATION_EXPORT_DIR = (
    workflow_cfg.FULL_MODEL_HYDROGEN_TRANSFORMATION_EXPORT_DIR
)
HYDROGEN_TRANSFORMATION_FILENAME_TEMPLATE = (
    workflow_cfg.FULL_MODEL_HYDROGEN_TRANSFORMATION_FILENAME_TEMPLATE
)

#%%
# --- Supply workflow config ---
SUPPLY_ECONOMIES = list(workflow_cfg.FULL_MODEL_SUPPLY_ECONOMIES)
SUPPLY_SCENARIOS = list(workflow_cfg.FULL_MODEL_SUPPLY_SCENARIOS)
SUPPLY_INCLUDE_LEAP_IMPORT = (
    (LEAP_API_AVAILABLE if ANALYSIS_INPUT_WRITE_MODE == "api" else True)
    if workflow_cfg.FULL_MODEL_SUPPLY_INCLUDE_LEAP_IMPORT is None
    else workflow_cfg.FULL_MODEL_SUPPLY_INCLUDE_LEAP_IMPORT
)
SUPPLY_IMPORT_SCENARIOS = _default_import_scenarios(SUPPLY_SCENARIOS)
SUPPLY_EXPORT_DATASET_KEY = workflow_cfg.FULL_MODEL_SUPPLY_EXPORT_DATASET_KEY

# --- Results supply-link workflow config (replaces supply step when enabled) ---
supply_reconciliation_ECONOMIES = ["20_USA"]
supply_reconciliation_SCENARIOS = list(workflow_cfg.SUPPLY_NOTEBOOK_SCENARIOS)
supply_reconciliation_CAPACITY_UNMET_PASS_MODE = "baseline_seed"  # baseline_seed|results_update
supply_reconciliation_SCRAPE_LEAP_RESULTS = False
supply_reconciliation_RUN_LEAP_FUEL_BRANCH_PROBE_AT_START = False
supply_reconciliation_RESULTS_SINGLE_FILE_OUTPUT = True
supply_reconciliation_RESULTS_WRITE_LEGACY_SIDECAR_FILES = False
supply_reconciliation_KEEP_ALL_ZERO_SUPPLY_ROWS = True
supply_reconciliation_ENABLE_COMPLETION_BEEP = True
supply_reconciliation_COMPLETION_BEEP_ON_ERROR = True
supply_reconciliation_COMPLETION_BEEP_COUNT = 1
supply_reconciliation_COMPLETION_BEEP_FREQUENCY_HZ = 880
supply_reconciliation_COMPLETION_BEEP_DURATION_MS = 180
supply_reconciliation_COMPLETION_BEEP_PAUSE_SECONDS = 0.12
supply_reconciliation_RESULTS_SINGLE_FILE_ARCHIVE_MIN_HOURS = 24
supply_reconciliation_RESULTS_SINGLE_FILE_ARCHIVE_EVERY_RUN = True
supply_reconciliation_RUN_RESET_SUPPLY_AND_TRANSFORMATION_IMPORT_EXPORT = False

#%%
# --- Transfers workflow config ---
TRANSFERS_ECONOMIES = workflow_cfg.FULL_MODEL_TRANSFERS_ECONOMIES
TRANSFERS_SCENARIOS = list(workflow_cfg.FULL_MODEL_TRANSFERS_SCENARIOS)
TRANSFERS_INCLUDE_LEAP_IMPORT = (
    (LEAP_API_AVAILABLE if ANALYSIS_INPUT_WRITE_MODE == "api" else True)
    if workflow_cfg.FULL_MODEL_TRANSFERS_INCLUDE_LEAP_IMPORT is None
    else workflow_cfg.FULL_MODEL_TRANSFERS_INCLUDE_LEAP_IMPORT
)
TRANSFERS_IMPORT_SCENARIOS = _default_import_scenarios(TRANSFERS_SCENARIOS)
TRANSFERS_HANDLE_CURRENT_ACCOUNTS = workflow_cfg.FULL_MODEL_TRANSFERS_HANDLE_CURRENT_ACCOUNTS
TRANSFERS_INCLUDE_OUTPUT_SERIES = workflow_cfg.FULL_MODEL_TRANSFERS_INCLUDE_OUTPUT_SERIES
TRANSFERS_USE_OUTPUT_TARGETS = workflow_cfg.FULL_MODEL_TRANSFERS_USE_OUTPUT_TARGETS
TRANSFERS_AGGREGATE_ECONOMY_LABEL = workflow_cfg.FULL_MODEL_TRANSFERS_AGGREGATE_ECONOMY_LABEL

#%%
# --- Minor demand workflow config ---
MINOR_DEMAND_ECONOMIES = list(workflow_cfg.FULL_MODEL_MINOR_DEMAND_ECONOMIES)
MINOR_DEMAND_SCENARIOS = list(workflow_cfg.FULL_MODEL_MINOR_DEMAND_SCENARIOS)
MINOR_DEMAND_IMPORT_SCENARIOS = _default_import_scenarios(MINOR_DEMAND_SCENARIOS)
MINOR_DEMAND_REGION = workflow_cfg.FULL_MODEL_MINOR_DEMAND_REGION
MINOR_DEMAND_INCLUDE_LEAP_IMPORT = (
    (LEAP_API_AVAILABLE if ANALYSIS_INPUT_WRITE_MODE == "api" else True)
    if workflow_cfg.FULL_MODEL_MINOR_DEMAND_INCLUDE_LEAP_IMPORT is None
    else workflow_cfg.FULL_MODEL_MINOR_DEMAND_INCLUDE_LEAP_IMPORT
)
MINOR_DEMAND_AGGREGATE_ECONOMY_LABEL = (
    workflow_cfg.FULL_MODEL_MINOR_DEMAND_AGGREGATE_ECONOMY_LABEL
)
MINOR_DEMAND_EXPORT_FILENAME = workflow_cfg.FULL_MODEL_MINOR_DEMAND_EXPORT_FILENAME

#%%
# --- Industry mapping workflow config ---
INDUSTRY_EXPORT_PATH = workflow_cfg.FULL_MODEL_INDUSTRY_EXPORT_PATH
INDUSTRY_SHEET_NAME = workflow_cfg.FULL_MODEL_INDUSTRY_SHEET_NAME
INDUSTRY_ECONOMY = workflow_cfg.FULL_MODEL_INDUSTRY_ECONOMY
INDUSTRY_BASE_YEAR = workflow_cfg.FULL_MODEL_INDUSTRY_BASE_YEAR
INDUSTRY_SCENARIO = workflow_cfg.FULL_MODEL_INDUSTRY_SCENARIO
INDUSTRY_REGION = workflow_cfg.FULL_MODEL_INDUSTRY_REGION
INDUSTRY_ENSURE_BASE_YEAR_FROM_CURRENT_ACCOUNTS = (
    workflow_cfg.FULL_MODEL_INDUSTRY_ENSURE_BASE_YEAR_FROM_CURRENT_ACCOUNTS
)
INDUSTRY_ENFORCE_BASE_YEAR_PRESENCE = (
    workflow_cfg.FULL_MODEL_INDUSTRY_ENFORCE_BASE_YEAR_PRESENCE
)


def _safe_filename_segment(value: str) -> str:
    return "".join(ch if ch.isalnum() or ch in {"_", "-"} else "_" for ch in str(value).strip())

INDUSTRY_REMAP_FUELS = True
INDUSTRY_MAPPING_CSV_PATH = REPO_ROOT / "intermediate_data" / "industry_fuel_mapping.csv"
INDUSTRY_ESTO_DATA_PATH = REPO_ROOT / "data" / "00APEC_2024_low.csv"
INDUSTRY_NINTH_DATA_PATH = REPO_ROOT / "data" / "merged_file_energy_ALL_20251106.csv"
INDUSTRY_ESTO_SUBTOTAL_MAPPING_PATH = REPO_ROOT / "config" / "ESTO_subtotal_mapping.xlsx"
INDUSTRY_REMAP_OUTPUT_PATH = (
    STANDALONE_LEAP_EXPORTS_ROOT
    / f"industry_export_remapped_{_safe_filename_segment(INDUSTRY_ECONOMY)}_{_safe_filename_segment(INDUSTRY_SCENARIO)}.xlsx"
)
INDUSTRY_REMAP_REPORT_PATH = REPO_ROOT / "intermediate_data" / "industry_fuel_remap_report.csv"
INDUSTRY_REMAP_VALIDATION_PATH = REPO_ROOT / "intermediate_data" / "industry_fuel_remap_validation.csv"
INDUSTRY_NINTH_SCENARIO = "reference"

INDUSTRY_CREATE_BRANCHES = False
INDUSTRY_FILL_BRANCHES = True
INDUSTRY_SET_UNITS = True
INDUSTRY_HANDLE_CURRENT_ACCOUNTS_TOO = True

#%%
# --- Combined workbook output config ---
RUN_COMBINE_ALL_WORKBOOKS = True
COMBINED_WORKBOOK_OUTPUT_PATH = (
    COMBINED_LEAP_EXPORTS_ROOT / "full_model_combined_leap_imports.xlsx"
)
COMBINED_WORKBOOK_ID_REFERENCE_PATH = REPO_ROOT / "data" / "full model export.xlsx"
COMBINED_WORKBOOK_ID_REFERENCE_SHEET = "Export"
COMBINED_WORKBOOK_REPLACE_YEARS_WITH_EXPRESSION = True

#%%
# --- Workflow runners ---

def run_transformation_workflow():
    from codebase import transformation_entry

    return transformation_entry.run_transformation_workflow(
        economies=TRANSFORMATION_ECONOMIES,
        scenarios=TRANSFORMATION_SCENARIOS,
        include_leap_import=TRANSFORMATION_INCLUDE_LEAP_IMPORT,
        import_scenario=TRANSFORMATION_IMPORT_SCENARIOS,
        feedstock_method=FEEDSTOCK_METHOD,
        aggregate_economy_label=TRANSFORMATION_AGGREGATE_ECONOMY_LABEL,
        export_output_dir=TRANSFORMATION_EXPORT_DIR,
        filename_template=TRANSFORMATION_FILENAME_TEMPLATE,
    )


def run_hydrogen_transformation_workflow():
    from codebase import hydrogen_transformation_workflow

    return hydrogen_transformation_workflow.run_hydrogen_export_and_import(
        economies=HYDROGEN_TRANSFORMATION_ECONOMIES,
        scenarios=HYDROGEN_TRANSFORMATION_SCENARIOS,
        include_leap_import=HYDROGEN_TRANSFORMATION_INCLUDE_LEAP_IMPORT,
        import_scenario=HYDROGEN_TRANSFORMATION_IMPORT_SCENARIOS,
        handle_current_accounts=HYDROGEN_TRANSFORMATION_HANDLE_CURRENT_ACCOUNTS,
        feedstock_method=FEEDSTOCK_METHOD,
        aggregate_economy_label=HYDROGEN_TRANSFORMATION_AGGREGATE_ECONOMY_LABEL,
        export_output_dir=HYDROGEN_TRANSFORMATION_EXPORT_DIR,
        filename_template=HYDROGEN_TRANSFORMATION_FILENAME_TEMPLATE,
    )


def run_supply_reconciliation_workflow():
    from codebase import supply_reconciliation_workflow

    supply_reconciliation_workflow.ECONOMIES = list(supply_reconciliation_ECONOMIES)
    supply_reconciliation_workflow.SCENARIOS = list(supply_reconciliation_SCENARIOS)
    supply_reconciliation_workflow.CAPACITY_UNMET_PASS_MODE = (
        supply_reconciliation_CAPACITY_UNMET_PASS_MODE
    )
    supply_reconciliation_workflow.SCRAPE_LEAP_RESULTS = (
        supply_reconciliation_SCRAPE_LEAP_RESULTS
    )
    supply_reconciliation_workflow.RUN_LEAP_FUEL_BRANCH_PROBE_AT_START = (
        supply_reconciliation_RUN_LEAP_FUEL_BRANCH_PROBE_AT_START
    )
    supply_reconciliation_workflow.RESULTS_SINGLE_FILE_OUTPUT = (
        supply_reconciliation_RESULTS_SINGLE_FILE_OUTPUT
    )
    supply_reconciliation_workflow.RESULTS_WRITE_LEGACY_SIDECAR_FILES = (
        supply_reconciliation_RESULTS_WRITE_LEGACY_SIDECAR_FILES
    )
    supply_reconciliation_workflow.KEEP_ALL_ZERO_SUPPLY_ROWS = (
        supply_reconciliation_KEEP_ALL_ZERO_SUPPLY_ROWS
    )
    supply_reconciliation_workflow.ENABLE_COMPLETION_BEEP = (
        supply_reconciliation_ENABLE_COMPLETION_BEEP
    )
    supply_reconciliation_workflow.COMPLETION_BEEP_ON_ERROR = (
        supply_reconciliation_COMPLETION_BEEP_ON_ERROR
    )
    supply_reconciliation_workflow.COMPLETION_BEEP_COUNT = (
        supply_reconciliation_COMPLETION_BEEP_COUNT
    )
    supply_reconciliation_workflow.COMPLETION_BEEP_FREQUENCY_HZ = (
        supply_reconciliation_COMPLETION_BEEP_FREQUENCY_HZ
    )
    supply_reconciliation_workflow.COMPLETION_BEEP_DURATION_MS = (
        supply_reconciliation_COMPLETION_BEEP_DURATION_MS
    )
    supply_reconciliation_workflow.COMPLETION_BEEP_PAUSE_SECONDS = (
        supply_reconciliation_COMPLETION_BEEP_PAUSE_SECONDS
    )
    supply_reconciliation_workflow.RESULTS_SINGLE_FILE_ARCHIVE_MIN_HOURS = (
        supply_reconciliation_RESULTS_SINGLE_FILE_ARCHIVE_MIN_HOURS
    )
    supply_reconciliation_workflow.RESULTS_SINGLE_FILE_ARCHIVE_EVERY_RUN = (
        supply_reconciliation_RESULTS_SINGLE_FILE_ARCHIVE_EVERY_RUN
    )
    supply_reconciliation_workflow.RUN_RESET_SUPPLY_AND_TRANSFORMATION_IMPORT_EXPORT = (
        supply_reconciliation_RUN_RESET_SUPPLY_AND_TRANSFORMATION_IMPORT_EXPORT
    )
    return supply_reconciliation_workflow.run_with_config()


def run_supply_workflow():
    if USE_supply_reconciliation_FOR_SUPPLY:
        return run_supply_reconciliation_workflow()

    from codebase import supply_workflow

    return supply_workflow.run_supply_export_and_import(
        economies=SUPPLY_ECONOMIES,
        export_dataset_key=SUPPLY_EXPORT_DATASET_KEY,
        scenario_names=SUPPLY_SCENARIOS,
        include_leap_import=SUPPLY_INCLUDE_LEAP_IMPORT,
        import_scenario=SUPPLY_IMPORT_SCENARIOS,
    )


def run_transfers_workflow():
    from codebase import transfers_workflow

    return transfers_workflow.run_transfer_export_and_import(
        economies=TRANSFERS_ECONOMIES,
        scenarios=TRANSFERS_SCENARIOS,
        include_leap_import=TRANSFERS_INCLUDE_LEAP_IMPORT,
        import_scenario=TRANSFERS_IMPORT_SCENARIOS,
        handle_current_accounts=TRANSFERS_HANDLE_CURRENT_ACCOUNTS,
        include_output_series=TRANSFERS_INCLUDE_OUTPUT_SERIES,
        use_output_targets=TRANSFERS_USE_OUTPUT_TARGETS,
        feedstock_method=FEEDSTOCK_METHOD,
        aggregate_economy_label=TRANSFERS_AGGREGATE_ECONOMY_LABEL,
    )


def run_minor_demand_workflow():
    from codebase.archive import minor_demand_workflow

    run_economies = list(MINOR_DEMAND_ECONOMIES)
    should_aggregate, aggregate_label, _ = workflow_common.resolve_aggregate_economy(
        run_economies,
        aggregate_label=MINOR_DEMAND_AGGREGATE_ECONOMY_LABEL,
    )
    if should_aggregate:
        run_economies = [aggregate_label]

    outputs = []
    for economy in run_economies:
        output = minor_demand_workflow.assemble_minor_demand_workbook(
            economy=economy,
            export_filename=MINOR_DEMAND_EXPORT_FILENAME,
            include_leap_import=MINOR_DEMAND_INCLUDE_LEAP_IMPORT,
            scenarios=MINOR_DEMAND_SCENARIOS,
            import_scenario=MINOR_DEMAND_IMPORT_SCENARIOS,
            region=MINOR_DEMAND_REGION,
            aggregate_economy_label=MINOR_DEMAND_AGGREGATE_ECONOMY_LABEL,
        )
        outputs.append(output)
    return outputs


def run_industry_mapping_workflow():
    from codebase.configuration.config import (
        BRANCH_DEMAND_CATEGORY,
        BRANCH_DEMAND_TECHNOLOGY,
    )
    from codebase.functions.industry_fuel_remap import remap_industry_export_fuels
    from codebase.functions.leap_core import (
        connect_to_leap,
        create_branches_from_export_file,
        fill_branches_from_export_file,
    )
    from codebase.functions.analysis_input_write_dispatcher import (
        dispatch_analysis_input_write,
        get_analysis_input_write_mode,
    )

    export_path = INDUSTRY_EXPORT_PATH
    industry_sheet_name = INDUSTRY_SHEET_NAME
    if INDUSTRY_REMAP_FUELS:
        remap_industry_export_fuels(
            input_path=str(INDUSTRY_EXPORT_PATH),
            output_path=str(INDUSTRY_REMAP_OUTPUT_PATH),
            mapping_csv_path=str(INDUSTRY_MAPPING_CSV_PATH),
            esto_data_path=str(INDUSTRY_ESTO_DATA_PATH),
            ninth_data_path=str(INDUSTRY_NINTH_DATA_PATH),
            subtotal_mapping_path=str(INDUSTRY_ESTO_SUBTOTAL_MAPPING_PATH),
            economy=INDUSTRY_ECONOMY,
            base_year=INDUSTRY_BASE_YEAR,
            scenario=INDUSTRY_NINTH_SCENARIO,
            sheet_name=INDUSTRY_SHEET_NAME,
            include_extra_others=False,
            report_path=str(INDUSTRY_REMAP_REPORT_PATH),
            validation_path=str(INDUSTRY_REMAP_VALIDATION_PATH),
            ensure_base_year_from_current_accounts=(
                INDUSTRY_ENSURE_BASE_YEAR_FROM_CURRENT_ACCOUNTS
            ),
            enforce_base_year_presence=INDUSTRY_ENFORCE_BASE_YEAR_PRESENCE,
        )
        export_path = INDUSTRY_REMAP_OUTPUT_PATH
        industry_sheet_name = "LEAP"

    if not INDUSTRY_CREATE_BRANCHES and not INDUSTRY_FILL_BRANCHES:
        return export_path
    write_mode = get_analysis_input_write_mode()
    if write_mode == "workbook":
        dispatch_analysis_input_write(
            export_path=Path(export_path),
            sheet_name=industry_sheet_name,
            scenario=INDUSTRY_SCENARIO,
            region=INDUSTRY_REGION,
            context_label="full_model_workflow_notebook.run_industry_mapping_workflow",
        )
        return export_path
    if not LEAP_API_AVAILABLE:
        print("[WARN] LEAP API unavailable; skipping industry branch creation/fill.")
        return export_path

    L = connect_to_leap()

    if INDUSTRY_CREATE_BRANCHES:
        create_branches_from_export_file(
            L,
            str(export_path),
            sheet_name=industry_sheet_name,
            branch_path_col="Branch Path",
            scenario=INDUSTRY_SCENARIO,
            region=INDUSTRY_REGION,
            branch_type_mapping=None,
            default_branch_type=(
                BRANCH_DEMAND_CATEGORY,
                BRANCH_DEMAND_CATEGORY,
                BRANCH_DEMAND_TECHNOLOGY,
            ),
            RAISE_ERROR_ON_FAILED_BRANCH_CREATION=True,
        )

    if INDUSTRY_FILL_BRANCHES:
        fill_branches_from_export_file(
            L,
            str(export_path),
            sheet_name=industry_sheet_name,
            scenario=INDUSTRY_SCENARIO,
            region=INDUSTRY_REGION,
            RAISE_ERROR_ON_FAILED_SET=True,
            SET_UNITS=INDUSTRY_SET_UNITS,
            HANDLE_CURRENT_ACCOUNTS_TOO=INDUSTRY_HANDLE_CURRENT_ACCOUNTS_TOO,
        )

    return export_path

#%%
# --- Workbook combine helpers ---

def _normalize_template_header_value(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and pd.isna(value):
        return ""
    if isinstance(value, float) and float(value).is_integer():
        return str(int(value))
    return str(value)


def _read_workbook_sheet_with_header_detection(
    workbook_path: Path | str,
    sheet_name: str,
) -> tuple[pd.DataFrame, pd.DataFrame, list[object]]:
    """Return (preamble_rows, data_rows, header_values) for a LEAP-style sheet."""
    path = Path(workbook_path).expanduser().resolve()
    raw = pd.read_excel(path, sheet_name=sheet_name, header=None)
    header_row = None
    for idx in range(len(raw.index)):
        values = {
            _normalize_template_header_value(item).strip().lower()
            for item in raw.iloc[idx].tolist()
        }
        if "branch path" in values and "variable" in values:
            header_row = int(idx)
            break
    if header_row is None:
        raise ValueError(f"Could not locate LEAP header in {path.name}::{sheet_name}.")
    header_values = raw.iloc[header_row].tolist()
    preamble = raw.iloc[:header_row].copy()
    data = raw.iloc[header_row + 1 :].copy()
    data.columns = header_values
    data = data.dropna(how="all").reset_index(drop=True)
    return preamble, data, header_values


def _merge_workbook_sheets(
    workbook_paths: list[Path],
    sheet_name: str,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    source_paths: list[Path] = []
    for path in workbook_paths:
        if not path.exists():
            continue
        try:
            sheet_names = set(pd.ExcelFile(path).sheet_names)
        except Exception:
            continue
        if sheet_name not in sheet_names:
            continue
        source_paths.append(path)
    if not source_paths:
        return pd.DataFrame(), pd.DataFrame()

    preamble, first_data, _ = _read_workbook_sheet_with_header_detection(
        source_paths[0],
        sheet_name=sheet_name,
    )
    ordered_columns = list(first_data.columns)
    merged = [first_data]
    for path in source_paths[1:]:
        _, data, _ = _read_workbook_sheet_with_header_detection(path, sheet_name=sheet_name)
        for col in data.columns:
            if col not in ordered_columns:
                ordered_columns.append(col)
        merged.append(data)

    normalized = [frame.reindex(columns=ordered_columns) for frame in merged]
    merged_data = pd.concat(normalized, ignore_index=True, sort=False)
    dedupe_cols = [
        col
        for col in ["Branch Path", "Variable", "Scenario", "Region", "Expression"]
        if col in merged_data.columns
    ]
    if dedupe_cols:
        merged_data = merged_data.drop_duplicates(subset=dedupe_cols, keep="last")
    else:
        merged_data = merged_data.drop_duplicates(keep="last")

    if "Branch Path" in merged_data.columns and "Variable" in merged_data.columns:
        merged_data = merged_data.sort_values(["Branch Path", "Variable"]).reset_index(drop=True)
    return preamble, merged_data


def _prepare_preamble_for_export(preamble: pd.DataFrame) -> pd.DataFrame:
    """Return a 4-column preamble block to be written starting at column E."""
    out = pd.DataFrame(
        [["Area:", "LEAP Transformation Imports", "Ver:", "2"], ["", "", "", ""]]
    )
    if preamble is None or preamble.empty:
        return out

    existing = preamble.copy()
    if existing.shape[1] >= 4:
        existing = existing.iloc[:, :4].copy()
    else:
        existing = existing.reindex(columns=range(4))
    existing = existing.fillna("")

    # Keep any additional preamble rows below the standard two-row LEAP preamble.
    if len(existing.index) > 2:
        extra = existing.iloc[2:].copy()
        out = pd.concat([out, extra], ignore_index=True)
    return out


def _normalize_join_value(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and pd.isna(value):
        return ""
    text = str(value).strip()
    if text.lower() in {"", "nan", "none", "null", "<na>"}:
        return ""
    return text


def _load_combined_id_reference() -> pd.DataFrame:
    source_path = Path(COMBINED_WORKBOOK_ID_REFERENCE_PATH).expanduser().resolve()
    if not source_path.exists():
        print(
            "[WARN] ID reference workbook not found for combined output merge: "
            f"{source_path}"
        )
        return pd.DataFrame()
    try:
        _, reference_data, _ = _read_workbook_sheet_with_header_detection(
            source_path,
            COMBINED_WORKBOOK_ID_REFERENCE_SHEET,
        )
    except Exception as exc:
        print(
            "[WARN] Failed reading ID reference workbook for combined output merge: "
            f"{source_path} (sheet={COMBINED_WORKBOOK_ID_REFERENCE_SHEET}) -> {exc}"
        )
        return pd.DataFrame()
    return reference_data


def _load_combined_reference_export_columns() -> list[object]:
    source_path = Path(COMBINED_WORKBOOK_ID_REFERENCE_PATH).expanduser().resolve()
    if not source_path.exists():
        return []
    try:
        _, _, header_values = _read_workbook_sheet_with_header_detection(
            source_path,
            COMBINED_WORKBOOK_ID_REFERENCE_SHEET,
        )
    except Exception:
        return []
    return list(header_values)


def _canonical_header_token(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and pd.isna(value):
        return ""
    if isinstance(value, (int, float)) and float(value).is_integer():
        return str(int(value))
    text = str(value).strip()
    if text.lower() in {"", "nan", "none", "null", "<na>"}:
        return ""
    return text


def _reorder_columns_to_reference_template(
    data: pd.DataFrame,
    *,
    label: str,
    reference_columns: list[object],
) -> pd.DataFrame:
    if data is None or data.empty or not reference_columns:
        return data.copy()

    token_to_source: dict[str, object] = {}
    for col in data.columns:
        token = _canonical_header_token(col)
        if not token:
            continue
        if token not in token_to_source:
            token_to_source[token] = col

    ordered_series: list[pd.Series] = []
    missing_template_tokens: set[str] = set()
    used_source_columns: set[object] = set()
    row_count = len(data)
    for template_col in reference_columns:
        token = _canonical_header_token(template_col)
        if not token:
            ordered_series.append(pd.Series([""] * row_count, index=data.index))
            continue
        source_col = token_to_source.get(token)
        if source_col is None:
            ordered_series.append(pd.Series([""] * row_count, index=data.index))
            missing_template_tokens.add(token)
            continue
        used_source_columns.add(source_col)
        ordered_series.append(data[source_col].copy())

    ordered = pd.concat(ordered_series, axis=1)
    ordered.columns = reference_columns

    dropped = [
        _canonical_header_token(col)
        for col in data.columns
        if col not in used_source_columns and _canonical_header_token(col)
    ]
    if missing_template_tokens:
        print(
            f"[WARN] Combined workbook {label}: missing {len(missing_template_tokens)} "
            "template column(s) from source data; filled blanks."
        )
    if dropped:
        preview = ", ".join(sorted(set(dropped))[:8])
        print(
            f"[INFO] Combined workbook {label}: dropped {len(dropped)} non-template column(s) "
            f"to match full-model structure: {preview}"
        )
    return ordered


def _is_year_like_column(value: object) -> bool:
    token = _canonical_header_token(value)
    if not token or not token.isdigit():
        return False
    year = int(token)
    return 1900 <= year <= 2100


def _format_expression_number(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and pd.isna(value):
        return ""
    if isinstance(value, (int, float)):
        numeric = float(value)
        if pd.isna(numeric):
            return ""
        if numeric.is_integer():
            return str(int(numeric))
        return format(numeric, ".15g")
    text = str(value).strip()
    if text.lower() in {"", "nan", "none", "null", "<na>"}:
        return ""
    return text


def _build_expression_from_year_values(row: pd.Series, year_cols: list[object]) -> str:
    points: list[tuple[str, str]] = []
    for col in year_cols:
        year_token = _canonical_header_token(col)
        value_token = _format_expression_number(row.get(col))
        if not year_token or not value_token:
            continue
        points.append((year_token, value_token))
    if not points:
        return ""
    if len(points) == 1:
        return points[0][1]
    flattened = ",".join(f"{year},{value}" for year, value in points)
    return f"Interp({flattened})"


def _replace_year_columns_with_expression(
    data: pd.DataFrame,
    *,
    label: str,
    preferred_expression: pd.Series | None = None,
) -> pd.DataFrame:
    if data is None or data.empty:
        return data.copy()

    year_cols = [col for col in data.columns if _is_year_like_column(col)]
    if not year_cols:
        print(f"[WARN] Combined workbook {label}: no year columns found to replace with Expression.")
        return data.copy()

    original_columns = list(data.columns)
    insert_position = min(original_columns.index(col) for col in year_cols)

    expression_values = pd.Series([""] * len(data), index=data.index, dtype="object")
    if preferred_expression is not None:
        aligned = preferred_expression.reindex(data.index)
        expression_values = aligned.map(_format_expression_number)
    if "Expression" in data.columns:
        existing = data["Expression"].map(_format_expression_number)
        use_existing = expression_values == ""
        expression_values = expression_values.where(~use_existing, existing)

    missing_mask = expression_values == ""
    if bool(missing_mask.any()):
        built = data.loc[missing_mask, year_cols].apply(
            lambda row: _build_expression_from_year_values(row, year_cols),
            axis=1,
        )
        expression_values.loc[missing_mask] = built

    out = data.drop(columns=[*year_cols, "Expression"], errors="ignore").copy()
    out.insert(insert_position, "Expression", expression_values.fillna(""))
    print(
        f"[INFO] Combined workbook {label}: replaced {len(year_cols)} year column(s) with "
        "single Expression column."
    )
    return out


def _count_trailing_blank_template_columns(reference_columns: list[object]) -> int:
    trailing = 0
    for col in reversed(reference_columns):
        if _canonical_header_token(col):
            break
        trailing += 1
    return trailing


def _preserve_trailing_blank_template_columns(
    workbook_path: Path | str,
    *,
    sheet_names: list[str],
    reference_columns: list[object],
) -> None:
    """Keep trailing blank template columns from being dropped by Excel I/O."""
    if not reference_columns:
        return
    trailing_blank_cols = _count_trailing_blank_template_columns(reference_columns)
    if trailing_blank_cols <= 0:
        return

    target_col_count = len(reference_columns)
    trailing_start_col = target_col_count - trailing_blank_cols + 1
    path = Path(workbook_path).expanduser().resolve()

    try:
        workbook = load_workbook(path)
    except Exception as exc:
        print(
            "[WARN] Could not post-process combined workbook trailing blank template columns: "
            f"{path} -> {exc}"
        )
        return

    touched_sheets: list[str] = []
    for sheet_name in sheet_names:
        if sheet_name not in workbook.sheetnames:
            continue
        worksheet = workbook[sheet_name]
        sheet_touched = False
        for col_idx in range(trailing_start_col, target_col_count + 1):
            cell = worksheet.cell(row=1, column=col_idx)
            current = _canonical_header_token(cell.value)
            if current:
                continue
            cell.value = " "
            sheet_touched = True
        if sheet_touched:
            touched_sheets.append(sheet_name)

    if touched_sheets:
        workbook.save(path)
        joined = ", ".join(touched_sheets)
        print(
            "[INFO] Preserved trailing blank template column(s) in combined workbook "
            f"for sheets: {joined}."
        )
    workbook.close()


def _merge_id_columns_from_reference(
    data: pd.DataFrame,
    reference_data: pd.DataFrame,
    *,
    label: str,
) -> pd.DataFrame:
    out = data.copy()
    key_cols = ["Branch Path", "Variable", "Scenario", "Region"]
    id_cols = ["BranchID", "VariableID", "ScenarioID", "RegionID"]

    for col in reversed(id_cols):
        if col not in out.columns:
            out.insert(0, col, pd.NA)

    if reference_data is None or reference_data.empty:
        print(f"[WARN] Combined workbook {label}: reference data empty; IDs were not merged.")
        return out

    required_ref_cols = [*key_cols, *id_cols]
    if any(col not in reference_data.columns for col in required_ref_cols):
        missing = [col for col in required_ref_cols if col not in reference_data.columns]
        print(
            f"[WARN] Combined workbook {label}: reference data missing required columns {missing}; "
            "IDs were not merged."
        )
        return out

    ref = reference_data[required_ref_cols].copy()
    for col in key_cols:
        out[f"__k_{col}"] = out[col].map(_normalize_join_value)
        ref[f"__k_{col}"] = ref[col].map(_normalize_join_value)

    ref_key_cols = [f"__k_{col}" for col in key_cols]
    ref_dedup = ref.drop_duplicates(subset=ref_key_cols, keep="first").copy()
    duplicate_count = int(len(ref) - len(ref_dedup))
    if duplicate_count:
        print(
            f"[WARN] Combined workbook {label}: reference export has {duplicate_count} duplicate key rows; "
            "using first match."
        )

    merged = out.merge(
        ref_dedup[ref_key_cols + id_cols].rename(columns={col: f"__ref_{col}" for col in id_cols}),
        on=ref_key_cols,
        how="left",
    )

    unmatched_mask = merged["__ref_BranchID"].isna()
    unmatched_count = int(unmatched_mask.sum())
    matched_count = int(len(merged) - unmatched_count)
    print(
        f"[INFO] Combined workbook {label}: ID merge matched {matched_count}/{len(merged)} rows."
    )
    if unmatched_count:
        print(
            f"[WARN] Combined workbook {label}: {unmatched_count} rows had no ID join in "
            f"{Path(COMBINED_WORKBOOK_ID_REFERENCE_PATH).name}; keeping workbook generation."
        )
        sample = merged.loc[unmatched_mask, key_cols].drop_duplicates().head(8)
        for _, row in sample.iterrows():
            print(
                "  - {scenario} | {branch} | {variable} | {region}".format(
                    scenario=_normalize_join_value(row.get("Scenario")),
                    branch=_normalize_join_value(row.get("Branch Path")),
                    variable=_normalize_join_value(row.get("Variable")),
                    region=_normalize_join_value(row.get("Region")),
                )
            )

    for col in id_cols:
        ref_col = f"__ref_{col}"
        merged[col] = merged[ref_col].where(merged[ref_col].notna(), merged[col])

    cleanup_cols = [*ref_key_cols, *[f"__ref_{col}" for col in id_cols]]
    merged = merged.drop(columns=cleanup_cols, errors="ignore")
    return merged


def _collect_workbook_paths(value: object) -> list[Path]:
    paths: list[Path] = []

    def _visit(node: object) -> None:
        if node is None:
            return
        if isinstance(node, Path):
            candidate = node.expanduser().resolve()
            if candidate.suffix.lower() == ".xlsx" and not candidate.name.startswith("~$"):
                paths.append(candidate)
            return
        if isinstance(node, str):
            text = node.strip()
            if not text:
                return
            candidate = Path(text)
            if not candidate.is_absolute():
                candidate = (REPO_ROOT / candidate).resolve()
            else:
                candidate = candidate.expanduser().resolve()
            if candidate.suffix.lower() == ".xlsx" and not candidate.name.startswith("~$"):
                paths.append(candidate)
            return
        if isinstance(node, dict):
            for item in node.values():
                _visit(item)
            return
        if isinstance(node, (list, tuple, set)):
            for item in node:
                _visit(item)
            return

    _visit(value)
    deduped: list[Path] = []
    seen: set[str] = set()
    for item in paths:
        token = str(item)
        if token in seen:
            continue
        seen.add(token)
        if item.exists():
            deduped.append(item)
    return deduped


def combine_workbooks_into_one(
    workbook_paths: list[Path],
    *,
    output_path: Path | str = COMBINED_WORKBOOK_OUTPUT_PATH,
) -> Path | None:
    if not workbook_paths:
        print("[WARN] No workbook paths were collected; skipping combined workbook output.")
        return None

    leap_preamble, leap_data = _merge_workbook_sheets(workbook_paths, "LEAP")
    if leap_data.empty:
        print("[WARN] No LEAP-sheet rows found across collected workbooks; skipping combine.")
        return None
    viewing_preamble, viewing_data = _merge_workbook_sheets(workbook_paths, "FOR_VIEWING")
    if viewing_data.empty:
        viewing_preamble = leap_preamble.copy()
        viewing_data = leap_data.copy()

    id_reference = _load_combined_id_reference()
    reference_columns = _load_combined_reference_export_columns()
    leap_expression_source = leap_data["Expression"].copy() if "Expression" in leap_data.columns else None
    viewing_expression_source = (
        viewing_data["Expression"].copy() if "Expression" in viewing_data.columns else None
    )
    if not id_reference.empty:
        leap_data = _merge_id_columns_from_reference(
            leap_data,
            id_reference,
            label="LEAP",
        )
        viewing_data = _merge_id_columns_from_reference(
            viewing_data,
            id_reference,
            label="FOR_VIEWING",
        )
    if reference_columns:
        leap_data = _reorder_columns_to_reference_template(
            leap_data,
            label="LEAP",
            reference_columns=reference_columns,
        )
        viewing_data = _reorder_columns_to_reference_template(
            viewing_data,
            label="FOR_VIEWING",
            reference_columns=reference_columns,
        )
    if COMBINED_WORKBOOK_REPLACE_YEARS_WITH_EXPRESSION:
        leap_data = _replace_year_columns_with_expression(
            leap_data,
            label="LEAP",
            preferred_expression=leap_expression_source,
        )
        viewing_data = _replace_year_columns_with_expression(
            viewing_data,
            label="FOR_VIEWING",
            preferred_expression=viewing_expression_source,
        )

    leap_preamble = _prepare_preamble_for_export(leap_preamble)
    viewing_preamble = _prepare_preamble_for_export(viewing_preamble)

    out_path = Path(output_path).expanduser().resolve()
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(out_path, engine="openpyxl", mode="w") as writer:
        leap_preamble.to_excel(
            writer,
            sheet_name="LEAP",
            index=False,
            header=False,
            startcol=4,
        )
        pd.DataFrame([list(leap_data.columns)]).to_excel(
            writer,
            sheet_name="LEAP",
            index=False,
            header=False,
            startrow=len(leap_preamble),
        )
        leap_data.to_excel(
            writer,
            sheet_name="LEAP",
            index=False,
            header=False,
            startrow=len(leap_preamble) + 1,
        )
        viewing_preamble.to_excel(
            writer,
            sheet_name="FOR_VIEWING",
            index=False,
            header=False,
            startcol=4,
        )
        pd.DataFrame([list(viewing_data.columns)]).to_excel(
            writer,
            sheet_name="FOR_VIEWING",
            index=False,
            header=False,
            startrow=len(viewing_preamble),
        )
        viewing_data.to_excel(
            writer,
            sheet_name="FOR_VIEWING",
            index=False,
            header=False,
            startrow=len(viewing_preamble) + 1,
        )
    if reference_columns:
        _preserve_trailing_blank_template_columns(
            out_path,
            sheet_names=["LEAP", "FOR_VIEWING"],
            reference_columns=reference_columns,
        )
    print(
        "[INFO] Saved combined full-model workbook to "
        f"{out_path} (LEAP rows={len(leap_data)}, FOR_VIEWING rows={len(viewing_data)})."
    )
    return out_path

#%%
# --- Run all workflows ---

def run_all_workflows():
    results = {}
    start = time.time()
    print(
        "[WARN] Reset reminder: if you need stale trade targets cleared before filling LEAP, "
        "set MAIN_RUN_RESET_SUPPLY_AND_TRANSFORMATION_IMPORT_EXPORT=True in "
        "codebase/supply_reconciliation_workflow.py."
    )

    if RUN_TRANSFORMATION_WORKFLOW:
        print("[1/6] Running transformation workflow...")
        results["transformation"] = run_transformation_workflow()

    if RUN_HYDROGEN_TRANSFORMATION_WORKFLOW:
        print("[2/6] Running hydrogen transformation workflow...")
        results["hydrogen_transformation"] = run_hydrogen_transformation_workflow()

    if RUN_SUPPLY_WORKFLOW:
        if USE_supply_reconciliation_FOR_SUPPLY:
            print(
                "[3/6] Running supply_reconciliation workflow (replacing standalone supply workflow)..."
            )
        else:
            print("[3/6] Running supply workflow...")
        results["supply"] = run_supply_workflow()

    if RUN_TRANSFERS_WORKFLOW:
        if USE_supply_reconciliation_FOR_SUPPLY:
            print(
                "[4/6] Skipping standalone transfers workflow because "
                "supply_reconciliation already includes transfer exports."
            )
            results["transfers"] = {
                "skipped": True,
                "reason": "supply_reconciliation_workflow handles transfer exports.",
            }
        else:
            print("[4/6] Running transfers workflow...")
            results["transfers"] = run_transfers_workflow()

    if RUN_MINOR_DEMAND_WORKFLOW:
        print("[5/6] Running minor demand workflow...")
        results["minor_demand"] = run_minor_demand_workflow()

    if RUN_INDUSTRY_MAPPING_WORKFLOW:
        print("[6/6] Running industry mapping workflow...")
        results["industry_mapping"] = run_industry_mapping_workflow()

    if RUN_COMBINE_ALL_WORKBOOKS:
        workbook_paths = _collect_workbook_paths(results)
        combined_path = combine_workbooks_into_one(
            workbook_paths,
            output_path=COMBINED_WORKBOOK_OUTPUT_PATH,
        )
        results["combined_workbook_path"] = combined_path

    elapsed = time.time() - start
    print(f"Done in {elapsed:.1f}s")
    return results

#%%
# Execute everything in one click.
if __name__ == "__main__":
    run_all_workflows()
#%%
