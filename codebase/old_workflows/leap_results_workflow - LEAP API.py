#%%
"""
Workflow to export LEAP Results tables programmatically via the COM API.

This script drives LEAP Results favorites/tables and saves exported CSV or Excel
outputs for downstream dashboard and comparison workflows.
It contains the lower-level extraction and workbook-refresh utilities used when
LEAP needs to render Results tables before pandas can parse them.

Run order (per user preference):
Connect → set area → ensure calc (NeedsCalculation or force) → optionally activate
favorite → set axes/context → export via CSV (or fetch values directly later).

The UI-export path mirrors what you see in Results view: LEAP renders the table,
`ExportResultsCSV` writes it to disk, and pandas can convert it to Excel.
"""
from __future__ import annotations

import os
import sys
import hashlib
import csv
import difflib
import json
from pathlib import Path
from typing import Optional
from datetime import datetime, timedelta

import pandas as pd
from openpyxl import load_workbook

# Make repo importable even when run from a notebook elsewhere.
REPO_ROOT = Path(__file__).resolve().parents[1]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from codebase.functions.leap_results_functions import (
    activate_favorite,
    connect_leap,
    ensure_calculated,
    ensure_parent_dirs,
    export_results_csv,
    fetch_values_rs,
    list_dimensions,
    select_area,
    set_axes,
    set_context,
)
from codebase.utilities.ninth_to_esto_mapping_coverage import (
    run_mapping_coverage_check,
)
from codebase.utilities.output_paths import INTEGRATED_LEAP_EXPORTS_ROOT, LEAP_RESULTS_ROOT
from codebase.utilities.workflow_common import archive_config_dir_once_per_day
from codebase.utilities import fuel_catalog_preflight
from codebase.utilities.workflow_outputs import build_workflow_output_layout, write_output_manifest


# Stable constants (unlikely to change often)
DEFAULT_OUTPUT_DIR = LEAP_RESULTS_ROOT
TEMPLATE_PATHS: list[Path] = [
    Path("data/leap results tables/transformation_results_20_USA_Target.xlsx"),
    Path("data/leap results tables/transformation_results_20_USA_Reference.xlsx"),
    Path("data/leap results tables/supply_results_20_USA_Target.xlsx"),
    Path("data/leap results tables/supply_results_20_USA_Reference.xlsx"),
    Path("data/leap results tables/transport_results_20_USA_Target.xlsx"),
    Path("data/leap results tables/transport_results_20_USA_Reference.xlsx"),
    Path("data/leap results tables/industry_results_20_USA_Target.xlsx"),
    Path("data/leap results tables/industry_results_20_USA_Reference.xlsx"),
    Path("data/leap results tables/demand_others_results_20_USA_Target.xlsx"),
    Path("data/leap results tables/demand_others_results_20_USA_Reference.xlsx"),
    Path("data/leap results tables/buildings_results_20_USA_Target.xlsx"),
    Path("data/leap results tables/buildings_results_20_USA_Reference.xlsx"),
]
COMBINED_XLSX_PATH = DEFAULT_OUTPUT_DIR / "leap_results_combined.xlsx"
NINTH_TO_ESTO_MAPPING_PATH = REPO_ROOT / "config/ninth_pairs_to_esto_pairs.xlsx"
ESTO_DATA_PATH = REPO_ROOT / "data/00APEC_2025_low_with_subtotals.csv"
NINTH_DATA_PATH = REPO_ROOT / "data/merged_file_energy_ALL_20251106.csv"
MAPPING_COVERAGE_DIR = DEFAULT_OUTPUT_DIR / "supporting_files" / "checks" / "mapping_coverage"
MAPPING_COVERAGE_BASE_YEAR = 2022
MAPPING_COVERAGE_PROJECTION_YEARS = tuple(range(2023, 2071))
MAPPING_COVERAGE_SCENARIO = "reference"
FUEL_CATALOG_PATH = (
    INTEGRATED_LEAP_EXPORTS_ROOT
    / "supporting_files"
    / "checks"
    / "leap_fuel_branch_catalog.csv"
)
FUEL_CATALOG_COVERAGE_CSV = DEFAULT_OUTPUT_DIR / "supporting_files" / "checks" / "template_fuel_catalog_coverage.csv"
FUEL_CATALOG_COVERAGE_FAIL_ON_MISSING = False


# Frequently changed settings (edit these for your run)
FORCE_RECALC = False
WRITE_EXCEL = True

# Table specs: add one entry per table you want to create/extract.
# Paths default to outputs/leap_results/<name>.csv|xlsx
TABLE_SPECS = {
    "default": {
        "area": None,  # e.g., "US Transport Study"
        "scenario": None,
        "region": None,
        "year": None,
        "unit": None,
        "branch": None,  # full path, e.g., "Demand\\Transport"
        "variable": None,  # e.g., "Energy Demand"
        "favorite": None,  # e.g., "Results#Transport Energy"
        "x_axis": None,  # e.g., "Years"
        "legend": None,  # e.g., "Scenarios"
        "csv_path": DEFAULT_OUTPUT_DIR / "leap_results_default.csv",
        "sheet": "default",  # sheet name to use in combined workbook
    },
    # Add more tables here with their own context/outputs.
}

# Optional scaling to adjust LEAP default units to desired display units.
# Per-sheet overrides:
# UNIT_SCALES = {"Passenger road": {"target_unit": "Petajoules", "scale_factor": 1e-15}}
UNIT_SCALES: dict[str, dict] = {
    # Capacity is exported as Thousand Megawatts in the template workbook.
    # LEAP ValueRS commonly returns capacity in Megawatts by default, so scale to Thousand Megawatts.
    "elecgen capacity": {"target_unit": "Thousand Megawatts", "scale_factor": 1e-3},
}
# Per-variable defaults (variable name as shown in the sheet, e.g., "Final Energy Demand")
UNIT_SCALES_BY_VARIABLE: dict[str, dict] = {
    "Final Energy Demand": {"target_unit": "Petajoules", "scale_factor": 1e-6},
    "Inputs": {"target_unit": "Petajoules", "scale_factor": 1e-6},
    "Outputs by Feedstock Fuel": {"target_unit": "Petajoules", "scale_factor": 1e-6},
    "Outputs by Output Fuel": {"target_unit": "Petajoules", "scale_factor": 1e-6},
    "Indigenous Production": {"target_unit": "Petajoules", "scale_factor": 1e-6},
    "Imports": {"target_unit": "Petajoules", "scale_factor": 1e-6},
    "Exports": {"target_unit": "Petajoules", "scale_factor": 1e-6},
    # Add more: "Capacity": {...}, "Imports": {...}
}

# Template-driven extraction settings
USE_TEMPLATE = True
ARCHIVE_TEMPLATE = True
ARCHIVE_DIR = DEFAULT_OUTPUT_DIR / "supporting_files" / "template_archives"
ARCHIVE_RETENTION_DAYS: int | None = 120
TEMPLATE_EXPORT_TMP_DIR = DEFAULT_OUTPUT_DIR / "supporting_files" / "runtime" / "_tmp_template_exports"
# Extraction mode for template refill:
# - "values": direct Variable.ValueRS calls (no CSV export)
# - "csv": LEAP ExportResultsCSV per sheet (with ValueRS fallback on failure)
TEMPLATE_EXTRACT_METHOD = "values"
FALLBACK_SUSPICIOUS_VALUES_TO_CSV = True
VALUES_DISCOVER_LEGEND_MEMBERS_FROM_API = True
VALUES_DROP_ALL_ZERO_COMPONENT_ROWS = True
LOG_RENDERED_CSV_EXPORT_STEPS = True
PRECHECK_NO_RESULTS_WITH_RESULTVALUE = True
STRICT_TEMPLATE_VARIABLE_MATCH = True
USE_FAVORITES_FOR_TRANSFORMATION_INPUT_STYLE_SHEETS = True
TRANSFORMATION_INPUT_STYLE_FAVORITE_SUFFIXES: tuple[str, ...] = (
    "_feed_inputs",
    "_aux",
    "_aux_other",
)
STRICT_FAVORITE_FOR_TRANSFORMATION_INPUT_STYLE = False
# Favorite resolution safety:
# - "strict": exact + explicit safe variants only (recommended)
# - "stem": strict + unique same-stem candidate
# - "fuzzy": stem + high-confidence global fuzzy fallback
FAVORITE_RESOLUTION_MODE = "strict"
# On rendered CSV metadata mismatch (requested variable/branch vs exported),
# choose behavior:
# - "raise": fail fast
# - "empty_table": log warning and write a no-results table for this sheet
RENDERED_CSV_METADATA_MISMATCH_POLICY = "empty_table"
# When a rendered CSV extraction step fails for transformation sheets
# (e.g., missing branch/favorite context issue), choose behavior:
# - "raise": fail fast
# - "empty_table": write no-results table and continue
RENDERED_CSV_TRANSFORMATION_ERROR_POLICY = "empty_table"
TEMPLATE_EXPECTED_YEAR_START = 2022
TEMPLATE_EXPECTED_YEAR_END = 2060
TEMPLATE_EXPECTED_INCLUDE_TOTAL = True
TEMPLATE_YEAR_AXIS_AUDIT_CSV = DEFAULT_OUTPUT_DIR / "supporting_files" / "checks" / "template_year_axis_audit.csv"
EXPECTED_SHEETS_CONFIG_PATH = REPO_ROOT / "config" / "leap_results_expected_sheets.json"
ENFORCE_EXPECTED_SHEETS_CONTRACT = True
OPTIONAL_EXPECTED_SHEETS_BY_WORKBOOK: dict[str, set[str]] = {
    "transport_results_20_USA_Target.xlsx": {
        "Passenger road",
        "Freight road",
        "Demand total",
    },
    "transport_results_20_USA_Reference.xlsx": {
        "Passenger road",
        "Freight road",
        "Demand total",
    },
    "industry_results_20_USA_Target.xlsx": {"Manufacturing"},
    "industry_results_20_USA_Reference.xlsx": {"Manufacturing"},
}
STRICT_REQUIRED_FAVORITES_CONTRACT = True
SUSPICIOUS_TABLE_CONTEXT_EXCEPTIONS: dict[str, dict[str, str]] = {
    "transport_results_20_USA_Reference.xlsx": {
        "Freight non road ship": "Known USA freight-ship export currently resolves to a Total-only row.",
    },
    "transport_results_20_USA_Target.xlsx": {
        "Freight non road ship": "Known USA freight-ship export currently resolves to a Total-only row.",
    },
    "transformation_results_20_USA_Target.xlsx": {
        "non_spec_inputs": "Temporary exception while USA non-specified transformation inputs export resolves to a suspicious total-only table.",
        "non_spec_out_feed": "Temporary exception while USA non-specified transformation feedstock output export resolves to a suspicious total-only table.",
    },
}

# Derived parent sheets: skip direct LEAP extraction and sum from child sheets.
DERIVED_SHEET_CHILDREN: dict[str, dict[str, list[str]]] = {
    "transport_results_20_USA_Target.xlsx": {
        "Passenger road": ["Passenger road moto", "Passenger road bus", "Passenger road lpv"],
        "Freight road": ["Freight road trucks", "Freight road lcvs"],
        "Demand total": [
            "International transport air",
            "International transport ship",
            "Nonspecified transport",
            "Passenger road moto",
            "Passenger road bus",
            "Passenger road lpv",
            "Freight road trucks",
            "Freight road lcvs",
            "Passenger non road air",
            "Passenger non road rail",
            "Passenger non road ship",
            "Freight non road ship",
            "Freight non road rail",
            "Freight non road air",
            "Pipeline transport",
        ],
    },
    "transport_results_20_USA_Reference.xlsx": {
        "Passenger road": ["Passenger road moto", "Passenger road bus", "Passenger road lpv"],
        "Freight road": ["Freight road trucks", "Freight road lcvs"],
        "Demand total": [
            "International transport air",
            "International transport ship",
            "Nonspecified transport",
            "Passenger road moto",
            "Passenger road bus",
            "Passenger road lpv",
            "Freight road trucks",
            "Freight road lcvs",
            "Passenger non road air",
            "Passenger non road rail",
            "Passenger non road ship",
            "Freight non road ship",
            "Freight non road rail",
            "Freight non road air",
            "Pipeline transport",
        ],
    },
    "industry_results_20_USA_Target.xlsx": {
        "Manufacturing": [
            "Textiles and leather",
            "Wood and wood products",
            "Pulp and paper",
            "Food and tobacco",
            "Machinery",
            "Transport equipment",
            "Non-metallic minerals",
            "Non-ferrous metals",
            "Chemicals",
            "Iron and steel",
        ],
    },
    "industry_results_20_USA_Reference.xlsx": {
        "Manufacturing": [
            "Textiles and leather",
            "Wood and wood products",
            "Pulp and paper",
            "Food and tobacco",
            "Machinery",
            "Transport equipment",
            "Non-metallic minerals",
            "Non-ferrous metals",
            "Chemicals",
            "Iron and steel",
        ],
    },
}
USE_DERIVED_PARENT_SHEETS = False

_RENDERED_CSV_METADATA_MISMATCH_EVENTS: list[dict[str, str]] = []


def reset_rendered_csv_metadata_mismatch_events() -> None:
    _RENDERED_CSV_METADATA_MISMATCH_EVENTS.clear()


def get_rendered_csv_metadata_mismatch_events() -> list[dict[str, str]]:
    return list(_RENDERED_CSV_METADATA_MISMATCH_EVENTS)


def ensure_repo_root() -> None:
    """Make sure we run from repository root for relative paths."""
    cwd = Path.cwd()
    if cwd != REPO_ROOT:
        os.chdir(REPO_ROOT)


def _lookup_suspicious_table_exception(context: str) -> str:
    """Return the configured exception reason for a workbook/sheet context, if any."""
    text = str(context or "").strip()
    if not text:
        return ""
    workbook_name, _, sheet_name = text.partition("/")
    workbook_exceptions = SUSPICIOUS_TABLE_CONTEXT_EXCEPTIONS.get(workbook_name.strip(), {})
    return str(workbook_exceptions.get(sheet_name.strip(), "")).strip()


def _log_rendered_csv_export_step(context: str, step: str) -> None:
    """Emit a step-level log for rendered CSV export debugging."""
    if not LOG_RENDERED_CSV_EXPORT_STEPS:
        return
    prefix = f"[LEAP CSV] {context}" if context else "[LEAP CSV]"
    print(f"{prefix}: {step}", flush=True)


def _run_rendered_csv_export_step(context: str, step: str, func):
    """Run one rendered-export step and raise with step context on failure."""
    _log_rendered_csv_export_step(context, step)
    try:
        return func()
    except Exception as exc:  # noqa: BLE001
        suffix = f" for {context}" if context else ""
        raise RuntimeError(f"LEAP rendered CSV step failed{suffix} at '{step}': {exc}") from exc


def _set_active_variable(app, variable_obj) -> None:
    """Set the LEAP results variable by name to avoid stale-table exports."""
    variable_name = str(getattr(variable_obj, "Name", "") or "").strip()
    if variable_name:
        app.ActiveVariable = variable_name
        return
    app.ActiveVariable = variable_obj


def _show_results_view_table(app) -> None:
    """Switch LEAP to the Results table view."""
    app.ShowResultsViewTable()


def _export_results_csv_file(app, output_path: Path) -> Path:
    """Export the current LEAP Results table to CSV."""
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    app.ExportResultsCSV(str(output_path))
    return output_path


def _extract_year_items(meta: dict[str, object]) -> list[int]:
    """Return valid year tokens from template metadata."""
    years: list[int] = []
    seen: set[int] = set()
    for raw_value in meta.get("x_items", []) or []:
        try:
            year = int(float(raw_value))
        except Exception:
            continue
        if not (1900 <= year <= 2200):
            continue
        if year in seen:
            continue
        seen.add(year)
        years.append(year)
    return years


def _extract_template_axis_details(x_items: list[object] | tuple[object, ...]) -> tuple[list[int], bool]:
    """Return parsed year items and whether the axis includes a Total column."""
    years: list[int] = []
    seen: set[int] = set()
    has_total = False
    for raw_value in x_items or []:
        text = str(raw_value or "").strip()
        if not text:
            continue
        if text.lower() == "total":
            has_total = True
            continue
        try:
            year = int(float(text))
        except Exception:
            continue
        if year in seen:
            continue
        seen.add(year)
        years.append(year)
    return years, has_total


def _expected_template_x_items() -> list[object]:
    """Return the canonical year axis for LEAP template workbooks."""
    years: list[object] = list(range(TEMPLATE_EXPECTED_YEAR_START, TEMPLATE_EXPECTED_YEAR_END + 1))
    if TEMPLATE_EXPECTED_INCLUDE_TOTAL:
        years.append("Total")
    return years


def _summarize_template_axis(x_items: list[object] | tuple[object, ...]) -> str:
    """Build a compact summary string for a template year axis."""
    years, has_total = _extract_template_axis_details(x_items)
    if years:
        summary = f"{min(years)}-{max(years)}"
    else:
        summary = "<none>"
    if has_total:
        summary = f"{summary} + Total"
    return summary


def normalize_template_year_axis(
    meta: dict[str, object],
    *,
    template_name: str = "",
    sheet_name: str = "",
) -> tuple[dict[str, object], dict[str, object]]:
    """
    Normalize damaged template headers back to the canonical annual axis.

    A bad refill can overwrite workbook headers with sparse years (for example
    every fifth year only). Once that happens, later refills keep reusing the
    damaged header unless we repair it before extraction.
    """
    normalized = dict(meta)
    raw_x_items = list(meta.get("x_items", []) or [])
    actual_years, has_total = _extract_template_axis_details(raw_x_items)
    expected_x_items = _expected_template_x_items()
    expected_years, expected_has_total = _extract_template_axis_details(expected_x_items)

    consecutive = bool(actual_years) and all(
        later - earlier == 1 for earlier, later in zip(actual_years, actual_years[1:])
    )
    needs_normalization = (actual_years != expected_years) or (has_total != expected_has_total)
    if needs_normalization:
        normalized["x_items"] = list(expected_x_items)

    audit_row = {
        "template": template_name,
        "sheet": sheet_name,
        "year_count_found": len(actual_years),
        "year_count_expected": len(expected_years),
        "years_consecutive_found": consecutive,
        "has_total_found": has_total,
        "has_total_expected": expected_has_total,
        "axis_found": _summarize_template_axis(raw_x_items),
        "axis_expected": _summarize_template_axis(expected_x_items),
        "normalized": needs_normalization,
    }
    return normalized, audit_row


def audit_template_year_axes(
    paths: list[Path] | None = None,
    *,
    output_csv_path: Path = TEMPLATE_YEAR_AXIS_AUDIT_CSV,
) -> tuple[pd.DataFrame, Path]:
    """Audit template workbook year axes and write a CSV summary."""
    ensure_repo_root()
    rows: list[dict[str, object]] = []
    for tpl in paths or TEMPLATE_PATHS:
        workbook_path = (REPO_ROOT / tpl).resolve()
        if not workbook_path.exists():
            rows.append(
                {
                    "template": tpl.name,
                    "sheet": "",
                    "year_count_found": 0,
                    "year_count_expected": TEMPLATE_EXPECTED_YEAR_END - TEMPLATE_EXPECTED_YEAR_START + 1,
                    "years_consecutive_found": False,
                    "has_total_found": False,
                    "has_total_expected": TEMPLATE_EXPECTED_INCLUDE_TOTAL,
                    "axis_found": "<missing workbook>",
                    "axis_expected": _summarize_template_axis(_expected_template_x_items()),
                    "normalized": False,
                    "exists": False,
                }
            )
            continue

        wb = load_workbook(workbook_path, read_only=True, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            meta = parse_template_worksheet(ws)
            _, audit_row = normalize_template_year_axis(
                meta,
                template_name=tpl.name,
                sheet_name=sheet_name,
            )
            audit_row["exists"] = True
            rows.append(audit_row)

    audit_df = pd.DataFrame(rows)
    output_csv_path = output_csv_path.resolve()
    output_csv_path.parent.mkdir(parents=True, exist_ok=True)
    audit_df.to_csv(output_csv_path, index=False)
    return audit_df, output_csv_path


def _probe_no_results_via_value_rs(
    app,
    *,
    meta: dict[str, object],
    variable_obj,
    context: str = "",
) -> bool | None:
    """
    Return False when ValueRS fails for every probe year.

    Return True when any probe year succeeds, even if the value is zero.
    Return None when the probe itself cannot be run safely.
    """
    years = _extract_year_items(meta)
    if not years:
        return None

    try:
        set_context(
            app,
            scenario=meta.get("scenario"),
            region=meta.get("region"),
        )
    except Exception as exc:  # noqa: BLE001
        _log_rendered_csv_export_step(context, f"result_probe skipped set_context failed: {exc}")
        return None

    region = meta.get("region")
    scenario = meta.get("scenario")
    if not region or not scenario:
        _log_rendered_csv_export_step(context, "result_probe skipped missing region_or_scenario")
        return None

    for year in years:
        try:
            value = fetch_values_rs(
                variable_obj,
                region=str(region),
                scenario=str(scenario),
                year=year,
                unit="",
                filter_str="",
            )
        except Exception:
            continue
        if not pd.isna(value):
            _log_rendered_csv_export_step(context, f"result_probe success year={year} value={value}")
            return True

    _log_rendered_csv_export_step(context, "result_probe no successful ValueRS calls")
    return False


def _build_no_results_table(meta: dict[str, object]) -> pd.DataFrame:
    """Return a template-shaped table with metadata and header only."""
    header = [meta.get("legend_label", "")] + list(meta.get("x_items", []) or [])
    rows = [
        [meta.get("variable", "")],
        [f"Scenario: {meta.get('scenario','')}, Region: {meta.get('region','')}"],
        [f"Branch: {meta.get('branch','')}"],
        [f"Units: {meta.get('units','')}"],
        [""],
        header,
    ]
    return pd.DataFrame(rows)


class MissingRequestedVariableError(ValueError):
    """Raised when a sheet requests a variable that does not exist on the target LEAP branch."""


def run_mapping_coverage_check_for_workflow() -> dict[str, object]:
    """Validate canonical mapping coverage before LEAP export work begins."""
    archive_config_dir_once_per_day()
    result = run_mapping_coverage_check(
        mapping_path=NINTH_TO_ESTO_MAPPING_PATH,
        esto_data_path=ESTO_DATA_PATH,
        ninth_data_path=NINTH_DATA_PATH,
        output_dir=MAPPING_COVERAGE_DIR,
        base_year=MAPPING_COVERAGE_BASE_YEAR,
        projection_years=MAPPING_COVERAGE_PROJECTION_YEARS,
        scenario=MAPPING_COVERAGE_SCENARIO,
    )
    summary = result["summary"]
    print(
        "Mapping coverage check: "
        f"missing ESTO pairs={summary['missing_esto_pairs']}, "
        f"missing 9th pairs={summary['missing_ninth_pairs']}",
        flush=True,
    )
    return result


def write_excel_from_csv(csv_path: Path, xlsx_path: Path, sheet_name: str = "Results") -> Path:
    """Load CSV and save to Excel for parity with manual exports."""
    df = pd.read_csv(csv_path)
    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)
    return xlsx_path


def parse_template_sheet(sheet: pd.DataFrame) -> dict:
    """Extract metadata (variable, scenario, region, branch, unit, legend, axis items) from a template sheet."""
    meta = {}
    meta["variable"] = str(sheet.iloc[0, 0]).strip()
    # Scenario and Region are in row 2 (index 1) as "Scenario: X, Region: Y"
    scenario_region = str(sheet.iloc[1, 0])
    for part in scenario_region.split(","):
        if "Scenario:" in part:
            meta["scenario"] = part.split(":", 1)[1].strip()
        if "Region:" in part:
            meta["region"] = part.split(":", 1)[1].strip()
    # Branch row
    branch_line = str(sheet.iloc[2, 0])
    meta["branch"] = branch_line.split(":", 1)[1].strip() if ":" in branch_line else branch_line.strip()
    # Units row
    units_line = str(sheet.iloc[3, 0])
    meta["units"] = units_line.split(":", 1)[1].strip() if ":" in units_line else units_line.strip()
    # Legend label at A6 (index 5, col 0)
    meta["legend_label"] = str(sheet.iloc[5, 0]).strip()
    # Axis (X) items are row 6 (index 5) columns 1+
    axis_items = []
    for val in sheet.iloc[5, 1:]:
        if pd.isna(val):
            continue
        axis_items.append(val)
    meta["x_items"] = axis_items
    # Legend members are col 0 from row 7 (index 6) downward until blank
    legend_members = []
    for val in sheet.iloc[6:, 0]:
        if pd.isna(val) or str(val).strip() == "":
            break
        legend_members.append(str(val).strip())
    meta["legend_members"] = legend_members
    return meta


def parse_template_worksheet(ws) -> dict:
    """Extract template metadata directly from an openpyxl worksheet."""
    meta: dict[str, object] = {}
    meta["variable"] = str(ws.cell(row=1, column=1).value or "").strip()

    scenario_region = str(ws.cell(row=2, column=1).value or "")
    for part in scenario_region.split(","):
        if "Scenario:" in part:
            meta["scenario"] = part.split(":", 1)[1].strip()
        if "Region:" in part:
            meta["region"] = part.split(":", 1)[1].strip()

    branch_line = str(ws.cell(row=3, column=1).value or "")
    meta["branch"] = branch_line.split(":", 1)[1].strip() if ":" in branch_line else branch_line.strip()

    units_line = str(ws.cell(row=4, column=1).value or "")
    meta["units"] = units_line.split(":", 1)[1].strip() if ":" in units_line else units_line.strip()

    meta["legend_label"] = str(ws.cell(row=6, column=1).value or "").strip()

    axis_items: list[object] = []
    for col in range(2, ws.max_column + 1):
        val = ws.cell(row=6, column=col).value
        if val is None or str(val).strip() == "":
            continue
        axis_items.append(val)
    meta["x_items"] = axis_items

    legend_members: list[str] = []
    for row in range(7, ws.max_row + 1):
        val = ws.cell(row=row, column=1).value
        if val is None or str(val).strip() == "":
            break
        legend_members.append(str(val).strip())
    meta["legend_members"] = legend_members
    return meta


def write_table_values_preserve_format(ws, table_df: pd.DataFrame) -> None:
    """Replace worksheet values while preserving existing cell formatting/layout."""
    old_max_row = ws.max_row
    old_max_col = ws.max_column

    for row in range(1, old_max_row + 1):
        for col in range(1, old_max_col + 1):
            ws.cell(row=row, column=col).value = None

    for row_idx, row_vals in enumerate(table_df.itertuples(index=False, name=None), start=1):
        for col_idx, val in enumerate(row_vals, start=1):
            ws.cell(row=row_idx, column=col_idx).value = None if pd.isna(val) else val


def _snapshot_worksheet_values(ws) -> pd.DataFrame:
    """Capture the current worksheet values as a DataFrame for safe fallback writes."""
    rows = ws.max_row
    cols = ws.max_column
    values = [
        [ws.cell(row=row_idx, column=col_idx).value for col_idx in range(1, cols + 1)]
        for row_idx in range(1, rows + 1)
    ]
    return pd.DataFrame(values)


def _data_rows_from_template_table(table_df: pd.DataFrame) -> tuple[list[str], pd.DataFrame]:
    """Return (year columns, data rows) from a template-shaped table."""
    if table_df.empty or table_df.shape[0] < 6:
        return [], pd.DataFrame(columns=["label"])
    header = [str(x).strip() for x in table_df.iloc[5, :].tolist()]
    year_cols = [col for col in header[1:] if col]
    body = table_df.iloc[6:, : len(header)].copy()
    body.columns = header
    label_col = header[0]
    body[label_col] = body[label_col].fillna("").astype(str).str.strip()
    body = body[body[label_col] != ""]
    if body.empty:
        return year_cols, pd.DataFrame(columns=["label"] + year_cols)
    data = body.rename(columns={label_col: "label"})
    for col in year_cols:
        data[col] = pd.to_numeric(data[col], errors="coerce").fillna(0.0)
    return year_cols, data[["label"] + year_cols]


def _build_derived_table_from_children(
    *,
    target_meta: dict[str, object],
    child_tables: list[pd.DataFrame],
) -> pd.DataFrame:
    """Build a template-shaped derived table by summing child tables."""
    if not child_tables:
        return _build_no_results_table(target_meta)

    year_cols = [str(x).strip() for x in list(target_meta.get("x_items", []) or []) if str(x).strip() and str(x).strip().lower() != "total"]
    combined = pd.DataFrame(columns=["label"] + year_cols)
    for child in child_tables:
        _, data = _data_rows_from_template_table(child)
        if data.empty:
            continue
        data = data[data["label"].str.lower() != "total"].copy()
        # keep only years expected on target sheet
        keep_cols = ["label"] + [c for c in year_cols if c in data.columns]
        data = data[keep_cols]
        for col in year_cols:
            if col not in data.columns:
                data[col] = 0.0
        data = data[["label"] + year_cols]
        if combined.empty:
            combined = data.copy()
        else:
            combined = (
                combined.merge(data, on="label", how="outer", suffixes=("_l", "_r"))
                .fillna(0.0)
            )
            rebuilt = pd.DataFrame({"label": combined["label"]})
            for col in year_cols:
                rebuilt[col] = pd.to_numeric(combined.get(f"{col}_l", 0.0), errors="coerce").fillna(0.0) + pd.to_numeric(combined.get(f"{col}_r", 0.0), errors="coerce").fillna(0.0)
            combined = rebuilt

    if combined.empty:
        return _build_no_results_table(target_meta)

    combined = combined.groupby("label", as_index=False)[year_cols].sum()
    total_vals = {col: float(pd.to_numeric(combined[col], errors="coerce").fillna(0.0).sum()) for col in year_cols}
    total_row = pd.DataFrame([{"label": "Total", **total_vals}])
    final_data = pd.concat([combined, total_row], ignore_index=True)

    header = [str(target_meta.get("legend_label", "") or "")] + year_cols
    if any(str(x).strip().lower() == "total" for x in list(target_meta.get("x_items", []) or [])):
        header.append("Total")
        final_data["Total"] = pd.to_numeric(final_data[year_cols], errors="coerce").fillna(0.0).sum(axis=1)
        # For Total row, Total column equals sum across years already.

    meta_rows = [
        [target_meta.get("variable", "")],
        [f"Scenario: {target_meta.get('scenario','')}, Region: {target_meta.get('region','')}"],
        [f"Branch: {target_meta.get('branch','')}"],
        [f"Units: {target_meta.get('units','')}"],
        [""],
    ]
    data_rows = [header]
    for _, row in final_data.iterrows():
        data_rows.append([row.get("label", "")] + [row.get(col, 0.0) for col in header[1:]])
    return pd.DataFrame(meta_rows + data_rows)


def _record_catalog_coverage_row(
    *,
    catalog_rows: list[dict[str, object]],
    catalog_df: pd.DataFrame,
    template_name: str,
    sheet_name: str,
    meta: dict[str, object],
    table_df: pd.DataFrame,
) -> None:
    scope_type, module_or_root = _catalog_scope_from_meta(meta)
    if not (scope_type and module_or_root):
        return
    expected_norm = _expected_fuels_from_catalog(
        catalog_df,
        catalog_type=scope_type,
        module_or_root=module_or_root,
        scenario=str(meta.get("scenario") or ""),
        variable_name=str(meta.get("variable") or ""),
    )
    actual_labels = _extract_table_legend_rows(table_df)
    actual_norm = {
        _normalize_fuel_label(label)
        for label in actual_labels
        if _normalize_fuel_label(label) and _normalize_fuel_label(label) != "total"
    }
    missing_norm = sorted(expected_norm - actual_norm)
    extra_norm = sorted(actual_norm - expected_norm) if expected_norm else []
    catalog_rows.append(
        {
            "template": template_name,
            "sheet": sheet_name,
            "scenario": str(meta.get("scenario") or ""),
            "branch": str(meta.get("branch") or ""),
            "catalog_type": scope_type,
            "module_or_root": module_or_root,
            "expected_fuels_count": int(len(expected_norm)),
            "actual_fuels_count": int(len(actual_norm)),
            "missing_fuels_count": int(len(missing_norm)),
            "extra_fuels_count": int(len(extra_norm)),
            "missing_fuels": "; ".join(missing_norm),
            "extra_fuels": "; ".join(extra_norm),
        }
    )
    if expected_norm and missing_norm:
        print(
            f"[WARN] Catalog fuel coverage gap for {template_name}/{sheet_name}: "
            f"{len(missing_norm)} expected fuel(s) missing from table.",
            flush=True,
        )


def _resolve_existing_branch_path(app, branch_path: str) -> str:
    """Return a live branch path, raising cleanly when the template path is stale."""
    requested = str(branch_path or "").strip()
    if not requested:
        raise ValueError("Branch path is required")
    branches = app.Branches
    try:
        exists = bool(branches.Exists(requested))
    except Exception:
        exists = False
    if exists:
        return requested
    raise ValueError(
        f"No LEAP branch found for '{requested}'. Update the template workbook branch metadata."
    )


def _resolve_branch_variable(
    app,
    branch_path: str,
    variable_name: str | None,
    *,
    allow_substitution: bool = False,
):
    """Return the requested LEAP variable object for a branch.

    When ``allow_substitution`` is False, this function raises instead of silently
    switching the requested measure to another variable.
    """
    resolved_branch_path = _resolve_existing_branch_path(app, branch_path)
    branch_obj = app.Branches.Item(resolved_branch_path)
    requested = str(variable_name or "").strip()
    variables = branch_obj.Variables
    var_count = int(getattr(variables, "Count", 0))
    if var_count <= 0:
        raise ValueError(f"No variables available on branch '{resolved_branch_path}'")

    # Build a case-insensitive lookup first so we avoid triggering LEAP errors
    # from invalid Variables.Item(name) calls.
    by_name: dict[str, int] = {}
    for idx in range(1, var_count + 1):
        name = str(variables.Item(idx).Name).strip()
        if name:
            by_name.setdefault(name.lower(), idx)

    if requested:
        req_norm = requested.lower()
        requested_idx = by_name.get(req_norm)
        if requested_idx is not None:
            return variables.Item(requested_idx)

        available_names = sorted({str(variables.Item(i).Name).strip() for i in by_name.values()})
        available = ", ".join(available_names)
        if allow_substitution:
            candidate_names = [requested]
            if req_norm == "inputs":
                # Some transformation results branches expose fuel-disaggregated
                # inputs only through feedstock/output result variables.
                candidate_names.extend(["Outputs by Feedstock Fuel", "Outputs by Output Fuel"])
            elif req_norm == "outputs by feedstock fuel":
                candidate_names.append("Inputs")

            for candidate in candidate_names:
                idx = by_name.get(str(candidate).strip().lower())
                if idx is not None:
                    if str(candidate).strip().lower() != req_norm:
                        print(
                            f"Info: variable '{requested}' not found on branch '{resolved_branch_path}'; "
                            f"using '{candidate}' instead.",
                            flush=True,
                        )
                    return variables.Item(idx)

            print(
                f"Warning: variable '{requested}' not found on branch '{resolved_branch_path}'; "
                f"available variables: {available or '<none>'}. Using first variable instead.",
                flush=True,
            )
            return variables.Item(1)

        raise MissingRequestedVariableError(
            f"Requested variable '{requested}' not found on branch '{resolved_branch_path}'. "
            f"Available variables: {available or '<none>'}."
        )
    return variables.Item(1)


def _preferred_fuel_dimension_names(meta: dict) -> list[str]:
    """Return likely LEAP filter dimension names for a fuel-oriented table."""
    variable = str(meta.get("variable") or "").strip().lower()
    legend = str(meta.get("legend_label") or "").strip()
    preferred: list[str] = []
    if "feedstock" in variable or variable == "inputs":
        preferred.append("Feedstock Fuel")
    if "output fuel" in variable:
        preferred.append("Output Fuel")
    if legend:
        preferred.append(legend)
    preferred.extend(["Fuel", "Feedstock Fuel", "Output Fuel"])

    seen: set[str] = set()
    out: list[str] = []
    for name in preferred:
        key = name.strip().lower()
        if not key or key in seen:
            continue
        seen.add(key)
        out.append(name)
    return out


def _is_supply_energy_table(meta: dict[str, object]) -> bool:
    """Return True for supply/resource result sheets that should avoid Results-view variable switching."""
    variable = str(meta.get("variable") or "").strip()
    branch = str(meta.get("branch") or "").strip().lower()
    return (
        variable in {"Indigenous Production", "Imports", "Exports"}
        and (
            "supply" in branch
            or branch.startswith("resources\\")
            or branch.startswith("resources/")
        )
    )


def _is_transformation_energy_table(meta: dict[str, object]) -> bool:
    """Return True for transformation result sheets that still require rendered CSV export."""
    variable = str(meta.get("variable") or "").strip()
    branch = str(meta.get("branch") or "").strip().lower()
    return (
        variable in {"Inputs", "Outputs by Feedstock Fuel", "Outputs by Output Fuel"}
        and "transformation" in branch
    )


def _is_transformation_input_style_sheet(template_path: Path, sheet_name: str) -> bool:
    """True when this sheet should be driven by a same-name LEAP favorite."""
    if "transformation_results_" not in str(template_path.name).lower():
        return False
    sheet = str(sheet_name or "").strip().lower()
    return any(sheet.endswith(suffix) for suffix in TRANSFORMATION_INPUT_STYLE_FAVORITE_SUFFIXES)


def _expected_input_type_qualifier_for_sheet(sheet_name: str) -> str | None:
    """
    Expected Results qualifier for transformation input-style sheets.

    Returned text is matched against LEAP CSV metadata line 2
    (Scenario/Region/... qualifier).
    """
    name = str(sheet_name or "").strip().lower()
    if name.endswith("_feed_inputs"):
        return "Feedstock Fuels"
    if name.endswith("_aux"):
        return "Auxiliary Fuels From Outputs"
    if name.endswith("_aux_other"):
        return "Auxiliary Fuels From Other Modules"
    return None


def _favorite_candidates_for_sheet(sheet_name: str) -> list[str]:
    """
    Return candidate favorite names for a sheet.

    Strict mode now expects exact canonical names.
    """
    name = str(sheet_name or "").strip()
    if not name:
        return []
    return [name]


def _favorite_sheet_stem(name: str) -> str:
    """Best-effort stem used to resolve sheet/favorite truncation variants."""
    text = str(name or "").strip().lower()
    suffixes = (
        "_feed_inputs",
        "_aux_other",
        "_aux",
        "_out_feed",
        "_out_fuel",
        "_inputs",
    )
    for suffix in suffixes:
        if text.endswith(suffix):
            return text[: -len(suffix)]
    return text


def _favorite_suffix_family(name: str) -> str:
    text = str(name or "").strip().lower()
    if text.endswith("_feed_inputs"):
        return "feed_inputs"
    if text.endswith("_aux"):
        return "aux"
    if text.endswith("_aux_other"):
        return "aux_other"
    if text.endswith("_out_feed"):
        return "out_feed"
    if text.endswith("_out_fuel"):
        return "out_fuel"
    if text.endswith("_inputs"):
        return "inputs"
    return "other"


def _list_favorite_names(app) -> list[str]:
    names: list[str] = []
    try:
        favorites = app.Favorites
        count = int(getattr(favorites, "Count", 0))
        for idx in range(1, count + 1):
            try:
                fav = favorites.Item(idx)
                name = str(getattr(fav, "Name", "")).strip()
            except Exception:
                name = ""
            if name:
                names.append(name)
    except Exception:
        return []
    return names


def resolve_favorite_name_for_sheet(app, sheet_name: str) -> tuple[str | None, str]:
    """
    Resolve a favorite name for a worksheet safely.

    Strategy:
    1) Exact case-insensitive name
    2) Candidate variants from known truncation patterns
    3) Same-stem favorites with strong prefix/fuzzy score
    """
    requested = str(sheet_name or "").strip()
    if not requested:
        return None, "empty_sheet_name"

    available = _list_favorite_names(app)
    if not available:
        return None, "no_favorites_available"

    req_lower = requested.lower()
    by_lower = {name.lower(): name for name in available}
    if req_lower in by_lower:
        return by_lower[req_lower], "exact"

    for candidate in _favorite_candidates_for_sheet(requested):
        key = candidate.lower()
        if key in by_lower:
            return by_lower[key], "candidate_variant"

    mode = str(FAVORITE_RESOLUTION_MODE or "strict").strip().lower()
    if mode in {"stem", "fuzzy"}:
        req_stem = _favorite_sheet_stem(requested)
        req_family = _favorite_suffix_family(requested)
        same_stem = [
            name for name in available
            if _favorite_sheet_stem(name) == req_stem
            and _favorite_suffix_family(name) == req_family
        ]
        if len(same_stem) == 1:
            return same_stem[0], "same_stem_unique"

    if mode == "fuzzy":
        # Global high-confidence fuzzy fallback (opt-in only)
        fuzzy = difflib.get_close_matches(requested, available, n=1, cutoff=0.93)
        if fuzzy:
            return fuzzy[0], "global_fuzzy"
    return None, "unresolved"


def activate_favorite_for_sheet(app, sheet_name: str) -> tuple[str | None, str, str]:
    """
    Resolve and activate a favorite for sheet_name.

    Returns: (resolved_name, resolution_mode, status_message)
    """
    resolved, mode = resolve_favorite_name_for_sheet(app, sheet_name)
    if not resolved:
        return None, mode, f"Favorite unresolved for sheet '{sheet_name}' (mode={mode})"
    status = activate_favorite(app, resolved)
    return resolved, mode, str(status or "")


def validate_transformation_sheet_naming(sheet_names: list[str], *, context: str = "") -> list[str]:
    """
    Enforce canonical transformation sheet suffixes.

    Input-style canonical suffixes:
    - _feed_inputs
    - _aux
    - _aux_other
    """
    issues: list[str] = []
    for name in sheet_names:
        lower = str(name or "").strip().lower()
        if not lower:
            continue
        if lower.endswith(("_aux_outputs", "_aux_other_outputs", "_aux_other_out")):
            issues.append(f"{name}: legacy/truncated aux suffix detected")
    if issues and context:
        issues = [f"{context} | {msg}" for msg in issues]
    return issues


def _load_expected_sheets_contract(config_path: Path) -> dict[str, list[str]]:
    if not config_path.exists():
        raise FileNotFoundError(f"Expected-sheets contract not found: {config_path}")
    raw = json.loads(config_path.read_text(encoding="utf-8"))
    if not isinstance(raw, dict):
        raise ValueError("Expected-sheets contract must be an object keyed by workbook filename.")
    out: dict[str, list[str]] = {}
    for workbook_name, spec in raw.items():
        sheets: list[str] = []
        if isinstance(spec, dict):
            items = spec.get("sheets", [])
            if isinstance(items, list):
                sheets = [str(x) for x in items]
        elif isinstance(spec, list):
            sheets = [str(x) for x in spec]
        if not sheets:
            raise ValueError(f"Expected-sheets contract entry has no sheets: {workbook_name}")
        out[str(workbook_name)] = sheets
    return out


def _validate_expected_sheets_for_workbook(
    *,
    workbook_name: str,
    actual_sheet_names: list[str],
    expected_by_workbook: dict[str, list[str]],
) -> tuple[list[str], list[str], list[str]]:
    if workbook_name not in expected_by_workbook:
        raise RuntimeError(
            f"Workbook '{workbook_name}' missing from expected-sheets contract: "
            f"{EXPECTED_SHEETS_CONFIG_PATH}"
        )
    expected = list(expected_by_workbook[workbook_name])
    actual = list(actual_sheet_names)
    optional_sheets = OPTIONAL_EXPECTED_SHEETS_BY_WORKBOOK.get(workbook_name, set())
    if optional_sheets:
        expected = [s for s in expected if s not in optional_sheets]
        actual = [s for s in actual if s not in optional_sheets]
    missing = [s for s in expected if s not in actual]
    unexpected = [s for s in actual if s not in expected]
    return expected, missing, unexpected


def _is_supply_or_transformation_energy_table(meta: dict[str, object]) -> bool:
    """Return True for energy result sheets that need extra collapse validation."""
    return _is_supply_energy_table(meta) or _is_transformation_energy_table(meta)


def _merge_legend_members(template_members: list[str], live_members: list[str]) -> list[str]:
    """Merge template and live legend members, preferring live LEAP order while retaining template-only fallbacks."""
    merged: list[str] = []
    seen: set[str] = set()

    def _append(member: object) -> None:
        text = str(member or "").strip()
        if not text:
            return
        key = text.lower()
        if key == "total":
            return
        if key in seen:
            return
        seen.add(key)
        merged.append(text)

    for member in live_members:
        _append(member)
    for member in template_members:
        _append(member)
    merged.append("Total")
    return merged


_RESULTS_UNIT_PREFIX_SCALE = {
    "thousand": 1e3,
    "million": 1e6,
    "billion": 1e9,
}

_RESULTS_UNIT_BASE_SCALE = {
    "joule": 1.0,
    "joules": 1.0,
    "gigajoule": 1e9,
    "gigajoules": 1e9,
    "petajoule": 1e15,
    "petajoules": 1e15,
    "watt": 1.0,
    "watts": 1.0,
    "kilowatt": 1e3,
    "kilowatts": 1e3,
    "megawatt": 1e6,
    "megawatts": 1e6,
    "gigawatt": 1e9,
    "gigawatts": 1e9,
}


def _normalize_unit_text(unit_text: object) -> str:
    return " ".join(str(unit_text or "").strip().lower().split())


def _results_unit_to_si_scale(unit_text: object) -> float | None:
    normalized = _normalize_unit_text(unit_text)
    if not normalized:
        return None
    prefix_scale = 1.0
    base_unit = normalized
    parts = normalized.split(" ", 1)
    if len(parts) == 2 and parts[0] in _RESULTS_UNIT_PREFIX_SCALE:
        prefix_scale = _RESULTS_UNIT_PREFIX_SCALE[parts[0]]
        base_unit = parts[1]
    base_scale = _RESULTS_UNIT_BASE_SCALE.get(base_unit)
    if base_scale is None:
        return None
    return prefix_scale * base_scale


def _conversion_factor_between_results_units(from_unit: object, to_unit: object) -> float | None:
    from_norm = _normalize_unit_text(from_unit)
    to_norm = _normalize_unit_text(to_unit)
    if not from_norm or not to_norm:
        return None
    if from_norm == to_norm:
        return 1.0
    from_scale = _results_unit_to_si_scale(from_norm)
    to_scale = _results_unit_to_si_scale(to_norm)
    if from_scale is None or to_scale is None:
        return None
    return float(from_scale) / float(to_scale)


def _read_exported_results_csv_unit(csv_path: Path) -> str:
    """Extract the 'Units: ...' line from a LEAP Results CSV export."""
    try:
        with csv_path.open("r", encoding="utf-8-sig", newline="") as f:
            reader = csv.reader(f)
            for idx, row in enumerate(reader):
                if idx > 5:
                    break
                if not row:
                    continue
                first = str(row[0] or "").strip()
                if first.lower().startswith("units:"):
                    return first.split(":", 1)[1].strip()
    except Exception:
        return ""
    return ""


def _read_exported_results_csv_metadata(csv_path: Path) -> dict[str, str]:
    """Read leading LEAP CSV metadata lines (variable/scenario/branch/units)."""
    meta = {"variable": "", "scenario_region": "", "branch": "", "units": ""}
    try:
        with csv_path.open("r", encoding="utf-8-sig", newline="") as f:
            reader = csv.reader(f)
            rows: list[list[str]] = []
            for idx, row in enumerate(reader):
                rows.append(row)
                if idx >= 4:
                    break
        if len(rows) >= 1 and rows[0]:
            meta["variable"] = str(rows[0][0] or "").strip()
        if len(rows) >= 2 and rows[1]:
            meta["scenario_region"] = str(rows[1][0] or "").strip()
        if len(rows) >= 3 and rows[2]:
            meta["branch"] = str(rows[2][0] or "").strip()
        if len(rows) >= 4 and rows[3]:
            meta["units"] = str(rows[3][0] or "").strip()
    except Exception:
        return meta
    return meta


def _extract_member_name(member_obj) -> str:
    """Best-effort member name extraction across COM collection shapes."""
    for attr in ("Name", "name", "Label", "label"):
        try:
            val = getattr(member_obj, attr)
            text = str(val).strip()
            if text:
                return text
        except Exception:
            continue
    text = str(member_obj).strip()
    return text


def discover_legend_members_from_api(
    app,
    legend_label: str | None,
    preferred_dimension_names: list[str] | None = None,
) -> tuple[str | None, list[str]]:
    """Best-effort discovery of live legend members (e.g., fuels) from LEAP API.

    Returns (resolved_dimension_name, members).
    """
    label = str(legend_label or "").strip().lower()
    if not label:
        return None, []
    try:
        dims = app.Dimensions
        dim_count = int(getattr(dims, "Count", 0))
    except Exception:
        return None, []

    dim_candidates: list[tuple[str, object]] = []
    for i in range(1, dim_count + 1):
        try:
            dim = dims.Item(i)
            raw_name = str(getattr(dim, "Name", "")).strip()
            dim_name = raw_name.lower()
        except Exception:
            continue
        dim_candidates.append((raw_name, dim))

    target_dim = None
    target_dim_name: str | None = None
    preferred = [str(name).strip().lower() for name in (preferred_dimension_names or []) if str(name).strip()]
    if label and label not in preferred:
        preferred.append(label)

    for preferred_name in preferred:
        for raw_name, dim in dim_candidates:
            if raw_name.strip().lower() == preferred_name:
                target_dim = dim
                target_dim_name = raw_name
                break
        if target_dim is not None:
            break

    if target_dim is None:
        for preferred_name in preferred:
            for raw_name, dim in dim_candidates:
                if preferred_name in raw_name.strip().lower():
                    target_dim = dim
                    target_dim_name = raw_name
                    break
            if target_dim is not None:
                break

    if target_dim is None and label == "fuel":
        # Common mismatch: legend displays "Fuel" while underlying dimension is
        # "Output Fuel" / "Feedstock Fuel" etc.
        for raw_name, dim in dim_candidates:
            if "fuel" in raw_name.strip().lower():
                target_dim = dim
                target_dim_name = raw_name
                break

    if target_dim is not None:
        dim = target_dim
        dim_name = str(target_dim_name or "").strip()
        if not dim_name:
            dim_name = str(legend_label or "").strip()
        if not dim_name:
            return None, []

        member_names: list[str] = []
        # Common COM patterns: .Members collection with .Count and .Item(index)
        try:
            members = getattr(dim, "Members")
            m_count = int(getattr(members, "Count", 0))
            for j in range(1, m_count + 1):
                name = _extract_member_name(members.Item(j))
                if name:
                    member_names.append(name)
        except Exception:
            pass

        # Fallback patterns: iterate directly over dim or dim.Items
        if not member_names:
            try:
                for obj in dim:
                    name = _extract_member_name(obj)
                    if name:
                        member_names.append(name)
            except Exception:
                pass
        if not member_names:
            try:
                items = getattr(dim, "Items")
                m_count = int(getattr(items, "Count", 0))
                for j in range(1, m_count + 1):
                    name = _extract_member_name(items.Item(j))
                    if name:
                        member_names.append(name)
            except Exception:
                pass

        # Keep order, remove duplicates, and skip explicit total labels.
        seen: set[str] = set()
        out: list[str] = []
        for name in member_names:
            key = name.strip().lower()
            if not key or key == "total" or key in seen:
                continue
            seen.add(key)
            out.append(name.strip())
        return dim_name, out

    return None, []


def build_fresh_table_from_export(
    app,
    meta: dict,
    *,
    export_csv_path: Path,
    scale_spec: dict | None = None,
    context: str = "",
    allow_variable_substitution: bool = False,
    trust_active_results_view: bool = False,
    expected_input_type_qualifier: str | None = None,
) -> pd.DataFrame:
    """Fetch a full Results table via LEAP CSV export for one sheet context."""
    resolved_branch = _run_rendered_csv_export_step(
        context,
        "resolve_branch",
        lambda: _resolve_existing_branch_path(app, str(meta.get("branch") or "")),
    )
    meta = dict(meta)
    meta["branch"] = resolved_branch
    _run_rendered_csv_export_step(
        context,
        "show_results_view_table",
        lambda: _show_results_view_table(app),
    )
    _run_rendered_csv_export_step(
        context,
        f"set_axes x=Years legend={meta.get('legend_label') or ''}",
        lambda: set_axes(app, x_axis="Years", legend=meta.get("legend_label") or None),
    )
    _run_rendered_csv_export_step(
        context,
        f"set_context scenario={meta.get('scenario') or ''} region={meta.get('region') or ''} branch={meta.get('branch') or ''}",
        lambda: set_context(
            app,
            scenario=meta.get("scenario"),
            region=meta.get("region"),
            branch_path=meta.get("branch"),
        ),
    )
    variable_obj = None
    target_variable_name = ""
    if not trust_active_results_view:
        variable_obj = _run_rendered_csv_export_step(
            context,
            f"resolve_variable variable={meta.get('variable') or ''}",
            lambda: _resolve_branch_variable(
                app,
                str(meta.get("branch") or ""),
                meta.get("variable"),
                allow_substitution=allow_variable_substitution,
            ),
        )
        if PRECHECK_NO_RESULTS_WITH_RESULTVALUE:
            probe_result = _probe_no_results_via_value_rs(
                app,
                meta=meta,
                variable_obj=variable_obj,
                context=context,
            )
            if probe_result is False:
                _log_rendered_csv_export_step(context, "skipping rendered export no results available")
                return _build_no_results_table(meta)
        target_variable_name = str(getattr(variable_obj, "Name", "") or "").strip()
    else:
        _log_rendered_csv_export_step(
            context,
            "trust_active_results_view=True; skipping variable resolve/ValueRS probe",
        )
    export_path_abs = export_csv_path.resolve()
    exported_unit = ""
    exported_df = pd.DataFrame()
    exported_meta: dict[str, str] = {}
    mismatch_error = ""
    for attempt in range(2):
        if variable_obj is not None:
            _run_rendered_csv_export_step(
                context,
                f"set_active_variable {target_variable_name or '<blank>'} (attempt {attempt + 1})",
                lambda: _set_active_variable(app, variable_obj),
            )
        _run_rendered_csv_export_step(
            context,
            f"refresh_results_view_table (attempt {attempt + 1})",
            lambda: _show_results_view_table(app),
        )
        _run_rendered_csv_export_step(
            context,
            f"export_results_csv_file path={export_path_abs} (attempt {attempt + 1})",
            lambda: _export_results_csv_file(app, export_path_abs),
        )
        exported_unit = _run_rendered_csv_export_step(
            context,
            "read_exported_unit",
            lambda: _read_exported_results_csv_unit(export_path_abs),
        )
        exported_df = _run_rendered_csv_export_step(
            context,
            "parse_exported_results_csv",
            lambda: parse_exported_results_csv(export_path_abs),
        )
        exported_meta = _run_rendered_csv_export_step(
            context,
            "read_exported_metadata",
            lambda: _read_exported_results_csv_metadata(export_path_abs),
        )
        _log_rendered_csv_export_step(
            context,
            f"parsed_csv rows={len(exported_df)} cols={len(exported_df.columns)} unit={exported_unit or '<blank>'}",
        )

        requested_variable = str(meta.get("variable") or "").strip().lower()
        exported_variable = str(exported_meta.get("variable") or "").strip().lower()
        requested_branch = str(meta.get("branch") or "").strip().lower()
        exported_branch_line = str(exported_meta.get("branch") or "").strip().lower()
        exported_scenario_region = str(exported_meta.get("scenario_region") or "").strip().lower()
        requested_scenario = str(meta.get("scenario") or "").strip().lower()
        requested_region = str(meta.get("region") or "").strip().lower()

        mismatch_error = ""
        if requested_scenario and requested_scenario not in exported_scenario_region:
            mismatch_error = (
                "Rendered LEAP CSV scenario mismatch: "
                f"requested='{meta.get('scenario')}', exported_scenario_region='{exported_meta.get('scenario_region')}'."
            )
        elif requested_region and requested_region not in exported_scenario_region:
            mismatch_error = (
                "Rendered LEAP CSV region mismatch: "
                f"requested='{meta.get('region')}', exported_scenario_region='{exported_meta.get('scenario_region')}'."
            )
        elif (not trust_active_results_view) and requested_variable and exported_variable and requested_variable != exported_variable:
            mismatch_error = (
                "Rendered LEAP CSV variable mismatch: "
                f"requested='{meta.get('variable')}', exported='{exported_meta.get('variable')}'."
            )
        elif requested_branch and exported_branch_line and requested_branch not in exported_branch_line:
            mismatch_error = (
                "Rendered LEAP CSV branch mismatch: "
                f"requested='{meta.get('branch')}', exported_line='{exported_meta.get('branch')}'."
            )
        elif expected_input_type_qualifier:
            expected_lower = str(expected_input_type_qualifier).strip().lower()
            if expected_lower and expected_lower not in exported_scenario_region:
                mismatch_error = (
                    "Rendered LEAP CSV input-type qualifier mismatch: "
                    f"expected='{expected_input_type_qualifier}', "
                    f"exported_scenario_region='{exported_meta.get('scenario_region')}'."
                )

        if not mismatch_error:
            break
        _log_rendered_csv_export_step(
            context,
            f"metadata_mismatch attempt {attempt + 1}: {mismatch_error}",
        )

    if mismatch_error:
        policy = str(RENDERED_CSV_METADATA_MISMATCH_POLICY or "raise").strip().lower()
        detail = (
            f"{mismatch_error} requested_variable='{meta.get('variable')}', "
            f"requested_branch='{meta.get('branch')}', "
            f"exported_variable='{exported_meta.get('variable')}', "
            f"exported_branch='{exported_meta.get('branch')}'."
        )
        if policy == "empty_table":
            _RENDERED_CSV_METADATA_MISMATCH_EVENTS.append(
                {
                    "context": str(context or ""),
                    "requested_variable": str(meta.get("variable") or ""),
                    "requested_branch": str(meta.get("branch") or ""),
                    "exported_variable": str(exported_meta.get("variable") or ""),
                    "exported_branch": str(exported_meta.get("branch") or ""),
                    "detail": str(detail),
                }
            )
            print(
                f"[WARN][CSV_METADATA_MISMATCH][EMPTY_TABLE] {context}: {detail} "
                "Writing no-results table and continuing.",
                flush=True,
            )
            return _build_no_results_table(meta)
        raise RuntimeError(
            f"{detail} This usually indicates stale Results filters/favorites in LEAP."
        )

    # Rebuild as table-like rows so it matches template layout expectations.
    data_df = pd.DataFrame([list(exported_df.columns)] + exported_df.values.tolist())

    target_unit = scale_spec.get("target_unit") if scale_spec else None
    scale_factor = None
    if target_unit:
        scale_factor = _conversion_factor_between_results_units(exported_unit, target_unit)
        if scale_factor is None and scale_spec is not None:
            scale_factor = scale_spec.get("scale_factor")
            print(
                f"Warning: could not convert CSV export units '{exported_unit}' to '{target_unit}' directly; "
                f"falling back to configured scale factor {scale_factor}.",
                flush=True,
            )
    elif scale_spec is not None:
        scale_factor = scale_spec.get("scale_factor")
    if scale_factor is not None and data_df.shape[1] > 1:
        value_cols = list(data_df.columns[1:])
        for col in value_cols:
            data_df.loc[1:, col] = pd.to_numeric(data_df.loc[1:, col], errors="coerce") * float(scale_factor)

    meta_rows = [
        [meta.get("variable", "")],
        [f"Scenario: {meta.get('scenario','')}, Region: {meta.get('region','')}"],
        [f"Branch: {meta.get('branch','')}"],
        [f"Units: {target_unit or exported_unit or meta.get('units','')}"],
        [""],
    ]
    final_df = pd.DataFrame(meta_rows + data_df.values.tolist())
    return final_df


def build_fresh_table(
    app,
    meta: dict,
    scale_spec: dict | None = None,
    *,
    allow_variable_substitution: bool = False,
) -> pd.DataFrame:
    """Fetch values for the given template metadata using LEAP direct ValueRS calls."""
    resolved_branch = _resolve_existing_branch_path(app, str(meta.get("branch") or ""))
    meta = dict(meta)
    meta["branch"] = resolved_branch
    legend_label = meta["legend_label"]
    # Set context
    set_axes(app, x_axis="Years", legend=legend_label or None)
    set_context(
        app,
        scenario=meta.get("scenario"),
        region=meta.get("region"),
        # Do not set ActiveUnit/ActiveVariable here; we will use defaults to avoid unit/variable name mismatches.
        branch_path=meta.get("branch"),
    )
    variable_obj = _resolve_branch_variable(
        app,
        str(meta["branch"]),
        meta.get("variable"),
        allow_substitution=allow_variable_substitution,
    )

    x_items = meta["x_items"]
    filter_dim_label = legend_label
    legend_members = list(meta["legend_members"])
    if VALUES_DISCOVER_LEGEND_MEMBERS_FROM_API:
        resolved_dim_name, live_members = discover_legend_members_from_api(
            app,
            legend_label,
            preferred_dimension_names=_preferred_fuel_dimension_names(meta),
        )
        if resolved_dim_name:
            filter_dim_label = resolved_dim_name
        if live_members:
            legend_members = _merge_legend_members(legend_members, live_members)

    # Build data rows: first row is header
    header = [legend_label] + x_items
    rows = [header]
    for member in legend_members:
        row = [member]
        filter_str = f"{filter_dim_label}={member}" if filter_dim_label else ""
        for year in x_items:
            try:
                # Use default units by passing empty string to avoid unit-name mismatches (e.g., "Petajoules" vs LEAP unit code).
                val = variable_obj.ValueRS(meta["region"], meta["scenario"], int(year), "", filter_str)
            except Exception:
                val = float("nan")
            row.append(val)
        if VALUES_DROP_ALL_ZERO_COMPONENT_ROWS and str(member).strip().lower() != "total":
            series = pd.to_numeric(pd.Series(row[1:]), errors="coerce").fillna(0.0)
            if float(series.abs().sum()) == 0.0:
                continue
        rows.append(row)
    data_df = pd.DataFrame(rows)

    # Apply optional scaling to numeric data cells only.
    # Do NOT scale the header row (years), or year labels become values like 2.022e-12.
    scale_factor = scale_spec.get("scale_factor") if scale_spec else None
    target_unit = scale_spec.get("target_unit") if scale_spec else None
    if scale_factor is not None:
        value_cols = list(data_df.columns[1:])  # first column is legend labels
        for col in value_cols:
            data_df.loc[1:, col] = pd.to_numeric(data_df.loc[1:, col], errors="coerce") * float(scale_factor)

    # Prepend the metadata rows to match template structure
    meta_rows = [
        [meta["variable"]],
        [f"Scenario: {meta.get('scenario','')}, Region: {meta.get('region','')}"],
        [f"Branch: {meta.get('branch','')}"],
        [f"Units: {target_unit or meta.get('units','')}"],
        [""],
    ]
    final_df = pd.DataFrame(meta_rows + data_df.values.tolist())
    return final_df


def parse_exported_results_csv(csv_path: Path) -> pd.DataFrame:
    """Parse LEAP-exported CSV that may contain variable-width metadata rows."""
    rows: list[list[str]] = []
    with csv_path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f)
        for row in reader:
            # trim trailing empties but keep interior empties
            while row and str(row[-1]).strip() == "":
                row.pop()
            rows.append([str(c) for c in row])

    if not rows:
        return pd.DataFrame()

    def _score_header(r: list[str]) -> int:
        if len(r) < 3:
            return -1
        numeric = 0
        for cell in r[1:]:
            s = str(cell).strip()
            if not s:
                continue
            try:
                float(s)
                numeric += 1
            except Exception:
                if s.lower() == "total":
                    numeric += 1
        return numeric

    best_idx = max(range(len(rows)), key=lambda i: _score_header(rows[i]))
    if _score_header(rows[best_idx]) <= 0:
        # Fallback: use widest row as header if no numeric-looking header found.
        best_idx = max(range(len(rows)), key=lambda i: len(rows[i]))

    header = rows[best_idx]
    data_rows = []
    for r in rows[best_idx + 1 :]:
        if not r or str(r[0]).strip() == "":
            break
        if len(r) < len(header):
            r = r + [""] * (len(header) - len(r))
        data_rows.append(r[: len(header)])

    if not data_rows:
        return pd.DataFrame(columns=header)
    return pd.DataFrame(data_rows, columns=header)


def is_suspicious_values_table(table_df: pd.DataFrame) -> bool:
    """Flag ValueRS-filled tables where component rows look invalid vs total."""
    if table_df.shape[0] < 7 or table_df.shape[1] < 3:
        return False
    body = table_df.iloc[5:, :].copy()
    if body.empty:
        return False

    header = body.iloc[0].tolist()
    year_cols: list[int] = []
    for idx, val in enumerate(header[1:], start=1):
        try:
            float(val)
            year_cols.append(idx)
        except Exception:
            continue
    if not year_cols:
        return False

    data = body.iloc[1:, [0] + year_cols].copy()
    data.columns = ["label"] + [str(header[i]) for i in year_cols]
    labels = data["label"].astype(str).str.strip()
    valid_mask = labels != ""
    if valid_mask.sum() == 0:
        return False
    data = data.loc[valid_mask].reset_index(drop=True)
    labels = data["label"].astype(str).str.strip()
    vals = data.iloc[:, 1:].apply(pd.to_numeric, errors="coerce").fillna(0.0)

    total_mask = labels.str.lower().eq("total")
    if not total_mask.any():
        return False
    total_vals = vals.loc[total_mask].iloc[0]
    total_nonzero = bool(total_vals.abs().sum() > 0)
    comp_vals = vals.loc[~total_mask]
    if comp_vals.empty:
        # A nonzero Total-only table means the legend breakout collapsed entirely.
        return total_nonzero

    nonzero_comp_rows = int((comp_vals.abs().sum(axis=1) > 0).sum())
    duplicate_total_components = int(comp_vals.apply(lambda r: r.equals(total_vals), axis=1).sum())

    # Suspicious patterns observed in bad outputs:
    # 1) Total has values but no component rows do.
    # 2) Multiple component rows are exactly identical to total row.
    if total_nonzero and nonzero_comp_rows == 0:
        return True
    if duplicate_total_components >= 2:
        return True
    return False


def has_only_total_row(table_df: pd.DataFrame) -> bool:
    """Return True when the table contains a header and only a single Total row."""
    if table_df.shape[0] < 7 or table_df.shape[1] < 2:
        return False
    body = table_df.iloc[5:, :].copy()
    if body.empty:
        return False

    data = body.iloc[1:, 0].fillna("").astype(str).str.strip()
    data = data[data.ne("")]
    if data.empty:
        return False
    return data.str.lower().eq("total").all()


def validate_component_rows_present(table_df: pd.DataFrame, *, context: str = "") -> None:
    """
    Fail when a nonzero Total row exists but all component rows are zero/blank.

    This should never be written to the LEAP result workbooks for split-fuel
    tables such as outputs by feedstock/output fuel, because it means the data
    was not broken out into the expected fuel categories.
    """
    if not is_suspicious_values_table(table_df):
        return
    exception_reason = _lookup_suspicious_table_exception(context)
    if exception_reason:
        print(
            f"Warning: allowed suspicious LEAP results table for {context}: {exception_reason}",
            flush=True,
        )
        return
    suffix = f" for {context}" if context else ""
    raise RuntimeError(
        "Suspicious LEAP results table detected"
        f"{suffix}: nonzero Total row without valid component rows, or repeated component rows equal to Total."
    )


def validate_refilled_table(table_df: pd.DataFrame, *, meta: dict[str, object], context: str = "") -> None:
    """
    Fail fast when a supply/transformation energy table collapses to a Total-only layout.

    These template workbooks are overwritten in place. Once a sheet loses its
    component rows, later refills inherit `legend_members=['Total']` from the
    damaged template and silently keep the degraded layout.
    """
    validate_component_rows_present(table_df, context=context)

    if not _is_supply_or_transformation_energy_table(meta):
        return
    if not has_only_total_row(table_df):
        return

    suffix = f" for {context}" if context else ""
    raise RuntimeError(
        "Refilled LEAP results table collapsed to a Total-only layout"
        f"{suffix}. This usually means live legend-member discovery failed and the in-place template would lose "
        "its fuel rows."
    )


def _normalize_fuel_label(text: object) -> str:
    """Normalize fuel labels for catalog-vs-table comparison."""
    token = str(text or "").strip().lower()
    if not token:
        return ""
    return " ".join(token.split())


def _extract_table_legend_rows(table_df: pd.DataFrame) -> list[str]:
    """Extract row labels from a refilled LEAP table body (first column)."""
    if table_df.shape[0] < 7 or table_df.shape[1] < 1:
        return []
    body_labels = table_df.iloc[6:, 0].fillna("").astype(str).str.strip().tolist()
    labels: list[str] = []
    for label in body_labels:
        if not label:
            break
        labels.append(label)
    return labels


def _load_fuel_catalog(path: Path) -> pd.DataFrame:
    """Load the transformation/supply fuel catalog CSV when available."""
    df = fuel_catalog_preflight.load_fuel_catalog(path)
    if df.empty:
        return pd.DataFrame()
    # Align normalization to template matching in this workflow.
    df = df.copy()
    df["fuel_name_norm"] = df["fuel_name"].map(_normalize_fuel_label)
    return df


def _expected_fuels_from_catalog(
    catalog_df: pd.DataFrame,
    *,
    catalog_type: str,
    module_or_root: str,
    scenario: str,
    variable_name: str | None = None,
) -> set[str]:
    """Return expected normalized fuel labels for a catalog scope."""
    if catalog_df.empty:
        return set()
    subset = catalog_df[
        catalog_df["catalog_type"].str.lower() == str(catalog_type or "").strip().lower()
    ].copy()
    if subset.empty:
        return set()
    subset = subset[
        subset["module_or_root_norm"] == str(module_or_root or "").strip().lower()
    ].copy()
    if subset.empty:
        return set()

    scenario_norm = str(scenario or "").strip().lower()
    by_scenario = subset[subset["scenario_norm"] == scenario_norm]
    use_df = by_scenario if not by_scenario.empty else subset

    if str(catalog_type or "").strip().lower() == "transformation":
        var_norm = str(variable_name or "").strip().lower()
        wanted_groups: set[str] | None = None
        if var_norm == "inputs":
            wanted_groups = {"feedstock fuels", "auxiliary fuels"}
        elif var_norm == "outputs by output fuel":
            wanted_groups = {"output fuels"}
        elif var_norm == "outputs by feedstock fuel":
            wanted_groups = {"feedstock fuels"}
        if wanted_groups:
            group_norm = use_df["fuel_group"].str.lower()
            grouped = use_df[group_norm.isin(wanted_groups)]
            if not grouped.empty:
                use_df = grouped
    return {
        str(item)
        for item in use_df["fuel_name_norm"].tolist()
        if str(item).strip()
    }


def _catalog_scope_from_meta(meta: dict[str, object]) -> tuple[str | None, str | None]:
    """Map template metadata branch to catalog scope."""
    branch_path = str(meta.get("branch") or "").strip()
    parts = [part.strip() for part in branch_path.split("\\") if part and part.strip()]
    if len(parts) < 2:
        return None, None
    if parts[0].lower() == "transformation":
        return "transformation", parts[1]
    if parts[0].lower() == "resources":
        root = parts[1]
        if root.lower() in {"primary", "secondary"}:
            return "supply", root
    return None, None


def _sheet_tmp_export_path(tpl: Path, sheet_name: str, sheet_index: int) -> Path:
    """Build a short deterministic temp CSV filename to avoid LEAP path I/O issues."""
    key = f"{tpl.name}|{sheet_name}|{sheet_index}"
    token = hashlib.md5(key.encode("utf-8")).hexdigest()[:10]
    return TEMPLATE_EXPORT_TMP_DIR / f"sheet_{sheet_index:03d}_{token}.csv"


def archive_file(path: Path, archive_dir: Path) -> Optional[Path]:
    """Copy an existing file to an archive directory with a date-stamped suffix."""
    if not path.exists():
        return None
    now_utc = datetime.utcnow()
    stamp = now_utc.strftime("%Y%m%d_%H%M%S")
    day_stamp = now_utc.strftime("%Y%m%d")
    daily_dir = archive_dir / "daily_archives" / day_stamp
    daily_dir.mkdir(parents=True, exist_ok=True)
    archived = daily_dir / f"{path.stem}_{stamp}{path.suffix}"
    archived.write_bytes(path.read_bytes())

    if ARCHIVE_RETENTION_DAYS is not None:
        cutoff = now_utc.date() - timedelta(days=int(ARCHIVE_RETENTION_DAYS))
        base_daily = archive_dir / "daily_archives"
        for daily_path in base_daily.glob("*"):
            if not daily_path.is_dir():
                continue
            try:
                day = datetime.strptime(daily_path.name, "%Y%m%d").date()
            except ValueError:
                continue
            if day >= cutoff:
                continue
            for old_file in daily_path.glob("*"):
                try:
                    old_file.unlink()
                except OSError:
                    pass
            try:
                daily_path.rmdir()
            except OSError:
                pass
    return archived


def run_results_export() -> dict:
    """Perform UI exports for each table spec and return a log dict."""
    ensure_repo_root()
    layout = build_workflow_output_layout(DEFAULT_OUTPUT_DIR)
    coverage_result = run_mapping_coverage_check_for_workflow()
    app = connect_leap()

    results = []
    dfs_for_combined = []
    for name, spec in TABLE_SPECS.items():
        csv_path = Path(spec.get("csv_path", DEFAULT_OUTPUT_DIR / f"{name}.csv"))
        ensure_parent_dirs([csv_path, COMBINED_XLSX_PATH])

        select_area(app, spec.get("area"))
        ensure_calculated(app, force=FORCE_RECALC)

        fave_status = activate_favorite(app, spec.get("favorite"))

        set_axes(app, x_axis=spec.get("x_axis"), legend=spec.get("legend"))
        set_context(
            app,
            scenario=spec.get("scenario"),
            region=spec.get("region"),
            year=spec.get("year"),
            unit=spec.get("unit"),
            branch_path=spec.get("branch"),
        )
        requested_variable = spec.get("variable")
        requested_branch = spec.get("branch")
        if requested_variable:
            if requested_branch:
                variable_obj = _resolve_branch_variable(
                    app,
                    str(requested_branch),
                    requested_variable,
                    allow_substitution=True,
                )
                try:
                    app.ActiveVariable = str(variable_obj.Name)
                except Exception:
                    pass
            else:
                print(
                    f"Warning: table '{name}' requested variable '{requested_variable}' without a branch; "
                    "skipping ActiveVariable assignment.",
                    flush=True,
                )

        dims = list_dimensions(app)
        csv_path = export_results_csv(app, csv_path)
        entry = {
            "table": name,
            "csv": str(csv_path),
            "favorite_status": fave_status,
            "dimensions": dims,
            "sheet": spec.get("sheet", name),
        }

        if WRITE_EXCEL:
            df = pd.read_csv(csv_path)
            dfs_for_combined.append((spec.get("sheet", name), df))

        results.append(entry)

    combined_path = None
    if WRITE_EXCEL and dfs_for_combined:
        COMBINED_XLSX_PATH.parent.mkdir(parents=True, exist_ok=True)
        with pd.ExcelWriter(COMBINED_XLSX_PATH, engine="openpyxl") as writer:
            for sheet_name, df in dfs_for_combined:
                df.to_excel(writer, sheet_name=sheet_name, index=False)
        combined_path = str(COMBINED_XLSX_PATH)

    primary_outputs = {entry["table"]: entry["csv"] for entry in results if entry.get("csv")}
    if combined_path:
        primary_outputs["combined_xlsx"] = combined_path
    supporting_outputs = {"mapping_coverage_dir": str(MAPPING_COVERAGE_DIR)}
    manifest_path = write_output_manifest(
        out_dir=layout.root,
        primary_outputs=primary_outputs,
        supporting_outputs=supporting_outputs,
        primary_output_descriptions={
            "combined_xlsx": "Combined workbook containing the exported result tables.",
        },
        supporting_output_descriptions={
            "mapping_coverage_dir": "Coverage check outputs for the ninth-to-ESTO mapping used by this workflow.",
        },
        notes=[
            "Direct exported tables stay at the workflow root.",
            "Diagnostics and temp artifacts are grouped under supporting_files/.",
        ],
    )
    return {
        "tables": results,
        "combined_xlsx": combined_path,
        "mapping_coverage": coverage_result,
        "output_manifest_json": str(manifest_path),
    }


def run_template_fill() -> dict:
    """Read one or more LEAP Results template workbooks and refill each sheet from LEAP."""
    if not TEMPLATE_PATHS:
        raise ValueError("TEMPLATE_PATHS must be set when USE_TEMPLATE is True.")

    paths = TEMPLATE_PATHS
    layout = build_workflow_output_layout(DEFAULT_OUTPUT_DIR)

    print("Template refill: starting", flush=True)
    ensure_repo_root()
    coverage_result = run_mapping_coverage_check_for_workflow()
    app = connect_leap()
    print("Connected to LEAP", flush=True)
    default_area = next(iter(TABLE_SPECS.values()), {}).get("area") if TABLE_SPECS else None
    select_area(app, default_area)
    ensure_calculated(app, force=FORCE_RECALC)
    print("LEAP calculation check done", flush=True)

    outputs = []
    year_axis_audit_rows: list[dict[str, object]] = []
    catalog_coverage_rows: list[dict[str, object]] = []
    expected_by_workbook: dict[str, list[str]] = {}
    if ENFORCE_EXPECTED_SHEETS_CONTRACT:
        expected_by_workbook = _load_expected_sheets_contract(EXPECTED_SHEETS_CONFIG_PATH.resolve())
        missing_templates_in_contract = [
            tpl.name for tpl in paths if tpl.name not in expected_by_workbook
        ]
        if missing_templates_in_contract:
            raise RuntimeError(
                "Expected-sheets contract missing template(s): "
                + ", ".join(missing_templates_in_contract)
            )
    fuel_catalog_preflight.ensure_fuel_catalog_current(
        catalog_path=FUEL_CATALOG_PATH.resolve(),
        max_age_days=7,
        leap_app=app,
        context="leap_results_workflow.run_template_fill",
        auto_refresh=True,
        fail_on_refresh_error=True,
    )
    catalog_df = _load_fuel_catalog(FUEL_CATALOG_PATH.resolve())
    if catalog_df.empty:
        print(
            f"Warning: fuel catalog not found or empty at {FUEL_CATALOG_PATH}. "
            "Template fuel-coverage checks will be skipped.",
            flush=True,
        )
    TEMPLATE_EXPORT_TMP_DIR.mkdir(parents=True, exist_ok=True)

    for tpl in paths:
        print(f"Starting template {tpl}", flush=True)
        archive_path = None
        if ARCHIVE_TEMPLATE:
            archive_path = archive_file(tpl, ARCHIVE_DIR)

        # Overwrite template in-place so downstream workflows that read data/leap results tables
        # always see the latest exported values.
        output_path = tpl
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb = load_workbook(output_path)
        expected_sheet_order: list[str] = list(wb.sheetnames)
        if ENFORCE_EXPECTED_SHEETS_CONTRACT:
            _expected_contract_order, missing_sheets, unexpected_sheets = _validate_expected_sheets_for_workbook(
                workbook_name=tpl.name,
                actual_sheet_names=list(wb.sheetnames),
                expected_by_workbook=expected_by_workbook,
            )
            if missing_sheets or unexpected_sheets:
                lines: list[str] = [
                    f"Expected-sheets contract mismatch for {tpl.name}:",
                ]
                if missing_sheets:
                    lines.append("missing: " + ", ".join(missing_sheets))
                if unexpected_sheets:
                    lines.append("unexpected: " + ", ".join(unexpected_sheets))
                raise RuntimeError("\n".join(lines))
            # Iterate in workbook order (as stored in the workbook), while still
            # enforcing that the contract set matches exactly.
            expected_sheet_order = list(wb.sheetnames)
        naming_issues = validate_transformation_sheet_naming(
            wb.sheetnames,
            context=str(tpl.name),
        )
        if naming_issues:
            raise RuntimeError(
                "Transformation sheet naming validation failed:\n - "
                + "\n - ".join(naming_issues)
            )
        derived_sheet_children = DERIVED_SHEET_CHILDREN.get(tpl.name, {}) if USE_DERIVED_PARENT_SHEETS else {}
        derived_targets = set(derived_sheet_children.keys())
        meta_by_sheet: dict[str, dict[str, object]] = {}
        table_by_sheet: dict[str, pd.DataFrame] = {}
        sheet_logs = []
        available_favorites_lower = {name.lower() for name in _list_favorite_names(app)}
        for sheet_index, sheet_name in enumerate(expected_sheet_order, start=1):
            print(f"Processing {tpl} / {sheet_name}", flush=True)
            ws = wb[sheet_name]
            existing_sheet_df = _snapshot_worksheet_values(ws)
            meta = parse_template_worksheet(ws)
            meta_by_sheet[sheet_name] = meta
            favorite_status = None
            favorite_activated_for_sheet = False
            expected_input_type_qualifier = _expected_input_type_qualifier_for_sheet(sheet_name)
            if (
                USE_FAVORITES_FOR_TRANSFORMATION_INPUT_STYLE_SHEETS
                and _is_transformation_input_style_sheet(tpl, sheet_name)
            ):
                if STRICT_REQUIRED_FAVORITES_CONTRACT and sheet_name.lower() not in available_favorites_lower:
                    raise RuntimeError(
                        f"Required favorite '{sheet_name}' missing for {tpl.name}/{sheet_name}. "
                        "Update LEAP favorites to match sheet names."
                    )
                resolved_name, resolve_mode, favorite_status = activate_favorite_for_sheet(app, sheet_name)
                if resolved_name and "not found/activated" not in str(favorite_status).lower():
                    favorite_activated_for_sheet = True
                    print(
                        f"[FAVORITE] Activated '{resolved_name}' for {tpl.name}/{sheet_name} "
                        f"(mode={resolve_mode}): {favorite_status}",
                        flush=True,
                    )
                else:
                    message = (
                        f"Failed to activate required transformation input-style favorite "
                        f"for sheet '{sheet_name}' (mode={resolve_mode})."
                    )
                    if favorite_status:
                        message += f" Status: {favorite_status}"
                    if STRICT_FAVORITE_FOR_TRANSFORMATION_INPUT_STYLE:
                        raise RuntimeError(message)
                    print(f"[WARN] {message}", flush=True)
            meta, year_axis_audit = normalize_template_year_axis(
                meta,
                template_name=tpl.name,
                sheet_name=sheet_name,
            )
            year_axis_audit_rows.append(year_axis_audit)
            if year_axis_audit["normalized"]:
                print(
                    f"Warning: normalized template year axis for {tpl.name}/{sheet_name} "
                    f"from {year_axis_audit['axis_found']} to {year_axis_audit['axis_expected']}.",
                    flush=True,
                )
            meta_by_sheet[sheet_name] = meta
            if sheet_name in derived_targets:
                continue
            scale_spec = UNIT_SCALES.get(sheet_name) or UNIT_SCALES_BY_VARIABLE.get(meta.get("variable"))
            is_supply_sheet = _is_supply_energy_table(meta)
            requires_rendered_csv = _is_transformation_energy_table(meta)
            use_rendered_csv = requires_rendered_csv or (
                TEMPLATE_EXTRACT_METHOD.lower() == "csv" and not is_supply_sheet
            )
            if use_rendered_csv:
                export_csv_path = _sheet_tmp_export_path(tpl, sheet_name, sheet_index)
                try:
                    fresh_df = build_fresh_table_from_export(
                        app,
                        meta,
                        export_csv_path=export_csv_path,
                        scale_spec=scale_spec,
                        context=f"{tpl.name}/{sheet_name}",
                        allow_variable_substitution=not STRICT_TEMPLATE_VARIABLE_MATCH,
                        trust_active_results_view=(
                            favorite_activated_for_sheet
                            and _is_transformation_input_style_sheet(tpl, sheet_name)
                        ),
                        expected_input_type_qualifier=expected_input_type_qualifier,
                    )
                except MissingRequestedVariableError as exc:
                    print(
                        f"Warning: {exc} Writing no-results table for {tpl.name}/{sheet_name}.",
                        flush=True,
                    )
                    fresh_df = _build_no_results_table(meta)
                except Exception as exc:  # noqa: BLE001
                    if requires_rendered_csv:
                        policy = str(RENDERED_CSV_TRANSFORMATION_ERROR_POLICY or "raise").strip().lower()
                        if policy == "empty_table":
                            print(
                                f"[WARN][RENDERED_CSV_TRANSFORMATION_ERROR][EMPTY_TABLE] "
                                f"{tpl.name}/{sheet_name}: {exc}. "
                                "Writing no-results table and continuing.",
                                flush=True,
                            )
                            fresh_df = _build_no_results_table(meta)
                        else:
                            raise RuntimeError(
                                f"Rendered LEAP CSV export failed for {tpl.name}/{sheet_name}; "
                                "refusing to fall back to template-driven ValueRS extraction for this sheet."
                            ) from exc
                        # Skip ValueRS fallback for transformation sheets by design.
                        # Continue to validation/write with the empty table.
                        pass
                    else:
                        print(
                            f"Warning: CSV export failed for {tpl.name}/{sheet_name} ({exc}); "
                            "falling back to ValueRS extraction.",
                            flush=True,
                        )
                        try:
                            fresh_df = build_fresh_table(
                                app,
                                meta,
                                scale_spec=scale_spec,
                                allow_variable_substitution=not STRICT_TEMPLATE_VARIABLE_MATCH,
                            )
                        except MissingRequestedVariableError as missing_exc:
                            print(
                                f"Warning: {missing_exc} Writing no-results table for {tpl.name}/{sheet_name}.",
                                flush=True,
                            )
                            fresh_df = _build_no_results_table(meta)
            else:
                try:
                    fresh_df = build_fresh_table(
                        app,
                        meta,
                        scale_spec=scale_spec,
                        allow_variable_substitution=not STRICT_TEMPLATE_VARIABLE_MATCH,
                    )
                except MissingRequestedVariableError as exc:
                    print(
                        f"Warning: {exc} Writing no-results table for {tpl.name}/{sheet_name}.",
                        flush=True,
                    )
                    fresh_df = _build_no_results_table(meta)
                if (
                    FALLBACK_SUSPICIOUS_VALUES_TO_CSV
                    and not is_supply_sheet
                    and is_suspicious_values_table(fresh_df)
                ):
                    export_csv_path = _sheet_tmp_export_path(tpl, sheet_name, sheet_index)
                    try:
                        csv_df = build_fresh_table_from_export(
                            app,
                            meta,
                            export_csv_path=export_csv_path,
                            scale_spec=scale_spec,
                            context=f"{tpl.name}/{sheet_name}",
                            allow_variable_substitution=not STRICT_TEMPLATE_VARIABLE_MATCH,
                            trust_active_results_view=(
                                favorite_activated_for_sheet
                                and _is_transformation_input_style_sheet(tpl, sheet_name)
                            ),
                            expected_input_type_qualifier=expected_input_type_qualifier,
                        )
                        if not csv_df.empty:
                            print(
                                f"Info: replaced suspicious ValueRS table with CSV export for "
                                f"{tpl.name}/{sheet_name}.",
                                flush=True,
                            )
                            fresh_df = csv_df
                    except Exception as exc:  # noqa: BLE001
                        print(
                            f"Warning: suspicious ValueRS table for {tpl.name}/{sheet_name}; "
                            f"CSV recovery failed ({exc}).",
                            flush=True,
                        )
            context = f"{tpl.name}/{sheet_name}"
            try:
                validate_refilled_table(
                    fresh_df,
                    meta=meta,
                    context=context,
                )
            except RuntimeError as exc:
                message = str(exc or "")
                if "collapsed to a Total-only layout" not in message:
                    raise
                recovered = False
                if not use_rendered_csv:
                    export_csv_path = _sheet_tmp_export_path(tpl, sheet_name, sheet_index)
                    try:
                        recovered_df = build_fresh_table_from_export(
                            app,
                            meta,
                            export_csv_path=export_csv_path,
                            scale_spec=scale_spec,
                            context=context,
                            allow_variable_substitution=not STRICT_TEMPLATE_VARIABLE_MATCH,
                            trust_active_results_view=(
                                favorite_activated_for_sheet
                                and _is_transformation_input_style_sheet(tpl, sheet_name)
                            ),
                            expected_input_type_qualifier=expected_input_type_qualifier,
                        )
                        validate_refilled_table(
                            recovered_df,
                            meta=meta,
                            context=context,
                        )
                        fresh_df = recovered_df
                        recovered = True
                        print(
                            f"Warning: recovered Total-only refill via CSV export for {context}.",
                            flush=True,
                        )
                    except Exception as recovery_exc:  # noqa: BLE001
                        print(
                            f"Warning: Total-only refill persisted after CSV recovery for {context} "
                            f"({recovery_exc}). Preserving existing sheet values.",
                            flush=True,
                        )
                if not recovered:
                    print(
                        f"Warning: preserving existing sheet values for {context} due to Total-only refill.",
                        flush=True,
                    )
                    fresh_df = existing_sheet_df
            _record_catalog_coverage_row(
                catalog_rows=catalog_coverage_rows,
                catalog_df=catalog_df,
                template_name=tpl.name,
                sheet_name=sheet_name,
                meta=meta,
                table_df=fresh_df,
            )
            write_table_values_preserve_format(ws, fresh_df)
            table_by_sheet[sheet_name] = fresh_df
            # Summarize X items (years) as a range for compact logging
            x_items = [v for v in meta["x_items"] if not pd.isna(v)]
            year_values = [int(v) for v in x_items if isinstance(v, (int, float))]
            x_summary = None
            if year_values:
                x_summary = f"{min(year_values)}-{max(year_values)}"
            if x_items and str(x_items[-1]).lower() == "total":
                x_summary = f"{x_summary} + Total" if x_summary else "Total"
            sheet_logs.append(
                {
                    "sheet": sheet_name,
                    "legend": meta["legend_label"],
                    "x_range": x_summary,
                    "scale_used": scale_spec,
                    "favorite_status": favorite_status,
                }
            )

        # Build derived parent sheets from child sheets after base sheets are extracted.
        if derived_targets:
            for sheet_name in expected_sheet_order:
                if sheet_name not in derived_targets:
                    continue
                ws = wb[sheet_name]
                target_meta = meta_by_sheet[sheet_name]
                child_names = derived_sheet_children.get(sheet_name, [])
                missing_children = [name for name in child_names if name not in table_by_sheet]
                if missing_children:
                    raise RuntimeError(
                        f"Derived sheet '{sheet_name}' in {tpl.name} missing child table(s): "
                        + ", ".join(missing_children)
                    )
                child_tables = [table_by_sheet[name] for name in child_names]
                derived_df = _build_derived_table_from_children(
                    target_meta=target_meta,
                    child_tables=child_tables,
                )
                _record_catalog_coverage_row(
                    catalog_rows=catalog_coverage_rows,
                    catalog_df=catalog_df,
                    template_name=tpl.name,
                    sheet_name=sheet_name,
                    meta=target_meta,
                    table_df=derived_df,
                )
                write_table_values_preserve_format(ws, derived_df)
                table_by_sheet[sheet_name] = derived_df
                x_items = [v for v in target_meta["x_items"] if not pd.isna(v)]
                year_values = [int(v) for v in x_items if isinstance(v, (int, float))]
                x_summary = None
                if year_values:
                    x_summary = f"{min(year_values)}-{max(year_values)}"
                if x_items and str(x_items[-1]).lower() == "total":
                    x_summary = f"{x_summary} + Total" if x_summary else "Total"
                sheet_logs.append(
                    {
                        "sheet": sheet_name,
                        "legend": target_meta["legend_label"],
                        "x_range": x_summary,
                        "scale_used": UNIT_SCALES.get(sheet_name) or UNIT_SCALES_BY_VARIABLE.get(target_meta.get("variable")),
                        "favorite_status": "derived_from_children",
                        "derived_children": "; ".join(child_names),
                    }
                )
        wb.save(output_path)

        outputs.append(
            {
                "template": str(tpl),
                "archived_copy": str(archive_path) if archive_path else None,
                "output": str(output_path),
                "sheets": sheet_logs,
            }
        )

    year_axis_audit_path = TEMPLATE_YEAR_AXIS_AUDIT_CSV.resolve()
    year_axis_audit_path.parent.mkdir(parents=True, exist_ok=True)
    pd.DataFrame(year_axis_audit_rows).to_csv(year_axis_audit_path, index=False)
    catalog_coverage_path = FUEL_CATALOG_COVERAGE_CSV.resolve()
    catalog_coverage_path.parent.mkdir(parents=True, exist_ok=True)
    catalog_coverage_df = pd.DataFrame(catalog_coverage_rows)
    catalog_coverage_df.to_csv(catalog_coverage_path, index=False)
    missing_total = int(
        pd.to_numeric(
            catalog_coverage_df.get("missing_fuels_count", pd.Series(dtype=float)),
            errors="coerce",
        ).fillna(0).sum()
    ) if not catalog_coverage_df.empty else 0
    if missing_total > 0:
        print(
            "[WARN] Template fuel catalog coverage detected "
            f"{missing_total} missing expected fuel row(s)."
        )
    if FUEL_CATALOG_COVERAGE_FAIL_ON_MISSING and missing_total > 0:
        raise RuntimeError(
            "Template refill failed due to missing expected catalog fuels. "
            f"See {catalog_coverage_path}."
        )

    manifest_path = write_output_manifest(
        out_dir=layout.root,
        primary_outputs={},
        supporting_outputs={
            "template_year_axis_audit": str(year_axis_audit_path),
            "template_fuel_catalog_coverage": str(catalog_coverage_path),
            "mapping_coverage_dir": str(MAPPING_COVERAGE_DIR),
            "template_archives_dir": str(ARCHIVE_DIR),
            "tmp_template_exports_dir": str(TEMPLATE_EXPORT_TMP_DIR),
        },
        supporting_output_descriptions={
            "template_year_axis_audit": "Audit of normalized year axes found in template sheets.",
            "template_fuel_catalog_coverage": "Coverage check comparing template rows to the prepared fuel catalog.",
            "mapping_coverage_dir": "Coverage check outputs for the ninth-to-ESTO mapping used by this workflow.",
            "template_archives_dir": "Archived copies of templates before refill.",
            "tmp_template_exports_dir": "Temporary CSV exports used during template refill and recovery.",
        },
        notes=[
            "Refilled template workbooks are written in place under data/leap results tables.",
            "Workflow-generated diagnostics and temporary files live under outputs/leap_results/supporting_files/.",
        ],
    )
    return {
        "templates_processed": outputs,
        "mapping_coverage": coverage_result,
        "template_year_axis_audit": str(year_axis_audit_path),
        "template_fuel_catalog_coverage": str(catalog_coverage_path),
        "output_manifest_json": str(manifest_path),
    }


#%% Bottom run block (edit toggles above, then run this cell)
if __name__ == "__main__":
    try:
        if USE_TEMPLATE:
            run_log = run_template_fill()
            print("Template refill complete:")
        else:
            run_log = run_results_export()
            print("LEAP Results export complete:")
        for key, val in run_log.items():
            print(f"  {key}: {val}")
    except Exception as exc:  # noqa: BLE001
        print(f"Export failed: {exc}")

#%%


try:
    from codebase.utilities.workflow_common import emit_completion_beep as _emit_completion_beep
except Exception:  # pragma: no cover
    def _emit_completion_beep(*, success: bool = True) -> None:  # noqa: ARG001
        return


if __name__ == "__main__":  # pragma: no cover
    _emit_completion_beep(success=True, style="chime")
