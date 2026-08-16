"""Plain-language validation of the inputs a supported command needs.

Every command validates its whole input set before it does any work, so a
colleague finds out that a file is missing, is the wrong kind of file, or does
not cover the economy they asked for *before* a long render starts — and reads
about it in a sentence, not a stack trace.

The report this module produces is written into the run manifest and into the
support bundle, so a failed run can be diagnosed without re-running it.
"""

from __future__ import annotations

import csv
import json
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Iterable, Sequence


#: Diagnostic artifacts the balance-review command reads. The two required files
#: are produced together by the balance-diagnostics step
#: (``run_baseline_seed_balance_diagnostics``, which reads a LEAP balance export
#: against the ESTO and 9th-edition source tables). The mapping-issue file is
#: optional because a clean run may not produce one.
BALANCE_REVIEW_REQUIRED_ARTIFACTS = (
    "leap_balance_source_review.csv",
    "leap_balance_source_differences.csv",
)
BALANCE_REVIEW_OPTIONAL_ARTIFACTS = ("leap_balance_mapping_issues.csv",)

BALANCE_REVIEW_SHARED_COLUMNS = (
    "economy",
    "scenario",
    "year",
    "esto_flow",
    "esto_product",
    "leap_sector_names",
    "leap_fuel_names",
    "status",
    "difference_pj",
)
BALANCE_REVIEW_REVIEW_ONLY_COLUMNS = (
    "leap_value_pj",
    "source_value_pj",
    "leap_balance_row",
    "leap_balance_fuel",
    "no_direct_projection_comparator",
)

DASHBOARD_COMPARISON_COLUMNS = (
    "comparison_scope",
    "source_system",
    "economy",
    "scenario",
    "year",
    "common_flow_label",
    "common_product_label",
    "common_row_id",
    "value",
)
DASHBOARD_ROWS_COLUMNS = ("common_row_id",)

_SCENARIO_ALIASES = {
    "ref": "Reference",
    "reference": "Reference",
    "tgt": "Target",
    "target": "Target",
}


class InputValidationError(ValueError):
    """Raised when a command's inputs cannot support a run."""

    def __init__(self, report: "ValidationReport") -> None:
        self.report = report
        super().__init__(report.failure_message())


@dataclass
class ValidationCheck:
    name: str
    ok: bool
    detail: str

    def as_dict(self) -> dict[str, Any]:
        return {"name": self.name, "ok": self.ok, "detail": self.detail}


@dataclass
class ValidationReport:
    """Result of validating one command's inputs."""

    command: str
    checks: list[ValidationCheck] = field(default_factory=list)
    facts: dict[str, Any] = field(default_factory=dict)

    def add(self, name: str, ok: bool, detail: str) -> bool:
        self.checks.append(ValidationCheck(name=name, ok=ok, detail=detail))
        return ok

    @property
    def ok(self) -> bool:
        return all(check.ok for check in self.checks)

    @property
    def failures(self) -> list[ValidationCheck]:
        return [check for check in self.checks if not check.ok]

    def failure_message(self) -> str:
        problems = self.failures
        if not problems:
            return f"{self.command}: inputs are valid."
        header = (
            f"{self.command} cannot run because "
            f"{len(problems)} input {'problem was' if len(problems) == 1 else 'problems were'} found:"
        )
        body = "\n".join(f"  - {check.detail}" for check in problems)
        return f"{header}\n{body}"

    def as_dict(self) -> dict[str, Any]:
        return {
            "command": self.command,
            "ok": self.ok,
            "checks": [check.as_dict() for check in self.checks],
            "facts": self.facts,
        }

    def as_text(self) -> str:
        lines = [f"Input validation report: {self.command}", "=" * 72]
        for check in self.checks:
            lines.append(f"[{'PASS' if check.ok else 'FAIL'}] {check.name}")
            lines.append(f"       {check.detail}")
        lines.append("")
        lines.append(f"Overall: {'PASS' if self.ok else 'FAIL'}")
        if self.facts:
            lines.append("")
            lines.append("Facts observed")
            lines.append("-" * 72)
            lines.append(json.dumps(self.facts, indent=2, default=str))
        return "\n".join(lines)

    def raise_if_failed(self) -> "ValidationReport":
        if not self.ok:
            raise InputValidationError(self)
        return self


def normalize_scenario(scenario: object) -> str:
    """Return the full scenario name LEAP writes into a balance sheet."""
    text = str(scenario or "").strip()
    if not text:
        raise ValueError("A scenario is required (for example 'Target' or 'TGT').")
    return _SCENARIO_ALIASES.get(text.lower(), text)


def normalize_economy(economy: object) -> str:
    """Return the underscore-normalized economy code (``20_USA``)."""
    text = str(economy or "").strip().upper().replace("-", "_")
    if not text:
        raise ValueError("An economy code is required (for example '20_USA').")
    if "_" not in text and len(text) > 2 and text[:2].isdigit():
        text = f"{text[:2]}_{text[2:]}"
    return text


def _read_header(path: Path) -> list[str]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        for row in csv.reader(handle):
            return [str(cell).strip() for cell in row]
    return []


def _check_readable_file(
    report: ValidationReport,
    path: Path | None,
    *,
    name: str,
    description: str,
    suffixes: Sequence[str] = (),
) -> bool:
    if path is None:
        return report.add(name, False, f"No path was given for the {description}.")
    if not path.exists():
        return report.add(
            name,
            False,
            f"The {description} does not exist at:\n      {path}\n"
            f"    Check the path, or copy the file into the release's input folder.",
        )
    if not path.is_file():
        return report.add(
            name,
            False,
            f"The {description} path is a folder, not a file:\n      {path}",
        )
    if suffixes and path.suffix.lower() not in {s.lower() for s in suffixes}:
        return report.add(
            name,
            False,
            f"The {description} must be a {'/'.join(suffixes)} file, but "
            f"{path.name!r} is a {path.suffix or 'no-extension'} file.",
        )
    if path.stat().st_size == 0:
        return report.add(name, False, f"The {description} is empty:\n      {path}")
    return report.add(name, True, f"Found the {description}: {path.name}")


def _check_columns(
    report: ValidationReport,
    path: Path,
    required: Iterable[str],
    *,
    name: str,
    description: str,
) -> bool:
    header = _read_header(path)
    missing = [column for column in required if column not in header]
    if missing:
        return report.add(
            name,
            False,
            f"The {description} ({path.name}) is missing required "
            f"{'column' if len(missing) == 1 else 'columns'}: {', '.join(missing)}. "
            "This usually means the file came from a different tool or an older "
            "version of the pipeline.",
        )
    return report.add(
        name,
        True,
        f"The {description} has all {len(list(required))} required columns.",
    )


# ---------------------------------------------------------------------------
# Balance review
# ---------------------------------------------------------------------------


def validate_balance_review_inputs(
    *,
    economy: str,
    scenario: str,
    year: int,
    balance_export_workbook: Path | str | None,
    diagnostics_directory: Path | str | None,
) -> ValidationReport:
    """Validate the inputs for one balance-review workbook.

    This is the *existing diagnostic artifacts* input mode: the caller already
    has a diagnostics directory produced by the balance-diagnostics step, plus
    the LEAP balance export those diagnostics were computed against.
    """
    report = ValidationReport(command="balance-review")

    try:
        economy_code = normalize_economy(economy)
        report.facts["economy"] = economy_code
    except ValueError as exc:
        report.add("economy", False, str(exc))
        economy_code = ""
    try:
        scenario_name = normalize_scenario(scenario)
        report.facts["scenario"] = scenario_name
    except ValueError as exc:
        report.add("scenario", False, str(exc))
        scenario_name = ""

    try:
        year_value = int(year)
    except (TypeError, ValueError):
        report.add(
            "year",
            False,
            f"The year must be a four-digit number, but {year!r} was given.",
        )
        year_value = 0
    else:
        if not 1990 <= year_value <= 2100:
            report.add(
                "year",
                False,
                f"The year {year_value} is outside the supported range 1990-2100.",
            )
        else:
            report.add("year", True, f"Review year {year_value}.")
        report.facts["year"] = year_value

    workbook_path = Path(str(balance_export_workbook)) if balance_export_workbook else None
    workbook_ok = _check_readable_file(
        report,
        workbook_path,
        name="balance_export_workbook",
        description="LEAP balance-export workbook",
        suffixes=(".xlsx",),
    )

    diagnostics_dir = Path(str(diagnostics_directory)) if diagnostics_directory else None
    if diagnostics_dir is None or not diagnostics_dir.is_dir():
        report.add(
            "diagnostics_directory",
            False,
            "The diagnostics folder does not exist at:\n"
            f"      {diagnostics_dir}\n"
            "    This command builds the workbook from an existing diagnostics "
            "folder. Generating one from a LEAP balance export is a separate "
            "step that reads the ESTO and 9th-edition source tables, and is not "
            "part of this release.",
        )
    else:
        report.add(
            "diagnostics_directory",
            True,
            f"Found the diagnostics folder: {diagnostics_dir}",
        )
        for filename in BALANCE_REVIEW_REQUIRED_ARTIFACTS:
            artifact = diagnostics_dir / filename
            if not _check_readable_file(
                report,
                artifact,
                name=f"artifact:{filename}",
                description=f"diagnostic artifact {filename}",
                suffixes=(".csv",),
            ):
                continue
            required = list(BALANCE_REVIEW_SHARED_COLUMNS)
            if filename == "leap_balance_source_review.csv":
                required += list(BALANCE_REVIEW_REVIEW_ONLY_COLUMNS)
            _check_columns(
                report,
                artifact,
                required,
                name=f"columns:{filename}",
                description=f"diagnostic artifact {filename}",
            )
        for filename in BALANCE_REVIEW_OPTIONAL_ARTIFACTS:
            artifact = diagnostics_dir / filename
            report.add(
                f"optional:{filename}",
                True,
                f"Optional artifact {filename} is "
                f"{'present and will be included' if artifact.is_file() else 'absent; mapping issues remain available in the diagnostic CSVs'}.",
            )

        review_path = diagnostics_dir / "leap_balance_source_review.csv"
        if review_path.is_file() and economy_code and scenario_name and year_value:
            matched, available = _scan_diagnostic_scope(
                review_path,
                economy=economy_code,
                scenario=scenario_name,
                year=year_value,
            )
            report.facts["diagnostic_rows_matching_request"] = matched
            report.facts["diagnostic_scope_available"] = available
            if matched:
                report.add(
                    "diagnostic_scope",
                    True,
                    f"The diagnostics cover {economy_code} / {scenario_name} / "
                    f"{year_value} ({matched:,} comparison rows).",
                )
            else:
                offered = ", ".join(available) if available else "nothing"
                report.add(
                    "diagnostic_scope",
                    False,
                    f"The diagnostics contain no rows for {economy_code} / "
                    f"{scenario_name} / {year_value}. They cover: {offered}. "
                    "Either request one of those, or use a diagnostics folder "
                    "produced for the economy, scenario, and year you want.",
                )

    if workbook_ok and workbook_path is not None and scenario_name and year_value:
        _check_balance_export_sheet(
            report,
            workbook_path,
            scenario=scenario_name,
            year=year_value,
        )

    return report


def _scan_diagnostic_scope(
    review_path: Path,
    *,
    economy: str,
    scenario: str,
    year: int,
) -> tuple[int, list[str]]:
    """Count rows matching the request and list the scopes actually present."""
    matched = 0
    available: set[str] = set()
    with review_path.open("r", encoding="utf-8-sig", newline="") as handle:
        for row in csv.DictReader(handle):
            row_economy = str(row.get("economy") or "").strip()
            row_scenario = str(row.get("scenario") or "").strip()
            row_year = str(row.get("year") or "").strip()
            available.add(f"{row_economy}/{row_scenario}/{row_year.split('.')[0]}")
            if (
                row_economy == economy
                and row_scenario.lower() == scenario.lower()
                and row_year.split(".")[0] == str(year)
            ):
                matched += 1
    return matched, sorted(available)


def _check_balance_export_sheet(
    report: ValidationReport,
    workbook_path: Path,
    *,
    scenario: str,
    year: int,
) -> None:
    """Confirm the export workbook has exactly one sheet for the scenario/year."""
    try:
        from openpyxl import load_workbook
    except ImportError:  # pragma: no cover - openpyxl is a declared dependency
        report.add(
            "balance_export_sheet",
            False,
            "openpyxl is not available, so the balance export cannot be inspected.",
        )
        return

    import re

    try:
        workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    except Exception as exc:
        report.add(
            "balance_export_sheet",
            False,
            f"The balance-export workbook could not be opened: {exc}. "
            "Re-export it from LEAP, or check that the file is not still open in Excel.",
        )
        return

    found: list[str] = []
    described: set[str] = set()
    try:
        for sheet in workbook.worksheets:
            metadata = str(sheet.cell(2, 1).value or "").strip()
            scenario_match = re.search(r"Scenario:\s*([^,]+)", metadata, flags=re.I)
            year_match = re.search(r"Year:\s*(\d{4})", metadata, flags=re.I)
            units_match = re.search(r"Units:\s*(.+)$", metadata, flags=re.I)
            if not (scenario_match and year_match and units_match):
                continue
            sheet_scenario = scenario_match.group(1).strip()
            sheet_year = int(year_match.group(1))
            described.add(f"{sheet_scenario}/{sheet_year}")
            if sheet_scenario.lower() == scenario.lower() and sheet_year == year:
                found.append(str(sheet.title))
                report.facts["balance_export_units"] = units_match.group(1).strip().rstrip(".")
    finally:
        workbook.close()

    if not described:
        report.add(
            "balance_export_sheet",
            False,
            "No sheet in the balance-export workbook declares Scenario, Year, and "
            "Units in cell A2. This does not look like a LEAP Energy Balance "
            "export; re-export using the Energy Balance results view.",
        )
        return
    if len(found) > 1:
        report.add(
            "balance_export_sheet",
            False,
            f"The workbook has {len(found)} sheets for {scenario} {year} "
            f"({', '.join(found)}). Exactly one is required.",
        )
        return
    if not found:
        report.add(
            "balance_export_sheet",
            False,
            f"The workbook has no sheet for {scenario} {year}. It contains: "
            f"{', '.join(sorted(described))}.",
        )
        return
    report.facts["balance_export_sheet"] = found[0]
    report.add(
        "balance_export_sheet",
        True,
        f"Sheet {found[0]!r} holds {scenario} {year} in "
        f"{report.facts.get('balance_export_units', 'unknown units')}.",
    )


# ---------------------------------------------------------------------------
# Balance review from a LEAP export
# ---------------------------------------------------------------------------

#: Columns an ESTO base table must carry to be usable as a source.
ESTO_TABLE_COLUMNS = ("economy", "flows", "products", "is_subtotal")


def validate_balance_review_from_export_inputs(
    *,
    economy: str,
    scenario: str,
    year: int,
    balance_export_workbook: Path | str | None,
    esto_table_path: Path | str | None,
    mapping_workbook_path: Path | str | None,
    bundled_esto_table: Path | str | None,
    projection_table: Path | str | None,
    extra_years: Sequence[int] = (),
) -> ValidationReport:
    """Validate the inputs for a LEAP-export-to-workbook run.

    The extra work over :func:`validate_balance_review_inputs` is the source
    tables: the run needs an ESTO base table and the 9th-edition projection
    table, and when a user substitutes their own ESTO table it is checked
    against the flow/product vocabulary the canonical mapping expects.
    """
    report = ValidationReport(command="balance-review-from-export")

    mapping_path = Path(str(mapping_workbook_path)) if mapping_workbook_path else None
    if mapping_path is None:
        report.add("mapping_workbook_contract", False, "No mapping workbook was supplied.")
    else:
        try:
            from codebase.utilities.mapping_workbook_resolver import validate_mapping_workbook

            validate_mapping_workbook(mapping_path)
        except Exception as exc:  # noqa: BLE001
            report.add(
                "mapping_workbook_contract",
                False,
                f"The staged mapping workbook does not satisfy the initialisation contract: {exc}",
            )
        else:
            report.add(
                "mapping_workbook_contract",
                True,
                f"The staged mapping workbook satisfies the initialisation contract: {mapping_path}",
            )

    try:
        economy_code = normalize_economy(economy)
        report.facts["economy"] = economy_code
    except ValueError as exc:
        report.add("economy", False, str(exc))
    try:
        scenario_name = normalize_scenario(scenario)
        report.facts["scenario"] = scenario_name
    except ValueError as exc:
        report.add("scenario", False, str(exc))
        scenario_name = ""

    try:
        year_value = int(year)
    except (TypeError, ValueError):
        report.add("year", False, f"The year must be a four-digit number, got {year!r}.")
        year_value = 0
    else:
        if not 1990 <= year_value <= 2100:
            report.add(
                "year", False, f"The year {year_value} is outside the supported range 1990-2100."
            )
        else:
            report.add("year", True, f"Review year {year_value}.")
        report.facts["year"] = year_value

    # Additional years are checked here too. Left to the run, a bad third year
    # surfaces only after the first two have each taken several minutes.
    extra_values: list[int] = []
    for extra in extra_years:
        try:
            extra_value = int(extra)
        except (TypeError, ValueError):
            report.add(
                f"year_{extra}", False,
                f"The year must be a four-digit number, got {extra!r}.",
            )
            continue
        if not 1990 <= extra_value <= 2100:
            report.add(
                f"year_{extra_value}", False,
                f"The year {extra_value} is outside the supported range 1990-2100.",
            )
            continue
        report.add(f"year_{extra_value}", True, f"Review year {extra_value}.")
        extra_values.append(extra_value)
    if extra_years:
        report.facts["years"] = [year_value, *extra_values]

    workbook_path = Path(str(balance_export_workbook)) if balance_export_workbook else None
    workbook_ok = _check_readable_file(
        report,
        workbook_path,
        name="balance_export_workbook",
        description="LEAP balance-export workbook",
        suffixes=(".xlsx",),
    )
    if workbook_ok and workbook_path is not None and scenario_name and year_value:
        for value in [year_value, *extra_values]:
            _check_balance_export_sheet(
                report, workbook_path, scenario=scenario_name, year=value
            )
        _check_balance_export_level2(report, workbook_path)

    _check_readable_file(
        report,
        Path(str(projection_table)) if projection_table else None,
        name="ninth_projection_table",
        description="9th-edition projection table shipped with this release",
        suffixes=(".csv",),
    )

    active_esto = esto_table_path or bundled_esto_table
    esto_ok = _check_readable_file(
        report,
        Path(str(active_esto)) if active_esto else None,
        name="esto_base_table",
        description=(
            "ESTO base table you supplied"
            if esto_table_path
            else "ESTO base table shipped with this release"
        ),
        suffixes=(".csv",),
    )
    report.facts["esto_table_is_user_supplied"] = bool(esto_table_path)

    if esto_ok and active_esto is not None:
        _check_esto_vintage(
            report,
            Path(str(active_esto)),
            packaged_esto_table=bundled_esto_table,
            is_user_supplied=bool(esto_table_path),
        )

    if esto_ok and active_esto is not None:
        esto_path = Path(str(active_esto))
        if _check_columns(
            report,
            esto_path,
            ESTO_TABLE_COLUMNS,
            name="columns:esto_base_table",
            description="ESTO base table",
        ):
            _check_esto_vocabulary(
                report,
                esto_path,
                mapping_workbook_path=mapping_workbook_path,
            )

    return report


def _check_balance_export_level2(report: ValidationReport, workbook_path: Path) -> None:
    """Confirm the export carries Level 2+ detail, which diagnostics require.

    LEAP writes Level 2 child rows with leading spaces in column A. A Level 1
    export imports and looks fine, then produces a comparison with no sector
    detail at all, so this is checked up front rather than deep in the run.
    """
    try:
        from openpyxl import load_workbook

        workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001
        report.add("balance_export_detail", False, f"Could not inspect the export: {exc}")
        return
    try:
        for sheet in workbook.worksheets:
            for row in range(3, min(sheet.max_row or 3, 200) + 1):
                value = sheet.cell(row=row, column=1).value
                if isinstance(value, str) and value.startswith(" ") and value.strip():
                    report.add(
                        "balance_export_detail",
                        True,
                        "The export has Level 2+ detail (indented branch rows found).",
                    )
                    return
    finally:
        workbook.close()
    report.add(
        "balance_export_detail",
        False,
        "This export looks like a LEAP Level 1 Energy Balance (no indented "
        "branch rows). Re-export it with at least Level 2 detail, or the "
        "comparison will have no sector detail to work with.",
    )


def _check_esto_vintage(
    report: ValidationReport,
    active_esto: Path,
    *,
    packaged_esto_table: Path | str | None,
    is_user_supplied: bool,
) -> None:
    """Report the base year in use, and refuse a run that mixes ESTO issues.

    The base year is derived from the table rather than configured, so a new
    ESTO issue moves it automatically. The failure this guards against is
    subtler: a supplied table from a newer issue works for the balance review
    but not for the dashboard, whose comparison data was prepared from the
    issue the release was built on. Without this the mismatch is invisible.
    """
    from codebase.portable_release import esto_vintage as vintage_module

    try:
        vintage = vintage_module.infer_esto_vintage(active_esto)
    except vintage_module.EstoVintageError as exc:
        report.add("esto_vintage", False, str(exc))
        return

    report.facts["esto_base_year"] = vintage.base_year
    report.facts["esto_vintage_years"] = vintage.label
    report.add("esto_vintage", True, vintage.describe())

    # A new ESTO issue moves the base year, and the diagnostics step cannot
    # review a year before it. Left to the step itself, the refusal says
    # "pre-base historical years are not yet supported" without mentioning that
    # the base year moved or what to ask for instead - which is exactly the
    # situation a user is in the first time they supply a newer table.
    requested_year = report.facts.get("year")
    if isinstance(requested_year, int) and requested_year < vintage.base_year:
        report.add(
            "year_within_esto_vintage",
            False,
            f"You asked to review {requested_year}, but this ESTO data starts its "
            f"comparison at {vintage.base_year}.\n"
            f"    {vintage.path.name} covers {vintage.label}, so {vintage.base_year} "
            "is its base year and the earliest year that can be reviewed.\n"
            f"    Re-run with --year {vintage.base_year}, or use an ESTO file whose "
            f"base year is {requested_year}.",
        )

    if not is_user_supplied or not packaged_esto_table:
        return
    try:
        packaged = vintage_module.infer_esto_vintage(Path(str(packaged_esto_table)))
    except vintage_module.EstoVintageError:
        return
    # Reported, not refused: the worker re-derives the comparison rows from
    # whichever table it is given, so a different issue is a supported choice
    # rather than a mismatch to block.
    for note in vintage_module.describe_vintage_change(
        supplied=vintage,
        packaged_base_year=packaged.base_year,
    ):
        report.add("esto_vintage_change", True, note)


def _check_esto_vocabulary(
    report: ValidationReport,
    esto_path: Path,
    *,
    mapping_workbook_path: Path | str | None,
) -> None:
    """Compare a supplied ESTO table against the mapping's expected vocabulary.

    The authoritative reference is the ``leap_combined_esto`` sheet of the
    canonical mapping workbook — the flows and products the mapping actually
    reads from the source table. (The workbook's ``ESTO unique flows and
    products`` sheet is a reference listing and is not kept in step with it.)

    A missing flow or product is reported, not fatal: the diagnostics step
    applies ``synthetic_reference_rows.csv``, which is what adds the zero-valued
    rows a newer ESTO vintage introduced. This check exists so that what was
    filled in, and what could not be, is visible before the run rather than
    inferred from the numbers afterwards.
    """
    if not mapping_workbook_path or not Path(str(mapping_workbook_path)).is_file():
        report.add(
            "esto_vocabulary",
            True,
            "Skipped the ESTO vocabulary check: the mapping workbook is not "
            "available to compare against.",
        )
        return
    try:
        import pandas as pd

        sheet = pd.read_excel(Path(str(mapping_workbook_path)), sheet_name="leap_combined_esto")
    except Exception as exc:  # noqa: BLE001
        report.add(
            "esto_vocabulary",
            True,
            f"Skipped the ESTO vocabulary check: could not read the mapping workbook ({exc}).",
        )
        return

    def tokens(value: object) -> list[str]:
        return [part.strip() for part in str(value or "").split("|") if part.strip()]

    expected_flows: set[str] = set()
    expected_products: set[str] = set()
    for _, row in sheet.iterrows():
        expected_flows.update(tokens(row.get("esto_flow")))
        expected_products.update(tokens(row.get("esto_product")))

    present_flows: set[str] = set()
    present_products: set[str] = set()
    with esto_path.open("r", encoding="utf-8-sig", newline="") as handle:
        for record in csv.DictReader(handle):
            present_flows.add(str(record.get("flows") or "").strip())
            present_products.add(str(record.get("products") or "").strip())

    # Combined categories are built by the mapping layer and never appear as raw
    # source rows, so they must not be reported as missing. Three shapes exist:
    # hyphen ranges ("04-05 International transport (bunkers)"), comma lists
    # ("09.01.01,09.02.01 Electricity plants"), and own-use rollups
    # ("09.08.01 Coke ovens (including own use)").
    def is_raw(entry: str) -> bool:
        code = entry.split(" ", 1)[0]
        if "-" in code or "," in code:
            return False
        return "(including own use)" not in entry.lower()

    missing_flows = sorted(f for f in expected_flows - present_flows if is_raw(f))
    missing_products = sorted(p for p in expected_products - present_products if is_raw(p))
    report.facts["esto_expected_flows"] = len(expected_flows)
    report.facts["esto_missing_flows"] = missing_flows
    report.facts["esto_missing_products"] = missing_products

    if not missing_flows and not missing_products:
        report.add(
            "esto_vocabulary",
            True,
            f"The ESTO table carries every flow and product the mapping reads "
            f"({len(present_flows)} flows, {len(present_products)} products).",
        )
        return

    def preview(values: list[str]) -> str:
        head = ", ".join(values[:8])
        return head + (f" (and {len(values) - 8} more)" if len(values) > 8 else "")

    parts = []
    if missing_flows:
        parts.append(f"{len(missing_flows)} flow(s): {preview(missing_flows)}")
    if missing_products:
        parts.append(f"{len(missing_products)} product(s): {preview(missing_products)}")
    report.add(
        "esto_vocabulary",
        True,
        "The ESTO table is missing some entries the mapping reads — "
        + "; ".join(parts)
        + ". The diagnostics step will fill any of these that its "
        "synthetic-reference-row rules cover (this is how Datacentres and the "
        "hydrogen rows are handled); anything left over simply has no source "
        "comparator and is reported as unavailable rather than as zero.",
    )


# ---------------------------------------------------------------------------
# Dashboard
# ---------------------------------------------------------------------------


def validate_dashboard_inputs(
    *,
    economy: str,
    comparison_data_path: Path | str | None,
    common_rows_path: Path | str | None,
    template_path: Path | str | None,
    series_config_path: Path | str | None,
    comparison_scope: str = "esto_leap_ninth",
) -> ValidationReport:
    """Validate the inputs for one Common ESTO dashboard render."""
    report = ValidationReport(command="dashboard")

    try:
        economy_code = normalize_economy(economy)
    except ValueError as exc:
        report.add("economy", False, str(exc))
        economy_code = ""
    compact = economy_code.replace("_", "")
    report.facts["economy"] = economy_code
    report.facts["economy_dashboard_key"] = compact
    report.facts["comparison_scope"] = comparison_scope

    comparison_path = Path(str(comparison_data_path)) if comparison_data_path else None
    rows_path = Path(str(common_rows_path)) if common_rows_path else None

    comparison_ok = _check_readable_file(
        report,
        comparison_path,
        name="comparison_data",
        description="Common ESTO comparison data file",
        suffixes=(".csv",),
    )
    if comparison_ok and comparison_path is not None:
        _check_columns(
            report,
            comparison_path,
            DASHBOARD_COMPARISON_COLUMNS,
            name="columns:comparison_data",
            description="Common ESTO comparison data file",
        )

    rows_ok = _check_readable_file(
        report,
        rows_path,
        name="common_rows",
        description="Common ESTO row metadata file",
        suffixes=(".csv",),
    )
    if rows_ok and rows_path is not None:
        _check_columns(
            report,
            rows_path,
            DASHBOARD_ROWS_COLUMNS,
            name="columns:common_rows",
            description="Common ESTO row metadata file",
        )

    for label, path, key in [
        ("dashboard template", template_path, "template"),
        ("dashboard series configuration", series_config_path, "series_config"),
    ]:
        candidate = Path(str(path)) if path else None
        if _check_readable_file(
            report,
            candidate,
            name=key,
            description=label,
            suffixes=(".json",),
        ) and candidate is not None:
            try:
                json.loads(candidate.read_text(encoding="utf-8"))
            except json.JSONDecodeError as exc:
                report.add(
                    f"parse:{key}",
                    False,
                    f"The {label} is not valid JSON ({exc.msg} at line {exc.lineno}). "
                    "Restore it from the release, or fix the edit that broke it.",
                )
            else:
                report.add(f"parse:{key}", True, f"The {label} parsed successfully.")

    if comparison_ok and comparison_path is not None and economy_code:
        matched, available = _scan_dashboard_scope(
            comparison_path,
            economy_compact=compact,
            comparison_scope=comparison_scope,
        )
        report.facts["comparison_rows_matching_economy"] = matched
        report.facts["economies_available"] = available
        if matched:
            report.add(
                "economy_coverage",
                True,
                f"The comparison data covers {economy_code} in scope "
                f"{comparison_scope!r} ({matched:,} rows).",
            )
        else:
            offered = ", ".join(available[:25]) if available else "no economies"
            more = "" if len(available) <= 25 else f" (and {len(available) - 25} more)"
            report.add(
                "economy_coverage",
                False,
                f"The comparison data has no rows for {economy_code} in scope "
                f"{comparison_scope!r}. It covers: {offered}{more}.",
            )

    return report


def validate_dashboard_from_export_inputs(
    *,
    economy: str,
    export_dir: Path | str | None,
    template_path: Path | str | None,
    series_config_path: Path | str | None,
    mapping_workbook_path: Path | str | None,
    source_branch_fallback_rules_path: Path | str | None,
    all_demand_components_path: Path | str | None,
    mapping_chain_data_assets: dict[str, Path | str | None],
) -> ValidationReport:
    """Validate the inputs for a LEAP-export-to-dashboard run.

    This is the mapping-chain input mode: the caller supplies a folder of LEAP
    balance-export workbooks for one economy, and the pre-built mapping
    artifacts (§2 of the handover) drive the chain that turns them into Common
    ESTO comparison data before the dashboard render, itself validated by
    :func:`validate_dashboard_inputs`.
    """
    report = ValidationReport(command="dashboard-from-export")

    try:
        economy_code = normalize_economy(economy)
        report.facts["economy"] = economy_code
    except ValueError as exc:
        report.add("economy", False, str(exc))

    export_path = Path(str(export_dir)) if export_dir else None
    if export_path is None or not export_path.is_dir():
        report.add(
            "export_dir",
            False,
            f"The LEAP balance export folder does not exist at:\n      {export_path}",
        )
    elif not any(export_path.glob("*.xlsx")):
        report.add(
            "export_dir",
            False,
            f"No .xlsx LEAP balance export files were found in: {export_path}",
        )
    else:
        report.add("export_dir", True, f"Found the LEAP balance export folder: {export_path}")

    for label, path, key in [
        ("dashboard template", template_path, "template"),
        ("dashboard series configuration", series_config_path, "series_config"),
        ("mapping workbook", mapping_workbook_path, "mapping_workbook"),
        (
            "source-branch fallback rules",
            source_branch_fallback_rules_path,
            "source_branch_fallback_rules",
        ),
        (
            "all-demand aggregated components",
            all_demand_components_path,
            "all_demand_aggregated_components",
        ),
    ]:
        candidate = Path(str(path)) if path else None
        _check_readable_file(report, candidate, name=key, description=label)

    for role, path in sorted(mapping_chain_data_assets.items()):
        candidate = Path(str(path)) if path else None
        _check_readable_file(
            report,
            candidate,
            name=f"data:{role}",
            description=f"mapping-chain artifact {role}",
        )

    return report


def _scan_dashboard_scope(
    comparison_path: Path,
    *,
    economy_compact: str,
    comparison_scope: str,
) -> tuple[int, list[str]]:
    """Count matching rows and list the economies present, streaming the file.

    The production comparison file is close to a gigabyte, so this reads it row
    by row rather than loading it into memory just to answer a yes/no question.
    """
    matched = 0
    available: set[str] = set()
    with comparison_path.open("r", encoding="utf-8-sig", newline="") as handle:
        for row in csv.DictReader(handle):
            row_economy = str(row.get("economy") or "").replace("_", "").strip()
            available.add(row_economy)
            if row_economy != economy_compact:
                continue
            if str(row.get("comparison_scope") or "").strip() == comparison_scope:
                matched += 1
    available.discard("")
    return matched, sorted(available)
