#%%
"""Build balance-review workbooks with Python and openpyxl only."""

from __future__ import annotations

import csv
import hashlib
import re
from collections import Counter
from copy import copy
from pathlib import Path
from typing import Callable

from openpyxl import load_workbook
from openpyxl.cell import Cell
from openpyxl.styles import Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from codebase.utilities.leap_balance_export_resolver import (
    balance_export_unit_to_petajoule_multiplier,
)


LEAP_SHEET_NAME = "LEAP Values"
ERROR_SHEET_NAME = "LEAP - Source Error"
FULL_EXPECTED_SHEET_NAME = "Full Expected Source"
OUTPUT_UNITS = "Petajoule"

RED_FONT = "9C0006"
RED_FILL = "FCE8E6"
BLUE_FONT = "1F4E78"
BLUE_FILL = "DDEBF7"
YELLOW_FILL = "FFF2CC"
NO_COMPARATOR_FONT = "5B2C83"
NO_COMPARATOR_FILL = "E4DFEC"
AFFECTED_SUPPLY_FONT = "375623"
AFFECTED_SUPPLY_FILL = "E2F0D9"
AFFECTED_SUPPLY_BORDER = "70AD47"
NEUTRAL_FILL = "F2F2F2"
PJ_NUMBER_FORMAT = "#,##0.00;-#,##0.00;-"
FORMULA_ERROR_PATTERN = re.compile(r"#REF!|#DIV/0!|#VALUE!|#NAME\?|#N/A", re.I)


def _file_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _normalize_label(value: object) -> str:
    return str(value or "").strip()


def _safe_filename_token(value: object) -> str:
    token = re.sub(r"[^A-Za-z0-9_-]+", "_", str(value).strip())
    return token.strip("_") or "unknown"


def _as_number(value: object) -> float | None:
    if value is None or str(value).strip() == "":
        return None
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    return number


def _as_boolean(value: object) -> bool:
    return str(value or "").strip().lower() in {"true", "1", "yes"}


def _load_csv_rows(path: Path, *, optional: bool = False) -> list[dict[str, str]]:
    if optional and not path.exists():
        return []
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def _parse_source_metadata(metadata: object) -> dict[str, object]:
    text = str(metadata or "")
    scenario_match = re.search(r"Scenario:\s*([^,]+)", text, flags=re.I)
    year_match = re.search(r"Year:\s*(\d{4})", text, flags=re.I)
    units_match = re.search(r"Units:\s*(.+)$", text, flags=re.I)
    if not scenario_match or not year_match or not units_match:
        raise ValueError(
            "Source metadata must declare Scenario, Year, and Units: "
            f"{text}"
        )
    return {
        "scenario": scenario_match.group(1).strip(),
        "year": int(year_match.group(1)),
        "units": units_match.group(1).strip().rstrip("."),
    }


def _source_unit_to_pj_multiplier(units: object) -> float:
    return balance_export_unit_to_petajoule_multiplier(units)


def _convert_source_sheet_to_petajoule(
    sheet: Worksheet,
    *,
    source_rows: int,
    source_columns: int,
    multiplier: float,
) -> None:
    if multiplier == 1:
        return
    for row in sheet.iter_rows(
        min_row=4,
        max_row=source_rows,
        min_col=2,
        max_col=source_columns,
    ):
        for cell in row:
            if isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
                cell.value *= multiplier


def _copy_sheet_view(source: Worksheet, destination: Worksheet) -> None:
    destination.sheet_view.showGridLines = source.sheet_view.showGridLines
    destination.freeze_panes = "B4"
    destination.sheet_properties = copy(source.sheet_properties)
    destination.page_margins = copy(source.page_margins)
    destination.page_setup = copy(source.page_setup)
    destination.print_options = copy(source.print_options)


def _clear_diagnostic_values(
    sheet: Worksheet,
    *,
    source_rows: int,
    source_columns: int,
    neutral_fill: bool = False,
) -> None:
    fill = PatternFill("solid", fgColor=NEUTRAL_FILL)
    for row in sheet.iter_rows(
        min_row=4,
        max_row=source_rows,
        min_col=2,
        max_col=source_columns,
    ):
        for cell in row:
            cell.value = None
            if neutral_fill:
                cell.fill = copy(fill)


def _set_diagnostic_title(sheet: Worksheet, title: str, explanation: str) -> None:
    sheet["A1"] = title
    sheet["A2"] = explanation
    sheet["A1"].font = Font(bold=True, color="1F1F1F")
    sheet["A2"].font = Font(italic=True, color="595959")


def _style_cell(
    cell: Cell,
    *,
    fill_color: str,
    font_color: str,
    bold: bool = False,
) -> None:
    cell.fill = PatternFill("solid", fgColor=fill_color)
    cell.font = Font(color=font_color, bold=bold)
    cell.number_format = PJ_NUMBER_FORMAT


def _make_structure_resolver(
    sheet: Worksheet,
    *,
    source_rows: int,
    source_columns: int,
) -> Callable[[object, object], dict[str, object]]:
    row_index: dict[str, list[int]] = {}
    for row in range(4, source_rows + 1):
        normalized = _normalize_label(sheet.cell(row=row, column=1).value)
        row_index.setdefault(normalized, []).append(row)

    fuel_index: dict[str, list[int]] = {}
    for column in range(2, source_columns + 1):
        normalized = _normalize_label(sheet.cell(row=3, column=column).value)
        fuel_index.setdefault(normalized, []).append(column)

    def resolve_row(label: object) -> tuple[list[int], str]:
        raw_label = str(label or "")
        exact = row_index.get(_normalize_label(raw_label), [])
        if len(exact) == 1:
            return exact, "exact_visible_label"
        if len(exact) > 1:
            return exact, "ambiguous_visible_label"

        path_parts = [
            _normalize_label(part)
            for part in raw_label.split("/")
            if _normalize_label(part)
        ]
        if len(path_parts) > 1:
            leaves = row_index.get(path_parts[-1], [])
            indented = [
                row
                for row in leaves
                if re.match(r"^\s+", str(sheet.cell(row=row, column=1).value or ""))
            ]
            if len(indented) == 1:
                return indented, "hierarchy_leaf_indent"
            return leaves, "absent" if not leaves else "ambiguous_hierarchy_leaf"
        return [], "absent"

    def resolve(row_label: object, fuel_label: object) -> dict[str, object]:
        row_candidates, mode = resolve_row(row_label)
        fuel_candidates = fuel_index.get(_normalize_label(fuel_label), [])
        candidate_count = len(row_candidates) * len(fuel_candidates)
        if candidate_count == 1:
            row = row_candidates[0]
            column = fuel_candidates[0]
            return {
                "row": row,
                "column": column,
                "candidate_count": 1,
                "status": "unique",
                "mode": mode,
                "address": f"{get_column_letter(column)}{row}",
            }
        return {
            "row": None,
            "column": None,
            "candidate_count": candidate_count,
            "status": "absent" if candidate_count == 0 else "ambiguous",
            "mode": mode,
            "address": "",
        }

    return resolve


def _preferred_comparison_labels(row: dict[str, str]) -> tuple[str, str]:
    return (
        row.get("leap_balance_row") or row.get("leap_sector_names") or "",
        row.get("leap_balance_fuel") or row.get("leap_fuel_names") or "",
    )


def _preferred_mapping_issue_labels(row: dict[str, str]) -> tuple[str, str]:
    return (
        row.get("leap_flow_name")
        or row.get("leap_flow")
        or row.get("leap_sector_name_full_path")
        or "",
        row.get("leap_product_name") or row.get("leap_product") or "",
    )


def _make_missing_record(
    *,
    category: str,
    economy: object,
    scenario: object,
    year: object,
    row_label: object,
    fuel_label: object,
    leap_value: object,
    status: object,
    details: object,
    resolution: dict[str, object],
    recommendation: object,
) -> list[object]:
    present = (
        "Yes"
        if resolution["status"] == "unique"
        else "Ambiguous"
        if resolution["status"] == "ambiguous"
        else "No"
    )
    return [
        category,
        economy,
        scenario,
        _as_number(year) if _as_number(year) is not None else year,
        row_label,
        fuel_label,
        _as_number(leap_value),
        status,
        details,
        present,
        resolution["candidate_count"],
        resolution["address"],
        resolution["mode"],
        recommendation,
    ]


def _scan_formula_errors(workbook) -> list[str]:
    errors: list[str] = []
    for sheet_name in [ERROR_SHEET_NAME, FULL_EXPECTED_SHEET_NAME]:
        sheet = workbook[sheet_name]
        for row in sheet.iter_rows():
            for cell in row:
                if FORMULA_ERROR_PATTERN.search(str(cell.value or "")):
                    errors.append(f"{sheet_name}!{cell.coordinate}")
    return errors


def build_balance_structure_review_workbook(
    *,
    economy: str,
    source_workbook: Path | str,
    source_sheet_name: str,
    diagnostics_directory: Path | str,
    output_workbook: Path | str,
) -> dict[str, object]:
    """Create the three-sheet balance-review workbook without Node dependencies."""
    source_path = Path(source_workbook)
    diagnostics_dir = Path(diagnostics_directory)
    output_path = Path(output_workbook)
    if not source_path.exists():
        raise FileNotFoundError(source_path)
    source_before = _file_sha256(source_path)

    workbook = load_workbook(source_path, data_only=False)
    if source_sheet_name not in workbook.sheetnames:
        raise KeyError(
            f"Source sheet {source_sheet_name!r} is not present in {source_path}."
        )
    source_sheet = workbook[source_sheet_name]
    for sheet in list(workbook.worksheets):
        if sheet is not source_sheet:
            workbook.remove(sheet)
    source_sheet.title = LEAP_SHEET_NAME
    source_rows = source_sheet.max_row
    source_columns = source_sheet.max_column
    if source_rows < 4 or source_columns < 2:
        raise ValueError(
            f"Expected a populated balance structure, found {source_rows}x{source_columns}"
        )

    title = str(source_sheet["A1"].value or "")
    metadata_text = str(source_sheet["A2"].value or "")
    source_metadata = _parse_source_metadata(metadata_text)
    source_multiplier = _source_unit_to_pj_multiplier(source_metadata["units"])
    _convert_source_sheet_to_petajoule(
        source_sheet,
        source_rows=source_rows,
        source_columns=source_columns,
        multiplier=source_multiplier,
    )
    source_sheet["A2"] = re.sub(
        r"Units:\s*(.+)$",
        f"Units: {OUTPUT_UNITS}",
        metadata_text,
        flags=re.I,
    )
    source_sheet.freeze_panes = "B4"

    differences = _load_csv_rows(diagnostics_dir / "leap_balance_source_differences.csv")
    reviews = _load_csv_rows(diagnostics_dir / "leap_balance_source_review.csv")
    mapping_issues = _load_csv_rows(
        diagnostics_dir / "leap_balance_mapping_issues.csv",
        optional=True,
    )
    projection_diagnostics = _load_csv_rows(
        diagnostics_dir / "ninth_projection_allocation_diagnostics.csv",
        optional=True,
    )

    def matches_source(row: dict[str, str]) -> bool:
        return (
            str(row.get("economy") or "") == economy
            and str(row.get("scenario") or "").lower()
            == str(source_metadata["scenario"]).lower()
            and _as_number(row.get("year")) == float(source_metadata["year"])
        )

    differences = [row for row in differences if matches_source(row)]
    reviews = [row for row in reviews if matches_source(row)]
    mapping_issues = [
        row
        for row in mapping_issues
        if (not row.get("economy") or row.get("economy") == economy)
        and (
            not row.get("scenario")
            or row["scenario"].lower() == str(source_metadata["scenario"]).lower()
        )
        and (
            not row.get("year")
            or _as_number(row["year"]) == float(source_metadata["year"])
        )
    ]
    if not differences or len(differences) != len(reviews):
        raise ValueError(
            "Expected matching non-empty difference/review rows; found "
            f"{len(differences)} differences and {len(reviews)} review rows"
        )

    identity_fields = [
        "economy",
        "scenario",
        "year",
        "esto_flow",
        "esto_product",
        "leap_sector_names",
        "leap_fuel_names",
        "status",
        "difference_pj",
    ]
    for row_number, (difference, review) in enumerate(
        zip(differences, reviews, strict=True),
        start=1,
    ):
        for field in identity_fields:
            if str(difference.get(field) or "") != str(review.get(field) or ""):
                raise ValueError(
                    f"Differences/review comparison row {row_number} disagrees on {field}"
                )

    error_sheet = workbook.copy_worksheet(source_sheet)
    error_sheet.title = ERROR_SHEET_NAME
    full_expected_sheet = workbook.copy_worksheet(source_sheet)
    full_expected_sheet.title = FULL_EXPECTED_SHEET_NAME
    for sheet in [source_sheet, error_sheet, full_expected_sheet]:
        _copy_sheet_view(source_sheet, sheet)
    _clear_diagnostic_values(
        error_sheet,
        source_rows=source_rows,
        source_columns=source_columns,
    )
    _clear_diagnostic_values(
        full_expected_sheet,
        source_rows=source_rows,
        source_columns=source_columns,
        neutral_fill=True,
    )
    _set_diagnostic_title(
        error_sheet,
        f'LEAP - Source Error for Area "{economy}"',
        f"Scenario: {source_metadata['scenario']}, Year: {source_metadata['year']}, "
        f"Units: {OUTPUT_UNITS} | Red = LEAP minus source; purple blank = "
        "seed/carry-forward process; yellow blank = other unavailable comparator; "
        "green outline = supply affected by a seed-derived transformation",
    )
    _set_diagnostic_title(
        full_expected_sheet,
        f'Full Expected Source for Area "{economy}"',
        f"Scenario: {source_metadata['scenario']}, Year: {source_metadata['year']}, "
        f"Units: {OUTPUT_UNITS} | Blue = source-backed expected; purple = "
        "seed/carry-forward process; yellow = other unavailable comparator; "
        "green outline = affected supply; grey = structurally absent",
    )

    resolve = _make_structure_resolver(
        source_sheet,
        source_rows=source_rows,
        source_columns=source_columns,
    )
    missing_records: list[list[object]] = []
    comparison_state_counts = {
        "mapped": 0,
        "no_direct_projection_comparator": 0,
        "reference_unavailable": 0,
        "missing_in_leap": 0,
        "missing_visible_structure": 0,
        "ambiguous_structure_resolution": 0,
    }
    comparable_states: dict[str, dict[str, object]] = {}
    yellow_keys: set[str] = set()
    no_comparator_keys: set[str] = set()
    affected_supply_keys: set[str] = set()
    reconciliation_samples: list[dict[str, object]] = []
    diagnostic_collisions: list[dict[str, object]] = []
    parent_child_mismatch_rows = [
        row
        for row in projection_diagnostics
        if row.get("diagnostic_type") == "parent_child_reconciliation_mismatch"
        and (not row.get("economy") or row.get("economy") == economy)
    ]

    for review in reviews:
        row_label, fuel_label = _preferred_comparison_labels(review)
        resolution = resolve(row_label, fuel_label)
        if _as_boolean(review.get("no_direct_projection_comparator")):
            comparison_state_counts["no_direct_projection_comparator"] += 1
            missing_records.append(
                _make_missing_record(
                    category="no_direct_projection_comparator",
                    economy=review.get("economy"),
                    scenario=review.get("scenario"),
                    year=review.get("year"),
                    row_label=row_label,
                    fuel_label=fuel_label,
                    leap_value=review.get("leap_value_pj"),
                    status=review.get("primary_classification") or review.get("status"),
                    details=review.get("evidence_note")
                    or "No direct projection comparator; process generated from a seed/carry-forward rule.",
                    resolution=resolution,
                    recommendation=review.get("next_action")
                    or "Leave process efficiency and auxiliary values unchanged pending an explicit comparator.",
                )
            )
            if resolution["status"] == "unique":
                no_comparator_keys.add(str(resolution["address"]))
            continue

        status = review.get("status")
        if status in {"missing_in_leap", "reference_unavailable"}:
            comparison_state_counts[status] += 1
            details = review.get("evidence_note") or (
                "No valid raw source comparator is available."
                if status == "reference_unavailable"
                else "The source has a value, but the LEAP balance export has no numeric value for this mapped combination."
            )
            recommendation = review.get("next_action") or (
                "Leave diagnostic cells blank; do not interpret unavailable as zero."
                if status == "reference_unavailable"
                else "Review the mapping and LEAP branch state; do not treat the absent LEAP value as a comparable zero."
            )
            missing_records.append(
                _make_missing_record(
                    category=status,
                    economy=review.get("economy"),
                    scenario=review.get("scenario"),
                    year=review.get("year"),
                    row_label=row_label,
                    fuel_label=fuel_label,
                    leap_value=review.get("leap_value_pj"),
                    status=status,
                    details=details,
                    resolution=resolution,
                    recommendation=recommendation,
                )
            )
            if resolution["status"] == "unique":
                yellow_keys.add(str(resolution["address"]))
            continue

        if resolution["status"] != "unique":
            state = (
                "ambiguous_structure_resolution"
                if resolution["status"] == "ambiguous"
                else "missing_visible_structure"
            )
            comparison_state_counts[state] += 1
            missing_records.append(
                _make_missing_record(
                    category="structure_unresolved",
                    economy=review.get("economy"),
                    scenario=review.get("scenario"),
                    year=review.get("year"),
                    row_label=row_label,
                    fuel_label=fuel_label,
                    leap_value=review.get("leap_value_pj"),
                    status=state,
                    details=(
                        "Diagnostic comparator could not resolve to exactly one visible "
                        f"balance cell ({resolution['mode']})."
                    ),
                    resolution=resolution,
                    recommendation=(
                        "Review the balance row naming/structure; do not allocate the difference."
                    ),
                )
            )
            continue

        comparison_state_counts["mapped"] += 1
        address = str(resolution["address"])
        leap_value = _as_number(review.get("leap_value_pj"))
        source_value = _as_number(review.get("source_value_pj"))
        reported_difference = _as_number(review.get("difference_pj"))
        if leap_value is None or source_value is None or reported_difference is None:
            raise ValueError(
                f"Comparable diagnostic for {address} has a non-numeric value."
            )
        state = comparable_states.setdefault(
            address,
            {
                "leap_value": 0.0,
                "source_value": 0.0,
                "reported_difference": 0.0,
                "has_mismatch": False,
                "component_count": 0,
                "row": resolution["row"],
                "column": resolution["column"],
            },
        )
        state["leap_value"] = float(state["leap_value"]) + leap_value
        state["source_value"] = float(state["source_value"]) + source_value
        state["reported_difference"] = (
            float(state["reported_difference"]) + reported_difference
        )
        state["has_mismatch"] = bool(state["has_mismatch"]) or status == "value_mismatch"
        state["component_count"] = int(state["component_count"]) + 1
        displayed_error = state["reported_difference"] if state["has_mismatch"] else 0

        error_cell = error_sheet[address]
        full_expected_cell = full_expected_sheet[address]
        error_cell.value = displayed_error
        if state["has_mismatch"]:
            _style_cell(
                error_cell,
                fill_color=RED_FILL,
                font_color=RED_FONT,
                bold=True,
            )
            # This sheet is a source-of-truth view, not a derived reconciliation
            # view.  Store the resolved source value directly so it remains
            # auditable even before Excel recalculates workbook formulas.
            full_expected_cell.value = state["source_value"]
        else:
            _style_cell(
                error_cell,
                fill_color=NEUTRAL_FILL,
                font_color="666666",
            )
            full_expected_cell.value = state["source_value"]
        _style_cell(full_expected_cell, fill_color=BLUE_FILL, font_color=BLUE_FONT)

        sample = {
            "owner": review.get("preliminary_owner"),
            "row": row_label,
            "fuel": fuel_label,
            "address": address,
            "status": "value_mismatch" if state["has_mismatch"] else "match",
            "leapValue": state["leap_value"],
            "displayedError": displayed_error,
            "sourceValue": state["source_value"],
        }
        existing_sample = next(
            (item for item in reconciliation_samples if item["address"] == address),
            None,
        )
        if existing_sample:
            existing_sample.update(sample)
        elif len(reconciliation_samples) < 50:
            reconciliation_samples.append(sample)

    for address, state in comparable_states.items():
        if int(state["component_count"]) <= 1:
            continue
        visible = _as_number(
            source_sheet.cell(
                row=int(state["row"]),
                column=int(state["column"]),
            ).value
        )
        tolerance = 1e-6 * max(1, abs(visible or 0))
        if visible is None or abs(float(state["leap_value"]) - visible) > tolerance:
            # Two independently mapped ESTO comparisons can occasionally point
            # to one visible Level-2 LEAP balance cell. Do not add their values:
            # that would present a fabricated source/error value. Keep the
            # source LEAP value visible and mark both diagnostic sheets yellow
            # so the workbook remains usable while the mapping is reviewed.
            yellow_keys.add(address)
            diagnostic_collisions.append(
                {
                    "address": address,
                    "visibleLeapValue": visible,
                    "diagnosticLeapValueSum": state["leap_value"],
                    "componentCount": state["component_count"],
                }
            )

    for review in reviews:
        if not _as_boolean(review.get("no_direct_projection_comparator")):
            continue
        fuel_label = (
            review.get("leap_balance_fuel")
            or review.get("leap_fuel_names")
            or review.get("fuel_label")
        )
        for supply_row in ["Production", "Imports", "Exports"]:
            resolution = resolve(supply_row, fuel_label)
            if resolution["status"] == "unique":
                affected_supply_keys.add(str(resolution["address"]))

    for issue in mapping_issues:
        row_label, fuel_label = _preferred_mapping_issue_labels(issue)
        resolution = resolve(row_label, fuel_label)
        is_boundary = issue.get("reason") == "total_balance_mapping_check"
        missing_records.append(
            _make_missing_record(
                category="aggregate_boundary_error" if is_boundary else "missing_esto_pair",
                economy=issue.get("economy") or economy,
                scenario=issue.get("scenario") or source_metadata["scenario"],
                year=issue.get("year") or source_metadata["year"],
                row_label=row_label,
                fuel_label=fuel_label,
                leap_value=issue.get("value_petajoule"),
                status=issue.get("reason"),
                details=issue.get("details"),
                resolution=resolution,
                recommendation=(
                    "Treat as an aggregate comparison-boundary error, not an ordinary cell mapping."
                    if is_boundary
                    else "Add or review the explicit LEAP-to-ESTO pair before using this cell as a comparator."
                ),
            )
        )
        if not is_boundary and resolution["status"] == "unique":
            yellow_keys.add(str(resolution["address"]))

    for address in yellow_keys:
        for sheet in [error_sheet, full_expected_sheet]:
            sheet[address] = '=""'
            sheet[address].fill = PatternFill("solid", fgColor=YELLOW_FILL)
    for address in no_comparator_keys:
        _style_cell(
            source_sheet[address],
            fill_color=NO_COMPARATOR_FILL,
            font_color=NO_COMPARATOR_FONT,
            bold=True,
        )
        for sheet in [error_sheet, full_expected_sheet]:
            sheet[address] = '=""'
            _style_cell(
                sheet[address],
                fill_color=NO_COMPARATOR_FILL,
                font_color=NO_COMPARATOR_FONT,
                bold=True,
            )

    # Projection parent/child mismatches are a separate contract from the
    # ordinary LEAP-minus-source comparison.  When the projection diagnostic
    # is available, mark the affected child cells red so a seed/carry-forward
    # row is not mistaken for a benign purple comparator gap.
    parent_child_red_keys: dict[str, float] = {}
    for mismatch in parent_child_mismatch_rows:
        child_flows = {
            token.strip()
            for token in str(mismatch.get("child_flows") or "").split(";")
            if token.strip()
        }
        child_products = {
            token.strip()
            for token in str(mismatch.get("esto_products") or "").split(";")
            if token.strip()
        }
        mismatch_value = _as_number(mismatch.get("reconciliation_error")) or 0.0
        for review in reviews:
            review_flow = str(review.get("esto_flow") or "").replace(
                " (including own use)", ""
            ).strip()
            review_product = str(review.get("esto_product") or "").strip()
            if child_flows and review_flow not in child_flows:
                continue
            if child_products and review_product not in child_products:
                continue
            resolution = resolve(*_preferred_comparison_labels(review))
            if resolution["status"] == "unique":
                address = str(resolution["address"])
                parent_child_red_keys[address] = (
                    parent_child_red_keys.get(address, 0.0) + mismatch_value
                )
    for address, mismatch_value in parent_child_red_keys.items():
        _style_cell(
            error_sheet[address],
            fill_color=RED_FILL,
            font_color=RED_FONT,
            bold=True,
        )
        error_sheet[address].value = mismatch_value

    medium_green = Side(style="medium", color=AFFECTED_SUPPLY_BORDER)
    green_border = Border(
        left=medium_green,
        right=medium_green,
        top=medium_green,
        bottom=medium_green,
    )
    for address in affected_supply_keys:
        _style_cell(
            source_sheet[address],
            fill_color=AFFECTED_SUPPLY_FILL,
            font_color=AFFECTED_SUPPLY_FONT,
            bold=True,
        )
        for sheet in [error_sheet, full_expected_sheet]:
            sheet[address].border = copy(green_border)

    accounted = sum(comparison_state_counts.values())
    if accounted != len(reviews):
        raise ValueError(
            f"Comparison audit lost rows: accounted for {accounted} of {len(reviews)}"
        )
    expected_missing = (
        comparison_state_counts["no_direct_projection_comparator"]
        + comparison_state_counts["reference_unavailable"]
        + comparison_state_counts["missing_in_leap"]
        + comparison_state_counts["missing_visible_structure"]
        + comparison_state_counts["ambiguous_structure_resolution"]
        + len(mapping_issues)
    )
    if len(missing_records) != expected_missing:
        raise ValueError(
            f"Missing/audit reconciliation expected {expected_missing} rows, "
            f"found {len(missing_records)}"
        )

    status_counts = Counter(row.get("status") or "" for row in reviews)
    issue_counts = Counter(row.get("reason") or "" for row in mapping_issues)
    formula_error_cells = _scan_formula_errors(workbook)
    if formula_error_cells:
        raise ValueError(f"Formula errors found: {formula_error_cells}")
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcMode = "auto"
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    workbook.close()

    if source_before != _file_sha256(source_path):
        raise RuntimeError("Source workbook changed during build")
    return {
        "sourceWorkbook": str(source_path),
        "sourceSheet": source_sheet_name,
        "outputWorkbook": str(output_path),
        "metadata": {
            "title": title,
            "scenario": source_metadata["scenario"],
            "year": source_metadata["year"],
            "sourceUnits": source_metadata["units"],
            "outputUnits": OUTPUT_UNITS,
            "sourceUnitMultiplier": source_multiplier,
        },
        "sourceShape": {"rows": source_rows, "columns": source_columns},
        "statusCounts": dict(status_counts),
        "issueCounts": dict(issue_counts),
        "comparisonStateCounts": comparison_state_counts,
        "missingAuditRows": len(missing_records),
        "formulaErrorCells": formula_error_cells,
        "formulaPolicy": (
            "Full Expected Source stores resolved source values directly; the "
            "error sheet intentionally displays within-tolerance differences as zero."
        ),
        "reconciliationSamples": reconciliation_samples,
        "diagnosticCollisions": diagnostic_collisions,
        "renders": {},
    }


#%%
