from __future__ import annotations

import copy
import importlib
import json
import os
import re
import shutil
import sys
import time
from contextlib import contextmanager
from datetime import datetime, timezone
from functools import lru_cache
from pathlib import Path
from typing import Iterable

import pandas as pd
from openpyxl.styles import Font, PatternFill

from codebase.supply_reconciliation_config import *  # noqa: F401,F403
from codebase.supply_reconciliation_config import (
    _ModuleCapRule,
    _resolve_module_cap_rule,
    _use_legacy_trade_split_mode,
    _use_output_share_supply_exports_mode,
    _use_capacity_unmet_iterative_mode,
    _use_capacity_unmet_iterative_balanced_mode,
    _use_capacity_unmet_iterative_any_mode,
    _use_capacity_constrained_mode,
    _use_capacity_like_mode,
)
from codebase.utilities.workflow_utils import _resolve
from codebase.utilities import workflow_common
from codebase.utilities.output_paths import BALANCE_TABLES_ROOT, INTEGRATED_LEAP_EXPORTS_ROOT
from codebase.utilities.master_config import MASTER_CONFIG_PATH, config_table_exists, read_config_table
from codebase.configuration import workflow_config as workflow_cfg
from codebase.configuration.all_products_and_flows import ESTO_PRODUCT_LIST, ESTO_SECTORS
from codebase.mappings.canonical_mapping import (
    DEFAULT_BACKUP_LEAP_MAPPINGS,
    DEFAULT_CODEBOOK,
    DEFAULT_NINTH_TO_ESTO,
    DEFAULT_SHEET_MAP,
    build_sector_to_esto_flow_lookup,
    load_canonical_pairs,
    load_fuel_aliases,
    load_sheet_map,
)
from codebase.functions import supply_data_pipeline, leap_api, patch_baseline_seeds
from codebase.functions.analysis_input_write_dispatcher import get_analysis_input_write_mode
from codebase import (
    electricity_heat_interim_workflow,
    other_loss_own_use_proxy_workflow,
    transformation_workflow,
    transfers_workflow,
)
from codebase.utilities.leap_results_dashboard_balance import (
    DEFAULT_BACKUP_MAPPINGS_PATH as DEFAULT_BALANCE_BACKUP_MAPPINGS_PATH,
    DEFAULT_BASE_TABLE_PATH as DEFAULT_BALANCE_BASE_TABLE_PATH,
    DEFAULT_CODEBOOK_PATH as DEFAULT_BALANCE_CODEBOOK_PATH,
    DEFAULT_EXPLICIT_MAPPINGS_PATH as DEFAULT_BALANCE_EXPLICIT_MAPPINGS_PATH,
    DEFAULT_EXPLICIT_REASSIGNMENTS_PATH as DEFAULT_BALANCE_EXPLICIT_REASSIGNMENTS_PATH,
    DEFAULT_MAPPING_PAIRS_PATH as DEFAULT_BALANCE_MAPPING_PAIRS_PATH,
    DEFAULT_PROJECTION_TABLE_PATH as DEFAULT_BALANCE_PROJECTION_TABLE_PATH,
    DEFAULT_REF_WORKBOOK_PATH as DEFAULT_BALANCE_REF_WORKBOOK_PATH,
    DEFAULT_SHEET_MAP_PATH as DEFAULT_BALANCE_SHEET_MAP_PATH,
    DEFAULT_SYNTHETIC_REFERENCE_ROWS_PATH as DEFAULT_BALANCE_SYNTHETIC_REFERENCE_ROWS_PATH,
    DEFAULT_TGT_WORKBOOK_PATH as DEFAULT_BALANCE_TGT_WORKBOOK_PATH,
    build_balance_comparison_esto_axis,
    build_esto_axis_structure_from_dashboard_template,
    convert_leap_balances_to_esto_long_table,
)
from codebase.utilities.leap_balance_export_resolver import resolve_balance_export_workbook
from codebase.utilities.leap_results_dashboard_utils import (
    DEFAULT_EXPLICIT_LEAP_MAPPINGS,
    DEFAULT_EXPLICIT_LEAP_REASSIGNMENTS,
    apply_explicit_sector_reassignments,
    build_comparisons,
    load_explicit_sector_fuel_mappings,
    load_explicit_sector_reassignments,
    load_leap_workbook,
    map_fuel_label,
)
from codebase.scrapbook.utilities import load_augmented_reference_tables
from codebase.utilities.workflow_common import archive_config_dir_once_per_day
from codebase.supply_reconciliation_utils import (
    _canonical_transformation_fuel_label,
    _load_code_to_name_table,
    _normalize_label_for_lookup,
    _normalize_esto_product_for_match,
    _build_label_to_esto_product_lookup,
    _iter_year_value_items,
    _sort_output_frame_for_csv,
    _normalize_template_header_value,
)
from codebase.supply_reconciliation_history import (
    _state_token,
    _capacity_addition_state_key,
    _output_addition_state_key,
    _results_signature_state_key,
    _capacity_unmet_default_state,
    _resolve_capacity_unmet_pass_mode,
    _is_capacity_unmet_baseline_seed_pass,
    _read_capacity_unmet_state,
    _write_capacity_unmet_state,
    _build_results_signature,
    _lookup_runtime_capacity_additions_for_record,
    _lookup_runtime_primary_addition,
    _lookup_runtime_export_adjustment,
)
from codebase.supply_reconciliation_results import (
    _parse_year_column_token,
    _find_supply_results_header_row,
    _read_supply_results_trade_sheet,
    _read_supply_results_import_sheet,
    _read_supply_results_export_sheet,
    _balance_table_csv_candidates,
    _collect_observed_trade_from_balance_tables,
    _select_supply_results_workbook,
    _scenario_filename_candidates,
    _abbreviate_scenario,
    _resolve_refinery_results_workbook,
    _resolve_transformation_results_workbook,
)
from codebase.supply_reconciliation_balance_tables import (
    build_year_balance_table,
    save_year_balance_tables,
    build_conventional_balance_matrix,
    build_reference_conventional_balance_matrix,
    build_conventional_balance_diff_matrix,
    save_conventional_balance_tables,
    _get_refinery_fallback_rows_for_balance,
    _split_sector_codes,
    _sector_code_sequence,
    _select_primary_sector_code,
    _safe_filename_token,
    _filter_balance_scenarios,
    _ensure_current_accounts_scenario,
    _zero_small_numeric_values,
)
import codebase.supply_reconciliation_allocation as _sra

# NOTE: _load_reset_scope_from_full_model_export uses a late import of
#       _extract_catalog_rows_from_full_model_export to avoid circular imports.

# Default scenario list used as the fallback when a caller does not pass
# scenario_names (see run_preflight_compressed_projection).  Mirrors the
# sibling supply modules (supply_leap_io, supply_results_saver) that share this
# `from supply_reconciliation_config import *` header; without it the bare
# `SCENARIOS` reference in run_preflight_compressed_projection raises NameError.
SCENARIOS = list(workflow_cfg.SUPPLY_NOTEBOOK_SCENARIOS)

# Preflight toggle also referenced as a bare global by
# run_preflight_compressed_projection.  It is authored as a notebook-runtime
# override in supply_reconciliation_workflow; this module-level default keeps
# the extracted function self-contained (callers may still override the
# attribute before invoking).  Without it the reference raises NameError.
PREFLIGHT_COMPRESSED_INCLUDE_CURRENT_ACCOUNTS = True

# Module-level cache for _load_reset_scope_from_full_model_export.  That function
# declares `global _RESET_SCOPE_FROM_EXPORT_CACHE` and reads it on entry before
# any assignment runs, so the name must exist at import time or the first call
# raises NameError.
_RESET_SCOPE_FROM_EXPORT_CACHE = None


@contextmanager
def _keep_windows_pc_awake(enabled: bool = True):
    """Prevent Windows system sleep while the wrapped workflow is running."""
    if not enabled or not sys.platform.startswith("win"):
        yield
        return

    try:
        import ctypes

        ES_CONTINUOUS = 0x80000000
        ES_SYSTEM_REQUIRED = 0x00000001

        result = ctypes.windll.kernel32.SetThreadExecutionState(
            ES_CONTINUOUS | ES_SYSTEM_REQUIRED
        )
    except Exception as exc:
        # Only setup failures fall back to a plain yield. Exceptions raised by the
        # wrapped workflow (thrown into this generator at the yield below) must not
        # be caught here, or contextlib sees a second yield after throw() and raises
        # "generator didn't stop after throw()", masking the real error.
        print(f"[WARN] Windows sleep prevention unavailable: {exc}")
        yield
        return

    if result:
        print("[INFO] Windows sleep prevention enabled for this workflow run.")
    else:
        print("[WARN] Could not enable Windows sleep prevention for this workflow run.")
    try:
        yield
    finally:
        ctypes.windll.kernel32.SetThreadExecutionState(ES_CONTINUOUS)
        if result:
            print("[INFO] Windows sleep prevention released.")


def _emit_completion_beep(*, success: bool = True, style: str = "simple") -> None:
    """Emit an audible completion signal (winsound, notebook audio, terminal bell)."""
    if not bool(ENABLE_COMPLETION_BEEP):
        return
    count = max(int(COMPLETION_BEEP_COUNT), 1)
    frequency = max(int(COMPLETION_BEEP_FREQUENCY_HZ), 37)
    duration = max(int(COMPLETION_BEEP_DURATION_MS), 50)
    pause_seconds = max(float(COMPLETION_BEEP_PAUSE_SECONDS), 0.0)
    if not success:
        count = max(count, 2)
        frequency = max(frequency - 180, 37)
        if style == "chime":
            style = "error"

    if style == "chime":
        tone_plan = [(659, 90), (784, 90), (988, 140)]  # E5, G5, B5
        gap_ms = 40
    elif style == "error":
        tone_plan = [(440, 140), (330, 180)]  # A4 -> E4 (descending)
        gap_ms = 60
    else:
        tone_plan = [(frequency, duration)] * count
        gap_ms = int(pause_seconds * 1000)

    try:
        import winsound  # type: ignore

        for index, (freq_hz, tone_duration_ms) in enumerate(tone_plan):
            try:
                winsound.Beep(max(int(freq_hz), 37), max(int(tone_duration_ms), 50))
            except Exception:
                winsound.MessageBeep()
            if gap_ms > 0 and index < len(tone_plan) - 1:
                time.sleep(gap_ms / 1000.0)
        return
    except Exception:
        pass

    # Jupyter kernels often ignore terminal bells; use browser audio when possible.
    try:
        from IPython import get_ipython  # type: ignore
        from IPython.display import Javascript, display  # type: ignore

        ip = get_ipython()
        shell_name = type(ip).__name__ if ip is not None else ""
        if shell_name == "ZMQInteractiveShell":
            tones_js = ", ".join(
                f"{{freq: {max(int(freq_hz), 37)}, durMs: {max(int(tone_duration_ms), 50)}}}"
                for freq_hz, tone_duration_ms in tone_plan
            )
            js = f"""
            (() => {{
              const AudioCtx = window.AudioContext || window.webkitAudioContext;
              if (!AudioCtx) return;
              const tones = [{tones_js}];
              const gapMs = {int(gap_ms)};
              const playOne = (delayMs, freq, durMs) => {{
                setTimeout(() => {{
                  const ctx = new AudioCtx();
                  const osc = ctx.createOscillator();
                  const gain = ctx.createGain();
                  osc.type = "sine";
                  osc.frequency.value = freq;
                  gain.gain.value = 0.045;
                  osc.connect(gain);
                  gain.connect(ctx.destination);
                  osc.start();
                  osc.stop(ctx.currentTime + (durMs / 1000));
                  osc.onended = () => ctx.close();
                }}, delayMs);
              }};
              let cursor = 0;
              for (const tone of tones) {{
                playOne(cursor, tone.freq, tone.durMs);
                cursor += tone.durMs + gapMs;
              }}
            }})();
            """
            display(Javascript(js))
            return
    except Exception:
        pass

    for index, _ in enumerate(tone_plan):
        print("\a", end="", flush=True)
        if gap_ms > 0 and index < len(tone_plan) - 1:
            time.sleep(gap_ms / 1000.0)
    print("", flush=True)


# -----------------------------------------------------------------------------
# Workflow configuration
# -----------------------------------------------------------------------------

"""
HOW THIS WORKFLOW OPERATES
==========================

The goal of this workflow is to build internally consistent supply and
transformation LEAP import files, using LEAP's own computed results to
iteratively close any energy balance gaps.  It runs as a multi-pass loop:
each pass produces workbooks that are imported into LEAP, LEAP is
recalculated, its balance results are exported, and the next pass uses
those results to correct remaining gaps.

-- INPUTS ------------------------------------------------------------------

  Three data sources feed the workflow:

  1. ESTO energy balance data (via supply_data_pipeline and
     transformation_workflow).  The APEC historical energy balance, used
     for base-year anchoring and as the starting projection.

  2. 9th Outlook projection data (the 'ninth' dataset).  Supply and
     transformation projections built from regression models against ESTO.
     These drive the initial baseline estimates for each scenario year.

  3. LEAP balance results (comparison_long.csv, balance tables in
     YEARLY_BALANCE_DIR).  Outputs from a previous LEAP calculation run,
     used in results_update passes to observe what gaps remain after LEAP
     solves with the previous iteration's imports.

-- PASS 1: BASELINE SEED ---------------------------------------------------

  Set CAPACITY_UNMET_PASS_MODE = 'baseline_seed' for the first run.

  The workflow builds a reconciliation table that merges ESTO demand,
  transformation sector activity, and supply projections into a single
  long-format frame indexed by economy / scenario / product / year.  This
  is the master ledger for the run.

  Supply exports are written with imports set to zero.  The idea is to
  show LEAP only what the economy produces domestically and let LEAP's
  own balance calculation reveal where supply falls short of demand.
  Transformation exports are written with process shares, efficiencies,
  feedstock allocations, and auxiliary fuel use derived from ESTO and the
  projection models.  Transfer exports capture upstream/refinery fuel flows.

  Before writing, a zero-reset can be applied
  (RUN_RESET_SUPPLY_AND_TRANSFORMATION_IMPORT_EXPORT) to clear any stale
  values from a prior run in the LEAP branches this workflow owns.  Zero-
  fill rows are also generated for transformation and transfer branches that
  this workflow is responsible for but for which a given economy has no ESTO
  data, so those branches are explicitly set to 0 rather than left with
  whatever was there before.

  The output is a set of LEAP workbooks (supply, transformation, transfers,
  combined) written to EXPORT_OUTPUT_DIR.  These are imported into LEAP and
  LEAP is recalculated manually.

-- BETWEEN PASSES: LEAP RECALCULATION --------------------------------------

  After importing the workbooks and recalculating LEAP:
  - LEAP produces balance results showing, for each product, whether supply
    met demand or whether there is an unmet requirement (a residual gap that
    LEAP could not satisfy from the supplied sources).
  - The results dashboard workflow is run to export those balance results as
    yearly balance tables (CSV files in YEARLY_BALANCE_DIR).
  - The workflow is then re-run with CAPACITY_UNMET_PASS_MODE = 'results_update'.

-- PASS 2+: RESULTS UPDATE -------------------------------------------------

  In results_update mode the workflow reads the LEAP balance tables from
  the previous calculation and, for each economy / scenario / product,
  computes the remaining gap (the observed import volume that LEAP pulled
  in because domestic supply was insufficient).

  It then attempts to close each gap by allocating additions across three
  levers, tried in order:

  1. Transformation capacity additions.  For products with a priority module
     list (CAPACITY_UNMET_PRIORITY_BY_PRODUCT), the workflow adds exogenous
     capacity to the relevant transformation modules in priority order.  Each
     module's headroom is bounded by CAPACITY_UNMET_MODULE_CAPACITY_UPPER_LIMITS
     (expressed as sentinels like KEEP_EXOGENOUS_CAP_SAME_AS_BASE_YEAR_
     ENERGY_OUTPUT so caps stay meaningful as data updates).  Cumulative
     additions across passes are persisted in a JSON state file so each pass
     builds on the previous one rather than starting over.

  2. Primary production additions.  If transformation cannot absorb the full
     gap (e.g. for primary fuels with no relevant transformation module),
     the workflow adds to primary resource production directly.  Also bounded
     by CAPACITY_UNMET_PRODUCTION_UPPER_LIMITS.  Secondary fuels (refined
     products, processed gases, etc.) are treated as transformation outputs
     only -- they cannot be provisioned via primary production, so any
     residual gap for a secondary fuel that exhausts transformation headroom
     goes straight to the import fallback.

  3. Import fallback.  If neither lever can close the remaining gap and
     CAPACITY_UNMET_UNRESOLVED_POSITIVE_POLICY = 'imports_fallback', the
     residual is written as an import.  This is the safety valve -- LEAP
     will balance correctly even if domestic allocation is incomplete.

  The updated workbooks are written and the process repeats: import into
  LEAP -> recalculate -> export results -> run workflow in results_update.
  Repeat until gaps are within tolerance.

-- OWN-USE AND LOSS PROXY --------------------------------------------------

  Not all transformation own-use and losses are modelled directly -- some
  are approximated with a proxy written to the Demand\\Other loss and own use
  LEAP branch.  These are the sectors whose own-use is handled by the proxy:
  oil refineries (10.01.11), oil & gas extraction (10.01.12), liquefaction /
  regasification (10.01.03), patent fuel plants (10.01.08), BKB/PB plants
  (10.01.09), coal-to-oil liquefaction (10.01.10), charcoal processing
  (10.01.15), biogas gasification (10.01.16), pump storage (10.01.13),
  nuclear (10.01.14), and electricity/CHP/heat plants (10.01.01).
  Transmission & distribution losses (10.02) are also proxied for all fuels
  except electricity itself.

  By contrast, the following sectors have their own-use embedded directly as
  auxiliary fuel use rows in their Transformation module processes (via the
  loss_flow_codes mechanism in transformation_analysis_utils.py): gas works
  (10.01.02), coke ovens (10.01.05), blast furnaces (10.01.07), coal mines
  (10.01.06), oil refineries (10.01.11), and non-specified transformation
  (10.01.17).  These are all disabled in the proxy config.

  Electricity distribution losses are handled separately in the Transmission
  & Distribution transformation module rather than via the proxy, because the
  activity driver is production rather than distribution output.

  In the 'first' stage (baseline_seed) the proxy uses ESTO base-year own-use
  ratios extrapolated with 9th projection activity scaling.  In the 'second'
  stage (results_update) the proxy switches to using LEAP balance table values
  for other losses and own use, so the proxy tracks what LEAP actually
  computed in the previous pass rather than the ESTO-derived estimate.

-- TRANSFORMATION REFRESH (CURRENTLY DISABLED -- LEAP API BUGS) ------------

  If REFRESH_TRANSFORMATION_MEASURES_FROM_LEAP_RESULTS = True, the workflow
  queries LEAP's Results layer via the COM API before building transformation
  exports and replaces ESTO-derived efficiency and process share values with
  what LEAP last computed.  Useful in principle when LEAP has been calibrated
  by hand, but currently kept False because of reliability issues in the
  LEAP COM API results layer.

-- CONSTRAINT TABLES (CURRENTLY DORMANT) -----------------------------------

  load_leap_constraint_tables() can read additional fuel-level capacity caps
  from LEAP-format workbooks listed in CONSTRAINT_TEMPLATE_PATHS.  That list
  is currently empty so no file-based caps are active.  The system exists so
  future constraints can be added without code changes.

-- FEEDSTOCK FUEL SHARE ZERO-FILL AND THE 100.0 FALLBACK ------------------

  After transformation rows are built, build_aux_fuel_zero_rows() in
  transformation_analysis_utils.py writes zero-value rows for any catalog
  feedstock fuel branches that were not explicitly set this run.  For LEAP to
  accept an import, at least one Feedstock Fuel Share value per process must be
  non-zero.  The function handles this in two tiers:

  Tier 1 — processes where we actually wrote Feedstock Fuel Share rows (i.e.
  the process had ESTO activity data).  For any sibling feedstock branches that
  were NOT set this run, the first one is written as 100.0 and the rest as 0.0.
  This ensures no process ends up with all-zero shares.

  Tier 2 — in-scope sectors where we had NO ESTO data at all (the workflow
  produced zero process records for that sector).  All catalog branches under
  those sectors are zeroed out, but the 100.0 fallback is intentionally NOT
  applied, because there is no data to anchor which fuel should hold the share.

  Tier 2 now applies the same 100.0 anchor for Feedstock Fuel Share: branches
  are grouped by process prefix and the first branch per process gets 100.0,
  the rest 0.0.  This covers sectors like Gas works plants that exist in LEAP's
  catalog but have zero ESTO activity — previously these fell through to
  all-zero shares which LEAP rejected on import.

-- OUTPUTS -----------------------------------------------------------------

  Each run writes:
  - Supply workbook:           LEAP imports for Resources\\ branches
  - Transformation workbook:   LEAP imports for Transformation\\ branches
  - Transfers workbook:        LEAP imports for Transformation\\Transfers\\ branches
  - Combined workbook:         all three merged for easier review
  - Yearly balance CSVs:       per-product reconciliation tables (also used as
                               input for the next results_update pass)
  - Consolidated run workbook: everything bundled into one Excel for QA
  - Diagnostic reports:        unmatched rows, metadata mismatches, demand
                               issues, timing -- written to RESULTS_CHECKS_DIR
"""

# Runtime mutable globals have moved to supply_reconciliation_allocation.py.
# Access them as _sra._CAPACITY_UNMET_RUNTIME_* from this module.


def _format_scope_preview(
    values: Iterable[object] | None,
    *,
    default_label: str,
    limit: int = 6,
) -> str:
    """Return a compact preview string for scope filters."""
    if not values:
        return default_label
    normalized = [str(item).strip() for item in values if str(item or "").strip()]
    if not normalized:
        return default_label
    if len(normalized) <= limit:
        return ", ".join(normalized)
    head = ", ".join(normalized[:limit])
    return f"{head}, ... (+{len(normalized) - limit} more)"


def _print_reset_mode_reminder(
    *,
    run_economies: Iterable[str],
    run_scenarios: Iterable[str],
) -> None:
    """Warn users when reset mode is off; summarize active scope when on."""
    run_economy_preview = _format_scope_preview(
        run_economies,
        default_label="all run economies",
    )
    run_scenario_preview = _format_scope_preview(
        run_scenarios,
        default_label="all run scenarios",
    )
    if RUN_RESET_SUPPLY_AND_TRANSFORMATION_IMPORT_EXPORT:
        economy_preview = _format_scope_preview(
            RESET_SCOPE_ECONOMIES,
            default_label=run_economy_preview,
        )
        scenario_preview = _format_scope_preview(
            RESET_SCOPE_SCENARIOS,
            default_label=run_scenario_preview,
        )
        year_preview = _format_scope_preview(
            RESET_SCOPE_YEARS,
            default_label=f"{BASE_YEAR}-{FINAL_YEAR}",
        )
        print(
            "[INFO] Reset reminder: supply/transformation import-export reset is ENABLED "
            f"for economies [{economy_preview}], scenarios [{scenario_preview}], years [{year_preview}]."
        )
        return
    print(
        "[WARN] Reset reminder: supply/transformation import-export reset is DISABLED. "
        "Stale LEAP Imports/Exports/targets may persist across runs. "
        "Set RUN_RESET_SUPPLY_AND_TRANSFORMATION_IMPORT_EXPORT=True "
        "to force zero reset before filling."
    )



def _flatten_reset_scope_values(values_by_group: dict[str, list[str]] | None) -> list[str]:
    """Flatten grouped reset-scope values to a de-duplicated ordered list."""
    ordered: list[str] = []
    seen: set[str] = set()
    for values in (values_by_group or {}).values():
        for item in (values or []):
            token = str(item or "").strip()
            if not token:
                continue
            key = token.lower()
            if key in seen:
                continue
            seen.add(key)
            ordered.append(token)
    return ordered


def _load_reset_scope_from_full_model_export() -> tuple[list[str], list[str]]:
    """Return transformation module/fuel reset scope derived from full-model export."""
    from codebase.functions.supply_results_saver import _extract_catalog_rows_from_full_model_export  # late import — avoids circular
    global _RESET_SCOPE_FROM_EXPORT_CACHE
    if isinstance(_RESET_SCOPE_FROM_EXPORT_CACHE, dict):
        return (
            list(_RESET_SCOPE_FROM_EXPORT_CACHE.get("modules") or []),
            list(_RESET_SCOPE_FROM_EXPORT_CACHE.get("fuels") or []),
        )

    if not RESET_SCOPE_USE_FULL_MODEL_EXPORT:
        _RESET_SCOPE_FROM_EXPORT_CACHE = {
            "modules": [],
            "fuels": [],
            "module_output_fuels": {},
        }
        return [], []

    try:
        rows = _extract_catalog_rows_from_full_model_export(
            source_path=RESULTS_VERIFICATION_EXPORT_PATH,
            sheet_name=RESULTS_VERIFICATION_EXPORT_SHEET,
        )
    except Exception as exc:
        print(
            "[WARN] Failed deriving reset scope from full model export: "
            f"{exc}"
        )
        rows = []

    modules: list[str] = []
    fuels: list[str] = []
    module_output_fuels: dict[str, list[str]] = {}
    seen_modules: set[str] = set()
    seen_fuels: set[str] = set()
    seen_module_output_fuels: dict[str, set[str]] = {}
    for row in rows:
        if str(row.get("catalog_type") or "").strip().lower() != "transformation":
            continue
        module = str(row.get("module_or_root") or "").strip()
        fuel = str(row.get("fuel_name") or "").strip()
        fuel_group = str(row.get("fuel_group") or "").strip().lower()
        if module:
            key = module.lower()
            if key not in seen_modules:
                seen_modules.add(key)
                modules.append(module)
        if fuel:
            key = fuel.lower()
            if key not in seen_fuels:
                seen_fuels.add(key)
                fuels.append(fuel)
        if module and fuel and fuel_group == "output fuels":
            canonical_fuel = _canonical_transformation_fuel_label(fuel)
            if canonical_fuel:
                module_key = module.lower()
                module_bucket = module_output_fuels.setdefault(module_key, [])
                seen_bucket = seen_module_output_fuels.setdefault(module_key, set())
                canonical_key = canonical_fuel.lower()
                if canonical_key not in seen_bucket:
                    seen_bucket.add(canonical_key)
                    module_bucket.append(canonical_fuel)

    if modules or fuels:
        module_scoped_count = sum(
            1 for labels in module_output_fuels.values() if labels
        )
        print(
            "[INFO] Derived reset scope from full model export: "
            f"modules={len(modules)}, fuels={len(fuels)}, "
            f"module_output_scopes={module_scoped_count} "
            f"(source={_resolve(RESULTS_VERIFICATION_EXPORT_PATH)})"
        )
    else:
        print(
            "[WARN] No transformation reset scope derived from full model export "
            f"(source={_resolve(RESULTS_VERIFICATION_EXPORT_PATH)})."
        )

    _RESET_SCOPE_FROM_EXPORT_CACHE = {
        "modules": modules,
        "fuels": fuels,
        "module_output_fuels": module_output_fuels,
    }
    return modules, fuels


def _configured_reset_module_names() -> set[str]:
    """Return normalized module names configured for reset operations."""
    legacy_modules = _flatten_reset_scope_values(TRANSFORMATION_RESET_MODULES)
    manual_modules = _flatten_reset_scope_values(TRANSFORMATION_RESET_MODULES_MANUAL_OVERRIDES)
    derived_modules, _ = _load_reset_scope_from_full_model_export()

    if RESET_SCOPE_USE_FULL_MODEL_EXPORT and derived_modules:
        base_modules = derived_modules
    elif RESET_SCOPE_USE_FULL_MODEL_EXPORT and RESET_SCOPE_REQUIRE_FULL_MODEL_EXPORT:
        raise ValueError(
            "Reset scope requires full model export derivation, but no module scope "
            f"was derived from {RESULTS_VERIFICATION_EXPORT_PATH} "
            f"(sheet={RESULTS_VERIFICATION_EXPORT_SHEET})."
        )
    else:
        base_modules = legacy_modules

    tokens = [
        str(item).strip()
        for item in [*base_modules, *manual_modules]
        if str(item or "").strip()
    ]
    return {token.lower() for token in tokens}


def _configured_reset_fuel_labels() -> list[str]:
    """Return unique configured reset fuel labels (preserve first-seen order)."""
    legacy_fuels = _flatten_reset_scope_values(TRANSFORMATION_RESET_FUELS)
    manual_fuels = _flatten_reset_scope_values(TRANSFORMATION_RESET_FUELS_MANUAL_OVERRIDES)
    _, derived_fuels = _load_reset_scope_from_full_model_export()

    if RESET_SCOPE_USE_FULL_MODEL_EXPORT and derived_fuels:
        base_fuels = derived_fuels
    elif RESET_SCOPE_USE_FULL_MODEL_EXPORT and RESET_SCOPE_REQUIRE_FULL_MODEL_EXPORT:
        raise ValueError(
            "Reset scope requires full model export derivation, but no fuel scope "
            f"was derived from {RESULTS_VERIFICATION_EXPORT_PATH} "
            f"(sheet={RESULTS_VERIFICATION_EXPORT_SHEET})."
        )
    else:
        base_fuels = legacy_fuels

    labels: list[str] = []
    seen: set[str] = set()
    for item in [*base_fuels, *manual_fuels]:
        token = str(item or "").strip()
        if not token:
            continue
        key = token.lower()
        if key in seen:
            continue
        seen.add(key)
        labels.append(token)
    return labels


def _configured_reset_output_fuel_labels_by_module(
    module_names: Iterable[str] | None = None,
) -> dict[str, list[str]]:
    """
    Return module-specific Output Fuels reset labels.

    Keys are lower-cased transformation module names.
    Values are canonicalized fuel labels in first-seen order.
    """
    requested_modules = {
        str(item or "").strip().lower()
        for item in (module_names or [])
        if str(item or "").strip()
    }
    mapping: dict[str, list[str]] = {}
    seen: dict[str, set[str]] = {}

    def _append(module_key: str, fuel_label: object) -> None:
        module_token = str(module_key or "").strip().lower()
        if not module_token:
            return
        canonical = _canonical_transformation_fuel_label(str(fuel_label or ""))
        if not canonical:
            return
        module_seen = seen.setdefault(module_token, set())
        canonical_key = canonical.lower()
        if canonical_key in module_seen:
            return
        module_seen.add(canonical_key)
        mapping.setdefault(module_token, []).append(canonical)

    _load_reset_scope_from_full_model_export()
    cached = _RESET_SCOPE_FROM_EXPORT_CACHE if isinstance(_RESET_SCOPE_FROM_EXPORT_CACHE, dict) else {}
    raw_module_map = cached.get("module_output_fuels") if isinstance(cached, dict) else {}
    if isinstance(raw_module_map, dict):
        for module_key, labels in raw_module_map.items():
            module_token = str(module_key or "").strip().lower()
            if not module_token:
                continue
            if requested_modules and module_token not in requested_modules:
                continue
            for label in (labels or []):
                _append(module_token, label)

    manual_overrides = TRANSFORMATION_RESET_FUELS_MANUAL_OVERRIDES or {}
    if not isinstance(manual_overrides, dict):
        return mapping
    for scope_key, labels in manual_overrides.items():
        scope_token = str(scope_key or "").strip().lower()
        if not scope_token:
            continue
        if scope_token == "all":
            target_modules = (
                set(requested_modules)
                if requested_modules
                else (
                    set(mapping.keys())
                    or set(_configured_reset_module_names())
                )
            )
        else:
            target_modules = {scope_token}
            if requested_modules and scope_token not in requested_modules:
                continue
        for module_token in sorted(target_modules):
            for label in (labels or []):
                _append(module_token, label)
    return mapping


def _is_year_header(value: object) -> bool:
    """Return True when a column header represents a four-digit year."""
    text = str(value).strip()
    if not text:
        return False
    if re.fullmatch(r"\d{4}", text):
        return True
    if re.fullmatch(r"\d{4}\.0", text):
        return True
    return False


SOURCE_DIAGNOSTIC_COLUMNS = [
    "issue_type",
    "workflow",
    "source",
    "economy",
    "scenario",
    "branch_path",
    "esto_flow",
    "esto_product",
    "ninth_sector",
    "ninth_fuel",
    "year",
    "value",
    "suggested_fix",
]


def _build_source_diagnostics(
    *,
    balance_demand_issues: pd.DataFrame | None = None,
    nonzero_missing_id_rows: pd.DataFrame | None = None,
) -> pd.DataFrame:
    """Build one concise diagnostics table for source/mapping issues."""
    rows: list[dict[str, object]] = []

    if balance_demand_issues is not None and not balance_demand_issues.empty:
        for _, row in balance_demand_issues.iterrows():
            rows.append(
                {
                    "issue_type": str(row.get("reason") or row.get("issue_type") or "mapping_issue").strip(),
                    "workflow": "supply_reconciliation",
                    "source": str(row.get("source") or row.get("mapping_source") or "balance_demand").strip(),
                    "economy": str(row.get("economy") or "").strip(),
                    "scenario": str(row.get("scenario") or "").strip(),
                    "branch_path": str(
                        row.get("branch_path")
                        or row.get("Branch Path")
                        or row.get("leap_sector_name_full_path")
                        or ""
                    ).strip(),
                    "esto_flow": str(row.get("esto_flow") or "").strip(),
                    "esto_product": str(row.get("esto_product") or row.get("leap_product_name") or "").strip(),
                    "ninth_sector": str(row.get("ninth_sector") or row.get("mapping_key_sector") or "").strip(),
                    "ninth_fuel": str(row.get("ninth_fuel") or row.get("mapping_key_fuel") or "").strip(),
                    "year": row.get("year", ""),
                    "value": row.get("value", ""),
                    "suggested_fix": str(row.get("suggested_fix") or row.get("note") or "").strip(),
                }
            )

    if nonzero_missing_id_rows is not None and not nonzero_missing_id_rows.empty:
        for _, row in nonzero_missing_id_rows.iterrows():
            year_value = ""
            value = ""
            for col in row.index:
                if _is_year_header(col):
                    year_value = str(col)
                    value = row.get(col, "")
                    break
            rows.append(
                {
                    "issue_type": "missing_full_model_export_branch",
                    "workflow": "supply_reconciliation",
                    "source": "full_model_export",
                    "economy": "",
                    "scenario": str(row.get("Scenario") or "").strip(),
                    "branch_path": str(row.get("Branch Path") or "").strip(),
                    "esto_flow": "",
                    "esto_product": "",
                    "ninth_sector": "",
                    "ninth_fuel": "",
                    "year": year_value,
                    "value": value,
                    "suggested_fix": "Refresh data/full model export.xlsx from LEAP so it includes this branch.",
                }
            )

    diagnostics = pd.DataFrame(rows, columns=SOURCE_DIAGNOSTIC_COLUMNS)
    return diagnostics.drop_duplicates().reset_index(drop=True)


def _write_source_diagnostics(diagnostics: pd.DataFrame) -> Path:
    """Save concise source diagnostics and print a compact end-of-run summary."""
    path = _resolve(RESULTS_CHECKS_DIR) / RESULTS_SOURCE_DIAGNOSTICS_FILENAME
    path.parent.mkdir(parents=True, exist_ok=True)
    _sort_output_frame_for_csv(diagnostics).to_csv(path, index=False)
    if diagnostics.empty:
        print(f"[INFO] Source diagnostics: no issues. Wrote empty diagnostics CSV to {path}.")
        return path

    counts = (
        diagnostics.groupby("issue_type", dropna=False)
        .size()
        .reset_index(name="row_count")
        .sort_values(["row_count", "issue_type"], ascending=[False, True])
    )
    summary = ", ".join(f"{row.issue_type}: {int(row.row_count)}" for row in counts.itertuples(index=False))
    print(f"[WARN] Source diagnostics written to {path}. Counts: {summary}")
    for _, row in diagnostics.head(5).iterrows():
        print(
            "  - {issue_type} | source={source} | scenario={scenario} | branch={branch}".format(
                issue_type=str(row.get("issue_type") or ""),
                source=str(row.get("source") or ""),
                scenario=str(row.get("scenario") or ""),
                branch=str(row.get("branch_path") or ""),
            )
        )
    return path

def _scenario_to_ninth_label(scenario: object) -> str:
    """Map LEAP scenario labels to ninth CSV scenario labels."""
    text = str(scenario or "").strip().lower()
    if text in {"target", "tgt"}:
        return "target"
    return "reference"


def _create_preflight_compressed_source_files(
    *,
    output_dir: Path | str,
    scenario_names: Iterable[str],
    preserve_source_scenarios: bool = False,
    economy_filter: Iterable[str] | None = None,
    file_prefix: str = "preflight",
) -> dict[str, object]:
    """
    Create preflight source files where BASE_YEAR+1 compresses all projection years.

    The ESTO file is copied unchanged so BASE_YEAR remains the configured
    observed year. The ninth file is reduced to non-year columns plus one
    compressed projection year equal to the signed sum across all projection
    years. A companion diagnostics CSV stores abs_sum so future-year activity
    hidden by signed cancellation is still visible.

    Two scenario modes are supported:

    * ``preserve_source_scenarios=False`` (default, used by the baseline
      compressed-projection preflight): future years are summed across *all*
      source scenarios and the single combined series is replicated into each
      requested output scenario. This is fine for the 00_APEC baseline-seed
      approximation, which never compares Reference against Target.
    * ``preserve_source_scenarios=True`` (used by the compressed
      results-update preflight): the source ``scenarios`` column is normalized
      to ``reference``/``target`` and the compression is grouped *within* each
      scenario, so Reference source rows compress into Reference and Target
      source rows compress into Target. Reference and Target are never summed
      together and replicated.

    ``economy_filter`` optionally restricts the compressed ninth output to a
    subset of economies (e.g. ``["20_USA"]``) so the results-update preflight
    stays fast and scoped.
    """
    source_cfg = workflow_cfg.get_energy_source_config()
    base_year = int(source_cfg.esto_base_year)
    compressed_year = base_year + 1
    out_dir = _resolve(output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    esto_out = out_dir / f"{file_prefix}_esto_base_{base_year}.csv"
    ninth_out = out_dir / f"{file_prefix}_ninth_projection_compressed_{compressed_year}.csv"
    abs_diag_out = out_dir / f"{file_prefix}_ninth_projection_compressed_abs_sum_{compressed_year}.csv"

    shutil.copy2(source_cfg.esto_base_table_path, esto_out)

    economy_filter_set = {
        str(item).strip()
        for item in (economy_filter or [])
        if str(item or "").strip()
    }

    header = pd.read_csv(source_cfg.ninth_projection_table_path, nrows=0)
    ninth_columns = list(header.columns)
    if "scenarios" not in ninth_columns:
        raise KeyError("Configured ninth projection table is missing required 'scenarios' column.")
    year_cols = sorted(
        [
            col
            for col in ninth_columns
            if str(col).isdigit() and int(col) > base_year
        ],
        key=lambda col: int(col),
    )
    if not year_cols:
        raise ValueError(
            "Preflight compressed projection could not find any ninth projection "
            f"year columns after base_year={base_year} in {source_cfg.ninth_projection_table_path}."
        )
    non_year_cols = [col for col in ninth_columns if col not in year_cols]
    if preserve_source_scenarios:
        # Keep 'scenarios' in the grouping so Reference and Target stay separate.
        group_cols = list(non_year_cols)
    else:
        group_cols = [col for col in non_year_cols if col != "scenarios"]
    # Always read the full set of non-year columns so economy filtering and
    # scenario normalisation have the columns they need regardless of mode.
    usecols = [*non_year_cols, *year_cols]
    value_cols = [str(compressed_year), f"{compressed_year}_abs_sum"]
    grouped_parts: list[pd.DataFrame] = []
    for chunk in pd.read_csv(
        source_cfg.ninth_projection_table_path,
        usecols=usecols,
        low_memory=False,
        chunksize=200_000,
    ):
        if economy_filter_set and "economy" in chunk.columns:
            chunk = chunk[chunk["economy"].astype(str).str.strip().isin(economy_filter_set)]
            if chunk.empty:
                continue
        values = chunk[year_cols].apply(pd.to_numeric, errors="coerce").fillna(0.0)
        work = chunk[group_cols].copy()
        if preserve_source_scenarios:
            work["scenarios"] = work["scenarios"].map(_scenario_to_ninth_label)
        work[str(compressed_year)] = values.sum(axis=1)
        work[f"{compressed_year}_abs_sum"] = values.abs().sum(axis=1)
        grouped_parts.append(
            work.groupby(group_cols, dropna=False, as_index=False)[value_cols].sum()
        )
    if grouped_parts:
        grouped = (
            pd.concat(grouped_parts, ignore_index=True)
            .groupby(group_cols, dropna=False, as_index=False)[value_cols]
            .sum()
        )
    else:
        grouped = pd.DataFrame(columns=[*group_cols, *value_cols])

    if preserve_source_scenarios:
        # 'scenarios' is already a grouping column with normalized labels.
        compressed = grouped
    else:
        csv_scenarios = sorted(
            {
                _scenario_to_ninth_label(scenario)
                for scenario in scenario_names
                if str(scenario or "").strip().lower() not in {"current accounts", "current account"}
            }
        ) or ["reference"]
        parts = []
        for csv_scenario in csv_scenarios:
            scenario_frame = grouped.copy()
            scenario_frame["scenarios"] = csv_scenario
            parts.append(scenario_frame)
        compressed = pd.concat(parts, ignore_index=True)

    # Restore a familiar column order for downstream readers.
    ordered_non_year = []
    for col in non_year_cols:
        if col == "scenarios":
            ordered_non_year.append(col)
        elif col in compressed.columns:
            ordered_non_year.append(col)
    for col in compressed.columns:
        if col not in ordered_non_year and col not in set(value_cols):
            ordered_non_year.append(col)
    compressed[[*ordered_non_year, str(compressed_year)]].to_csv(ninth_out, index=False)
    compressed[[*ordered_non_year, f"{compressed_year}_abs_sum"]].rename(
        columns={f"{compressed_year}_abs_sum": "compressed_projection_abs_sum"}
    ).to_csv(abs_diag_out, index=False)

    return {
        "esto_path": esto_out,
        "ninth_path": ninth_out,
        "ninth_abs_diagnostics_path": abs_diag_out,
        "base_year": base_year,
        "compressed_year": compressed_year,
    }


def _snapshot_preflight_state() -> dict[str, object]:
    """Capture mutable module/config globals changed by compressed preflight."""
    names = [
        "ECONOMIES",
        "SCENARIOS",
        "FINAL_YEAR",
        "BALANCE_EXPORT_YEARS",
        "OUTPUT_DIR",
        "EXPORT_OUTPUT_DIR",
        "TRANSFORMATION_EXPORT_OUTPUT_DIR",
        "YEARLY_BALANCE_DIR",
        "CONVENTIONAL_BALANCE_DIR",
        "RESULTS_CHECKS_DIR",
        "RESULTS_RUNTIME_DIR",
        "RESULTS_SINGLE_FILE_NAME",
        "LEAP_IMPORT_SUPPLY_TO_LEAP",
        "LEAP_IMPORT_TRANSFORMATION_TO_LEAP",
        "LEAP_IMPORT_TRANSFERS_TO_LEAP",
        "LEAP_IMPORT_CREATE_BRANCHES",
        "LEAP_IMPORT_FILL_BRANCHES",
        "LEAP_IMPORT_INCLUDE_CURRENT_ACCOUNTS",
        "AGGREGATED_DEMAND_INCLUDE_IN_LEAP_IMPORT",
        "OTHER_LOSS_OWN_USE_INCLUDE_IN_LEAP_IMPORT",
        "ZERO_OTHER_DEMAND_INCLUDE_IN_LEAP_IMPORT",
        "RUN_LEAP_FUEL_BRANCH_PROBE_AT_START",
        "SCRAPE_LEAP_RESULTS",
        "TRANSFORMATION_SUPPLY_CACHE_ENABLED",
        "SKIP_ECONOMIES_WITH_EXISTING_EXPORTS",
        "DIRECT_DEMAND_BASE_TABLE_PATH",
        "DIRECT_DEMAND_PROJECTION_TABLE_PATH",
        "DIRECT_DEMAND_BASE_YEAR",
        "DIRECT_DEMAND_PROJECTION_YEARS",
        "BALANCE_DEMAND_BASE_TABLE_PATH",
        "BALANCE_DEMAND_PROJECTION_TABLE_PATH",
    ]
    workflow_cfg_names = [
        "ENERGY_SOURCE_ESTO_BASE_TABLE_PATH",
        "ENERGY_SOURCE_ESTO_BASE_YEAR",
        "ENERGY_SOURCE_NINTH_PROJECTION_TABLE_PATH",
        "ENERGY_SOURCE_PROJECTION_START_YEAR",
        "ENERGY_SOURCE_PROJECTION_FINAL_YEAR",
        "GLOBAL_FINAL_YEAR",
        "TRANSFORMATION_EXPORT_FINAL_YEAR",
        "MINOR_DEMAND_EXPORT_FINAL_YEAR",
        "BASELINE_SEED_VALIDATION_BASE_YEAR",
        "BASELINE_SEED_VALIDATION_FINAL_YEAR",
    ]
    return {
        "globals": {name: globals().get(name) for name in names},
        "workflow_cfg": {name: getattr(workflow_cfg, name, None) for name in workflow_cfg_names},
    }


def _restore_preflight_state(state: dict[str, object]) -> None:
    """Restore globals/config changed for compressed preflight."""
    for name, value in state.get("workflow_cfg", {}).items():
        setattr(workflow_cfg, name, value)
    for name, value in state.get("globals", {}).items():
        globals()[name] = value
    importlib.reload(supply_data_pipeline)
    importlib.reload(transformation_workflow.core)
    importlib.reload(transformation_workflow)
    importlib.reload(transfers_workflow)
    importlib.reload(electricity_heat_interim_workflow)
    importlib.reload(other_loss_own_use_proxy_workflow)


def _apply_preflight_compressed_state(
    *,
    source_files: dict[str, object],
    preflight_root: Path,
    scenarios: list[str],
) -> list[tuple[str, str, object]]:
    """Patch module globals for the compressed two-year preflight run."""
    base_year = int(source_files["base_year"])
    compressed_year = int(source_files["compressed_year"])

    workflow_cfg.ENERGY_SOURCE_ESTO_BASE_TABLE_PATH = str(source_files["esto_path"])
    workflow_cfg.ENERGY_SOURCE_ESTO_BASE_YEAR = base_year
    workflow_cfg.ENERGY_SOURCE_NINTH_PROJECTION_TABLE_PATH = str(source_files["ninth_path"])
    workflow_cfg.ENERGY_SOURCE_PROJECTION_START_YEAR = compressed_year
    workflow_cfg.ENERGY_SOURCE_PROJECTION_FINAL_YEAR = compressed_year
    workflow_cfg.GLOBAL_FINAL_YEAR = compressed_year
    workflow_cfg.TRANSFORMATION_EXPORT_FINAL_YEAR = compressed_year
    workflow_cfg.MINOR_DEMAND_EXPORT_FINAL_YEAR = compressed_year
    workflow_cfg.BASELINE_SEED_VALIDATION_BASE_YEAR = base_year
    workflow_cfg.BASELINE_SEED_VALIDATION_FINAL_YEAR = compressed_year

    importlib.reload(supply_data_pipeline)
    importlib.reload(transformation_workflow.core)
    importlib.reload(transformation_workflow)
    importlib.reload(transfers_workflow)
    importlib.reload(electricity_heat_interim_workflow)
    importlib.reload(other_loss_own_use_proxy_workflow)

    overrides: dict[str, object] = {
            "ECONOMIES": ["00_APEC"],
            "SCENARIOS": list(scenarios),
            "FINAL_YEAR": compressed_year,
            "BALANCE_EXPORT_YEARS": [base_year, compressed_year],
            "OUTPUT_DIR": preflight_root,
            "EXPORT_OUTPUT_DIR": preflight_root / "workbooks",
            "TRANSFORMATION_EXPORT_OUTPUT_DIR": preflight_root / "workbooks",
            "YEARLY_BALANCE_DIR": preflight_root / "yearly_balance_tables",
            "CONVENTIONAL_BALANCE_DIR": preflight_root / "conventional_balance_tables",
            "RESULTS_CHECKS_DIR": preflight_root / "checks",
            "RESULTS_RUNTIME_DIR": preflight_root / "runtime",
            "RESULTS_SINGLE_FILE_NAME": "preflight_compressed_projection_run.xlsx",
            "LEAP_IMPORT_SUPPLY_TO_LEAP": False,
            "LEAP_IMPORT_TRANSFORMATION_TO_LEAP": False,
            "LEAP_IMPORT_TRANSFERS_TO_LEAP": False,
            "LEAP_IMPORT_CREATE_BRANCHES": False,
            "LEAP_IMPORT_FILL_BRANCHES": False,
            "LEAP_IMPORT_INCLUDE_CURRENT_ACCOUNTS": False,
            "AGGREGATED_DEMAND_INCLUDE_IN_LEAP_IMPORT": False,
            "OTHER_LOSS_OWN_USE_INCLUDE_IN_LEAP_IMPORT": False,
            "ZERO_OTHER_DEMAND_INCLUDE_IN_LEAP_IMPORT": False,
            "RUN_LEAP_FUEL_BRANCH_PROBE_AT_START": False,
            "SCRAPE_LEAP_RESULTS": False,
            "TRANSFORMATION_SUPPLY_CACHE_ENABLED": False,
            "SKIP_ECONOMIES_WITH_EXISTING_EXPORTS": False,
            "DIRECT_DEMAND_BASE_TABLE_PATH": source_files["esto_path"],
            "DIRECT_DEMAND_PROJECTION_TABLE_PATH": source_files["ninth_path"],
            "DIRECT_DEMAND_BASE_YEAR": base_year,
            "DIRECT_DEMAND_PROJECTION_YEARS": (compressed_year,),
            "BALANCE_DEMAND_BASE_TABLE_PATH": source_files["esto_path"],
            "BALANCE_DEMAND_PROJECTION_TABLE_PATH": source_files["ninth_path"],
    }
    return _broadcast_config_overrides(overrides)


def run_preflight_compressed_projection(
    *,
    scenario_names: Iterable[str] | None = None,
) -> dict[str, object]:
    """Run a compressed two-year 00_APEC integration preflight before full runs."""
    scenario_list = workflow_common.normalize_workflow_scenarios(
        scenario_names,
        SCENARIOS,
    )
    if not scenario_list:
        scenario_list = ["Reference", "Target"]
    if bool(PREFLIGHT_COMPRESSED_INCLUDE_CURRENT_ACCOUNTS):
        scenario_list = _ensure_current_accounts_scenario(scenario_list)

    preflight_root = _resolve(OUTPUT_DIR) / "preflight_compressed_projection"
    source_dir = preflight_root / "runtime" / "compressed_sources"
    print(
        "[INFO] Starting preflight_compressed_projection: "
        f"economies=['00_APEC'], scenarios={scenario_list}, "
        f"years={BASE_YEAR}-{BASE_YEAR + 1}."
    )
    source_files = _create_preflight_compressed_source_files(
        output_dir=source_dir,
        scenario_names=scenario_list,
    )
    print(
        "[INFO] Compressed ninth projection source written to "
        f"{source_files['ninth_path']} "
        f"(abs-sum diagnostics: {source_files['ninth_abs_diagnostics_path']})."
    )

    # Late import (avoids the supply_reconciliation_workflow <-> supply_preflight
    # import cycle).  Use the workflow wrapper rather than the raw
    # supply_results_saver entry point so its _sync_results_saver_overrides()
    # step still runs — that is the callable the bare global resolved to before
    # this function was extracted out of supply_reconciliation_workflow.
    from codebase.supply_reconciliation_workflow import (
        run_results_linked_transformation_supply_workflow,
    )

    state = _snapshot_preflight_state()
    broadcast_snapshot: list[tuple[str, str, object]] | None = None
    try:
        broadcast_snapshot = _apply_preflight_compressed_state(
            source_files=source_files,
            preflight_root=preflight_root,
            scenarios=scenario_list,
        )
        result = run_results_linked_transformation_supply_workflow(
            economies=["00_APEC"],
            scenario_names=scenario_list,
            export_dataset_key=EXPORT_DATASET_KEY,
            include_leap_import=False,
            import_scenarios=[],
            scrape_leap_results=False,
        )
        result["preflight_compressed_ninth_path"] = source_files["ninth_path"]
        result["preflight_compressed_abs_diagnostics_path"] = source_files["ninth_abs_diagnostics_path"]
        print("[INFO] preflight_compressed_projection completed.")
        return result
    finally:
        if broadcast_snapshot is not None:
            _restore_config_overrides(broadcast_snapshot)
        _restore_preflight_state(state)


# =============================================================================
# Compressed results-update preflight
# =============================================================================
#
# The compressed *results-update* preflight complements the compressed
# *projection* preflight above. Where the projection preflight exercises the
# baseline_seed path with synthetic 00_APEC projection-only demand, this one
# exercises the majority of the results_update path against the *real* 20_USA
# LEAP balance-export structure -- but compressed to only two effective years
# (the ESTO base year and a synthetic BASE_YEAR+1) so it runs in minutes.
#
# It builds temporary two-sheet reduced REF/TGT balance workbooks
# (EBal|<base> copied verbatim, EBal|<base+1> a signed sum of every
# post-base-year source sheet), then runs the workflow in results_update mode
# with all LEAP imports/scraping/caching disabled and every output redirected
# into an isolated preflight root. See
# docs/supply_reconciliation_workflow_guide.md ("Fast preflight checks") and
# docs/special_rules_and_design_decisions.md (INIT entry) for the rationale.


def _is_numeric_cell(value: object) -> bool:
    """Return True for a real numeric cell (excludes bools, NaN, and text)."""
    import numbers

    if value is None or isinstance(value, bool):
        return False
    if isinstance(value, numbers.Real):
        # NaN is the only Real that is not equal to itself.
        return value == value
    return False


def _normalize_balance_grid_label(value: object) -> str:
    """Collapse a balance-sheet label the same way the LEAP balance reader does."""
    return " ".join(str(value if value is not None else "").strip().lower().split())


def _read_balance_sheet_grid(workbook_path: Path | str, sheet_name: str) -> list[list[object]]:
    """Return a LEAP balance sheet as a plain 2D grid (blanks -> None)."""
    raw = pd.read_excel(_resolve(workbook_path), sheet_name=sheet_name, header=None)
    grid = raw.astype(object).where(pd.notna(raw), None).values.tolist()
    return grid


def _detect_balance_header_row(grid: list[list[object]]) -> int:
    """Return the 0-indexed header row (first row within the top 8 whose col0 is blank)."""
    for idx in range(min(8, len(grid))):
        first_cell = grid[idx][0] if grid[idx] else None
        if first_cell is None or not str(first_cell).strip():
            return idx
    # Fall back to the conventional LEAP balance header row.
    return 2 if len(grid) > 2 else 0


def _balance_fuel_columns(grid: list[list[object]], header_row_idx: int) -> dict[str, int]:
    """Map each non-empty fuel header label to its column index (must be unique)."""
    header = grid[header_row_idx] if header_row_idx < len(grid) else []
    fuel_columns: dict[str, int] = {}
    for col_idx, value in enumerate(header):
        if col_idx == 0:
            continue
        label = _normalize_balance_grid_label(value)
        if not label:
            continue
        if label in fuel_columns:
            raise ValueError(
                f"Duplicate fuel column label {label!r} in a future balance sheet header; "
                "cannot align fuel columns safely."
            )
        fuel_columns[label] = col_idx
    return fuel_columns


def _sum_future_balance_grids(
    future_grids: list[tuple[int, list[list[object]]]],
) -> tuple[list[list[object]], list[list[object]], str]:
    """Signed-sum future balance grids into one synthetic grid + an abs-sum grid.

    Real 20_USA balance sheets share an identical balance-row sequence across
    years but carry a *variable* set of fuel columns, so the summation:

    * proves every future sheet has the *same* balance-row identities (col0)
      below the header row -- rows are then matched positionally (the safe,
      unambiguous key, since indented/non-indented labels can repeat);
    * unions the fuel columns by their unique header labels and matches each
      future sheet's fuels by label, so a fuel absent from one year simply
      contributes zero rather than being dropped or positionally corrupted.

    Structural cells (metadata rows, the fuel header row, and col0 labels) come
    from the first future grid (the template). Signs are preserved in
    ``signed_grid``; ``abs_grid`` sums absolute values so categories hidden by
    signed cancellation stay visible. A structural mismatch (differing balance
    rows or duplicate identifying labels) raises a clear diagnostic rather than
    summing by position blindly.
    """
    if not future_grids:
        raise ValueError("Cannot build a synthetic future balance sheet from zero future sheets.")

    header_idx = _detect_balance_header_row(future_grids[0][1])
    template = future_grids[0][1]
    n_rows = len(template)

    # 1. Prove the balance-row sequence (col0 at/below the header) is identical
    #    across every future sheet before matching rows positionally.
    template_row_labels = [_normalize_balance_grid_label(row[0]) for row in template[header_idx + 1 :]]
    for year, grid in future_grids:
        if _detect_balance_header_row(grid) != header_idx or len(grid) != n_rows:
            raise ValueError(
                f"Future balance sheet EBal|{year} has a different row/header layout than the "
                f"template (rows={len(grid)} vs {n_rows}); refusing to sum by position blindly."
            )
        grid_row_labels = [_normalize_balance_grid_label(row[0]) for row in grid[header_idx + 1 :]]
        if grid_row_labels != template_row_labels:
            first_diff = next(
                (
                    idx
                    for idx, (a, b) in enumerate(zip(template_row_labels, grid_row_labels))
                    if a != b
                ),
                min(len(template_row_labels), len(grid_row_labels)),
            )
            raise ValueError(
                f"Future balance sheet EBal|{year} balance-row identities differ from the "
                f"template (first difference at data row {first_diff}); cannot align rows "
                "safely. Refusing to invent or discard unmatched rows silently."
            )

    # 2. Union the fuel columns by their unique header labels (template order
    #    first, then any additional fuels seen only in later years).
    fuel_columns_by_grid: list[dict[str, int]] = [
        _balance_fuel_columns(grid, header_idx) for _, grid in future_grids
    ]
    canonical_fuels: list[str] = []
    seen_fuels: set[str] = set()
    fuel_label_source: dict[str, object] = {}
    for (year, grid), fuel_columns in zip(future_grids, fuel_columns_by_grid):
        for label, col_idx in fuel_columns.items():
            if label not in seen_fuels:
                seen_fuels.add(label)
                canonical_fuels.append(label)
                # Preserve the original (un-normalized) header text for display.
                fuel_label_source[label] = grid[header_idx][col_idx]
    out_cols = 1 + len(canonical_fuels)
    all_identical = all(
        list(fuel_columns.keys()) == canonical_fuels for fuel_columns in fuel_columns_by_grid
    )
    method = "positional_identical" if all_identical else "label_union_aligned"

    # 3. Build structural scaffold (metadata + header + col0 labels) from template.
    signed_grid: list[list[object]] = []
    abs_grid: list[list[object]] = []
    for r in range(n_rows):
        signed_row: list[object] = [None] * out_cols
        # col0 structural label (metadata rows keep their own text).
        signed_row[0] = template[r][0] if template[r] else None
        signed_grid.append(signed_row)
        abs_grid.append(list(signed_row))
    # Rewrite the header row with the canonical fuel labels.
    for fuel_idx, label in enumerate(canonical_fuels):
        signed_grid[header_idx][fuel_idx + 1] = fuel_label_source[label]
        abs_grid[header_idx][fuel_idx + 1] = fuel_label_source[label]
    # Copy metadata rows above the header verbatim (padded to out_cols).
    for r in range(header_idx):
        for c in range(1, out_cols):
            value = template[r][c] if c < len(template[r]) else None
            signed_grid[r][c] = value
            abs_grid[r][c] = value

    # 4. Sum numeric data cells across future years, matched by fuel label.
    for r in range(header_idx + 1, n_rows):
        for fuel_idx, label in enumerate(canonical_fuels):
            out_col = fuel_idx + 1
            signed = 0.0
            absolute = 0.0
            saw_number = False
            for (year, grid), fuel_columns in zip(future_grids, fuel_columns_by_grid):
                col_idx = fuel_columns.get(label)
                if col_idx is None:
                    continue
                cell = grid[r][col_idx] if col_idx < len(grid[r]) else None
                if _is_numeric_cell(cell):
                    saw_number = True
                    signed += float(cell)
                    absolute += abs(float(cell))
            if saw_number:
                signed_grid[r][out_col] = signed
                abs_grid[r][out_col] = absolute
    return signed_grid, abs_grid, method


def _write_balance_grids_workbook(
    output_path: Path,
    ordered_sheets: list[tuple[str, list[list[object]]]],
) -> None:
    """Write ordered (sheet_name, grid) pairs to a plain .xlsx for the balance reader."""
    import numbers as _numbers

    import openpyxl

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for sheet_name, grid in ordered_sheets:
        ws = wb.create_sheet(title=sheet_name)
        for row in grid:
            ws.append(
                [
                    float(cell)
                    if (isinstance(cell, _numbers.Real) and not isinstance(cell, bool))
                    else cell
                    for cell in row
                ]
            )
    wb.save(output_path)
    wb.close()


def _write_balance_abs_sum_diagnostic(diagnostic_path: Path, abs_grid: list[list[object]]) -> None:
    """Write the signed-cancellation abs-sum diagnostic grid as a separate CSV."""
    diagnostic_path.parent.mkdir(parents=True, exist_ok=True)
    frame = pd.DataFrame(abs_grid)
    frame.to_csv(diagnostic_path, index=False, header=False)


def _build_reduced_preflight_balance_workbook(
    *,
    source_path: Path | str,
    output_path: Path,
    base_year: int,
    synthetic_year: int,
    scenario_code: str,
    abs_diagnostic_path: Path,
) -> dict[str, object]:
    """Build a two-sheet reduced LEAP balance workbook for the results-update preflight.

    The reduced workbook contains exactly:

    * ``EBal|<base_year>`` -- copied verbatim from the source workbook.
    * ``EBal|<synthetic_year>`` -- a synthetic future balance sheet built as the
      signed sum of every source ``EBal|YYYY`` sheet where ``YYYY > base_year``.

    The source workbook is never modified. A separate abs-sum diagnostic CSV is
    written alongside so signed cancellation stays visible.
    """
    import openpyxl

    source = _resolve(source_path)
    if not source.exists():
        raise FileNotFoundError(f"Missing source LEAP balance workbook: {source}")

    wb = openpyxl.load_workbook(source, read_only=True, data_only=True)
    try:
        ebal_years: dict[int, str] = {}
        for name in wb.sheetnames:
            if not str(name).strip().lower().startswith("ebal|"):
                continue
            try:
                year = int(str(name).split("|", 1)[1])
            except (ValueError, IndexError):
                continue
            ebal_years[year] = name
    finally:
        wb.close()

    if base_year not in ebal_years:
        raise ValueError(
            f"Source workbook {source.name} has no EBal|{base_year} sheet; cannot build the "
            f"reduced {scenario_code} preflight workbook."
        )
    future_years = sorted(year for year in ebal_years if year > base_year)
    if not future_years:
        raise ValueError(
            f"Source workbook {source.name} has no post-base-year EBal|YYYY sheets "
            f"(base_year={base_year}); cannot build a synthetic future balance sheet."
        )

    base_grid = _read_balance_sheet_grid(source, ebal_years[base_year])
    future_grids = [(year, _read_balance_sheet_grid(source, ebal_years[year])) for year in future_years]
    signed_grid, abs_grid, method = _sum_future_balance_grids(future_grids)

    base_sheet_name = f"EBal|{base_year}"
    synthetic_sheet_name = f"EBal|{synthetic_year}"
    _write_balance_grids_workbook(
        output_path,
        [
            (base_sheet_name, base_grid),
            (synthetic_sheet_name, signed_grid),
        ],
    )
    _write_balance_abs_sum_diagnostic(abs_diagnostic_path, abs_grid)

    print(
        f"[INFO] Built reduced {scenario_code} preflight balance workbook {output_path.name}: "
        f"sheets=[{base_sheet_name}, {synthetic_sheet_name}], synthetic sheet summed "
        f"{len(future_years)} source sheets (EBal|{future_years[0]}..EBal|{future_years[-1]}) "
        f"via {method}. Abs-sum diagnostic: {abs_diagnostic_path.name}."
    )
    return {
        "workbook_path": output_path,
        "abs_diagnostic_path": abs_diagnostic_path,
        "sheet_names": [base_sheet_name, synthetic_sheet_name],
        "source_path": source,
        "future_years": future_years,
        "sum_method": method,
        "scenario_code": scenario_code,
    }


# --- Cross-module config broadcast -------------------------------------------
#
# Config values such as RESULTS_CHECKS_DIR, ECONOMIES, CAPACITY_UNMET_PASS_MODE
# and BALANCE_DEMAND_REF_WORKBOOK_PATH are pulled into many modules via
# `from supply_reconciliation_config import *`, so each consuming module holds
# its own copy. To redirect the whole results_update pipeline into an isolated
# preflight scope we broadcast the overrides to every already-imported
# `codebase.*` module that defines the name, and snapshot the prior values so
# they can be restored exactly (including after a failed preflight).

_BROADCAST_OVERRIDE_MISSING = object()


def _broadcast_config_overrides(overrides: dict[str, object]) -> list[tuple[str, str, object]]:
    """Set config overrides on every loaded codebase module that defines them.

    Returns a snapshot list of ``(module_name, attribute, previous_value)`` for
    restoration. Only names that already exist on a module are overridden, so no
    spurious attributes are created.
    """
    snapshot: list[tuple[str, str, object]] = []
    for module_name, module in list(sys.modules.items()):
        if module is None:
            continue
        if module_name != "codebase" and not module_name.startswith("codebase."):
            continue
        module_dict = getattr(module, "__dict__", None)
        if module_dict is None:
            continue
        for name, value in overrides.items():
            if name in module_dict:
                snapshot.append((module_name, name, module_dict[name]))
                module_dict[name] = value
    return snapshot


def _restore_config_overrides(snapshot: list[tuple[str, str, object]]) -> None:
    """Restore module attributes captured by _broadcast_config_overrides."""
    for module_name, name, previous in reversed(snapshot):
        module = sys.modules.get(module_name)
        if module is not None:
            setattr(module, name, previous)


def _apply_preflight_results_update_state(
    *,
    source_files: dict[str, object],
    reduced_ref_path: Path,
    reduced_tgt_path: Path,
    preflight_root: Path,
    scenarios: list[str],
    economy: str,
) -> list[tuple[str, str, object]]:
    """Patch config for the compressed two-year 20_USA results-update preflight.

    Mirrors the compressed-projection preflight's two-year source setup (compressed
    ESTO/9th + module reloads) but additionally routes the balance-demand loader to
    the temporary reduced REF/TGT workbooks, forces results_update mode, and
    broadcasts every isolation/routing override across all consuming modules.
    Returns the broadcast snapshot for restoration.
    """
    base_year = int(source_files["base_year"])
    compressed_year = int(source_files["compressed_year"])

    workflow_cfg.ENERGY_SOURCE_ESTO_BASE_TABLE_PATH = str(source_files["esto_path"])
    workflow_cfg.ENERGY_SOURCE_ESTO_BASE_YEAR = base_year
    workflow_cfg.ENERGY_SOURCE_NINTH_PROJECTION_TABLE_PATH = str(source_files["ninth_path"])
    workflow_cfg.ENERGY_SOURCE_PROJECTION_START_YEAR = compressed_year
    workflow_cfg.ENERGY_SOURCE_PROJECTION_FINAL_YEAR = compressed_year
    workflow_cfg.GLOBAL_FINAL_YEAR = compressed_year
    workflow_cfg.TRANSFORMATION_EXPORT_FINAL_YEAR = compressed_year
    workflow_cfg.MINOR_DEMAND_EXPORT_FINAL_YEAR = compressed_year
    # Baseline-seed validation must use the same deliberately compressed
    # two-year horizon as the generated preflight workbooks. Production values
    # are restored in ``finally`` by _restore_preflight_state.
    workflow_cfg.BASELINE_SEED_VALIDATION_BASE_YEAR = base_year
    workflow_cfg.BASELINE_SEED_VALIDATION_FINAL_YEAR = compressed_year

    importlib.reload(supply_data_pipeline)
    importlib.reload(transformation_workflow.core)
    importlib.reload(transformation_workflow)
    importlib.reload(transfers_workflow)
    importlib.reload(electricity_heat_interim_workflow)
    importlib.reload(other_loss_own_use_proxy_workflow)

    workbooks_dir = preflight_root / "workbooks"
    overrides: dict[str, object] = {
        # --- scope ---
        "ECONOMIES": [economy],
        "SCENARIOS": list(scenarios),
        "FINAL_YEAR": compressed_year,
        "BALANCE_EXPORT_YEARS": [base_year, compressed_year],
        # --- isolated outputs / runtime / checks ---
        "OUTPUT_DIR": preflight_root,
        "EXPORT_OUTPUT_DIR": workbooks_dir,
        "TRANSFORMATION_EXPORT_OUTPUT_DIR": workbooks_dir,
        "YEARLY_BALANCE_DIR": preflight_root / "yearly_balance_tables",
        "CONVENTIONAL_BALANCE_DIR": preflight_root / "conventional_balance_tables",
        "RESULTS_CHECKS_DIR": preflight_root / "checks",
        "RESULTS_RUNTIME_DIR": preflight_root / "runtime",
        "RESULTS_SINGLE_FILE_NAME": "preflight_compressed_results_update_run.xlsx",
        # --- disable LEAP imports / branch creation / scraping / caches ---
        "LEAP_IMPORT_SUPPLY_TO_LEAP": False,
        "LEAP_IMPORT_TRANSFORMATION_TO_LEAP": False,
        "LEAP_IMPORT_TRANSFERS_TO_LEAP": False,
        "LEAP_IMPORT_CREATE_BRANCHES": False,
        "LEAP_IMPORT_FILL_BRANCHES": False,
        "LEAP_IMPORT_INCLUDE_CURRENT_ACCOUNTS": False,
        "AGGREGATED_DEMAND_INCLUDE_IN_LEAP_IMPORT": False,
        "OTHER_LOSS_OWN_USE_INCLUDE_IN_LEAP_IMPORT": False,
        "ZERO_OTHER_DEMAND_INCLUDE_IN_LEAP_IMPORT": False,
        "RUN_LEAP_FUEL_BRANCH_PROBE_AT_START": False,
        "SCRAPE_LEAP_RESULTS": False,
        "TRANSFORMATION_SUPPLY_CACHE_ENABLED": False,
        "SKIP_ECONOMIES_WITH_EXISTING_EXPORTS": False,
        # --- results_update pass semantics ---
        "CAPACITY_UNMET_PASS_MODE": "results_update",
        "USE_AGGREGATED_DEMAND_AS_DUMMY": False,
        "OTHER_LOSS_OWN_USE_PROXY_STAGE": "second",
        "RUN_RESET_SUPPLY_AND_TRANSFORMATION_IMPORT_EXPORT": False,
        "RUN_ELECTRICITY_HEAT_INTERIM": False,
        # --- balance-demand routing to the temporary reduced workbooks ---
        "BALANCE_DEMAND_REF_WORKBOOK_PATH": reduced_ref_path,
        "BALANCE_DEMAND_TGT_WORKBOOK_PATH": reduced_tgt_path,
        "BALANCE_DEMAND_EXPORTS_ROOT": preflight_root / "reduced_balance_exports",
        "DIRECT_DEMAND_PROJECTION_ECONOMY": economy,
        "DIRECT_DEMAND_BASE_ECONOMY": economy.replace("_", ""),
        "DIRECT_DEMAND_BASE_YEAR": base_year,
        "DIRECT_DEMAND_PROJECTION_YEARS": (compressed_year,),
        "BALANCE_DEMAND_BASE_TABLE_PATH": source_files["esto_path"],
        "BALANCE_DEMAND_PROJECTION_TABLE_PATH": source_files["ninth_path"],
        "DIRECT_DEMAND_BASE_TABLE_PATH": source_files["esto_path"],
        "DIRECT_DEMAND_PROJECTION_TABLE_PATH": source_files["ninth_path"],
    }
    return _broadcast_config_overrides(overrides)


_BALANCE_DEMAND_ISSUE_SCHEMA_COLUMNS = [
    "reason",
    "details",
    "scenario",
    "year",
    "source_sheet",
    "leap_sector_name_full_path",
    "leap_flow",
    "leap_flow_name",
    "leap_product",
    "leap_product_name",
    "mapping_failed",
    "mapping_key_sector",
    "mapping_key_fuel",
    "mapping_candidate_rule",
    "esto_flow",
    "esto_product",
    "value_petajoule",
    "severity",
    "source_workbook",
    "economy",
    "issue_sector_key",
    "issue_fuel_key",
    "issue_fuel_is_non_actionable",
    "pair_is_demand",
    "sector_is_demand",
    "pair_scope_matched",
    "sector_scope_matched",
    "demand_relevant",
    "demand_relevance_basis",
]


def _finalize_balance_demand_issue_report(
    *,
    checks_dir: Path,
    run_error: Exception | None = None,
) -> dict[str, object]:
    """Guarantee a deterministic issue-report artifact and print a compact summary.

    The results_update runner only writes the balance-demand issue CSV when issues
    exist. This helper reads that CSV if present, otherwise writes a header-only CSV
    using the current schema (never leaving a stale report), then prints a compact
    summary of the demand-mapping outcome.
    """
    checks_dir = _resolve(checks_dir)
    report_path = checks_dir / RESULTS_BALANCE_DEMAND_ISSUES_FILENAME

    if report_path.exists() and report_path.stat().st_size > 0:
        try:
            issues = pd.read_csv(report_path)
        except pd.errors.EmptyDataError:
            issues = pd.DataFrame(columns=_BALANCE_DEMAND_ISSUE_SCHEMA_COLUMNS)
    else:
        issues = pd.DataFrame(columns=_BALANCE_DEMAND_ISSUE_SCHEMA_COLUMNS)
        checks_dir.mkdir(parents=True, exist_ok=True)
        issues.to_csv(report_path, index=False)
        print(
            f"[INFO] No balance-demand mapping issues: wrote header-only report to {report_path}."
        )

    def _bool_series(frame: pd.DataFrame, column: str) -> pd.Series:
        if column not in frame.columns:
            return pd.Series([False] * len(frame), index=frame.index)
        return frame[column].map(
            lambda v: str(v).strip().lower() in {"true", "1", "1.0", "yes"}
            if not isinstance(v, bool)
            else v
        ).fillna(False).astype(bool)

    total_rows = int(len(issues))
    demand_relevant_mask = _bool_series(issues, "demand_relevant")
    non_actionable_mask = _bool_series(issues, "issue_fuel_is_non_actionable")
    actionable_rows = int(demand_relevant_mask.sum())
    ignored_rows = int(total_rows - actionable_rows)

    reason_counts: dict[str, int] = {}
    if "reason" in issues.columns and total_rows:
        reason_counts = {
            str(reason): int(count)
            for reason, count in issues["reason"].fillna("(none)").value_counts().items()
        }

    unique_sector_fuel_keys = 0
    if {"issue_sector_key", "issue_fuel_key"}.issubset(issues.columns) and total_rows:
        unique_sector_fuel_keys = int(
            issues[["issue_sector_key", "issue_fuel_key"]]
            .fillna("")
            .astype(str)
            .drop_duplicates()
            .shape[0]
        )

    total_fuel_rows = 0
    if "issue_fuel_key" in issues.columns and total_rows:
        total_fuel_rows = int(
            issues["issue_fuel_key"].map(_normalize_balance_grid_label).eq("total").sum()
        )

    # Known-label-exception rescues and rollup-resolved rows are recorded as
    # resolution bases rather than as issues; surface them from any available
    # resolution/basis column so successful rescues stay visible.
    known_label_exception_rescues = 0
    rollup_resolved_rows = 0
    basis_text_cols = [
        col
        for col in ("demand_relevance_basis", "mapping_candidate_rule", "details", "reason")
        if col in issues.columns
    ]
    if basis_text_cols and total_rows:
        joined = issues[basis_text_cols].fillna("").astype(str).agg(" ".join, axis=1).str.lower()
        known_label_exception_rescues = int(
            joined.str.contains("known").__and__(joined.str.contains("label")).sum()
        )
        rollup_resolved_rows = int(joined.str.contains("rollup").sum())

    # Unresolved rows: demand-relevant, non-rescued mapping failures that remain.
    unresolved_rows = actionable_rows

    print("[INFO] Compressed results-update preflight balance-demand summary:")
    print(f"  - report: {report_path}")
    print(f"  - total issue rows:              {total_rows}")
    print(f"  - actionable demand issue rows:  {actionable_rows}")
    print(f"  - ignored / non-demand rows:     {ignored_rows}")
    print(f"  - non-actionable-fuel rows:      {int(non_actionable_mask.sum())}")
    print(f"  - unique sector/fuel keys:       {unique_sector_fuel_keys}")
    print(f"  - 'Total' fuel rows:             {total_fuel_rows}")
    print(f"  - known-label-exception rescues: {known_label_exception_rescues}")
    print(f"  - rollup-resolved rows:          {rollup_resolved_rows}")
    print(f"  - unresolved demand-relevant:    {unresolved_rows}")
    if reason_counts:
        counts_text = ", ".join(f"{reason}: {count}" for reason, count in reason_counts.items())
        print(f"  - counts by reason:              {counts_text}")
    if run_error is not None:
        print(f"  - NOTE: the workflow raised before completing: {run_error}")

    return {
        "report_path": report_path,
        "total_issue_rows": total_rows,
        "actionable_demand_issue_rows": actionable_rows,
        "ignored_non_demand_rows": ignored_rows,
        "unique_sector_fuel_keys": unique_sector_fuel_keys,
        "total_fuel_rows": total_fuel_rows,
        "known_label_exception_rescues": known_label_exception_rescues,
        "rollup_resolved_rows": rollup_resolved_rows,
        "unresolved_demand_relevant_rows": unresolved_rows,
        "reason_counts": reason_counts,
    }


def run_preflight_compressed_results_update(
    *,
    scenario_names: Iterable[str] | None = None,
) -> dict[str, object]:
    """Run a compressed two-year 20_USA results-update integration preflight.

    Builds temporary reduced REF/TGT LEAP balance workbooks (real 20_USA balance
    structure, two effective years) and runs the workflow in results_update mode
    with LEAP imports/scraping/caches disabled and every output isolated. State is
    snapshotted and restored in ``finally`` so a normal production run afterwards is
    unaffected -- even if the preflight fails.
    """
    economy = "20_USA"
    scenario_list = workflow_common.normalize_workflow_scenarios(scenario_names, SCENARIOS)
    # The balance-demand comparison only understands Reference/Target.
    scenario_list = [s for s in scenario_list if str(s).strip().lower() in {"reference", "target"}]
    if not scenario_list:
        scenario_list = ["Reference", "Target"]

    source_cfg = workflow_cfg.get_energy_source_config()
    base_year = int(source_cfg.esto_base_year)
    synthetic_year = base_year + 1

    preflight_root = _resolve(OUTPUT_DIR) / "preflight_compressed_results_update"
    runtime_dir = preflight_root / "runtime"
    source_dir = runtime_dir / "compressed_sources"
    reduced_dir = runtime_dir / "reduced_balance_workbooks"

    print(
        "[INFO] Starting preflight_compressed_results_update: "
        f"economies=['{economy}'], scenarios={scenario_list}, "
        f"years={base_year}-{synthetic_year} (real reduced LEAP balance structure)."
    )

    # Build temporary reduced REF/TGT workbooks from the production source workbooks
    # (read-only). The REF synthetic sheet is built only from REF; TGT only from TGT.
    reduced_ref = _build_reduced_preflight_balance_workbook(
        source_path=BALANCE_DEMAND_REF_WORKBOOK_PATH,
        output_path=reduced_dir / f"reduced_{economy}_REF_EBal_{base_year}_{synthetic_year}.xlsx",
        base_year=base_year,
        synthetic_year=synthetic_year,
        scenario_code="REF",
        abs_diagnostic_path=reduced_dir / f"reduced_{economy}_REF_abs_sum_{synthetic_year}.csv",
    )
    reduced_tgt = _build_reduced_preflight_balance_workbook(
        source_path=BALANCE_DEMAND_TGT_WORKBOOK_PATH,
        output_path=reduced_dir / f"reduced_{economy}_TGT_EBal_{base_year}_{synthetic_year}.xlsx",
        base_year=base_year,
        synthetic_year=synthetic_year,
        scenario_code="TGT",
        abs_diagnostic_path=reduced_dir / f"reduced_{economy}_TGT_abs_sum_{synthetic_year}.csv",
    )

    # Scenario-separated compressed ESTO/9th sources for 20_USA (Reference stays
    # Reference, Target stays Target -- never summed together and replicated).
    source_files = _create_preflight_compressed_source_files(
        output_dir=source_dir,
        scenario_names=scenario_list,
        preserve_source_scenarios=True,
        economy_filter=[economy],
        file_prefix="preflight_results_update",
    )
    print(
        "[INFO] Compressed scenario-separated 9th source written to "
        f"{source_files['ninth_path']} "
        f"(abs-sum diagnostics: {source_files['ninth_abs_diagnostics_path']})."
    )

    checks_dir = preflight_root / "checks"

    # Late import (avoids the supply_reconciliation_workflow <-> supply_preflight cycle).
    from codebase.supply_reconciliation_workflow import (
        run_results_linked_transformation_supply_workflow,
    )

    state = _snapshot_preflight_state()
    broadcast_snapshot: list[tuple[str, str, object]] | None = None
    try:
        broadcast_snapshot = _apply_preflight_results_update_state(
            source_files=source_files,
            reduced_ref_path=reduced_ref["workbook_path"],
            reduced_tgt_path=reduced_tgt["workbook_path"],
            preflight_root=preflight_root,
            scenarios=scenario_list,
            economy=economy,
        )
        run_error: Exception | None = None
        try:
            result = run_results_linked_transformation_supply_workflow(
                economies=[economy],
                scenario_names=scenario_list,
                export_dataset_key=EXPORT_DATASET_KEY,
                include_leap_import=False,
                import_scenarios=[],
                scrape_leap_results=False,
            )
        except Exception as exc:  # noqa: BLE001 — still finalize the deterministic report
            run_error = exc
            _finalize_balance_demand_issue_report(checks_dir=checks_dir, run_error=exc)
            raise
        report_summary = _finalize_balance_demand_issue_report(checks_dir=checks_dir)
        result["preflight_results_update_report_summary"] = report_summary
        result["preflight_results_update_reduced_ref"] = reduced_ref
        result["preflight_results_update_reduced_tgt"] = reduced_tgt
        result["preflight_results_update_compressed_ninth_path"] = source_files["ninth_path"]
        result["preflight_results_update_compressed_abs_diagnostics_path"] = source_files[
            "ninth_abs_diagnostics_path"
        ]
        result["preflight_results_update_base_year"] = base_year
        result["preflight_results_update_synthetic_year"] = synthetic_year
        result["preflight_results_update_scenarios"] = scenario_list
        print("[INFO] preflight_compressed_results_update completed.")
        return result
    finally:
        if broadcast_snapshot is not None:
            _restore_config_overrides(broadcast_snapshot)
        _restore_preflight_state(state)


