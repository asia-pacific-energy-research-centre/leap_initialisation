"""
Build limited-year LEAP balance comparisons with ESTO and the 9th Outlook.

This first diagnostic stage is read-only. It reports LEAP minus source values
on the shared ESTO flow/product axis and flags comparisons that would need an
allocation rule before a later update workflow could act on them.
"""

from __future__ import annotations

import json
import re
import tempfile
from contextlib import contextmanager
from pathlib import Path
from typing import Any, Iterable, Sequence

import pandas as pd

from codebase.configuration import workflow_config as workflow_cfg
from codebase.mappings.canonical_mapping import ConfigTableRef
from codebase.utilities.leap_balance_export_resolver import (
    require_level2_balance_export_detail,
    resolve_balance_export_workbook,
)
from codebase.functions.transformation_analysis_utils import MAJOR_SECTOR_CONFIG


REPO_ROOT = Path(__file__).resolve().parents[2]


def _find_workspace_sibling_repo(repo_name: str) -> Path:
    """Find a workspace sibling from either a normal checkout or nested worktree."""
    candidates = [REPO_ROOT.parent / repo_name]
    for ancestor in REPO_ROOT.parents:
        if ancestor.name == "leap_initialisation":
            candidates.append(ancestor.parent / repo_name)
    for candidate in candidates:
        if candidate.exists():
            return candidate
    return candidates[0]


LEAP_MAPPINGS_REPO_ROOT = _find_workspace_sibling_repo("leap_mappings")
OUTLOOK_MAPPINGS_MASTER_PATH = (
    LEAP_MAPPINGS_REPO_ROOT / "config" / "outlook_mappings_master.xlsx"
)
DEFAULT_EXPORTS_ROOT = REPO_ROOT / "data" / "leap balances exports"
DEFAULT_OUTPUT_DIR = (
    REPO_ROOT
    / "outputs"
    / "leap_exports"
    / "supply_reconciliation"
    / "supporting_files"
    / "baseline_seed_balance_diagnostics"
)
DEFAULT_KNOWN_ISSUES_PATH = REPO_ROOT / "config" / "leap_results_balance_known_issues.json"
DEFAULT_TEMPLATE_SHEET = "EBal|2060"
DEFAULT_BASE_YEAR = 2022
DEFAULT_TOLERANCE_PJ = 1e-6
DEFAULT_MAPPING_PAIRS_PATH: ConfigTableRef = (
    OUTLOOK_MAPPINGS_MASTER_PATH,
    "ninth_pairs_to_esto_pairs",
)
DEFAULT_CODEBOOK_PATH = OUTLOOK_MAPPINGS_MASTER_PATH
DEFAULT_SHEET_MAP_PATH = REPO_ROOT / "config" / "runtime_tables" / "leap_results_sheet_map.csv"
DEFAULT_BACKUP_MAPPINGS_PATH = REPO_ROOT / "config" / "backup_leap_mappings.xlsx"
DEFAULT_EXPLICIT_MAPPINGS_PATH = REPO_ROOT / "config" / "leap_results_explicit_mappings.csv"
DEFAULT_EXPLICIT_REASSIGNMENTS_PATH = (
    REPO_ROOT / "config" / "runtime_tables" / "leap_explicit_reassignments.csv"
)
DEFAULT_SYNTHETIC_REFERENCE_ROWS_PATH = (
    REPO_ROOT / "config" / "runtime_tables" / "synthetic_reference_rows.csv"
)
DEFAULT_BALANCE_VARIABLE_RULES_PATH = (
    REPO_ROOT / "config" / "runtime_tables" / "balance_error_signal_rules.csv"
)
DEFAULT_BASE_TABLE_PATH = workflow_cfg.get_energy_source_config().esto_base_table_path
DEFAULT_PROJECTION_TABLE_PATH = REPO_ROOT / "data" / "merged_file_energy_ALL_20251106.csv"

DIFFERENCE_OUTPUT_COLUMNS = [
    "economy",
    "scenario",
    "year",
    "esto_flow",
    "esto_product",
    "leap_sector_names",
    "leap_fuel_names",
    "ninth_sector_codes",
    "ninth_fuel_codes",
    "reference_source",
    "leap_value_pj",
    "source_value_pj",
    "difference_pj",
    "absolute_difference_pj",
    "difference_percent",
    "correction_to_match_source_pj",
    "status",
    "is_mismatch",
    "comparison_grain",
    "leap_component_count",
    "ninth_pair_count",
    "ninth_pair_max_esto_claimants",
    "projection_allocation_complete",
    "projection_target_pair_count",
    "projection_matched_pair_count",
    "projection_allocation_methods",
    "projection_share_sources",
    "update_allocation_required",
    "update_allocation_reason",
    "sheet",
    "measure",
    "fuel_label",
]

REVIEW_ADDED_COLUMNS = [
    "leap_balance_row",
    "leap_balance_fuel",
    "material_for_review",
    "balance_variable_role",
    "allowed_to_change",
    "error_signal_name",
    "balance_variable_rule_reason",
    "balance_contract_issue",
    "requires_issue_review",
    "update_signal_eligible",
    "placeholder_scope",
    "placeholder_scope_reason",
    "preliminary_owner",
    "primary_classification",
    "evidence_note",
    "next_action",
]

BALANCE_VARIABLE_RULE_COLUMNS = [
    "economy",
    "scenario",
    "esto_product",
    "esto_flow",
    "balance_variable_role",
    "error_signal_name",
    "reason",
    "enabled",
]

PLACEHOLDER_SECTOR_PATTERN = re.compile(
    r"(?:^|\|)(?:Electricity interim/|CHP interim/|Heat plant interim/|"
    r"All demand aggregated(?:/|$))",
    flags=re.IGNORECASE,
)

def _resolve(path: Path | str) -> Path:
    """Resolve a notebook-friendly path against this repository."""
    normalized = str(path).replace("\\", "/")
    candidate = Path(normalized)
    return candidate if candidate.is_absolute() else REPO_ROOT / candidate


def _resolve_config_table_ref(
    value: ConfigTableRef,
) -> ConfigTableRef:
    """Resolve a path or ``(workbook, sheet)`` mapping-table reference."""
    if isinstance(value, tuple):
        return (_resolve(value[0]), str(value[1]))
    return _resolve(value)


def _clean_token(value: object) -> str:
    if value is None or value is pd.NA:
        return ""
    return str(value).strip()


def load_balance_variable_rules(
    path: Path | str = DEFAULT_BALANCE_VARIABLE_RULES_PATH,
) -> pd.DataFrame:
    """Load the explicit balance-variable/error-signal contract."""
    rules = pd.read_csv(_resolve(path))
    missing = sorted(set(BALANCE_VARIABLE_RULE_COLUMNS) - set(rules.columns))
    if missing:
        raise KeyError(f"Balance-variable rules are missing columns: {missing}")
    enabled = rules["enabled"].fillna(False).astype(str).str.lower().isin(
        {"true", "1", "yes"}
    )
    return rules[enabled].reset_index(drop=True)


def classify_balance_variable(
    *,
    economy: object,
    scenario: object,
    esto_product: object,
    esto_flow: object,
    rules: pd.DataFrame,
) -> dict[str, str]:
    """Resolve the most-specific rule, defaulting unlisted flows to protected."""
    values = {
        "economy": _clean_token(economy),
        "scenario": _clean_token(scenario).lower(),
        "esto_product": _clean_token(esto_product),
        "esto_flow": _clean_token(esto_flow),
    }
    matches: list[tuple[int, int, pd.Series]] = []
    for index, rule in rules.iterrows():
        specificity = 0
        matched = True
        for column, value in values.items():
            rule_value = _clean_token(rule.get(column))
            if column == "scenario":
                rule_value = rule_value.lower()
            if rule_value == "*":
                continue
            if rule_value != value:
                matched = False
                break
            specificity += 1
        if matched:
            matches.append((specificity, int(index), rule))
    if not matches:
        return {
            "balance_variable_role": "protected",
            "error_signal_name": "",
            "balance_variable_rule_reason": (
                "No allowed-change rule applies; differences are protected by default."
            ),
        }
    _, _, selected = sorted(matches, key=lambda item: (-item[0], item[1]))[0]
    return {
        "balance_variable_role": _clean_token(
            selected.get("balance_variable_role")
        ),
        "error_signal_name": _clean_token(selected.get("error_signal_name")),
        "balance_variable_rule_reason": _clean_token(selected.get("reason")),
    }


def _split_pipe_tokens(value: object) -> list[str]:
    return [token.strip() for token in _clean_token(value).split("|") if token.strip()]


def _unique_pipe(values: Iterable[object]) -> str:
    tokens: set[str] = set()
    for value in values:
        tokens.update(_split_pipe_tokens(value))
    return "|".join(sorted(tokens))


def _iter_paired_tokens(left: object, right: object) -> list[tuple[str, str]]:
    """Expand paired mapping tokens without hiding non-one-to-one cardinality."""
    left_tokens = _split_pipe_tokens(left)
    right_tokens = _split_pipe_tokens(right)
    if not left_tokens or not right_tokens:
        return []
    if len(left_tokens) == len(right_tokens):
        return list(dict.fromkeys(zip(left_tokens, right_tokens)))
    if len(left_tokens) == 1:
        return [(left_tokens[0], token) for token in right_tokens]
    if len(right_tokens) == 1:
        return [(token, right_tokens[0]) for token in left_tokens]
    return [(left_token, right_token) for left_token in left_tokens for right_token in right_tokens]


def _count_mapping_pairs(frame: pd.DataFrame, left_col: str, right_col: str) -> int:
    pairs: set[tuple[str, str]] = set()
    if frame.empty:
        return 0
    for row in frame[[left_col, right_col]].itertuples(index=False, name=None):
        pairs.update(_iter_paired_tokens(row[0], row[1]))
    return len(pairs)


def _compact_economy_code(economy: str) -> str:
    return _clean_token(economy).replace("_", "")


def _normalize_scenarios(scenarios: Sequence[str]) -> list[str]:
    normalized: list[str] = []
    for scenario in scenarios:
        label = _clean_token(scenario)
        if label.lower() not in {"reference", "target"}:
            raise ValueError(
                f"Unsupported scenario {scenario!r}. Step 1 accepts Reference and Target."
            )
        canonical = label.title()
        if canonical not in normalized:
            normalized.append(canonical)
    if not normalized:
        raise ValueError("At least one of Reference or Target is required.")
    return normalized


def _validate_years(years: Sequence[int], *, base_year: int) -> list[int]:
    normalized = sorted({int(year) for year in years})
    if not normalized:
        raise ValueError("At least one diagnostic year is required.")
    historical = [year for year in normalized if year < int(base_year)]
    if historical:
        raise ValueError(
            "Step 1 currently supports the ESTO base year and later 9th Outlook "
            f"projection years. Pre-base historical years are not yet supported: {historical}."
        )
    return normalized


def _read_direct_workbook_scope(
    workbook_path: Path | str,
    *,
    base_year: int,
) -> tuple[Path, list[int], list[str]]:
    """Read the diagnostic year/scenario from one explicit balance workbook."""
    path = _resolve(workbook_path)
    if not path.exists():
        raise FileNotFoundError(f"Direct LEAP balance workbook does not exist: {path}")

    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    metadata_rows: list[tuple[str, int, str]] = []
    try:
        for sheet in workbook.worksheets:
            metadata = str(sheet.cell(2, 1).value or "").strip()
            scenario_match = re.search(
                r"Scenario:\s*([^,]+)",
                metadata,
                flags=re.IGNORECASE,
            )
            year_match = re.search(r"Year:\s*(\d{4})", metadata, flags=re.IGNORECASE)
            units_match = re.search(r"Units:\s*(.+)$", metadata, flags=re.IGNORECASE)
            if not scenario_match or not year_match or not units_match:
                raise ValueError(
                    "Direct LEAP balance workbook metadata must declare Scenario, "
                    f"Year, and Units on every sheet: {path} [{sheet.title}]"
                )
            metadata_rows.append(
                (
                    scenario_match.group(1).strip(),
                    int(year_match.group(1)),
                    units_match.group(1).strip().rstrip("."),
                )
            )
    finally:
        workbook.close()

    units = sorted({unit.lower() for _, _, unit in metadata_rows})
    supported_units = {"petajoule", "thousand petajoule"}
    unsupported_units = sorted(set(units) - supported_units)
    if unsupported_units:
        raise ValueError(
            "Direct LEAP balance diagnostics currently require Petajoule or "
            "Thousand Petajoule workbook metadata; "
            f"found unsupported units {unsupported_units} in {path}."
        )
    scenarios = _normalize_scenarios([scenario for scenario, _, _ in metadata_rows])
    years = _validate_years([year for _, year, _ in metadata_rows], base_year=base_year)
    return path, years, scenarios


def _load_optional_json(path: Path | str | None) -> dict[str, Any]:
    if path is None:
        return {}
    resolved = _resolve(path)
    if not resolved.exists():
        return {}
    try:
        payload = json.loads(resolved.read_text(encoding="utf-8-sig"))
    except Exception as exc:
        raise ValueError(f"Could not read JSON configuration {resolved}: {exc}") from exc
    if not isinstance(payload, dict):
        raise ValueError(f"Expected a JSON object in {resolved}.")
    return payload


@contextmanager
def _temporary_balance_runtime_paths(
    *,
    codebook_path: Path,
    sheet_map_path: Path,
    exports_root: Path,
):
    """Route shared loader globals for a worktree run and restore them."""
    from codebase.utilities import master_config

    master_snapshot = {
        "OUTLOOK_MAPPINGS_MASTER_PATH": master_config.OUTLOOK_MAPPINGS_MASTER_PATH,
        "RUNTIME_TABLE_DIR": master_config.RUNTIME_TABLE_DIR,
    }
    resolver_defaults = dict(resolve_balance_export_workbook.__kwdefaults__ or {})
    master_config.OUTLOOK_MAPPINGS_MASTER_PATH = codebook_path
    master_config.RUNTIME_TABLE_DIR = sheet_map_path.parent
    if resolve_balance_export_workbook.__kwdefaults__ is not None:
        resolve_balance_export_workbook.__kwdefaults__["exports_root"] = exports_root

    from codebase.utilities.leap_results_dashboard_balance import (
        build_balance_comparison_esto_axis,
        convert_leap_balances_to_esto_long_table,
    )
    from codebase.utilities import energy_balance_template_extractor

    extractor_snapshot = {
        "LEAP_MAPPINGS_REPO_ROOT": energy_balance_template_extractor.LEAP_MAPPINGS_REPO_ROOT,
        "SUBTOTAL_MISMATCH_EXCEPTIONS_PATH": (
            energy_balance_template_extractor.SUBTOTAL_MISMATCH_EXCEPTIONS_PATH
        ),
        "_SUBTOTAL_MISMATCH_EXCEPTION_SETS": (
            energy_balance_template_extractor._SUBTOTAL_MISMATCH_EXCEPTION_SETS
        ),
        "_DUPLICATE_MAPPING_EXCEPTION_SETS": (
            energy_balance_template_extractor._DUPLICATE_MAPPING_EXCEPTION_SETS
        ),
    }
    energy_balance_template_extractor.LEAP_MAPPINGS_REPO_ROOT = LEAP_MAPPINGS_REPO_ROOT
    energy_balance_template_extractor.SUBTOTAL_MISMATCH_EXCEPTIONS_PATH = (
        LEAP_MAPPINGS_REPO_ROOT / "config" / "mapping_issue_exception_sets.xlsx"
    )
    energy_balance_template_extractor._SUBTOTAL_MISMATCH_EXCEPTION_SETS = None
    energy_balance_template_extractor._DUPLICATE_MAPPING_EXCEPTION_SETS = None
    try:
        yield build_balance_comparison_esto_axis, convert_leap_balances_to_esto_long_table
    finally:
        master_config.OUTLOOK_MAPPINGS_MASTER_PATH = master_snapshot[
            "OUTLOOK_MAPPINGS_MASTER_PATH"
        ]
        master_config.RUNTIME_TABLE_DIR = master_snapshot["RUNTIME_TABLE_DIR"]
        if resolve_balance_export_workbook.__kwdefaults__ is not None:
            resolve_balance_export_workbook.__kwdefaults__.clear()
            resolve_balance_export_workbook.__kwdefaults__.update(resolver_defaults)
        for name, value in extractor_snapshot.items():
            setattr(energy_balance_template_extractor, name, value)


def _scope_rows_to_diagnostic_window(
    frame: pd.DataFrame,
    *,
    years: Sequence[int],
    scenarios: Sequence[str],
) -> pd.DataFrame:
    """Limit supporting diagnostics to the same small review window."""
    if frame is None or frame.empty:
        return pd.DataFrame() if frame is None else frame.copy()
    scoped = frame.copy()
    if "year" in scoped.columns:
        numeric_year = pd.to_numeric(scoped["year"], errors="coerce")
        scoped = scoped[numeric_year.isin({int(year) for year in years})].copy()
    if "scenario" in scoped.columns:
        wanted = {scenario.lower() for scenario in _normalize_scenarios(scenarios)}
        scoped = scoped[
            scoped["scenario"].fillna("").astype(str).str.strip().str.lower().isin(wanted)
        ].copy()
    return scoped.reset_index(drop=True)


def _preliminary_owner(esto_flow: object) -> str:
    """Return the likely workflow owner from the ESTO flow family."""
    flow = _clean_token(esto_flow)
    if flow.startswith(("01 ", "02 ", "03 ", "04 ", "05 ", "06 ", "07 ")):
        return "supply"
    if flow.startswith("08"):
        return "transfers"
    if flow.startswith("09"):
        return "transformation"
    if flow.startswith("10"):
        return "other_loss_and_own_use"
    if flow.startswith(("12 ", "13 ", "14 ", "15 ", "16 ", "17 ")):
        return "aggregated_demand"
    return "mapping_or_diagnostic"


def build_balance_review_table(
    differences: pd.DataFrame,
    *,
    material_threshold_pj: float = 1.0,
    balance_variable_rules: pd.DataFrame | None = None,
) -> pd.DataFrame:
    """Classify differences using the explicit allowed balance-variable contract."""
    review = differences.copy()
    for column in DIFFERENCE_OUTPUT_COLUMNS:
        if column not in review.columns:
            review[column] = pd.NA
    review["leap_balance_row"] = review["leap_sector_names"].fillna("")
    review["leap_balance_fuel"] = review["leap_fuel_names"].fillna("")
    review["material_for_review"] = (
        pd.to_numeric(review["absolute_difference_pj"], errors="coerce")
        .fillna(0.0)
        .ge(float(material_threshold_pj))
    )
    rules = (
        load_balance_variable_rules()
        if balance_variable_rules is None
        else balance_variable_rules.copy()
    )
    variable_classifications = [
        classify_balance_variable(
            economy=row.get("economy"),
            scenario=row.get("scenario"),
            esto_product=row.get("esto_product"),
            esto_flow=row.get("esto_flow"),
            rules=rules,
        )
        for row in review.to_dict("records")
    ]
    review["balance_variable_role"] = [
        row["balance_variable_role"] for row in variable_classifications
    ]
    review["allowed_to_change"] = review["balance_variable_role"].eq("error_signal")
    review["error_signal_name"] = [
        row["error_signal_name"] for row in variable_classifications
    ]
    review["balance_variable_rule_reason"] = [
        row["balance_variable_rule_reason"] for row in variable_classifications
    ]
    review["balance_contract_issue"] = ""
    review["requires_issue_review"] = False
    review["update_signal_eligible"] = False
    review["placeholder_scope"] = (
        review["leap_sector_names"]
        .fillna("")
        .astype(str)
        .str.contains(PLACEHOLDER_SECTOR_PATTERN, regex=True)
    )
    review["placeholder_scope_reason"] = ""
    review.loc[review["placeholder_scope"], "placeholder_scope_reason"] = (
        "Interim/placeholder activity may need a combined placeholder-and-replacement "
        "comparison boundary; it is not automatically excluded."
    )
    review["preliminary_owner"] = review["esto_flow"].map(_preliminary_owner)
    review["primary_classification"] = ""
    review["evidence_note"] = ""
    review["next_action"] = ""

    discrepancy_mask = review["status"].ne("match")
    unavailable = discrepancy_mask & review["status"].isin(
        {"reference_unavailable", "missing_in_reference", "missing_in_leap"}
    )
    allocation_required = discrepancy_mask & review[
        "update_allocation_required"
    ].fillna(False).astype(bool)
    error_signal_difference = (
        discrepancy_mask
        & ~unavailable
        & ~allocation_required
        & review["balance_variable_role"].eq("error_signal")
    )
    derived_difference = (
        discrepancy_mask
        & ~unavailable
        & ~allocation_required
        & review["balance_variable_role"].eq("derived_check")
    )
    protected_difference = (
        discrepancy_mask
        & ~unavailable
        & ~allocation_required
        & review["balance_variable_role"].eq("protected")
    )

    review.loc[unavailable, "balance_contract_issue"] = (
        "comparison_unavailable_or_mapping_issue"
    )
    review.loc[allocation_required, "balance_contract_issue"] = (
        "mapping_allocation_rule_required"
    )
    review.loc[error_signal_difference, "balance_contract_issue"] = (
        "expected_error_signal_difference"
    )
    review.loc[derived_difference, "balance_contract_issue"] = (
        "derived_balance_difference"
    )
    review.loc[protected_difference, "balance_contract_issue"] = (
        "protected_flow_difference"
    )
    review.loc[
        unavailable | allocation_required | derived_difference | protected_difference,
        "requires_issue_review",
    ] = True
    review.loc[error_signal_difference, "update_signal_eligible"] = True

    review.loc[discrepancy_mask, "primary_classification"] = "unresolved"
    review.loc[error_signal_difference, "primary_classification"] = (
        "expected_error_signal"
    )
    review.loc[derived_difference, "primary_classification"] = (
        "derived_balance_difference"
    )
    review.loc[protected_difference, "primary_classification"] = (
        "protected_flow_difference"
    )
    review.loc[discrepancy_mask, "evidence_note"] = (
        "Classified against the allowed balance-variable contract."
    )
    review.loc[error_signal_difference, "evidence_note"] = (
        "Imports are the configured balancing error signal for this fuel."
    )
    review.loc[error_signal_difference, "next_action"] = (
        "Use this difference as updater input, subject to allocator and mapping checks."
    )
    review.loc[
        protected_difference | derived_difference,
        "evidence_note",
    ] = (
        "This variable is not configured as an allowed balancing signal. Investigate "
        "the baseline seed, LEAP balancing rules, or the variable contract."
    )
    review.loc[
        protected_difference | derived_difference,
        "next_action",
    ] = (
        "Raise an issue; do not convert this row directly into a numeric update."
    )
    review.loc[unavailable | allocation_required, "next_action"] = (
        "Resolve comparison coverage or allocation grain before using this row."
    )

    total_final_boundary = (
        discrepancy_mask
        & review["esto_flow"].fillna("").eq("13 Total final energy consumption")
        & review["leap_sector_names"]
        .fillna("")
        .str.lower()
        .eq("total final energy consumption")
    )
    review.loc[total_final_boundary, "primary_classification"] = "diagnostic_bug"
    review.loc[total_final_boundary, "balance_contract_issue"] = (
        "diagnostic_comparison_boundary_bug"
    )
    review.loc[total_final_boundary, "requires_issue_review"] = True
    review.loc[total_final_boundary, "update_signal_eligible"] = False
    review.loc[total_final_boundary, "preliminary_owner"] = "mapping_or_diagnostic"
    review.loc[total_final_boundary, "evidence_note"] = (
        "LEAP Total Final Energy Demand includes the separate Other loss and own "
        "use demand branch, so it is not a direct ESTO flow-13 comparator."
    )
    review.loc[total_final_boundary, "next_action"] = (
        "Define a rollup-aware comparison boundary; do not update seed rows from "
        "this aggregate difference."
    )

    missing_rollup = (
        review["status"].eq("reference_unavailable")
        & review["esto_flow"].fillna("").str.contains(
            r"\(including own use\)",
            regex=True,
        )
    )
    review.loc[
        missing_rollup,
        "primary_classification",
    ] = "mapping_grain_or_allocation_required"
    review.loc[missing_rollup, "preliminary_owner"] = "mapping_or_diagnostic"
    review.loc[missing_rollup, "evidence_note"] = (
        "The named ESTO comparison row is a synthetic own-use boundary rollup, "
        "not a literal row in the raw ESTO table."
    )
    review.loc[missing_rollup, "next_action"] = (
        "Apply the maintained rollup components before numerical comparison."
    )

    return review[[*DIFFERENCE_OUTPUT_COLUMNS, *REVIEW_ADDED_COLUMNS]].reset_index(
        drop=True
    )


def build_balance_diagnostic_counts(
    differences: pd.DataFrame,
    mapping_issues: pd.DataFrame,
) -> dict[str, int]:
    """Return the separate review counts required before prioritisation."""
    statuses = differences.get("status", pd.Series(dtype=str)).fillna("").astype(str)
    issue_reasons = mapping_issues.get("reason", pd.Series(dtype=str)).fillna("").astype(str)
    issue_severity = (
        mapping_issues.get("severity", pd.Series(dtype=str)).fillna("").astype(str)
    )
    direct_mask = differences.get(
        "comparison_grain",
        pd.Series("", index=differences.index, dtype=str),
    ).eq("direct_leap_to_esto_pair")
    unsafe_mask = differences.get(
        "update_allocation_required",
        pd.Series(False, index=differences.index, dtype=bool),
    ).fillna(False).astype(bool)
    return {
        "value_mismatches": int(statuses.eq("value_mismatch").sum()),
        "rows_missing_from_leap": int(statuses.eq("missing_in_leap").sum()),
        "rows_missing_from_comparator": int(
            statuses.isin(["reference_unavailable", "missing_in_source"]).sum()
        ),
        "unmapped_rows": int(issue_reasons.eq("missing_esto_pair").sum()),
        "total_balance_check_failures": int(
            (
                issue_reasons.eq("total_balance_mapping_check")
                & issue_severity.eq("error")
            ).sum()
        ),
        "direct_one_to_one_comparisons": int(direct_mask.sum()),
        "aggregate_or_shared_unsafe_comparisons": int(unsafe_mask.sum()),
    }


def _write_esto_axis_extraction_mapping_workbook(
    *,
    output_path: Path,
    codebook_path: Path,
) -> Path:
    """Write the strict LEAP->ESTO subset needed by the Step 1 extractor.

    The ESTO-axis diagnostic does not consume ``leap_combined_ninth`` during
    extraction. Future-year source values are joined later through the separate
    canonical ninth_pairs_to_esto_pairs bridge. Keeping a header-only ninth
    sheet prevents unrelated LEAP->9th validation findings from blocking this
    read-only ESTO-axis report without marking those findings as accepted.
    """
    from codebase.utilities.master_config import read_config_table

    esto_mapping = read_config_table(
        codebook_path,
        sheet_name="leap_combined_esto",
        dtype=str,
    )
    ninth_columns = read_config_table(
        codebook_path,
        sheet_name="leap_combined_ninth",
        dtype=str,
    ).columns
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_path) as writer:
        esto_mapping.to_excel(writer, sheet_name="leap_combined_esto", index=False)
        pd.DataFrame(columns=ninth_columns).to_excel(
            writer,
            sheet_name="leap_combined_ninth",
            index=False,
        )
    return output_path


def _build_mapping_metadata(
    mapping_status: pd.DataFrame,
    leap_long: pd.DataFrame | None,
) -> pd.DataFrame:
    """Return one mapping/cardinality record per displayed comparison row."""
    key_columns = ["sheet", "measure", "fuel_label"]
    metadata_columns = [
        *key_columns,
        "esto_flow",
        "esto_product",
        "ninth_sector_codes",
        "ninth_fuel_codes",
        "ninth_pair_count",
        "ninth_pair_max_esto_claimants",
        "leap_sector_names",
        "leap_fuel_names",
        "leap_component_count",
    ]
    if mapping_status is None or mapping_status.empty:
        return pd.DataFrame(columns=metadata_columns)

    status = mapping_status.copy()
    for column in [
        *key_columns,
        "esto_flow",
        "esto_product",
        "sector_code_9th",
        "ninth_fuel_code",
    ]:
        if column not in status.columns:
            status[column] = ""
        status[column] = status[column].fillna("").astype(str).str.strip()

    ninth_pair_claimants: dict[tuple[str, str], set[tuple[str, str]]] = {}
    for row in status[
        ["esto_flow", "esto_product", "sector_code_9th", "ninth_fuel_code"]
    ].itertuples(index=False, name=None):
        esto_pairs = _iter_paired_tokens(row[0], row[1])
        ninth_pairs = _iter_paired_tokens(row[2], row[3])
        for ninth_pair in ninth_pairs:
            ninth_pair_claimants.setdefault(ninth_pair, set()).update(esto_pairs)

    records: list[dict[str, Any]] = []
    for key, group in status.groupby(key_columns, dropna=False, sort=False):
        group_ninth_pairs: set[tuple[str, str]] = set()
        for row in group[["sector_code_9th", "ninth_fuel_code"]].itertuples(
            index=False,
            name=None,
        ):
            group_ninth_pairs.update(_iter_paired_tokens(row[0], row[1]))
        records.append(
            {
                "sheet": key[0],
                "measure": key[1],
                "fuel_label": key[2],
                "esto_flow": _unique_pipe(group["esto_flow"]),
                "esto_product": _unique_pipe(group["esto_product"]),
                "ninth_sector_codes": _unique_pipe(group["sector_code_9th"]),
                "ninth_fuel_codes": _unique_pipe(group["ninth_fuel_code"]),
                "ninth_pair_count": _count_mapping_pairs(
                    group,
                    "sector_code_9th",
                    "ninth_fuel_code",
                ),
                "ninth_pair_max_esto_claimants": max(
                    (
                        len(ninth_pair_claimants.get(pair, set()))
                        for pair in group_ninth_pairs
                    ),
                    default=0,
                ),
            }
        )
    metadata = pd.DataFrame(records)

    if leap_long is None or leap_long.empty:
        metadata["leap_sector_names"] = ""
        metadata["leap_fuel_names"] = ""
        metadata["leap_component_count"] = 0
        return metadata[metadata_columns]

    leap = leap_long.copy()
    if "sheet" not in leap.columns and "sheet_name" in leap.columns:
        leap["sheet"] = leap["sheet_name"]
    for column in [*key_columns, "leap_sector_name", "leap_fuel_name", "leap_sector", "leap_fuel"]:
        if column not in leap.columns:
            leap[column] = ""
        leap[column] = leap[column].fillna("").astype(str).str.strip()

    component_records: list[dict[str, Any]] = []
    for key, group in leap.groupby(key_columns, dropna=False, sort=False):
        sector_column = "leap_sector_name" if group["leap_sector_name"].ne("").any() else "leap_sector"
        fuel_column = "leap_fuel_name" if group["leap_fuel_name"].ne("").any() else "leap_fuel"
        component_records.append(
            {
                "sheet": key[0],
                "measure": key[1],
                "fuel_label": key[2],
                "leap_sector_names": _unique_pipe(group[sector_column]),
                "leap_fuel_names": _unique_pipe(group[fuel_column]),
                "leap_component_count": _count_mapping_pairs(
                    group,
                    sector_column,
                    fuel_column,
                ),
            }
        )
    components = pd.DataFrame(component_records)
    metadata = metadata.merge(components, on=key_columns, how="left")
    metadata["leap_sector_names"] = metadata["leap_sector_names"].fillna("")
    metadata["leap_fuel_names"] = metadata["leap_fuel_names"].fillna("")
    metadata["leap_component_count"] = (
        pd.to_numeric(metadata["leap_component_count"], errors="coerce").fillna(0).astype(int)
    )
    return metadata[metadata_columns]


def _projection_pair_records(mapping_status: pd.DataFrame) -> pd.DataFrame:
    """Expand displayed comparison keys to distinct ESTO target pairs."""
    columns = ["sheet", "measure", "fuel_label", "esto_flow", "esto_product"]
    if mapping_status is None or mapping_status.empty:
        return pd.DataFrame(columns=columns)
    status = mapping_status.copy()
    if "sheet" not in status.columns and "sheet_name" in status.columns:
        status["sheet"] = status["sheet_name"]
    for column in columns:
        if column not in status.columns:
            status[column] = ""
        status[column] = status[column].fillna("").astype(str).str.strip()

    records: list[dict[str, str]] = []
    for row in status[columns].itertuples(index=False, name=None):
        for esto_flow, esto_product in _iter_paired_tokens(row[3], row[4]):
            records.append(
                {
                    "sheet": row[0],
                    "measure": row[1],
                    "fuel_label": row[2],
                    "esto_flow": esto_flow,
                    "esto_product": esto_product,
                }
            )
    if not records:
        return pd.DataFrame(columns=columns)
    return pd.DataFrame(records).drop_duplicates().reset_index(drop=True)


def apply_canonical_projection_comparators(
    *,
    comparison_long: pd.DataFrame,
    mapping_status: pd.DataFrame,
    projection_tables: pd.DataFrame,
    allocation_provenance: pd.DataFrame,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Replace raw 9th projection claims with allocated ESTO-pair values.

    ``projection_tables`` contains one row per scenario/economy/ESTO pair and
    year columns. ``allocation_provenance`` is optional review detail from the
    same canonical allocation. Only complete displayed comparison groups are
    replaced; incomplete groups retain their original value and stay subject
    to the conservative raw-cardinality gate.
    """
    status_columns = [
        "scenario",
        "sheet",
        "measure",
        "fuel_label",
        "year",
        "projection_allocation_complete",
        "projection_target_pair_count",
        "projection_matched_pair_count",
        "projection_allocation_methods",
        "projection_share_sources",
    ]
    if (
        comparison_long is None
        or comparison_long.empty
        or projection_tables is None
        or projection_tables.empty
    ):
        return comparison_long.copy(), pd.DataFrame(columns=status_columns)

    keys = ["scenario", "sheet", "measure", "fuel_label", "year"]
    targets = comparison_long.copy()
    for column in ["scenario", "sheet", "measure", "fuel_label", "source"]:
        if column not in targets.columns:
            targets[column] = ""
        targets[column] = targets[column].fillna("").astype(str).str.strip()
    targets["scenario"] = targets["scenario"].str.title()
    targets["year"] = pd.to_numeric(targets["year"], errors="coerce").astype("Int64")
    targets = targets[
        targets["source"].str.lower().eq("projection") & targets["year"].notna()
    ][keys].drop_duplicates()
    if targets.empty:
        return comparison_long.copy(), pd.DataFrame(columns=status_columns)

    pair_records = _projection_pair_records(mapping_status)
    if pair_records.empty:
        return comparison_long.copy(), pd.DataFrame(columns=status_columns)
    expanded = targets.merge(
        pair_records,
        on=["sheet", "measure", "fuel_label"],
        how="left",
    )
    expanded = expanded[
        expanded["esto_flow"].fillna("").astype(str).str.strip().ne("")
        & expanded["esto_product"].fillna("").astype(str).str.strip().ne("")
    ].drop_duplicates()
    if expanded.empty:
        return comparison_long.copy(), pd.DataFrame(columns=status_columns)

    projection = projection_tables.copy()
    projection["scenario"] = (
        projection.get("scenario", "").fillna("").astype(str).str.strip().str.title()
    )
    for column in ["esto_flow", "esto_product"]:
        projection[column] = projection.get(column, "").fillna("").astype(str).str.strip()
    year_columns = [
        column
        for column in projection.columns
        if str(column).isdigit()
    ]
    if not year_columns:
        return comparison_long.copy(), pd.DataFrame(columns=status_columns)
    projection_long = projection.melt(
        id_vars=["scenario", "esto_flow", "esto_product"],
        value_vars=year_columns,
        var_name="year",
        value_name="_allocated_projection_value",
    )
    projection_long["year"] = pd.to_numeric(
        projection_long["year"], errors="coerce"
    ).astype("Int64")
    projection_long["_allocated_projection_value"] = pd.to_numeric(
        projection_long["_allocated_projection_value"], errors="coerce"
    )
    projection_long = (
        projection_long.groupby(
            ["scenario", "esto_flow", "esto_product", "year"],
            as_index=False,
            dropna=False,
        )["_allocated_projection_value"]
        .sum(min_count=1)
    )

    expanded = expanded.merge(
        projection_long,
        on=["scenario", "esto_flow", "esto_product", "year"],
        how="left",
        indicator="_projection_merge",
    )
    expanded["_pair_key"] = (
        expanded["esto_flow"].astype(str) + "\x1f" + expanded["esto_product"].astype(str)
    )
    pair_counts = (
        expanded.groupby(keys, dropna=False, as_index=False)
        .agg(
            projection_target_pair_count=("_pair_key", "nunique"),
            projection_matched_pair_count=(
                "_projection_merge",
                lambda values: int((values == "both").sum()),
            ),
            _allocated_projection_value=(
                "_allocated_projection_value",
                lambda values: values.sum(min_count=1),
            ),
        )
    )
    pair_counts["projection_allocation_complete"] = (
        pair_counts["projection_target_pair_count"].gt(0)
        & pair_counts["projection_matched_pair_count"].eq(
            pair_counts["projection_target_pair_count"]
        )
    )

    provenance_summary = pd.DataFrame(
        columns=[*keys, "projection_allocation_methods", "projection_share_sources"]
    )
    if allocation_provenance is not None and not allocation_provenance.empty:
        provenance = allocation_provenance.copy()
        provenance["scenario"] = (
            provenance.get("scenario", "")
            .fillna("")
            .astype(str)
            .str.strip()
            .str.title()
        )
        provenance["year"] = pd.to_numeric(
            provenance.get("year"), errors="coerce"
        ).astype("Int64")
        for column in [
            "esto_flow",
            "esto_product",
            "allocation_method",
            "share_source",
        ]:
            provenance[column] = (
                provenance.get(column, "").fillna("").astype(str).str.strip()
            )
        provenance_join = expanded[
            [*keys, "esto_flow", "esto_product"]
        ].drop_duplicates().merge(
            provenance[
                [
                    "scenario",
                    "year",
                    "esto_flow",
                    "esto_product",
                    "allocation_method",
                    "share_source",
                ]
            ],
            on=["scenario", "year", "esto_flow", "esto_product"],
            how="left",
        )
        provenance_summary = (
            provenance_join.groupby(keys, as_index=False, dropna=False)
            .agg(
                projection_allocation_methods=(
                    "allocation_method",
                    _unique_pipe,
                ),
                projection_share_sources=("share_source", _unique_pipe),
            )
        )

    status = pair_counts.merge(provenance_summary, on=keys, how="left")
    status["projection_allocation_methods"] = (
        status["projection_allocation_methods"].fillna("")
    )
    status["projection_share_sources"] = status["projection_share_sources"].fillna("")

    replacement = status[
        [*keys, "projection_allocation_complete", "_allocated_projection_value"]
    ]
    output = comparison_long.copy()
    for column in ["scenario", "sheet", "measure", "fuel_label", "source"]:
        if column not in output.columns:
            output[column] = ""
        output[column] = output[column].fillna("").astype(str).str.strip()
    output["scenario"] = output["scenario"].str.title()
    output["year"] = pd.to_numeric(output["year"], errors="coerce").astype("Int64")
    output = output.merge(replacement, on=keys, how="left")
    replace_mask = (
        output["source"].str.lower().eq("projection")
        & output["projection_allocation_complete"].fillna(False).astype(bool)
    )
    output.loc[replace_mask, "value"] = output.loc[
        replace_mask, "_allocated_projection_value"
    ]
    output = output.drop(
        columns=["projection_allocation_complete", "_allocated_projection_value"]
    )
    return output, status[status_columns].reset_index(drop=True)


def build_canonical_projection_inputs(
    *,
    base_df: pd.DataFrame,
    ninth_df: pd.DataFrame,
    mapping_pairs_path: ConfigTableRef,
    base_year: int,
    projection_years: Sequence[int],
    scenarios: Sequence[str],
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """Build allocated ESTO projections once per selected 9th scenario."""
    if not projection_years:
        return pd.DataFrame(), pd.DataFrame(), pd.DataFrame()
    from codebase.functions.ninth_projection_mapping import (
        build_esto_projection_table,
    )

    projection_parts: list[pd.DataFrame] = []
    diagnostic_parts: list[pd.DataFrame] = []
    provenance_parts: list[pd.DataFrame] = []
    resolved_mapping = _resolve_config_table_ref(mapping_pairs_path)
    for scenario in _normalize_scenarios(scenarios):
        projection, allocation_diagnostics, provenance = build_esto_projection_table(
            ninth_data=ninth_df,
            esto_data=base_df,
            mapping_path=resolved_mapping,
            base_year=int(base_year),
            projection_years=projection_years,
            scenario=scenario.lower(),
            sign_stable_flows="all",
            strict_conservation=False,
            return_allocation_provenance=True,
        )
        for frame in (projection, allocation_diagnostics, provenance):
            if frame is not None and not frame.empty:
                frame["scenario"] = scenario
        if projection is not None and not projection.empty:
            projection_parts.append(projection)
        if allocation_diagnostics is not None and not allocation_diagnostics.empty:
            diagnostic_parts.append(allocation_diagnostics)
        if provenance is not None and not provenance.empty:
            provenance_parts.append(provenance)
    return (
        pd.concat(projection_parts, ignore_index=True, sort=False)
        if projection_parts
        else pd.DataFrame(),
        pd.concat(diagnostic_parts, ignore_index=True, sort=False)
        if diagnostic_parts
        else pd.DataFrame(),
        pd.concat(provenance_parts, ignore_index=True, sort=False)
        if provenance_parts
        else pd.DataFrame(),
    )


def _comparison_grain(row: pd.Series) -> str:
    leap_count = int(row.get("leap_component_count", 0) or 0)
    ninth_count = int(row.get("ninth_pair_count", 0) or 0)
    ninth_claimants = int(row.get("ninth_pair_max_esto_claimants", 0) or 0)
    source = _clean_token(row.get("reference_source", "")).lower()
    if source == "9th outlook":
        if bool(row.get("projection_allocation_complete", False)):
            return "canonical_allocated_ninth_to_esto_pair"
        if ninth_claimants > 1 and leap_count > 1:
            return "aggregate_many_leap_via_shared_ninth_pair"
        if ninth_claimants > 1:
            return "aggregate_shared_ninth_pair_across_esto_rows"
        if leap_count > 1 and ninth_count > 1:
            return "aggregate_many_leap_to_many_ninth"
        if leap_count > 1:
            return "aggregate_many_leap_to_one_ninth"
        if ninth_count > 1:
            return "aggregate_one_leap_to_many_ninth"
        return "direct_leap_to_ninth_via_esto_pair"
    if leap_count > 1:
        return "aggregate_many_leap_to_one_esto"
    return "direct_leap_to_esto_pair"


def _allocation_reason(row: pd.Series) -> str:
    reasons: list[str] = []
    if int(row.get("leap_component_count", 0) or 0) > 1:
        reasons.append("multiple_leap_components_share_the_esto_pair")
    projection_allocation_complete = bool(
        row.get("projection_allocation_complete", False)
    )
    if (
        _clean_token(row.get("reference_source", "")).lower() == "9th outlook"
        and not projection_allocation_complete
        and int(row.get("ninth_pair_count", 0) or 0) > 1
    ):
        reasons.append("esto_pair_sums_multiple_ninth_pairs")
    if (
        _clean_token(row.get("reference_source", "")).lower() == "9th outlook"
        and not projection_allocation_complete
        and int(row.get("ninth_pair_max_esto_claimants", 0) or 0) > 1
    ):
        reasons.append("ninth_pair_is_shared_by_multiple_esto_pairs")
    return ";".join(reasons)


def _add_refinery_auxiliary_own_use_to_base_reference(
    *,
    wide: pd.DataFrame,
    base_df: pd.DataFrame | None,
    economy: str,
) -> pd.DataFrame:
    """Add the configured refinery own-use flow to its ESTO base comparator.

    The LEAP Oil Refining balance row is the net-by-fuel module boundary. Its
    source comparator therefore needs the transformation flow plus the exact
    own-use flow maintained for that module in ``MAJOR_SECTOR_CONFIG``.
    """
    if base_df is None or base_df.empty or wide.empty:
        return wide

    config = MAJOR_SECTOR_CONFIG["oil_refineries"]
    transformation_flows = list(config.get("transformation_flow_codes", []))
    auxiliary_flows = list(config.get("loss_flow_codes", []))
    if len(transformation_flows) != 1 or not auxiliary_flows:
        raise ValueError(
            "Oil-refinery diagnostic expects one transformation flow and at "
            "least one configured own-use flow."
        )

    required = {"economy", "flows", "products"}
    missing = sorted(required - set(base_df.columns))
    if missing:
        raise KeyError(f"base_df is missing refinery comparison columns: {missing}")

    year_columns = {
        int(str(column)): column
        for column in base_df.columns
        if str(column).strip().isdigit()
    }
    scoped = base_df.copy()
    scoped["_economy_key"] = (
        scoped["economy"].fillna("").astype(str).str.replace("_", "", regex=False).str.upper()
    )
    scoped = scoped[
        scoped["_economy_key"].eq(_clean_token(economy).replace("_", "").upper())
        & scoped["flows"].fillna("").astype(str).str.strip().isin(auxiliary_flows)
    ].copy()
    if "is_subtotal" in scoped.columns:
        scoped = scoped[~scoped["is_subtotal"].fillna(False).astype(bool)].copy()
    if scoped.empty:
        return wide

    adjustment_rows: list[dict[str, Any]] = []
    for _, row in scoped.iterrows():
        product = _clean_token(row.get("products", ""))
        for year, column in year_columns.items():
            value = pd.to_numeric(pd.Series([row.get(column)]), errors="coerce").iloc[0]
            if pd.notna(value):
                adjustment_rows.append(
                    {
                        "esto_product": product,
                        "year": year,
                        "_refinery_auxiliary_own_use_pj": float(value),
                    }
                )
    if not adjustment_rows:
        return wide

    adjustments = (
        pd.DataFrame(adjustment_rows)
        .groupby(["esto_product", "year"], as_index=False)["_refinery_auxiliary_own_use_pj"]
        .sum()
    )
    out = wide.merge(adjustments, on=["esto_product", "year"], how="left")
    refinery_mask = (
        out["esto_flow"].eq(transformation_flows[0])
        & out["base"].notna()
        & out["_refinery_auxiliary_own_use_pj"].notna()
    )
    out.loc[refinery_mask, "base"] = (
        out.loc[refinery_mask, "base"]
        + out.loc[refinery_mask, "_refinery_auxiliary_own_use_pj"]
    )
    return out.drop(columns="_refinery_auxiliary_own_use_pj")


def build_leap_source_difference_table(
    *,
    comparison_long: pd.DataFrame,
    mapping_status: pd.DataFrame,
    leap_long: pd.DataFrame | None = None,
    projection_allocation_status: pd.DataFrame | None = None,
    base_df: pd.DataFrame | None = None,
    economy: str,
    years: Sequence[int],
    scenarios: Sequence[str],
    tolerance_pj: float = DEFAULT_TOLERANCE_PJ,
) -> pd.DataFrame:
    """Build the narrow Step 1 LEAP-versus-source diagnostic table."""
    required = {"scenario", "sheet", "measure", "fuel_label", "source", "year", "value"}
    missing = sorted(required - set(comparison_long.columns))
    if missing:
        raise KeyError(f"comparison_long is missing required columns: {missing}")

    selected_years = {int(year) for year in years}
    selected_scenarios = {scenario.lower() for scenario in _normalize_scenarios(scenarios)}
    working = comparison_long.copy()
    working["scenario"] = working["scenario"].fillna("").astype(str).str.strip().str.title()
    working["year"] = pd.to_numeric(working["year"], errors="coerce").astype("Int64")
    working["value"] = pd.to_numeric(working["value"], errors="coerce")
    working["source"] = working["source"].fillna("").astype(str).str.strip().str.lower()
    working = working[
        working["year"].isin(selected_years)
        & working["scenario"].str.lower().isin(selected_scenarios)
        & working["source"].isin({"leap", "base", "projection"})
    ].copy()
    if working.empty:
        return pd.DataFrame(columns=DIFFERENCE_OUTPUT_COLUMNS)

    key_columns = ["scenario", "sheet", "measure", "fuel_label", "year"]
    grouped = (
        working.groupby([*key_columns, "source"], dropna=False, as_index=False)["value"]
        .sum(min_count=1)
    )
    wide = grouped.pivot_table(
        index=key_columns,
        columns="source",
        values="value",
        aggfunc="first",
    ).reset_index()
    wide.columns.name = None
    for column in ["leap", "base", "projection"]:
        if column not in wide.columns:
            wide[column] = pd.NA
        wide[column] = pd.to_numeric(wide[column], errors="coerce")

    metadata = _build_mapping_metadata(mapping_status, leap_long)
    wide = wide.merge(metadata, on=["sheet", "measure", "fuel_label"], how="left")
    wide = _add_refinery_auxiliary_own_use_to_base_reference(
        wide=wide,
        base_df=base_df,
        economy=economy,
    )

    has_base = wide["base"].notna()
    has_projection = wide["projection"].notna()
    wide["reference_source"] = ""
    wide.loc[has_base & ~has_projection, "reference_source"] = "ESTO"
    wide.loc[has_projection & ~has_base, "reference_source"] = "9th Outlook"
    wide.loc[has_base & has_projection, "reference_source"] = "ambiguous"
    wide["source_value_pj"] = wide["base"].combine_first(wide["projection"])
    wide["leap_value_pj"] = wide["leap"]

    both_present = wide["leap_value_pj"].notna() & wide["source_value_pj"].notna()
    wide["difference_pj"] = wide["leap_value_pj"] - wide["source_value_pj"]
    wide["absolute_difference_pj"] = wide["difference_pj"].abs()
    wide["correction_to_match_source_pj"] = -wide["difference_pj"]
    nonzero_reference = wide["source_value_pj"].abs().gt(float(tolerance_pj))
    wide["difference_percent"] = pd.NA
    wide.loc[both_present & nonzero_reference, "difference_percent"] = (
        wide.loc[both_present & nonzero_reference, "difference_pj"]
        / wide.loc[both_present & nonzero_reference, "source_value_pj"]
        * 100.0
    )

    wide["status"] = "reference_unavailable"
    wide.loc[wide["leap_value_pj"].isna() & wide["source_value_pj"].notna(), "status"] = "missing_in_leap"
    wide.loc[
        both_present & wide["absolute_difference_pj"].le(float(tolerance_pj)),
        "status",
    ] = "match"
    wide.loc[
        both_present & wide["absolute_difference_pj"].gt(float(tolerance_pj)),
        "status",
    ] = "value_mismatch"
    wide.loc[has_base & has_projection, "status"] = "ambiguous_reference"
    wide["is_mismatch"] = wide["status"].isin({"value_mismatch", "missing_in_leap"})

    allocation_keys = ["scenario", "sheet", "measure", "fuel_label", "year"]
    if (
        projection_allocation_status is not None
        and not projection_allocation_status.empty
    ):
        allocation_status = projection_allocation_status.copy()
        allocation_status["scenario"] = (
            allocation_status["scenario"]
            .fillna("")
            .astype(str)
            .str.strip()
            .str.title()
        )
        allocation_status["year"] = pd.to_numeric(
            allocation_status["year"], errors="coerce"
        ).astype("Int64")
        wide = wide.merge(
            allocation_status,
            on=allocation_keys,
            how="left",
        )
    for column in [
        "esto_flow",
        "esto_product",
        "ninth_sector_codes",
        "ninth_fuel_codes",
        "leap_sector_names",
        "leap_fuel_names",
    ]:
        wide[column] = wide.get(column, "").fillna("").astype(str)
    for column in [
        "leap_component_count",
        "ninth_pair_count",
        "ninth_pair_max_esto_claimants",
    ]:
        wide[column] = pd.to_numeric(wide.get(column, 0), errors="coerce").fillna(0).astype(int)
    if "projection_allocation_complete" not in wide.columns:
        wide["projection_allocation_complete"] = False
    wide["projection_allocation_complete"] = wide[
        "projection_allocation_complete"
    ].fillna(False).astype(bool)
    for column in [
        "projection_target_pair_count",
        "projection_matched_pair_count",
    ]:
        if column not in wide.columns:
            wide[column] = 0
        wide[column] = (
            pd.to_numeric(wide[column], errors="coerce")
            .fillna(0)
            .astype(int)
        )
    for column in [
        "projection_allocation_methods",
        "projection_share_sources",
    ]:
        if column not in wide.columns:
            wide[column] = ""
        wide[column] = wide[column].fillna("").astype(str)

    wide["economy"] = _clean_token(economy)
    wide["comparison_grain"] = wide.apply(_comparison_grain, axis=1)
    wide["update_allocation_reason"] = wide.apply(_allocation_reason, axis=1)
    wide["update_allocation_required"] = wide["update_allocation_reason"].ne("")

    status_order = {
        "value_mismatch": 0,
        "missing_in_leap": 1,
        "ambiguous_reference": 2,
        "reference_unavailable": 3,
        "match": 4,
    }
    wide["_status_order"] = wide["status"].map(status_order).fillna(99)
    wide = wide.sort_values(
        ["_status_order", "absolute_difference_pj", "scenario", "year", "esto_flow", "esto_product"],
        ascending=[True, False, True, True, True, True],
        na_position="last",
        kind="mergesort",
    ).drop(columns=["_status_order", "leap", "base", "projection"])
    return wide[DIFFERENCE_OUTPUT_COLUMNS].reset_index(drop=True)


def run_economy_balance_diagnostic(
    *,
    economy: str,
    years: Sequence[int] | None,
    scenarios: Sequence[str] | None = ("Reference", "Target"),
    base_year: int = DEFAULT_BASE_YEAR,
    exports_root: Path | str = DEFAULT_EXPORTS_ROOT,
    workbook_path: Path | str | None = None,
    ref_date_id: str | None = None,
    tgt_date_id: str | None = None,
    template_sheet: str = DEFAULT_TEMPLATE_SHEET,
    tolerance_pj: float = DEFAULT_TOLERANCE_PJ,
    mapping_pairs_path: Any = DEFAULT_MAPPING_PAIRS_PATH,
    codebook_path: Path | str = DEFAULT_CODEBOOK_PATH,
    sheet_map_path: Path | str = DEFAULT_SHEET_MAP_PATH,
    backup_mappings_path: Path | str = DEFAULT_BACKUP_MAPPINGS_PATH,
    explicit_mappings_path: Path | str = DEFAULT_EXPLICIT_MAPPINGS_PATH,
    explicit_reassignments_path: Path | str = DEFAULT_EXPLICIT_REASSIGNMENTS_PATH,
    synthetic_reference_rows_path: Path | str = DEFAULT_SYNTHETIC_REFERENCE_ROWS_PATH,
    esto_table_path: Path | str = DEFAULT_BASE_TABLE_PATH,
    projection_table_path: Path | str = DEFAULT_PROJECTION_TABLE_PATH,
    known_issues_path: Path | str | None = DEFAULT_KNOWN_ISSUES_PATH,
) -> dict[str, Any]:
    """Run the read-only Step 1 diagnostic for one economy."""
    resolved_codebook_path = _resolve(codebook_path)
    resolved_sheet_map_path = _resolve(sheet_map_path)

    if workbook_path is not None:
        direct_path, selected_years, selected_scenarios = _read_direct_workbook_scope(
            workbook_path,
            base_year=base_year,
        )
        ref_path = direct_path if "Reference" in selected_scenarios else None
        tgt_path = direct_path if "Target" in selected_scenarios else None
    else:
        if years is None or scenarios is None:
            raise ValueError(
                "years and scenarios are required when no direct workbook_path is supplied."
            )
        selected_years = _validate_years(years, base_year=base_year)
        selected_scenarios = _normalize_scenarios(scenarios)
        ref_path = (
            resolve_balance_export_workbook(
                economy=economy,
                scenario="REF",
                date_id=ref_date_id,
                exports_root=exports_root,
            )
            if "Reference" in selected_scenarios
            else None
        )
        tgt_path = (
            resolve_balance_export_workbook(
                economy=economy,
                scenario="TGT",
                date_id=tgt_date_id,
                exports_root=exports_root,
            )
            if "Target" in selected_scenarios
            else None
        )
    projection_years = [year for year in selected_years if year > int(base_year)]
    scenario_map = {scenario: scenario.lower() for scenario in selected_scenarios}

    workbook_paths = [path for path in (ref_path, tgt_path) if path is not None]
    detail_inspections = require_level2_balance_export_detail(workbook_paths)
    known_issues = _load_optional_json(known_issues_path)

    with _temporary_balance_runtime_paths(
        codebook_path=resolved_codebook_path,
        sheet_map_path=resolved_sheet_map_path,
        exports_root=_resolve(exports_root),
    ) as (build_balance_comparison_esto_axis, convert_leap_balances_to_esto_long_table):
        # Windows can briefly retain the generated mapping workbook after
        # pandas/openpyxl finishes reading it. Cleanup failure must not discard
        # a completed diagnostic conversion.
        with tempfile.TemporaryDirectory(
            prefix="leap_balance_esto_mapping_",
            ignore_cleanup_errors=True,
        ) as temp_dir:
            extraction_mapping_path = _write_esto_axis_extraction_mapping_workbook(
                output_path=Path(temp_dir) / "esto_axis_extraction_mappings.xlsx",
                codebook_path=resolved_codebook_path,
            )
            conversion = convert_leap_balances_to_esto_long_table(
                ref_workbook_path=ref_path,
                tgt_workbook_path=tgt_path,
                template_sheet=template_sheet,
                mapping_pairs_path=extraction_mapping_path,
                codebook_path=resolved_codebook_path,
                known_issues=known_issues,
                projection_economy=economy,
                max_output_year=max(selected_years),
                explicit_pair_mappings_only=True,
                allow_descendant_mapping_expansion=False,
            )
        comparison = build_balance_comparison_esto_axis(
            leap_long=conversion["leap_long"],
            mapping_status=conversion["mapping_status"],
            base_year=int(base_year),
            projection_years=projection_years,
            base_economy=_compact_economy_code(economy),
            projection_economy=economy,
            scenario_map=scenario_map,
            sheet_map_path=resolved_sheet_map_path,
            backup_mappings_path=backup_mappings_path,
            codebook_path=resolved_codebook_path,
            canonical_pairs_path=mapping_pairs_path,
            explicit_mappings_path=explicit_mappings_path,
            explicit_reassignments_path=explicit_reassignments_path,
            synthetic_reference_rows_path=synthetic_reference_rows_path,
            esto_table_path=esto_table_path,
            projection_table_path=projection_table_path,
            chart_navigation_guide_path=None,
            balance_mapping_workbook_path=resolved_codebook_path,
            known_issues=known_issues,
        )
    canonical_projection, projection_allocation_diagnostics, allocation_provenance = (
        build_canonical_projection_inputs(
            base_df=comparison.get("base_df", pd.DataFrame()),
            ninth_df=comparison.get("ninth_df", pd.DataFrame()),
            mapping_pairs_path=mapping_pairs_path,
            base_year=int(base_year),
            projection_years=projection_years,
            scenarios=selected_scenarios,
        )
    )
    comparison_long_raw = comparison["comparison_long"]
    allocated_comparison_long, projection_allocation_status = (
        apply_canonical_projection_comparators(
            comparison_long=comparison_long_raw,
            mapping_status=comparison["mapping_status"],
            projection_tables=canonical_projection,
            allocation_provenance=allocation_provenance,
        )
    )
    comparison["comparison_long_raw"] = comparison_long_raw
    comparison["comparison_long"] = allocated_comparison_long
    comparison["canonical_projection"] = canonical_projection
    comparison["projection_allocation_status"] = projection_allocation_status
    comparison["projection_allocation_diagnostics"] = (
        projection_allocation_diagnostics
    )
    comparison["projection_allocation_provenance"] = allocation_provenance
    difference_table = build_leap_source_difference_table(
        comparison_long=allocated_comparison_long,
        mapping_status=comparison["mapping_status"],
        leap_long=conversion["leap_long"],
        projection_allocation_status=projection_allocation_status,
        base_df=comparison.get("base_df"),
        economy=economy,
        years=selected_years,
        scenarios=selected_scenarios,
        tolerance_pj=tolerance_pj,
    )
    mapping_issues = _scope_rows_to_diagnostic_window(
        conversion.get("issues", pd.DataFrame()),
        years=selected_years,
        scenarios=selected_scenarios,
    )
    total_balance_checks = _scope_rows_to_diagnostic_window(
        conversion.get("total_balance_checks", pd.DataFrame()),
        years=selected_years,
        scenarios=selected_scenarios,
    )
    matching_diagnostics = _scope_rows_to_diagnostic_window(
        conversion.get("matching_diagnostics", pd.DataFrame()),
        years=selected_years,
        scenarios=selected_scenarios,
    )
    return {
        "economy": economy,
        "years": selected_years,
        "scenarios": selected_scenarios,
        "ref_workbook_path": ref_path,
        "tgt_workbook_path": tgt_path,
        "detail_inspections": detail_inspections,
        "difference_table": difference_table,
        "mapping_issues": mapping_issues,
        "total_balance_checks": total_balance_checks,
        "matching_diagnostics": matching_diagnostics,
        "projection_allocation_status": projection_allocation_status,
        "projection_allocation_diagnostics": projection_allocation_diagnostics,
        "projection_allocation_provenance": allocation_provenance,
        "conversion": conversion,
        "comparison": comparison,
    }


def run_baseline_seed_balance_diagnostics(
    *,
    economies: Sequence[str],
    years: Sequence[int] | None,
    scenarios: Sequence[str] | None = ("Reference", "Target"),
    output_dir: Path | str = DEFAULT_OUTPUT_DIR,
    exports_root: Path | str = DEFAULT_EXPORTS_ROOT,
    date_ids_by_economy: dict[str, dict[str, str | None]] | None = None,
    workbook_paths_by_economy: dict[str, Path | str] | None = None,
    base_year: int = DEFAULT_BASE_YEAR,
    tolerance_pj: float = DEFAULT_TOLERANCE_PJ,
    **diagnostic_paths: Any,
) -> dict[str, Any]:
    """Run Step 1 for several economies and write one combined CSV table."""
    economy_list = [_clean_token(economy) for economy in economies if _clean_token(economy)]
    if not economy_list:
        raise ValueError("At least one economy is required.")
    resolved_output_dir = _resolve(output_dir)
    resolved_output_dir.mkdir(parents=True, exist_ok=True)

    results: dict[str, dict[str, Any]] = {}
    difference_parts: list[pd.DataFrame] = []
    issue_parts: list[pd.DataFrame] = []
    for economy in economy_list:
        date_ids = (date_ids_by_economy or {}).get(economy, {})
        direct_workbook_path = (workbook_paths_by_economy or {}).get(economy)
        result = run_economy_balance_diagnostic(
            economy=economy,
            years=None if direct_workbook_path is not None else years,
            scenarios=None if direct_workbook_path is not None else scenarios,
            base_year=base_year,
            exports_root=exports_root,
            workbook_path=direct_workbook_path,
            ref_date_id=date_ids.get("REF"),
            tgt_date_id=date_ids.get("TGT"),
            tolerance_pj=tolerance_pj,
            **diagnostic_paths,
        )
        results[economy] = result
        difference_parts.append(result["difference_table"])
        issues = result["mapping_issues"].copy()
        if not issues.empty:
            if "economy" not in issues.columns:
                issues["economy"] = economy
            issue_parts.append(issues)

    differences = (
        pd.concat(difference_parts, ignore_index=True, sort=False)
        if difference_parts
        else pd.DataFrame(columns=DIFFERENCE_OUTPUT_COLUMNS)
    )
    differences_path = resolved_output_dir / "leap_balance_source_differences.csv"
    differences.to_csv(differences_path, index=False)

    mapping_issues = pd.concat(issue_parts, ignore_index=True, sort=False) if issue_parts else pd.DataFrame()
    mapping_issues_path: Path | None = None
    if not mapping_issues.empty:
        mapping_issues_path = resolved_output_dir / "leap_balance_mapping_issues.csv"
        mapping_issues.to_csv(mapping_issues_path, index=False)

    review = build_balance_review_table(differences)
    review_path = resolved_output_dir / "leap_balance_source_review.csv"
    review.to_csv(review_path, index=False)
    diagnostic_counts = build_balance_diagnostic_counts(differences, mapping_issues)

    mismatch_count = int(differences["is_mismatch"].fillna(False).astype(bool).sum())
    allocation_count = int(
        differences["update_allocation_required"].fillna(False).astype(bool).sum()
    )
    print(
        "[INFO] baseline-seed balance diagnostic: "
        f"{len(differences):,} comparison rows, {mismatch_count:,} mismatches, "
        f"{allocation_count:,} rows needing a future allocation rule."
    )
    print(f"[INFO] Wrote {differences_path}")
    print(f"[INFO] Wrote {review_path}")
    if mapping_issues_path is not None:
        print(f"[INFO] Wrote {mapping_issues_path}")

    return {
        "differences": differences,
        "differences_path": differences_path,
        "review": review,
        "review_path": review_path,
        "mapping_issues": mapping_issues,
        "mapping_issues_path": mapping_issues_path,
        "economy_results": results,
        "summary": {
            "comparison_rows": int(len(differences)),
            "mismatch_rows": mismatch_count,
            "future_allocation_rule_rows": allocation_count,
            "mapping_issue_rows": int(len(mapping_issues)),
            **diagnostic_counts,
        },
    }
