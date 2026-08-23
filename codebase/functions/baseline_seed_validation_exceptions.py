#%%
"""Maintain the reviewable missing-branch exception workbook.

Rows in the workbook are a warning ledger, not permission to import into a
branch that does not yet exist in LEAP.  The derived PJ columns are populated
only when blank so modeller notes and an already reviewed estimate are kept.
"""
from __future__ import annotations

import os
from pathlib import Path
from typing import Iterable

import pandas as pd
from openpyxl import load_workbook

from codebase.analysis.missing_branch_esto_vintage_impact import VINTAGES as ESTO_VINTAGES
from codebase.mapping_tools.missing_branch_registry_materiality_workflow import (
    _esto_base_materiality,
    _projection_materiality,
    _registry_source_keys,
)
from codebase.utilities import leap_export_template_resolver


REPO_ROOT = Path(__file__).resolve().parents[2]
WORKBOOK_PATH = REPO_ROOT / "config" / "baseline_seed_validation_exception_sets.xlsx"
SHEET_NAME = "branch_exceptions"
BASE_COLUMNS = ["enabled", "branch_path", "notes", "economies_that_need_it", "economies_resolved_in_templates"]
ESTO_VALUE_COLUMNS = [
    f"esto_{vintage}_last_year_signed_pj_all_economies"
    for vintage in ESTO_VINTAGES
]
NINTH_VALUE_COLUMN = "ninth_reference_average_pj_per_year_all_economies"
RELEVANCE_AUDIT_COLUMN = "relevance_audit"
ZERO_FILTER_COLUMN = "zero filter"
MATERIALITY_VALUE_COLUMNS = [*ESTO_VALUE_COLUMNS, NINTH_VALUE_COLUMN]
REQUIRED_COLUMNS = [*BASE_COLUMNS, *MATERIALITY_VALUE_COLUMNS, RELEVANCE_AUDIT_COLUMN, ZERO_FILTER_COLUMN]


def _blank(value: object) -> bool:
    return value is None or pd.isna(value) or not str(value).strip()


def _truthy(value: object) -> bool:
    return str(value or "").strip().casefold() in {"1", "true", "t", "yes", "y", "on"}


def _normalise_path(value: object) -> str:
    return "\\".join(part.strip() for part in str(value or "").replace("/", "\\").split("\\") if part.strip()).casefold()


def _economies(value: object) -> set[str]:
    """Read the workbook's compact, reviewable pipe-separated economy list."""
    return {part.strip() for part in str(value or "").split("|") if part.strip() and part.strip().casefold() != "all"}


def _append_unique_note(existing: object, message: str) -> str:
    """Append one audit sentence once, without erasing modeller-authored notes."""
    note = str(existing or "").strip()
    return note if message in note else f"{note} {message}".strip()


def _has_material_seed_trigger(row: pd.Series) -> bool:
    """The completed-vintage audit is stronger evidence than a simplified mapping zero."""
    return "triggered for" in str(row.get(RELEVANCE_AUDIT_COLUMN, "")).casefold()


def apply_zero_filter(rows: pd.DataFrame) -> pd.DataFrame:
    """Mark simplified all-zero materiality results that contradict seed evidence.

    A material seed expression can be nonzero through a proxy, aggregation, or
    multi-row mapping even where the current one-flow/one-fuel materiality
    helper returns zero.  Such zeros must remain blank and visibly flagged
    until the canonical mapping is added; they are not evidence for pruning.
    """
    for column in [*MATERIALITY_VALUE_COLUMNS, RELEVANCE_AUDIT_COLUMN, ZERO_FILTER_COLUMN]:
        if column not in rows:
            rows[column] = ""
    rows[MATERIALITY_VALUE_COLUMNS] = rows[MATERIALITY_VALUE_COLUMNS].astype(object)
    for index, row in rows.iterrows():
        values = pd.to_numeric(row[MATERIALITY_VALUE_COLUMNS], errors="coerce")
        all_zero = values.notna().all() and values.eq(0).all()
        any_blank = values.isna().any()
        triggered = _has_material_seed_trigger(row)
        enabled = _truthy(row.get("enabled", ""))
        if triggered and all_zero:
            rows.loc[index, MATERIALITY_VALUE_COLUMNS] = ""
            rows.at[index, ZERO_FILTER_COLUMN] = "MAPPING INCOMPLETE — seed triggered"
        elif triggered and any_blank:
            rows.at[index, ZERO_FILTER_COLUMN] = "MAPPING INCOMPLETE — seed triggered"
        elif not enabled:
            rows.at[index, ZERO_FILTER_COLUMN] = "DISABLED — no trigger across audited vintages"
        elif values.notna().all() and values.ne(0).any():
            rows.at[index, ZERO_FILTER_COLUMN] = "CONFIRMED NONZERO"
        elif all_zero:
            rows.at[index, ZERO_FILTER_COLUMN] = "ZERO VALUES — review seed relevance"
        else:
            rows.at[index, ZERO_FILTER_COLUMN] = "MAPPING INCOMPLETE — source mapping required"
    return rows


def _read_rows(workbook_path: Path = WORKBOOK_PATH) -> pd.DataFrame:
    if not workbook_path.exists():
        return pd.DataFrame(columns=REQUIRED_COLUMNS)
    rows = pd.read_excel(workbook_path, sheet_name=SHEET_NAME, dtype=object).fillna("")
    missing = {"enabled", "branch_path", "notes"}.difference(rows.columns)
    if missing:
        raise ValueError(f"{workbook_path.name} is missing required columns: {sorted(missing)}")
    for column in REQUIRED_COLUMNS:
        if column not in rows:
            rows[column] = ""
    return rows.reindex(columns=[*REQUIRED_COLUMNS, *[c for c in rows.columns if c not in REQUIRED_COLUMNS]])


def _write_rows(rows: pd.DataFrame, workbook_path: Path = WORKBOOK_PATH) -> None:
    workbook = load_workbook(workbook_path)
    if SHEET_NAME not in workbook.sheetnames:
        raise ValueError(f"{workbook_path.name} has no {SHEET_NAME!r} sheet")
    sheet = workbook[SHEET_NAME]
    sheet.delete_rows(1, sheet.max_row)
    for row in [list(rows.columns), *rows.fillna("").itertuples(index=False, name=None)]:
        sheet.append(list(row))
    sheet.freeze_panes = "A2"
    workbook.save(workbook_path)


def refresh_exception_materiality(
    workbook_path: Path | str = WORKBOOK_PATH,
    *,
    ninth_path: Path | str = REPO_ROOT / "data" / "merged_file_energy_ALL_20251106.csv",
    projection_start_year: int = 2023,
    projection_final_year: int = 2060,
    esto_vintages: dict[str, tuple[Path, int]] = ESTO_VINTAGES,
    retry_mapping_incomplete: bool = False,
) -> pd.DataFrame:
    """Fill blank derived values from all maintained ESTO vintages and Ninth.

    ESTO values use each file's final historical year.  The Ninth value is the
    signed Reference total across all economies divided by the model years.
    ``retry_mapping_incomplete`` is for an explicit recheck after a canonical
    mapping repair; normal refreshes preserve known mapping-incomplete blanks.
    """
    workbook = Path(workbook_path)
    rows = _read_rows(workbook)
    value_columns = MATERIALITY_VALUE_COLUMNS
    needs_values = rows[rows["branch_path"].map(lambda value: bool(str(value).strip()))]
    if needs_values.empty:
        rows = apply_zero_filter(rows)
        _write_rows(rows, workbook)
        return rows
    needs_values = needs_values[needs_values[value_columns].map(_blank).any(axis=1)]
    if not retry_mapping_incomplete:
        needs_values = needs_values[
            ~needs_values[ZERO_FILTER_COLUMN].astype(str).str.startswith("MAPPING INCOMPLETE — seed triggered")
        ]
    if needs_values.empty:
        rows = apply_zero_filter(rows)
        _write_rows(rows, workbook)
        return rows

    mapped_rows = []
    for item in needs_values[["branch_path"]].to_dict("records"):
        try:
            mapped_rows.extend(_registry_source_keys([item]).to_dict("records"))
        except ValueError as exc:
            note_index = rows.index[rows["branch_path"].map(_normalise_path).eq(_normalise_path(item["branch_path"]))]
            for index in note_index:
                note = str(rows.at[index, "notes"] or "").strip()
                if "materiality mapping needs review" not in note.casefold():
                    rows.at[index, "notes"] = (note + " " if note else "") + f"Materiality mapping needs review: {exc}"
    if not mapped_rows:
        rows = apply_zero_filter(rows)
        _write_rows(rows, workbook)
        return rows
    keys = pd.DataFrame(mapped_rows)
    for vintage, (esto_path, final_year) in esto_vintages.items():
        values = _esto_base_materiality(keys, esto_path=Path(esto_path), base_year=int(final_year))
        column = f"esto_{vintage}_last_year_signed_pj_all_economies"
        for index in needs_values.index:
            if rows.at[index, "branch_path"] in set(keys["branch_path"]) and _blank(rows.at[index, column]):
                rows.at[index, column] = values[rows.at[index, "branch_path"]][0]

    years = list(range(projection_start_year, projection_final_year + 1))
    projections = _projection_materiality(keys, ninth_path=Path(ninth_path), projection_years=years)
    for index in needs_values.index:
        if rows.at[index, "branch_path"] in set(keys["branch_path"]) and _blank(rows.at[index, NINTH_VALUE_COLUMN]):
            rows.at[index, NINTH_VALUE_COLUMN] = (
                projections[rows.at[index, "branch_path"]]["reference"][0] / len(years)
            )
    rows = apply_zero_filter(rows)
    _write_rows(rows, workbook)
    return rows


def register_missing_branch_paths(
    branch_paths: Iterable[str],
    *,
    economy: str,
    workbook_path: Path | str = WORKBOOK_PATH,
) -> list[str]:
    """Append newly observed paths as enabled warning rows and return additions."""
    workbook = Path(workbook_path)
    rows = _read_rows(workbook)
    known = {_normalise_path(value) for value in rows["branch_path"]}
    additions = []
    for branch_path in sorted({str(path).strip() for path in branch_paths if str(path).strip()}):
        if _normalise_path(branch_path) in known:
            continue
        additions.append(branch_path)
        rows.loc[len(rows)] = {
            "enabled": True,
            "branch_path": branch_path,
            "notes": "Automatically recorded from baseline-seed validation; review/create in LEAP when appropriate.",
            "economies_that_need_it": str(economy).strip(),
            "economies_resolved_in_templates": "",
        }
        known.add(_normalise_path(branch_path))
    if additions:
        _write_rows(rows, workbook)
        if workbook.resolve() == WORKBOOK_PATH.resolve():
            refresh_placeholder_review_workbook(workbook)
    return additions


def register_material_missing_branch_findings(
    findings: pd.DataFrame,
    *,
    workbook_path: Path | str = WORKBOOK_PATH,
) -> list[str]:
    """Record all material unknown paths, retaining the economies that triggered them.

    This is the final-stage registration boundary: a material unknown path is a
    warning-ledger entry, never a reason to block the baseline seed.  Existing
    rows gain newly observed economies instead of being duplicated.
    """
    required = {"economy", "branch_path"}
    missing = required.difference(findings.columns)
    if missing:
        raise ValueError(f"Material findings are missing columns: {sorted(missing)}")
    workbook = Path(workbook_path)
    rows = _read_rows(workbook)
    observed: dict[str, tuple[str, set[str]]] = {}
    for item in findings[["economy", "branch_path"]].dropna().itertuples(index=False):
        path = str(item.branch_path).strip()
        economy = str(item.economy).strip()
        if path and economy:
            key = _normalise_path(path)
            canonical, economies = observed.get(key, (path, set()))
            economies.add(economy)
            observed[key] = (canonical, economies)

    existing = {_normalise_path(row.branch_path): index for index, row in rows.iterrows() if str(row.branch_path).strip()}
    additions: list[str] = []
    for key, (path, economies) in sorted(observed.items(), key=lambda item: item[1][0]):
        if key in existing:
            index = existing[key]
            rows.at[index, "economies_that_need_it"] = "|".join(sorted(_economies(rows.at[index, "economies_that_need_it"]) | economies))
            continue
        additions.append(path)
        rows.loc[len(rows)] = {
            "enabled": True,
            "branch_path": path,
            "notes": "Automatically recorded from material baseline-seed validation; review/create in LEAP when appropriate. Warnings only: this row never blocks seed generation.",
            "economies_that_need_it": "|".join(sorted(economies)),
            "economies_resolved_in_templates": "",
        }
    if additions or not findings.empty:
        _write_rows(rows, workbook)
        if workbook.resolve() == WORKBOOK_PATH.resolve():
            refresh_placeholder_review_workbook(workbook)
    return additions


def audit_exception_relevance(
    findings_by_vintage: dict[str, pd.DataFrame],
    *,
    workbook_path: Path | str = WORKBOOK_PATH,
    prune_after_all_vintages: bool = False,
) -> pd.DataFrame:
    """Record whether enabled exception economies actually triggered in completed runs.

    A single-vintage audit only records evidence.  When the caller explicitly
    confirms that every relevant vintage was checked, stale economy memberships
    are removed and a row with no remaining economies is disabled.  This avoids
    incorrectly disabling a branch merely because it is zero in one vintage.
    """
    workbook = Path(workbook_path)
    rows = _read_rows(workbook)
    observed_by_vintage: dict[str, dict[str, set[str]]] = {}
    for vintage, findings in findings_by_vintage.items():
        required = {"economy", "branch_path"}
        if required.difference(findings.columns):
            raise ValueError(f"{vintage} material findings are missing {sorted(required.difference(findings.columns))}")
        observed: dict[str, set[str]] = {}
        for item in findings[["economy", "branch_path"]].dropna().itertuples(index=False):
            path, economy = str(item.branch_path).strip(), str(item.economy).strip()
            if path and economy:
                observed.setdefault(_normalise_path(path), set()).add(economy)
        observed_by_vintage[str(vintage)] = observed

    for index, row in rows.iterrows():
        path = str(row["branch_path"]).strip()
        if not path:
            continue
        configured = _economies(row["economies_that_need_it"])
        key = _normalise_path(path)
        triggered_union: set[str] = set()
        audit_parts: list[str] = []
        for vintage, observed in observed_by_vintage.items():
            triggered = observed.get(key, set())
            triggered_union.update(triggered)
            no_longer_needed = configured - triggered
            if triggered:
                audit_parts.append(f"{vintage}: triggered for {'|'.join(sorted(triggered))}.")
            if no_longer_needed:
                audit_parts.append(
                    f"{vintage}: found not needed for {'|'.join(sorted(no_longer_needed))} (no material unknown-path finding)."
                )
        summary = " ".join(audit_parts) or "No supplied run contained this exception path."
        rows.at[index, RELEVANCE_AUDIT_COLUMN] = summary
        rows.at[index, "notes"] = _append_unique_note(
            row["notes"], f"Relevance audit: {summary}"
        )
        if prune_after_all_vintages and configured:
            retained = configured & triggered_union
            rows.at[index, "economies_that_need_it"] = "|".join(sorted(retained))
            if not retained:
                rows.at[index, "enabled"] = False
                rows.at[index, "notes"] = _append_unique_note(
                    rows.at[index, "notes"],
                    "Disabled after the completed all-vintage relevance audit: it did not trigger for any previously marked economy.",
                )
    rows = apply_zero_filter(rows)
    _write_rows(rows, workbook)
    if workbook.resolve() == WORKBOOK_PATH.resolve():
        refresh_placeholder_review_workbook(workbook)
    return rows


def refresh_placeholder_review_workbook(workbook_path: Path | str = WORKBOOK_PATH) -> Path:
    """Refresh the read-only per-economy placeholder review after ledger changes.

    This never changes a template.  It also makes added and removed/disabled
    ledger paths immediately visible to the modeller before any apply request.
    """
    from codebase.mapping_tools.add_validation_exception_template_rows import (
        write_material_exception_placeholder_review_workbook,
    )

    return write_material_exception_placeholder_review_workbook(
        exception_workbook_path=workbook_path,
    )


def sync_exception_template_coverage(
    workbook_path: Path | str = WORKBOOK_PATH,
    *,
    templates_root: Path | str | None = None,
) -> pd.DataFrame:
    """Update exact per-economy template coverage; use ``all`` only when real everywhere."""
    workbook = Path(workbook_path)
    rows = _read_rows(workbook)
    resolved_templates_root = (
        leap_export_template_resolver.DEFAULT_LEAP_EXPORT_TEMPLATES_ROOT
        if templates_root is None else templates_root
    )
    templates = list(leap_export_template_resolver.iter_leap_export_templates(resolved_templates_root))
    if not templates:
        raise FileNotFoundError(f"No LEAP export templates found under {resolved_templates_root}")
    from codebase.mapping_tools.add_validation_exception_template_rows import _exception_rows, PLACEHOLDER_BRANCH_ID

    paths = {str(value).strip() for value in rows["branch_path"] if str(value).strip()}
    template_paths: dict[str, dict[str, bool]] = {}
    for template in templates:
        matches_by_path: dict[str, list[dict[str, str]]] = {path: [] for path in paths}
        for item in _exception_rows(template.path, paths):
            matches_by_path[item["path"]].append(item)
        template_paths[template.economy] = {
            path: bool(matches) and all(
                str(item["branch_id"]) not in {str(PLACEHOLDER_BRANCH_ID), "100"}
                for item in matches
            )
            for path, matches in matches_by_path.items()
        }
    for index, row in rows.iterrows():
        path = str(row["branch_path"]).strip()
        if not path:
            continue
        resolved = sorted(economy for economy, statuses in template_paths.items() if statuses[path])
        needed = sorted(economy for economy, statuses in template_paths.items() if not statuses[path])
        # ``economies_that_need_it`` is an observed materiality/relevance list,
        # not a list of every template that happens not to have the branch.
        rows.at[index, "economies_resolved_in_templates"] = "all" if not needed else "|".join(resolved)
    _write_rows(rows, workbook)
    return rows


def load_enabled_exception_notes(
    workbook_path: Path | str = WORKBOOK_PATH,
    *,
    refresh_materiality: bool = True,
) -> dict[str, str]:
    """Return enabled exact exceptions after filling any missing materiality."""
    workbook = Path(workbook_path)
    if refresh_materiality and not os.environ.get("LEAP_SKIP_EXCEPTION_MATERIALITY_REFRESH"):
        refresh_exception_materiality(workbook)
    rows = _read_rows(workbook)
    return {
        str(row.branch_path).strip(): str(row.notes).strip()
        for row in rows.itertuples(index=False)
        if _truthy(row.enabled) and str(row.branch_path).strip()
    }


#%%
