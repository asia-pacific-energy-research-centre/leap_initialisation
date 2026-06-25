from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

import pandas as pd
from openpyxl import load_workbook


@dataclass
class ReplicaSheetResult:
    template: str
    sheet: str
    scenario: str
    branch: str
    variable_template: str
    variable_used: str
    legend_template: str
    legend_used: str
    mode: str
    source_csv: str
    status: str
    notes: str
    golden_status: str
    max_abs_diff: float


def _resolve(path: Path | str) -> Path:
    raw = str(path).replace("\\", "/")
    candidate = Path(raw)
    if candidate.is_absolute():
        return candidate
    return (Path(__file__).resolve().parents[2] / candidate).resolve()


def _strict_transformation_variable(sheet_name: str, fallback: str) -> str:
    name = str(sheet_name or "").strip().lower()
    if (
        name.endswith("_inputs")
        or name.endswith("_feed_inputs")
        or name.endswith("_aux")
        or name.endswith("_aux_other")
    ):
        return "Inputs"
    if name.endswith("_out_feed"):
        return "Outputs by Feedstock Fuel"
    if name.endswith("_out_fuel"):
        return "Outputs by Output Fuel"
    return str(fallback or "").strip()


def _strict_transformation_legend(sheet_name: str, fallback: str) -> str:
    name = str(sheet_name or "").strip().lower()
    if (
        name.endswith("_inputs")
        or name.endswith("_feed_inputs")
        or name.endswith("_aux")
        or name.endswith("_aux_other")
        or name.endswith("_out_feed")
        or name.endswith("_out_fuel")
    ):
        return "Fuel"
    return str(fallback or "").strip()


def apply_strict_meta(template_path: Path, sheet_name: str, meta: dict[str, object]) -> dict[str, object]:
    updated = dict(meta)
    if "transformation_results_" in template_path.name.lower():
        updated["variable"] = _strict_transformation_variable(sheet_name, str(meta.get("variable") or ""))
        updated["legend_label"] = _strict_transformation_legend(sheet_name, str(meta.get("legend_label") or ""))
    return updated


def build_table_from_export_csv(meta: dict[str, object], export_csv_path: Path) -> pd.DataFrame:
    from codebase import leap_results_workflow  # noqa: PLC0415
    exported_df = leap_results_workflow.parse_exported_results_csv(export_csv_path)
    data_df = pd.DataFrame([list(exported_df.columns)] + exported_df.values.tolist())
    rows = [
        [meta.get("variable", "")],
        [f"Scenario: {meta.get('scenario','')}, Region: {meta.get('region','')}"] ,
        [f"Branch: {meta.get('branch','')}"] ,
        [f"Units: {meta.get('units','')}"] ,
        [""],
    ]
    return pd.DataFrame(rows + data_df.values.tolist())


def _normalize_table_for_compare(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    for col in out.columns:
        out[col] = out[col].map(lambda v: "" if pd.isna(v) else v)
    return out


def is_effectively_empty_results_table(table_df: pd.DataFrame) -> bool:
    """
    Return True when a refilled LEAP table has no meaningful component rows.

    Treat a table as effectively empty when:
    - there are no rows below the header, or
    - only a Total row exists, or
    - component rows exist but all numeric values are zero/blank.
    """
    if table_df.shape[0] < 7 or table_df.shape[1] < 2:
        return True

    body = table_df.iloc[6:, :].copy()
    if body.empty:
        return True

    labels = body.iloc[:, 0].fillna("").astype(str).str.strip()
    nonblank = labels != ""
    if int(nonblank.sum()) == 0:
        return True

    data = body.loc[nonblank].copy()
    data_labels = data.iloc[:, 0].astype(str).str.strip()
    component_mask = data_labels.str.lower() != "total"
    components = data.loc[component_mask]
    if components.empty:
        return True

    numeric = components.iloc[:, 1:].apply(pd.to_numeric, errors="coerce").fillna(0.0)
    return float(numeric.abs().sum().sum()) == 0.0


def _should_enforce_empty_clear(template_path: Path, sheet_name: str) -> bool:
    name = template_path.name.lower()
    sheet = str(sheet_name or "").strip().lower()
    if "transformation_results_" not in name:
        return False
    return (
        sheet.endswith("_inputs")
        or sheet.endswith("_feed_inputs")
        or sheet.endswith("_aux")
        or sheet.endswith("_aux_other")
        or sheet.endswith("_out_feed")
        or sheet.endswith("_out_fuel")
    )


def _is_transformation_input_style_sheet(template_path: Path, sheet_name: str) -> bool:
    """
    Sheets that should be extracted via same-name LEAP favorites only.

    Example: refining_feed_inputs, refining_aux, refining_aux_other
    """
    if "transformation_results_" not in str(template_path.name).lower():
        return False
    sheet = str(sheet_name or "").strip().lower()
    return (
        sheet.endswith("_feed_inputs")
        or sheet.endswith("_aux")
        or sheet.endswith("_aux_other")
    )


def compare_tables(candidate: pd.DataFrame, golden: pd.DataFrame, *, tol: float = 1e-6) -> tuple[str, float, str]:
    cand = _normalize_table_for_compare(candidate)
    gold = _normalize_table_for_compare(golden)
    if cand.shape != gold.shape:
        return "shape_mismatch", float("inf"), f"shape candidate={cand.shape}, golden={gold.shape}"

    max_abs_diff = 0.0
    for r in range(cand.shape[0]):
        for c in range(cand.shape[1]):
            a = cand.iat[r, c]
            b = gold.iat[r, c]
            na = pd.to_numeric(a, errors="coerce")
            nb = pd.to_numeric(b, errors="coerce")
            if not pd.isna(na) and not pd.isna(nb):
                diff = abs(float(na) - float(nb))
                if diff > max_abs_diff:
                    max_abs_diff = diff
                if diff > tol:
                    return "value_mismatch", max_abs_diff, f"first mismatch at r={r+1}, c={c+1}: {a} vs {b}"
            else:
                if str(a).strip() != str(b).strip():
                    return "label_mismatch", max_abs_diff, f"first mismatch at r={r+1}, c={c+1}: {a!r} vs {b!r}"
    return "match", max_abs_diff, ""


def run_replica_extraction(
    *,
    template_paths: Iterable[Path | str],
    output_dir: Path | str,
    mode: str = "replay_csv",
    golden_workbook: Path | str | None = None,
    clear_on_effective_empty: bool = True,
) -> dict[str, object]:
    from codebase import leap_results_workflow  # noqa: PLC0415
    mode = str(mode or "").strip().lower()
    if mode not in {"replay_csv", "live"}:
        raise ValueError("mode must be 'replay_csv' or 'live'")

    output_root = _resolve(output_dir)
    output_root.mkdir(parents=True, exist_ok=True)
    run_results: list[ReplicaSheetResult] = []
    written_workbooks: list[str] = []

    app = None
    if mode == "live":
        app = leap_results_workflow.connect_leap()
        leap_results_workflow.ensure_calculated(app, force=False)

    golden_xl = pd.ExcelFile(_resolve(golden_workbook)) if golden_workbook else None

    for tpl_raw in template_paths:
        template_path = _resolve(tpl_raw)
        wb = load_workbook(template_path)
        out_path = output_root / f"{template_path.stem}.replica.xlsx"

        for idx, sheet_name in enumerate(wb.sheetnames, start=1):
            ws = wb[sheet_name]
            template_meta = leap_results_workflow.parse_template_worksheet(ws)
            strict_meta = apply_strict_meta(template_path, sheet_name, template_meta)

            source_csv = ""
            status = "ok"
            notes = ""
            fresh_df = None

            try:
                if mode == "replay_csv":
                    export_csv_path = leap_results_workflow._sheet_tmp_export_path(template_path, sheet_name, idx)
                    source_csv = str(export_csv_path)
                    if not export_csv_path.exists():
                        raise FileNotFoundError(f"CSV replay source missing: {export_csv_path}")
                    fresh_df = build_table_from_export_csv(strict_meta, export_csv_path)
                else:
                    export_csv_path = leap_results_workflow._sheet_tmp_export_path(template_path, sheet_name, idx)
                    source_csv = str(export_csv_path)
                    input_style = _is_transformation_input_style_sheet(template_path, sheet_name)
                    favorite_activated = False
                    expected_input_type_qualifier = leap_results_workflow._expected_input_type_qualifier_for_sheet(
                        sheet_name
                    )
                    if _is_transformation_input_style_sheet(template_path, sheet_name):
                        resolved_name, resolve_mode, favorite_status = leap_results_workflow.activate_favorite_for_sheet(
                            app,
                            sheet_name,
                        )
                        favorite_activated = bool(resolved_name) and "not found/activated" not in str(favorite_status).lower()
                        if favorite_status:
                            notes = f"{notes}; favorite={favorite_status}; favorite_resolve={resolve_mode}; favorite_name={resolved_name}".strip("; ")
                        if input_style and not favorite_activated:
                            notes = f"{notes}; input_style_favorite_missing_for_sheet={sheet_name}".strip("; ")
                    fresh_df = leap_results_workflow.build_fresh_table_from_export(
                        app,
                        strict_meta,
                        export_csv_path=export_csv_path,
                        context=f"replica/{template_path.name}/{sheet_name}",
                        allow_variable_substitution=False,
                        trust_active_results_view=(input_style and favorite_activated),
                        expected_input_type_qualifier=expected_input_type_qualifier,
                    )
            except Exception as exc:  # noqa: BLE001
                status = "error"
                notes = str(exc)
                fresh_df = pd.DataFrame([[f"ERROR: {exc}"]])

            if (
                status == "ok"
                and bool(clear_on_effective_empty)
                and _should_enforce_empty_clear(template_path, sheet_name)
                and is_effectively_empty_results_table(fresh_df)
            ):
                fresh_df = leap_results_workflow._build_no_results_table(strict_meta)
                notes = f"{notes}; effective_empty_cleared".strip("; ")

            golden_status = "not_checked"
            max_abs_diff = 0.0
            if golden_xl and sheet_name in golden_xl.sheet_names and status == "ok":
                golden_df = pd.read_excel(_resolve(golden_workbook), sheet_name=sheet_name, header=None)
                golden_status, max_abs_diff, compare_notes = compare_tables(fresh_df, golden_df)
                if compare_notes:
                    notes = f"{notes}; {compare_notes}".strip("; ")

            leap_results_workflow.write_table_values_preserve_format(ws, fresh_df)
            run_results.append(
                ReplicaSheetResult(
                    template=template_path.name,
                    sheet=sheet_name,
                    scenario=str(strict_meta.get("scenario") or ""),
                    branch=str(strict_meta.get("branch") or ""),
                    variable_template=str(template_meta.get("variable") or ""),
                    variable_used=str(strict_meta.get("variable") or ""),
                    legend_template=str(template_meta.get("legend_label") or ""),
                    legend_used=str(strict_meta.get("legend_label") or ""),
                    mode=mode,
                    source_csv=source_csv,
                    status=status,
                    notes=notes,
                    golden_status=golden_status,
                    max_abs_diff=float(max_abs_diff),
                )
            )

        wb.save(out_path)
        written_workbooks.append(str(out_path))

    report_path = output_root / "replica_extraction_report.csv"
    pd.DataFrame([r.__dict__ for r in run_results]).to_csv(report_path, index=False)

    summary = {
        "mode": mode,
        "written_workbooks": written_workbooks,
        "report_path": str(report_path),
        "total_sheets": len(run_results),
        "error_sheets": sum(1 for r in run_results if r.status != "ok"),
        "golden_matches": sum(1 for r in run_results if r.golden_status == "match"),
        "golden_mismatches": sum(1 for r in run_results if r.golden_status not in {"match", "not_checked"}),
    }
    return summary
