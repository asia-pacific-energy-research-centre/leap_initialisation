#%%
"""
Clone a LEAP results template workbook for multiple economies and scenarios.

Usage pattern (edit the CONSTANTS section below):
- Point TEMPLATE_PATH to an existing template workbook (e.g., the Reference USA file).
- Fill SCENARIOS (list of scenario names) and ECONOMIES (list of tuples with a
  short code and full region name).
- Run the bottom block to create copies in OUTPUT_DIR, with filenames of the form
  industry_results_20_<ECON>_<SCENARIO>.xlsx.

Notes:
- Only the metadata row "Scenario: X, Region: Y" (row 2, col A) is modified.
- Sheet names and all data/layout are preserved as-is.
"""
from __future__ import annotations

import os
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

# Stable constants
REPO_ROOT = Path(__file__).resolve().parents[2]

# Frequently changed constants
TEMPLATE_PATH = Path("data/leap results tables/demand_others_results_20_USA_Reference.xlsx")
OUTPUT_DIR = Path("data/leap results tables")
SCENARIOS: list[str] = ["Reference", "Target"]
# List of (economy_code, region_display_name)
ECONOMIES: list[tuple[str, str]] = [
    ("USA", "United States"),
    # Add more, e.g., ("CAN", "Canada")
]
FILENAME_TEMPLATE = "demand_others_results_20_{econ}_{scenario}.xlsx"


def ensure_repo_root() -> None:
    """Ensure cwd is the repo root for relative paths and return it."""
    cwd = Path.cwd()
    if cwd != REPO_ROOT:
        os.chdir(REPO_ROOT)
    return REPO_ROOT


def update_scenario_region_cell(text: str, scenario: str, region: str) -> str:
    """Return updated 'Scenario: X, Region: Y' text."""
    parts = text.split(",")
    out_parts: list[str] = []
    for part in parts:
        part = part.strip()
        if part.lower().startswith("scenario:"):
            out_parts.append(f"Scenario: {scenario}")
        elif part.lower().startswith("region:"):
            out_parts.append(f"Region: {region}")
        else:
            out_parts.append(part)
    return ", ".join(out_parts)


def clone_workbook(template_path: Path, out_path: Path, scenario: str, region: str) -> None:
    """Load template, update scenario/region in each sheet, save to out_path."""
    wb = load_workbook(template_path)
    for ws in wb.worksheets:
        cell = ws.cell(row=2, column=1)
        if cell.value and isinstance(cell.value, str) and "Scenario:" in cell.value:
            cell.value = update_scenario_region_cell(cell.value, scenario, region)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def run_bulk_clone(
    template_path: Path,
    output_dir: Path,
    scenarios: Iterable[str],
    economies: Iterable[tuple[str, str]],
    filename_template: str,
) -> list[Path]:
    """Create all combinations of scenario x economy; return list of created paths."""
    created: list[Path] = []
    for econ_code, region_name in economies:
        for scenario in scenarios:
            out_name = filename_template.format(econ=econ_code, scenario=scenario)
            out_path = output_dir / out_name
            clone_workbook(template_path, out_path, scenario, region_name)
            created.append(out_path)
    return created


#%%
# Bottom run block
if __name__ == "__main__":
    ensure_repo_root()
    files = run_bulk_clone(
        template_path=TEMPLATE_PATH,
        output_dir=OUTPUT_DIR,
        scenarios=SCENARIOS,
        economies=ECONOMIES,
        filename_template=FILENAME_TEMPLATE,
    )
    print("Created files:")
    for f in files:
        print(f" - {f}")

#%%
