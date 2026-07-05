from __future__ import annotations

import re
import shutil
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

import openpyxl
import pandas as pd

from codebase.utilities.master_config import (
    OUTLOOK_MAPPINGS_MASTER_PATH,
    config_table_exists,
    read_config_table,
)


DEFAULT_TEMPLATE_PATH = Path("data/detailed balance table output example.xlsx")
DEFAULT_ESTOPATH = Path("data/00APEC_2024_low_with_subtotals.csv")
DEFAULT_CODEBOOK_PATH = OUTLOOK_MAPPINGS_MASTER_PATH

DEFAULT_SHEET_NAME = "Energy Balance"
DEFAULT_ECONOMY = "20USA"
DEFAULT_YEAR = 2022
DEFAULT_SCENARIO = "Reference"
DEFAULT_UNITS = "Billion Gigajoule"
DEFAULT_AREA_LABEL = "USA (ESTO reconstructed)"

MAX_SCAN_COLS = 400
MAX_SCAN_ROWS = 3000


def _normalize_text(value: object) -> str:
    text = str(value or "").strip().lower()
    if not text or text in {"nan", "none"}:
        return ""
    text = text.replace("&", "and")
    text = text.replace("_", " ")
    text = text.replace("-", " ")
    text = re.sub(r"[^a-z0-9.]+", " ", text)
    return " ".join(text.split())


def _normalize_economy_key(value: object) -> str:
    text = str(value or "").strip().upper()
    return text.replace("_", "")


def _parse_leading_indent(text: str) -> int:
    return len(text) - len(text.lstrip(" "))


def _extract_name_from_esto_label(label: str) -> str:
    text = str(label or "").strip()
    return re.sub(r"^\d{2}(?:\.\d{2})*\s*", "", text).strip()


def _resolve_path(path: str | Path, repo_root: Path) -> Path:
    if isinstance(path, Path):
        candidate = path
    else:
        normalized = str(path).replace("\\", "/")
        drive_match = re.match(r"^([a-zA-Z]):/(.*)$", normalized)
        if drive_match:
            drive = drive_match.group(1).lower()
            rest = drive_match.group(2)
            return Path(f"/mnt/{drive}/{rest}")
        candidate = Path(normalized)
    if candidate.is_absolute():
        return candidate
    return repo_root / candidate


def _dedupe_preserve_order(values: Iterable[str]) -> list[str]:
    out: list[str] = []
    seen: set[str] = set()
    for value in values:
        token = str(value).strip()
        if not token or token in seen:
            continue
        seen.add(token)
        out.append(token)
    return out


def _expand_alias_key_map(raw: dict[str, str]) -> dict[str, str]:
    out: dict[str, str] = {}
    for key, value in raw.items():
        norm = _normalize_text(key)
        if norm:
            out[norm] = value
    return out


FLOW_ALIASES = _expand_alias_key_map(
    {
        "Total Primary Supply": "07 Total primary energy supply",
        "Total Transformation": "09 Total transformation sector",
        "Total Final Energy Demand": "13 Total final energy consumption",
        "Transmission and Distribution": "10.02 Transmission and distribution losses",
        "Oil Refining": "09.07 Oil refineries",
        "Heat Production": "09.01.03 Heat plants",
        "Hydrogen transformation": "09.12 Non-specified transformation",
        "BKB and PB plants": "09.08.04 BKB/PB plants",
        "NG Liquefaction": "09.06.02.01 Liquefaction",
        "LNG regasification": "09.06.02.02 Regasification",
        "Electricity Generation": "09.01.01 Electricity plants",
        "Industry": "14 Industry sector",
        "Other sector": "16 Other sector",
        "Buildings": "16 Other sector",
        "Services": "16.01 Commercial and public services",
        "Residential": "16.02 Residential",
        "Transport non road": "15 Transport sector",
        "Freight road": "15.02 Road",
        "Passenger road": "15.02 Road",
        "Blast furnaces": "09.08.02 Blast furnaces",
        "Upstream liquids transfers": "08.04 Gas separation",
        "Refinery and blending transfers": "08.03 Products transferred",
        "Transfers unallocated": "08.99 Transfers nonspecified",
        "Liquefaction coal to oil": "09.08.05 Liquefaction (coal to oil)",
        "Charcoal processing": "09.11 Charcoal processing",
        "Non specified transformation": "09.12 Non-specified transformation",
        "Coke ovens": "09.08.01 Coke ovens",
        "Patent fuel plants": "09.08.03 Patent fuel plants",
        "Natural gas blending plants": "09.06.03 Natural gas blending plants",
        "Gas works plants": "09.06.01 Gas works plants",
    }
)

PRODUCT_ALIASES = _expand_alias_key_map(
    {
        "Coal Bituminous": "01.02 Other bituminous coal",
        "Coal Sub bituminous": "01.03 Sub-bituminous coal",
        "Crude Oil and NGL": "06 Crude oil & NGL",
        "Biomass": "15 Solid biomass",
        "Ammonia": "16.09 Other sources",
        "Others": "16 Others",
        "Black liqour": "15.04 Black liqour",
        "Fuelwood and woodwaste": "15.01 Fuelwood & woodwaste",
        "Gas and diesel oil": "07.07 Gas/diesel oil",
        "Municipal solid waste renewable": "16.03 Municipal solid waste (renewable)",
        "Municipal solid waste non renewable": "16.04 Municipal solid waste (non-renewable)",
        "Other biomass": "15.05 Other biomass",
        "Other liquid biofuels": "16.08 Other liquid biofuels",
        "Other sources": "16.09 Other sources",
        "PetProd nonspecified": "07.17 Other products",
        "Paraffin waxes": "07.17 Other products",
        "White spirit SBP": "07.17 Other products",
        "BKB and PB": "02.08 BKB/PB",
    }
)


@dataclass
class HeaderSpec:
    col_idx: int
    label: str
    mapped_products: list[str]
    mapping_status: str


@dataclass
class RowSpec:
    row_idx: int
    raw_label: str
    stripped_label: str
    indent: int
    direct_flows: list[str]
    direct_products: list[str]
    resolved_flows: list[str]
    mode: str
    resolved_from: str


def _load_template_structure(
    workbook_path: Path,
    *,
    sheet_name: str,
) -> tuple[list[tuple[int, str]], list[tuple[int, str]]]:
    wb = openpyxl.load_workbook(workbook_path, read_only=False, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Template sheet {sheet_name!r} not found in {workbook_path}")
    ws = wb[sheet_name]

    header_cells: list[tuple[int, str]] = []
    col = 2
    while col <= MAX_SCAN_COLS:
        value = ws.cell(3, col).value
        text = str(value).strip() if value not in (None, "") else ""
        if not text:
            break
        header_cells.append((col, text))
        col += 1
    if not header_cells:
        raise ValueError("No template fuel headers found on row 3.")

    row_cells: list[tuple[int, str]] = []
    row = 4
    while row <= MAX_SCAN_ROWS:
        value = ws.cell(row, 1).value
        if value not in (None, ""):
            text = str(value)
            if text.strip():
                row_cells.append((row, text))
        row += 1
    if not row_cells:
        raise ValueError("No template data rows found from row 4 downward.")
    return header_cells, row_cells


def _build_name_lookups(
    codebook_path: Path,
    esto_flows: Iterable[str],
    esto_products: Iterable[str],
) -> tuple[dict[str, list[str]], dict[str, list[str]]]:
    flow_lookup: dict[str, list[str]] = {}
    product_lookup: dict[str, list[str]] = {}

    codebook = read_config_table(codebook_path, sheet_name="code_to_name", dtype=str).fillna("")
    for _, row in codebook.iterrows():
        name = _normalize_text(row.get("name", ""))
        if not name:
            continue
        esto_label = str(row.get("esto_label", "")).strip()
        esto_column = _normalize_text(row.get("esto_column", ""))
        if not esto_label:
            continue
        if esto_column in {"flows", "flow"}:
            flow_lookup.setdefault(name, [])
            flow_lookup[name].append(esto_label)
        elif esto_column in {"products", "product"}:
            product_lookup.setdefault(name, [])
            product_lookup[name].append(esto_label)

    leap_names = read_config_table(codebook_path, sheet_name="ESTO_LEAP_names", dtype=str).fillna("")
    for _, row in leap_names.iterrows():
        if _normalize_text(row.get("category", "")) != "products":
            continue
        leap_name = _normalize_text(row.get("leap_name", ""))
        esto_product = str(row.get("original_label", "")).strip()
        if leap_name and esto_product:
            product_lookup.setdefault(leap_name, [])
            product_lookup[leap_name].append(esto_product)

    # Also map raw ESTO labels and ESTO labels without numeric prefix.
    for flow in esto_flows:
        flow_text = str(flow).strip()
        if not flow_text:
            continue
        for key in {_normalize_text(flow_text), _normalize_text(_extract_name_from_esto_label(flow_text))}:
            if key:
                flow_lookup.setdefault(key, [])
                flow_lookup[key].append(flow_text)
    for product in esto_products:
        product_text = str(product).strip()
        if not product_text:
            continue
        for key in {
            _normalize_text(product_text),
            _normalize_text(_extract_name_from_esto_label(product_text)),
        }:
            if key:
                product_lookup.setdefault(key, [])
                product_lookup[key].append(product_text)

    for key, vals in list(flow_lookup.items()):
        flow_lookup[key] = _dedupe_preserve_order(vals)
    for key, vals in list(product_lookup.items()):
        product_lookup[key] = _dedupe_preserve_order(vals)

    return flow_lookup, product_lookup


def _apply_aliases(
    label: str,
    lookup: dict[str, list[str]],
    aliases: dict[str, str],
) -> list[str]:
    norm = _normalize_text(label)
    if not norm:
        return []
    direct = lookup.get(norm, [])
    if direct:
        return _dedupe_preserve_order(direct)
    alias_target = aliases.get(norm, "")
    if not alias_target:
        return []
    alias_norm = _normalize_text(alias_target)
    if alias_norm in lookup:
        return _dedupe_preserve_order(lookup[alias_norm])
    return [alias_target]


def _build_headers(
    header_cells: list[tuple[int, str]],
    *,
    product_lookup: dict[str, list[str]],
) -> list[HeaderSpec]:
    headers: list[HeaderSpec] = []
    for col_idx, label in header_cells:
        mapped_products = _apply_aliases(label, product_lookup, PRODUCT_ALIASES)
        status = "mapped" if mapped_products else "unmapped"
        headers.append(
            HeaderSpec(
                col_idx=col_idx,
                label=label,
                mapped_products=mapped_products,
                mapping_status=status,
            )
        )
    return headers


def _build_rows(
    row_cells: list[tuple[int, str]],
    *,
    flow_lookup: dict[str, list[str]],
    product_lookup: dict[str, list[str]],
) -> list[RowSpec]:
    rows: list[RowSpec] = []
    flow_stack: list[tuple[int, list[str]]] = []

    for row_idx, raw_label in row_cells:
        stripped = raw_label.strip()
        indent = _parse_leading_indent(raw_label)

        while flow_stack and flow_stack[-1][0] >= indent:
            flow_stack.pop()

        inherited_flows = list(flow_stack[-1][1]) if flow_stack else []
        direct_flows = _apply_aliases(stripped, flow_lookup, FLOW_ALIASES)
        direct_products = _apply_aliases(stripped, product_lookup, PRODUCT_ALIASES)

        mode = "unmapped"
        resolved_from = ""
        resolved_flows: list[str] = []
        if direct_flows:
            mode = "flow_row"
            resolved_from = "direct_flow"
            resolved_flows = direct_flows
        elif direct_products and inherited_flows:
            mode = "fuel_leaf_row"
            resolved_from = "inherited_flow"
            resolved_flows = inherited_flows

        if direct_flows:
            flow_stack.append((indent, direct_flows))

        rows.append(
            RowSpec(
                row_idx=row_idx,
                raw_label=raw_label,
                stripped_label=stripped,
                indent=indent,
                direct_flows=direct_flows,
                direct_products=direct_products,
                resolved_flows=resolved_flows,
                mode=mode,
                resolved_from=resolved_from,
            )
        )
    return rows


def _coerce_bool(series: pd.Series) -> pd.Series:
    return (
        series.fillna(False)
        .astype(str)
        .str.strip()
        .str.lower()
        .isin({"1", "true", "yes", "y", "t"})
    )


def _build_esto_pair_values(
    esto_data_path: Path,
    *,
    economy: str,
    year: int,
    include_subtotals: bool = True,
) -> tuple[dict[tuple[str, str], float], pd.DataFrame]:
    year_col = str(year)
    usecols = ["economy", "flows", "products", year_col]
    probe = read_config_table(esto_data_path, nrows=1)
    if "is_subtotal" in probe.columns:
        usecols.append("is_subtotal")
    df = read_config_table(esto_data_path, usecols=usecols, dtype={year_col: "float64"})
    df = df.fillna("")
    df["economy_norm"] = df["economy"].map(_normalize_economy_key)
    target_key = _normalize_economy_key(economy)
    df = df[df["economy_norm"] == target_key].copy()
    if df.empty:
        raise ValueError(f"No ESTO rows found for economy {economy!r}.")

    if not include_subtotals and "is_subtotal" in df.columns:
        df = df[~_coerce_bool(df["is_subtotal"])].copy()
    df["flows"] = df["flows"].astype(str).str.strip()
    df["products"] = df["products"].astype(str).str.strip()
    df[year_col] = pd.to_numeric(df[year_col], errors="coerce").fillna(0.0)
    df = df[df["flows"].ne("") & df["products"].ne("")]
    grouped = (
        df.groupby(["flows", "products"], dropna=False, as_index=False)[year_col]
        .sum()
        .rename(columns={year_col: "value"})
    )
    pair_values = {
        (str(row["flows"]).strip(), str(row["products"]).strip()): float(row["value"])
        for _, row in grouped.iterrows()
    }
    return pair_values, grouped


def _render_values(
    rows: list[RowSpec],
    headers: list[HeaderSpec],
    pair_values: dict[tuple[str, str], float],
) -> dict[tuple[int, int], float]:
    value_map: dict[tuple[int, int], float] = {}
    for row in rows:
        for header in headers:
            value = 0.0
            if row.mode == "flow_row":
                if row.resolved_flows and header.mapped_products:
                    for flow in row.resolved_flows:
                        for product in header.mapped_products:
                            value += pair_values.get((flow, product), 0.0)
            elif row.mode == "fuel_leaf_row":
                if row.resolved_flows and row.direct_products and header.mapped_products:
                    match_products = set(row.direct_products).intersection(header.mapped_products)
                    if match_products:
                        for flow in row.resolved_flows:
                            for product in match_products:
                                value += pair_values.get((flow, product), 0.0)
            value_map[(row.row_idx, header.col_idx)] = float(value)
    return value_map


def _write_output_workbook(
    *,
    template_workbook_path: Path,
    output_workbook_path: Path,
    sheet_name: str,
    area_label: str,
    scenario_label: str,
    year: int,
    units: str,
    value_map: dict[tuple[int, int], float],
) -> None:
    output_workbook_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copyfile(template_workbook_path, output_workbook_path)
    wb = openpyxl.load_workbook(output_workbook_path)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet {sheet_name!r} missing in output workbook.")
    ws = wb[sheet_name]
    ws.cell(1, 1).value = f'Energy Balance for Area "{area_label}"'
    ws.cell(2, 1).value = f"Scenario: {scenario_label}, Year: {int(year)}, Units: {units}"
    for (row_idx, col_idx), value in value_map.items():
        ws.cell(row_idx, col_idx).value = float(value)
    wb.save(output_workbook_path)


def build_detailed_balance_from_esto(
    *,
    template_workbook_path: str | Path = DEFAULT_TEMPLATE_PATH,
    esto_data_path: str | Path = DEFAULT_ESTOPATH,
    codebook_path: str | Path = DEFAULT_CODEBOOK_PATH,
    output_workbook_path: str | Path,
    output_dir: str | Path,
    sheet_name: str = DEFAULT_SHEET_NAME,
    economy: str = DEFAULT_ECONOMY,
    year: int = DEFAULT_YEAR,
    scenario_label: str = DEFAULT_SCENARIO,
    units: str = DEFAULT_UNITS,
    area_label: str = DEFAULT_AREA_LABEL,
    include_subtotals: bool = True,
) -> dict[str, object]:
    repo_root = Path(__file__).resolve().parents[2]
    template_path = _resolve_path(template_workbook_path, repo_root)
    esto_path = _resolve_path(esto_data_path, repo_root)
    codebook = _resolve_path(codebook_path, repo_root)
    output_path = _resolve_path(output_workbook_path, repo_root)
    diagnostics_dir = _resolve_path(output_dir, repo_root)
    diagnostics_dir.mkdir(parents=True, exist_ok=True)

    for path in [template_path, esto_path]:
        if not path.exists():
            raise FileNotFoundError(f"Missing required input: {path}")
    if not config_table_exists(codebook):
        raise FileNotFoundError(f"Missing required input: {codebook}")

    pair_values, grouped = _build_esto_pair_values(
        esto_path,
        economy=economy,
        year=int(year),
        include_subtotals=include_subtotals,
    )
    esto_flows = grouped["flows"].astype(str).unique().tolist()
    esto_products = grouped["products"].astype(str).unique().tolist()

    header_cells, row_cells = _load_template_structure(template_path, sheet_name=sheet_name)
    flow_lookup, product_lookup = _build_name_lookups(codebook, esto_flows, esto_products)
    headers = _build_headers(header_cells, product_lookup=product_lookup)
    rows = _build_rows(row_cells, flow_lookup=flow_lookup, product_lookup=product_lookup)
    value_map = _render_values(rows, headers, pair_values)
    _write_output_workbook(
        template_workbook_path=template_path,
        output_workbook_path=output_path,
        sheet_name=sheet_name,
        area_label=area_label,
        scenario_label=scenario_label,
        year=int(year),
        units=units,
        value_map=value_map,
    )

    header_diag = pd.DataFrame(
        [
            {
                "col_idx": h.col_idx,
                "header_label": h.label,
                "mapped_products": "|".join(h.mapped_products),
                "mapping_status": h.mapping_status,
            }
            for h in headers
        ]
    )
    row_diag = pd.DataFrame(
        [
            {
                "row_idx": r.row_idx,
                "raw_label": r.raw_label,
                "stripped_label": r.stripped_label,
                "indent": r.indent,
                "direct_flows": "|".join(r.direct_flows),
                "direct_products": "|".join(r.direct_products),
                "resolved_flows": "|".join(r.resolved_flows),
                "mode": r.mode,
                "resolved_from": r.resolved_from,
            }
            for r in rows
        ]
    )
    summary = pd.DataFrame(
        [
            {
                "economy": economy,
                "year": int(year),
                "template_sheet": sheet_name,
                "row_count": len(rows),
                "header_count": len(headers),
                "mapped_headers": int((header_diag["mapping_status"] == "mapped").sum()),
                "flow_rows": int((row_diag["mode"] == "flow_row").sum()),
                "fuel_leaf_rows": int((row_diag["mode"] == "fuel_leaf_row").sum()),
                "unmapped_rows": int((row_diag["mode"] == "unmapped").sum()),
                "output_workbook": str(output_path),
            }
        ]
    )

    header_diag_path = diagnostics_dir / "detailed_balance_header_mapping.csv"
    row_diag_path = diagnostics_dir / "detailed_balance_row_mapping.csv"
    summary_path = diagnostics_dir / "detailed_balance_summary.csv"
    values_preview_path = diagnostics_dir / "detailed_balance_nonzero_pairs.csv"

    header_diag.to_csv(header_diag_path, index=False)
    row_diag.to_csv(row_diag_path, index=False)
    summary.to_csv(summary_path, index=False)
    grouped[grouped["value"] != 0].sort_values(["flows", "products"]).to_csv(values_preview_path, index=False)

    return {
        "output_workbook": str(output_path),
        "header_mapping_csv": str(header_diag_path),
        "row_mapping_csv": str(row_diag_path),
        "summary_csv": str(summary_path),
        "nonzero_pairs_csv": str(values_preview_path),
        "summary": summary.iloc[0].to_dict(),
    }
