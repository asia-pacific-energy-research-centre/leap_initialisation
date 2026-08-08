from __future__ import annotations

import os
import re
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Iterable, Sequence

import pandas as pd

from codebase.configuration import workflow_config as workflow_cfg

REPO_ROOT = Path(__file__).resolve().parents[2]
DEFAULT_BALANCE_EXPORTS_ROOT = REPO_ROOT / "data" / "leap balances exports"

BALANCE_EXPORT_FILENAME_PATTERN = re.compile(
    r"^(?:"
    r"full model output all years (?P<date_first>\d{4,8}) (?P<scenario_second>[A-Za-z]+)"
    r"|"
    r"(?P<scenario_first>REF|TGT|Reference|Target) (?P<date_second>\d{4,8})"
    r")"
    r"(?:\s[^.]*)?\.xlsx$",
    re.IGNORECASE,
)


_ECONOMY_AREA_ALIASES = {
    "01_AUS": ("aus", "australia"),
    "02_BD": ("bd", "brunei", "brunei darussalam"),
    "03_CDA": ("cda", "canada"),
    "04_CHL": ("chl", "chile"),
    "05_PRC": ("prc", "china", "people's republic of china"),
    "06_HKC": ("hkc", "hong kong", "hong kong china"),
    "07_INA": ("ina", "indonesia"),
    "08_JPN": ("jpn", "japan"),
    "09_ROK": ("rok", "korea", "republic of korea", "south korea"),
    "10_MAS": ("mas", "malaysia"),
    "11_MEX": ("mex", "mexico"),
    "12_NZ": ("nz", "new zealand"),
    "13_PNG": ("png", "papua new guinea"),
    "14_PE": ("pe", "peru"),
    "15_PHL": ("phl", "philippines", "the philippines"),
    "16_RUS": ("rus", "russia", "russian federation"),
    "17_SGP": ("sgp", "singapore"),
    "18_CT": ("ct", "chinese taipei"),
    "19_THA": ("tha", "thailand"),
    "20_USA": ("usa", "united states", "united states of america"),
    "21_VN": ("vn", "vietnam", "viet nam"),
}

# LEAP writes the area name into the title row of every balance sheet.
_ECONOMY_AREA_TITLE_PATTERN = re.compile(
    r"Energy Balance for Area\s*[\"'](?P<area>.+?)[\"']",
    re.IGNORECASE,
)


SCENARIO_CODE_ALIASES = {
    "ref": "REF",
    "reference": "REF",
    "tgt": "TGT",
    "target": "TGT",
}


@dataclass(frozen=True)
class BalanceExportWorkbook:
    path: Path
    economy: str
    scenario_code: str
    date_id: str
    parsed_date: date | None


@dataclass(frozen=True)
class BalanceExportDetailInspection:
    """Minimum LEAP Energy Balance detail that can be proven from a workbook."""

    path: Path
    detected_level_label: str
    has_level2_detail: bool
    sample_indented_label: str | None


@dataclass(frozen=True)
class BalanceExportSheet:
    """One scenario/year balance sheet available inside an export workbook."""

    path: Path
    sheet_name: str
    scenario: str
    scenario_code: str
    year: int
    units: str


@dataclass(frozen=True)
class BalanceExportIdentity:
    """The economy, scenario, and years a LEAP export declares about itself."""

    path: Path
    economy: str
    area_name: str
    scenario: str
    scenario_code: str
    years: tuple[int, ...]
    units: str


BALANCE_LABEL_ALIASES = {
    "total transformation sector": "total transformation",
}


def normalize_balance_scenario_code(scenario: str) -> str:
    """Return the balance-export filename scenario token."""
    text = str(scenario).strip()
    if not text:
        raise ValueError("Balance-export scenario cannot be blank.")
    return SCENARIO_CODE_ALIASES.get(text.lower(), text.upper())


def normalize_balance_label(value: object) -> str:
    """Return a compact lowercase key for LEAP balance row/fuel matching."""
    normalized = " ".join(str(value or "").strip().lower().split())
    return BALANCE_LABEL_ALIASES.get(normalized, normalized)


def inspect_balance_export_detail(
    workbook_path: Path | str,
) -> BalanceExportDetailInspection:
    """Distinguish a Level 1 export from an export with Level 2+ detail.

    LEAP writes Level 2 child rows with leading spaces in column A. Higher
    export settings do not include reliable metadata that distinguishes Levels
    2-5, so the strongest honest result from the workbook is ``Level 2+``.
    """
    path = _resolve_path(workbook_path)
    if not path.exists():
        raise FileNotFoundError(f"LEAP balance-export workbook does not exist: {path}")

    try:
        from openpyxl import load_workbook

        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError(
            f"Could not inspect LEAP balance-export detail in {path}: {exc}"
        ) from exc

    sample_indented_label: str | None = None
    try:
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows(min_row=4, max_col=1, values_only=True):
                value = row[0]
                if (
                    isinstance(value, str)
                    and value.strip()
                    and value != value.lstrip(" ")
                ):
                    sample_indented_label = value.strip()
                    break
            if sample_indented_label is not None:
                break
    except Exception as exc:
        raise ValueError(
            f"Could not inspect LEAP balance-export detail rows in {path}: {exc}"
        ) from exc
    finally:
        workbook.close()

    has_level2_detail = sample_indented_label is not None
    return BalanceExportDetailInspection(
        path=path,
        detected_level_label="Level 2+" if has_level2_detail else "Level 1",
        has_level2_detail=has_level2_detail,
        sample_indented_label=sample_indented_label,
    )


def require_level2_balance_export_detail(
    workbook_paths: Sequence[Path | str],
) -> list[BalanceExportDetailInspection]:
    """Inspect workbooks and reject any that cannot prove Level 2 detail."""
    inspections = [inspect_balance_export_detail(path) for path in workbook_paths]
    insufficient = [
        inspection for inspection in inspections if not inspection.has_level2_detail
    ]
    if insufficient:
        paths = ", ".join(str(inspection.path) for inspection in insufficient)
        raise ValueError(
            "LEAP Energy Balance export workbook(s) were detected as Level 1 "
            "(no indented branch rows). Re-export with at least Level 2 detail "
            f"before running balance diagnostics or results_update: {paths}"
        )
    return inspections


# LEAP's own-use/loss balance rows are exported as positive quantities, but they
# represent energy consumed out of the demand side of the balance (not supply),
# so ESTO/NINTH and the balance tables represent the same category as negative.
#
# This module's own-use/loss handling is intentionally narrow:
#   - `load_leap_balance_activity_table` below negates these rows via
#     `is_leap_balance_own_use_or_loss_row` because its only consumer (the
#     own-use/loss proxy workflow's activity-driver path) needs a value that's
#     directly comparable to ESTO/NINTH's signed convention.
#   - LEAP-native arithmetic elsewhere in this repo (`supply_conservation.py`,
#     `supply_reconciliation_tables.py`) intentionally treats own-use/loss
#     `demand_value` as a *positive* consumption quantity, and must keep doing
#     so — those modules sum it unconditionally (no `abs()`), so sign-flipping
#     it at the parse source would throw off `closure_residual`/
#     `adjusted_balance` and produce false conservation-violation errors. Do
#     not extend the negation below to that path.
#   - The dashboard's ESTO/NINTH-comparable pipeline is a fully separate
#     codebase (`leap_mappings`) that reads the same raw balance export files
#     independently; its own sign handling lives in
#     `leap_mappings/codebase/mapping_tools/convert_leap_results_to_esto.py`
#     and has no dependency on this module.
LEAP_BALANCE_LOSS_AND_OWN_USE_ROWS = (
    "Other loss and own use",
    "Coal mines",
    "Electricity CHP and heat plants",
    "Liquefaction and regasification plants",
    "Non specified own uses",
    "Non-specified own uses",
    "Oil and gas extraction",
    "Oil refineries",
    "Pump storage plants",
    "Transmission and distribution loss",
)
LEAP_BALANCE_LOSS_AND_OWN_USE_ROW_KEYS = frozenset(
    normalize_balance_label(row) for row in LEAP_BALANCE_LOSS_AND_OWN_USE_ROWS
)


def is_leap_balance_own_use_or_loss_row(row_label: object) -> bool:
    """Return True when a LEAP balance row is one of the own-use/loss rows above."""
    return normalize_balance_label(row_label) in LEAP_BALANCE_LOSS_AND_OWN_USE_ROW_KEYS


def _resolve_path(path: Path | str) -> Path:
    """Resolve repo-relative paths while leaving absolute paths unchanged."""
    raw = str(path).replace("\\", "/")
    drive_match = re.match(r"^([a-zA-Z]):/(.*)$", raw)
    if drive_match:
        drive = drive_match.group(1).lower()
        rest = drive_match.group(2)
        if os.name == "nt":
            return Path(f"{drive.upper()}:/{rest}")
        return Path(f"/mnt/{drive}/{rest}")
    candidate = Path(raw)
    return candidate if candidate.is_absolute() else (REPO_ROOT / candidate)


def _parse_balance_export_date_id(
    date_id: str,
    *,
    fallback_year: int | None = None,
) -> date | None:
    """Parse compact workbook date ids such as 2907, 492026, or 4212026."""
    token = str(date_id).strip()
    if not token.isdigit():
        return None

    if len(token) == 4 and fallback_year is not None:
        try:
            return date(int(fallback_year), int(token[2:4]), int(token[:2]))
        except ValueError:
            return None

    if len(token) == 8:
        for year, month, day in (
            (token[:4], token[4:6], token[6:8]),
            (token[4:8], token[:2], token[2:4]),
            (token[4:8], token[2:4], token[:2]),
        ):
            try:
                return date(int(year), int(month), int(day))
            except ValueError:
                continue

    if len(token) in {6, 7}:
        year_text = token[-4:]
        month_day = token[:-4]
        month_day_splits: list[tuple[str, str]] = []
        if len(month_day) >= 3 and month_day[:2] in {"10", "11", "12"}:
            month_day_splits.append((month_day[:2], month_day[2:]))
        month_day_splits.append((month_day[:1], month_day[1:]))
        if len(month_day) == 4:
            month_day_splits.append((month_day[:2], month_day[2:]))
        for month_text, day_text in month_day_splits:
            try:
                return date(int(year_text), int(month_text), int(day_text))
            except ValueError:
                continue

    return None


def _balance_export_filename_parts(path: Path) -> tuple[str, str] | None:
    """Return ``(date_id, scenario_code)`` for either maintained filename form."""
    match = BALANCE_EXPORT_FILENAME_PATTERN.match(path.name)
    if not match:
        return None
    date_id = match.group("date_first") or match.group("date_second")
    scenario = match.group("scenario_first") or match.group("scenario_second")
    return date_id, normalize_balance_scenario_code(scenario)


def _inspect_balance_export_identity(
    path: Path,
    *,
    economy: str,
    scenario_code: str,
) -> tuple[str, str] | None:
    """Read economy/scenario identity from workbook headers for nonstandard names."""
    from openpyxl import load_workbook

    aliases = _ECONOMY_AREA_ALIASES.get(economy, (economy.rsplit("_", 1)[-1],))
    alias_pattern = re.compile(
        r"(?:^|[^a-z])(?:" + "|".join(re.escape(alias) for alias in aliases) + r")(?:[^a-z]|$)",
        re.IGNORECASE,
    )
    area_pattern = re.compile(
        r"Energy Balance for Area\s*[\"'](?P<area>.+?)[\"']",
        re.IGNORECASE,
    )
    scenario_pattern = re.compile(
        r"Scenario:\s*(?P<scenario>[^,]+)",
        re.IGNORECASE,
    )

    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return None
    try:
        area_matches: list[str] = []
        scenario_matches: set[str] = set()
        for sheet in workbook.worksheets:
            for row_number in range(1, min(sheet.max_row, 3) + 1):
                values = [sheet.cell(row_number, column).value for column in range(1, 5)]
                text = " ".join(str(value).strip() for value in values if value is not None)
                area_match = area_pattern.search(text)
                if area_match:
                    area_matches.append(area_match.group("area"))
                scenario_match = scenario_pattern.search(text)
                if scenario_match:
                    scenario_matches.add(normalize_balance_scenario_code(scenario_match.group("scenario")))
            if area_matches and scenario_matches:
                break
    finally:
        workbook.close()

    if not area_matches or not any(alias_pattern.search(area) for area in area_matches):
        return None
    if scenario_code not in scenario_matches:
        return None

    date_id = next(
        (match.group(0) for match in re.finditer(r"\d{4,8}", path.stem)),
        str(int(path.stat().st_mtime)),
    )
    return date_id, scenario_code


def _iter_balance_export_workbooks(
    export_dir: Path,
    *,
    economy: str,
    scenario_code: str,
) -> Iterable[BalanceExportWorkbook]:
    if not export_dir.exists():
        return
    for path in export_dir.glob("*.xlsx"):
        if path.name.startswith("~$"):
            continue
        parts = _balance_export_filename_parts(path)
        if parts is None:
            parts = _inspect_balance_export_identity(
                path,
                economy=economy,
                scenario_code=scenario_code,
            )
            if parts is None:
                continue
        date_id, candidate_scenario_code = parts
        if candidate_scenario_code != scenario_code:
            continue
        yield BalanceExportWorkbook(
            path=path,
            economy=economy,
            scenario_code=scenario_code,
            date_id=date_id,
            parsed_date=_parse_balance_export_date_id(
                date_id,
                fallback_year=date.fromtimestamp(path.stat().st_mtime).year,
            ),
        )


def resolve_balance_export_workbook(
    *,
    economy: str,
    scenario: str,
    date_id: str | None = None,
    exports_root: Path | str = DEFAULT_BALANCE_EXPORTS_ROOT,
) -> Path:
    """Resolve a LEAP balance-export workbook by economy, scenario, and optional date id."""
    economy_text = str(economy).strip()
    if not economy_text:
        raise ValueError("Balance-export economy cannot be blank.")
    scenario_code = normalize_balance_scenario_code(scenario)
    export_dir = _resolve_path(exports_root) / economy_text
    candidates = list(
        _iter_balance_export_workbooks(
            export_dir,
            economy=economy_text,
            scenario_code=scenario_code,
        )
    )

    if date_id is not None:
        date_text = str(date_id).strip()
        candidates = [candidate for candidate in candidates if candidate.date_id == date_text]
        if not candidates:
            raise FileNotFoundError(
                "No LEAP balance-export workbook matched "
                f"economy={economy_text!r}, scenario={scenario_code!r}, date_id={date_text!r} "
                f"under {export_dir}."
            )
        if len(candidates) > 1:
            paths = "\n".join(
                f"- {candidate.path}"
                for candidate in sorted(candidates, key=lambda item: item.path.name)
            )
            raise ValueError(
                "Multiple LEAP balance-export workbooks matched "
                f"economy={economy_text!r}, scenario={scenario_code!r}, date_id={date_text!r}:\n{paths}"
            )
        return candidates[0].path

    if not candidates:
        raise FileNotFoundError(
            "No LEAP balance-export workbook matched "
            f"economy={economy_text!r}, scenario={scenario_code!r} under {export_dir}."
        )

    sortable = [
        candidate
        for candidate in candidates
        if candidate.parsed_date is not None
    ]
    if sortable:
        latest_date = max(candidate.parsed_date for candidate in sortable)
        latest = [candidate for candidate in sortable if candidate.parsed_date == latest_date]
    else:
        latest = candidates

    if len(latest) > 1:
        paths = "\n".join(
            f"- {candidate.path}"
            for candidate in sorted(latest, key=lambda item: item.path.name)
        )
        raise ValueError(
            "Multiple LEAP balance-export workbooks matched the latest date for "
            f"economy={economy_text!r}, scenario={scenario_code!r}. Set date_id explicitly.\n{paths}"
        )

    return latest[0].path


def list_balance_export_sheets(
    workbook_path: Path | str,
) -> list[BalanceExportSheet]:
    """Read scenario/year/unit metadata for every balance sheet in a workbook."""
    path = _resolve_path(workbook_path)
    if not path.exists():
        raise FileNotFoundError(f"LEAP balance-export workbook does not exist: {path}")

    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheets: list[BalanceExportSheet] = []
    try:
        for sheet in workbook.worksheets:
            metadata = str(sheet.cell(2, 1).value or "").strip()
            scenario_match = re.search(
                r"Scenario:\s*([^,]+)",
                metadata,
                flags=re.IGNORECASE,
            )
            year_match = re.search(r"Year:\s*(\d{4})", metadata, flags=re.IGNORECASE)
            units_match = re.search(r"Units:\s*(.+)$", metadata, flags=re.IGNORECASE)
            if not scenario_match or not year_match or not units_match:
                continue
            scenario_code = normalize_balance_scenario_code(
                scenario_match.group(1).strip()
            )
            sheets.append(
                BalanceExportSheet(
                    path=path,
                    sheet_name=str(sheet.title),
                    scenario="Reference" if scenario_code == "REF" else "Target",
                    scenario_code=scenario_code,
                    year=int(year_match.group(1)),
                    units=units_match.group(1).strip().rstrip("."),
                )
            )
    finally:
        workbook.close()
    if not sheets:
        raise ValueError(
            "No sheets declared Scenario, Year, and Units metadata in "
            f"{path}."
        )
    return sheets


def read_balance_export_area_name(workbook_path: Path | str) -> str:
    """Return the LEAP area name declared in an export's sheet titles."""
    path = _resolve_path(workbook_path)
    if not path.exists():
        raise FileNotFoundError(f"LEAP balance-export workbook does not exist: {path}")

    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        for sheet in workbook.worksheets:
            for row_number in range(1, min(sheet.max_row, 3) + 1):
                values = [sheet.cell(row_number, column).value for column in range(1, 5)]
                text = " ".join(
                    str(value).strip() for value in values if value is not None
                )
                match = _ECONOMY_AREA_TITLE_PATTERN.search(text)
                if match:
                    return match.group("area").strip()
    finally:
        workbook.close()
    return ""


def economy_from_balance_export_area(area_name: str) -> str:
    """Return the APEC economy code a LEAP area name identifies, or "" if unclear.

    The name a modeller gives an area is free text, but in practice it opens
    with the economy's short code ("aus clean slate 29_07"). A leading alias
    therefore wins outright, and only when no alias leads the name are
    whole-word matches elsewhere considered — so a short alias such as "pe" or
    "ct" cannot claim an unrelated name. Either tier must produce exactly one
    economy, so an ambiguous name is reported as unknown rather than guessed at.
    """
    text = " ".join(str(area_name or "").strip().lower().split())
    if not text:
        return ""

    leading: set[str] = set()
    contained: set[str] = set()
    for economy, aliases in _ECONOMY_AREA_ALIASES.items():
        for alias in aliases:
            if re.match(rf"{re.escape(alias)}(?:[^a-z]|$)", text):
                leading.add(economy)
            if re.search(rf"(?:^|[^a-z]){re.escape(alias)}(?:[^a-z]|$)", text):
                contained.add(economy)

    candidates = leading or contained
    return next(iter(candidates)) if len(candidates) == 1 else ""


def infer_balance_export_identity(
    workbook_path: Path | str,
) -> BalanceExportIdentity:
    """Read the economy, scenario, and years a LEAP export declares.

    LEAP writes its area name and the scenario/year/units of each sheet into
    the sheet headers, so a caller does not need to be told any of them. A
    workbook covering more than one scenario is rejected rather than resolved
    arbitrarily: nothing in the export says which one a review should use.
    """
    path = _resolve_path(workbook_path)
    sheets = list_balance_export_sheets(path)

    scenario_codes = sorted({sheet.scenario_code for sheet in sheets})
    if len(scenario_codes) > 1:
        scenarios = ", ".join(
            "Reference" if code == "REF" else "Target" if code == "TGT" else code
            for code in scenario_codes
        )
        raise ValueError(
            f"This export covers more than one scenario ({scenarios}). "
            "Export one scenario at a time so the review knows which to use."
        )
    scenario_code = scenario_codes[0]
    if scenario_code not in {"REF", "TGT"}:
        raise ValueError(
            f"This export declares the scenario {scenario_code!r}. A balance "
            "review compares a Reference or Target export."
        )

    area_name = read_balance_export_area_name(path)
    economy = economy_from_balance_export_area(area_name)

    return BalanceExportIdentity(
        path=path,
        economy=economy,
        area_name=area_name,
        scenario="Reference" if scenario_code == "REF" else "Target",
        scenario_code=scenario_code,
        years=tuple(sorted({sheet.year for sheet in sheets})),
        units=sheets[0].units,
    )


def select_balance_export_sheets(
    workbook_path: Path | str,
    *,
    years: Sequence[int],
    scenarios: Sequence[str],
) -> list[BalanceExportSheet]:
    """Return exact requested scenario/year sheets and reject missing/duplicates."""
    wanted_years = sorted({int(year) for year in years})
    wanted_codes = {
        normalize_balance_scenario_code(scenario) for scenario in scenarios
    }
    if not wanted_years:
        raise ValueError("At least one balance-export year is required.")
    if not wanted_codes:
        raise ValueError("At least one balance-export scenario is required.")

    available_sheets = list_balance_export_sheets(workbook_path)
    selected = [
        sheet
        for sheet in available_sheets
        if sheet.year in wanted_years and sheet.scenario_code in wanted_codes
    ]
    counts: dict[tuple[str, int], int] = {}
    for sheet in selected:
        key = (sheet.scenario_code, sheet.year)
        counts[key] = counts.get(key, 0) + 1
    duplicates = sorted(key for key, count in counts.items() if count > 1)
    if duplicates:
        raise ValueError(
            f"Duplicate scenario/year balance sheets in {workbook_path}: {duplicates}"
        )

    expected = {
        (scenario_code, year)
        for scenario_code in wanted_codes
        for year in wanted_years
    }
    missing = sorted(expected - set(counts))
    if missing:
        raise ValueError(
            f"Requested balance sheets are missing from {workbook_path}: {missing}"
        )
    return sorted(selected, key=lambda sheet: (sheet.scenario_code, sheet.year))


def _leap_balance_sheet_unit_to_pj_multiplier(raw: pd.DataFrame) -> float:
    """Return the multiplier needed to convert a LEAP balance sheet to PJ."""
    unit_text = ""
    for row_idx in range(min(4, len(raw))):
        for value in raw.iloc[row_idx].tolist():
            text = str(value or "").strip()
            match = re.search(r"units:\s*(.+)$", text, flags=re.IGNORECASE)
            if match:
                unit_text = match.group(1).strip().lower()
                break
        if unit_text:
            break
    if not unit_text:
        return 1.0

    unit_text = unit_text.rstrip(".")
    if unit_text.startswith("thousand petajoule"):
        return 1000.0
    if unit_text.startswith("petajoule"):
        return 1.0
    if unit_text.startswith("terajoule"):
        return 0.001
    if unit_text.startswith("gigajoule"):
        return 0.000001
    if unit_text.startswith("million gigajoule"):
        return 1.0
    return 1.0


SCENARIO_LABEL_ALIASES = {
    "reference": "REF",
    "ref": "REF",
    "target": "TGT",
    "tgt": "TGT",
}


def _leap_balance_sheet_scenario_label(raw: pd.DataFrame) -> str:
    """Return the normalized REF/TGT scenario code from a sheet's "Scenario:" subtitle, or "" if absent/unrecognized."""
    for row_idx in range(min(4, len(raw))):
        for value in raw.iloc[row_idx].tolist():
            text = str(value or "").strip()
            match = re.search(r"scenario:\s*([^,]+)", text, flags=re.IGNORECASE)
            if match:
                label = match.group(1).strip().lower()
                return SCENARIO_LABEL_ALIASES.get(label, "")
    return ""


def _scenario_code_from_balance_export_filename(workbook: Path) -> str:
    """Return the REF/TGT scenario code implied by a balance-export filename, or "" if unrecognized."""
    parts = _balance_export_filename_parts(workbook)
    return parts[1] if parts is not None else ""


def load_leap_balance_activity_table(
    workbook_path: Path | str,
    *,
    balance_rows: Sequence[str],
    fuels: Sequence[str],
) -> pd.DataFrame:
    """Return long LEAP balance values for selected row labels and fuel columns.

    Values are normalized to petajoules when the LEAP sheet subtitle declares a
    recognized energy unit.
    """
    workbook = _resolve_path(workbook_path)
    if not workbook.exists():
        raise FileNotFoundError(f"Missing LEAP balance workbook: {workbook}")

    wanted_rows = {normalize_balance_label(row) for row in balance_rows}
    wanted_fuels = {normalize_balance_label(fuel) for fuel in fuels}
    rows: list[dict[str, object]] = []
    xls = pd.ExcelFile(workbook)
    for sheet_name in xls.sheet_names:
        name = str(sheet_name).strip()
        if name.lower().startswith("ebal|"):
            year_text = name.split("|", 1)[1]
        elif name.isdigit():
            # Raw "full model output" exports from LEAP use bare-year sheet
            # names (e.g. "2022") rather than the "EBal|2022" convention used
            # by this pipeline's own generated balance workbooks.
            year_text = name
        else:
            continue
        try:
            year = int(year_text)
        except Exception:
            continue
        raw = pd.read_excel(workbook, sheet_name=sheet_name, header=None)
        if raw.shape[0] < 3 or raw.shape[1] < 2:
            continue
        unit_multiplier = _leap_balance_sheet_unit_to_pj_multiplier(raw)
        header_row_idx = None
        best_match_count = 0
        for candidate_idx in range(min(8, len(raw))):
            candidate_labels = raw.iloc[candidate_idx].fillna("").astype(str).tolist()
            match_count = sum(
                1
                for label in candidate_labels[1:]
                if normalize_balance_label(label) in wanted_fuels
            )
            if match_count > best_match_count:
                best_match_count = match_count
                header_row_idx = candidate_idx
        if header_row_idx is None or best_match_count == 0:
            continue
        fuel_labels = raw.iloc[header_row_idx].fillna("").astype(str).tolist()
        fuel_columns = {
            col_idx: label
            for col_idx, label in enumerate(fuel_labels)
            if col_idx > 0 and normalize_balance_label(label) in wanted_fuels
        }
        if not fuel_columns:
            continue
        for row_idx in range(header_row_idx + 1, len(raw)):
            row_label = str(raw.iat[row_idx, 0] or "").strip()
            if normalize_balance_label(row_label) not in wanted_rows:
                continue
            for col_idx, fuel_label in fuel_columns.items():
                value = pd.to_numeric(raw.iat[row_idx, col_idx], errors="coerce")
                if pd.isna(value):
                    value = 0.0
                value = float(value) * unit_multiplier
                if is_leap_balance_own_use_or_loss_row(row_label):
                    value = -value
                rows.append(
                    {
                        "source_dataset": "leap_balance",
                        "year": int(year),
                        "balance_row": row_label,
                        "fuel_label": str(fuel_label).strip(),
                        "value": value,
                    }
                )
    columns = ["source_dataset", "year", "balance_row", "fuel_label", "value"]
    return pd.DataFrame(rows, columns=columns)


def build_leap_balance_activity_series(
    leap_balance_activity: pd.DataFrame,
    *,
    balance_rows: Sequence[str],
    fuels: Sequence[str],
    value_mode: str = "signed_sum",
    base_year: int,
    final_year: int,
) -> dict[int, float]:
    """Sum selected LEAP balance rows/fuels into one yearly activity series."""
    wanted_fuels = {normalize_balance_label(fuel) for fuel in fuels}
    wanted_rows = {normalize_balance_label(row) for row in balance_rows}
    year_range = range(int(base_year), int(final_year) + 1)
    if leap_balance_activity.empty or not wanted_fuels or not wanted_rows:
        return {year: 0.0 for year in year_range}

    subset = leap_balance_activity[
        leap_balance_activity["fuel_label"].map(normalize_balance_label).isin(wanted_fuels)
        & leap_balance_activity["balance_row"].map(normalize_balance_label).isin(wanted_rows)
    ].copy()
    if subset.empty:
        return {year: 0.0 for year in year_range}

    mode = str(value_mode or "signed_sum").strip().lower()
    values = pd.to_numeric(subset["value"], errors="coerce").fillna(0.0)
    if mode in {"signed", "signed_sum", ""}:
        subset["activity_value"] = values
    elif mode in {"positive", "positive_only", "outputs"}:
        subset["activity_value"] = values.where(values > 0.0, 0.0)
    elif mode in {"negative_abs", "input_abs", "inputs_abs"}:
        subset["activity_value"] = values.where(values < 0.0, 0.0).abs()
    elif mode in {"absolute", "abs"}:
        subset["activity_value"] = values.abs()
    else:
        raise ValueError(f"Invalid LEAP balance value_mode={mode!r}.")

    grouped = subset.groupby("year", dropna=False)["activity_value"].sum()
    return {
        year: float(grouped.get(year, 0.0))
        for year in year_range
    }
