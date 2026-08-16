#%%
"""Resolve and validate the generated Outlook mapping master.

Normal developer checkouts use the live workbook in the sibling
``leap_mappings`` repository.  Standalone checkouts use the committed,
read-only fallback snapshot in this repository.  A present but broken live
workbook is never hidden by falling back to the snapshot.
"""
from __future__ import annotations

from dataclasses import asdict, dataclass
import hashlib
import json
from pathlib import Path
import subprocess
import warnings

from openpyxl import load_workbook


REPO_ROOT = Path(__file__).resolve().parents[2]
LEAP_MAPPINGS_REPO_ROOT = REPO_ROOT.parent / "leap_mappings"
LIVE_MAPPING_WORKBOOK_PATH = (
    LEAP_MAPPINGS_REPO_ROOT / "config" / "outlook_mappings_master.xlsx"
)
FALLBACK_MAPPING_WORKBOOK_PATH = (
    REPO_ROOT / "config" / "outlook_mappings_master_DO_NOT_EDIT.xlsx"
)
FALLBACK_PROVENANCE_PATH = (
    REPO_ROOT / "config" / "outlook_mappings_master_DO_NOT_EDIT.provenance.json"
)
MAPPING_CONTRACT_VERSION = "initialisation-canonical-mapping-v1"

# This boundary contract is deliberately small: it names every sheet and
# column initialisation consumes while allowing mapping authors to add fields.
REQUIRED_SHEET_COLUMNS: dict[str, tuple[str, ...]] = {
    "leap_combined_esto": (
        "leap_sector_name_full_path", "raw_leap_fuel_name", "esto_flow", "esto_product",
    ),
    "leap_combined_ninth": (
        "leap_sector_name_full_path", "raw_leap_fuel_name", "ninth_sector", "ninth_fuel",
    ),
    "ninth_pairs_to_esto_pairs": (
        "ninth_sector", "ninth_fuel", "esto_flow", "esto_product",
    ),
    "leap_display_names": ("code", "leap_display_name"),
    "ninth fuel to esto product": ("ninth_fuel", "esto_product"),
    "leap_rollup_rules": (
        "rollup_context", "input_leap_sector_name_full_path", "input_raw_leap_fuel_name",
        "rolled_leap_sector_name_full_path", "rolled_raw_leap_fuel_name",
        "rollup_group_id", "rollup_reason", "priority", "include", "Note",
    ),
    "esto_rollup_rules": (
        "rollup_context", "input_esto_flow", "input_esto_product", "rolled_esto_flow",
        "rolled_esto_product", "rollup_group_id", "rollup_reason", "priority", "include", "Note",
    ),
    "ninth_rollup_rules": (
        "rollup_context", "input_ninth_sector", "input_ninth_fuel", "rolled_ninth_sector",
        "rolled_ninth_fuel", "rollup_group_id", "rollup_reason", "priority", "include", "Note",
    ),
}


class MappingWorkbookResolutionError(RuntimeError):
    """Raised when neither an acceptable live nor fallback workbook is usable."""


@dataclass(frozen=True)
class MappingWorkbookSelection:
    path: Path
    selected_source: str
    sha256: str
    source_commit: str
    contract_version: str = MAPPING_CONTRACT_VERSION
    fallback_refresh_required: bool = False

    def as_manifest_record(self) -> dict[str, object]:
        record = asdict(self)
        record["path"] = str(self.path.resolve())
        return record


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for block in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def validate_mapping_workbook(path: Path) -> None:
    """Fail clearly when a workbook cannot satisfy the consumed schema."""
    if not path.is_file():
        raise MappingWorkbookResolutionError(f"Mapping workbook not found: {path}")
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        raise MappingWorkbookResolutionError(
            f"Could not open mapping workbook {path}: {type(exc).__name__}: {exc}"
        ) from exc
    try:
        for sheet_name, required_columns in REQUIRED_SHEET_COLUMNS.items():
            if sheet_name not in workbook.sheetnames:
                raise MappingWorkbookResolutionError(
                    f"Mapping workbook {path} is missing required sheet {sheet_name!r}."
                )
            row = next(workbook[sheet_name].iter_rows(min_row=1, max_row=1, values_only=True), ())
            present = {str(value).strip() for value in row if value is not None}
            missing = [column for column in required_columns if column not in present]
            if missing:
                raise MappingWorkbookResolutionError(
                    f"Mapping sheet {sheet_name!r} in {path} is missing required columns {missing}."
                )
    finally:
        workbook.close()


def _read_fallback_provenance() -> dict[str, object]:
    if not FALLBACK_PROVENANCE_PATH.is_file():
        raise MappingWorkbookResolutionError(
            f"Fallback provenance file not found: {FALLBACK_PROVENANCE_PATH}"
        )
    try:
        return json.loads(FALLBACK_PROVENANCE_PATH.read_text(encoding="utf-8"))
    except Exception as exc:
        raise MappingWorkbookResolutionError(
            f"Could not read fallback provenance {FALLBACK_PROVENANCE_PATH}: {exc}"
        ) from exc


def _git_commit(repo_root: Path) -> str:
    try:
        result = subprocess.run(
            ["git", "-C", str(repo_root), "rev-parse", "HEAD"],
            check=True,
            capture_output=True,
            text=True,
            timeout=10,
        )
        return result.stdout.strip()
    except Exception:
        return ""


def resolve_mapping_workbook() -> MappingWorkbookSelection:
    """Select live sibling first, otherwise the verified vendored fallback."""
    if LIVE_MAPPING_WORKBOOK_PATH.exists():
        validate_mapping_workbook(LIVE_MAPPING_WORKBOOK_PATH)
        live_hash = sha256_file(LIVE_MAPPING_WORKBOOK_PATH)
        fallback_differs = False
        if FALLBACK_MAPPING_WORKBOOK_PATH.is_file():
            fallback_differs = sha256_file(FALLBACK_MAPPING_WORKBOOK_PATH) != live_hash
            if fallback_differs:
                warnings.warn(
                    "The live mapping master differs from the vendored fallback. "
                    "The live workbook was selected; run the fallback sync helper before release.",
                    RuntimeWarning,
                    stacklevel=2,
                )
        return MappingWorkbookSelection(
            path=LIVE_MAPPING_WORKBOOK_PATH,
            selected_source="sibling_live",
            sha256=live_hash,
            source_commit=_git_commit(LEAP_MAPPINGS_REPO_ROOT),
            fallback_refresh_required=fallback_differs,
        )

    validate_mapping_workbook(FALLBACK_MAPPING_WORKBOOK_PATH)
    provenance = _read_fallback_provenance()
    fallback_hash = sha256_file(FALLBACK_MAPPING_WORKBOOK_PATH)
    declared_hash = str(provenance.get("workbook_sha256", "")).strip().lower()
    if not declared_hash or fallback_hash != declared_hash:
        raise MappingWorkbookResolutionError(
            "Fallback mapping workbook hash does not match its provenance sidecar: "
            f"actual={fallback_hash}, declared={declared_hash or '<missing>'}."
        )
    declared_contract = str(provenance.get("contract_version", "")).strip()
    if declared_contract != MAPPING_CONTRACT_VERSION:
        raise MappingWorkbookResolutionError(
            "Fallback mapping contract version is incompatible: "
            f"expected={MAPPING_CONTRACT_VERSION}, declared={declared_contract or '<missing>'}."
        )
    return MappingWorkbookSelection(
        path=FALLBACK_MAPPING_WORKBOOK_PATH,
        selected_source="vendored_fallback",
        sha256=fallback_hash,
        source_commit=str(provenance.get("source_repository_commit", "")).strip(),
    )


SELECTED_MAPPING_WORKBOOK = resolve_mapping_workbook()

#%%
