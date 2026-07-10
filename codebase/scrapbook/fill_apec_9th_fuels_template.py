"""One-off filler for workbook sheet: 'APEC 9th fuels' (FED + TPES)."""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from codebase.utilities.master_config import read_config_table


REPO_ROOT = Path(__file__).resolve().parents[2]
CSV_PATH = REPO_ROOT / "data" / "merged_file_energy_00_APEC_20251106.csv"
MAP_PATH = REPO_ROOT / "config" / "sector_fuel_codes_to_names.xlsx"
WORKBOOK_PATH = REPO_ROOT / "data" / "Data for comparison  - APERC outlooks .xlsx"
TARGET_SHEET = "APEC 9th fuels"
ECONOMY = "00_APEC"
SCENARIO = "reference"

SECTOR_MAP = {
    "FED": "12_total_final_consumption",
    "TPES": "07_total_primary_energy_supply",
}

TARGET_YEARS = list(range(2000, 2060 + 1))
TARGET_BUCKETS = [
    "Coal",
    "Oil",
    "Gas",
    "Electricity",
    "Heat",
    "Hydrogen",
    "Other",
    "Renewables_total",
    "Biomass",
    "Other renewables",
]


def _norm(text: object) -> str:
    return str(text).strip().lower()


def _bucket_from_codes_and_name(fuel_code: str, subfuel_code: str, name: str) -> str:
    n = _norm(name)

    # Direct buckets first.
    if fuel_code == "17_electricity":
        return "Electricity"
    if fuel_code == "18_heat":
        return "Heat"
    if "hydrogen" in n or subfuel_code == "16_x_hydrogen":
        return "Hydrogen"
    if fuel_code.startswith("08_"):
        return "Gas"
    if fuel_code.startswith("01_") or fuel_code.startswith("02_"):
        return "Coal"
    if fuel_code.startswith("06_") or fuel_code.startswith("07_") or fuel_code.startswith("05_"):
        return "Oil"

    # Renewables split.
    if (
        fuel_code.startswith("15_")
        or "biomass" in n
        or "bagasse" in n
        or "charcoal" in n
        or "fuelwood" in n
        or "black liq" in n
        or "biogas" in n
        or "biodiesel" in n
        or "biogasoline" in n
        or "bio jet" in n
        or "liquid biofuel" in n
        or "municipal solid waste (renewable)" in n
    ):
        return "Biomass"

    if (
        fuel_code.startswith("10_")
        or fuel_code.startswith("11_")
        or fuel_code.startswith("12_")
        or fuel_code.startswith("13_")
        or fuel_code.startswith("14_")
        or "solar" in n
        or "wind" in n
        or "hydro" in n
        or "geothermal" in n
        or "tide wave ocean" in n
        or "photovoltaic" in n
    ):
        return "Other renewables"

    return "Other"


def main() -> None:
    df = pd.read_csv(CSV_PATH)
    code_map = read_config_table(MAP_PATH, sheet_name="code_to_name")[["ninth_label", "ninth_column", "name"]]
    code_map = code_map.dropna(subset=["ninth_label", "ninth_column"]).copy()
    code_map["ninth_label"] = code_map["ninth_label"].astype(str)
    code_map["ninth_column"] = code_map["ninth_column"].astype(str)

    fuel_name_map = (
        code_map[code_map["ninth_column"] == "fuels"][["ninth_label", "name"]]
        .drop_duplicates()
        .set_index("ninth_label")["name"]
        .to_dict()
    )
    subfuel_name_map = (
        code_map[code_map["ninth_column"] == "subfuels"][["ninth_label", "name"]]
        .drop_duplicates()
        .set_index("ninth_label")["name"]
        .to_dict()
    )

    year_cols = [str(y) for y in TARGET_YEARS]

    def build_bucket_rows(sector_code: str) -> dict[str, dict[str, float]]:
        f = df[
            (df["scenarios"] == SCENARIO)
            & (df["economy"] == ECONOMY)
            & (df["sectors"] == sector_code)
            & (df["subtotal_layout"] == False)
            & (df["subtotal_results"] == False)
        ].copy()
        if f.empty:
            raise RuntimeError(f"No rows found for sector {sector_code}.")

        # Avoid double-counting: where detailed subfuels exist for a fuel, drop the 'x' subtotal row.
        keep_parts: list[pd.DataFrame] = []
        for _, grp in f.groupby("fuels"):
            has_nonx = (grp["subfuels"].astype(str) != "x").any()
            keep_parts.append(grp[grp["subfuels"].astype(str) != "x"] if has_nonx else grp)
        f = pd.concat(keep_parts, ignore_index=True)

        def resolve_name(row: pd.Series) -> str:
            subfuel = str(row["subfuels"])
            fuel = str(row["fuels"])
            if subfuel != "x":
                return str(subfuel_name_map.get(subfuel, subfuel))
            return str(fuel_name_map.get(fuel, fuel))

        f["mapped_name"] = f.apply(resolve_name, axis=1)
        f["bucket"] = f.apply(
            lambda r: _bucket_from_codes_and_name(str(r["fuels"]), str(r["subfuels"]), str(r["mapped_name"])),
            axis=1,
        )
        grouped = f.groupby("bucket", as_index=False)[year_cols].sum(min_count=1)
        bucket_to_values: dict[str, dict[str, float]] = {}
        for _, row in grouped.iterrows():
            bucket_to_values[str(row["bucket"])] = {y: float(row[y]) if pd.notna(row[y]) else 0.0 for y in year_cols}

        # Renewables_total = Biomass + Other renewables
        bio = bucket_to_values.get("Biomass", {y: 0.0 for y in year_cols})
        oth = bucket_to_values.get("Other renewables", {y: 0.0 for y in year_cols})
        bucket_to_values["Renewables_total"] = {y: float(bio.get(y, 0.0)) + float(oth.get(y, 0.0)) for y in year_cols}
        return bucket_to_values

    # Build both sector datasets.
    sector_bucket_values = {label: build_bucket_rows(code) for label, code in SECTOR_MAP.items()}

    wb = load_workbook(WORKBOOK_PATH)
    if TARGET_SHEET in wb.sheetnames:
        wb.remove(wb[TARGET_SHEET])
    ws = wb.create_sheet(TARGET_SHEET)

    headers = ["scenario", "sector", "fuel", "units"] + year_cols
    for c, h in enumerate(headers, start=1):
        ws.cell(1, c, h)
    ws.freeze_panes = "A2"

    row_idx = 2
    for sector_label in ("FED", "TPES"):
        values = sector_bucket_values[sector_label]
        for fuel in TARGET_BUCKETS:
            ws.cell(row_idx, 1, "Reference")
            ws.cell(row_idx, 2, sector_label)
            ws.cell(row_idx, 3, fuel)
            ws.cell(row_idx, 4, "PJ")
            for j, y in enumerate(year_cols, start=5):
                ws.cell(row_idx, j, float(values.get(fuel, {}).get(y, 0.0)))
            row_idx += 1

    wb.save(WORKBOOK_PATH)
    print("Filled template sheet:", TARGET_SHEET)
    print("Workbook:", WORKBOOK_PATH)


if __name__ == "__main__":
    main()
