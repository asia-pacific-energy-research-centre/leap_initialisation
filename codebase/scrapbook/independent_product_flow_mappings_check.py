#%%
# Summary: Validate independent product/flow mappings for completeness and many-to-many issues.
import os
import pandas as pd

from codebase.utilities.master_config import read_config_table
#%%

#%%
######### CONSTANTS (UNLIKELY TO CHANGE) #########
ENABLE_DEBUG_BREAKPOINTS = True
NINTH_SECTOR_COLUMNS = [
    "sectors",
    "sub1sectors",
    "sub2sectors",
    "sub3sectors",
    "sub4sectors",
]
NINTH_FUEL_COLUMNS = ["fuels", "subfuels"]
ESTO_IGNORE_LABELS = {
    "19 Total",
    "20 Total Renewables",
    "21 Modern renewables",
}
#%%

#%%
######### FUNCTIONS #########
def ensure_repo_root():
    """Move to repo root if running from the scrapbook folder."""
    try:
        if os.getcwd().endswith("scrapbook"):
            os.chdir("../../")
    except Exception as exc:
        print(f"Failed to set repo root: {exc}")
        try_debug_breakpoint()
        raise


def try_debug_breakpoint():
    """Trigger a debug breakpoint when enabled (safe to call anywhere)."""
    if not ENABLE_DEBUG_BREAKPOINTS:
        return
    try:
        breakpoint()
    except Exception as breakpoint_exc:
        print(f"Debug breakpoint failed: {breakpoint_exc}")


def _normalize_series(series, drop_x=False):
    values = series.fillna("").astype(str).str.strip()
    if drop_x:
        values = values[(values != "") & (values.str.lower() != "x")]
    else:
        values = values[values != ""]
    return values


def _strip_esto_ignored(values):
    return values[~values.isin(ESTO_IGNORE_LABELS)]


def load_mapping_workbook(workbook_path):
    """Load mapping sheets for product and flow."""
    try:
        return {
            sheet_name: read_config_table(workbook_path, sheet_name=sheet_name, dtype=str)
            for sheet_name in ["product", "flow"]
        }
    except Exception as exc:
        print(f"Failed to read workbook {workbook_path}: {exc}")
        try_debug_breakpoint()
        raise


def collect_reference_labels(workbook_path):
    """Collect reference labels from the sector_fuel_codes_to_names workbook."""
    try:
        ninth_df = read_config_table(workbook_path, sheet_name="9th", dtype=str)
        esto_df = read_config_table(workbook_path, sheet_name="ESTO", dtype=str)

        ninth_sectors = pd.concat(
            [_normalize_series(ninth_df[col], drop_x=True) for col in NINTH_SECTOR_COLUMNS if col in ninth_df],
            ignore_index=True,
        ).drop_duplicates()
        ninth_fuels = pd.concat(
            [_normalize_series(ninth_df[col], drop_x=True) for col in NINTH_FUEL_COLUMNS if col in ninth_df],
            ignore_index=True,
        ).drop_duplicates()

        esto_flows = _strip_esto_ignored(
            _normalize_series(esto_df.get("flows", pd.Series([], dtype=str)))
        )
        esto_products = _strip_esto_ignored(
            _normalize_series(esto_df.get("products", pd.Series([], dtype=str)))
        )

        return {
            "ninth_sectors": set(ninth_sectors.tolist()),
            "ninth_fuels": set(ninth_fuels.tolist()),
            "esto_flows": set(esto_flows.tolist()),
            "esto_products": set(esto_products.tolist()),
        }
    except Exception as exc:
        print(f"Failed to collect reference labels from {workbook_path}: {exc}")
        try_debug_breakpoint()
        raise


def report_duplicates(df, sheet_name):
    """Report exact duplicate rows."""
    dupes = df[df.duplicated(keep=False)]
    print(f"{sheet_name}: duplicate rows = {len(dupes)}")
    if len(dupes):
        print(dupes.to_string(index=False))
    return dupes


def report_missing_labels(df, left_col, right_col, left_reference, right_reference, sheet_name, drop_x_left=False):
    """Report missing labels vs reference sets."""
    left_values = set(_normalize_series(df[left_col], drop_x=drop_x_left).tolist())
    right_values = set(_strip_esto_ignored(_normalize_series(df[right_col])).tolist())

    missing_left = sorted(set(left_reference) - left_values)
    missing_right = sorted(set(right_reference) - right_values)

    print(f"{sheet_name}: missing {left_col} labels = {len(missing_left)}")
    if missing_left:
        print(pd.DataFrame({left_col: missing_left}).to_string(index=False))

    print(f"{sheet_name}: missing {right_col} labels = {len(missing_right)}")
    if missing_right:
        print(pd.DataFrame({right_col: missing_right}).to_string(index=False))

    return missing_left, missing_right


def report_unknown_labels(df, left_col, right_col, left_reference, right_reference, sheet_name, drop_x_left=False):
    """Report labels present in mapping but missing from reference sets."""
    left_values = set(_normalize_series(df[left_col], drop_x=drop_x_left).tolist())
    right_values = set(_strip_esto_ignored(_normalize_series(df[right_col])).tolist())

    unknown_left = sorted(left_values - set(left_reference))
    unknown_right = sorted(right_values - set(right_reference))

    print(f"{sheet_name}: unknown {left_col} labels = {len(unknown_left)}")
    if unknown_left:
        print(pd.DataFrame({left_col: unknown_left}).to_string(index=False))

    print(f"{sheet_name}: unknown {right_col} labels = {len(unknown_right)}")
    if unknown_right:
        print(pd.DataFrame({right_col: unknown_right}).to_string(index=False))

    return unknown_left, unknown_right


def report_many_to_many(df, left_col, right_col, sheet_name):
    """Report many-to-many mappings where both sides have degree > 1."""
    working = df.copy()
    working[left_col] = working[left_col].fillna("").astype(str).str.strip()
    working[right_col] = working[right_col].fillna("").astype(str).str.strip()
    working = working[(working[left_col] != "") & (working[right_col] != "")]
    working = working[~working[right_col].isin(ESTO_IGNORE_LABELS)]
    working = working.drop_duplicates(subset=[left_col, right_col])

    if working.empty:
        print(f"{sheet_name}: no rows with both {left_col} and {right_col}")
        return pd.DataFrame()

    left_counts = working[left_col].value_counts()
    right_counts = working[right_col].value_counts()
    working["left_degree"] = working[left_col].map(left_counts)
    working["right_degree"] = working[right_col].map(right_counts)

    many_to_many = (working["left_degree"] > 1) & (working["right_degree"] > 1)
    many_to_many_df = working.loc[many_to_many, [left_col, right_col]].copy()
    print(f"{sheet_name}: many-to-many mappings = {len(many_to_many_df)}")
    if len(many_to_many_df):
        print(many_to_many_df.to_string(index=False))
    return many_to_many_df


def resolve_sheet_columns(df, sheet_name):
    """Resolve mapping column names for product/flow sheets."""
    columns = [str(col).strip() for col in df.columns]
    column_set = {col.lower() for col in columns}
    if sheet_name == PRODUCT_SHEET_NAME:
        if {"ninth_fuel", "esto_product"}.issubset(column_set):
            return "ninth_fuel", "esto_product"
        if {"ninth_label", "esto_label"}.issubset(column_set):
            return "ninth_label", "esto_label"
    if sheet_name == FLOW_SHEET_NAME:
        if {"ninth_sector", "esto_flow"}.issubset(column_set):
            return "ninth_sector", "esto_flow"
        if {"ninth_label", "esto_label"}.issubset(column_set):
            return "ninth_label", "esto_label"
    raise ValueError(
        f"{sheet_name} sheet missing expected columns; found: {columns}"
    )


def add_missing_values(df, left_col, right_col, left_reference, right_reference, drop_x_left=False):
    left_values = _normalize_series(df[left_col], drop_x=drop_x_left)
    right_values = _strip_esto_ignored(_normalize_series(df[right_col]))
    missing_left = sorted(set(left_reference) - set(left_values.tolist()))
    missing_right = sorted(set(right_reference) - set(right_values.tolist()))

    if not missing_left and not missing_right:
        return df.copy(), 0, 0

    new_rows = []
    for value in missing_left:
        row = {col: "" for col in df.columns}
        row[left_col] = value
        new_rows.append(row)
    for value in missing_right:
        row = {col: "" for col in df.columns}
        row[right_col] = value
        new_rows.append(row)

    updated = pd.concat([df, pd.DataFrame(new_rows)], ignore_index=True)
    return updated, len(missing_left), len(missing_right)


def remove_duplicate_rows(df):
    return df.drop_duplicates().copy()


def write_output_with_highlights(output_path, sheets, many_to_many_pairs):
    try:
        output_dir = os.path.dirname(output_path)
        if output_dir:
            os.makedirs(output_dir, exist_ok=True)
        with pd.ExcelWriter(output_path, engine="openpyxl", mode="w") as writer:
            for sheet_name, df in sheets.items():
                df.to_excel(writer, sheet_name=sheet_name, index=False)

        if not HIGHLIGHT_MANY_TO_MANY:
            return

        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill

        wb = load_workbook(output_path)
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        for sheet_name, pairs in many_to_many_pairs.items():
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            if not pairs:
                continue

            header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            if sheet_name == PRODUCT_SHEET_NAME:
                left_col, right_col = resolve_sheet_columns(
                    pd.DataFrame(columns=header), sheet_name
                )
            else:
                left_col, right_col = resolve_sheet_columns(
                    pd.DataFrame(columns=header), sheet_name
                )
            try:
                left_idx = header.index(left_col) + 1
                right_idx = header.index(right_col) + 1
            except ValueError:
                continue
            notes_idx = None
            for idx, col_name in enumerate(header, start=1):
                if str(col_name).strip().lower() == "notes":
                    notes_idx = idx
                    break

            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                left_val = str(row[left_idx - 1].value or "").strip()
                right_val = str(row[right_idx - 1].value or "").strip()
                if (left_val, right_val) in pairs:
                    for cell in row:
                        cell.fill = red_fill
                    if notes_idx is not None:
                        notes_cell = row[notes_idx - 1]
                        existing = str(notes_cell.value or "").strip()
                        tag = "many-to-many"
                        if tag not in existing.lower():
                            notes_cell.value = (
                                f"{existing}; {tag}".strip("; ") if existing else tag
                            )

        wb.save(output_path)
    except Exception as exc:
        print(f"Failed to write highlighted output workbook {output_path}: {exc}")
        try_debug_breakpoint()
        raise

#%%

#%%
######### CONSTANTS (LIKELY TO CHANGE) #########
MAPPING_WORKBOOK_PATH = "config/independent product flow mappings.xlsx"
REFERENCE_WORKBOOK_PATH = "config/sector_fuel_codes_to_names.xlsx"
OUTPUT_WORKBOOK_PATH = "outputs/independent product flow mappings_checked.xlsx"
PRODUCT_SHEET_NAME = "product"
FLOW_SHEET_NAME = "flow"
RUN_CHECKS = True
REMOVE_DUPLICATE_ROWS = True
ADD_MISSING_VALUES = True
HIGHLIGHT_MANY_TO_MANY = True
WRITE_OUTPUT = True
#%%

#%%
######### RUN CHECKS (TOGGLE) #########
if RUN_CHECKS:
    try:
        ensure_repo_root()
        print(
            "Running independent product/flow mapping checks: duplicates, missing labels, "
            "unknown labels, and many-to-many mappings."
        )
        sheets = load_mapping_workbook(MAPPING_WORKBOOK_PATH)
        references = collect_reference_labels(REFERENCE_WORKBOOK_PATH)
        updated_sheets = dict(sheets)
        many_to_many_pairs = {}

        product_df = sheets.get(PRODUCT_SHEET_NAME)
        if product_df is None:
            raise ValueError(f"Missing sheet: {PRODUCT_SHEET_NAME}")
        left_col, right_col = resolve_sheet_columns(product_df, PRODUCT_SHEET_NAME)
        if REMOVE_DUPLICATE_ROWS:
            product_df = remove_duplicate_rows(product_df)
        report_duplicates(product_df, PRODUCT_SHEET_NAME)
        if ADD_MISSING_VALUES:
            product_df, added_left, added_right = add_missing_values(
                product_df,
                left_col,
                right_col,
                references["ninth_fuels"],
                references["esto_products"],
                drop_x_left=True,
            )
            print(
                f"{PRODUCT_SHEET_NAME}: added {added_left} {left_col} and {added_right} {right_col}"
            )
        report_missing_labels(
            product_df,
            left_col,
            right_col,
            references["ninth_fuels"],
            references["esto_products"],
            PRODUCT_SHEET_NAME,
            drop_x_left=True,
        )
        report_unknown_labels(
            product_df,
            left_col,
            right_col,
            references["ninth_fuels"],
            references["esto_products"],
            PRODUCT_SHEET_NAME,
            drop_x_left=True,
        )
        many_to_many_df = report_many_to_many(
            product_df, left_col, right_col, PRODUCT_SHEET_NAME
        )
        many_to_many_pairs[PRODUCT_SHEET_NAME] = set(
            zip(many_to_many_df[left_col], many_to_many_df[right_col])
        ) if not many_to_many_df.empty else set()
        updated_sheets[PRODUCT_SHEET_NAME] = product_df

        flow_df = sheets.get(FLOW_SHEET_NAME)
        if flow_df is None:
            raise ValueError(f"Missing sheet: {FLOW_SHEET_NAME}")
        left_col, right_col = resolve_sheet_columns(flow_df, FLOW_SHEET_NAME)
        if REMOVE_DUPLICATE_ROWS:
            flow_df = remove_duplicate_rows(flow_df)
        report_duplicates(flow_df, FLOW_SHEET_NAME)
        if ADD_MISSING_VALUES:
            flow_df, added_left, added_right = add_missing_values(
                flow_df,
                left_col,
                right_col,
                references["ninth_sectors"],
                references["esto_flows"],
                drop_x_left=True,
            )
            print(
                f"{FLOW_SHEET_NAME}: added {added_left} {left_col} and {added_right} {right_col}"
            )
        report_missing_labels(
            flow_df,
            left_col,
            right_col,
            references["ninth_sectors"],
            references["esto_flows"],
            FLOW_SHEET_NAME,
            drop_x_left=True,
        )
        report_unknown_labels(
            flow_df,
            left_col,
            right_col,
            references["ninth_sectors"],
            references["esto_flows"],
            FLOW_SHEET_NAME,
            drop_x_left=True,
        )
        many_to_many_df = report_many_to_many(
            flow_df, left_col, right_col, FLOW_SHEET_NAME
        )
        many_to_many_pairs[FLOW_SHEET_NAME] = set(
            zip(many_to_many_df[left_col], many_to_many_df[right_col])
        ) if not many_to_many_df.empty else set()
        updated_sheets[FLOW_SHEET_NAME] = flow_df

        if WRITE_OUTPUT:
            write_output_with_highlights(
                OUTPUT_WORKBOOK_PATH,
                updated_sheets,
                many_to_many_pairs,
            )
            print(f"Output written to: {OUTPUT_WORKBOOK_PATH}")
    except Exception as exc:
        print(f"Failed to run independent product/flow mapping checks: {exc}")
        try_debug_breakpoint()
#%%
