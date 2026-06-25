#%%
"""
Reconcile mapping coverage across leap_mappings and master_config, then write
analysis outputs for ninth_pairs_to_esto_pairs consistency.

What this workflow does
-----------------------
1) Ensures these mapping sheets include full source + LEAP universes with blank
   counterpart columns when unmapped:
   - leap_mappings.xlsx :: fuel_ninth_final_proposed
   - leap_mappings.xlsx :: fuel_product_final_proposed
   - leap_mappings.xlsx :: sector_flow_final_proposed
   - leap_mappings.xlsx :: sector_ninth_final_proposed

2) Updates master_config.xlsx sheets:
   - sector_fuel_code_to_name
   - sector_fuel_ESTO_LEAP_names
   using LEAP names inferred from leap_combined_esto / leap_combined_ninth.

3) Writes analysis of ninth_pairs_to_esto_pairs carry-through vs implied mapping
   from joining leap_combined_ninth and leap_combined_esto on LEAP source pair.

Outputs
-------
- outputs/mappings/mapping_checks/mapping_alignment_summary.json
- outputs/mappings/mapping_checks/ninth_pairs_join_vs_master_status.csv
- outputs/mappings/mapping_checks/ninth_pairs_join_vs_master_mismatches.csv
- outputs/mappings/mapping_checks/sector_fuel_name_conflicts.csv

Backups
-------
Creates timestamped backups in:
- config/archive/leap_mappings.before_mapping_alignment_*.xlsx
- config/archive/master_config.before_mapping_alignment_*.xlsx
"""

from __future__ import annotations

import json
import re
import shutil
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows


REPO_ROOT = Path(__file__).resolve().parents[2]


def _resolve(path: Path | str) -> Path:
    raw = str(path).replace("\\", "/")
    candidate = Path(raw)
    return candidate if candidate.is_absolute() else (REPO_ROOT / candidate)


LEAP_MAPPINGS_PATH = _resolve("config/leap_mappings.xlsx")
MASTER_CONFIG_PATH = _resolve("config/master_config.xlsx")
ESTO_DATA_PATH = _resolve("data/00APEC_2025_low_with_subtotals.csv")
NINTH_DATA_PATH = _resolve("data/merged_file_energy_ALL_20251106.csv")

OUTPUT_DIR = _resolve("outputs/mappings/mapping_checks")
ARCHIVE_DIR = _resolve("config/archive")

LEAP_COMBINED_ESTO = "leap_combined_esto"
LEAP_COMBINED_NINTH = "leap_combined_ninth"

FUEL_NINTH_SHEET = "fuel_ninth_final_proposed"
FUEL_ESTO_SHEET = "fuel_product_final_proposed"
SECTOR_ESTO_SHEET = "sector_flow_final_proposed"
SECTOR_NINTH_SHEET = "sector_ninth_final_proposed"

MASTER_CODE_TO_NAME_SHEET = "sector_fuel_code_to_name"
MASTER_ESTO_LEAP_NAMES_SHEET = "sector_fuel_ESTO_LEAP_names"
MASTER_NINTH_ESTO_PAIRS_SHEET = "ninth_pairs_to_esto_pairs"


#%%
def _clean(value: object) -> str:
    text = str(value or "").strip()
    return "" if text.lower() in {"", "nan", "none", "<na>"} else text


def _norm(value: object) -> str:
    return " ".join(_clean(value).lower().split())


def _truthy(value: object) -> bool:
    return _norm(value) in {"1", "true", "t", "yes", "y", "on"}


def _active_mask(df: pd.DataFrame) -> pd.Series:
    remove_col = "remove_row" if "remove_row" in df.columns else None
    dup_col = "duplicate_to_remove" if "duplicate_to_remove" in df.columns else None
    remove_mask = df[remove_col].map(_truthy) if remove_col else pd.Series(False, index=df.index)
    dup_mask = df[dup_col].map(_truthy) if dup_col else pd.Series(False, index=df.index)
    return ~(remove_mask | dup_mask)


def _path_key(path: object) -> str:
    parts = [part.strip() for part in _clean(path).replace("\\", "/").split("/") if part.strip()]
    return "/".join(_norm(part) for part in parts)


def _sheet_df(path: Path, sheet_name: str) -> pd.DataFrame:
    return pd.read_excel(path, sheet_name=sheet_name, dtype=object).fillna("")


def _replace_sheet(path: Path, sheet_name: str, frame: pd.DataFrame) -> None:
    wb = load_workbook(path)
    if sheet_name in wb.sheetnames:
        idx = wb.sheetnames.index(sheet_name)
        del wb[sheet_name]
        ws = wb.create_sheet(sheet_name, idx)
    else:
        ws = wb.create_sheet(sheet_name)
    for row in dataframe_to_rows(frame, index=False, header=True):
        ws.append(row)
    wb.save(path)


def _backup(path: Path, suffix: str) -> Path:
    ARCHIVE_DIR.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup = ARCHIVE_DIR / f"{path.stem}.{suffix}_{timestamp}{path.suffix}"
    shutil.copy2(path, backup)
    return backup


#%%
def _extract_ninth_codes(df: pd.DataFrame) -> tuple[set[str], dict[str, str], set[str], dict[str, str]]:
    sector_cols = ["sectors", "sub1sectors", "sub2sectors", "sub3sectors", "sub4sectors"]
    fuel_cols = ["fuels", "subfuels"]
    for col in sector_cols + fuel_cols:
        if col not in df.columns:
            df[col] = ""
        df[col] = df[col].map(_clean)

    sectors: set[str] = set()
    fuels: set[str] = set()
    sector_to_col: dict[str, str] = {}
    fuel_to_col: dict[str, str] = {}

    for row in df.itertuples(index=False):
        values = {col: _clean(getattr(row, col, "")) for col in sector_cols + fuel_cols}

        deepest_sector = ""
        deepest_sector_col = ""
        for col in reversed(sector_cols):
            val = values[col]
            if val and _norm(val) != "x":
                deepest_sector = val
                deepest_sector_col = col
                break

        deepest_fuel = ""
        deepest_fuel_col = ""
        for col in reversed(fuel_cols):
            val = values[col]
            if val and _norm(val) != "x":
                deepest_fuel = val
                deepest_fuel_col = col
                break

        if deepest_sector:
            sectors.add(deepest_sector)
            sector_to_col.setdefault(deepest_sector, deepest_sector_col)
        if deepest_fuel:
            fuels.add(deepest_fuel)
            fuel_to_col.setdefault(deepest_fuel, deepest_fuel_col)

    return sectors, sector_to_col, fuels, fuel_to_col


def _extract_esto_codes(df: pd.DataFrame) -> tuple[set[str], set[str]]:
    flows = set(df.get("flows", pd.Series(dtype=str)).map(_clean))
    products = set(df.get("products", pd.Series(dtype=str)).map(_clean))
    flows.discard("")
    products.discard("")
    return flows, products


def _build_esto_pair_abs_sum_lookup(df: pd.DataFrame) -> dict[tuple[str, str], float]:
    work = df.copy().fillna("")
    for col in ["flows", "products"]:
        if col not in work.columns:
            work[col] = ""
        work[col] = work[col].map(_clean)

    year_cols = [col for col in work.columns if str(col).isdigit()]
    if not year_cols:
        return {}

    numeric = work[year_cols].apply(pd.to_numeric, errors="coerce").fillna(0.0)
    work["_abs_sum_row"] = numeric.abs().sum(axis=1)
    grouped = (
        work[(work["flows"] != "") & (work["products"] != "")]
        .groupby(["flows", "products"], as_index=False)["_abs_sum_row"]
        .sum()
    )
    return {
        (_clean(row["flows"]), _clean(row["products"])): float(row["_abs_sum_row"])
        for _, row in grouped.iterrows()
    }


def _build_ninth_pair_abs_sum_lookup(df: pd.DataFrame) -> dict[tuple[str, str], float]:
    work = df.copy().fillna("")
    sector_cols = ["sectors", "sub1sectors", "sub2sectors", "sub3sectors", "sub4sectors"]
    fuel_cols = ["fuels", "subfuels"]
    for col in sector_cols + fuel_cols:
        if col not in work.columns:
            work[col] = ""
        work[col] = work[col].map(_clean)

    year_cols = [col for col in work.columns if str(col).isdigit()]
    if not year_cols:
        return {}
    numeric = work[year_cols].apply(pd.to_numeric, errors="coerce").fillna(0.0)
    work["_abs_sum_row"] = numeric.abs().sum(axis=1)

    def _deepest_non_x(row: pd.Series, cols: list[str]) -> str:
        for col in reversed(cols):
            value = _clean(row.get(col, ""))
            if value and _norm(value) != "x":
                return value
        return ""

    work["ninth_sector"] = work.apply(lambda row: _deepest_non_x(row, sector_cols), axis=1)
    work["ninth_fuel"] = work.apply(lambda row: _deepest_non_x(row, fuel_cols), axis=1)

    grouped = (
        work[(work["ninth_sector"] != "") & (work["ninth_fuel"] != "")]
        .groupby(["ninth_sector", "ninth_fuel"], as_index=False)["_abs_sum_row"]
        .sum()
    )
    return {
        (_clean(row["ninth_sector"]), _clean(row["ninth_fuel"])): float(row["_abs_sum_row"])
        for _, row in grouped.iterrows()
    }


def _pair_counts(df: pd.DataFrame, left_col: str, right_col: str) -> tuple[dict[str, int], dict[str, int]]:
    left_counts: dict[str, int] = defaultdict(int)
    right_counts: dict[str, int] = defaultdict(int)
    work = df.copy()
    work[left_col] = work[left_col].map(_clean)
    work[right_col] = work[right_col].map(_clean)
    work = work[(work[left_col] != "") & (work[right_col] != "")].drop_duplicates(subset=[left_col, right_col])

    for left, grp in work.groupby(left_col):
        left_counts[left] = int(grp[right_col].nunique())
    for right, grp in work.groupby(right_col):
        right_counts[right] = int(grp[left_col].nunique())
    return left_counts, right_counts


def _map_cardinality(left_target_count: int, right_source_count: int) -> str:
    if left_target_count <= 0 or right_source_count <= 0:
        return ""
    if left_target_count == 1 and right_source_count == 1:
        return "one_to_one"
    if left_target_count > 1 and right_source_count == 1:
        return "one_to_many"
    if left_target_count == 1 and right_source_count > 1:
        return "many_to_one"
    return "many_to_many"


def _ensure_two_col_coverage(
    frame: pd.DataFrame,
    left_col: str,
    right_col: str,
    left_universe: set[str],
    right_universe: set[str],
) -> tuple[pd.DataFrame, dict[str, int]]:
    out = frame.copy().fillna("")
    if left_col not in out.columns:
        out[left_col] = ""
    if right_col not in out.columns:
        out[right_col] = ""
    out[left_col] = out[left_col].map(_clean)
    out[right_col] = out[right_col].map(_clean)

    existing_left = {v for v in out[left_col].tolist() if v}
    existing_right = {v for v in out[right_col].tolist() if v}

    missing_left = sorted(v for v in left_universe if v and v not in existing_left)
    missing_right = sorted(v for v in right_universe if v and v not in existing_right)

    template = {col: "" for col in out.columns}
    add_rows = []
    for value in missing_left:
        row = template.copy()
        row[left_col] = value
        add_rows.append(row)
    for value in missing_right:
        row = template.copy()
        row[right_col] = value
        add_rows.append(row)

    if add_rows:
        out = pd.concat([out, pd.DataFrame(add_rows)], ignore_index=True)

    left_counts, right_counts = _pair_counts(out, left_col, right_col)
    if "mapping_cardinality" in out.columns:
        out["mapping_cardinality"] = out.apply(
            lambda row: _map_cardinality(
                left_counts.get(_clean(row[left_col]), 0),
                right_counts.get(_clean(row[right_col]), 0),
            )
            if _clean(row[left_col]) and _clean(row[right_col])
            else "",
            axis=1,
        )

    sort_cols = [c for c in [left_col, right_col] if c in out.columns]
    if sort_cols:
        out = out.sort_values(
            sort_cols,
            key=lambda col: col.map(_clean).str.lower(),
            kind="stable",
            na_position="last",
        ).reset_index(drop=True)

    diagnostics = {
        "added_missing_left": len(missing_left),
        "added_missing_right": len(missing_right),
        "final_rows": len(out),
    }
    return out, diagnostics


#%%
def _build_name_lookups(leap_esto: pd.DataFrame, leap_ninth: pd.DataFrame) -> dict[str, dict[str, str]]:
    lookups = {
        "esto_flow": {},
        "esto_product": {},
        "ninth_sector": {},
        "ninth_fuel": {},
    }

    def _choose_name(counter: Counter[str]) -> str:
        if not counter:
            return ""
        return sorted(counter.items(), key=lambda item: (-item[1], item[0].lower()))[0][0]

    for target_col, name_col, source_df in [
        ("esto_flow", "leap_sector_name_original", leap_esto),
        ("esto_product", "raw_leap_fuel_name", leap_esto),
        ("ninth_sector", "leap_sector_name_original", leap_ninth),
        ("ninth_fuel", "raw_leap_fuel_name", leap_ninth),
    ]:
        counts: dict[str, Counter[str]] = defaultdict(Counter)
        work = source_df.copy()
        if target_col not in work.columns or name_col not in work.columns:
            continue
        if "remove_row" in work.columns or "duplicate_to_remove" in work.columns:
            work = work[_active_mask(work)].copy()
        work[target_col] = work[target_col].map(_clean)
        work[name_col] = work[name_col].map(_clean)
        work = work[(work[target_col] != "") & (work[name_col] != "")]
        for _, row in work.iterrows():
            counts[row[target_col]][row[name_col]] += 1
        lookups[target_col] = {code: _choose_name(counter) for code, counter in counts.items()}

    return lookups


def _update_sector_fuel_code_to_name(
    frame: pd.DataFrame,
    name_lookups: dict[str, dict[str, str]],
    ninth_sector_col_map: dict[str, str],
    ninth_fuel_col_map: dict[str, str],
    esto_flows: set[str],
    esto_products: set[str],
    ninth_sectors: set[str],
    ninth_fuels: set[str],
) -> tuple[pd.DataFrame, pd.DataFrame, dict[str, int]]:
    out = frame.copy().fillna("")
    for col in ["9th_label", "9th_column", "esto_label", "esto_column", "name"]:
        if col not in out.columns:
            out[col] = ""
        out[col] = out[col].map(_clean)

    conflicts: list[dict[str, str]] = []
    updates = 0

    for idx, row in out.iterrows():
        ninth_label = _clean(row.get("9th_label", ""))
        esto_label = _clean(row.get("esto_label", ""))
        existing = _clean(row.get("name", ""))

        candidates: list[str] = []
        if ninth_label in name_lookups["ninth_sector"]:
            candidates.append(name_lookups["ninth_sector"][ninth_label])
        if ninth_label in name_lookups["ninth_fuel"]:
            candidates.append(name_lookups["ninth_fuel"][ninth_label])
        if esto_label in name_lookups["esto_flow"]:
            candidates.append(name_lookups["esto_flow"][esto_label])
        if esto_label in name_lookups["esto_product"]:
            candidates.append(name_lookups["esto_product"][esto_label])
        candidates = [c for c in candidates if c]

        chosen = ""
        if candidates:
            if existing and any(_norm(existing) == _norm(c) for c in candidates):
                chosen = existing
            elif existing and len({_norm(c) for c in candidates}) > 1:
                chosen = existing
                conflicts.append(
                    {
                        "row_number": str(idx + 2),
                        "9th_label": ninth_label,
                        "esto_label": esto_label,
                        "existing_name": existing,
                        "candidate_names": " | ".join(sorted(set(candidates))),
                        "decision": "kept_existing_due_to_conflict",
                    }
                )
            else:
                chosen = sorted(candidates, key=lambda x: (x.lower(), x))[0]
        else:
            chosen = existing

        if chosen != existing:
            out.at[idx, "name"] = chosen
            updates += 1

    # Add missing ninth sector/fuel codes
    existing_9th = {_clean(v) for v in out["9th_label"].tolist() if _clean(v)}
    add_rows: list[dict[str, str]] = []

    for code in sorted((ninth_sectors | ninth_fuels) - existing_9th):
        row = {col: "" for col in out.columns}
        row["9th_label"] = code
        row["9th_column"] = ninth_sector_col_map.get(code) or ninth_fuel_col_map.get(code) or ""
        name = (
            name_lookups["ninth_sector"].get(code)
            or name_lookups["ninth_fuel"].get(code)
            or ""
        )
        row["name"] = name
        add_rows.append(row)

    # Add missing ESTO flow/product labels
    existing_esto = {_clean(v) for v in out["esto_label"].tolist() if _clean(v)}
    for code in sorted((esto_flows | esto_products) - existing_esto):
        row = {col: "" for col in out.columns}
        row["esto_label"] = code
        row["esto_column"] = "flows" if code in esto_flows else "products" if code in esto_products else ""
        name = name_lookups["esto_flow"].get(code) or name_lookups["esto_product"].get(code) or ""
        row["name"] = name
        add_rows.append(row)

    if add_rows:
        out = pd.concat([out, pd.DataFrame(add_rows)], ignore_index=True)

    out = out.sort_values(
        ["9th_label", "esto_label", "name"],
        key=lambda col: col.map(_clean).str.lower(),
        kind="stable",
        na_position="last",
    ).reset_index(drop=True)

    conflict_df = pd.DataFrame(conflicts, columns=[
        "row_number", "9th_label", "esto_label", "existing_name", "candidate_names", "decision"
    ])

    diagnostics = {
        "name_updates": updates,
        "added_missing_code_rows": len(add_rows),
        "conflict_rows": len(conflict_df),
        "final_rows": len(out),
    }
    return out, conflict_df, diagnostics


def _update_sector_fuel_esto_leap_names(
    frame: pd.DataFrame,
    name_lookups: dict[str, dict[str, str]],
    esto_flows: set[str],
    esto_products: set[str],
) -> tuple[pd.DataFrame, dict[str, int]]:
    out = frame.copy().fillna("")
    for col in ["category", "original_label", "leap_name"]:
        if col not in out.columns:
            out[col] = ""
        out[col] = out[col].map(_clean)

    updates = 0
    for idx, row in out.iterrows():
        category = _norm(row.get("category", ""))
        original = _clean(row.get("original_label", ""))
        existing = _clean(row.get("leap_name", ""))

        suggested = ""
        if category == "flows":
            suggested = name_lookups["esto_flow"].get(original, "")
        elif category == "products":
            suggested = name_lookups["esto_product"].get(original, "")

        if suggested and _norm(existing) != _norm(suggested):
            out.at[idx, "leap_name"] = suggested
            updates += 1

    existing_keys = {
        (_norm(r["category"]), _clean(r["original_label"]))
        for _, r in out.iterrows()
        if _clean(r.get("original_label", ""))
    }

    add_rows: list[dict[str, str]] = []
    for flow in sorted(esto_flows):
        key = ("flows", flow)
        if key not in existing_keys:
            row = {col: "" for col in out.columns}
            row["category"] = "flows"
            row["original_label"] = flow
            row["leap_name"] = name_lookups["esto_flow"].get(flow, "")
            add_rows.append(row)
    for product in sorted(esto_products):
        key = ("products", product)
        if key not in existing_keys:
            row = {col: "" for col in out.columns}
            row["category"] = "products"
            row["original_label"] = product
            row["leap_name"] = name_lookups["esto_product"].get(product, "")
            add_rows.append(row)

    if add_rows:
        out = pd.concat([out, pd.DataFrame(add_rows)], ignore_index=True)

    out = out.sort_values(
        ["category", "original_label", "leap_name"],
        key=lambda col: col.map(_clean).str.lower(),
        kind="stable",
        na_position="last",
    ).reset_index(drop=True)

    diagnostics = {
        "name_updates": updates,
        "added_missing_rows": len(add_rows),
        "final_rows": len(out),
    }
    return out, diagnostics


#%%
def _active_complete(df: pd.DataFrame, required: list[str]) -> pd.DataFrame:
    work = df.copy().fillna("")
    for col in required:
        if col not in work.columns:
            work[col] = ""
        work[col] = work[col].map(_clean)
    work = work[_active_mask(work)] if ("remove_row" in work.columns or "duplicate_to_remove" in work.columns) else work
    for col in required:
        work = work[work[col] != ""]
    return work


def _build_ninth_esto_pair_analysis(
    leap_esto: pd.DataFrame,
    leap_ninth: pd.DataFrame,
    ninth_pairs_master: pd.DataFrame,
    esto_pair_abs_sum_lookup: dict[tuple[str, str], float],
    ninth_pair_abs_sum_lookup: dict[tuple[str, str], float],
) -> tuple[pd.DataFrame, pd.DataFrame, dict[str, int]]:
    esto_req = ["leap_sector_name_full_path", "raw_leap_fuel_name", "esto_flow", "esto_product"]
    ninth_req = ["leap_sector_name_full_path", "raw_leap_fuel_name", "ninth_sector", "ninth_fuel"]

    esto = _active_complete(leap_esto, esto_req).copy()
    ninth = _active_complete(leap_ninth, ninth_req).copy()

    esto["source_key"] = list(zip(esto["leap_sector_name_full_path"].map(_path_key), esto["raw_leap_fuel_name"].map(_norm)))
    ninth["source_key"] = list(zip(ninth["leap_sector_name_full_path"].map(_path_key), ninth["raw_leap_fuel_name"].map(_norm)))
    esto = esto[esto["source_key"].map(lambda k: bool(k[0] and k[1]))]
    ninth = ninth[ninth["source_key"].map(lambda k: bool(k[0] and k[1]))]

    join_df = ninth[["source_key", "ninth_sector", "ninth_fuel"]].drop_duplicates().merge(
        esto[["source_key", "esto_flow", "esto_product"]].drop_duplicates(),
        on="source_key",
        how="inner",
    )

    implied_pairs = join_df[["ninth_sector", "ninth_fuel", "esto_flow", "esto_product"]].drop_duplicates().copy()

    master = ninth_pairs_master.copy().fillna("")
    for col in ["9th_sector", "9th_fuel", "esto_flow", "esto_product"]:
        if col not in master.columns:
            master[col] = ""
        master[col] = master[col].map(_clean)

    master_pairs = master[
        (master["9th_sector"] != "")
        & (master["9th_fuel"] != "")
        & (master["esto_flow"] != "")
        & (master["esto_product"] != "")
    ][["9th_sector", "9th_fuel", "esto_flow", "esto_product"]].drop_duplicates().copy()
    master_pairs = master_pairs.rename(columns={"9th_sector": "ninth_sector", "9th_fuel": "ninth_fuel"})

    implied_key = set(
        zip(
            implied_pairs["ninth_sector"],
            implied_pairs["ninth_fuel"],
            implied_pairs["esto_flow"],
            implied_pairs["esto_product"],
        )
    )
    master_key = set(
        zip(
            master_pairs["ninth_sector"],
            master_pairs["ninth_fuel"],
            master_pairs["esto_flow"],
            master_pairs["esto_product"],
        )
    )

    rows: list[dict[str, str]] = []
    for pair in sorted(implied_key | master_key):
        in_implied = pair in implied_key
        in_master = pair in master_key
        status = "in_both" if in_implied and in_master else "only_in_join_implied" if in_implied else "only_in_master"
        ninth_key = (_clean(pair[0]), _clean(pair[1]))
        esto_key = (_clean(pair[2]), _clean(pair[3]))
        rows.append(
            {
                "ninth_sector": pair[0],
                "ninth_fuel": pair[1],
                "esto_flow": pair[2],
                "esto_product": pair[3],
                "status": status,
                "esto_all_years_abs_sum": float(esto_pair_abs_sum_lookup.get(esto_key, 0.0)),
                "ninth_all_years_abs_sum": float(ninth_pair_abs_sum_lookup.get(ninth_key, 0.0)),
            }
        )

    status_df = pd.DataFrame(rows).sort_values(
        ["status", "ninth_sector", "ninth_fuel", "esto_flow", "esto_product"],
        key=lambda col: col.map(_clean).str.lower(),
        kind="stable",
        na_position="last",
    ).reset_index(drop=True)

    mismatches_df = status_df[status_df["status"].isin(["only_in_join_implied", "only_in_master"])].copy()

    diagnostics = {
        "join_implied_pairs": len(implied_key),
        "master_pairs": len(master_key),
        "pairs_in_both": int((status_df["status"] == "in_both").sum()),
        "only_in_join_implied": int((status_df["status"] == "only_in_join_implied").sum()),
        "only_in_master": int((status_df["status"] == "only_in_master").sum()),
        "only_in_join_implied_nonzero_esto": int(
            (
                (status_df["status"] == "only_in_join_implied")
                & (pd.to_numeric(status_df["esto_all_years_abs_sum"], errors="coerce").fillna(0.0) > 0)
            ).sum()
        ),
        "only_in_join_implied_nonzero_ninth": int(
            (
                (status_df["status"] == "only_in_join_implied")
                & (pd.to_numeric(status_df["ninth_all_years_abs_sum"], errors="coerce").fillna(0.0) > 0)
            ).sum()
        ),
        "only_in_master_nonzero_esto": int(
            (
                (status_df["status"] == "only_in_master")
                & (pd.to_numeric(status_df["esto_all_years_abs_sum"], errors="coerce").fillna(0.0) > 0)
            ).sum()
        ),
        "only_in_master_nonzero_ninth": int(
            (
                (status_df["status"] == "only_in_master")
                & (pd.to_numeric(status_df["ninth_all_years_abs_sum"], errors="coerce").fillna(0.0) > 0)
            ).sum()
        ),
    }
    return status_df, mismatches_df, diagnostics


#%%
def run_workflow() -> dict[str, object]:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    if not LEAP_MAPPINGS_PATH.exists():
        raise FileNotFoundError(f"Missing workbook: {LEAP_MAPPINGS_PATH}")
    if not MASTER_CONFIG_PATH.exists():
        raise FileNotFoundError(f"Missing workbook: {MASTER_CONFIG_PATH}")

    leap_backup = _backup(LEAP_MAPPINGS_PATH, "before_mapping_alignment")
    master_backup = _backup(MASTER_CONFIG_PATH, "before_mapping_alignment")

    leap_esto = _sheet_df(LEAP_MAPPINGS_PATH, LEAP_COMBINED_ESTO)
    leap_ninth = _sheet_df(LEAP_MAPPINGS_PATH, LEAP_COMBINED_NINTH)

    fuel_ninth = _sheet_df(LEAP_MAPPINGS_PATH, FUEL_NINTH_SHEET)
    fuel_esto = _sheet_df(LEAP_MAPPINGS_PATH, FUEL_ESTO_SHEET)
    sector_esto = _sheet_df(LEAP_MAPPINGS_PATH, SECTOR_ESTO_SHEET)
    sector_ninth = _sheet_df(LEAP_MAPPINGS_PATH, SECTOR_NINTH_SHEET)

    master_code_name = _sheet_df(MASTER_CONFIG_PATH, MASTER_CODE_TO_NAME_SHEET)
    master_esto_leap = _sheet_df(MASTER_CONFIG_PATH, MASTER_ESTO_LEAP_NAMES_SHEET)
    master_ninth_esto_pairs = _sheet_df(MASTER_CONFIG_PATH, MASTER_NINTH_ESTO_PAIRS_SHEET)

    esto_data = pd.read_csv(ESTO_DATA_PATH, dtype=object).fillna("")
    ninth_data = pd.read_csv(NINTH_DATA_PATH, dtype=object, low_memory=False).fillna("")

    esto_pair_abs_sum_lookup = _build_esto_pair_abs_sum_lookup(esto_data)
    ninth_pair_abs_sum_lookup = _build_ninth_pair_abs_sum_lookup(ninth_data)

    esto_flows, esto_products = _extract_esto_codes(esto_data)
    ninth_sectors, ninth_sector_col_map, ninth_fuels, ninth_fuel_col_map = _extract_ninth_codes(ninth_data)

    leap_fuels = {
        _clean(v)
        for v in pd.concat([
            leap_esto.get("raw_leap_fuel_name", pd.Series(dtype=object)),
            leap_ninth.get("raw_leap_fuel_name", pd.Series(dtype=object)),
        ], ignore_index=True).tolist()
        if _clean(v)
    }
    leap_sectors = {
        _clean(v)
        for v in pd.concat([
            leap_esto.get("leap_sector_name_full_path", pd.Series(dtype=object)),
            leap_ninth.get("leap_sector_name_full_path", pd.Series(dtype=object)),
        ], ignore_index=True).tolist()
        if _clean(v)
    }

    fuel_ninth_updated, d_fuel_ninth = _ensure_two_col_coverage(
        fuel_ninth,
        left_col="leap_fuel_name",
        right_col="ninth_fuel",
        left_universe=leap_fuels,
        right_universe=ninth_fuels,
    )
    fuel_esto_updated, d_fuel_esto = _ensure_two_col_coverage(
        fuel_esto,
        left_col="leap_fuel_name",
        right_col="esto_product",
        left_universe=leap_fuels,
        right_universe=esto_products,
    )
    sector_esto_updated, d_sector_esto = _ensure_two_col_coverage(
        sector_esto,
        left_col="leap_sector_name",
        right_col="esto_flow",
        left_universe=leap_sectors,
        right_universe=esto_flows,
    )
    sector_ninth_updated, d_sector_ninth = _ensure_two_col_coverage(
        sector_ninth,
        left_col="leap_sector_name",
        right_col="ninth_sector",
        left_universe=leap_sectors,
        right_universe=ninth_sectors,
    )

    _replace_sheet(LEAP_MAPPINGS_PATH, FUEL_NINTH_SHEET, fuel_ninth_updated)
    _replace_sheet(LEAP_MAPPINGS_PATH, FUEL_ESTO_SHEET, fuel_esto_updated)
    _replace_sheet(LEAP_MAPPINGS_PATH, SECTOR_ESTO_SHEET, sector_esto_updated)
    _replace_sheet(LEAP_MAPPINGS_PATH, SECTOR_NINTH_SHEET, sector_ninth_updated)

    name_lookups = _build_name_lookups(leap_esto, leap_ninth)

    code_name_updated, conflict_df, d_code_name = _update_sector_fuel_code_to_name(
        master_code_name,
        name_lookups,
        ninth_sector_col_map,
        ninth_fuel_col_map,
        esto_flows,
        esto_products,
        ninth_sectors,
        ninth_fuels,
    )
    esto_leap_updated, d_esto_leap = _update_sector_fuel_esto_leap_names(
        master_esto_leap,
        name_lookups,
        esto_flows,
        esto_products,
    )

    _replace_sheet(MASTER_CONFIG_PATH, MASTER_CODE_TO_NAME_SHEET, code_name_updated)
    _replace_sheet(MASTER_CONFIG_PATH, MASTER_ESTO_LEAP_NAMES_SHEET, esto_leap_updated)

    status_df, mismatch_df, d_pair_analysis = _build_ninth_esto_pair_analysis(
        leap_esto,
        leap_ninth,
        master_ninth_esto_pairs,
        esto_pair_abs_sum_lookup,
        ninth_pair_abs_sum_lookup,
    )

    status_path = OUTPUT_DIR / "ninth_pairs_join_vs_master_status.csv"
    mismatch_path = OUTPUT_DIR / "ninth_pairs_join_vs_master_mismatches.csv"
    conflicts_path = OUTPUT_DIR / "sector_fuel_name_conflicts.csv"
    summary_path = OUTPUT_DIR / "mapping_alignment_summary.json"

    status_df.to_csv(status_path, index=False)
    mismatch_df.to_csv(mismatch_path, index=False)
    conflict_df.to_csv(conflicts_path, index=False)

    summary = {
        "timestamp": datetime.now().isoformat(timespec="seconds"),
        "backups": {
            "leap_mappings": str(leap_backup),
            "master_config": str(master_backup),
        },
        "coverage_updates": {
            FUEL_NINTH_SHEET: d_fuel_ninth,
            FUEL_ESTO_SHEET: d_fuel_esto,
            SECTOR_ESTO_SHEET: d_sector_esto,
            SECTOR_NINTH_SHEET: d_sector_ninth,
        },
        "master_updates": {
            MASTER_CODE_TO_NAME_SHEET: d_code_name,
            MASTER_ESTO_LEAP_NAMES_SHEET: d_esto_leap,
        },
        "ninth_pairs_join_analysis": d_pair_analysis,
        "outputs": {
            "status_csv": str(status_path),
            "mismatch_csv": str(mismatch_path),
            "name_conflicts_csv": str(conflicts_path),
        },
    }

    with open(summary_path, "w", encoding="utf-8") as f:
        json.dump(summary, f, indent=2)

    return summary


#%%
if __name__ == "__main__":
    result = run_workflow()
    print("[OK] Mapping alignment workflow completed")
    print(json.dumps(result, indent=2))
#%%
