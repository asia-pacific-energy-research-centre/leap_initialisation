"""
One-off script: backfill missing Level 1..N columns in the
leap_import_baseline_seed_*_20260603.xlsx files.

Rows where Level 1 is null have their Level columns filled from
the Branch Path value (split on backslash). Rows that already
have Level 1 filled are left untouched. Both LEAP and FOR_VIEWING
sheets are updated in-place.

Run from the repo root:
    python backfill_level_cols_20260603.py
"""
from __future__ import annotations

import sys
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

TARGET_DIR = Path(
    r"C:\Users\Work\github\leap_utilities\outputs\leap_exports\supply_reconciliation"
)
FILE_PATTERN = "leap_import_baseline_seed_*_20260603.xlsx"
SHEETS = ["LEAP", "FOR_VIEWING"]


def _fill_level_cols(df: pd.DataFrame) -> tuple[pd.DataFrame, int]:
    """
    Fill Level N columns from Branch Path for rows where Level 1 is null.
    Returns the updated DataFrame and a count of rows that were patched.
    """
    level_cols = sorted(
        [c for c in df.columns if str(c).startswith("Level ")],
        key=lambda c: int(str(c).split()[-1]),
    )
    if not level_cols:
        return df, 0

    null_mask = df["Level 1"].isnull() & df["Branch Path"].notna()
    patched = int(null_mask.sum())
    if patched == 0:
        return df, 0

    for idx, col in enumerate(level_cols):
        df.loc[null_mask, col] = df.loc[null_mask, "Branch Path"].apply(
            lambda path, i=idx: str(path).split("\\")[i]
            if len(str(path).split("\\")) > i
            else pd.NA
        )
    return df, patched


def _find_header_row(ws) -> int | None:
    """Return the 0-based row index of the LEAP column-header row."""
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if row and "Branch Path" in row:
            return i
    return None


def process_file(path: Path) -> None:
    xl = pd.ExcelFile(path, engine="openpyxl")
    sheets_present = [s for s in SHEETS if s in xl.sheet_names]
    if not sheets_present:
        print(f"  [SKIP] No matching sheets in {path.name}")
        return

    # Read both sheets, patch, write back with openpyxl to preserve other content
    wb = load_workbook(path)
    total_patched = 0

    for sheet in sheets_present:
        ws = wb[sheet]
        header_row_idx = _find_header_row(ws)
        if header_row_idx is None:
            print(f"  [WARN] Could not find header row in sheet '{sheet}' of {path.name}")
            continue

        # Pull into DataFrame starting at the header row
        data = list(ws.iter_rows(min_row=header_row_idx + 1, values_only=True))
        headers = data[0]
        rows = data[1:]
        df = pd.DataFrame(rows, columns=headers)

        # Convert column name to position mapping (0-based within the sheet row)
        col_positions = {name: i for i, name in enumerate(headers) if name is not None}

        if "Branch Path" not in col_positions or "Level 1" not in col_positions:
            print(f"  [WARN] Missing Branch Path or Level 1 in sheet '{sheet}' of {path.name}")
            continue

        df, patched = _fill_level_cols(df)
        total_patched += patched

        if patched == 0:
            continue

        # Write patched values back into the worksheet
        level_cols = sorted(
            [c for c in df.columns if str(c).startswith("Level ")],
            key=lambda c: int(str(c).split()[-1]),
        )
        data_start_row = header_row_idx + 2  # 1-based openpyxl row of first data row

        for df_row_idx, row in df.iterrows():
            ws_row = data_start_row + df_row_idx
            for col_name in level_cols:
                col_idx_0 = col_positions.get(col_name)
                if col_idx_0 is None:
                    continue
                ws.cell(row=ws_row, column=col_idx_0 + 1).value = (
                    row[col_name] if pd.notna(row[col_name]) else None
                )

    if total_patched > 0:
        wb.save(path)
        print(f"  [OK] {path.name}: patched {total_patched} cell-rows across sheets")
    else:
        print(f"  [SKIP] {path.name}: no null Level 1 rows found, nothing to do")


def main() -> None:
    files = sorted(TARGET_DIR.glob(FILE_PATTERN))
    if not files:
        print(f"No files matched {FILE_PATTERN} in {TARGET_DIR}")
        sys.exit(1)

    print(f"Found {len(files)} files to process.\n")
    for f in files:
        process_file(f)
    print("\nDone.")


if __name__ == "__main__":
    main()
