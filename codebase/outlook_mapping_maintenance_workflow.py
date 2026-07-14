#%%

"""
Maintain Outlook mapping audit columns and review outputs.

This workflow recomputes the lightweight audit columns used to maintain
`leap_combined_esto` and `leap_combined_ninth` without rerunning the full
dashboard process. It also builds mapping review outputs for crosswalk gaps,
target conflicts, duplicate rows, and researcher-facing checks.
"""

from __future__ import annotations

import os
import re
import shutil
import sys
import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path
from typing import Sequence

import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

from codebase.mapping_tools.excel_sheet_utils import safe_excel_sheet_name  # noqa: E402

REPO_ROOT = Path(__file__).resolve().parents[1]
try:
    if str(REPO_ROOT) not in sys.path:
        sys.path.insert(0, str(REPO_ROOT))
except Exception as exc:
    print(f"Failed to add repo root to sys.path: {exc}")

from codebase.utilities.master_config import OUTLOOK_MAPPINGS_MASTER_PATH  # noqa: E402
from codebase.utilities.leap_balance_export_resolver import resolve_balance_export_workbook  # noqa: E402
from codebase.utilities.energy_balance_template_extractor import (  # noqa: E402
    TemplateBalanceExtractor,
    _parse_unit_factor_to_petajoule,
)
from codebase.utilities.leap_results_dashboard_balance import _list_balance_sheets, _pick_template_sheet  # noqa: E402
from codebase.functions.outlook_mapping_maintenance_utils import (  # noqa: E402
    _clean,
    _norm_text,
    _truthy,
    _path_key,
    _mapping_cardinality,
    _subtotal_alignment,
    _active_mask,
    _drop_unnamed_columns,
    _drop_columns_if_present,
    _reorder_columns,
    _compute_leap_subtotals,
    _compute_pair_cardinality,
    _apply_auto_remove_rules,
    _refresh_esto_sheet,
    _refresh_ninth_sheet,
    _active_pairs,
    _active_leap_source_pairs,
    _leap_source_pair_presence_lookup,
    _build_duplicate_mappings,
    _build_trio_presence_check,
    _active_mapping_rows,
    _build_many_to_many_conflicts,
    _build_missing_between_sheet_conflicts,
    _build_crosswalk_target_conflicts,
    _active_ninth_to_esto_pairs,
    _build_implied_missing_crosswalk_pairs,
    build_mapping_conflict_report,
    _filter_researcher_rows,
    _pair_cardinality_for_columns,
    _researcher_export_frame,
)


#%%
def _resolve(path: Path | str) -> Path:
    raw = str(path).replace("\\", "/")
    drive_match = re.match(r"^([a-zA-Z]):/(.*)$", raw)
    if drive_match:
        drive = drive_match.group(1).lower()
        rest = drive_match.group(2)
        if os.name == "nt":
            return Path(f"{drive.upper()}:/{rest}")
        return Path(f"/mnt/{drive}/{rest}")
    candidate = Path(raw)
    return candidate if candidate.is_absolute() else (REPO_ROOT / candidate)


MAPPING_WORKBOOK_PATH = OUTLOOK_MAPPINGS_MASTER_PATH
CODEBOOK_PATH = OUTLOOK_MAPPINGS_MASTER_PATH
ESTO_TABLE_PATH = _resolve("data/00APEC_2025_low_with_subtotals.csv")
NINTH_TABLE_PATH = _resolve("data/merged_file_energy_ALL_20251106.csv")
OUTPUT_DIR = _resolve("outputs/mappings/mapping_checks")
RESEARCHER_MAPPINGS_PATH = _resolve("outputs/mappings/researcher_mappings.xlsx")
MISSING_PAIRS_CSV_PATH = OUTPUT_DIR / "leap_mapping_missing_pairs.csv"
DUPLICATE_MAPPINGS_CSV_PATH = OUTPUT_DIR / "leap_mapping_duplicate_mappings.csv"
TRIO_PRESENCE_CSV_PATH = OUTPUT_DIR / "leap_mapping_trio_presence_check.csv"
MAPPING_CONFLICTS_WORKBOOK_PATH = OUTPUT_DIR / "leap_mapping_conflicts.xlsx"
MAPPING_REFRESH_REPORT_PATH = OUTPUT_DIR / "leap_mapping_refresh_report.xlsx"

TRIO_PRESENCE_OUTPUT_NOTE = (
    "Filter trio_presence_csv by presence_status first. Focus most on "
    "ninth_active_esto_removed, esto_active_ninth_removed, "
    "esto_removed_ninth_active, and ninth_active_esto_missing; these are mapped "
    "rows that can change expected dashboard results. Treat fuel=Total rows, "
    "same-target parent/child mappings, old incorrect fuel rows, expected "
    "losses-sector removals like 10.01.02/10.01.03, and detailed transport "
    "rows kept remove_row=True as low priority. Sort previous_runs ascending "
    "to find rows that are new or least frequently repeated across refreshes."
)

ESTO_SHEET = "leap_combined_esto"
NINTH_SHEET = "leap_combined_ninth"
SECTOR_FUEL_CODE_TO_NAME_SHEET = "leap_display_names"
NINTH_PAIRS_TO_ESTO_PAIRS_SHEET = "ninth_pairs_to_esto_pairs"

BASE_YEAR = 2022
PROJECTION_YEARS: Sequence[int] = tuple(range(2023, 2061))
PROJECTION_SCENARIOS: Sequence[str] = ("reference", "target")
BALANCE_EXPORT_ECONOMY = "20_USA"
REF_BALANCE_EXPORT_DATE_ID: str | None = None
TGT_BALANCE_EXPORT_DATE_ID: str | None = None
BALANCE_TEMPLATE_SHEET = "EBal|2060"



#%%
def _load_esto_lookup() -> pd.DataFrame:
    base_df = pd.read_csv(ESTO_TABLE_PATH)
    work = base_df.copy()
    if "is_subtotal" not in work.columns:
        work["is_subtotal"] = False
    for col in ["economy", "flows", "products", str(BASE_YEAR), "is_subtotal"]:
        if col not in work.columns:
            work[col] = ""
    work["esto_flow"] = work["flows"].fillna("").astype(str).str.strip()
    work["esto_product"] = work["products"].fillna("").astype(str).str.strip()
    work["value"] = pd.to_numeric(work[str(BASE_YEAR)], errors="coerce").fillna(0.0)
    work["is_subtotal"] = work["is_subtotal"].fillna(False).map(_truthy)
    work = work[work["esto_flow"].ne("") & work["esto_product"].ne("")].copy()
    grouped = (
        work.groupby(["esto_flow", "esto_product"], as_index=False)
        .agg(
            pair_value_sum=("value", "sum"),
            esto_pair_is_subtotal=("is_subtotal", "max"),
        )
        .reset_index(drop=True)
    )
    grouped["esto_pair_abs_sum"] = grouped["pair_value_sum"].abs()
    return grouped


def _load_ninth_lookup() -> pd.DataFrame:
    ninth_df = pd.read_csv(NINTH_TABLE_PATH)
    work = ninth_df.copy()
    for col in [
        "economy",
        "scenarios",
        "sectors",
        "sub1sectors",
        "sub2sectors",
        "sub3sectors",
        "sub4sectors",
        "fuels",
        "subfuels",
        "subtotal_layout",
        "subtotal_results",
    ]:
        if col not in work.columns:
            work[col] = ""
    for col in ["subtotal_layout", "subtotal_results"]:
        work[col] = work[col].fillna(False).map(_truthy)
    scenario_set = {str(value).strip().lower() for value in PROJECTION_SCENARIOS}
    work = work[work["scenarios"].fillna("").astype(str).str.strip().str.lower().isin(scenario_set)].copy()
    year_cols = [str(year) for year in PROJECTION_YEARS if str(year) in work.columns]
    if not year_cols or work.empty:
        return pd.DataFrame(
            columns=[
                "ninth_sector",
                "ninth_fuel",
                "ninth_pair_is_subtotal",
                "ninth_pair_abs_sum",
            ]
        )
    values = work[year_cols].apply(pd.to_numeric, errors="coerce").fillna(0.0)
    work["ninth_sector"] = work.apply(
        lambda row: next(
            (
                _clean(row.get(col, ""))
                for col in ["sub4sectors", "sub3sectors", "sub2sectors", "sub1sectors", "sectors"]
                if _clean(row.get(col, ""))
            ),
            "",
        ),
        axis=1,
    )
    work["ninth_fuel"] = work.apply(
        lambda row: next(
            (
                _clean(row.get(col, ""))
                for col in ["subfuels", "fuels"]
                if _clean(row.get(col, ""))
            ),
            "",
        ),
        axis=1,
    )
    work["value_abs_sum_row"] = values.abs().sum(axis=1)
    work = work[work["ninth_sector"].ne("") & work["ninth_fuel"].ne("")].copy()
    grouped = (
        work.groupby(["ninth_sector", "ninth_fuel"], as_index=False)
        .agg(
            subtotal_layout=("subtotal_layout", "max"),
            subtotal_results=("subtotal_results", "max"),
            ninth_pair_abs_sum=("value_abs_sum_row", "sum"),
        )
        .reset_index(drop=True)
    )
    grouped["ninth_pair_is_subtotal"] = (
        grouped["subtotal_layout"].fillna(False).astype(bool)
        | grouped["subtotal_results"].fillna(False).astype(bool)
    )
    return grouped


def _write_mapping_conflict_report(
    report_sheets: dict[str, pd.DataFrame],
    output_path: Path = MAPPING_CONFLICTS_WORKBOOK_PATH,
) -> None:
    """Write the mapping conflict report workbook."""
    output_path.parent.mkdir(parents=True, exist_ok=True)
    if output_path.exists():
        _assert_not_open(output_path)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        used_sheet_names: set[str] = set()
        for sheet_name, sheet_df in report_sheets.items():
            safe_sheet_name = safe_excel_sheet_name(sheet_name, used_sheet_names)
            sheet_df.to_excel(writer, sheet_name=safe_sheet_name, index=False)


def _write_excel_report(output_path: Path, sheets: dict[str, pd.DataFrame]) -> None:
    """Write an Excel workbook with basic filter/freeze formatting."""
    output_path.parent.mkdir(parents=True, exist_ok=True)
    if output_path.exists():
        _assert_not_open(output_path)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        used_sheet_names: set[str] = set()
        for sheet_name, sheet_df in sheets.items():
            safe_name = safe_excel_sheet_name(sheet_name, used_sheet_names)
            sheet_df.to_excel(writer, sheet_name=safe_name, index=False)
            worksheet = writer.book[safe_name]
            worksheet.freeze_panes = "A2"
            if worksheet.max_row >= 1 and worksheet.max_column >= 1:
                worksheet.auto_filter.ref = worksheet.dimensions
            for column_cells in worksheet.columns:
                header = str(column_cells[0].value or "")
                width = min(max(len(header) + 2, 12), 42)
                worksheet.column_dimensions[column_cells[0].column_letter].width = width


def _build_mapping_refresh_report_sheets(
    *,
    gaps: pd.DataFrame,
    duplicate_mappings: pd.DataFrame,
    trio_presence_issues: pd.DataFrame,
    mapping_conflict_report: dict[str, pd.DataFrame],
) -> dict[str, pd.DataFrame]:
    """Build the one-workbook mapping refresh report sheets."""
    conflict_summary = mapping_conflict_report.get("summary", pd.DataFrame())
    report_rows = [
        {
            "section": "coverage_gaps",
            "row_count": int(len(gaps)),
            "sheet_name": "coverage_gaps",
            "description": "Source or target pairs missing from active mapping rows.",
        },
        {
            "section": "duplicate_mappings",
            "row_count": int(len(duplicate_mappings)),
            "sheet_name": "duplicate_mappings",
            "description": "Exact duplicate active mapping rows.",
        },
        {
            "section": "trio_presence_issues",
            "row_count": int(len(trio_presence_issues)),
            "sheet_name": "trio_presence_issues",
            "description": "Active/removed/missing mismatches between leap_combined_esto and leap_combined_ninth.",
        },
    ]
    for row in conflict_summary.to_dict("records"):
        report_rows.append(
            {
                "section": str(row.get("check_name", "")),
                "row_count": int(row.get("row_count", 0) or 0),
                "sheet_name": str(row.get("check_name", ""))[:31],
                "description": "Mapping conflict check.",
            }
        )

    readme = pd.DataFrame(
        [
            {
                "item": "start_here",
                "detail": "Open the summary sheet first, then inspect only sheets with nonzero row_count.",
            },
            {
                "item": "most_actionable_sheets",
                "detail": "coverage_gaps, trio_presence_issues, strict one-to-one rows in crosswalk_target_conflicts.",
            },
            {
                "item": "many_to_many",
                "detail": "These rows are not always wrong, but they need explicit review because one LEAP pair and one target pair both map multiple ways.",
            },
            {
                "item": "non_strict_cardinality_target_review",
                "detail": "These target differences involve one-to-many or many-to-one cardinality, so they are review items rather than strict conflicts.",
            },
            {
                "item": "implied_missing_crosswalk",
                "detail": "Candidate ninth_pairs_to_esto_pairs rows inferred from active LEAP combined mappings. Review rows marked would_create_many_to_many before adding.",
            },
            {
                "item": "trio_presence_note",
                "detail": TRIO_PRESENCE_OUTPUT_NOTE,
            },
        ]
    )
    summary = pd.DataFrame(report_rows)
    output_inventory = pd.DataFrame(
        [
            {
                "output_type": "primary_excel_report",
                "path": str(MAPPING_REFRESH_REPORT_PATH),
                "note": "One workbook with all current mapping refresh checks.",
            },
            {
                "output_type": "coverage_gaps_csv",
                "path": str(COVERAGE_GAPS_PATH),
                "note": "Kept for compatibility with existing workflows.",
            },
            {
                "output_type": "duplicate_mappings_csv",
                "path": str(DUPLICATE_MAPPINGS_CSV_PATH),
                "note": "Kept for compatibility with existing workflows.",
            },
            {
                "output_type": "trio_presence_csv",
                "path": str(TRIO_PRESENCE_CSV_PATH),
                "note": "Kept for compatibility with existing workflows.",
            },
        ]
    )

    sheets: dict[str, pd.DataFrame] = {
        "README": readme,
        "summary": summary,
        "output_inventory": output_inventory,
        "coverage_gaps": gaps,
        "duplicate_mappings": duplicate_mappings,
        "trio_presence_issues": trio_presence_issues,
    }
    for sheet_name, sheet_df in mapping_conflict_report.items():
        if sheet_name == "summary":
            continue
        sheets[sheet_name] = sheet_df
    return sheets


def _write_mapping_refresh_report(
    *,
    gaps: pd.DataFrame,
    duplicate_mappings: pd.DataFrame,
    trio_presence_issues: pd.DataFrame,
    mapping_conflict_report: dict[str, pd.DataFrame],
    output_path: Path = MAPPING_REFRESH_REPORT_PATH,
) -> None:
    """Write the primary one-workbook report for mapping refresh checks."""
    sheets = _build_mapping_refresh_report_sheets(
        gaps=gaps,
        duplicate_mappings=duplicate_mappings,
        trio_presence_issues=trio_presence_issues,
        mapping_conflict_report=mapping_conflict_report,
    )
    _write_excel_report(output_path, sheets)


def _resolve_balance_workbook_for_mapping_check(*, scenario: str, date_id: str | None) -> Path:
    return resolve_balance_export_workbook(
        economy=BALANCE_EXPORT_ECONOMY,
        scenario=scenario,
        date_id=date_id,
    )


def _extract_raw_balance_workbook(workbook_path: Path) -> pd.DataFrame:
    """
    Extract raw nonzero LEAP balance rows without loading or applying mappings.

    The mapping workbook may be incomplete or temporarily invalid while this
    maintenance workflow is being used to find gaps, so this raw extraction must
    not depend on mapping rows being valid.
    """
    chosen_template = _pick_template_sheet(workbook_path, BALANCE_TEMPLATE_SHEET)
    extractor = TemplateBalanceExtractor(
        template_sheet=chosen_template,
        mapping_pairs_path=MAPPING_WORKBOOK_PATH,
        codebook_path=CODEBOOK_PATH,
        reinterpret_fuel_rows_as_parent_sector=False,
        explicit_pair_mappings_only=True,
    )
    workbook = openpyxl.load_workbook(workbook_path, data_only=True, read_only=False)
    if chosen_template not in workbook.sheetnames:
        raise ValueError(f"Template sheet {chosen_template!r} not found in workbook: {workbook_path}")

    template_layout = extractor._extract_layout(workbook[chosen_template])
    selected_sheets = _list_balance_sheets(workbook_path)
    if not selected_sheets:
        raise ValueError(f"No balance sheets found in workbook: {workbook_path}")

    frames: list[pd.DataFrame] = []
    for sheet_name in selected_sheets:
        worksheet = workbook[sheet_name]
        meta = extractor._extract_metadata(worksheet)
        try:
            sheet_layout = extractor._extract_layout(worksheet)
        except ValueError:
            sheet_layout = template_layout
        extracted = extractor._extract_sheet_matrix(worksheet, template=sheet_layout)
        if extracted.empty:
            continue
        extracted.insert(0, "source_sheet", sheet_name)
        extracted.insert(1, "source_workbook", str(workbook_path))
        extracted["area"] = str(meta.get("area", ""))
        extracted["scenario"] = str(meta.get("scenario", ""))
        extracted["year"] = meta.get("year")
        extracted["units"] = str(meta.get("units", ""))
        factor, parse_status, prefix_label, base_label = _parse_unit_factor_to_petajoule(
            str(meta.get("units", ""))
        )
        extracted["value_original"] = extracted["value"]
        extracted["units_original"] = extracted["units"]
        extracted["unit_to_petajoule_factor"] = factor
        extracted["unit_parse_status"] = parse_status
        extracted["unit_prefix"] = prefix_label
        extracted["unit_base"] = base_label
        if factor is not None:
            extracted["value_petajoule"] = pd.to_numeric(extracted["value"], errors="coerce") * float(factor)
        else:
            extracted["value_petajoule"] = pd.NA
        extracted["units_petajoule"] = "Petajoule"
        extracted = extracted[
            pd.to_numeric(extracted.get("value_petajoule", pd.Series(index=extracted.index)), errors="coerce")
            .fillna(0.0)
            .ne(0.0)
        ].copy()
        if not extracted.empty:
            frames.append(extracted)

    if not frames:
        return pd.DataFrame()
    raw_long = pd.concat(frames, ignore_index=True, sort=False)
    dedupe_cols = [
        "source_sheet",
        "leap_sector_name",
        "leap_sector_name_original",
        "leap_sector_name_full_path",
        "leap_fuel_name",
        "value",
    ]
    dedupe_cols_present = [col for col in dedupe_cols if col in raw_long.columns]
    if dedupe_cols_present:
        raw_long = raw_long.drop_duplicates(subset=dedupe_cols_present).reset_index(drop=True)
    return raw_long


def _load_raw_leap_balance_lookup() -> pd.DataFrame:
    """
    Return nonzero raw LEAP balance sector/fuel pairs from REF and TGT exports.

    This intentionally uses raw extractor output rather than mapped rows, so it
    can find LEAP source pairs that are absent from both mapping sheets.
    """
    workbooks = [
        ("Reference", _resolve_balance_workbook_for_mapping_check(scenario="REF", date_id=REF_BALANCE_EXPORT_DATE_ID)),
        ("Target", _resolve_balance_workbook_for_mapping_check(scenario="TGT", date_id=TGT_BALANCE_EXPORT_DATE_ID)),
    ]
    frames: list[pd.DataFrame] = []
    for scenario_label, workbook_path in workbooks:
        raw = _extract_raw_balance_workbook(workbook_path)
        if raw.empty:
            continue
        raw["scenario"] = raw.get("scenario", "").fillna("").astype(str).str.strip()
        raw.loc[raw["scenario"].eq(""), "scenario"] = scenario_label
        if "raw_leap_fuel_name" not in raw.columns:
            raw["raw_leap_fuel_name"] = raw.get("leap_fuel_name", "")
        for col in ["source_sheet", "leap_sector_name_full_path", "raw_leap_fuel_name"]:
            if col not in raw.columns:
                raw[col] = ""
            raw[col] = raw[col].fillna("").astype(str).str.strip()
        if "year" not in raw.columns:
            raw["year"] = pd.NA
        raw["year"] = pd.to_numeric(raw["year"], errors="coerce")
        raw["value_petajoule"] = pd.to_numeric(
            raw.get("value_petajoule", raw.get("value", pd.NA)),
            errors="coerce",
        ).fillna(0.0)
        raw = raw[
            raw["leap_sector_name_full_path"].ne("")
            & raw["raw_leap_fuel_name"].ne("")
            & raw["value_petajoule"].ne(0)
        ].copy()
        if raw.empty:
            continue
        frames.append(
            raw[
                [
                    "source_sheet",
                    "scenario",
                    "year",
                    "leap_sector_name_full_path",
                    "raw_leap_fuel_name",
                    "value_petajoule",
                ]
            ]
        )

    columns = [
        "leap_sector_name_full_path",
        "raw_leap_fuel_name",
        "leap_pair_is_subtotal",
        "leap_pair_abs_sum",
        "raw_leap_row_count",
        "raw_leap_source_sheet_count",
        "raw_leap_scenarios",
        "raw_leap_year_min",
        "raw_leap_year_max",
    ]
    if not frames:
        return pd.DataFrame(columns=columns)

    combined = pd.concat(frames, ignore_index=True, sort=False)
    combined["_value_abs"] = pd.to_numeric(combined["value_petajoule"], errors="coerce").fillna(0.0).abs()
    combined = combined[combined["_value_abs"].gt(0)].copy()
    if combined.empty:
        return pd.DataFrame(columns=columns)
    subtotal_frame = _compute_leap_subtotals(
        combined[
            ["leap_sector_name_full_path", "raw_leap_fuel_name"]
        ].assign(remove_row=False, duplicate_to_remove=False)
    )
    combined["leap_pair_is_subtotal"] = subtotal_frame["leap_is_subtotal"].fillna(False).astype(bool)

    grouped = (
        combined.groupby(["leap_sector_name_full_path", "raw_leap_fuel_name"], as_index=False)
        .agg(
            leap_pair_is_subtotal=("leap_pair_is_subtotal", "max"),
            leap_pair_abs_sum=("_value_abs", "sum"),
            raw_leap_row_count=("_value_abs", "size"),
            raw_leap_source_sheet_count=("source_sheet", lambda values: int(values.astype(str).str.strip().nunique())),
            raw_leap_scenarios=("scenario", lambda values: "|".join(sorted({str(value).strip() for value in values if str(value).strip()}))),
            raw_leap_year_min=("year", "min"),
            raw_leap_year_max=("year", "max"),
        )
        .reset_index(drop=True)
    )
    return grouped[columns]


def _build_coverage_gaps(
    esto_sheet: pd.DataFrame,
    ninth_sheet: pd.DataFrame,
    esto_lookup: pd.DataFrame,
    ninth_lookup: pd.DataFrame,
    raw_leap_lookup: pd.DataFrame | None = None,
) -> pd.DataFrame:
    """
    Return a DataFrame of all coverage gaps: pairs with abs values > 0 that are
    missing from the active mapping rows.

    Columns: gap_type, sheet_name, original_dataset, original_pair_is_subtotal,
    key_col_1, key_col_2, pair_1, pair_2, abs_sum, mapping_presence_state,
    mapping_presence_detail
      gap_type values:
        "esto_missing"   - esto data pair not in any active esto mapping row
        "ninth_missing"  - 9th data pair not in any active ninth mapping row
        "leap_unmapped_esto"  - LEAP pair with value > 0 but no esto target in esto sheet
        "leap_unmapped_ninth" - LEAP pair with value > 0 but no ninth target in ninth sheet
        "raw_leap_missing_esto_mapping" - raw LEAP export pair not present in active esto mapping rows
        "raw_leap_missing_ninth_mapping" - raw LEAP export pair not present in active ninth mapping rows
    """
    records: list[dict] = []

    def _lookup_subtotal_flag(frame: pd.DataFrame, key_a: str, key_b: str, value_a: str, value_b: str, subtotal_col: str) -> bool:
        if subtotal_col not in frame.columns:
            return False
        mask = frame[key_a].astype(str).str.strip().eq(value_a) & frame[key_b].astype(str).str.strip().eq(value_b)
        if not bool(mask.any()):
            return False
        return bool(frame.loc[mask, subtotal_col].fillna(False).astype(bool).any())

    # --- 1. ESTO data pairs missing from the esto mapping ---
    esto_data_pairs = set(
        zip(
            esto_lookup.loc[esto_lookup["esto_pair_abs_sum"] > 0, "esto_flow"].astype(str).str.strip(),
            esto_lookup.loc[esto_lookup["esto_pair_abs_sum"] > 0, "esto_product"].astype(str).str.strip(),
        )
    )
    esto_mapped_pairs = _active_pairs(esto_sheet, "esto_flow", "esto_product")
    for flow, product in sorted(esto_data_pairs - esto_mapped_pairs):
        abs_sum = float(
            esto_lookup.loc[
                esto_lookup["esto_flow"].astype(str).str.strip().eq(flow)
                & esto_lookup["esto_product"].astype(str).str.strip().eq(product),
                "esto_pair_abs_sum",
            ].sum()
        )
        records.append(
            {
                "gap_type": "esto_missing",
                "sheet_name": ESTO_SHEET,
                "original_dataset": "esto",
                "original_pair_is_subtotal": _lookup_subtotal_flag(
                    esto_lookup,
                    "esto_flow",
                    "esto_product",
                    flow,
                    product,
                    "esto_pair_is_subtotal",
                ),
                "key_col_1": "esto_flow",
                "key_col_2": "esto_product",
                "pair_1": flow,
                "pair_2": product,
                "abs_sum": abs_sum,
                "mapping_presence_state": "target_pair_missing_from_active_mappings",
                "mapping_presence_detail": "",
            }
        )

    # --- 2. 9th data pairs missing from the ninth mapping ---
    ninth_data_pairs = set(
        zip(
            ninth_lookup.loc[ninth_lookup["ninth_pair_abs_sum"] > 0, "ninth_sector"].astype(str).str.strip(),
            ninth_lookup.loc[ninth_lookup["ninth_pair_abs_sum"] > 0, "ninth_fuel"].astype(str).str.strip(),
        )
    )
    ninth_mapped_pairs = _active_pairs(ninth_sheet, "ninth_sector", "ninth_fuel")
    for sector, fuel in sorted(ninth_data_pairs - ninth_mapped_pairs):
        abs_sum = float(
            ninth_lookup.loc[
                ninth_lookup["ninth_sector"].astype(str).str.strip().eq(sector)
                & ninth_lookup["ninth_fuel"].astype(str).str.strip().eq(fuel),
                "ninth_pair_abs_sum",
            ].sum()
        )
        records.append(
            {
                "gap_type": "ninth_missing",
                "sheet_name": NINTH_SHEET,
                "original_dataset": "ninth",
                "original_pair_is_subtotal": _lookup_subtotal_flag(
                    ninth_lookup,
                    "ninth_sector",
                    "ninth_fuel",
                    sector,
                    fuel,
                    "ninth_pair_is_subtotal",
                ),
                "key_col_1": "ninth_sector",
                "key_col_2": "ninth_fuel",
                "pair_1": sector,
                "pair_2": fuel,
                "abs_sum": abs_sum,
                "mapping_presence_state": "target_pair_missing_from_active_mappings",
                "mapping_presence_detail": "",
            }
        )

    # --- 3. LEAP pairs with abs(value) > 0 that are unmapped ---
    for sheet, gap_type, sheet_name, target_a, target_b in [
        (esto_sheet, "leap_unmapped_esto", ESTO_SHEET, "esto_flow", "esto_product"),
        (ninth_sheet, "leap_unmapped_ninth", NINTH_SHEET, "ninth_sector", "ninth_fuel"),
    ]:
        active = _compute_leap_subtotals(sheet)[_active_mask(sheet)].copy()
        if "value" not in active.columns:
            continue
        values = pd.to_numeric(active["value"], errors="coerce").fillna(0.0).abs()
        leap_sector = active["leap_sector_name_full_path"].fillna("").astype(str).str.strip() if "leap_sector_name_full_path" in active.columns else pd.Series("", index=active.index)
        leap_fuel = active["raw_leap_fuel_name"].fillna("").astype(str).str.strip() if "raw_leap_fuel_name" in active.columns else pd.Series("", index=active.index)
        leap_is_subtotal = active["leap_is_subtotal"].fillna(False).astype(bool) if "leap_is_subtotal" in active.columns else pd.Series(False, index=active.index)
        ta = active[target_a].fillna("").astype(str).str.strip() if target_a in active.columns else pd.Series("", index=active.index)
        tb = active[target_b].fillna("").astype(str).str.strip() if target_b in active.columns else pd.Series("", index=active.index)
        unmapped_mask = (values > 0) & (ta.eq("") | tb.eq(""))
        for sector, fuel in sorted(set(zip(leap_sector[unmapped_mask], leap_fuel[unmapped_mask]))):
            abs_sum = float(values[unmapped_mask & leap_sector.eq(sector) & leap_fuel.eq(fuel)].sum())
            is_subtotal = bool(leap_is_subtotal[unmapped_mask & leap_sector.eq(sector) & leap_fuel.eq(fuel)].any())
            records.append(
                {
                    "gap_type": gap_type,
                    "sheet_name": sheet_name,
                    "original_dataset": "leap",
                    "original_pair_is_subtotal": is_subtotal,
                    "key_col_1": "leap_sector_name_full_path",
                    "key_col_2": "raw_leap_fuel_name",
                    "pair_1": sector,
                    "pair_2": fuel,
                    "abs_sum": abs_sum,
                    "mapping_presence_state": "active_source_row_missing_target",
                    "mapping_presence_detail": "",
                }
            )

    # --- 4. Raw LEAP export pairs absent from the mapping sheets entirely ---
    if raw_leap_lookup is not None and not raw_leap_lookup.empty:
        raw = raw_leap_lookup.copy()
        for col in ["leap_sector_name_full_path", "raw_leap_fuel_name"]:
            if col not in raw.columns:
                raw[col] = ""
            raw[col] = raw[col].fillna("").astype(str).str.strip()
        if "leap_pair_abs_sum" not in raw.columns:
            raw["leap_pair_abs_sum"] = 0.0
        raw["leap_pair_abs_sum"] = pd.to_numeric(raw["leap_pair_abs_sum"], errors="coerce").fillna(0.0)
        raw = raw[
            raw["leap_sector_name_full_path"].ne("")
            & raw["raw_leap_fuel_name"].ne("")
            & raw["leap_pair_abs_sum"].gt(0)
        ].copy()

        active_source_pairs = {
            "raw_leap_missing_esto_mapping": (
                ESTO_SHEET,
                _active_leap_source_pairs(esto_sheet),
                _leap_source_pair_presence_lookup(esto_sheet),
            ),
            "raw_leap_missing_ninth_mapping": (
                NINTH_SHEET,
                _active_leap_source_pairs(ninth_sheet),
                _leap_source_pair_presence_lookup(ninth_sheet),
            ),
        }
        for gap_type, (sheet_name, mapped_source_pairs, presence_lookup) in active_source_pairs.items():
            for row in raw.itertuples(index=False):
                sector = str(getattr(row, "leap_sector_name_full_path", "")).strip()
                fuel = str(getattr(row, "raw_leap_fuel_name", "")).strip()
                if not sector or not fuel:
                    continue
                source_key = (_path_key(sector), _norm_text(fuel))
                if source_key in mapped_source_pairs:
                    continue
                presence = presence_lookup.get(source_key, {})
                records.append(
                    {
                        "gap_type": gap_type,
                        "sheet_name": sheet_name,
                        "original_dataset": "leap_balance_export",
                        "original_pair_is_subtotal": bool(getattr(row, "leap_pair_is_subtotal", False)),
                        "key_col_1": "leap_sector_name_full_path",
                        "key_col_2": "raw_leap_fuel_name",
                        "pair_1": sector,
                        "pair_2": fuel,
                        "abs_sum": float(getattr(row, "leap_pair_abs_sum", 0.0) or 0.0),
                        "mapping_presence_state": str(presence.get("state", "actually_missing")),
                        "mapping_presence_detail": str(presence.get("detail", "")),
                    }
                )

    return pd.DataFrame(
        records,
        columns=[
            "gap_type",
            "sheet_name",
            "original_dataset",
            "original_pair_is_subtotal",
            "key_col_1",
            "key_col_2",
            "pair_1",
            "pair_2",
            "abs_sum",
            "mapping_presence_state",
            "mapping_presence_detail",
        ],
    )


COVERAGE_GAPS_PATH = MISSING_PAIRS_CSV_PATH


def _report_coverage_gaps(gaps: pd.DataFrame, *, error_on_gaps: bool) -> None:
    """Write gaps CSV and either raise or warn depending on *error_on_gaps*."""
    COVERAGE_GAPS_PATH.parent.mkdir(parents=True, exist_ok=True)
    gaps.to_csv(COVERAGE_GAPS_PATH, index=False)

    if gaps.empty:
        return

    summary_lines: list[str] = []
    for gap_type, group in gaps.groupby("gap_type"):
        summary_lines.append(f"  {gap_type}: {len(group)} pair(s)")
    try:
        report_path = COVERAGE_GAPS_PATH.relative_to(REPO_ROOT)
    except ValueError:
        report_path = COVERAGE_GAPS_PATH
    summary = (
        f"{len(gaps)} coverage gap(s) found in leap_mappings.xlsx "
        f"(written to {report_path}):\n" + "\n".join(summary_lines)
    )

    if error_on_gaps:
        raise ValueError(summary)
    else:
        import warnings
        warnings.warn(summary, stacklevel=3)


def _report_duplicate_mappings(duplicates: pd.DataFrame) -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    duplicates.to_csv(DUPLICATE_MAPPINGS_CSV_PATH, index=False)
    if duplicates.empty:
        return
    summary = (
        f"{len(duplicates)} exact duplicate mapping row(s) found in leap_mappings.xlsx "
        f"(written to {DUPLICATE_MAPPINGS_CSV_PATH.relative_to(REPO_ROOT)})."
    )
    import warnings

    warnings.warn(summary, stacklevel=3)


TRIO_PRESENCE_HISTORY_KEY_COLUMNS = [
    "sheet_name",
    "presence_status",
    "leap_sector_name_original",
    "leap_sector_name_full_path",
    "raw_leap_fuel_name",
    "esto_flow",
    "esto_product",
    "ninth_sector",
    "ninth_fuel",
]


def _trio_presence_history_key(frame: pd.DataFrame) -> pd.Series:
    """Build a stable row key for carrying trio-presence history between runs."""
    work = frame.copy()
    for col in TRIO_PRESENCE_HISTORY_KEY_COLUMNS:
        if col not in work.columns:
            work[col] = ""
        work[col] = work[col].fillna("").astype(str).map(_norm_text)
    return work[TRIO_PRESENCE_HISTORY_KEY_COLUMNS].agg("|||".join, axis=1)


def _add_trio_presence_previous_runs(row_exclusive: pd.DataFrame) -> pd.DataFrame:
    """Add the count of prior refresh outputs where each issue row appeared."""
    out = row_exclusive.copy()
    if "previous_runs" in out.columns:
        out = out.drop(columns=["previous_runs"])
    out.insert(2, "previous_runs", 0)
    if out.empty or not TRIO_PRESENCE_CSV_PATH.exists():
        return out

    previous = pd.read_csv(TRIO_PRESENCE_CSV_PATH).fillna("")
    if previous.empty:
        return out

    previous_key = _trio_presence_history_key(previous)
    # If the previous file had no history column, appearing there still means
    # the row was present in one previous refresh output.
    if "previous_runs" in previous.columns:
        previous_count = pd.to_numeric(previous["previous_runs"], errors="coerce").fillna(0).astype(int) + 1
    else:
        previous_count = pd.Series(1, index=previous.index)

    prior_counts = previous_count.groupby(previous_key).max().to_dict()
    out["previous_runs"] = _trio_presence_history_key(out).map(prior_counts).fillna(0).astype(int)
    return out


def _write_trio_presence_csv(trio_presence: pd.DataFrame) -> pd.DataFrame:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    row_exclusive = trio_presence[~trio_presence["presence_status"].isin({"both_active"})].copy()
    row_exclusive = _add_trio_presence_previous_runs(row_exclusive)
    row_exclusive.to_csv(TRIO_PRESENCE_CSV_PATH, index=False)
    if row_exclusive.empty:
        return row_exclusive
    summary = row_exclusive.groupby(["sheet_name", "presence_status"], as_index=False).size().rename(columns={"size": "row_count"}).sort_values(["sheet_name", "presence_status"])
    summary_parts = [
        f"{row.sheet_name}:{row.presence_status}: {int(row.row_count)}"
        for row in summary.itertuples(index=False)
    ]
    import warnings

    warnings.warn(
        "Row presence mismatches found in leap_combined_esto and leap_combined_ninth "
        f"(written to {TRIO_PRESENCE_CSV_PATH.relative_to(REPO_ROOT)}): "
        + ", ".join(summary_parts),
        stacklevel=3,
    )
    return row_exclusive


ARCHIVE_DIR = MAPPING_WORKBOOK_PATH.parent / "archive"


def _backup_workbook(path: Path) -> Path:
    ARCHIVE_DIR.mkdir(parents=True, exist_ok=True)
    backup_path = ARCHIVE_DIR / f"{path.stem}.before_refresh_mapping_maintenance_columns_{pd.Timestamp.now():%Y%m%d_%H%M%S}{path.suffix}"
    shutil.copy2(path, backup_path)
    return backup_path


def _assert_not_open(path: Path) -> None:
    """Raise a clear error if the file is locked (open in Excel or another process)."""
    try:
        with open(path, "r+b"):
            pass
    except PermissionError:
        raise PermissionError(
            f"{path.name} is open in another application (e.g. Excel). "
            "Close it and re-run the workflow."
        ) from None


def _replace_sheet_with_dataframe(workbook_path: Path, sheet_name: str, frame: pd.DataFrame) -> None:
    """Replace one sheet in-place while preserving every other sheet in the workbook."""
    try:
        workbook = load_workbook(workbook_path)
    except IndexError as exc:
        raise RuntimeError(
            f"{workbook_path.name} can be read in streaming mode but openpyxl cannot "
            "load it in editable mode. Close and re-save the workbook in Excel, or "
            "restore a recent copy from config/archive, then re-run this workflow."
        ) from exc
    safe_sheet_name = safe_excel_sheet_name(sheet_name)
    existing_sheet_name = sheet_name if sheet_name in workbook.sheetnames else safe_sheet_name
    if existing_sheet_name in workbook.sheetnames:
        sheet_index = workbook.sheetnames.index(existing_sheet_name)
        del workbook[existing_sheet_name]
        worksheet = workbook.create_sheet(title=safe_sheet_name, index=sheet_index)
    else:
        worksheet = workbook.create_sheet(title=safe_sheet_name)
    for row in dataframe_to_rows(frame, index=False, header=True):
        worksheet.append(row)
    workbook.save(workbook_path)


def _read_mapping_sheet(sheet_name: str) -> pd.DataFrame:
    """Read a mapping workbook sheet, falling back to openpyxl read-only mode."""
    try:
        return pd.read_excel(MAPPING_WORKBOOK_PATH, sheet_name=sheet_name, dtype=object).fillna("")
    except IndexError:
        try:
            workbook = load_workbook(MAPPING_WORKBOOK_PATH, read_only=True, data_only=False)
            if sheet_name not in workbook.sheetnames:
                raise ValueError(f"Sheet {sheet_name!r} not found in {MAPPING_WORKBOOK_PATH}")
            rows = list(workbook[sheet_name].iter_rows(values_only=True))
        except IndexError:
            rows = _read_xlsx_sheet_values_xml(MAPPING_WORKBOOK_PATH, sheet_name)
        if not rows:
            return pd.DataFrame()
        headers = [str(value or "").strip() for value in rows[0]]
        data = list(rows[1:])
        width = len(headers)
        padded = [tuple(list(row[:width]) + [""] * max(0, width - len(row))) for row in data]
        return pd.DataFrame(padded, columns=headers).fillna("")


def _read_xlsx_sheet_values_xml(workbook_path: Path, sheet_name: str) -> list[tuple[object, ...]]:
    """Read one XLSX sheet directly from XML, ignoring styles that can break openpyxl."""
    ns = {
        "main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
        "rel": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
        "pkgrel": "http://schemas.openxmlformats.org/package/2006/relationships",
    }

    def _col_index(cell_ref: str) -> int:
        letters = "".join(ch for ch in str(cell_ref) if ch.isalpha()).upper()
        value = 0
        for letter in letters:
            value = value * 26 + (ord(letter) - ord("A") + 1)
        return max(value - 1, 0)

    with zipfile.ZipFile(workbook_path) as archive:
        shared_strings: list[str] = []
        if "xl/sharedStrings.xml" in archive.namelist():
            shared_root = ET.fromstring(archive.read("xl/sharedStrings.xml"))
            for si in shared_root.findall("main:si", ns):
                shared_strings.append("".join(t.text or "" for t in si.findall(".//main:t", ns)))

        workbook_root = ET.fromstring(archive.read("xl/workbook.xml"))
        rel_root = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
        rel_targets = {
            rel.attrib.get("Id", ""): rel.attrib.get("Target", "")
            for rel in rel_root.findall("pkgrel:Relationship", ns)
        }
        target = ""
        for sheet in workbook_root.findall(".//main:sheet", ns):
            if sheet.attrib.get("name", "") == sheet_name:
                target = rel_targets.get(sheet.attrib.get(f"{{{ns['rel']}}}id", ""), "")
                break
        if not target:
            raise ValueError(f"Sheet {sheet_name!r} not found in {workbook_path}")
        sheet_path = target.lstrip("/")
        if not sheet_path.startswith("xl/"):
            sheet_path = f"xl/{sheet_path}"

        sheet_root = ET.fromstring(archive.read(sheet_path))
        rows: list[tuple[object, ...]] = []
        for row in sheet_root.findall(".//main:sheetData/main:row", ns):
            values: list[object] = []
            for cell in row.findall("main:c", ns):
                col_idx = _col_index(cell.attrib.get("r", ""))
                while len(values) <= col_idx:
                    values.append("")
                cell_type = cell.attrib.get("t", "")
                raw_value = cell.findtext("main:v", default="", namespaces=ns)
                if cell_type == "s" and str(raw_value).strip().isdigit():
                    shared_idx = int(raw_value)
                    value = shared_strings[shared_idx] if shared_idx < len(shared_strings) else ""
                elif cell_type == "inlineStr":
                    value = "".join(t.text or "" for t in cell.findall(".//main:t", ns))
                else:
                    value = raw_value
                values[col_idx] = value
            rows.append(tuple(values))
        return rows


def _read_canonical_sheet(sheet_name: str) -> pd.DataFrame:
    """Read one sheet from the canonical Outlook mapping workbook."""
    if not MAPPING_WORKBOOK_PATH.exists():
        raise FileNotFoundError(f"Missing canonical mapping workbook: {MAPPING_WORKBOOK_PATH}")
    return pd.read_excel(MAPPING_WORKBOOK_PATH, sheet_name=sheet_name, dtype=object).fillna("")


def _canonical_display_names_as_researcher_rows() -> pd.DataFrame:
    """Convert canonical display-name rows to the legacy researcher export shape."""
    source = _read_canonical_sheet(SECTOR_FUEL_CODE_TO_NAME_SHEET)
    rows: list[dict[str, str]] = []
    known_columns = {"ninth_sector", "ninth_fuel", "esto_flow", "esto_product"}
    for _, row in source.iterrows():
        code_type = str(row.get("code_type", "")).strip().lower()
        column = code_type if code_type in known_columns else None
        code = str(row.get("code", "")).strip()
        name = str(row.get("leap_display_name", "") or row.get("auto_name", "")).strip()
        if column and code:
            rows.append({column: code, "name": name})
    return pd.DataFrame(rows)


def build_researcher_mappings_workbook(
    output_path: Path | str = RESEARCHER_MAPPINGS_PATH,
) -> dict[str, object]:
    """
    Write a narrow researcher-facing mapping workbook.

    The output keeps active rows only and writes four sheets:
    leap_combined_esto, leap_combined_ninth, sector_fuel_code_to_name, and
    ninth_pairs_to_esto_pairs.
    """
    output_path = _resolve(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    if output_path.exists():
        _assert_not_open(output_path)

    esto = _compute_pair_cardinality(_read_mapping_sheet(ESTO_SHEET), "esto_flow", "esto_product")
    ninth = _compute_pair_cardinality(_read_mapping_sheet(NINTH_SHEET), "ninth_sector", "ninth_fuel")
    code_to_name = _canonical_display_names_as_researcher_rows()
    ninth_to_esto = _pair_cardinality_for_columns(
        _read_canonical_sheet(NINTH_PAIRS_TO_ESTO_PAIRS_SHEET),
        source_cols=["ninth_sector", "ninth_fuel"],
        target_cols=["esto_flow", "esto_product"],
    )

    sheets = {
        ESTO_SHEET: _researcher_export_frame(
            esto,
            column_rename={
                "leap_sector_name_full_path": "leap_flow",
                "raw_leap_fuel_name": "leap_product",
            },
        ),
        NINTH_SHEET: _researcher_export_frame(
            ninth,
            column_rename={
                "leap_sector_name_full_path": "leap_flow",
                "raw_leap_fuel_name": "leap_product",
            },
        ),
        SECTOR_FUEL_CODE_TO_NAME_SHEET: _researcher_export_frame(
            code_to_name,
            column_rename={},
            include_name=True,
        ),
        NINTH_PAIRS_TO_ESTO_PAIRS_SHEET: _researcher_export_frame(
            ninth_to_esto,
            column_rename={},
        ),
    }

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        used_sheet_names: set[str] = set()
        for sheet_name, sheet_df in sheets.items():
            safe_sheet_name = safe_excel_sheet_name(sheet_name, used_sheet_names)
            sheet_df.to_excel(writer, sheet_name=safe_sheet_name, index=False)

    return {
        "researcher_mappings_workbook": str(output_path),
        "sheet_rows": {sheet_name: int(len(sheet_df)) for sheet_name, sheet_df in sheets.items()},
    }


def build_mapping_conflicts_workbook(
    output_path: Path | str = MAPPING_CONFLICTS_WORKBOOK_PATH,
) -> dict[str, object]:
    """Build the mapping conflict report from the current workbook state."""
    output_path = _resolve(output_path)
    esto_sheet = _compute_pair_cardinality(_read_mapping_sheet(ESTO_SHEET), "esto_flow", "esto_product")
    ninth_sheet = _compute_pair_cardinality(_read_mapping_sheet(NINTH_SHEET), "ninth_sector", "ninth_fuel")
    ninth_to_esto_pairs = _read_canonical_sheet(NINTH_PAIRS_TO_ESTO_PAIRS_SHEET)
    report_sheets = build_mapping_conflict_report(esto_sheet, ninth_sheet, ninth_to_esto_pairs)
    _write_mapping_conflict_report(report_sheets, output_path)
    return {
        "mapping_conflicts_workbook": str(output_path),
        "sheet_rows": {sheet_name: int(len(sheet_df)) for sheet_name, sheet_df in report_sheets.items()},
    }


def run_workflow(*, error_on_gaps: bool = True, check_raw_leap_coverage: bool = True) -> dict[str, object]:
    """
    Refresh mapping maintenance columns.

    Parameters
    ----------
    error_on_gaps:
        If True (default), raise a ValueError when coverage gaps are found.
        If False, emit a warning and continue; gaps are still written to
        outputs/mappings/mapping_checks/leap_mapping_missing_pairs.csv.
    check_raw_leap_coverage:
        If True (default), read the latest REF/TGT LEAP balance exports and
        report nonzero raw LEAP sector/fuel pairs missing from the mapping
        workbook.
    """
    if not MAPPING_WORKBOOK_PATH.exists():
        raise FileNotFoundError(f"Missing mapping workbook: {MAPPING_WORKBOOK_PATH}")
    _assert_not_open(MAPPING_WORKBOOK_PATH)
    backup_path = _backup_workbook(MAPPING_WORKBOOK_PATH)

    esto_sheet = _read_mapping_sheet(ESTO_SHEET)
    ninth_sheet = _read_mapping_sheet(NINTH_SHEET)

    esto_lookup = _load_esto_lookup()
    ninth_lookup = _load_ninth_lookup()

    refreshed_esto = _refresh_esto_sheet(esto_sheet, esto_lookup)
    refreshed_ninth = _refresh_ninth_sheet(ninth_sheet, ninth_lookup)

    refreshed_esto, esto_auto_remove = _apply_auto_remove_rules(refreshed_esto)
    refreshed_ninth, ninth_auto_remove = _apply_auto_remove_rules(refreshed_ninth)
    refreshed_esto = _compute_pair_cardinality(refreshed_esto, "esto_flow", "esto_product")
    refreshed_ninth = _compute_pair_cardinality(refreshed_ninth, "ninth_sector", "ninth_fuel")

    auto_remove_summary = (
        "Auto-remove rules applied: "
        f"ESTO total fuels={esto_auto_remove['auto_remove_total_fuel_rows']}, "
        f"ESTO suffix matches={esto_auto_remove['auto_remove_sector_fuel_suffix_rows']}, "
        f"9th total fuels={ninth_auto_remove['auto_remove_total_fuel_rows']}, "
        f"9th suffix matches={ninth_auto_remove['auto_remove_sector_fuel_suffix_rows']}, "
        f"newly marked rows={esto_auto_remove['auto_removed_new_rows'] + ninth_auto_remove['auto_removed_new_rows']}."
    )
    import warnings

    warnings.warn(auto_remove_summary, stacklevel=3)

    raw_leap_lookup = _load_raw_leap_balance_lookup() if check_raw_leap_coverage else pd.DataFrame()
    gaps = _build_coverage_gaps(refreshed_esto, refreshed_ninth, esto_lookup, ninth_lookup, raw_leap_lookup)

    refreshed_esto = refreshed_esto.sort_values(
        ["leap_sector_name_full_path", "raw_leap_fuel_name", "esto_flow", "esto_product"],
        key=lambda col: col.fillna("").astype(str).str.lower(),
        na_position="last",
    ).reset_index(drop=True)
    refreshed_ninth = refreshed_ninth.sort_values(
        ["leap_sector_name_full_path", "raw_leap_fuel_name", "ninth_sector", "ninth_fuel"],
        key=lambda col: col.fillna("").astype(str).str.lower(),
        na_position="last",
    ).reset_index(drop=True)

    duplicate_esto = _build_duplicate_mappings(
        refreshed_esto,
        sheet_name=ESTO_SHEET,
        target_a="esto_flow",
        target_b="esto_product",
    )
    duplicate_ninth = _build_duplicate_mappings(
        refreshed_ninth,
        sheet_name=NINTH_SHEET,
        target_a="ninth_sector",
        target_b="ninth_fuel",
    )
    duplicate_mappings = pd.concat([duplicate_esto, duplicate_ninth], ignore_index=True)
    _report_duplicate_mappings(duplicate_mappings)

    trio_presence = _build_trio_presence_check(refreshed_esto, refreshed_ninth)
    trio_presence_issues = _write_trio_presence_csv(trio_presence)
    mapping_conflict_report = build_mapping_conflict_report(
        refreshed_esto,
        refreshed_ninth,
        _read_canonical_sheet(NINTH_PAIRS_TO_ESTO_PAIRS_SHEET),
    )
    _write_mapping_refresh_report(
        gaps=gaps,
        duplicate_mappings=duplicate_mappings,
        trio_presence_issues=trio_presence_issues,
        mapping_conflict_report=mapping_conflict_report,
    )
    _report_coverage_gaps(gaps, error_on_gaps=error_on_gaps)

    print("\n" + "=" * 70)
    print("CONFIRMATION REQUIRED")
    print("=" * 70)
    print(f"About to overwrite sheets {ESTO_SHEET!r} and {NINTH_SHEET!r} in:")
    print(f"  {MAPPING_WORKBOOK_PATH}")
    print(f"A timestamped backup was already created at:\n  {backup_path}")
    answer = input("\nType 'yes' to proceed, anything else to abort: ").strip().lower()
    if answer != "yes":
        print("Aborted -- no changes written.")
        sys.exit(0)

    _replace_sheet_with_dataframe(MAPPING_WORKBOOK_PATH, ESTO_SHEET, refreshed_esto)
    _replace_sheet_with_dataframe(MAPPING_WORKBOOK_PATH, NINTH_SHEET, refreshed_ninth)

    return {
        "mapping_workbook": str(MAPPING_WORKBOOK_PATH),
        "backup_workbook": str(backup_path),
        "coverage_gaps_csv": str(COVERAGE_GAPS_PATH),
        "coverage_gaps_count": int(len(gaps)),
        "raw_leap_source_pairs_checked": int(len(raw_leap_lookup)),
        "raw_leap_missing_mapping_count": int(
            gaps["gap_type"].isin(
                {"raw_leap_missing_esto_mapping", "raw_leap_missing_ninth_mapping"}
            ).sum()
        ) if "gap_type" in gaps.columns else 0,
        "duplicate_mappings_csv": str(DUPLICATE_MAPPINGS_CSV_PATH),
        "duplicate_mappings_count": int(len(duplicate_mappings)),
        "trio_presence_csv": str(TRIO_PRESENCE_CSV_PATH),
        "trio_presence_count": int(len(trio_presence)),
        "mapping_refresh_report": str(MAPPING_REFRESH_REPORT_PATH),
        "mapping_conflicts_count": int(
            sum(
                len(sheet_df)
                for sheet_name, sheet_df in mapping_conflict_report.items()
                if sheet_name != "summary"
            )
        ),
        "auto_remove_total_fuel_rows_esto": int(esto_auto_remove["auto_remove_total_fuel_rows"]),
        "auto_remove_sector_fuel_suffix_rows_esto": int(esto_auto_remove["auto_remove_sector_fuel_suffix_rows"]),
        "auto_removed_new_rows_esto": int(esto_auto_remove["auto_removed_new_rows"]),
        "auto_remove_total_fuel_rows_ninth": int(ninth_auto_remove["auto_remove_total_fuel_rows"]),
        "auto_remove_sector_fuel_suffix_rows_ninth": int(ninth_auto_remove["auto_remove_sector_fuel_suffix_rows"]),
        "auto_removed_new_rows_ninth": int(ninth_auto_remove["auto_removed_new_rows"]),
        "leap_combined_esto_rows": int(len(refreshed_esto)),
        "leap_combined_ninth_rows": int(len(refreshed_ninth)),
    }


#%%
RUN_WORKFLOW = False
# Set to False to emit a warning instead of raising an error when coverage gaps are found.
ERROR_ON_GAPS = False
CREATE_RESEARCHER_MAPPINGS = False
CREATE_MAPPING_CONFLICTS = False

WORKFLOW_RESULT: dict[str, object] | None = None
if __name__ == "__main__" and RUN_WORKFLOW:
    WORKFLOW_RESULT = run_workflow(error_on_gaps=ERROR_ON_GAPS)
    print("[OK] Mapping maintenance columns refreshed.")
    for key, value in WORKFLOW_RESULT.items():
        print(f"- {key}: {value}")
        if key == "trio_presence_csv":
            print(f"- trio_presence_note: {TRIO_PRESENCE_OUTPUT_NOTE}")

RESEARCHER_MAPPINGS_RESULT: dict[str, object] | None = None
if __name__ == "__main__" and CREATE_RESEARCHER_MAPPINGS:
    RESEARCHER_MAPPINGS_RESULT = build_researcher_mappings_workbook()
    print("[OK] Researcher mappings workbook written.")
    for key, value in RESEARCHER_MAPPINGS_RESULT.items():
        print(f"- {key}: {value}")

MAPPING_CONFLICTS_RESULT: dict[str, object] | None = None
if __name__ == "__main__" and CREATE_MAPPING_CONFLICTS:
    MAPPING_CONFLICTS_RESULT = build_mapping_conflicts_workbook()
    print("[OK] Mapping conflicts workbook written.")
    for key, value in MAPPING_CONFLICTS_RESULT.items():
        print(f"- {key}: {value}")
#%%
