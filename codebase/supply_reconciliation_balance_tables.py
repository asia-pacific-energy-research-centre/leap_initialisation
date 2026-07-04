from __future__ import annotations

from datetime import datetime, timezone
from functools import lru_cache
from pathlib import Path
from typing import Iterable
import re
import shutil

import pandas as pd
from openpyxl.styles import Font, PatternFill

from codebase.supply_reconciliation_config import *  # noqa: F401,F403
from codebase.utilities.workflow_utils import _resolve
from codebase.supply_reconciliation_utils import (
    _build_label_to_esto_product_lookup,
    _normalize_template_header_value,
)
from codebase.supply_reconciliation_results import _resolve_refinery_results_workbook
from codebase.configuration.all_products_and_flows import ESTO_PRODUCT_LIST, ESTO_SECTORS

def _find_refinery_sheet_header_row(raw: pd.DataFrame) -> int | None:
    """Find the header row containing Fuel + year columns in refinery output sheet."""
    for idx in range(len(raw.index)):
        values = [_normalize_template_header_value(item) for item in raw.iloc[idx].tolist()]
        lowered = {item.strip().lower() for item in values if str(item).strip()}
        has_fuel = "fuel" in lowered
        has_year = any(str(item).isdigit() for item in values)
        if has_fuel and has_year:
            return int(idx)
    return None


@lru_cache(maxsize=32)
def _load_refinery_fallback_table(economy: str, scenario: str) -> pd.DataFrame:
    """Load refinery output rows from LEAP results workbook into long format."""
    workbook = _resolve_refinery_results_workbook(economy, scenario)
    if workbook is None:
        return pd.DataFrame(columns=["economy", "scenario", "year", "sector", "esto_product", "value"])
    try:
        raw = pd.read_excel(workbook, sheet_name=REFINERY_RESULTS_SHEET_NAME, header=None)
    except Exception:
        return pd.DataFrame(columns=["economy", "scenario", "year", "sector", "esto_product", "value"])

    header_row = _find_refinery_sheet_header_row(raw)
    if header_row is None:
        return pd.DataFrame(columns=["economy", "scenario", "year", "sector", "esto_product", "value"])

    header_values = [_normalize_template_header_value(item) for item in raw.iloc[header_row].tolist()]
    data = raw.iloc[header_row + 1 :].copy()
    data.columns = header_values
    data = data.dropna(how="all").reset_index(drop=True)
    if "Fuel" not in data.columns:
        return pd.DataFrame(columns=["economy", "scenario", "year", "sector", "esto_product", "value"])

    label_to_product = _build_label_to_esto_product_lookup()
    year_columns = [str(col) for col in data.columns if str(col).isdigit()]
    rows: list[dict[str, object]] = []
    for _, record in data.iterrows():
        fuel_label = str(record.get("Fuel") or "").strip()
        if not fuel_label or fuel_label.lower() == "total":
            continue
        esto_product = label_to_product.get(fuel_label) or label_to_product.get(fuel_label.lower())
        if not esto_product:
            alias = REFINERY_FUEL_LABEL_ALIASES.get(fuel_label)
            if alias:
                esto_product = label_to_product.get(alias) or label_to_product.get(alias.lower())
        if not esto_product:
            continue
        for year_col in year_columns:
            value = pd.to_numeric(record.get(year_col), errors="coerce")
            if pd.isna(value):
                continue
            rows.append(
                {
                    "economy": str(economy),
                    "scenario": str(scenario),
                    "year": int(year_col),
                    "sector": REFINERY_SECTOR_NAME,
                    "esto_product": str(esto_product),
                    "value": float(value),
                }
            )
    if not rows:
        return pd.DataFrame(columns=["economy", "scenario", "year", "sector", "esto_product", "value"])
    return pd.DataFrame(rows)


def _get_refinery_fallback_rows_for_balance(
    *,
    economy: str,
    scenario: str,
    year: int,
) -> pd.DataFrame:
    """Return refinery fallback rows for one economy/scenario/year."""
    table = _load_refinery_fallback_table(str(economy), str(scenario))
    if table.empty:
        return pd.DataFrame(columns=["sector", "esto_product", "value"])
    year_value = int(year)
    filtered = table[
        (table["economy"].astype(str) == str(economy))
        & (table["scenario"].astype(str) == str(scenario))
        & (pd.to_numeric(table["year"], errors="coerce").astype("Int64") == year_value)
    ].copy()
    if filtered.empty:
        return pd.DataFrame(columns=["sector", "esto_product", "value"])
    return (
        filtered.groupby(["sector", "esto_product"], dropna=False, as_index=False)["value"]
        .sum(min_count=1)
    )


def _split_sector_codes(raw_value: object) -> list[str]:
    """Split one-or-many mapped sector codes using the same separators as the dashboard."""
    text = str(raw_value or "").strip()
    if not text or text.lower() == "nan":
        return []
    parts = re.split(r"\s*(?:,|;|\||\band\b)\s*", text, flags=re.IGNORECASE)
    out: list[str] = []
    seen: set[str] = set()
    for part in parts:
        token = str(part or "").strip()
        if not token:
            continue
        key = token.lower()
        if key in seen:
            continue
        seen.add(key)
        out.append(token)
    return out


def _sector_code_sequence(value: object) -> tuple[int, ...]:
    """Return the numeric hierarchy sequence from a 9th sector code."""
    token = str(value or "").strip()
    if not token:
        return ()
    parts = [part for part in token.split("_") if part]
    seq: list[int] = []
    for part in parts:
        if not part.isdigit():
            break
        seq.append(int(part))
    return tuple(seq)


def _select_primary_sector_code(raw_value: object) -> str:
    """Pick the deepest mapped sector code for hierarchy comparisons."""
    codes = _split_sector_codes(raw_value)
    if not codes:
        return ""
    ranked = sorted(
        codes,
        key=lambda item: (len(_sector_code_sequence(item)), len(str(item))),
        reverse=True,
    )
    return str(ranked[0])


def _normalize_template_header_value(value: object) -> str:
    """Normalize LEAP import header cells into stable string column names."""
    if value is None or pd.isna(value):
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def build_year_balance_table(
    reconciliation_table: pd.DataFrame,
    year: int,
    economies: Iterable[str] | None = None,
    scenarios: Iterable[str] | None = None,
) -> pd.DataFrame:
    """Return a balance-table style long table for one year."""
    if reconciliation_table.empty:
        return pd.DataFrame(
            columns=[
                "economy",
                "scenario",
                "year",
                "esto_product",
                "balance_component",
                "value",
                "sign_convention",
            ]
        )

    working = reconciliation_table.copy()
    working["year"] = pd.to_numeric(working["year"], errors="coerce").astype("Int64")
    working = working[working["year"] == int(year)].copy()
    if economies:
        economy_list = {str(item) for item in economies}
        working = working[working["economy"].astype(str).isin(economy_list)].copy()
    if scenarios:
        scenario_list = {str(item) for item in scenarios}
        working = working[working["scenario"].astype(str).isin(scenario_list)].copy()
    if working.empty:
        return pd.DataFrame(
            columns=[
                "economy",
                "scenario",
                "year",
                "esto_product",
                "balance_component",
                "value",
                "sign_convention",
            ]
        )

    component_specs = [
        ("demand", "demand_value", "use_negative"),
        ("transformation_input", "transformation_input", "use_negative"),
        ("transformation_output", "constrained_transformation_output", "supply_positive"),
        ("transformation_losses", "transformation_losses", "use_negative"),
        ("production", "constrained_production", "supply_positive"),
        ("stock_changes", "stock_changes", "net_positive"),
        ("projected_imports", "projected_imports", "supply_positive"),
        ("projected_exports", "projected_exports", "use_negative"),
        ("adjusted_imports", "adjusted_imports", "supply_positive"),
        ("adjusted_exports", "adjusted_exports", "use_negative"),
        ("required_net_imports", "required_net_imports", "net_positive"),
        ("adjusted_net_imports", "adjusted_net_imports", "net_positive"),
        ("projected_net_imports", "projected_net_imports", "net_positive"),
        ("trade_adjustment", "trade_adjustment", "net_positive"),
        ("balance_residual", "adjusted_balance", "near_zero"),
    ]

    def _safe_number(value: object) -> float:
        numeric = pd.to_numeric(value, errors="coerce")
        if pd.isna(numeric):
            return 0.0
        return float(numeric)

    rows: list[dict[str, object]] = []
    for _, row in working.iterrows():
        base_record = {
            "economy": row["economy"],
            "scenario": row["scenario"],
            "year": int(row["year"]),
            "esto_product": row["esto_product"],
        }
        for component_name, column_name, sign_convention in component_specs:
            value = pd.to_numeric(row.get(column_name), errors="coerce")
            if pd.isna(value):
                continue
            signed_value = float(value)
            if sign_convention == "use_negative":
                signed_value = -abs(signed_value)
            elif sign_convention == "supply_positive":
                signed_value = abs(signed_value)
            rows.append(
                {
                    **base_record,
                    "balance_component": component_name,
                    "value": signed_value,
                    "raw_value": float(value),
                    "sign_convention": sign_convention,
                }
            )

        tpes_value = (
            _safe_number(row.get("constrained_production"))
            + _safe_number(row.get("stock_changes"))
            + _safe_number(row.get("adjusted_net_imports"))
            + _safe_number(row.get("constrained_transformation_output"))
            - abs(_safe_number(row.get("transformation_input")))
            - abs(_safe_number(row.get("transformation_losses")))
        )
        rows.append(
            {
                **base_record,
                "balance_component": "total_primary_energy_supply",
                "value": tpes_value,
                "raw_value": tpes_value,
                "sign_convention": "net_positive",
            }
        )
        final_balance = tpes_value - abs(_safe_number(row.get("demand_value")))
        rows.append(
            {
                **base_record,
                "balance_component": "final_balance_check",
                "value": final_balance,
                "raw_value": final_balance,
                "sign_convention": "near_zero",
            }
        )

    balance_table = pd.DataFrame(rows)
    if balance_table.empty:
        return balance_table

    totals = (
        balance_table.groupby(
            ["economy", "scenario", "year", "balance_component"],
            dropna=False,
            as_index=False,
        )
        .agg(value=("value", "sum"), raw_value=("raw_value", "sum"))
    )
    totals["esto_product"] = "Total"
    sign_lookup = {
        name: sign
        for name, _, sign in component_specs
    }
    sign_lookup["total_primary_energy_supply"] = "net_positive"
    sign_lookup["final_balance_check"] = "near_zero"
    totals["sign_convention"] = totals["balance_component"].map(sign_lookup).fillna("")

    balance_table = pd.concat([balance_table, totals], ignore_index=True, sort=False)
    balance_table = balance_table.sort_values(
        ["economy", "scenario", "esto_product", "balance_component"]
    ).reset_index(drop=True)
    return _zero_small_numeric_values(
        balance_table,
        label_columns=["economy", "scenario", "year", "esto_product", "balance_component", "sign_convention"],
        threshold=0.01,
    )


def save_year_balance_tables(
    reconciliation_table: pd.DataFrame,
    years: Iterable[int],
    output_dir: Path | str = YEARLY_BALANCE_DIR,
    economies: Iterable[str] | None = None,
    scenarios: Iterable[str] | None = None,
) -> list[Path]:
    """Write scenario/date-scoped CSV balance tables for the simple balance view."""
    output_path = _resolve(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)

    economy_list = [
        str(item).strip()
        for item in (
            economies
            if economies is not None
            else reconciliation_table.get("economy", pd.Series(dtype=str)).dropna().unique()
        )
        if str(item).strip()
    ]
    scenario_list = [
        str(item).strip()
        for item in (
            scenarios
            if scenarios is not None
            else reconciliation_table.get("scenario", pd.Series(dtype=str)).dropna().unique()
        )
        if str(item).strip()
    ]

    output_jobs: list[tuple[int, str, str, str, str, Path]] = []
    for year in years:
        year_int = int(year)
        for economy in economy_list:
            economy_token = _safe_filename_token(economy)
            for scenario in scenario_list:
                date_id, scenario_code = _balance_export_parts_for_scenario(scenario)
                stem = (
                    f"balance_table_{economy_token}_{_safe_filename_token(date_id)}_"
                    f"{_safe_filename_token(scenario_code)}_{year_int}"
                )
                output_jobs.append(
                    (
                        year_int,
                        economy,
                        scenario,
                        economy_token,
                        _safe_filename_token(scenario_code),
                        output_path / f"{stem}.csv",
                    )
                )

    saved_paths: list[Path] = []
    for year, economy, scenario, economy_token, scenario_code, csv_path in output_jobs:
        _archive_prior_year_balance_tables(
            output_path=output_path,
            economy_token=economy_token,
            scenario_code=scenario_code,
            year=int(year),
            current_csv_path=csv_path,
        )
        if csv_path.exists():
            saved_paths.append(csv_path)
            print(f"[INFO] Reusing existing yearly balance table CSV without overwrite: {csv_path}")
            continue
        table = build_year_balance_table(
            reconciliation_table,
            year=year,
            economies=[economy],
            scenarios=[scenario],
        )
        if table.empty:
            continue
        table.to_csv(csv_path, index=False)
        saved_paths.append(csv_path)
        print(f"Saved year balance table CSV to {csv_path}")
    return saved_paths


def _safe_filename_token(value: object) -> str:
    """Return a filesystem-safe token."""
    text = str(value or "").strip()
    if not text:
        return "item"
    safe = "".join(ch if ch.isalnum() or ch in {"_", "-"} else "_" for ch in text)
    return safe.strip("_") or "item"


def _archive_prior_year_balance_tables(
    *,
    output_path: Path,
    economy_token: str,
    scenario_code: str,
    year: int,
    current_csv_path: Path,
) -> list[Path]:
    """Move older matching yearly balance files out of the active output folder."""
    archive_dir = output_path / "archive"
    archived_paths: list[Path] = []
    current_stem = current_csv_path.stem
    patterns = [
        f"balance_table_{economy_token}_*_{scenario_code}_{int(year)}.csv",
        f"balance_table_{economy_token}_*_{scenario_code}_{int(year)}.xlsx",
    ]
    for pattern in patterns:
        for path in sorted(output_path.glob(pattern)):
            if path.name.startswith("~$"):
                continue
            if path.resolve() == current_csv_path.resolve():
                continue
            if path.suffix.lower() == ".xlsx" and path.stem == current_stem:
                pass
            elif path.stem == current_stem:
                continue
            archive_dir.mkdir(parents=True, exist_ok=True)
            target = archive_dir / path.name
            if target.exists():
                stamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
                target = archive_dir / f"{path.stem}_{stamp}{path.suffix}"
            shutil.move(str(path), str(target))
            archived_paths.append(target)
            print(f"[INFO] Archived older yearly balance table to {target}")
    return archived_paths


def _balance_export_filename_parts(path: Path | str) -> tuple[str, str]:
    """Return (date_id, scenario_code) from a LEAP balance-export workbook name."""
    match = re.match(
        r"^full model output all years (?P<date_id>\d{5,8}) (?P<scenario>[A-Za-z]+)(?:\s[^.]*)?\.xlsx$",
        Path(path).name,
        flags=re.IGNORECASE,
    )
    if not match:
        return "unknown_date", "unknown_scenario"
    return match.group("date_id"), match.group("scenario").upper()


def _balance_export_parts_for_scenario(scenario: object) -> tuple[str, str]:
    """Return filename provenance for Reference/Target balance-demand source workbooks."""
    scenario_key = str(scenario or "").strip().lower()
    if scenario_key == "reference":
        date_id, scenario_code = _balance_export_filename_parts(
            BALANCE_DEMAND_REF_WORKBOOK_PATH
        )
    elif scenario_key == "target":
        date_id, scenario_code = _balance_export_filename_parts(
            BALANCE_DEMAND_TGT_WORKBOOK_PATH
        )
    else:
        return "unknown_date", _safe_filename_token(scenario).upper()

    # Compressed preflight workbooks intentionally have generated filenames
    # rather than production LEAP-export names. Preserve unknown date provenance,
    # but never collapse distinct requested scenarios into ``unknown_scenario``.
    if scenario_code == "unknown_scenario":
        scenario_code = _safe_filename_token(scenario).upper()
    return date_id, scenario_code


def _zero_small_numeric_values(
    df: pd.DataFrame,
    *,
    label_columns: Iterable[str],
    threshold: float = 0.01,
) -> pd.DataFrame:
    """Set tiny numeric values to exactly zero for readability."""
    if df.empty:
        return df.copy()
    out = df.copy()
    label_set = {str(col) for col in label_columns}
    for column in out.columns:
        if str(column) in label_set:
            continue
        numeric = pd.to_numeric(out[column], errors="coerce")
        if numeric.notna().any():
            out[column] = numeric.where(numeric.abs() >= float(threshold), 0.0)
    return out


def _filter_balance_scenarios(scenarios: Iterable[str] | None) -> list[str]:
    """Return scenario labels excluding current-accounts style entries."""
    if scenarios is None:
        return []
    filtered: list[str] = []
    for value in scenarios:
        label = str(value or "").strip()
        if not label:
            continue
        if label.lower() in {"current accounts", "current account"}:
            continue
        filtered.append(label)
    return filtered


def _ensure_current_accounts_scenario(scenarios: Iterable[str] | None) -> list[str]:
    """Return scenarios with a canonical Current Accounts label appended if missing."""
    ordered: list[str] = []
    seen: set[str] = set()
    has_current_accounts = False
    for value in (scenarios or []):
        label = str(value or "").strip()
        if not label:
            continue
        key = label.lower()
        if key in seen:
            continue
        seen.add(key)
        ordered.append(label)
        if key in {"current accounts", "current account"}:
            has_current_accounts = True
    if not has_current_accounts:
        ordered.append("Current Accounts")
    return ordered


def _get_projection_value_for_flow_product(
    *,
    economy: str,
    flow: object,
    product: object,
    year: int,
) -> float:
    """Return the projected value for an ESTO flow/product pair in one year."""
    lookup = supply_data_pipeline.SUPPLY_PROJECTION_LOOKUP
    if lookup is None:
        return 0.0
    key = (
        supply_data_pipeline.normalize_economy_key(economy),
        str(flow or "").strip(),
        str(product or "").strip(),
    )
    if key not in lookup.index:
        return 0.0
    row = lookup.loc[key]
    if isinstance(row, pd.DataFrame):
        row = row.sum()
    value = pd.to_numeric(row.get(int(year), 0.0), errors="coerce")
    if pd.isna(value):
        return 0.0
    return float(value)


def _get_base_value_for_flow_product(
    *,
    base_df: pd.DataFrame,
    economy: str,
    flow: object,
    product: object,
    year: int,
) -> float:
    """Return the ESTO base-year value for an ESTO flow/product pair."""
    if base_df.empty:
        return 0.0
    if int(year) not in [col for col in base_df.columns if isinstance(col, int)]:
        return 0.0
    mask = (
        base_df.get("economy", pd.Series(index=base_df.index)).astype(str).eq(str(economy))
        & base_df.get("flows", pd.Series(index=base_df.index)).astype(str).eq(str(flow))
        & base_df.get("products", pd.Series(index=base_df.index)).astype(str).eq(str(product))
    )
    if not mask.any():
        return 0.0
    values = pd.to_numeric(base_df.loc[mask, int(year)], errors="coerce").fillna(0.0)
    return float(values.sum())


def _strip_esto_sector_prefix(label: object) -> str:
    """Remove leading numeric ESTO code prefixes for display rows."""
    text = str(label or "").strip()
    if not text:
        return ""
    parts = text.split(" ", 1)
    if len(parts) == 2 and all(part.isdigit() for part in parts[0].split(".")):
        return parts[1].strip()
    return text


def _build_conventional_row_backbone() -> list[str]:
    """Return the standard conventional balance row order."""
    preferred_rows = [
        "Production",
        "Imports",
        "Exports",
        "International marine bunkers",
        "International aviation bunkers",
        "Stock changes",
        "Total primary energy supply",
        "Transfers",
        "Upstream liquids transfers",
        "Refinery & blending transfers",
        "Transfers unallocated",
        "Recycled products",
        "Interproduct transfers",
        "Products transferred",
        "Gas separation",
        "Transfers nonspecified",
        "Total transformation sector",
        "Main activity producer",
        "Electricity plants",
        "CHP plants",
        "Heat plants",
        "Autoproducers",
        "Electricity plants (autoproducers)",
        "CHP plants (autoproducers)",
        "Heat plants (autoproducers)",
        "Heat pumps",
        "Electric boilers",
        "Chemical heat for electricity production",
        "Gas processing plants",
        "Gas works plants",
        "Liquefaction/regasification plants",
        "Natural gas blending plants",
        "Gas-to-liquids plants",
        "Oil Refining",
        "Coal transformation",
        "Coke ovens",
        "Blast furnaces",
        "Patent fuel plants",
        "BKB/PB plants",
        "Liquefaction (coal to oil)",
        "Petrochemical industry",
        "Biofuels processing",
        "Charcoal processing",
        "Hydrogen transformation",
        "Non-specified transformation",
        "Losses & own use",
        "Own Use",
        "Electricity, CHP and heat plants",
        "Gas works plants (own-use)",
        "Liquefaction/regasification plants (own-use)",
        "Gas-to-liquids plants (own-use)",
        "Coke ovens (own-use)",
        "Coal mines",
        "Blast furnaces (own-use)",
        "Patent fuel plants (own-use)",
        "BKB/PB plants (own-use)",
        "Liquefaction plants (Coal to Oil)",
        "Oil Refining (own-use)",
        "Oil and gas extraction",
        "Pump storage plants",
        "Nuclear industry",
        "Charcoal production plants",
        "Gasification plants for biogases",
        "Non-specified own uses",
        "Transmission and distribution losses",
        "Statistical discrepancy",
        "Total final consumption",
        "Total final energy consumption",
        "Industry sector",
        "Transport sector",
        "Other sector",
        "Non-energy use",
        "Total Final Energy Demand",
        "Unmet Requirements",
    ]

    dynamic_rows = [
        _strip_esto_sector_prefix(item)
        for item in ESTO_SECTORS
        if str(item or "").strip()
    ]

    normalized: list[str] = []
    seen: set[str] = set()
    for row in preferred_rows + dynamic_rows:
        name = str(row or "").strip()
        if not name or name in seen:
            continue
        seen.add(name)
        normalized.append(name)
    return normalized


def _normalize_conventional_sector_name(label: object) -> str:
    """Map internal labels to conventional balance row names."""
    text = _strip_esto_sector_prefix(label)
    replacements = {
        "NG Liquefaction": "Liquefaction/regasification plants",
        "LNG regasification": "Liquefaction/regasification plants",
        "09_13_hydrogen_transformation": "Hydrogen transformation",
        "Total Primary Supply": "Total primary energy supply",
        "Total Transformation": "Total transformation sector",
        "Transmission and Distribution": "Transmission and distribution losses",
        "Upstream & refinery transfers": "Transfers unallocated",
    }
    return replacements.get(text, text)


def _get_conventional_section_layout() -> list[tuple[str, list[str]]]:
    """Return ordered report sections and their canonical row labels."""
    return [
        (
            "Supply",
            [
                "Production",
                "Imports",
                "Exports",
                "International marine bunkers",
                "International aviation bunkers",
                "Stock changes",
                "Total primary energy supply",
            ],
        ),
        (
            "Transfers",
            [
                "Transfers",
                "Upstream liquids transfers",
                "Refinery & blending transfers",
                "Transfers unallocated",
                "Recycled products",
                "Interproduct transfers",
                "Products transferred",
                "Gas separation",
                "Transfers nonspecified",
            ],
        ),
        (
            "Transformation",
            [
                "Main activity producer",
                "Electricity plants",
                "CHP plants",
                "Heat plants",
                "Autoproducers",
                "Electricity plants (autoproducers)",
                "CHP plants (autoproducers)",
                "Heat plants (autoproducers)",
                "Heat pumps",
                "Electric boilers",
                "Chemical heat for electricity production",
                "Gas processing plants",
                "Gas works plants",
                "Liquefaction/regasification plants",
                "Natural gas blending plants",
                "Gas-to-liquids plants",
                "Oil Refining",
                "Coal transformation",
                "Coke ovens",
                "Blast furnaces",
                "Patent fuel plants",
                "BKB/PB plants",
                "Liquefaction (coal to oil)",
                "Petrochemical industry",
                "Biofuels processing",
                "Charcoal processing",
                "Hydrogen transformation",
                "Non-specified transformation",
                "Total transformation sector",
            ],
        ),
        (
            "Losses",
            [
                "Losses & own use",
                "Own Use",
                "Electricity, CHP and heat plants",
                "Gas works plants (own-use)",
                "Liquefaction/regasification plants (own-use)",
                "Gas-to-liquids plants (own-use)",
                "Coke ovens (own-use)",
                "Coal mines",
                "Blast furnaces (own-use)",
                "Patent fuel plants (own-use)",
                "BKB/PB plants (own-use)",
                "Liquefaction plants (Coal to Oil)",
                "Oil Refining (own-use)",
                "Oil and gas extraction",
                "Pump storage plants",
                "Nuclear industry",
                "Charcoal production plants",
                "Gasification plants for biogases",
                "Non-specified own uses",
                "Transmission and distribution losses",
                "Statistical discrepancy",
            ],
        ),
        (
            "Demand",
            [
                "Total final consumption",
                "Total final energy consumption",
                "Industry sector",
                "Transport sector",
                "Other sector",
                "Non-energy use",
                "Industry",
                "Transport",
                "Buildings",
                "Agriculture",
                "Other demand",
                "Total Final Energy Demand",
            ],
        ),
        (
            "Checks",
            [
                "Unmet Requirements",
            ],
        ),
    ]


def _build_conventional_section_lookup() -> tuple[dict[str, str], list[str]]:
    """Return row->section mapping and ordered row list from the report layout."""
    layout = _get_conventional_section_layout()
    row_to_section: dict[str, str] = {}
    ordered_rows: list[str] = []
    for section_name, rows in layout:
        for row in rows:
            row_to_section[row] = section_name
            ordered_rows.append(row)
    return row_to_section, ordered_rows


def _infer_top_level_demand_category(
    *,
    primary_sector_code: object,
    esto_flow: object,
) -> str:
    """Map a demand row to a broad category using mapped 9th/ESTO sector levels."""
    # Late import — these lookups live in supply_demand_mapping, which imports
    # this module at top level, so a top-level import here would be circular.
    from codebase.functions.supply_demand_mapping import SECTOR_TO_ESTO_FLOW_LOOKUP

    flow_label = str(esto_flow or "").strip()
    if not flow_label:
        flow_label = SECTOR_TO_ESTO_FLOW_LOOKUP.get(str(primary_sector_code or "").strip(), "")
    esto_text = flow_label.lower()
    seq = _sector_code_sequence(primary_sector_code)
    if esto_text.startswith("14.") or (seq and seq[0] == 14):
        return "Industry"
    if esto_text.startswith("15.") or (seq and seq[0] == 15):
        return "Transport"
    if esto_text.startswith("16.01") or esto_text.startswith("16.02"):
        return "Buildings"
    if esto_text.startswith("16.03") or esto_text.startswith("16.04"):
        return "Agriculture"
    if esto_text.startswith("17.") or (seq and seq[0] == 17):
        return "Non-energy use"
    if esto_text.startswith("16.05") or (seq and seq[0] == 16):
        return "Other demand"
    return "Other demand"


def _prepare_demand_rows_for_balance(
    demand: pd.DataFrame,
    *,
    drop_parent_rows: bool = DROP_PARENT_DEMAND_ROWS_WHEN_CHILDREN_PRESENT,
    include_top_level_categories: bool = INCLUDE_TOP_LEVEL_DEMAND_CATEGORY_ROWS,
    drop_disaggregated_rows: bool = DROP_DISAGGREGATED_DEMAND_SECTORS,
) -> pd.DataFrame:
    """Normalize demand rows using mapped hierarchy, drop parent rows, and add top-level aggregates."""
    if demand.empty:
        return pd.DataFrame(columns=["sector", "esto_product", "value", "is_top_level_aggregate"])

    working = (
        demand.groupby(
            ["sheet", "sector_code_9th", "esto_flow", "esto_product"],
            dropna=False,
            as_index=False,
        )["demand_value"]
        .sum(min_count=1)
    )
    working["sector"] = working["sheet"].map(_normalize_conventional_sector_name)
    working["primary_sector_code"] = working["sector_code_9th"].map(_select_primary_sector_code)
    working["sector_seq"] = working["primary_sector_code"].map(_sector_code_sequence)
    working["value"] = -working["demand_value"].abs()
    working["is_top_level_aggregate"] = False
    working = working[
        [
            "sector",
            "esto_product",
            "value",
            "primary_sector_code",
            "sector_seq",
            "esto_flow",
            "is_top_level_aggregate",
        ]
    ].copy()

    if drop_parent_rows:
        unique_codes = [
            seq
            for seq in {
                tuple(value) for value in working["sector_seq"].tolist()
                if isinstance(value, tuple) and value
            }
        ]
        parent_codes: set[tuple[int, ...]] = set()
        for seq in unique_codes:
            if any(
                len(other) > len(seq) and other[: len(seq)] == seq
                for other in unique_codes
            ):
                parent_codes.add(seq)
        if parent_codes:
            working = working[
                ~working["sector_seq"].map(
                    lambda seq: isinstance(seq, tuple) and tuple(seq) in parent_codes
                )
            ].copy()

    extra_rows: list[dict[str, object]] = []
    if include_top_level_categories and not working.empty:
        category_totals = (
            working.assign(
                top_level_category=working.apply(
                    lambda row: _infer_top_level_demand_category(
                        primary_sector_code=row.get("primary_sector_code"),
                        esto_flow=row.get("esto_flow"),
                    ),
                    axis=1,
                )
            )
            .groupby(["top_level_category", "esto_product"], dropna=False, as_index=False)["value"]
            .sum(min_count=1)
        )
        for _, row in category_totals.iterrows():
            extra_rows.append(
                {
                    "sector": str(row["top_level_category"]),
                    "esto_product": row["esto_product"],
                    "value": float(row["value"]),
                    "is_top_level_aggregate": True,
                }
            )
    if extra_rows:
        working = pd.concat([working, pd.DataFrame(extra_rows)], ignore_index=True, sort=False)

    if drop_disaggregated_rows:
        working = working[working["is_top_level_aggregate"].fillna(False)].copy()

    return (
        working.groupby(
            ["sector", "esto_product", "is_top_level_aggregate"],
            dropna=False,
            as_index=False,
        )["value"].sum(min_count=1)
    )


def build_reference_demand_rows_for_balance(
    sector_demand_table: pd.DataFrame,
    *,
    economy: str,
    scenario: str,
    year: int,
    base_df: pd.DataFrame,
) -> pd.DataFrame:
    """Build demand rows from ESTO base or projected values using dashboard mappings."""
    if sector_demand_table.empty:
        return pd.DataFrame(columns=["sector", "esto_product", "value", "is_top_level_aggregate"])

    working = sector_demand_table.copy()
    working["year"] = pd.to_numeric(working["year"], errors="coerce").astype("Int64")
    working = working[
        (working["year"] == int(year))
        & (working["economy"].astype(str) == str(economy))
        & (working["scenario"].astype(str) == str(scenario))
    ].copy()
    if working.empty:
        return pd.DataFrame(columns=["sector", "esto_product", "value", "is_top_level_aggregate"])

    def _resolve_source_value(row: pd.Series) -> float:
        if int(year) <= BASE_YEAR:
            return _get_base_value_for_flow_product(
                base_df=base_df,
                economy=str(economy),
                flow=row.get("esto_flow"),
                product=row.get("esto_product"),
                year=int(year),
            )
        return _get_projection_value_for_flow_product(
            economy=str(economy),
            flow=row.get("esto_flow"),
            product=row.get("esto_product"),
            year=int(year),
        )

    working["demand_value"] = working.apply(_resolve_source_value, axis=1)
    return _prepare_demand_rows_for_balance(
        working,
        drop_parent_rows=DROP_PARENT_DEMAND_ROWS_WHEN_CHILDREN_PRESENT,
        include_top_level_categories=INCLUDE_TOP_LEVEL_DEMAND_CATEGORY_ROWS,
        drop_disaggregated_rows=DROP_DISAGGREGATED_DEMAND_SECTORS,
    )


def build_reference_conventional_balance_matrix(
    *,
    reconciliation_table: pd.DataFrame,
    sector_demand_table: pd.DataFrame,
    transformation_sector_table: pd.DataFrame,
    supply_primary_table: pd.DataFrame,
    base_df: pd.DataFrame,
    year: int,
    economy: str,
    scenario: str,
) -> pd.DataFrame:
    """Build the same matrix shape using source ESTO/9th values before adjustments."""
    # Late import — lives in supply_demand_mapping, which imports this module at
    # top level (circular if imported at module scope here).
    from codebase.functions.supply_demand_mapping import ESTO_PARENT_PRODUCT_LOOKUP

    year_value = int(year)
    economy_value = str(economy)
    scenario_value = str(scenario)

    recon = reconciliation_table.copy()
    recon["year"] = pd.to_numeric(recon["year"], errors="coerce").astype("Int64")
    recon = recon[
        (recon["year"] == year_value)
        & (recon["economy"].astype(str) == economy_value)
        & (recon["scenario"].astype(str) == scenario_value)
    ].copy()

    trans = transformation_sector_table.copy()
    trans["year"] = pd.to_numeric(trans["year"], errors="coerce").astype("Int64")
    trans = trans[
        (trans["year"] == year_value)
        & (trans["economy"].astype(str) == economy_value)
    ].copy()

    supply = supply_primary_table.copy()
    supply["year"] = pd.to_numeric(supply["year"], errors="coerce").astype("Int64")
    supply = supply[
        (supply["year"] == year_value)
        & (supply["economy"].astype(str) == economy_value)
    ].copy()

    demand_grouped = build_reference_demand_rows_for_balance(
        sector_demand_table,
        economy=economy_value,
        scenario=scenario_value,
        year=year_value,
        base_df=base_df,
    )

    def _safe_number(value: object) -> float:
        numeric = pd.to_numeric(value, errors="coerce")
        if pd.isna(numeric):
            return 0.0
        return float(numeric)

    row_entries: list[dict[str, object]] = []

    if not supply.empty:
        for _, row in supply.iterrows():
            row_entries.append({"sector": "Production", "esto_product": row["esto_product"], "value": _safe_number(row.get("production"))})
            row_entries.append({"sector": "Stock changes", "esto_product": row["esto_product"], "value": _safe_number(row.get("stock_changes"))})

    if not recon.empty:
        for _, row in recon.iterrows():
            row_entries.append({"sector": "Imports", "esto_product": row["esto_product"], "value": abs(_safe_number(row.get("projected_imports")))})
            row_entries.append({"sector": "Exports", "esto_product": row["esto_product"], "value": -abs(_safe_number(row.get("projected_exports")))})

    if not supply.empty or not recon.empty:
        merged_primary = supply.merge(
            recon[["esto_product", "projected_imports", "projected_exports"]],
            on="esto_product",
            how="outer",
        )
        for _, row in merged_primary.iterrows():
            total_primary_supply = (
                _safe_number(row.get("production"))
                + _safe_number(row.get("projected_imports"))
                - _safe_number(row.get("projected_exports"))
                + _safe_number(row.get("stock_changes"))
            )
            row_entries.append({"sector": "Total primary energy supply", "esto_product": row["esto_product"], "value": total_primary_supply})

    trans_grouped = pd.DataFrame(columns=["sector", "esto_product", "value"])
    if not trans.empty:
        trans_grouped = trans.groupby(["sector", "esto_product"], dropna=False, as_index=False)["value"].sum(min_count=1)
    if trans_grouped.empty or not trans_grouped["sector"].astype(str).eq(REFINERY_SECTOR_NAME).any():
        refinery_fallback = _get_refinery_fallback_rows_for_balance(
            economy=economy_value,
            scenario=scenario_value,
            year=year_value,
        )
        if not refinery_fallback.empty:
            trans_grouped = pd.concat([trans_grouped, refinery_fallback], ignore_index=True, sort=False)
            trans_grouped = trans_grouped.groupby(["sector", "esto_product"], dropna=False, as_index=False)["value"].sum(min_count=1)
    if not trans_grouped.empty:
        for _, row in trans_grouped.iterrows():
            row_entries.append({"sector": row["sector"], "esto_product": row["esto_product"], "value": float(row["value"])})
        trans_totals = trans.groupby(["esto_product"], dropna=False, as_index=False)["value"].sum(min_count=1)
        if not trans_grouped.empty:
            trans_totals = trans_grouped.groupby(["esto_product"], dropna=False, as_index=False)["value"].sum(min_count=1)
        for _, row in trans_totals.iterrows():
            row_entries.append({"sector": "Total transformation sector", "esto_product": row["esto_product"], "value": float(row["value"])})

    if not demand_grouped.empty:
        for _, row in demand_grouped.iterrows():
            row_entries.append({"sector": _normalize_conventional_sector_name(row["sector"]), "esto_product": row["esto_product"], "value": float(row["value"])})
        demand_detail_rows = demand_grouped[~demand_grouped["is_top_level_aggregate"].fillna(False)].copy()
        demand_totals = demand_detail_rows.groupby(["esto_product"], dropna=False, as_index=False)["value"].sum(min_count=1)
        for _, row in demand_totals.iterrows():
            row_entries.append({"sector": "Total Final Energy Demand", "esto_product": row["esto_product"], "value": float(row["value"])})

    # Source datasets do not contain the post-adjustment unmet requirement concept; treat as zero baseline.
    if not row_entries:
        return pd.DataFrame()

    long_df = pd.DataFrame(row_entries)
    long_df["value"] = pd.to_numeric(long_df["value"], errors="coerce").fillna(0.0)
    long_df["fuel_group"] = long_df["esto_product"].map(lambda value: ESTO_PARENT_PRODUCT_LOOKUP.get(str(value), str(value)))
    long_df = long_df.groupby(["sector", "fuel_group"], dropna=False, as_index=False)["value"].sum(min_count=1).rename(columns={"fuel_group": "esto_product"})

    pivot = (
        long_df.pivot_table(index="sector", columns="esto_product", values="value", aggfunc="sum", fill_value=0.0)
        .reset_index()
    )
    pivot["sector"] = pivot["sector"].map(_normalize_conventional_sector_name)
    fuel_columns = [col for col in pivot.columns if col != "sector"]
    if fuel_columns:
        pivot["Total"] = pivot[fuel_columns].sum(axis=1)
    pivot = pivot.groupby("sector", as_index=False).sum(numeric_only=True)
    return _zero_small_numeric_values(pivot.rename(columns={"sector": "Sector"}), label_columns=["Sector"], threshold=0.01)


def build_conventional_balance_diff_matrix(
    shown_table: pd.DataFrame,
    reference_table: pd.DataFrame,
) -> pd.DataFrame:
    """Return shown minus reference with the same row/column layout as `shown_table`."""
    if shown_table.empty:
        return shown_table.copy()
    shown = shown_table.copy()
    if "Sector" not in shown.columns:
        return shown
    if reference_table.empty or "Sector" not in reference_table.columns:
        reference_aligned = pd.DataFrame()
    else:
        reference_aligned = reference_table.copy().set_index("Sector")

    value_columns = [col for col in shown.columns if col != "Sector"]
    diff = shown.copy()
    for column in value_columns:
        shown_values = pd.to_numeric(shown[column], errors="coerce").fillna(0.0)
        if not reference_aligned.empty and column in reference_aligned.columns:
            ref_values = pd.to_numeric(
                reference_aligned.reindex(shown["Sector"].astype(str).tolist())[column],
                errors="coerce",
            ).fillna(0.0)
        else:
            ref_values = pd.Series(0.0, index=shown.index)
        diff[column] = shown_values - ref_values.reset_index(drop=True)
    return _zero_small_numeric_values(diff, label_columns=["Sector"], threshold=0.01)


def build_conventional_balance_matrix(
    reconciliation_table: pd.DataFrame,
    sector_demand_table: pd.DataFrame,
    transformation_sector_table: pd.DataFrame,
    supply_primary_table: pd.DataFrame,
    year: int,
    economy: str,
    scenario: str,
) -> pd.DataFrame:
    """Return a conventional balance matrix: sectors on rows, fuels on columns."""
    # Late import — lives in supply_demand_mapping, which imports this module at
    # top level (circular if imported at module scope here).
    from codebase.functions.supply_demand_mapping import ESTO_PARENT_PRODUCT_LOOKUP

    year_value = int(year)
    economy_value = str(economy)
    scenario_value = str(scenario)

    recon = reconciliation_table.copy()
    recon["year"] = pd.to_numeric(recon["year"], errors="coerce").astype("Int64")
    recon = recon[
        (recon["year"] == year_value)
        & (recon["economy"].astype(str) == economy_value)
        & (recon["scenario"].astype(str) == scenario_value)
    ].copy()

    demand = sector_demand_table.copy()
    demand["year"] = pd.to_numeric(demand["year"], errors="coerce").astype("Int64")
    demand = demand[
        (demand["year"] == year_value)
        & (demand["economy"].astype(str) == economy_value)
        & (demand["scenario"].astype(str) == scenario_value)
    ].copy()

    trans = transformation_sector_table.copy()
    trans["year"] = pd.to_numeric(trans["year"], errors="coerce").astype("Int64")
    trans = trans[
        (trans["year"] == year_value)
        & (trans["economy"].astype(str) == economy_value)
    ].copy()

    supply = supply_primary_table.copy()
    supply["year"] = pd.to_numeric(supply["year"], errors="coerce").astype("Int64")
    supply = supply[
        (supply["year"] == year_value)
        & (supply["economy"].astype(str) == economy_value)
    ].copy()

    def _safe_number(value: object) -> float:
        numeric = pd.to_numeric(value, errors="coerce")
        if pd.isna(numeric):
            return 0.0
        return float(numeric)

    row_entries: list[dict[str, object]] = []

    if not supply.empty:
        production_override = None
        if not recon.empty and "constrained_production" in recon.columns:
            production_override = (
                recon[["esto_product", "constrained_production"]]
                .drop_duplicates(subset=["esto_product"], keep="last")
                .rename(columns={"constrained_production": "production_override"})
            )
            supply = supply.merge(production_override, on="esto_product", how="left")
        for _, row in supply.iterrows():
            production_value = row.get("production_override")
            if pd.isna(pd.to_numeric(production_value, errors="coerce")):
                production_value = row.get("production")
            row_entries.append(
                {
                    "sector": "Production",
                    "esto_product": row["esto_product"],
                    "value": _safe_number(production_value),
                }
            )
            row_entries.append(
                {
                    "sector": "Stock changes",
                    "esto_product": row["esto_product"],
                    "value": _safe_number(row.get("stock_changes")),
                }
            )

    if not recon.empty:
        for _, row in recon.iterrows():
            row_entries.append(
                {
                    "sector": "Imports",
                    "esto_product": row["esto_product"],
                    "value": abs(_safe_number(row.get("adjusted_imports"))),
                }
            )
            row_entries.append(
                {
                    "sector": "Exports",
                    "esto_product": row["esto_product"],
                    "value": -abs(_safe_number(row.get("adjusted_exports"))),
                }
            )

    if not supply.empty or not recon.empty:
        merged_primary = supply.merge(
            recon[["esto_product", "adjusted_imports", "adjusted_exports"]],
            on="esto_product",
            how="outer",
        )
        for _, row in merged_primary.iterrows():
            production = pd.to_numeric(row.get("production"), errors="coerce")
            adjusted_imports = pd.to_numeric(row.get("adjusted_imports"), errors="coerce")
            adjusted_exports = pd.to_numeric(row.get("adjusted_exports"), errors="coerce")
            stock_changes = pd.to_numeric(row.get("stock_changes"), errors="coerce")
            constrained_production = pd.to_numeric(row.get("constrained_production"), errors="coerce")
            if pd.notna(constrained_production):
                production_val = float(constrained_production)
            else:
                production_val = 0.0 if pd.isna(production) else float(production)
            imports_val = 0.0 if pd.isna(adjusted_imports) else float(adjusted_imports)
            exports_val = 0.0 if pd.isna(adjusted_exports) else float(adjusted_exports)
            stock_val = 0.0 if pd.isna(stock_changes) else float(stock_changes)
            total_primary_supply = production_val + imports_val - exports_val + stock_val
            row_entries.append(
                {
                    "sector": "Total primary energy supply",
                    "esto_product": row["esto_product"],
                    "value": total_primary_supply,
                }
            )

    trans_grouped = pd.DataFrame(columns=["sector", "esto_product", "value"])
    if not trans.empty:
        trans_grouped = (
            trans.groupby(["sector", "esto_product"], dropna=False, as_index=False)["value"]
            .sum(min_count=1)
        )
    if trans_grouped.empty or not trans_grouped["sector"].astype(str).eq(REFINERY_SECTOR_NAME).any():
        refinery_fallback = _get_refinery_fallback_rows_for_balance(
            economy=economy_value,
            scenario=scenario_value,
            year=year_value,
        )
        if not refinery_fallback.empty:
            trans_grouped = pd.concat([trans_grouped, refinery_fallback], ignore_index=True, sort=False)
            trans_grouped = (
                trans_grouped.groupby(["sector", "esto_product"], dropna=False, as_index=False)["value"]
                .sum(min_count=1)
            )
    if not trans_grouped.empty:
        for _, row in trans_grouped.iterrows():
            row_entries.append(
                {
                    "sector": row["sector"],
                    "esto_product": row["esto_product"],
                    "value": float(row["value"]),
                }
            )
        trans_totals = (
            trans_grouped.groupby(["esto_product"], dropna=False, as_index=False)["value"]
            .sum(min_count=1)
        )
        for _, row in trans_totals.iterrows():
            row_entries.append(
                {
                    "sector": "Total transformation sector",
                    "esto_product": row["esto_product"],
                    "value": float(row["value"]),
                }
            )

    if not demand.empty:
        demand_grouped = _prepare_demand_rows_for_balance(demand)
        for _, row in demand_grouped.iterrows():
            row_entries.append(
                {
                    "sector": _normalize_conventional_sector_name(row["sector"]),
                    "esto_product": row["esto_product"],
                    "value": float(row["value"]),
                }
            )
        demand_detail_rows = demand_grouped[~demand_grouped["is_top_level_aggregate"].fillna(False)].copy()
        demand_totals = (
            demand_detail_rows.groupby(["esto_product"], dropna=False, as_index=False)["value"]
            .sum(min_count=1)
        )
        for _, row in demand_totals.iterrows():
            row_entries.append(
                {
                    "sector": "Total Final Energy Demand",
                    "esto_product": row["esto_product"],
                    "value": float(row["value"]),
                }
            )

    if not recon.empty:
        residual_totals = (
            recon.groupby(["esto_product"], dropna=False, as_index=False)["adjusted_balance"]
            .sum(min_count=1)
        )
        for _, row in residual_totals.iterrows():
            row_entries.append(
                {
                    "sector": "Unmet Requirements",
                    "esto_product": row["esto_product"],
                    "value": -_safe_number(row["adjusted_balance"]),
                }
            )

    if not row_entries:
        return pd.DataFrame()

    long_df = pd.DataFrame(row_entries)
    long_df["value"] = pd.to_numeric(long_df["value"], errors="coerce").fillna(0.0)
    long_df["fuel_group"] = long_df["esto_product"].map(
        lambda value: ESTO_PARENT_PRODUCT_LOOKUP.get(str(value), str(value))
    )
    long_df = (
        long_df.groupby(["sector", "fuel_group"], dropna=False, as_index=False)["value"]
        .sum(min_count=1)
        .rename(columns={"fuel_group": "esto_product"})
    )

    pivot = (
        long_df.pivot_table(
            index="sector",
            columns="esto_product",
            values="value",
            aggfunc="sum",
            fill_value=0.0,
        )
        .reset_index()
    )
    pivot["sector"] = pivot["sector"].map(_normalize_conventional_sector_name)
    fuel_columns = [col for col in pivot.columns if col != "sector"]
    if fuel_columns:
        pivot["Total"] = pivot[fuel_columns].sum(axis=1)

    pivot = (
        pivot.groupby("sector", as_index=False)
        .sum(numeric_only=True)
    )
    row_to_section, ordered_backbone = _build_conventional_section_lookup()
    numeric_columns = [col for col in pivot.columns if col != "sector"]
    keep_always = {
        "Total primary energy supply",
        "Total transformation sector",
        "Total Final Energy Demand",
        "Unmet Requirements",
    }

    if numeric_columns:
        row_nonzero = (
            pivot[numeric_columns]
            .apply(pd.to_numeric, errors="coerce")
            .fillna(0.0)
            .abs()
            .sum(axis=1)
            > 0
        )
        pivot = pivot[row_nonzero | pivot["sector"].astype(str).isin(keep_always)].copy()

    fuel_columns = [col for col in pivot.columns if col not in {"sector", "Total"}]
    zero_fuel_columns = []
    for column in fuel_columns:
        column_values = pd.to_numeric(pivot[column], errors="coerce").fillna(0.0)
        if float(column_values.abs().sum()) == 0.0:
            zero_fuel_columns.append(column)
    if zero_fuel_columns:
        pivot = pivot.drop(columns=zero_fuel_columns)

    remaining_fuel_columns = [col for col in pivot.columns if col not in {"sector", "Total"}]
    if remaining_fuel_columns:
        pivot["Total"] = (
            pivot[remaining_fuel_columns]
            .apply(pd.to_numeric, errors="coerce")
            .fillna(0.0)
            .sum(axis=1)
        )
    elif "Total" in pivot.columns:
        pivot["Total"] = 0.0

    extra_rows = [
        name for name in pivot["sector"].astype(str).tolist()
        if name not in ordered_backbone
    ]
    demand_extra_rows = sorted(
        [name for name in extra_rows if row_to_section.get(name) is None]
    )
    ordered_rows = ordered_backbone + demand_extra_rows
    pivot = (
        pivot.set_index("sector")
        .reindex(ordered_rows, fill_value=0.0)
        .reset_index()
        .rename(columns={"sector": "Sector"})
    )
    if "Total" in pivot.columns:
        row_nonzero = (
            pivot[[col for col in pivot.columns if col != "Sector"]]
            .apply(pd.to_numeric, errors="coerce")
            .fillna(0.0)
            .abs()
            .sum(axis=1)
            > 0
        )
        pivot = pivot[row_nonzero | pivot["Sector"].astype(str).isin(keep_always)].copy()
    pivot = pivot.reset_index(drop=True)
    return _zero_small_numeric_values(
        pivot,
        label_columns=["Sector"],
        threshold=0.01,
    )


def _style_conventional_balance_worksheet(
    ws,
    table: pd.DataFrame,
    *,
    economy: str,
    scenario: str,
    year: int,
    is_diff: bool,
) -> None:
    """Apply report styling to a conventional balance worksheet."""
    subtotal_fill_by_row = {
        "Total primary energy supply": PatternFill(fill_type="solid", fgColor="FFFFFF"),
        "Total transformation sector": PatternFill(fill_type="solid", fgColor="E2F0D9"),
        "Total Final Energy Demand": PatternFill(fill_type="solid", fgColor="DDEBF7"),
        "Unmet Requirements": PatternFill(fill_type="solid", fgColor="FCE4D6"),
    }
    subtotal_rows = set(subtotal_fill_by_row)
    strong_rows: set[str] = set()
    subtotal_font = Font(bold=True, color="1F1F1F")
    strong_font = Font(bold=True, color="9C0006")
    section_fills = {
        "Supply": PatternFill(fill_type="solid", fgColor="F2F2F2"),
        "Transfers": PatternFill(fill_type="solid", fgColor="FFF2CC"),
        "Transformation": PatternFill(fill_type="solid", fgColor="E2F0D9"),
        "Losses": PatternFill(fill_type="solid", fgColor="FCE4D6"),
        "Demand": PatternFill(fill_type="solid", fgColor="DDEBF7"),
        "Checks": PatternFill(fill_type="solid", fgColor="E4DFEC"),
    }
    row_to_section, _ = _build_conventional_section_lookup()
    ws["A1"] = f"Energy Balance for {economy}"
    subtitle = f"Scenario: {scenario}, Year: {year}, Units: Petajoule"
    if is_diff:
        subtitle += " | Values shown minus source dataset values"
    ws["A2"] = subtitle

    for column_cells in ws.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            if len(value) > max_length:
                max_length = len(value)
        ws.column_dimensions[column_letter].width = min(max_length + 2, 36)

    header_row = 3
    for cell in ws[header_row]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(fill_type="solid", fgColor="EDEDED")

    for row_index in range(header_row + 1, header_row + 1 + len(table)):
        sector_value = ws.cell(row=row_index, column=1).value
        section_name = row_to_section.get(str(sector_value), "Demand")
        fill = section_fills.get(section_name)
        font = None
        if sector_value in subtotal_rows:
            fill = subtotal_fill_by_row.get(str(sector_value), fill)
            font = strong_font if sector_value in strong_rows else subtotal_font
        for column_index in range(1, len(table.columns) + 1):
            cell = ws.cell(row=row_index, column=column_index)
            if fill is not None:
                cell.fill = fill
            if font is not None:
                cell.font = font


def _write_formatted_conventional_balance_workbook(
    table: pd.DataFrame,
    path: Path,
    *,
    economy: str,
    scenario: str,
    year: int,
) -> None:
    """Write a single conventional balance worksheet to Excel."""
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        table.to_excel(writer, sheet_name="Balance", index=False, startrow=2)
        ws = writer.book["Balance"]
        _style_conventional_balance_worksheet(
            ws,
            table,
            economy=economy,
            scenario=scenario,
            year=year,
            is_diff=False,
        )


def save_conventional_balance_tables(
    reconciliation_table: pd.DataFrame,
    sector_demand_table: pd.DataFrame,
    transformation_sector_table: pd.DataFrame,
    supply_primary_table: pd.DataFrame,
    base_df: pd.DataFrame,
    years: Iterable[int],
    output_dir: Path | str = CONVENTIONAL_BALANCE_DIR,
    economies: Iterable[str] | None = None,
    scenarios: Iterable[str] | None = None,
) -> list[Path]:
    """Write one conventional balance workbook per economy/scenario with year and diff sheets."""
    output_path = _resolve(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)
    saved_paths: list[Path] = []

    economy_list = sorted(
        {
            str(value)
            for value in (economies or reconciliation_table.get("economy", pd.Series(dtype=str)).astype(str).unique())
            if str(value).strip()
        }
    )
    scenario_list = sorted(
        {
            str(value)
            for value in (scenarios or reconciliation_table.get("scenario", pd.Series(dtype=str)).astype(str).unique())
            if str(value).strip()
        }
    )

    for economy in economy_list:
        for scenario in scenario_list:
            workbook_path = output_path / (
                f"conventional_balance_{_safe_filename_token(economy)}_"
                f"{_safe_filename_token(scenario)}.xlsx"
            )
            sheets_written = 0
            with pd.ExcelWriter(workbook_path, engine="openpyxl") as writer:
                for year in years:
                    year_int = int(year)
                    table = build_conventional_balance_matrix(
                        reconciliation_table=reconciliation_table,
                        sector_demand_table=sector_demand_table,
                        transformation_sector_table=transformation_sector_table,
                        supply_primary_table=supply_primary_table,
                        year=year_int,
                        economy=economy,
                        scenario=scenario,
                    )
                    if table.empty:
                        continue
                    reference_table = build_reference_conventional_balance_matrix(
                        reconciliation_table=reconciliation_table,
                        sector_demand_table=sector_demand_table,
                        transformation_sector_table=transformation_sector_table,
                        supply_primary_table=supply_primary_table,
                        base_df=base_df,
                        year=year_int,
                        economy=economy,
                        scenario=scenario,
                    )
                    diff_table = build_conventional_balance_diff_matrix(
                        shown_table=table,
                        reference_table=reference_table,
                    )

                    balance_sheet_name = str(year_int)
                    diff_sheet_name = f"{year_int}-diffs"
                    table.to_excel(writer, sheet_name=balance_sheet_name, index=False, startrow=2)
                    diff_table.to_excel(writer, sheet_name=diff_sheet_name, index=False, startrow=2)

                    _style_conventional_balance_worksheet(
                        writer.book[balance_sheet_name],
                        table,
                        economy=economy,
                        scenario=scenario,
                        year=year_int,
                        is_diff=False,
                    )
                    _style_conventional_balance_worksheet(
                        writer.book[diff_sheet_name],
                        diff_table,
                        economy=economy,
                        scenario=scenario,
                        year=year_int,
                        is_diff=True,
                    )
                    sheets_written += 2
            if sheets_written:
                saved_paths.append(workbook_path)
                print(f"Saved conventional balance workbook to {workbook_path}")
            elif workbook_path.exists():
                workbook_path.unlink()
    return saved_paths


