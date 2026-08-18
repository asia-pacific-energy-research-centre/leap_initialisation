"""Summarise ESTO values behind missing baseline-seed branch findings.

The report keeps ESTO's signed PJ values, separates the 2022 base year from
years after 2022, and compares the 2024, 2025, and 2026 ESTO vintages.
"""

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


REPO_ROOT = Path(__file__).resolve().parents[2]
BASE_YEAR = 2022
VINTAGES = {
    "2024": (REPO_ROOT / "data" / "00APEC_2024_low_with_subtotals.csv", 2022),
    "2025": (REPO_ROOT / "data" / "00APEC_2025_low_with_subtotals.csv", 2023),
    "2026": (REPO_ROOT / "data" / "00APEC_2026_low_with_subtotals_PRELIMINARY.csv", 2024),
}
NINTH_DATA_PATH = REPO_ROOT / "data" / "merged_file_energy_ALL_20251106.csv"
BATCH_FINDINGS = [
    REPO_ROOT
    / "outputs/leap_exports/supply_reconciliation/baseline_seed/runs/"
    / "SEED_5521F03_BATCH1_AUS_USA_PRC_20260818_100306/supporting_files/"
    / "baseline_seed_validation/baseline_seed_20260818_consolidated_rule_findings.csv",
    REPO_ROOT.parent
    / "worktrees/leap_initialisation_seed_5521f03/outputs/leap_exports/"
    / "supply_reconciliation/baseline_seed/runs/"
    / "SEED_5521F03_BATCH2_BD_MAS_MEX_NZ_20260818_115901/supporting_files/"
    / "baseline_seed_validation/baseline_seed_20260818_consolidated_rule_findings.csv",
    REPO_ROOT.parent
    / "worktrees/leap_initialisation_seed_5521f03/outputs/leap_exports/"
    / "supply_reconciliation/baseline_seed/runs/"
    / "SEED_5521F03_BATCH3_PNG_PHL_RUS_THA_VN_20260818_145954/supporting_files/"
    / "baseline_seed_validation/baseline_seed_20260818_consolidated_rule_findings.csv",
]
SEED_RUNS = [
    REPO_ROOT
    / "outputs/leap_exports/supply_reconciliation/baseline_seed/runs/"
    / "SEED_5521F03_BATCH1_AUS_USA_PRC_20260818_100306",
    REPO_ROOT.parent
    / "worktrees/leap_initialisation_seed_5521f03/outputs/leap_exports/"
    / "supply_reconciliation/baseline_seed/runs/"
    / "SEED_5521F03_BATCH2_BD_MAS_MEX_NZ_20260818_115901",
    REPO_ROOT.parent
    / "worktrees/leap_initialisation_seed_5521f03/outputs/leap_exports/"
    / "supply_reconciliation/baseline_seed/runs/"
    / "SEED_5521F03_BATCH3_PNG_PHL_RUS_THA_VN_20260818_145954",
]
FLOW_BY_SECTOR = {
    "Coal mines": "10.01.06 Coal mines",
    "Electricity CHP and heat plants": "10.01.01 Electricity, CHP and heat plants",
    "Oil and gas extraction": "10.01.12 Oil and gas extraction",
    "Non specified own uses": "10.01.17 Non-specified own uses",
    "Transmission and distribution loss": "10.02 Transmission and distribution losses",
    "Non specified transformation": "09.12 Non-specified transformation",
    "LNG regasification": "09.06.02 Liquefaction/regasification plants",
    "NG liquefaction": "09.06.02 Liquefaction/regasification plants",
    "NG Liquefaction": "09.06.02 Liquefaction/regasification plants",
    "Heat plant interim": "09.01.03 Heat plants",
    "Coke ovens": "09.08.01 Coke ovens",
}
NINTH_SECTOR_LEVEL = {
    "Coal mines": ("sub2sectors", "10_01_06_coal_mines"),
    "Electricity CHP and heat plants": ("sub2sectors", "10_01_01_electricity_chp_and_heat_plants"),
    "Oil and gas extraction": ("sub2sectors", "10_01_12_oil_and_gas_extraction"),
    "Non specified own uses": ("sub2sectors", "10_01_17_nonspecified_own_uses"),
    "Transmission and distribution loss": ("sub1sectors", "10_02_transmission_and_distribution_losses"),
    "Non specified transformation": ("sub1sectors", "09_12_nonspecified_transformation"),
    "LNG regasification": ("sub2sectors", "09_06_02_liquefaction_regasification_plants"),
    "NG Liquefaction": ("sub2sectors", "09_06_02_liquefaction_regasification_plants"),
    "Heat plant interim": ("sub1sectors", "09_x_heat_plants"),
    "Coke ovens": ("sub2sectors", "09_08_01_coke_ovens"),
}
NINTH_FUEL = {
    "Anthracite": ("01_coal", "01_04_anthracite"),
    "BKB and PB": ("02_coal_products", "02_08_bkb_pb"),
    "Lignite": ("01_coal", "01_05_lignite"),
    "Bitumen": ("07_petroleum_products", "07_14_bitumen"),
    "Blast furnace gas": ("02_coal_products", "02_04_blast_furnace_gas"),
    "Lubricants": ("07_petroleum_products", "07_13_lubricants"),
    "Other recovered gases": ("02_coal_products", "02_05_other_recovered_gases"),
    "Petroleum coke": ("07_petroleum_products", "07_16_petroleum_coke"),
    "Refinery gas not liquefied": ("07_petroleum_products", "07_10_refinery_gas_not_liquefied"),
    "Patent fuel": ("02_coal_products", "02_06_patent_fuel"),
    "Kerosene type jet fuel": ("07_petroleum_products", "07_05_kerosene_type_jet_fuel"),
    "Coal tar": ("02_coal_products", "02_07_coal_tar"),
    "Coking coal": ("01_coal", "01_01_coking_coal"),
    "Gas works gas": ("08_gas", "08_03_gas_works_gas"),
    "Paraffin waxes": ("07_petroleum_products", "07_15_paraffin_waxes"),
    "Natural gas liquids": ("06_crude_oil_and_ngl", "06_02_natural_gas_liquids"),
    "Coke oven coke": ("02_coal_products", "02_01_coke_oven_coke"),
    "White spirit sbp": ("07_petroleum_products", "07_12_white_spirit_sbp"),
    "Other bituminous coal": ("01_coal", "01_02_other_bituminous_coal"),
    "Peat": ("03_peat", "x"),
    "Additives and oxygenates": ("06_crude_oil_and_ngl", "06_04_additives_oxygenates"),
    "Refinery feedstocks": ("06_crude_oil_and_ngl", "06_03_refinery_feedstocks"),
    "Gas": ("08_gas", "x"),
    "Crude oil": ("06_crude_oil_and_ngl", "06_01_crude_oil"),
}
PRODUCT_BY_FUEL = {
    "Anthracite": "01.04 Anthracite",
    "BKB and PB": "02.08 BKB/PB",
    "Lignite": "01.05 Lignite",
    "Bitumen": "07.14 Bitumen",
    "Blast furnace gas": "02.04 Blast furnace gas",
    "Lubricants": "07.13 Lubricants",
    "Other recovered gases": "02.05 Other recovered gases",
    "Petroleum coke": "07.16 Petroleum coke",
    "Refinery gas not liquefied": "07.10 Refinery gas (not liquefied)",
    "Patent fuel": "02.06 Patent fuel",
    "Kerosene type jet fuel": "07.05 Kerosene type jet fuel",
    "Coal tar": "02.07 Coal tar",
    "Coking coal": "01.01 Coking coal",
    "Gas works gas": "08.03 Gas works gas",
    "Paraffin waxes": "07.15 Paraffin waxes",
    "Natural gas liquids": "06.02 Natural gas liquids",
    "Crude oil": "06.01 Crude oil",
    "Coke oven coke": "02.01 Coke oven coke",
    "White spirit sbp": "07.12 White spirit SBP",
    "Other bituminous coal": "01.02 Other bituminous coal",
    "Peat": "03 Peat",
    "Additives and oxygenates": "06.04 Additives/ oxygenates",
    "Refinery feedstocks": "06.03 Refinery feedstocks",
    "Gas": "08 Gas",
}
CREATION_INSTRUCTION_COLUMNS = [
    "economy", "branch_path", "parent_path", "branch_label", "branch_kind",
    "source_flow", "esto_product", "actionable_source_nonzero", "seed_impact",
    "seed_matching_rows", "seed_activity_rows", "esto_2024_base_2022",
    "esto_2025_base_2023", "esto_2026_base_2024", "ninth_2024_projected_sum",
    "ninth_2025_projected_sum", "ninth_2026_projected_sum", "ninth_match_mode",
    "create_instruction",
]


def _normalise_economy(value: object) -> str:
    text = str(value).strip()
    return f"{text[:2]}_{text[2:]}" if len(text) > 2 and text[2] != "_" else text


def load_missing_branch_rows() -> pd.DataFrame:
    frames = [pd.read_csv(path) for path in BATCH_FINDINGS if path.exists()]
    if len(frames) != len(BATCH_FINDINGS):
        missing = [str(path) for path in BATCH_FINDINGS if not path.exists()]
        raise FileNotFoundError(f"Missing findings files: {missing}")
    return load_missing_branch_rows_from_findings(pd.concat(frames, ignore_index=True))


def load_missing_branch_rows_from_findings(findings: pd.DataFrame) -> pd.DataFrame:
    """Extract distinct missing-branch keys from one baseline-run findings frame."""
    required = {"rule_id", "status", "economy", "Branch Path"}
    missing_columns = required.difference(findings.columns)
    if missing_columns:
        raise ValueError(f"Findings frame is missing required columns: {sorted(missing_columns)}")
    findings = findings.copy()
    if findings.empty:
        return pd.DataFrame(columns=["economy", "sector", "fuel", "branch_path", "source_flow", "esto_product"])
    findings = findings[(findings["rule_id"] == "SEED-011") & (findings["status"] == "fail")]
    rows = []
    for _, row in findings[["economy", "Branch Path"]].drop_duplicates().iterrows():
        parts = str(row["Branch Path"]).split("\\")
        if parts[0] == "Demand":
            sector, fuel = parts[2], parts[-1]
        else:
            sector, fuel = parts[1], parts[-1]
        rows.append(
            {
                "economy": str(row["economy"]),
                "sector": sector,
                "fuel": fuel,
                "branch_path": row["Branch Path"],
                "source_flow": FLOW_BY_SECTOR.get(sector, ""),
                "esto_product": PRODUCT_BY_FUEL.get(fuel, ""),
            }
        )
    output = pd.DataFrame(rows).sort_values(["economy", "sector", "fuel"])
    unknown = output[(output["source_flow"] == "") | (output["esto_product"] == "")]
    if not unknown.empty:
        raise ValueError(f"Unmapped missing-branch rows: {unknown.to_dict('records')}")
    return output.reset_index(drop=True)


def _seed_presence_from_frames(keys: pd.DataFrame, seed_rows_by_economy: dict[str, pd.DataFrame]) -> pd.DataFrame:
    """Summarise actual in-memory seed rows without reopening Excel files."""
    def nonzero(value: object) -> bool:
        try:
            return abs(float(value)) > 1e-12
        except (TypeError, ValueError):
            return False

    rows = []
    for economy, branch_path in keys[["economy", "branch_path"]].drop_duplicates().itertuples(index=False):
        frame = seed_rows_by_economy.get(str(economy), pd.DataFrame())
        matching = frame[frame.get("Branch Path", pd.Series(dtype=object)).astype(str).eq(str(branch_path))] if not frame.empty else frame
        activity = matching[matching.get("Variable", pd.Series(dtype=object)).astype(str).str.strip().eq("Activity Level")] if not matching.empty else matching
        year_columns = {int(str(column)): column for column in matching.columns if str(column).isdigit()}
        rows.append({
            "economy": economy,
            "branch_path": branch_path,
            "seed_matching_rows": len(matching),
            "seed_activity_rows": len(activity),
            "seed_base_nonzero_any": any(nonzero(row.get(year_columns.get(2022))) for _, row in activity.iterrows()) if year_columns else False,
            "seed_projected_nonzero_any": any(
                nonzero(row.get(column)) for _, row in activity.iterrows()
                for year, column in year_columns.items() if year > 2022
            ) if year_columns else False,
        })
    return pd.DataFrame(rows)


def build_creation_instructions_for_run(
    findings: pd.DataFrame,
    *,
    seed_rows_by_economy: dict[str, pd.DataFrame],
    output_path: Path,
    esto_vintages: dict[str, tuple[Path, int]] | None = None,
    ninth_data_path: Path | None = None,
) -> pd.DataFrame:
    """Write the standard, source-energy-filtered branch creation CSV for a run."""
    keys = load_missing_branch_rows_from_findings(findings)
    if keys.empty:
        empty = pd.DataFrame(columns=CREATION_INSTRUCTION_COLUMNS)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        empty.to_csv(output_path, index=False)
        return empty
    report = keys.copy()
    for vintage, (path, base_year) in (esto_vintages or VINTAGES).items():
        report = report.merge(load_esto_values(path, keys, vintage, base_year), on=["economy", "branch_path"], how="left")
    report = report.merge(
        load_ninth_projection_values(keys, ninth_data_path),
        on=["economy", "branch_path"], how="left",
    )
    report = report.merge(_seed_presence_from_frames(keys, seed_rows_by_economy), on=["economy", "branch_path"], how="left")
    base_columns = [column for column in report if column.startswith("esto_") and "_base_" in column]
    projection_columns = [column for column in report if column.startswith("ninth_") and column.endswith("projected_sum")]
    report["actionable_source_nonzero"] = (
        report[base_columns].abs().gt(1e-12).any(axis=1)
        | (report["ninth_match_mode"].eq("exact_subfuel") & report[projection_columns].abs().gt(1e-12).any(axis=1))
    )
    report["seed_impact"] = report.apply(
        lambda row: "No matching seed rows; nonzero source energy omitted" if not row["seed_matching_rows"] and row["actionable_source_nonzero"]
        else "Rows retained in seed with nonzero source energy but unresolved IDs" if row["seed_matching_rows"] and row["actionable_source_nonzero"]
        else "Rows retained but source energy is zero; suppress metadata/activity-only rows" if row["seed_matching_rows"]
        else "No matching seed rows; source energy is zero", axis=1,
    )
    return build_creation_instructions(report, output_path)


def load_esto_values(path: Path, keys: pd.DataFrame, vintage: str, base_year: int) -> pd.DataFrame:
    years = pd.read_csv(path, nrows=0).columns.tolist()
    year_columns = [column for column in years if column.isdigit()]
    data = pd.read_csv(
        path,
        usecols=["economy", "flows", "products", "is_subtotal", *year_columns],
    )
    data["economy"] = data["economy"].map(_normalise_economy)
    data = data[data["is_subtotal"].fillna(False).eq(False)]
    joined = keys.merge(
        data,
        left_on=["economy", "source_flow", "esto_product"],
        right_on=["economy", "flows", "products"],
        how="left",
    )
    base_column = f"esto_{vintage}_base_{base_year}"
    available_projected = [year for year in year_columns if int(year) > base_year]
    projected_column = f"ninth_{vintage}_projected_sum"
    joined[base_column] = pd.to_numeric(joined.get(str(base_year)), errors="coerce").fillna(0.0)
    return joined[["economy", "branch_path", base_column]]


def load_ninth_projection_values(keys: pd.DataFrame, data_path: Path | None = None) -> pd.DataFrame:
    year_columns = [str(year) for year in range(2023, 2061)]
    usecols = [
        "scenarios", "economy", "sectors", "sub1sectors", "sub2sectors",
        "fuels", "subfuels", "subtotal_layout", "subtotal_results", *year_columns,
    ]
    economies = set(keys["economy"])
    chunks = []
    source_path = NINTH_DATA_PATH if data_path is None else Path(data_path)
    for chunk in pd.read_csv(source_path, usecols=usecols, chunksize=200_000):
        chunk = chunk[chunk["economy"].isin(economies)]
        if not chunk.empty:
            chunks.append(chunk)
    ninth = pd.concat(chunks, ignore_index=True)
    ninth = ninth[
        ninth["scenarios"].eq("reference")
        & ninth["subtotal_layout"].eq(False)
        & ninth["subtotal_results"].eq(False)
    ].copy()
    rows = []
    for _, key in keys.iterrows():
        level, sector_code = NINTH_SECTOR_LEVEL[key["sector"]]
        fuel_code, subfuel_code = NINTH_FUEL[key["fuel"]]
        subset = ninth[
            (ninth["economy"] == key["economy"])
            & ninth[level].eq(sector_code)
            & ninth["fuels"].eq(fuel_code)
        ]
        match_mode = "exact_subfuel"
        exact = subset[subset["subfuels"].eq(subfuel_code)]
        if exact.empty and subfuel_code != "x":
            exact = subset[subset["subfuels"].eq("x")]
            match_mode = "fuel_aggregate_fallback"
        if exact.empty:
            match_mode = "no_9th_match"
        values = exact[year_columns].apply(pd.to_numeric, errors="coerce").fillna(0.0).sum(axis=0)
        rows.append(
            {
                "economy": key["economy"],
                "branch_path": key["branch_path"],
                "ninth_2024_projected_sum": float(values.loc[[year for year in year_columns if int(year) > 2022]].sum()),
                "ninth_2025_projected_sum": float(values.loc[[year for year in year_columns if int(year) > 2023]].sum()),
                "ninth_2026_projected_sum": float(values.loc[[year for year in year_columns if int(year) > 2024]].sum()),
                "ninth_match_mode": match_mode,
                "ninth_sector_code": sector_code,
                "ninth_fuel_code": fuel_code,
                "ninth_subfuel_code": subfuel_code,
            }
        )
    return pd.DataFrame(rows)


def load_seed_presence(keys: pd.DataFrame) -> pd.DataFrame:
    def nonzero(value: object) -> bool:
        try:
            return abs(float(value)) > 1e-12
        except (TypeError, ValueError):
            return False

    rows = []
    for run in SEED_RUNS:
        for workbook_path in sorted(run.glob("leap_import_baseline_seed_*.xlsx")):
            economy = workbook_path.stem.split("_2026")[0].replace("leap_import_baseline_seed_", "")
            workbook = load_workbook(workbook_path, read_only=True, data_only=True)
            sheet = workbook["FOR_VIEWING"]
            header = [cell.value for cell in next(sheet.iter_rows(min_row=3, max_row=3))]
            index = {str(value): position for position, value in enumerate(header) if value is not None}
            branch_index = index.get("Branch Path")
            variable_index = index.get("Variable")
            year_indices = {
                year: index.get(str(year))
                for year in range(BASE_YEAR, 2061)
                if index.get(str(year)) is not None
            }
            matching = {}
            for values in sheet.iter_rows(min_row=4, values_only=True):
                branch_path = values[branch_index] if branch_index is not None else None
                if branch_path is None:
                    continue
                matching.setdefault(str(branch_path), []).append(values)
            for branch_path in keys.loc[keys["economy"] == economy, "branch_path"].unique():
                values = matching.get(branch_path, [])
                activity_rows = [
                    row for row in values
                    if variable_index is not None and str(row[variable_index]).strip() == "Activity Level"
                ]
                base_nonzero = any(nonzero(row[year_indices[BASE_YEAR]]) for row in activity_rows)
                projected_nonzero = any(
                    nonzero(row[year_index])
                    for row in activity_rows
                    for year, year_index in year_indices.items()
                    if year > BASE_YEAR
                )
                rows.append(
                    {
                        "economy": economy,
                        "branch_path": branch_path,
                        "seed_matching_rows": len(values),
                        "seed_activity_rows": len(activity_rows),
                        "seed_base_nonzero_any": base_nonzero,
                        "seed_projected_nonzero_any": projected_nonzero,
                    }
                )
            workbook.close()
    return pd.DataFrame(rows)


def build_report(output_path: Path) -> pd.DataFrame:
    keys = load_missing_branch_rows()
    report = keys.copy()
    for vintage, (path, base_year) in VINTAGES.items():
        values = load_esto_values(path, keys, vintage, base_year)
        report = report.merge(values, on=["economy", "branch_path"], how="left")
    report = report.merge(load_ninth_projection_values(keys), on=["economy", "branch_path"], how="left")
    report = report.merge(load_seed_presence(keys), on=["economy", "branch_path"], how="left")
    value_columns = [
        column
        for column in report
        if (column.startswith("esto_") and "_base_" in column)
        or (column.startswith("ninth_") and column.endswith("projected_sum"))
    ]
    report["nonzero_in_any_vintage"] = report[value_columns].abs().gt(1e-12).any(axis=1)
    base_columns = [column for column in report if column.startswith("esto_") and "_base_" in column]
    exact_projection = report["ninth_match_mode"].eq("exact_subfuel")
    projection_columns = [column for column in report if column.startswith("ninth_") and column.endswith("projected_sum")]
    report["actionable_source_nonzero"] = (
        report[base_columns].abs().gt(1e-12).any(axis=1)
        | (exact_projection & report[projection_columns].abs().gt(1e-12).any(axis=1))
    )
    report["base_year_convention"] = "ESTO 2024=2022; ESTO 2025=2023; ESTO 2026=2024"
    report["interpretation"] = report["nonzero_in_any_vintage"].map(
        {True: "ESTO has nonzero source value; investigate omission", False: "ESTO source is zero across reported vintage windows"}
    )
    report["seed_impact"] = report.apply(
        lambda row: (
            "No matching seed rows; nonzero source energy omitted"
            if not row["seed_matching_rows"] and row["actionable_source_nonzero"]
            else "Rows retained in seed with nonzero source energy but unresolved IDs"
            if row["seed_matching_rows"] and row["actionable_source_nonzero"]
            else "Rows retained but source energy is zero; suppress metadata/activity-only rows"
            if row["seed_matching_rows"]
            else "No matching seed rows; source energy is zero"
        ),
        axis=1,
    )
    report.to_csv(output_path, index=False, float_format="%.12g")
    return report


def build_creation_instructions(report: pd.DataFrame, output_path: Path) -> pd.DataFrame:
    instructions = report[
        report["actionable_source_nonzero"]
    ].copy()
    instructions["parent_path"] = instructions["branch_path"].str.rsplit("\\", n=1).str[0]
    instructions["branch_label"] = instructions["branch_path"].str.rsplit("\\", n=1).str[-1]
    instructions["branch_kind"] = instructions["branch_path"].map(
        lambda path: "transformation feedstock fuel leaf"
        if "\\Feedstock Fuels\\" in path
        else "transformation output fuel leaf"
        if "\\Output Fuels\\" in path
        else "demand loss/own-use fuel leaf"
        if path.startswith("Demand\\")
        else "transformation/process branch"
    )
    instructions["create_instruction"] = instructions.apply(
        lambda row: (
            f"Create '{row['branch_label']}' under '{row['parent_path']}' in the {row['economy']} LEAP area and export template; "
            "copy the parent/sibling variable, scenario, region, unit, and expression configuration; assign a real BranchID; "
            "then rerun baseline-seed generation and the missing-branch source-value check."
        ),
        axis=1,
    )
    columns = [
        "economy", "branch_path", "parent_path", "branch_label", "branch_kind",
        "source_flow", "esto_product", "actionable_source_nonzero", "seed_impact", "seed_matching_rows", "seed_activity_rows",
        "esto_2024_base_2022", "esto_2025_base_2023", "esto_2026_base_2024",
        "ninth_2024_projected_sum", "ninth_2025_projected_sum", "ninth_2026_projected_sum",
        "ninth_match_mode", "create_instruction",
    ]
    instructions = instructions[columns].sort_values(["economy", "branch_path"])
    instructions.to_csv(output_path, index=False, float_format="%.12g")
    return instructions


if __name__ == "__main__":
    output = REPO_ROOT / "outputs" / "leap_exports" / "supply_reconciliation" / "baseline_seed_missing_branch_esto_vintage_impact_20260818.csv"
    result = build_report(output)
    instructions_output = REPO_ROOT / "outputs" / "leap_exports" / "supply_reconciliation" / "missing_branch_creation_instructions_20260818.csv"
    build_creation_instructions(result, instructions_output)
    print(f"Wrote {len(result)} rows to {output}")
    print(f"Wrote branch creation instructions to {instructions_output}")
    print(result["nonzero_in_any_vintage"].value_counts(dropna=False).to_string())
