#%%
"""Append validated AUTO BALANCE rows to the active per-economy LEAP templates.

The supplied AUS workbook is a row-pattern reference only.  Each target gets
its own Region and its own VariableID/ScenarioID/RegionID values copied from a
matching existing fuel row; BranchID remains -1 because the new branch does
not yet exist in that area's exported LEAP template.
"""

from __future__ import annotations

import os
import shutil
from copy import copy
from pathlib import Path

from openpyxl import load_workbook


REPO_ROOT = Path(__file__).resolve().parents[2]
TEMPLATE_ROOT = REPO_ROOT / "data" / "leap_export_templates"
EXAMPLE_PATH = Path(r"C:\Users\Work\OneDrive - APERC\new rows transfers AUS.xlsx")
TARGET_ECONOMIES = [
    "AUS", "BD", "PRC", "MAS", "MEX", "NZ", "PNG", "PHL", "THA", "USA", "VN",
]
SHEET_NAME = "Export"
HEADER_ROW = 3
AUTO_BALANCE_LABEL = "AUTO BALANCE"
ID_COLUMNS = ("BranchID", "VariableID", "ScenarioID", "RegionID")
KEY_COLUMNS = ("Branch Path", "Variable", "Scenario", "Region")


def _header_lookup(sheet) -> dict[str, int]:
    """Return one-based column indexes for the fixed LEAP header row."""
    headers = {
        str(cell.value).strip(): cell.column
        for cell in sheet[HEADER_ROW]
        if cell.value is not None
    }
    required = set(ID_COLUMNS) | set(KEY_COLUMNS)
    missing = sorted(required - set(headers))
    if missing:
        raise ValueError(f"{sheet.title} is missing LEAP columns: {missing}")
    return headers


def _find_target_template(economy: str) -> Path:
    """Resolve exactly one active template by its economy letter token."""
    candidates = sorted(TEMPLATE_ROOT.glob(f"*{economy}*.xlsx"))
    if len(candidates) != 1:
        raise ValueError(f"Expected exactly one {economy} template, found: {candidates}")
    return candidates[0]


def _read_example_rows() -> tuple[list[object], list[dict[str, object]]]:
    """Read the 72 AUTO BALANCE reference rows without trusting their IDs."""
    workbook = load_workbook(EXAMPLE_PATH, read_only=True, data_only=False)
    try:
        sheet = workbook[workbook.sheetnames[0]]
        headers = [cell.value for cell in sheet[HEADER_ROW]]
        lookup = _header_lookup(sheet)
        rows = []
        for values in sheet.iter_rows(min_row=HEADER_ROW + 1, values_only=True):
            row = dict(zip(headers, values))
            if AUTO_BALANCE_LABEL in str(row.get("Branch Path") or ""):
                rows.append(row)
    finally:
        workbook.close()
    if len(rows) != 72:
        raise ValueError(f"Expected 72 example AUTO BALANCE rows, found {len(rows)}")
    if len({tuple(row[column] for column in KEY_COLUMNS) for row in rows}) != len(rows):
        raise ValueError("Example AUTO BALANCE rows have duplicate LEAP keys.")
    return headers, rows


def _target_region(sheet, lookup: dict[str, int]) -> str:
    """Require a single target region instead of guessing from a filename."""
    regions = {
        str(sheet.cell(row=row, column=lookup["Region"]).value).strip()
        for row in range(HEADER_ROW + 1, sheet.max_row + 1)
        if sheet.cell(row=row, column=lookup["Region"]).value not in (None, "")
    }
    if len(regions) != 1:
        raise ValueError(f"{sheet.title} must have one region; found {sorted(regions)}")
    return regions.pop()


def _matching_target_row_index(sheet, lookup: dict[str, int]) -> dict[tuple[str, object, object], int]:
    """Index existing non-AUTO-BALANCE fuel rows by process side/variable/scenario."""
    matches: dict[tuple[str, object, object], int] = {}
    for row in range(HEADER_ROW + 1, sheet.max_row + 1):
        path = str(sheet.cell(row=row, column=lookup["Branch Path"]).value or "")
        if not path or path.endswith(AUTO_BALANCE_LABEL):
            continue
        key = (
            path.rsplit("\\", 1)[0] + "\\",
            sheet.cell(row=row, column=lookup["Variable"]).value,
            sheet.cell(row=row, column=lookup["Scenario"]).value,
        )
        matches.setdefault(key, row)
    return matches


def _append_missing_rows(
    template_path: Path,
    headers: list[object],
    example_rows: list[dict[str, object]],
    apply_changes: bool,
) -> dict[str, object]:
    """Append only absent keys, atomically replacing the workbook when requested."""
    workbook = load_workbook(template_path)
    try:
        sheet = workbook[SHEET_NAME]
        lookup = _header_lookup(sheet)
        target_region = _target_region(sheet, lookup)
        existing_keys = {
            tuple(sheet.cell(row=row, column=lookup[column]).value for column in KEY_COLUMNS)
            for row in range(HEADER_ROW + 1, sheet.max_row + 1)
        }
        target_row_index = _matching_target_row_index(sheet, lookup)
        planned_rows = []
        for example_row in example_rows:
            row = dict(example_row)
            row["Region"] = target_region
            key = tuple(row[column] for column in KEY_COLUMNS)
            if key in existing_keys:
                continue
            match_key = (
                str(row["Branch Path"]).rsplit("\\", 1)[0] + "\\",
                row["Variable"],
                row["Scenario"],
            )
            matched_row = target_row_index.get(match_key)
            if matched_row is None:
                raise ValueError(f"No target row matches AUTO BALANCE key {match_key}")
            for column in ("VariableID", "ScenarioID", "RegionID"):
                row[column] = sheet.cell(row=matched_row, column=lookup[column]).value
            row["BranchID"] = -1
            planned_rows.append((row, matched_row))

        if len(planned_rows) not in {0, 72}:
            raise ValueError(
                f"{template_path.name} has a partial AUTO BALANCE set: "
                f"would append {len(planned_rows)} rows, not 0 or 72."
            )
        if apply_changes and planned_rows:
            destination_row = sheet.max_row + 1
            for row, matched_row in planned_rows:
                for column_index, header in enumerate(headers, start=1):
                    source_cell = sheet.cell(row=matched_row, column=column_index)
                    destination_cell = sheet.cell(row=destination_row, column=column_index)
                    destination_cell._style = copy(source_cell._style)
                    if source_cell.number_format:
                        destination_cell.number_format = source_cell.number_format
                    destination_cell.value = row.get(header)
                destination_row += 1
            temporary_path = template_path.with_suffix(".tmp.xlsx")
            workbook.save(temporary_path)
            os.replace(temporary_path, template_path)
        return {
            "template": template_path.name,
            "region": target_region,
            "rows_added": len(planned_rows),
        }
    finally:
        workbook.close()


def validate_auto_balance_rows(template_path: Path) -> None:
    """Confirm the added set is complete, key-unique, and explicitly unresolved."""
    workbook = load_workbook(template_path, read_only=True, data_only=False)
    try:
        sheet = workbook[SHEET_NAME]
        lookup = _header_lookup(sheet)
        rows = []
        for row_number in range(HEADER_ROW + 1, sheet.max_row + 1):
            path = str(sheet.cell(row=row_number, column=lookup["Branch Path"]).value or "")
            if path.endswith(AUTO_BALANCE_LABEL):
                rows.append({
                    column: sheet.cell(row=row_number, column=lookup[column]).value
                    for column in (*ID_COLUMNS, *KEY_COLUMNS)
                })
        if len(rows) != 72:
            raise ValueError(f"{template_path.name} has {len(rows)} AUTO BALANCE rows, expected 72")
        if len({tuple(row[column] for column in KEY_COLUMNS) for row in rows}) != 72:
            raise ValueError(f"{template_path.name} has duplicate AUTO BALANCE keys")
        if {row["BranchID"] for row in rows} != {-1}:
            raise ValueError(f"{template_path.name} AUTO BALANCE rows must keep BranchID=-1")
        if any(row[column] in (None, "") for row in rows for column in ("VariableID", "ScenarioID", "RegionID")):
            raise ValueError(f"{template_path.name} has missing non-branch IDs")
    finally:
        workbook.close()


def update_all_templates(apply_changes: bool = False) -> list[dict[str, object]]:
    """Dry-run or append the complete row set to every active economy template."""
    headers, example_rows = _read_example_rows()
    results = []
    for economy in TARGET_ECONOMIES:
        template_path = _find_target_template(economy)
        result = _append_missing_rows(
            template_path,
            headers=headers,
            example_rows=example_rows,
            apply_changes=apply_changes,
        )
        if apply_changes or result["rows_added"] == 0:
            validate_auto_balance_rows(template_path)
        results.append(result)
        print(result)
    return results


#%%
# Toggle from a Jupyter cell or run this module with the chosen value below.
APPLY_CHANGES = False

if __name__ == "__main__":
    update_all_templates(apply_changes=APPLY_CHANGES)

#%%
