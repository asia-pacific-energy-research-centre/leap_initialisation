#%%
"""Add reviewed missing baseline-seed branches to LEAP export templates.

The enabled paths in ``baseline_seed_validation_exception_sets.xlsx`` are
branches that baseline seeds can produce but the current LEAP areas do not
contain.  This notebook-safe helper makes the gap visible in each local export
template by cloning a complete sibling-leaf profile and setting ``BranchID`` to
``99``.  ID 99 is deliberately a display-only placeholder: it cannot make an
import target a branch that has not yet been created in the LEAP area.
"""
from __future__ import annotations

import os
import re
from copy import deepcopy
from pathlib import Path
from types import SimpleNamespace
from xml.etree import ElementTree
from zipfile import ZIP_DEFLATED, ZipFile
from xml.sax.saxutils import escape as xml_escape

from openpyxl import load_workbook

from codebase.functions.patch_baseline_seeds import load_validation_exception_branch_notes
from codebase.utilities import leap_export_template_resolver

REPO_ROOT = Path(__file__).resolve().parents[2]
TEMPLATE_ROOT = REPO_ROOT / "data" / "leap_export_templates"
PLACEHOLDER_BRANCH_ID = 99
SHEET_XML_PATH = "xl/worksheets/sheet1.xml"
WORKBOOK_XML_PATH = "xl/workbook.xml"
WORKBOOK_RELS_XML_PATH = "xl/_rels/workbook.xml.rels"
NAMESPACE = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
RELATIONSHIP_NAMESPACE = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}"
PACKAGE_RELATIONSHIP_NAMESPACE = "{http://schemas.openxmlformats.org/package/2006/relationships}"
REQUIRED_HEADERS = {"BranchID", "Branch Path", "Variable", "Scenario", "Region"}
EXCEPTION_SHEET_NAME = "branch_exceptions"
RESOLVED_IN_ALL_TEMPLATES_COLUMN = "resolved_in_all_templates"


def _column_name(cell) -> str:
    """Return Excel column letters from a cell reference such as ``AZ11252``."""
    match = re.match(r"[A-Z]+", cell.attrib.get("r", ""))
    return match.group(0) if match is not None else ""


def _read_template_xml(template_path: Path):
    """Read the Export sheet XML and resolve its optional shared-string table."""
    with ZipFile(template_path) as archive:
        sheet_xml_path = _export_sheet_xml_path(archive)
        root = ElementTree.fromstring(archive.read(sheet_xml_path))
        shared_strings = []
        if "xl/sharedStrings.xml" in archive.namelist():
            shared_root = ElementTree.fromstring(archive.read("xl/sharedStrings.xml"))
            shared_strings = ["".join(item.itertext()) for item in shared_root.iter(f"{NAMESPACE}si")]
    return root, shared_strings, sheet_xml_path


def _export_sheet_xml_path(archive: ZipFile) -> str:
    """Resolve the ``Export`` worksheet XML rather than assuming ``sheet1``.

    Some template workbooks put a human-facing LEAP sheet before the Export
    sheet.  Writing sheet1 in those files would silently modify the wrong
    worksheet, so only a deliberately tiny test fixture may use that fallback.
    """
    if WORKBOOK_XML_PATH not in archive.namelist() or WORKBOOK_RELS_XML_PATH not in archive.namelist():
        return SHEET_XML_PATH
    workbook = ElementTree.fromstring(archive.read(WORKBOOK_XML_PATH))
    export_sheet = next(
        (sheet for sheet in workbook.iter(f"{NAMESPACE}sheet") if sheet.attrib.get("name") == "Export"),
        None,
    )
    if export_sheet is None:
        raise ValueError("Workbook does not contain an Export sheet")
    relationship_id = export_sheet.attrib.get(f"{RELATIONSHIP_NAMESPACE}id")
    relationships = ElementTree.fromstring(archive.read(WORKBOOK_RELS_XML_PATH))
    relationship = next(
        (item for item in relationships.iter(f"{PACKAGE_RELATIONSHIP_NAMESPACE}Relationship") if item.attrib.get("Id") == relationship_id),
        None,
    )
    if relationship is None:
        raise ValueError("Could not resolve the Export sheet relationship")
    target = relationship.attrib.get("Target", "").lstrip("/")
    return target if target.startswith("xl/") else f"xl/{target}"


def _cell_value(cell, shared_strings: list[str]) -> str:
    if cell is None:
        return ""
    value_node = cell.find(f"{NAMESPACE}v")
    if cell.attrib.get("t") == "s" and value_node is not None:
        return shared_strings[int(value_node.text)]
    return "".join(cell.itertext())


def _header_columns(root, shared_strings: list[str]) -> dict[str, str]:
    """Return export header names keyed to their Excel column letters."""
    for row in root.iter(f"{NAMESPACE}row"):
        cells = {_column_name(cell): cell for cell in row}
        values = {_cell_value(cell, shared_strings): column for column, cell in cells.items()}
        if {"Branch Path", "Variable"}.issubset(values):
            return values
    raise ValueError("Could not locate the LEAP export header row")


def _set_inline_string(cell, value: str) -> None:
    cell.attrib["t"] = "inlineStr"
    for child in list(cell):
        cell.remove(child)
    inline = ElementTree.SubElement(cell, f"{NAMESPACE}is")
    ElementTree.SubElement(inline, f"{NAMESPACE}t").text = value


def _set_numeric_value(cell, value: int) -> None:
    cell.attrib.pop("t", None)
    for child in list(cell):
        cell.remove(child)
    ElementTree.SubElement(cell, f"{NAMESPACE}v").text = str(value)


def _raw_source_row_xml(sheet_xml: str, row_number: int) -> str:
    """Return one original row fragment without reserializing worksheet XML."""
    row_tag = r"(?:[A-Za-z_][\w.-]*:)?row"
    match = re.search(
        rf"(<{row_tag}\b[^>]*\br=\"{row_number}\"[^>]*>.*?</{row_tag}>)",
        sheet_xml,
        flags=re.DOTALL,
    )
    if match is None:
        raise ValueError(f"Could not locate source row {row_number} in worksheet XML")
    return match.group(1)


def _replace_raw_cell(row_xml: str, column: str, row_number: int, value: str, *, numeric: bool) -> str:
    """Replace one populated cell while retaining the source row's XML prefix/style."""
    cell_tag = r"(?:[A-Za-z_][\w.-]*:)?c"
    close_tag = r"(?:[A-Za-z_][\w.-]*:)?c"
    pattern = (
        rf"(?P<tag><{cell_tag}\b)(?P<attributes>[^>]*\br=\"{column}{row_number}\"[^>]*)>"
        rf"(?P<content>.*?)</{close_tag}>"
    )
    match = re.search(pattern, row_xml, flags=re.DOTALL)
    if match is None:
        raise ValueError(f"Source row lacks required cell {column}{row_number}")
    attributes = re.sub(r'\s+t="[^"]*"', "", match.group("attributes"))
    if numeric:
        replacement = f'{match.group("tag")}{attributes}><v>{value}</v></c>'
    else:
        replacement = f'{match.group("tag")}{attributes} t="inlineStr"><is><t>{xml_escape(value)}</t></is></c>'
    return row_xml[:match.start()] + replacement + row_xml[match.end():]


def _build_raw_placeholder_row(
    sheet_xml: str,
    source_row,
    target_path: str,
    headers: dict[str, str],
    next_row: int,
) -> str:
    """Clone a row in raw XML, preserving workbook namespace declarations."""
    source_row_number = int(source_row.attrib["r"])
    row_xml = _raw_source_row_xml(sheet_xml, source_row_number)
    row_xml = re.sub(rf'(\br="){source_row_number}("[^>]*>)', rf'\g<1>{next_row}\g<2>', row_xml, count=1)
    row_xml = re.sub(
        rf'([A-Z]+){source_row_number}(?=")',
        rf'\g<1>{next_row}',
        row_xml,
    )
    row_xml = _replace_raw_cell(
        row_xml, headers["BranchID"], next_row, str(PLACEHOLDER_BRANCH_ID), numeric=True
    )
    row_xml = _replace_raw_cell(row_xml, headers["Branch Path"], next_row, target_path, numeric=False)
    for index, part in enumerate(target_path.split("\\"), start=1):
        level_column = headers.get(f"Level {index}")
        if level_column is not None:
            row_xml = _replace_raw_cell(row_xml, level_column, next_row, part, numeric=False)
    return row_xml


def _append_raw_rows_to_sheet_xml(sheet_xml: str, rows: list[str], last_row: int) -> str:
    """Append row fragments before sheetData's closing tag without changing namespaces."""
    close_match = re.search(r"</(?:[A-Za-z_][\w.-]*:)?sheetData>", sheet_xml)
    if close_match is None:
        raise ValueError("Worksheet XML has no closing sheetData tag")
    updated = sheet_xml[:close_match.start()] + "".join(rows) + sheet_xml[close_match.start():]
    return re.sub(
        r'(<(?:[A-Za-z_][\w.-]*:)?dimension\b[^>]*\bref="[^"]*:[A-Z]+)\d+([^"]*")',
        rf'\g<1>{last_row}\g<2>',
        updated,
        count=1,
    )


def _leaf_profile_rows(
    template_rows: list,
    headers: dict[str, str],
    shared_strings: list[str],
    target_path: str,
) -> list:
    """Choose one complete direct-sibling leaf profile for a missing leaf.

    A leaf's profile is its complete Variable/Scenario/Region set.  Requiring
    an unambiguous largest profile prevents a partial or arbitrary row subset
    from being presented as a proposed LEAP branch.
    """
    parent = target_path.rsplit("\\", 1)[0]
    prefix = parent + "\\"
    sibling_rows: dict[str, list] = {}
    for row in template_rows:
        cells = {_column_name(cell): cell for cell in row}
        branch_path = _cell_value(cells.get(headers["Branch Path"]), shared_strings)
        if not branch_path.startswith(prefix):
            continue
        leaf = branch_path[len(prefix):]
        if not leaf or "\\" in leaf or branch_path == target_path:
            continue
        sibling_rows.setdefault(branch_path, []).append(row)

    if not sibling_rows:
        raise ValueError(f"No direct sibling leaf exists for {target_path!r}")

    def profile_key(rows: list) -> tuple[tuple[str, str, str], ...]:
        keys = set()
        for row in rows:
            cells = {_column_name(cell): cell for cell in row}
            keys.add(tuple(_cell_value(cells.get(headers[column]), shared_strings) for column in ("Variable", "Scenario", "Region")))
        if len(keys) != len(rows):
            raise ValueError(f"Sibling profile has duplicate Variable/Scenario/Region rows for {target_path!r}")
        return tuple(sorted(keys))

    candidates = [(path, rows, profile_key(rows)) for path, rows in sibling_rows.items()]
    maximum_size = max(len(rows) for _, rows, _ in candidates)
    largest = [(path, rows, profile) for path, rows, profile in candidates if len(rows) == maximum_size]
    profiles = {profile for _, _, profile in largest}
    if len(profiles) != 1:
        paths = sorted(path for path, _, _ in largest)
        raise ValueError(
            f"Ambiguous sibling profiles for {target_path!r}; explicitly create the branch in LEAP. "
            f"Candidates: {paths}"
        )
    return sorted(largest, key=lambda item: item[0])[0][1]


def _exception_rows(template_path: Path, branch_paths: set[str]) -> list[dict[str, str]]:
    root, shared_strings, _ = _read_template_xml(template_path)
    headers = _header_columns(root, shared_strings)
    rows = []
    for row in root.iter(f"{NAMESPACE}row"):
        cells = {_column_name(cell): cell for cell in row}
        path = _cell_value(cells.get(headers["Branch Path"]), shared_strings)
        if path in branch_paths:
            rows.append({
                "path": path,
                "branch_id": _cell_value(cells.get(headers["BranchID"]), shared_strings),
                "variable": _cell_value(cells.get(headers["Variable"]), shared_strings),
                "scenario": _cell_value(cells.get(headers["Scenario"]), shared_strings),
                "region": _cell_value(cells.get(headers["Region"]), shared_strings),
            })
    return rows


def validate_exception_placeholder_rows(template_path: Path, branch_paths: set[str]) -> None:
    """Confirm every requested branch is present and any proposal uses ID 99.

    An exception can already be a real branch in some economy templates.  Such
    rows retain their real local BranchID and are intentionally not rewritten.
    """
    rows = _exception_rows(template_path, branch_paths)
    found_paths = {row["path"] for row in rows}
    missing = sorted(branch_paths - found_paths)
    if missing:
        raise ValueError(f"{template_path.name} is missing proposed branch rows: {missing}")
    for branch_path in branch_paths:
        path_ids = {row["branch_id"] for row in rows if row["path"] == branch_path}
        if str(PLACEHOLDER_BRANCH_ID) in path_ids and path_ids != {str(PLACEHOLDER_BRANCH_ID)}:
            raise ValueError(
                f"{template_path.name} mixes BranchID={PLACEHOLDER_BRANCH_ID} with real IDs "
                f"for proposed path {branch_path!r}"
            )
    keys = {(row["path"], row["variable"], row["scenario"], row["region"]) for row in rows}
    if len(keys) != len(rows):
        raise ValueError(f"{template_path.name} has duplicate proposed branch row keys")


def insert_missing_exception_rows(
    template_path: Path | str,
    branch_paths: set[str],
    *,
    apply_changes: bool = False,
) -> dict[str, object]:
    """Append enabled missing leaves to one template, cloning local sibling profiles."""
    template_path = Path(template_path)
    branch_paths = {str(path).strip() for path in branch_paths if str(path).strip()}
    if not branch_paths:
        return {"template": template_path.name, "rows_added": 0}

    root, shared_strings, sheet_xml_path = _read_template_xml(template_path)
    headers = _header_columns(root, shared_strings)
    if not REQUIRED_HEADERS.issubset(headers):
        raise ValueError(f"{template_path.name} lacks required LEAP export columns")
    sheet_data = root.find(f"{NAMESPACE}sheetData")
    if sheet_data is None:
        raise ValueError(f"{template_path.name} has no sheet data")
    template_rows = list(sheet_data)
    existing_paths = {
        _cell_value({_column_name(cell): cell for cell in row}.get(headers["Branch Path"]), shared_strings)
        for row in template_rows
    }
    missing_paths = sorted(branch_paths - existing_paths)
    if not missing_paths:
        validate_exception_placeholder_rows(template_path, branch_paths)
        return {"template": template_path.name, "rows_added": 0}

    rows_to_add = [(source_row, target_path) for target_path in missing_paths for source_row in _leaf_profile_rows(template_rows, headers, shared_strings, target_path)]
    if apply_changes:
        next_row = max(int(row.attrib["r"]) for row in template_rows) + 1
        with ZipFile(template_path) as source_archive:
            original_xml = source_archive.read(sheet_xml_path).decode("utf-8")
        raw_rows = []
        for source_row, target_path in rows_to_add:
            raw_rows.append(_build_raw_placeholder_row(
                original_xml, source_row, target_path, headers, next_row,
            ))
            next_row += 1
        updated_xml = _append_raw_rows_to_sheet_xml(original_xml, raw_rows, next_row - 1).encode("utf-8")
        temporary_path = template_path.with_suffix(".tmp.xlsx")
        with ZipFile(template_path) as source_archive, ZipFile(temporary_path, "w", ZIP_DEFLATED) as destination_archive:
            for item in source_archive.infolist():
                content = updated_xml if item.filename == sheet_xml_path else source_archive.read(item.filename)
                destination_archive.writestr(item, content)
        os.replace(temporary_path, template_path)
        validate_exception_placeholder_rows(template_path, branch_paths)
    return {"template": template_path.name, "rows_added": len(rows_to_add)}


def update_all_templates(
    *,
    apply_changes: bool = False,
    templates_root: Path | str = TEMPLATE_ROOT,
) -> list[dict[str, object]]:
    """Dry-run or append every enabled validation-exception branch to each template."""
    branch_paths = set(load_validation_exception_branch_notes())
    results = []
    for template in leap_export_template_resolver.iter_leap_export_templates(templates_root):
        result = insert_missing_exception_rows(template.path, branch_paths, apply_changes=apply_changes)
        result["economy"] = template.economy
        results.append(result)
        print(result)
    return results


def _template_presence_status(
    branch_paths: set[str],
    templates: list[SimpleNamespace],
) -> dict[str, bool]:
    """Return whether each path is a real branch in every supplied template."""
    status = {path: True for path in branch_paths}
    for template in templates:
        rows = _exception_rows(template.path, branch_paths)
        for branch_path in branch_paths:
            matching_rows = [row for row in rows if row["path"] == branch_path]
            has_real_branch = bool(matching_rows) and all(
                row["branch_id"] != str(PLACEHOLDER_BRANCH_ID)
                for row in matching_rows
            )
            status[branch_path] = status[branch_path] and has_real_branch
    return status


def sync_exception_resolution_status(
    *,
    exception_workbook_path: Path = REPO_ROOT / "config" / "baseline_seed_validation_exception_sets.xlsx",
    templates_root: Path | str = TEMPLATE_ROOT,
) -> dict[str, bool]:
    """Record resolved exceptions and disable only those real in every template.

    This is intentionally a separate operation from row insertion.  Run it
    after refreshed real LEAP exports replace the ID-99 placeholders.  A path
    that is still absent or has any ID-99 row remains enabled for baseline-seed
    validation, even when it is real in some economies.
    """
    workbook_path = Path(exception_workbook_path)
    workbook = load_workbook(workbook_path)
    if EXCEPTION_SHEET_NAME not in workbook.sheetnames:
        raise ValueError(f"{workbook_path.name} has no {EXCEPTION_SHEET_NAME!r} sheet")
    sheet = workbook[EXCEPTION_SHEET_NAME]
    headers = {
        str(cell.value).strip(): cell.column
        for cell in sheet[1]
        if cell.value is not None and str(cell.value).strip()
    }
    required = {"enabled", "branch_path", "notes"}
    missing = required - set(headers)
    if missing:
        raise ValueError(f"{workbook_path.name} is missing required columns: {sorted(missing)}")
    if RESOLVED_IN_ALL_TEMPLATES_COLUMN not in headers:
        column = sheet.max_column + 1
        sheet.cell(row=1, column=column, value=RESOLVED_IN_ALL_TEMPLATES_COLUMN)
        headers[RESOLVED_IN_ALL_TEMPLATES_COLUMN] = column

    path_rows = {
        str(sheet.cell(row=row, column=headers["branch_path"]).value or "").strip(): row
        for row in range(2, sheet.max_row + 1)
    }
    path_rows = {path: row for path, row in path_rows.items() if path}
    templates = [SimpleNamespace(path=item.path, economy=item.economy) for item in leap_export_template_resolver.iter_leap_export_templates(templates_root)]
    if not templates:
        raise FileNotFoundError(f"No LEAP export templates found under {templates_root}")
    status = _template_presence_status(set(path_rows), templates)
    for branch_path, row in path_rows.items():
        resolved = status[branch_path]
        sheet.cell(row=row, column=headers[RESOLVED_IN_ALL_TEMPLATES_COLUMN], value=resolved)
        if resolved:
            sheet.cell(row=row, column=headers["enabled"], value=False)
    workbook.save(workbook_path)
    return status


#%%
APPLY_CHANGES = False

if __name__ == "__main__":
    update_all_templates(apply_changes=APPLY_CHANGES)
    # Run manually after refreshed real LEAP exports replace ID-99 rows:
    # sync_exception_resolution_status()

#%%
