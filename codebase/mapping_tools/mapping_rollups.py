#%%
"""
Shared rollup and cardinality helpers for LEAP / ESTO / 9th mapping workbooks.

The helpers keep researcher-maintained mapping rows simple, apply explicit
rollup rules before cardinality checks, and produce QA tables for unresolved
many-to-many mappings.
"""

#%%
import hashlib
import re
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

#%%
ROLLUP_SHEET_COLUMNS = {
    "leap_rollup_rules": [
        "rollup_context",
        "input_leap_sector_name_full_path",
        "input_raw_leap_fuel_name",
        "rolled_leap_sector_name_full_path",
        "rolled_raw_leap_fuel_name",
        "rollup_group_id",
        "rollup_reason",
        "priority",
        "include",
        "Note",
    ],
    "esto_rollup_rules": [
        "rollup_context",
        "input_esto_flow",
        "input_esto_product",
        "rolled_esto_flow",
        "rolled_esto_product",
        "rollup_group_id",
        "rollup_reason",
        "priority",
        "include",
        "Note",
    ],
    "ninth_rollup_rules": [
        "rollup_context",
        "input_9th_sector",
        "input_9th_fuel",
        "rolled_9th_sector",
        "rolled_9th_fuel",
        "rollup_group_id",
        "rollup_reason",
        "priority",
        "include",
        "Note",
    ],
    "rollup_label_overrides": [
        "rollup_group_id",
        "auto_rollup_code",
        "auto_rollup_name",
        "auto_rollup_label",
        "preferred_rollup_code",
        "preferred_rollup_name",
        "preferred_rollup_label",
        "Note",
    ],
}

INDIVIDUAL_MAPPING_EXCEPTION_COLUMNS = [
    "check_name",
    "source_value",
    "target_value",
    "include",
    "exception_reason",
    "Note",
]

ORIGINAL_LABEL_DATASET_BY_LABEL = {
    "raw_leap_fuel_name": "LEAP model export",
    "ninth_fuel": "9th Outlook data",
    "esto_product": "ESTO data",
}

SUBTOTAL_DATASET_BY_SYSTEM = {
    "LEAP": "LEAP model export / mapping hierarchy",
    "NINTH": "9th Outlook data",
    "ESTO": "ESTO data",
}

MAPPING_SHEET_CONFIGS = [
    {
        "source_sheet": "leap_combined_esto",
        "use_case": "leap_to_esto_balance_conversion",
        "source_system": "LEAP",
        "target_system": "ESTO",
        "rollup_context": "leap_to_esto",
        "source_cols": ["leap_sector_name_full_path", "raw_leap_fuel_name"],
        "target_cols": ["esto_flow", "esto_product"],
        "source_rollup_sheet": "leap_rollup_rules",
        "target_rollup_sheet": "esto_rollup_rules",
    },
    {
        "source_sheet": "ninth_pairs_to_esto_pairs",
        "use_case": "ninth_to_esto_balance_conversion",
        "source_system": "NINTH",
        "target_system": "ESTO",
        "rollup_context": "ninth_to_esto",
        "source_cols": ["9th_sector", "9th_fuel"],
        "target_cols": ["esto_flow", "esto_product"],
        "source_rollup_sheet": "ninth_rollup_rules",
        "target_rollup_sheet": "esto_rollup_rules",
    },
    {
        "source_sheet": "leap_combined_ninth",
        "use_case": "leap_to_ninth_comparison",
        "source_system": "LEAP",
        "target_system": "NINTH",
        "rollup_context": "leap_to_ninth",
        "source_cols": ["leap_sector_name_full_path", "raw_leap_fuel_name"],
        "target_cols": ["ninth_sector", "ninth_fuel"],
        "source_rollup_sheet": "leap_rollup_rules",
        "target_rollup_sheet": "ninth_rollup_rules",
    },
    {
        "source_sheet": "leap_combined_ninth",
        "use_case": "ninth_to_leap_initialisation",
        "source_system": "NINTH",
        "target_system": "LEAP",
        "rollup_context": "ninth_to_leap",
        "source_cols": ["ninth_sector", "ninth_fuel"],
        "target_cols": ["leap_sector_name_full_path", "raw_leap_fuel_name"],
        "source_rollup_sheet": "ninth_rollup_rules",
        "target_rollup_sheet": "leap_rollup_rules",
        "is_reverse": True,
    },
]

RELATIONSHIP_COLUMNS = [
    "relationship_id",
    "relationship_key",
    "use_case",
    "include_in_use_case",
    "source_system",
    "source_flow",
    "source_product",
    "target_system",
    "target_flow",
    "target_product",
    "original_source_flow",
    "original_source_product",
    "original_target_flow",
    "original_target_product",
    "rolled_source_flow",
    "rolled_source_product",
    "rolled_target_flow",
    "rolled_target_product",
    "rollup_applied",
    "source_rollup_group_id",
    "target_rollup_group_id",
    "rollup_reason",
    "pair_mapping_cardinality_raw",
    "pair_mapping_cardinality_after_rollup",
    "pair_mapping_cardinality",
    "cardinality",
    "allocation_method",
    "allocation_source",
    "allocation_share",
    "relationship_type",
    "relationship_level",
    "relationship_source",
    "source_sector_path",
    "source_fuel",
    "relationship_status",
    "exclude_reason",
    "review_required",
    "review_flags",
    "remove_row",
    "source_sheet",
    "source_row_number",
    "Note",
]

INDIVIDUAL_MAPPING_CHECKS = [
    {
        "check_name": "ninth_fuel_to_esto_product",
        "use_case": "ninth_to_esto_balance_conversion",
        "source_col": "rolled_source_product",
        "target_col": "rolled_target_product",
        "source_label": "9th_fuel",
        "target_label": "esto_product",
    },
    {
        "check_name": "raw_leap_fuel_name_to_ninth_fuel",
        "use_case": "leap_to_ninth_comparison",
        "source_col": "rolled_source_product",
        "target_col": "rolled_target_product",
        "source_label": "raw_leap_fuel_name",
        "target_label": "ninth_fuel",
    },
    {
        "check_name": "raw_leap_fuel_name_to_esto_product",
        "use_case": "leap_to_esto_balance_conversion",
        "source_col": "rolled_source_product",
        "target_col": "rolled_target_product",
        "source_label": "raw_leap_fuel_name",
        "target_label": "esto_product",
    },
]

#%%
def clean_text(value: Any) -> str:
    """Return a trimmed string while treating common null spellings as blank."""
    if pd.isna(value):
        return ""
    text = str(value).strip()
    return "" if text.lower() in {"nan", "none", "<na>"} else text


def normalise_key(value: Any) -> str:
    """Normalise mapping text for matching and grouping."""
    return re.sub(r"\s+", " ", clean_text(value).replace("\\", "/")).strip().lower()


def parse_bool(value: Any, default: bool = False) -> bool:
    """Interpret common spreadsheet boolean values."""
    text = clean_text(value).lower()
    if not text:
        return default
    if text in {"1", "true", "t", "yes", "y", "include", "included", "on"}:
        return True
    if text in {"0", "false", "f", "no", "n", "exclude", "excluded", "off"}:
        return False
    return default


def parse_priority(value: Any) -> float:
    """Return numeric priority where lower numbers win."""
    text = clean_text(value)
    if not text:
        return 9999.0
    try:
        return float(text)
    except ValueError:
        return 9999.0


def sanitize_identifier(value: Any) -> str:
    """Create a stable spreadsheet-friendly identifier."""
    text = normalise_key(value)
    text = re.sub(r"[^a-z0-9]+", "_", text)
    return text.strip("_")


def split_code_name(label: Any) -> tuple[str, str]:
    """Split a label into a leading code and a display name when possible."""
    text = clean_text(label)
    if not text:
        return "", ""
    match = re.match(r"^([0-9][0-9A-Za-z_.-]*(?:\.[0-9A-Za-z_.-]+)*(?:-[0-9A-Za-z_.-]+)?(?:,[0-9A-Za-z_.-]+(?:-[0-9A-Za-z_.-]+)?)*)\s+(.+)$", text)
    if match:
        return match.group(1).strip(), match.group(2).strip()
    return sanitize_identifier(text), text


def make_rollup_group_id(flow: Any, product: Any) -> str:
    """Create a deterministic group ID from a rolled flow/product pair."""
    label = " ".join(part for part in [clean_text(flow), clean_text(product)] if part)
    return sanitize_identifier(label)


def append_note(existing_note: Any, new_note: str) -> str:
    """Append a note once, preserving existing human notes."""
    existing = clean_text(existing_note)
    if not new_note:
        return existing
    parts = [part.strip() for part in existing.split(" | ") if part.strip()]
    if new_note in parts:
        return existing
    parts.append(new_note)
    return " | ".join(parts)


def make_relationship_id(row: pd.Series) -> str:
    """Create a stable ID for one effective source-to-target relationship."""
    parts = [
        clean_text(row.get("use_case", "")),
        clean_text(row.get("source_system", "")),
        normalise_key(row.get("rolled_source_flow", "")),
        normalise_key(row.get("rolled_source_product", "")),
        clean_text(row.get("target_system", "")),
        normalise_key(row.get("rolled_target_flow", "")),
        normalise_key(row.get("rolled_target_product", "")),
    ]
    digest = hashlib.sha1("||".join(parts).encode("utf-8")).hexdigest()[:16]
    return f"rel_{digest}"


def cardinality_label(source_target_count: int, target_source_count: int) -> str:
    """Label mapping cardinality from unique source/target counts."""
    if source_target_count <= 0 or target_source_count <= 0:
        return ""
    if source_target_count == 1 and target_source_count == 1:
        return "one_to_one"
    if source_target_count > 1 and target_source_count == 1:
        return "one_to_many"
    if source_target_count == 1 and target_source_count > 1:
        return "many_to_one"
    return "many_to_many"


def active_mask(frame: pd.DataFrame) -> pd.Series:
    """Return rows not marked remove_row or duplicate_to_remove."""
    mask = pd.Series(True, index=frame.index)
    for column in ["remove_row", "duplicate_to_remove", "removed", "is_removed"]:
        if column in frame.columns:
            mask &= ~frame[column].map(parse_bool)
    return mask


def context_applies(rule_context: Any, requested_context: str) -> bool:
    """Return True when a rollup rule context applies to this process."""
    text = normalise_key(rule_context)
    return text in {"", "all", normalise_key(requested_context)}


def value_matches(rule_value: Any, row_value: Any) -> bool:
    """Return True for exact or wildcard rule matching."""
    text = normalise_key(rule_value)
    return text in {"", "*", "all"} or text == normalise_key(row_value)


def rule_specificity(rule: pd.Series, input_flow_col: str, input_product_col: str) -> int:
    """Rank exact flow/product matches ahead of wildcard matches."""
    flow_is_wild = normalise_key(rule.get(input_flow_col, "")) in {"", "*", "all"}
    product_is_wild = normalise_key(rule.get(input_product_col, "")) in {"", "*", "all"}
    if not flow_is_wild and not product_is_wild:
        return 1
    if not flow_is_wild and product_is_wild:
        return 2
    if flow_is_wild and not product_is_wild:
        return 3
    return 4


def rollup_columns_for_sheet(rollup_sheet_name: str) -> dict[str, str]:
    """Return input/output columns for a rollup sheet."""
    if rollup_sheet_name == "leap_rollup_rules":
        return {
            "input_flow": "input_leap_sector_name_full_path",
            "input_product": "input_raw_leap_fuel_name",
            "rolled_flow": "rolled_leap_sector_name_full_path",
            "rolled_product": "rolled_raw_leap_fuel_name",
        }
    if rollup_sheet_name == "esto_rollup_rules":
        return {
            "input_flow": "input_esto_flow",
            "input_product": "input_esto_product",
            "rolled_flow": "rolled_esto_flow",
            "rolled_product": "rolled_esto_product",
        }
    if rollup_sheet_name == "ninth_rollup_rules":
        return {
            "input_flow": "input_9th_sector",
            "input_product": "input_9th_fuel",
            "rolled_flow": "rolled_9th_sector",
            "rolled_product": "rolled_9th_fuel",
        }
    raise ValueError(f"Unsupported rollup sheet: {rollup_sheet_name}")


#%%
def ensure_rollup_sheets(workbook_path: Path) -> None:
    """Create missing rollup-rule sheets with headers."""
    workbook = load_workbook(workbook_path)
    changed = False
    for sheet_name, columns in ROLLUP_SHEET_COLUMNS.items():
        if sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            existing_headers = [clean_text(cell.value) for cell in worksheet[1]]
            missing_columns = [column for column in columns if column not in existing_headers]
            if missing_columns:
                for column in missing_columns:
                    worksheet.cell(row=1, column=len(existing_headers) + 1, value=column)
                    existing_headers.append(column)
                changed = True
            continue
        worksheet = workbook.create_sheet(title=sheet_name)
        worksheet.append(columns)
        changed = True
    if changed:
        workbook.save(workbook_path)


def ensure_individual_mapping_exception_sheet(workbook_path: Path) -> None:
    """Create the editable individual-mapping exception sheet when missing."""
    workbook = load_workbook(workbook_path)
    if "individual_mapping_exceptions" in workbook.sheetnames:
        worksheet = workbook["individual_mapping_exceptions"]
        existing_headers = [clean_text(cell.value) for cell in worksheet[1]]
        missing_columns = [column for column in INDIVIDUAL_MAPPING_EXCEPTION_COLUMNS if column not in existing_headers]
        if not missing_columns:
            return
        for column in missing_columns:
            worksheet.cell(row=1, column=len(existing_headers) + 1, value=column)
            existing_headers.append(column)
        workbook.save(workbook_path)
        return
    worksheet = workbook.create_sheet(title="individual_mapping_exceptions")
    worksheet.append(INDIVIDUAL_MAPPING_EXCEPTION_COLUMNS)
    workbook.save(workbook_path)


def replace_sheet_with_dataframe(workbook_path: Path, sheet_name: str, frame: pd.DataFrame) -> None:
    """Replace one workbook sheet with a dataframe."""
    workbook = load_workbook(workbook_path)
    if sheet_name in workbook.sheetnames:
        sheet_index = workbook.sheetnames.index(sheet_name)
        del workbook[sheet_name]
        worksheet = workbook.create_sheet(title=sheet_name, index=sheet_index)
    else:
        worksheet = workbook.create_sheet(title=sheet_name)
    for row in dataframe_to_rows(frame, index=False, header=True):
        worksheet.append(row)
    workbook.save(workbook_path)


def read_required_sheet(workbook_path: Path, sheet_name: str) -> pd.DataFrame:
    """Read a workbook sheet as strings, raising a clear error if missing."""
    try:
        return pd.read_excel(workbook_path, sheet_name=sheet_name, dtype=object).fillna("")
    except ValueError as exc:
        raise ValueError(f"Required sheet {sheet_name!r} not found in {workbook_path}") from exc


def read_rollup_rules(workbook_path: Path) -> dict[str, pd.DataFrame]:
    """Read rollup sheets after ensuring they exist."""
    ensure_rollup_sheets(workbook_path)
    rules: dict[str, pd.DataFrame] = {}
    for sheet_name, columns in ROLLUP_SHEET_COLUMNS.items():
        frame = pd.read_excel(workbook_path, sheet_name=sheet_name, dtype=object).fillna("")
        for column in columns:
            if column not in frame.columns:
                frame[column] = ""
        rules[sheet_name] = frame.loc[:, columns].copy()
    return rules


def read_individual_mapping_exceptions(workbook_path: Path) -> pd.DataFrame:
    """Read optional approved exceptions for individual fuel/product checks."""
    columns = INDIVIDUAL_MAPPING_EXCEPTION_COLUMNS
    try:
        frame = pd.read_excel(workbook_path, sheet_name="individual_mapping_exceptions", dtype=object).fillna("")
    except ValueError:
        return pd.DataFrame(columns=columns)
    for column in columns:
        if column not in frame.columns:
            frame[column] = ""
    out = frame.loc[:, columns].copy()
    out = out[out["include"].map(lambda value: parse_bool(value, default=True))].copy()
    for column in ["check_name", "source_value", "target_value", "exception_reason", "Note"]:
        out[column] = out[column].fillna("").astype(str).str.strip()
    return out.reset_index(drop=True)


def active_rollup_rules(rules_df: pd.DataFrame, context: str) -> pd.DataFrame:
    """Filter rollup rules by include flag and context."""
    if rules_df.empty:
        return rules_df.copy()
    out = rules_df.copy()
    out["_rule_index"] = out.index + 2
    out = out[out["include"].map(lambda value: parse_bool(value, default=True))].copy()
    out = out[out["rollup_context"].map(lambda value: context_applies(value, context))].copy()
    return out.reset_index(drop=True)


def build_rollup_created_category_notes(
    rollup_rules: dict[str, pd.DataFrame],
    rollup_sheet_name: str,
    context: str,
) -> dict[str, str]:
    """Build notes for categories created by applicable rollup rules."""
    rules_df = active_rollup_rules(rollup_rules.get(rollup_sheet_name, pd.DataFrame()), context)
    if rules_df.empty:
        return {}
    columns = rollup_columns_for_sheet(rollup_sheet_name)
    notes: dict[str, str] = {}
    for _, rule in rules_df.iterrows():
        rolled_flow = clean_text(rule.get(columns["rolled_flow"], ""))
        rolled_product = clean_text(rule.get(columns["rolled_product"], ""))
        for rolled_value in [rolled_flow, rolled_product]:
            if not rolled_value:
                continue
            notes[normalise_key(rolled_value)] = (
                f"{rolled_value} is a rollup category from {rollup_sheet_name} "
                f"for {context}."
            )
    return notes


def annotate_rollup_created_categories(
    frame: pd.DataFrame,
    config: dict[str, Any],
    rollup_rules: dict[str, pd.DataFrame],
) -> pd.DataFrame:
    """Append Note text when mapping rows use rollup-created categories."""
    out = frame.copy()
    if "Note" not in out.columns:
        out["Note"] = ""
    sheet_pairs = [
        (config["source_rollup_sheet"], config["source_cols"]),
        (config["target_rollup_sheet"], config["target_cols"]),
    ]
    for rollup_sheet_name, mapping_cols in sheet_pairs:
        notes_by_value_and_col = build_rollup_created_category_notes(
            rollup_rules=rollup_rules,
            rollup_sheet_name=rollup_sheet_name,
            context=config["rollup_context"],
        )
        if not notes_by_value_and_col:
            continue
        for mapping_col in mapping_cols:
            if mapping_col not in out.columns:
                continue
            for idx, value in out[mapping_col].items():
                note_suffix = notes_by_value_and_col.get(normalise_key(value), "")
                if note_suffix:
                    note = f"{clean_text(value)} from {mapping_col} {note_suffix[len(clean_text(value)):].strip()}"
                    out.at[idx, "Note"] = append_note(out.at[idx, "Note"], note)
    return out


def choose_rollup_match(
    row: pd.Series,
    rules_df: pd.DataFrame,
    rollup_sheet_name: str,
    context: str,
    input_flow_value: Any,
    input_product_value: Any,
) -> tuple[pd.Series | None, pd.DataFrame]:
    """Choose the best rollup rule for a row and return ambiguous ties separately."""
    if rules_df.empty:
        return None, pd.DataFrame()
    columns = rollup_columns_for_sheet(rollup_sheet_name)
    candidates = rules_df[
        rules_df.apply(
            lambda rule: value_matches(rule.get(columns["input_flow"], ""), input_flow_value)
            and value_matches(rule.get(columns["input_product"], ""), input_product_value),
            axis=1,
        )
    ].copy()
    if candidates.empty:
        return None, pd.DataFrame()
    candidates["_specificity"] = candidates.apply(
        lambda rule: rule_specificity(rule, columns["input_flow"], columns["input_product"]),
        axis=1,
    )
    candidates["_priority"] = candidates["priority"].map(parse_priority)
    best_priority = candidates["_priority"].min()
    candidates = candidates[candidates["_priority"].eq(best_priority)].copy()
    best_specificity = candidates["_specificity"].min()
    best = candidates[candidates["_specificity"].eq(best_specificity)].copy()
    if len(best) > 1:
        ambiguous = best.copy()
        ambiguous["rollup_sheet"] = rollup_sheet_name
        ambiguous["rollup_context_used"] = context
        ambiguous["matched_flow"] = clean_text(input_flow_value)
        ambiguous["matched_product"] = clean_text(input_product_value)
        ambiguous["source_sheet"] = clean_text(row.get("source_sheet", ""))
        ambiguous["source_row_number"] = row.get("source_row_number", "")
        return None, ambiguous
    return best.iloc[0], pd.DataFrame()


def apply_rollup_axis(
    frame: pd.DataFrame,
    rules_df: pd.DataFrame,
    rollup_sheet_name: str,
    context: str,
    input_cols: list[str],
    output_prefix: str,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """Apply one side of rollup rules to a mapping table."""
    out = frame.copy()
    flow_col, product_col = input_cols
    rolled_flow_col = f"rolled_{output_prefix}_flow"
    rolled_product_col = f"rolled_{output_prefix}_product"
    group_col = f"{output_prefix}_rollup_group_id"
    reason_col = f"{output_prefix}_rollup_reason"
    rule_sheet_col = f"{output_prefix}_rollup_rule_sheet"
    rule_row_col = f"{output_prefix}_rollup_rule_row"

    active_rules = active_rollup_rules(rules_df, context)
    columns = rollup_columns_for_sheet(rollup_sheet_name)
    used_rows: list[dict[str, Any]] = []
    ambiguous_frames: list[pd.DataFrame] = []

    for column in [flow_col, product_col]:
        if column not in out.columns:
            out[column] = ""
        out[column] = out[column].fillna("").astype(str).str.strip()

    for idx, row in out.iterrows():
        input_flow = clean_text(row.get(flow_col, ""))
        input_product = clean_text(row.get(product_col, ""))
        chosen_rule, ambiguous_df = choose_rollup_match(
            row=row,
            rules_df=active_rules,
            rollup_sheet_name=rollup_sheet_name,
            context=context,
            input_flow_value=input_flow,
            input_product_value=input_product,
        )
        if not ambiguous_df.empty:
            ambiguous_frames.append(ambiguous_df)
            out.at[idx, rolled_flow_col] = input_flow
            out.at[idx, rolled_product_col] = input_product
            out.at[idx, group_col] = ""
            out.at[idx, reason_col] = "ambiguous_rollup_rule_not_applied"
            out.at[idx, rule_sheet_col] = rollup_sheet_name
            out.at[idx, rule_row_col] = ""
            continue
        if chosen_rule is None:
            out.at[idx, rolled_flow_col] = input_flow
            out.at[idx, rolled_product_col] = input_product
            out.at[idx, group_col] = ""
            out.at[idx, reason_col] = ""
            out.at[idx, rule_sheet_col] = ""
            out.at[idx, rule_row_col] = ""
            continue

        rolled_flow = clean_text(chosen_rule.get(columns["rolled_flow"], "")) or input_flow
        rolled_product = clean_text(chosen_rule.get(columns["rolled_product"], "")) or input_product
        group_id = clean_text(chosen_rule.get("rollup_group_id", "")) or make_rollup_group_id(rolled_flow, rolled_product)
        out.at[idx, rolled_flow_col] = rolled_flow
        out.at[idx, rolled_product_col] = rolled_product
        out.at[idx, group_col] = group_id
        out.at[idx, reason_col] = clean_text(chosen_rule.get("rollup_reason", ""))
        out.at[idx, rule_sheet_col] = rollup_sheet_name
        out.at[idx, rule_row_col] = clean_text(chosen_rule.get("_rule_index", ""))
        used_row = chosen_rule.to_dict()
        used_row["rollup_sheet"] = rollup_sheet_name
        used_row["rollup_context_used"] = context
        used_row["matched_flow"] = input_flow
        used_row["matched_product"] = input_product
        used_row["source_sheet"] = clean_text(row.get("source_sheet", ""))
        used_row["source_row_number"] = row.get("source_row_number", "")
        used_rows.append(used_row)

    used_df = pd.DataFrame(used_rows)
    ambiguous_df = pd.concat(ambiguous_frames, ignore_index=True) if ambiguous_frames else pd.DataFrame()
    return out, used_df, ambiguous_df


def add_cardinality(
    frame: pd.DataFrame,
    source_cols: list[str],
    target_cols: list[str],
    output_col: str,
) -> pd.DataFrame:
    """Calculate cardinality between source and target pairs."""
    out = frame.copy()
    for column in [*source_cols, *target_cols]:
        if column not in out.columns:
            out[column] = ""
        out[column] = out[column].fillna("").astype(str).str.strip()
    valid = active_mask(out)
    for column in [*source_cols, *target_cols]:
        valid &= out[column].ne("")
    pairs = out.loc[valid, [*source_cols, *target_cols]].drop_duplicates().copy()
    out[output_col] = ""
    if pairs.empty:
        return out
    pairs["_source_key"] = pairs[source_cols].agg("|||".join, axis=1)
    pairs["_target_key"] = pairs[target_cols].agg("|||".join, axis=1)
    source_target_counts = pairs.groupby("_source_key")["_target_key"].nunique()
    target_source_counts = pairs.groupby("_target_key")["_source_key"].nunique()
    out["_source_key"] = out[source_cols].agg("|||".join, axis=1)
    out["_target_key"] = out[target_cols].agg("|||".join, axis=1)
    valid_row_values = pd.Series(True, index=out.index)
    for column in [*source_cols, *target_cols]:
        valid_row_values &= out[column].ne("")
    out.loc[valid_row_values, output_col] = out.loc[valid_row_values].apply(
        lambda row: cardinality_label(
            int(source_target_counts.get(row["_source_key"], 0)),
            int(target_source_counts.get(row["_target_key"], 0)),
        ),
        axis=1,
    )
    return out.drop(columns=["_source_key", "_target_key"])


def build_effective_mapping(
    source_df: pd.DataFrame,
    config: dict[str, Any],
    rollup_rules: dict[str, pd.DataFrame],
) -> tuple[pd.DataFrame, dict[str, pd.DataFrame]]:
    """Apply rollups and cardinality to one configured mapping sheet/use case."""
    out = source_df.copy()
    out = out.loc[:, [column for column in out.columns if not str(column).startswith("Unnamed:")]].copy()
    out["source_sheet"] = config["source_sheet"]
    out["source_row_number"] = out.index + 2
    for column in [*config["source_cols"], *config["target_cols"]]:
        if column not in out.columns:
            out[column] = ""
        out[column] = out[column].fillna("").astype(str).str.strip()
    if "remove_row" not in out.columns:
        out["remove_row"] = False
    if "Note" not in out.columns:
        out["Note"] = ""
    out = annotate_rollup_created_categories(out, config, rollup_rules)

    out = add_cardinality(
        out,
        source_cols=config["source_cols"],
        target_cols=config["target_cols"],
        output_col="pair_mapping_cardinality_raw",
    )
    out, source_used, source_ambiguous = apply_rollup_axis(
        out,
        rules_df=rollup_rules.get(config["source_rollup_sheet"], pd.DataFrame()),
        rollup_sheet_name=config["source_rollup_sheet"],
        context=config["rollup_context"],
        input_cols=config["source_cols"],
        output_prefix="source",
    )
    out, target_used, target_ambiguous = apply_rollup_axis(
        out,
        rules_df=rollup_rules.get(config["target_rollup_sheet"], pd.DataFrame()),
        rollup_sheet_name=config["target_rollup_sheet"],
        context=config["rollup_context"],
        input_cols=config["target_cols"],
        output_prefix="target",
    )
    out = add_cardinality(
        out,
        source_cols=["rolled_source_flow", "rolled_source_product"],
        target_cols=["rolled_target_flow", "rolled_target_product"],
        output_col="pair_mapping_cardinality_after_rollup",
    )
    out["pair_mapping_cardinality"] = out["pair_mapping_cardinality_after_rollup"]
    out["use_case"] = config["use_case"]
    out["source_system"] = config["source_system"]
    out["target_system"] = config["target_system"]
    out["rollup_applied"] = (
        out["rolled_source_flow"].ne(out[config["source_cols"][0]])
        | out["rolled_source_product"].ne(out[config["source_cols"][1]])
        | out["rolled_target_flow"].ne(out[config["target_cols"][0]])
        | out["rolled_target_product"].ne(out[config["target_cols"][1]])
    )
    out["rollup_reason"] = out[["source_rollup_reason", "target_rollup_reason"]].apply(
        lambda row: " | ".join([clean_text(value) for value in row if clean_text(value)]),
        axis=1,
    )

    qa = {
        "rules_used": pd.concat([source_used, target_used], ignore_index=True),
        "rules_ambiguous": pd.concat([source_ambiguous, target_ambiguous], ignore_index=True),
    }
    return out, qa


def build_all_effective_mappings(
    workbook_path: Path,
    include_reverse: bool = True,
) -> tuple[dict[str, pd.DataFrame], dict[str, pd.DataFrame]]:
    """Read the workbook and build effective mappings for all configured use cases."""
    rollup_rules = read_rollup_rules(workbook_path)
    effective: dict[str, pd.DataFrame] = {}
    qa_frames: dict[str, list[pd.DataFrame]] = {"rules_used": [], "rules_ambiguous": []}
    configs = MAPPING_SHEET_CONFIGS if include_reverse else [config for config in MAPPING_SHEET_CONFIGS if not config.get("is_reverse")]
    for config in configs:
        source_df = read_required_sheet(workbook_path, config["source_sheet"])
        effective_df, qa = build_effective_mapping(source_df, config, rollup_rules)
        key = config["use_case"]
        effective[key] = effective_df
        for qa_name, qa_df in qa.items():
            if not qa_df.empty:
                qa_frames.setdefault(qa_name, []).append(qa_df)

    qa_tables = {
        qa_name: pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()
        for qa_name, frames in qa_frames.items()
    }
    qa_tables["rules_unused"] = build_unused_rules_qa(rollup_rules, qa_tables.get("rules_used", pd.DataFrame()))
    return effective, qa_tables


def build_unused_rules_qa(rollup_rules: dict[str, pd.DataFrame], used_df: pd.DataFrame) -> pd.DataFrame:
    """Return active rollup rules that did not match any mapping row."""
    used_keys = set()
    if not used_df.empty and {"rollup_sheet", "_rule_index"}.issubset(used_df.columns):
        used_keys = set(zip(used_df["rollup_sheet"].astype(str), used_df["_rule_index"].astype(str)))
    rows: list[dict[str, Any]] = []
    contexts = sorted({config["rollup_context"] for config in MAPPING_SHEET_CONFIGS})
    for sheet_name, rules_df in rollup_rules.items():
        if sheet_name == "rollup_label_overrides":
            continue
        for context in contexts:
            active_df = active_rollup_rules(rules_df, context)
            for _, row in active_df.iterrows():
                key = (sheet_name, str(row.get("_rule_index", "")))
                if key in used_keys:
                    continue
                output_row = row.to_dict()
                output_row["rollup_sheet"] = sheet_name
                output_row["rollup_context_checked"] = context
                rows.append(output_row)
    return pd.DataFrame(rows).drop_duplicates() if rows else pd.DataFrame()


def many_to_many_rows(frame: pd.DataFrame, cardinality_col: str) -> pd.DataFrame:
    """Return active rows with many-to-many cardinality."""
    if frame.empty or cardinality_col not in frame.columns:
        return pd.DataFrame()
    out = frame[active_mask(frame) & frame[cardinality_col].eq("many_to_many")].copy()
    if out.empty:
        return out
    out["qa_severity"] = "high" if cardinality_col == "pair_mapping_cardinality_after_rollup" else "warning"
    return out


def build_mapping_rows_changed_by_rollup(effective_tables: dict[str, pd.DataFrame]) -> pd.DataFrame:
    """Return rows whose effective category changed because of rollup rules."""
    frames = [df[df["rollup_applied"]].copy() for df in effective_tables.values() if "rollup_applied" in df.columns]
    return pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()


def build_duplicate_effective_relationships(relationship_df: pd.DataFrame) -> pd.DataFrame:
    """Find duplicate effective source-target relationships after rollup."""
    if relationship_df.empty:
        return pd.DataFrame()
    key_cols = ["use_case", "source_system", "rolled_source_flow", "rolled_source_product", "target_system", "rolled_target_flow", "rolled_target_product"]
    counts = (
        relationship_df[relationship_df["include_in_use_case"]]
        .groupby(key_cols, dropna=False)
        .agg(
            duplicate_row_count=("relationship_id", "size"),
            source_rows=("source_row_number", lambda values: "|".join(str(value) for value in values)),
        )
        .reset_index()
    )
    return counts[counts["duplicate_row_count"] > 1].copy()


def build_cardinality_summary(effective_tables: dict[str, pd.DataFrame]) -> pd.DataFrame:
    """Count rows by source sheet, use case, raw/effective cardinality, and include flag."""
    frames: list[pd.DataFrame] = []
    for df in effective_tables.values():
        work = df.copy()
        work["include_in_use_case"] = ~work["remove_row"].map(parse_bool)
        frames.append(work)
    if not frames:
        return pd.DataFrame()
    all_df = pd.concat(frames, ignore_index=True)
    return (
        all_df.groupby(
            [
                "source_sheet",
                "use_case",
                "pair_mapping_cardinality_raw",
                "pair_mapping_cardinality_after_rollup",
                "include_in_use_case",
            ],
            dropna=False,
        )
        .size()
        .reset_index(name="row_count")
    )


def _exception_matches(
    exceptions: pd.DataFrame,
    *,
    check_name: str,
    source_value: str,
    target_value: str,
) -> pd.DataFrame:
    """Return exception rows matching a check/source/target, with blanks as wildcards."""
    if exceptions.empty:
        return pd.DataFrame()
    work = exceptions.copy()
    check_key = normalise_key(check_name)
    source_key = normalise_key(source_value)
    target_key = normalise_key(target_value)
    return work[
        work["check_name"].map(normalise_key).eq(check_key)
        & work["source_value"].map(lambda value: normalise_key(value) in {"", "*", source_key})
        & work["target_value"].map(lambda value: normalise_key(value) in {"", "*", target_key})
    ].copy()


def build_individual_mapping_consistency_qa(
    effective_tables: dict[str, pd.DataFrame],
    exceptions: pd.DataFrame | None = None,
) -> pd.DataFrame:
    """
    Check individual fuel/product mappings after rollups.

    Pair-level mappings can legitimately be one-to-many because sectors/flows
    differ. This QA ignores those pair columns and checks whether the product
    mapping alone is still one-to-one unless it has been recorded as an exception.
    """
    exceptions = exceptions.copy() if exceptions is not None else pd.DataFrame()
    rows: list[dict[str, Any]] = []
    for check in INDIVIDUAL_MAPPING_CHECKS:
        table = effective_tables.get(check["use_case"], pd.DataFrame()).copy()
        if table.empty:
            continue
        source_col = check["source_col"]
        target_col = check["target_col"]
        for column in [source_col, target_col]:
            if column not in table.columns:
                table[column] = ""
            table[column] = table[column].fillna("").astype(str).str.strip()
        valid = active_mask(table) & table[source_col].ne("") & table[target_col].ne("")
        pairs = table.loc[valid, [source_col, target_col]].drop_duplicates().copy()
        if pairs.empty:
            continue
        pairs["_source_key"] = pairs[source_col].map(normalise_key)
        pairs["_target_key"] = pairs[target_col].map(normalise_key)
        source_target_counts = pairs.groupby("_source_key")["_target_key"].nunique()
        target_source_counts = pairs.groupby("_target_key")["_source_key"].nunique()
        pairs["individual_mapping_cardinality"] = pairs.apply(
            lambda row: cardinality_label(
                int(source_target_counts.get(row["_source_key"], 0)),
                int(target_source_counts.get(row["_target_key"], 0)),
            ),
            axis=1,
        )
        issue_pairs = pairs[pairs["individual_mapping_cardinality"].ne("one_to_one")].copy()
        for _, pair_row in issue_pairs.iterrows():
            source_value = clean_text(pair_row[source_col])
            target_value = clean_text(pair_row[target_col])
            matching_exceptions = _exception_matches(
                exceptions,
                check_name=check["check_name"],
                source_value=source_value,
                target_value=target_value,
            )
            source_group = pairs[pairs["_source_key"].eq(pair_row["_source_key"])]
            target_group = pairs[pairs["_target_key"].eq(pair_row["_target_key"])]
            row_filter = valid & table[source_col].map(normalise_key).eq(pair_row["_source_key"]) & table[target_col].map(normalise_key).eq(pair_row["_target_key"])
            source_rows = "|".join(str(value) for value in table.loc[row_filter, "source_row_number"].drop_duplicates().tolist())
            rows.append(
                {
                    "check_name": check["check_name"],
                    "use_case": check["use_case"],
                    "source_label": check["source_label"],
                    "target_label": check["target_label"],
                    "source_value": source_value,
                    "target_value": target_value,
                    "individual_mapping_cardinality": pair_row["individual_mapping_cardinality"],
                    "source_maps_to_target_count": int(source_target_counts.get(pair_row["_source_key"], 0)),
                    "target_maps_to_source_count": int(target_source_counts.get(pair_row["_target_key"], 0)),
                    "all_targets_for_source": " | ".join(sorted(source_group[target_col].map(clean_text).drop_duplicates())),
                    "all_sources_for_target": " | ".join(sorted(target_group[source_col].map(clean_text).drop_duplicates())),
                    "source_sheet": table.loc[row_filter, "source_sheet"].map(clean_text).drop_duplicates().str.cat(sep="|"),
                    "source_row_numbers": source_rows,
                    "exception_status": "recorded_exception" if not matching_exceptions.empty else "unrecorded_exception",
                    "exception_reason": " | ".join(
                        sorted({clean_text(value) for value in matching_exceptions.get("exception_reason", pd.Series(dtype=object)) if clean_text(value)})
                    ),
                    "qa_severity": "info" if not matching_exceptions.empty else "warning",
                }
            )
    columns = [
        "check_name",
        "use_case",
        "source_label",
        "target_label",
        "source_value",
        "target_value",
        "individual_mapping_cardinality",
        "source_maps_to_target_count",
        "target_maps_to_source_count",
        "all_targets_for_source",
        "all_sources_for_target",
        "source_sheet",
        "source_row_numbers",
        "exception_status",
        "exception_reason",
        "qa_severity",
    ]
    if not rows:
        return pd.DataFrame(columns=columns)
    return pd.DataFrame(rows, columns=columns).sort_values(
        ["exception_status", "check_name", "source_value", "target_value"]
    ).reset_index(drop=True)


def build_original_label_presence_qa(
    effective_tables: dict[str, pd.DataFrame],
    valid_labels_by_name: dict[str, set[str]] | None = None,
    exceptions: pd.DataFrame | None = None,
) -> pd.DataFrame:
    """Report mapping labels that are absent from their original source dataset."""
    valid_labels_by_name = valid_labels_by_name or {}
    exceptions = exceptions.copy() if exceptions is not None else pd.DataFrame()
    rows: list[dict[str, Any]] = []
    for check in INDIVIDUAL_MAPPING_CHECKS:
        table = effective_tables.get(check["use_case"], pd.DataFrame()).copy()
        if table.empty:
            continue
        for side in ["source", "target"]:
            value_col = check[f"{side}_col"]
            label_name = check[f"{side}_label"]
            valid_values = valid_labels_by_name.get(label_name)
            if valid_values is None:
                continue
            if value_col not in table.columns:
                table[value_col] = ""
            table[value_col] = table[value_col].fillna("").astype(str).str.strip()
            valid_rows = active_mask(table) & table[value_col].ne("")
            values = table.loc[valid_rows, value_col].drop_duplicates().sort_values()
            for label_value in values.tolist():
                label_key = normalise_key(label_value)
                if label_key in valid_values:
                    continue
                row_filter = valid_rows & table[value_col].map(normalise_key).eq(label_key)
                source_rows = "|".join(str(value) for value in table.loc[row_filter, "source_row_number"].drop_duplicates().tolist())
                exception_check_name = f"{check['check_name']}_{side}_label_presence"
                if side == "source":
                    matching_exceptions = _exception_matches(
                        exceptions,
                        check_name=exception_check_name,
                        source_value=label_value,
                        target_value="",
                    )
                else:
                    matching_exceptions = _exception_matches(
                        exceptions,
                        check_name=exception_check_name,
                        source_value="",
                        target_value=label_value,
                    )
                rows.append(
                    {
                        "check_name": check["check_name"],
                        "presence_check_name": exception_check_name,
                        "use_case": check["use_case"],
                        "label_side": side,
                        "label_name": label_name,
                        "label_value": label_value,
                        "original_dataset": ORIGINAL_LABEL_DATASET_BY_LABEL.get(label_name, ""),
                        "issue_type": f"{side}_label_missing_from_original_dataset",
                        "source_sheet": table.loc[row_filter, "source_sheet"].map(clean_text).drop_duplicates().str.cat(sep="|"),
                        "source_row_numbers": source_rows,
                        "exception_status": "recorded_exception" if not matching_exceptions.empty else "unrecorded_exception",
                        "exception_reason": " | ".join(
                            sorted({clean_text(value) for value in matching_exceptions.get("exception_reason", pd.Series(dtype=object)) if clean_text(value)})
                        ),
                        "qa_severity": "info" if not matching_exceptions.empty else "warning",
                    }
                )
    columns = [
        "check_name",
        "presence_check_name",
        "use_case",
        "label_side",
        "label_name",
        "label_value",
        "original_dataset",
        "issue_type",
        "source_sheet",
        "source_row_numbers",
        "exception_status",
        "exception_reason",
        "qa_severity",
    ]
    if not rows:
        return pd.DataFrame(columns=columns)
    return pd.DataFrame(rows, columns=columns).sort_values(
        ["exception_status", "label_name", "label_value", "check_name"]
    ).reset_index(drop=True)


def _lookup_pair_subtotal(
    lookup: dict[tuple[str, str], bool],
    flow_value: Any,
    product_value: Any,
) -> bool | None:
    key = (normalise_key(flow_value), normalise_key(product_value))
    if key in lookup:
        return bool(lookup[key])
    return None


def build_subtotal_alignment_qa(
    effective_tables: dict[str, pd.DataFrame],
    subtotal_lookup_by_system: dict[str, dict[tuple[str, str], bool]] | None = None,
    exceptions: pd.DataFrame | None = None,
) -> pd.DataFrame:
    """Report source-target rows where a subtotal maps to a non-subtotal."""
    subtotal_lookup_by_system = subtotal_lookup_by_system or {}
    exceptions = exceptions.copy() if exceptions is not None else pd.DataFrame()
    rows: list[dict[str, Any]] = []
    for use_case, table in effective_tables.items():
        if table.empty:
            continue
        required_cols = [
            "source_system",
            "target_system",
            "rolled_source_flow",
            "rolled_source_product",
            "rolled_target_flow",
            "rolled_target_product",
            "source_sheet",
            "source_row_number",
        ]
        work = table.copy()
        for column in required_cols:
            if column not in work.columns:
                work[column] = ""
            work[column] = work[column].fillna("").astype(str).str.strip()
        valid = (
            active_mask(work)
            & work["rolled_source_flow"].ne("")
            & work["rolled_source_product"].ne("")
            & work["rolled_target_flow"].ne("")
            & work["rolled_target_product"].ne("")
        )
        for _, row in work.loc[valid].iterrows():
            source_system = clean_text(row.get("source_system", ""))
            target_system = clean_text(row.get("target_system", ""))
            source_is_subtotal = _lookup_pair_subtotal(
                subtotal_lookup_by_system.get(source_system, {}),
                row.get("rolled_source_flow", ""),
                row.get("rolled_source_product", ""),
            )
            target_is_subtotal = _lookup_pair_subtotal(
                subtotal_lookup_by_system.get(target_system, {}),
                row.get("rolled_target_flow", ""),
                row.get("rolled_target_product", ""),
            )
            if source_is_subtotal is None or target_is_subtotal is None:
                continue
            if bool(source_is_subtotal) == bool(target_is_subtotal):
                continue
            source_pair = f"{clean_text(row.get('rolled_source_flow', ''))} || {clean_text(row.get('rolled_source_product', ''))}"
            target_pair = f"{clean_text(row.get('rolled_target_flow', ''))} || {clean_text(row.get('rolled_target_product', ''))}"
            exception_check_name = f"{use_case}_subtotal_alignment"
            matching_exceptions = _exception_matches(
                exceptions,
                check_name=exception_check_name,
                source_value=source_pair,
                target_value=target_pair,
            )
            issue_type = "source_subtotal_to_target_non_subtotal"
            if not bool(source_is_subtotal) and bool(target_is_subtotal):
                issue_type = "source_non_subtotal_to_target_subtotal"
            rows.append(
                {
                    "check_name": exception_check_name,
                    "use_case": use_case,
                    "source_system": source_system,
                    "target_system": target_system,
                    "source_flow": clean_text(row.get("rolled_source_flow", "")),
                    "source_product": clean_text(row.get("rolled_source_product", "")),
                    "target_flow": clean_text(row.get("rolled_target_flow", "")),
                    "target_product": clean_text(row.get("rolled_target_product", "")),
                    "source_is_subtotal": bool(source_is_subtotal),
                    "target_is_subtotal": bool(target_is_subtotal),
                    "issue_type": issue_type,
                    "source_dataset": SUBTOTAL_DATASET_BY_SYSTEM.get(source_system, ""),
                    "target_dataset": SUBTOTAL_DATASET_BY_SYSTEM.get(target_system, ""),
                    "source_sheet": clean_text(row.get("source_sheet", "")),
                    "source_row_number": clean_text(row.get("source_row_number", "")),
                    "exception_status": "recorded_exception" if not matching_exceptions.empty else "unrecorded_exception",
                    "exception_reason": " | ".join(
                        sorted({clean_text(value) for value in matching_exceptions.get("exception_reason", pd.Series(dtype=object)) if clean_text(value)})
                    ),
                    "qa_severity": "info" if not matching_exceptions.empty else "warning",
                }
            )
    columns = [
        "check_name",
        "use_case",
        "source_system",
        "target_system",
        "source_flow",
        "source_product",
        "target_flow",
        "target_product",
        "source_is_subtotal",
        "target_is_subtotal",
        "issue_type",
        "source_dataset",
        "target_dataset",
        "source_sheet",
        "source_row_number",
        "exception_status",
        "exception_reason",
        "qa_severity",
    ]
    if not rows:
        return pd.DataFrame(columns=columns)
    return pd.DataFrame(rows, columns=columns).drop_duplicates().sort_values(
        ["exception_status", "use_case", "source_flow", "source_product", "target_flow", "target_product"]
    ).reset_index(drop=True)


def build_relationship_rows(effective_tables: dict[str, pd.DataFrame]) -> pd.DataFrame:
    """Build the compiled relationship table from effective mapping rows."""
    rows: list[dict[str, Any]] = []
    for use_case, df in effective_tables.items():
        for _, row in df.iterrows():
            include = not parse_bool(row.get("remove_row", False))
            relationship_status = "included_in_use_case"
            exclude_reason = ""
            if not include:
                relationship_status = "excluded_by_remove_row_for_use_case"
                exclude_reason = "remove_row_true_in_source_mapping"
            review_flags: list[str] = []
            if clean_text(row.get("pair_mapping_cardinality_after_rollup", "")) == "many_to_many":
                review_flags.append("many_to_many_after_rollup")
            if clean_text(row.get("source_rollup_reason", "")) == "ambiguous_rollup_rule_not_applied":
                review_flags.append("ambiguous_source_rollup")
            if clean_text(row.get("target_rollup_reason", "")) == "ambiguous_rollup_rule_not_applied":
                review_flags.append("ambiguous_target_rollup")

            output_row = {
                "relationship_id": "",
                "relationship_key": "",
                "use_case": use_case,
                "include_in_use_case": include,
                "source_system": row.get("source_system", ""),
                "source_flow": row.get("rolled_source_flow", ""),
                "source_product": row.get("rolled_source_product", ""),
                "target_system": row.get("target_system", ""),
                "target_flow": row.get("rolled_target_flow", ""),
                "target_product": row.get("rolled_target_product", ""),
                "original_source_flow": row.get("rolled_source_flow", "") if False else row.get(row.attrs.get("unused", ""), ""),
                "original_source_product": "",
                "original_target_flow": "",
                "original_target_product": "",
                "rolled_source_flow": row.get("rolled_source_flow", ""),
                "rolled_source_product": row.get("rolled_source_product", ""),
                "rolled_target_flow": row.get("rolled_target_flow", ""),
                "rolled_target_product": row.get("rolled_target_product", ""),
                "rollup_applied": bool(row.get("rollup_applied", False)),
                "source_rollup_group_id": row.get("source_rollup_group_id", ""),
                "target_rollup_group_id": row.get("target_rollup_group_id", ""),
                "rollup_reason": row.get("rollup_reason", ""),
                "pair_mapping_cardinality_raw": row.get("pair_mapping_cardinality_raw", ""),
                "pair_mapping_cardinality_after_rollup": row.get("pair_mapping_cardinality_after_rollup", ""),
                "pair_mapping_cardinality": row.get("pair_mapping_cardinality", ""),
                "cardinality": row.get("pair_mapping_cardinality", ""),
                "allocation_method": "direct",
                "allocation_source": "",
                "allocation_share": "",
                "relationship_type": "direct_or_existing_mapping",
                "relationship_level": "",
                "relationship_source": row.get("source_sheet", ""),
                "source_sector_path": row.get("rolled_source_flow", ""),
                "source_fuel": row.get("rolled_source_product", ""),
                "relationship_status": relationship_status,
                "exclude_reason": exclude_reason,
                "review_required": bool(review_flags),
                "review_flags": "|".join(review_flags),
                "remove_row": row.get("remove_row", ""),
                "source_sheet": row.get("source_sheet", ""),
                "source_row_number": row.get("source_row_number", ""),
                "Note": row.get("Note", ""),
            }
            config = next(config for config in MAPPING_SHEET_CONFIGS if config["use_case"] == use_case)
            output_row["original_source_flow"] = row.get(config["source_cols"][0], "")
            output_row["original_source_product"] = row.get(config["source_cols"][1], "")
            output_row["original_target_flow"] = row.get(config["target_cols"][0], "")
            output_row["original_target_product"] = row.get(config["target_cols"][1], "")
            output_row["relationship_id"] = make_relationship_id(pd.Series(output_row))
            output_row["relationship_key"] = f"{output_row['relationship_id']}::{use_case}"
            rows.append(output_row)
    return pd.DataFrame(rows, columns=RELATIONSHIP_COLUMNS)


def build_relationship_catalogue(relationship_df: pd.DataFrame) -> pd.DataFrame:
    """Build one row per effective relationship ID with use-case summary."""
    if relationship_df.empty:
        return pd.DataFrame()
    rows: list[dict[str, Any]] = []
    for relationship_id, group_df in relationship_df.groupby("relationship_id", dropna=False):
        first = group_df.iloc[0]
        rows.append(
            {
                "relationship_id": relationship_id,
                "source_system": first["source_system"],
                "source_flow": first["source_flow"],
                "source_product": first["source_product"],
                "target_system": first["target_system"],
                "target_flow": first["target_flow"],
                "target_product": first["target_product"],
                "included_use_cases": "|".join(group_df.loc[group_df["include_in_use_case"], "use_case"].astype(str)),
                "excluded_use_cases": "|".join(group_df.loc[~group_df["include_in_use_case"], "use_case"].astype(str)),
                "pair_mapping_cardinality": "|".join(sorted(set(group_df["pair_mapping_cardinality"].astype(str)))),
                "rollup_applied": bool(group_df["rollup_applied"].any()),
                "review_required": bool(group_df["review_required"].any()),
                "review_flags": "|".join(sorted({flag for flags in group_df["review_flags"].astype(str) for flag in flags.split("|") if flag})),
                "Note": "|".join(sorted({clean_text(value) for value in group_df["Note"] if clean_text(value)})),
            }
        )
    return pd.DataFrame(rows)


def build_qa_tables(
    effective_tables: dict[str, pd.DataFrame],
    relationship_df: pd.DataFrame,
    rollup_qa: dict[str, pd.DataFrame],
    individual_mapping_exceptions: pd.DataFrame | None = None,
    valid_labels_by_name: dict[str, set[str]] | None = None,
    subtotal_lookup_by_system: dict[str, dict[tuple[str, str], bool]] | None = None,
    mapping_balance_coverage: pd.DataFrame | None = None,
) -> dict[str, pd.DataFrame]:
    """Build requested QA output tables."""
    before_frames = [many_to_many_rows(df, "pair_mapping_cardinality_raw") for df in effective_tables.values()]
    after_frames = [many_to_many_rows(df, "pair_mapping_cardinality_after_rollup") for df in effective_tables.values()]
    changed_df = build_mapping_rows_changed_by_rollup(effective_tables)
    after_df = pd.concat(after_frames, ignore_index=True) if after_frames else pd.DataFrame()
    missing_rollups_df = after_df.copy()
    if not missing_rollups_df.empty:
        missing_rollups_df["missing_rollup_reason"] = "many_to_many_after_rollup_needs_mapping_or_rollup_review"
    return {
        "qa_many_to_many_before_rollup": pd.concat(before_frames, ignore_index=True) if before_frames else pd.DataFrame(),
        "qa_many_to_many_after_rollup": after_df,
        "qa_rollup_rules_used": rollup_qa.get("rules_used", pd.DataFrame()),
        "qa_rollup_rules_unused": rollup_qa.get("rules_unused", pd.DataFrame()),
        "qa_rollup_rules_ambiguous": rollup_qa.get("rules_ambiguous", pd.DataFrame()),
        "qa_mapping_rows_changed_by_rollup": changed_df,
        "qa_missing_required_rollups": missing_rollups_df,
        "qa_duplicate_effective_relationships": build_duplicate_effective_relationships(relationship_df),
        "qa_cardinality_summary": build_cardinality_summary(effective_tables),
        "qa_individual_mapping_consistency": build_individual_mapping_consistency_qa(
            effective_tables,
            individual_mapping_exceptions,
        ),
        "qa_original_label_presence": build_original_label_presence_qa(
            effective_tables,
            valid_labels_by_name,
            individual_mapping_exceptions,
        ),
        "qa_subtotal_alignment": build_subtotal_alignment_qa(
            effective_tables,
            subtotal_lookup_by_system,
            individual_mapping_exceptions,
        ),
        "qa_mapping_balance_coverage": (
            mapping_balance_coverage.copy()
            if mapping_balance_coverage is not None
            else pd.DataFrame()
        ),
    }

#%%
